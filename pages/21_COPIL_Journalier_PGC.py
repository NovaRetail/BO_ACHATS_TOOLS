"""
COPIL Journalier PGC - SmartBuyer Hub
Generates a daily-COPIL-ready Excel workbook from the BI sales export.

Input : daily export (.xlsx) + optional D-1 export to flag new hot spots
Output: 4-sheet workbook (01_DIRECTION / 02_PGC_DETAIL / 03_POINTS_CHAUDS / 04_LECTURE) + hidden 90_DATA
"""

from __future__ import annotations

import io
from datetime import date, datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIGURATION
# ----------------------------------------------------------------------------

SHEET_NAME = "Export"

FORMAT_MAP = {
    "10301 - Hyper Marcory": "Hyper",
    "10202 - Hyper Palmeraie": "Hyper",
    "10203 - Hyper Yopougon": "Hyper",
    "10705 - Market 7 Décembre": "Market",
    "10208 - Market Riviera": "Market",
    "10209 - Market 2 Plateaux": "Market",
    "10206 - Market Kokoh Mall": "Market",
    "10604 - Market Cité verte": "Market",
    "10605 - Market Aboboté": "Market",
    "10601 - Supeco Niangon": "Supeco",
    "10602 - Supeco Terminus 47": "Supeco",
    "10603 - Supeco Toit rouge": "Supeco",
}

RAYON_LABELS = {
    "010 - BOISSON": "Boisson",
    "011 - DROGUERIE": "Droguerie",
    "012 - PARFUMERIE HYGIENE": "Parfumerie Hygiène",
    "014 - EPICERIE": "Épicerie",
}

DEPT_LABELS = {
    "01 - PGC": "PGC",
    "02 - PRODUITS FRAIS": "Produits Frais",
    "03 - BAZAR": "Bazar",
    "04 - EPCS": "EPCS",
    "06 - TEXTILE": "Textile",
}

FORMATS = ["Hyper", "Market", "Supeco"]
JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

# Thresholds (SPEC-07 / SPEC-15)
SEUIL_CROISSANCE_VERTE = 0.05
SEUIL_BUDGET_ROUGE = -0.03
SEUIL_MARGE_ROUGE_PTS = -1.0
SEUIL_ECART_FORMAT_PTS = 2.0
RATIO_DOMINANCE = 1.3      # trafic vs panier arbitration (SPEC-03)
TOP_PCT = 0.20             # SPEC-11/12 amended: relative top 20%...
MIN_LIGNES = 5             # ...with a floor of 5 lines
MAX_PAR_SITE = 3           # SPEC-11: cap per site

# Visual charter (SmartBuyer Hub)
BLUE, RED, GREEN, ORANGE = "007AFF", "FF3B30", "34C759", "FF9500"
DARK, GREY = "1C1C1E", "8E8E93"
FONT_NAME = "Arial"

FMT_CUR = '#,##0" FCFA"'
FMT_PCT = "0.0%;[RED]-0.0%"
FMT_PCT_SIGN = "+0.0%;[RED]-0.0%"
FMT_PTS = '+0.0" pts";[RED]-0.0" pts"'
FMT_NUM = "#,##0"


# ----------------------------------------------------------------------------
# DATA PREPARATION
# ----------------------------------------------------------------------------

def _niveau(row) -> str:
    """Tag each export row with its aggregation level."""
    if row["Département"] == "Total":
        return "GRAND_TOTAL"
    if row["Rayon"] == "Total" and pd.isna(row["Site"]):
        return "DEPT_TOTAL"
    if row["Département"] == "01 - PGC" and row["Site"] == "Total":
        return "PGC_RAYON_TOTAL"
    if row["Département"] == "01 - PGC" and row["Site"] in FORMAT_MAP:
        return "PGC_SITE"
    return "AUTRE"


def load_export(file) -> pd.DataFrame:
    """Read the BI export and drop its trailing blank / filter rows."""
    df = pd.read_excel(file, sheet_name=SHEET_NAME)
    df = df[df["Département"].notna()].copy()
    df = df[~df["Département"].astype(str).str.startswith("Filtres appliqués")].copy()
    df["Niveau"] = df.apply(_niveau, axis=1)
    df["Format"] = df["Site"].map(FORMAT_MAP)
    return df


def compute_site_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Derive every PGC site x rayon metric used for prioritisation."""
    s = df[(df["Département"] == "01 - PGC") & (df["Site"].isin(FORMAT_MAP))].copy()
    s["RayonL"] = s["Rayon"].map(RAYON_LABELS).fillna(s["Rayon"])
    s["SiteL"] = s["Site"].astype(str).str.split(" - ").str[-1]

    s["TM_N1"] = s["Taux de Marge N-1"]
    s["TM_N"] = s["Taux de Marge"]

    # SPEC-09: total margin at risk = rate dilution + volume effect
    s["dilution"] = (s["TM_N1"] - s["TM_N"]) * s["CA"]
    s["effet_volume"] = -(s["CA"] - s["CA N-1"]) * s["TM_N1"]
    s["marge_risque"] = s["dilution"] + s["effet_volume"]

    s["g_ca"] = s["CA"] / s["CA N-1"] - 1
    s["g_bgt"] = np.where(s["Budget"].notna() & (s["Budget"] != 0), s["CA"] / s["Budget"] - 1, np.nan)
    s["g_debit"] = s["Débit"] / s["Débit N-1"] - 1
    s["g_panier"] = s["Panier"] / s["Panier N-1"] - 1
    s["g_qte"] = s["Panier Qté"] / s["Panier Qté N-1"] - 1
    # SPEC-03: basket splits into quantity x average price
    s["g_prix"] = (1 + s["g_panier"]) / (1 + s["g_qte"]) - 1
    s["delta_tm_pts"] = (s["TM_N"] - s["TM_N1"]) * 100

    s["budget_dispo"] = s["Budget"].notna() & (s["Budget"] != 0)
    s["marge_negative"] = s["TM_N"] < 0

    s[["cause", "owner"]] = s.apply(lambda r: pd.Series(_cause_owner(r)), axis=1)
    s["signal"] = s.apply(_signal, axis=1)
    return s


def _cause_owner(r) -> tuple[str, str]:
    """SPEC-03/04/05: dominant effect vs N-1, and who owns it by default."""
    gd, gp, gq, gx = r["g_debit"], r["g_panier"], r["g_qte"], r["g_prix"]
    marge_baisse = r["delta_tm_pts"] < SEUIL_MARGE_ROUGE_PTS

    # SPEC-10: a negative margin is sold below cost — the traffic/basket split
    # would be misleading here, the question is pricing/costing, not footfall.
    if r["marge_negative"]:
        return "Vente à perte", "Achats"

    if pd.isna(gd) or pd.isna(gp):
        return "n/d", "Arbitrage"

    if abs(gd) >= abs(gp) * RATIO_DOMINANCE:
        if gd < 0:
            return "Trafic ⬇", "Magasin"
        # Traffic up but margin eroding -> pricing/mix issue, not a store issue
        return "Trafic ⬆", "Achats" if marge_baisse else "—"

    if abs(gp) >= abs(gd) * RATIO_DOMINANCE:
        if not pd.isna(gq) and not pd.isna(gx):
            if abs(gq) >= abs(gx) * RATIO_DOMINANCE:
                if gq < 0:
                    return "Quantité ⬇", "Magasin"
                return "Quantité ⬆", "Achats" if marge_baisse else "—"
            if abs(gx) >= abs(gq) * RATIO_DOMINANCE:
                if gx < 0:
                    return "Prix/mix ⬇", "Achats"
                return "Prix/mix ⬆", "Achats" if marge_baisse else "—"
        if gp < 0:
            return "Panier ⬇", "Arbitrage"
        return "Panier ⬆", "Achats" if marge_baisse else "—"

    return "Mixte", "Arbitrage"


def _signal(r) -> str:
    """SPEC-05: signal reads vs N-1 and margin; budget stays a separate column."""
    parts = []
    if r["marge_negative"]:
        parts.append("Marge négative")
    if not pd.isna(r["g_ca"]) and r["g_ca"] < -0.10:
        parts.append(f"CA/N-1 {r['g_ca']:+.1%}".replace(".", ","))
    if r["delta_tm_pts"] < -2:
        parts.append(f"TM {r['delta_tm_pts']:+.1f} pt".replace(".", ","))
    if r["budget_dispo"] and not pd.isna(r["g_bgt"]) and r["g_bgt"] < -0.20:
        parts.append(f"CA/Bgt {r['g_bgt']:+.1%}".replace(".", ","))
    return " · ".join(parts) if parts else "Dilution de marge"


def select_hot_spots(s: pd.DataFrame) -> pd.DataFrame:
    """SPEC-02/10/11/12: normalised scoring, negative margin first, relative cut."""
    s = s.copy()
    # SPEC-02: score normalised over the criteria actually available
    crit_ca = (s["g_ca"] < -0.10).astype(int)
    crit_tm = (s["delta_tm_pts"] < SEUIL_MARGE_ROUGE_PTS).astype(int)
    crit_bgt = np.where(s["budget_dispo"], (s["g_bgt"] < -0.20).astype(int), np.nan)
    nb_crit = np.where(s["budget_dispo"], 3, 2)
    total = crit_ca + crit_tm + np.nan_to_num(crit_bgt)
    s["score_norm"] = total / nb_crit

    ranked = s.sort_values("marge_risque", ascending=False)
    n = max(MIN_LIGNES, int(np.ceil(len(ranked) * TOP_PCT)))

    negatives = ranked[ranked["marge_negative"]]
    others = ranked[~ranked["marge_negative"]].head(n)
    sel = pd.concat([negatives, others]).drop_duplicates(subset=["Rayon", "Site"])
    # SPEC-11: never more than N lines for a single store
    sel = sel.groupby("SiteL", group_keys=False).head(MAX_PAR_SITE)
    return sel.sort_values(["marge_negative", "marge_risque"], ascending=[False, False])


def select_top_perf(s: pd.DataFrame) -> pd.DataFrame:
    """SPEC-06: three wins worth replicating."""
    p = s[(s["g_ca"] > 0) & (s["TM_N"] >= s["TM_N1"]) & (~s["marge_negative"])]
    return p.sort_values("CA", ascending=False).head(3)


def flag_new_lines(current: pd.DataFrame, previous: pd.DataFrame | None) -> pd.DataFrame:
    """Optional D-1 comparison: mark hot spots that were not there yesterday."""
    current = current.copy()
    if previous is None or previous.empty:
        current["nouveau"] = ""
        return current
    prev_keys = set(zip(previous["Rayon"], previous["Site"]))
    current["nouveau"] = [
        "🆕" if (r, s_) not in prev_keys else "" for r, s_ in zip(current["Rayon"], current["Site"])
    ]
    return current


# ----------------------------------------------------------------------------
# EXCEL STYLING HELPERS
# ----------------------------------------------------------------------------

def _styles():
    return {
        "title": Font(name=FONT_NAME, size=20, bold=True, color=DARK),
        "subtitle": Font(name=FONT_NAME, size=11, color=GREY, italic=True),
        "h1": Font(name=FONT_NAME, size=13, bold=True, color=BLUE),
        "h2": Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF"),
        "label": Font(name=FONT_NAME, size=10, bold=True, color=DARK),
        "val": Font(name=FONT_NAME, size=10, color=DARK),
        "big": Font(name=FONT_NAME, size=15, bold=True, color=DARK),
        "note": Font(name=FONT_NAME, size=9, italic=True, color=GREY),
        "quote": Font(name=FONT_NAME, size=11, italic=True, color=DARK),
        "say": Font(name=FONT_NAME, size=12, bold=True, color=DARK),
    }


FILLS = {
    "header": PatternFill("solid", fgColor=BLUE),
    "kpi": PatternFill("solid", fgColor="FFFFFF"),
    "band": PatternFill("solid", fgColor="F7F7FA"),
    "green": PatternFill("solid", fgColor="E4F8EA"),
    "orange": PatternFill("solid", fgColor="FFF2E0"),
    "red": PatternFill("solid", fgColor="FDE7E6"),
    "quote": PatternFill("solid", fgColor="F2F2F7"),
    "section": PatternFill("solid", fgColor="EDEFF7"),
    "saisie": PatternFill("solid", fgColor="FFF8EC"),
}

THIN = Side(style="thin", color="D1D1D6")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, c1, c2, st_):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = st_["h2"]
        cell.fill = FILLS["header"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def _data_row(ws, r, c1, c2, st_, saisie_from=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = st_["val"]
        cell.border = BOX
        cell.alignment = Alignment(horizontal="left" if c == c1 else "center", wrap_text=True)
        if saisie_from is not None and c >= saisie_from:
            cell.fill = FILLS["saisie"]
        elif r % 2 == 0:
            cell.fill = FILLS["band"]


def _section(ws, row, text, st_, note=None, last_col=10):
    ws.cell(row=row, column=2, value=text).font = st_["h1"]
    for c in range(2, last_col + 1):
        ws.cell(row=row, column=c).fill = FILLS["section"]
    if note:
        ws.cell(row=row + 1, column=2, value=note).font = st_["note"]
        return row + 2
    return row + 1


def _status_cf(ws, col, r1, r2):
    cl = get_column_letter(col)
    rng = f"{cl}{r1}:{cl}{r2}"
    for emoji, fill in (("🟢", "green"), ("🟠", "orange"), ("🔴", "red")):
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("{emoji}",{cl}{r1}))'], fill=FILLS[fill])
        )


# ----------------------------------------------------------------------------
# WORKBOOK BUILD
# ----------------------------------------------------------------------------

def build_workbook(df: pd.DataFrame, sites: pd.DataFrame, hot: pd.DataFrame,
                   perf: pd.DataFrame, jour_label: str) -> bytes:
    st_ = _styles()
    wb = Workbook()

    # ---------- 90_DATA ----------
    wsd = wb.active
    wsd.title = "90_DATA"
    cols = ["Niveau", "Département", "Rayon", "Site", "Format", "CA N-1", "Budget", "CA",
            "Marge N-1", "Marge", "Débit N-1", "Débit", "Panier N-1", "Panier",
            "Panier Qté N-1", "Panier Qté"]
    for j, h in enumerate(cols, start=1):
        wsd.cell(row=1, column=j, value=h)
    for i, (_, row) in enumerate(df[cols].iterrows()):
        for j, h in enumerate(cols, start=1):
            v = row[h]
            wsd.cell(row=i + 2, column=j, value=None if pd.isna(v) else v)

    n = len(df) + 1
    R = {k: f"90_DATA!${get_column_letter(i+1)}$2:${get_column_letter(i+1)}${n}"
         for i, k in enumerate(["NIV", "DEP", "RAY", "SIT", "FMT", "CAN1", "BUD", "CA",
                                "MGN1", "MG", "DBN1", "DB", "PANN1", "PAN", "PQN1", "PQ"])}

    # ---------- 01_DIRECTION ----------
    ws = wb.create_sheet("01_DIRECTION")
    ws.sheet_view.showGridLines = False
    _autosize(ws, [22, 16, 15, 15, 15, 15, 15, 20, 4])

    ws["B2"] = "COPIL JOURNALIER — PGC"
    ws["B2"].font = st_["title"]
    ws["B3"] = f"{jour_label} · Réseau Carrefour Côte d'Ivoire · photo du jour, non cumulée"
    ws["B3"].font = st_["subtitle"]
    ws["B4"] = ("Effet calendaire : les niveaux varient fortement selon le jour de semaine. "
                "Comparer à un jour équivalent avant de conclure à une tendance.")
    ws["B4"].font = st_["note"]
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=8)

    # --- Global KPI (needed before the "3 choses à dire" block references them) ---
    r = 25  # placed below; computed here so formulas can point at it
    kpi_row = None

    # 3 choses à dire (SPEC-18)
    r0 = _section(ws, 6, "📌 LES 3 CHOSES À DIRE", st_)
    say_rows = []
    for i in range(3):
        rr = r0 + i
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
        cell = ws.cell(row=rr, column=2)
        cell.font = st_["say"]
        cell.fill = FILLS["quote"]
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[rr].height = 20
        say_rows.append(rr)
    r_after_say = say_rows[-1] + 2

    # Alerte marge négative (SPEC-14)
    negs = sites[sites["marge_negative"]]
    r0 = _section(ws, r_after_say, "⛔ ALERTE MARGE NÉGATIVE", st_)
    if negs.empty:
        ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=8)
        c = ws.cell(row=r0, column=2, value="Aucune alerte — RAS")
        c.font = st_["label"]
        c.fill = FILLS["green"]
        c.border = BOX
        r_alert_last = r0
    else:
        heads = ["Rayon", "Site", "CA", "Taux de marge", "Marge (FCFA)", "Cause probable"]
        for j, h in enumerate(heads, start=2):
            ws.cell(row=r0, column=j, value=h)
        _header_row(ws, r0, 2, 2 + len(heads) - 1, st_)
        rr = r0 + 1
        for _, row in negs.iterrows():
            ws.cell(row=rr, column=2, value=row["RayonL"]).font = st_["label"]
            ws.cell(row=rr, column=3, value=row["SiteL"])
            ws.cell(row=rr, column=4, value=float(row["CA"])).number_format = FMT_CUR
            ws.cell(row=rr, column=5, value=float(row["TM_N"])).number_format = FMT_PCT
            ws.cell(row=rr, column=6, value=float(row["Marge"])).number_format = FMT_CUR
            ws.cell(row=rr, column=7, value=row["cause"])
            _data_row(ws, rr, 2, 7, st_)
            for c in range(2, 8):
                ws.cell(row=rr, column=c).fill = FILLS["red"]
            rr += 1
        r_alert_last = rr - 1

    # Bloc A — vue globale réseau
    r0 = _section(ws, r_alert_last + 2, "A — VUE GLOBALE RÉSEAU", st_)
    heads = ["CA Réseau", "Marge Réseau", "Taux de Marge", "Croissance CA vs N-1",
             "Croissance Marge vs N-1", "Vs Budget"]
    for j, h in enumerate(heads, start=2):
        ws.cell(row=r0, column=j, value=h)
    _header_row(ws, r0, 2, 2 + len(heads) - 1, st_)
    r = r0 + 1
    G = f'{R["NIV"]},"GRAND_TOTAL"'
    ws.cell(row=r, column=2, value=f'=SUMIFS({R["CA"]},{G})').number_format = FMT_CUR
    ws.cell(row=r, column=3, value=f'=SUMIFS({R["MG"]},{G})').number_format = FMT_CUR
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = FMT_PCT
    ws.cell(row=r, column=5, value=f'=B{r}/SUMIFS({R["CAN1"]},{G})-1').number_format = FMT_PCT_SIGN
    ws.cell(row=r, column=6, value=f'=C{r}/SUMIFS({R["MGN1"]},{G})-1').number_format = FMT_PCT_SIGN
    ws.cell(row=r, column=7, value=f'=B{r}/SUMIFS({R["BUD"]},{G})-1').number_format = FMT_PCT_SIGN
    for c in range(2, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = st_["big"]
        cell.border = BOX
        cell.fill = FILLS["kpi"]
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 22
    kpi_row = r

    # Bloc B — PGC vs tendance globale
    r0 = _section(ws, kpi_row + 2, "B — PGC VS TENDANCE GLOBALE", st_)
    heads = ["Département", "CA", "Vs N-1 (%)", "Vs Budget (%)", "Taux de Marge",
             "Marge vs N-1 (pts)", "Statut"]
    for j, h in enumerate(heads, start=2):
        ws.cell(row=r0, column=j, value=h)
    _header_row(ws, r0, 2, 2 + len(heads) - 1, st_)
    rr = r0 + 1
    pgc_row = rr
    order = [("01 - PGC", "PGC"), ("__G__", "Réseau Global")] + \
            [(k, v) for k, v in DEPT_LABELS.items() if k != "01 - PGC"]
    for code, label in order:
        crit = G if code == "__G__" else f'{R["NIV"]},"DEPT_TOTAL",{R["DEP"]},"{code}"'
        ws.cell(row=rr, column=2, value=label).font = st_["label"]
        ws.cell(row=rr, column=3, value=f'=SUMIFS({R["CA"]},{crit})').number_format = FMT_CUR
        ws.cell(row=rr, column=4, value=f'=C{rr}/SUMIFS({R["CAN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws.cell(row=rr, column=5, value=f'=IFERROR(C{rr}/SUMIFS({R["BUD"]},{crit})-1,"—")').number_format = FMT_PCT_SIGN
        ws.cell(row=rr, column=6, value=f'=SUMIFS({R["MG"]},{crit})/C{rr}').number_format = FMT_PCT
        ws.cell(row=rr, column=7,
                value=f'=(F{rr}-SUMIFS({R["MGN1"]},{crit})/SUMIFS({R["CAN1"]},{crit}))*100').number_format = FMT_PTS
        ws.cell(row=rr, column=8, value=(
            f'=IF(AND(D{rr}>={SEUIL_CROISSANCE_VERTE},IFERROR(E{rr},0)>=0,G{rr}>=0),"🟢 Performant",'
            f'IF(OR(D{rr}<0,IFERROR(E{rr},0)<{SEUIL_BUDGET_ROUGE},G{rr}<{SEUIL_MARGE_ROUGE_PTS}),'
            f'"🔴 Sous tendance","🟠 À surveiller"))'))
        _data_row(ws, rr, 2, 8, st_)
        rr += 1
    _status_cf(ws, 8, pgc_row, rr - 1)
    bloc_b_last = rr - 1

    # Message de synthèse
    r0 = _section(ws, bloc_b_last + 2, "MESSAGE DE SYNTHÈSE", st_)
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0 + 3, end_column=8)
    msg = ws.cell(row=r0, column=2)
    nb_neg = len(negs)
    alerte_txt = (f"{nb_neg} ligne(s) en marge négative à traiter" if nb_neg
                  else "aucune ligne en marge négative")
    msg.value = (
        f'="Le réseau est à "&TEXT(G{kpi_row},"+0.0%;-0.0%")&" vs budget et "'
        f'&TEXT(E{kpi_row},"+0.0%;-0.0%")&" vs N-1, marge à "&TEXT(D{kpi_row},"0.0%")&". '
        f'PGC à "&TEXT(D{pgc_row},"+0.0%;-0.0%")&" vs N-1 ; {alerte_txt} ; '
        f'{len(hot)} point(s) chaud(s) Site x Rayon identifiés (onglet 03)."'
    )
    msg.font = st_["quote"]
    msg.fill = FILLS["quote"]
    msg.border = BOX
    msg.alignment = Alignment(wrap_text=True, vertical="top")

    # Fill the "3 choses à dire" now that KPI rows exist
    ws.cell(row=say_rows[0], column=2, value=(
        f'="1. Réseau : "&TEXT(B{kpi_row},"#,##0")&" FCFA, "'
        f'&TEXT(E{kpi_row},"+0.0%;-0.0%")&" vs N-1, "&TEXT(G{kpi_row},"+0.0%;-0.0%")&" vs budget."'))
    if nb_neg:
        first = negs.iloc[0]
        ws.cell(row=say_rows[1], column=2, value=(
            f'="2. Point dur : {first["RayonL"]} — {first["SiteL"]}, marge négative à "'
            f'&TEXT({float(first["TM_N"])},"0.0%")&" (traitement immédiat)."'))
    else:
        ws.cell(row=say_rows[1], column=2,
                value='="2. Point dur : aucune marge négative aujourd\'hui."')
    ws.cell(row=say_rows[2], column=2,
            value=f'="3. À décider : {len(hot)} point(s) chaud(s) à arbitrer en séance (onglet 03)."')

    ws.freeze_panes = "B6"

    # ---------- 02_PGC_DETAIL ----------
    ws2 = wb.create_sheet("02_PGC_DETAIL")
    ws2.sheet_view.showGridLines = False
    _autosize(ws2, [24, 16, 13, 13, 13, 13, 15, 17, 20])

    ws2["B2"] = "DÉTAIL PGC — RAYONS & FORMATS"
    ws2["B2"].font = st_["title"]
    ws2["B3"] = jour_label
    ws2["B3"].font = st_["subtitle"]

    r0 = _section(ws2, 5, "C — RAYONS PGC : QUELS RAYONS EXPLIQUENT L'ÉCART ?", st_)
    heads = ["Rayon", "CA", "Poids PGC", "Vs N-1 (%)", "Vs Budget (%)", "Taux de Marge",
             "Marge vs N-1 (pts)", "Contribution marge", "Statut"]
    for j, h in enumerate(heads, start=2):
        ws2.cell(row=r0, column=j, value=h)
    _header_row(ws2, r0, 2, 2 + len(heads) - 1, st_)
    PGC_T = f'{R["NIV"]},"DEPT_TOTAL",{R["DEP"]},"01 - PGC"'
    rr = r0 + 1
    first_c = rr
    for code, label in RAYON_LABELS.items():
        crit = f'{R["NIV"]},"PGC_RAYON_TOTAL",{R["RAY"]},"{code}"'
        ws2.cell(row=rr, column=2, value=label).font = st_["label"]
        ws2.cell(row=rr, column=3, value=f'=SUMIFS({R["CA"]},{crit})').number_format = FMT_CUR
        ws2.cell(row=rr, column=4, value=f'=C{rr}/SUMIFS({R["CA"]},{PGC_T})').number_format = FMT_PCT
        ws2.cell(row=rr, column=5, value=f'=C{rr}/SUMIFS({R["CAN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=6, value=f'=C{rr}/SUMIFS({R["BUD"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=7, value=f'=SUMIFS({R["MG"]},{crit})/C{rr}').number_format = FMT_PCT
        ws2.cell(row=rr, column=8,
                 value=f'=(G{rr}-SUMIFS({R["MGN1"]},{crit})/SUMIFS({R["CAN1"]},{crit}))*100').number_format = FMT_PTS
        # SPEC-17: share of PGC margin, to be read against share of PGC sales
        ws2.cell(row=rr, column=9,
                 value=f'=SUMIFS({R["MG"]},{crit})/SUMIFS({R["MG"]},{PGC_T})').number_format = FMT_PCT
        ws2.cell(row=rr, column=10, value=(
            f'=IF(AND(E{rr}>={SEUIL_CROISSANCE_VERTE},F{rr}>=0,H{rr}>=0),"🟢 Performant",'
            f'IF(OR(E{rr}<0,F{rr}<{SEUIL_BUDGET_ROUGE},H{rr}<{SEUIL_MARGE_ROUGE_PTS}),'
            f'"🔴 Sous tendance","🟠 À surveiller"))'))
        _data_row(ws2, rr, 2, 10, st_)
        rr += 1
    _status_cf(ws2, 10, first_c, rr - 1)
    ws2.cell(row=rr, column=2, value=(
        "Lecture : comparer « Poids PGC » et « Contribution marge ». Un rayon dont la "
        "contribution marge est très inférieure à son poids CA tire la rentabilité vers le bas.")).font = st_["note"]
    ws2.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=10)
    ws2.row_dimensions[rr].height = 26
    last_c = rr

    r0 = _section(ws2, last_c + 2, "D — FORMATS DE MAGASIN", st_,
                  "Le taux de marge est affiché en niveau : un écart durable entre formats relève du modèle, pas de la journée.")
    heads = ["Format", "CA PGC", "Poids PGC", "Croissance vs N-1", "Écart vs PGC réseau",
             "Taux de Marge (niveau)", "Statut"]
    for j, h in enumerate(heads, start=2):
        ws2.cell(row=r0, column=j, value=h)
    _header_row(ws2, r0, 2, 2 + len(heads) - 1, st_)
    rr = r0 + 1
    first_d = rr
    pgc_growth = f'(SUMIFS({R["CA"]},{PGC_T})/SUMIFS({R["CAN1"]},{PGC_T})-1)'
    for fmt in FORMATS:
        crit = f'{R["NIV"]},"PGC_SITE",{R["FMT"]},"{fmt}"'
        ws2.cell(row=rr, column=2, value=fmt).font = st_["label"]
        ws2.cell(row=rr, column=3, value=f'=SUMIFS({R["CA"]},{crit})').number_format = FMT_CUR
        ws2.cell(row=rr, column=4, value=f'=C{rr}/SUMIFS({R["CA"]},{PGC_T})').number_format = FMT_PCT
        ws2.cell(row=rr, column=5, value=f'=C{rr}/SUMIFS({R["CAN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=6, value=f"=(E{rr}-{pgc_growth})*100").number_format = FMT_PTS
        ws2.cell(row=rr, column=7, value=f'=SUMIFS({R["MG"]},{crit})/C{rr}').number_format = FMT_PCT
        ws2.cell(row=rr, column=8, value=(
            f'=IF(F{rr}>={SEUIL_ECART_FORMAT_PTS},"🟢 Performant",'
            f'IF(F{rr}<-{SEUIL_ECART_FORMAT_PTS},"🔴 Sous tendance","🟠 À surveiller"))'))
        _data_row(ws2, rr, 2, 8, st_)
        rr += 1
    _status_cf(ws2, 8, first_d, rr - 1)

    # Trafic & panier par rayon
    r0 = _section(ws2, rr + 2, "G — TRAFIC & PANIER PAR RAYON", st_,
                  "D'où vient la variation : plus de clients (débit) ou plus de dépense par client (panier) ?")
    heads = ["Rayon", "Débit (clients)", "Vs N-1 Débit", "Panier Moyen", "Vs N-1 Panier",
             "Panier Qté", "Vs N-1 Qté", "Effet dominant"]
    for j, h in enumerate(heads, start=2):
        ws2.cell(row=r0, column=j, value=h)
    _header_row(ws2, r0, 2, 2 + len(heads) - 1, st_)
    rr = r0 + 1
    for code, label in RAYON_LABELS.items():
        crit = f'{R["NIV"]},"PGC_RAYON_TOTAL",{R["RAY"]},"{code}"'
        ws2.cell(row=rr, column=2, value=label).font = st_["label"]
        ws2.cell(row=rr, column=3, value=f'=SUMIFS({R["DB"]},{crit})').number_format = FMT_NUM
        ws2.cell(row=rr, column=4, value=f'=C{rr}/SUMIFS({R["DBN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=5, value=f'=SUMIFS({R["PAN"]},{crit})').number_format = FMT_CUR
        ws2.cell(row=rr, column=6, value=f'=E{rr}/SUMIFS({R["PANN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=7, value=f'=SUMIFS({R["PQ"]},{crit})').number_format = "0.0"
        ws2.cell(row=rr, column=8, value=f'=G{rr}/SUMIFS({R["PQN1"]},{crit})-1').number_format = FMT_PCT_SIGN
        ws2.cell(row=rr, column=9, value=(
            f'=IF(ABS(D{rr})>=ABS(F{rr})*{RATIO_DOMINANCE},IF(D{rr}>=0,"Trafic ⬆","Trafic ⬇"),'
            f'IF(ABS(F{rr})>=ABS(D{rr})*{RATIO_DOMINANCE},'
            f'IF(ABS(H{rr})>=ABS(F{rr}-H{rr})*{RATIO_DOMINANCE},IF(H{rr}>=0,"Quantité ⬆","Quantité ⬇"),'
            f'IF(F{rr}>=0,"Prix/mix ⬆","Prix/mix ⬇")),"Mixte"))'))
        _data_row(ws2, rr, 2, 9, st_)
        rr += 1

    ws2.freeze_panes = "B5"

    # ---------- 03_POINTS_CHAUDS ----------
    ws3 = wb.create_sheet("03_POINTS_CHAUDS")
    ws3.sheet_view.showGridLines = False
    _autosize(ws3, [5, 20, 22, 15, 12, 14, 13, 30, 26, 10, 14, 12, 13])

    ws3["B2"] = "POINTS CHAUDS & PLAN D'ACTIONS"
    ws3["B2"].font = st_["title"]
    ws3["B3"] = f"{jour_label} · priorisé par marge totale à risque"
    ws3["B3"].font = st_["subtitle"]

    ws3["B5"] = "ANALYSE — généré, lecture seule"
    ws3["B5"].font = Font(name=FONT_NAME, size=9, bold=True, color=BLUE)
    ws3["I5"] = "SAISIE — à remplir en séance"
    ws3["I5"].font = Font(name=FONT_NAME, size=9, bold=True, color=ORANGE)

    heads = ["", "Rayon", "Site", "Marge à risque", "Tx marge", "Cause probable",
             "Owner suggéré", "Signal", "Action décidée", "Rupture O/N", "Owner", "Échéance", "Statut"]
    hr = 6
    for j, h in enumerate(heads, start=2):
        ws3.cell(row=hr, column=j, value=h)
    _header_row(ws3, hr, 2, 2 + len(heads) - 1, st_)

    rr = hr + 1
    has_star = False
    for _, row in hot.iterrows():
        star = "" if row["budget_dispo"] else " *"
        if star:
            has_star = True
        prefix = "⛔" if row["marge_negative"] else (row.get("nouveau", "") or "")
        ws3.cell(row=rr, column=2, value=prefix)
        ws3.cell(row=rr, column=3, value=row["RayonL"]).font = st_["label"]
        ws3.cell(row=rr, column=4, value=f'{row["SiteL"]}{star}')
        ws3.cell(row=rr, column=5, value=float(row["marge_risque"])).number_format = FMT_CUR
        ws3.cell(row=rr, column=6, value=float(row["TM_N"])).number_format = FMT_PCT
        ws3.cell(row=rr, column=7, value=row["cause"])
        ws3.cell(row=rr, column=8, value=row["owner"])
        ws3.cell(row=rr, column=9, value=row["signal"])
        ws3.cell(row=rr, column=14, value="À analyser")
        _data_row(ws3, rr, 2, 14, st_, saisie_from=10)
        if row["marge_negative"]:
            for c in range(2, 10):
                ws3.cell(row=rr, column=c).fill = FILLS["red"]
        ws3.cell(row=rr, column=14).fill = FILLS["orange"]
        rr += 1
    hot_last = rr - 1

    note_r = rr
    notes = ["Marge à risque = marge perdue par dilution du taux + marge perdue sur baisse de volume, en FCFA."]
    if has_star:
        notes.append("* Sites Supeco : budget non disponible dans l'export. Le score de priorisation "
                     "est normalisé sur les critères disponibles (N-1 et marge).")
    notes.append("Owner suggéré est une proposition à confirmer en séance : une baisse de trafic peut "
                 "aussi venir d'une rupture fournisseur.")
    for i, t in enumerate(notes):
        ws3.cell(row=note_r + i, column=2, value=t).font = st_["note"]
        ws3.merge_cells(start_row=note_r + i, start_column=2, end_row=note_r + i, end_column=14)
    after_notes = note_r + len(notes)

    r0 = _section(ws3, after_notes + 1, "🟢 TOP 3 PERFORMANCES — À RÉPLIQUER", st_, last_col=14)
    heads = ["Rayon", "Site", "CA", "Vs N-1", "Taux de marge", "Marge vs N-1 (pts)", "Effet dominant"]
    for j, h in enumerate(heads, start=3):
        ws3.cell(row=r0, column=j, value=h)
    _header_row(ws3, r0, 3, 3 + len(heads) - 1, st_)
    rr = r0 + 1
    for _, row in perf.iterrows():
        ws3.cell(row=rr, column=3, value=row["RayonL"]).font = st_["label"]
        ws3.cell(row=rr, column=4, value=row["SiteL"])
        ws3.cell(row=rr, column=5, value=float(row["CA"])).number_format = FMT_CUR
        ws3.cell(row=rr, column=6, value=float(row["g_ca"])).number_format = FMT_PCT_SIGN
        ws3.cell(row=rr, column=7, value=float(row["TM_N"])).number_format = FMT_PCT
        ws3.cell(row=rr, column=8, value=float(row["delta_tm_pts"])).number_format = FMT_PTS
        ws3.cell(row=rr, column=9, value=row["cause"])
        _data_row(ws3, rr, 3, 9, st_)
        for c in range(3, 10):
            ws3.cell(row=rr, column=c).fill = FILLS["green"]
        rr += 1

    ws3.freeze_panes = "C7"

    # ---------- 04_LECTURE ----------
    ws4 = wb.create_sheet("04_LECTURE")
    ws4.sheet_view.showGridLines = False
    _autosize(ws4, [3, 30, 84, 3])
    ws4["B2"] = "GUIDE DE LECTURE"
    ws4["B2"].font = st_["title"]

    guide = [
        ("Objet", "Support de COPIL journalier PGC : lire le point du jour, arbitrer les points chauds, décider en séance."),
        ("Périmètre", "12 sites (3 Hyper / 6 Market / 3 Supeco). Le détail Rayon/Site n'existe que pour le département PGC."),
        ("Nature des données", "Snapshot d'une journée, non cumulé. Les niveaux varient fortement selon le jour de semaine."),
        ("", ""),
        ("🟢 Performant", f"Croissance CA ≥ +{SEUIL_CROISSANCE_VERTE:.0%}, budget atteint, marge stable ou en hausse"),
        ("🟠 À surveiller", "Zone intermédiaire"),
        ("🔴 Sous tendance", f"Croissance CA < 0%, écart budget < {SEUIL_BUDGET_ROUGE:.0%}, ou marge < {SEUIL_MARGE_ROUGE_PTS:.0f} pt"),
        ("⛔ Marge négative", "Vente à perte : remonte en tête quel que soit le montant, traitement immédiat"),
        ("🆕", "Point chaud absent de l'export de la veille (si l'export J-1 a été fourni)"),
        ("", ""),
        ("Marge à risque", "Marge perdue par dilution du taux + marge perdue sur baisse de volume, en FCFA. Sert au classement."),
        ("Trafic ⬇", "Moins de clients passés en caisse sur ce rayon — piste disponibilité / exécution magasin"),
        ("Quantité ⬇", "Autant de clients, moins d'articles par panier — piste disponibilité / implantation"),
        ("Prix/mix ⬇", "Autant d'articles, moins chers — piste assortiment / prix, côté achats"),
        ("Mixte", "Trafic et panier bougent ensemble — à arbitrer en séance"),
        ("", ""),
        ("Sélection des points chauds", f"Top {TOP_PCT:.0%} par marge à risque (minimum {MIN_LIGNES} lignes), "
                                        f"plafonné à {MAX_PAR_SITE} lignes par site. Les marges négatives passent toujours."),
        ("Budget Supeco", "Non disponible dans l'export : le score est normalisé sur les critères disponibles."),
        ("Historique", "Les tendances et récurrences se consultent dans Power BI, pas dans ce fichier."),
    ]
    rr = 4
    for label, desc in guide:
        if label:
            ws4.cell(row=rr, column=2, value=label).font = st_["label"]
            ws4.cell(row=rr, column=3, value=desc).font = st_["val"]
            ws4.cell(row=rr, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    wsd.sheet_state = "hidden"
    wb._sheets = [ws, ws2, ws3, ws4, wsd]
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------------
# STREAMLIT PAGE
# ----------------------------------------------------------------------------

def main():
    st.set_page_config(page_title="COPIL Journalier PGC", page_icon="📊", layout="centered")

    st.markdown(
        """
        <style>
        .stApp { background-color: #F2F2F7; }
        .sbh-card { background:#FFFFFF; border-radius:14px; padding:20px 24px; margin-bottom:16px; }
        .sbh-title { font-size:26px; font-weight:600; color:#1C1C1E; margin-bottom:4px; }
        .sbh-sub { font-size:14px; color:#8E8E93; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sbh-card"><div class="sbh-title">COPIL Journalier PGC</div>'
        '<div class="sbh-sub">Génère le classeur de séance à partir de l\'export ventes du jour</div></div>',
        unsafe_allow_html=True,
    )

    f_jour = st.file_uploader("Export du jour (obligatoire)", type=["xlsx"], key="jour")
    f_veille = st.file_uploader(
        "Export de la veille (optionnel — marque les nouveaux points chauds 🆕)",
        type=["xlsx"], key="veille",
    )
    date_rapport = st.date_input("Date du rapport", value=date.today())

    if f_jour is None:
        st.info("Déposez l'export du jour pour générer le classeur.")
        return

    try:
        df = load_export(f_jour)
        sites = compute_site_metrics(df)
        hot = select_hot_spots(sites)
        perf = select_top_perf(sites)
    except Exception as exc:
        st.error(f"Lecture impossible : {exc}")
        st.caption("Vérifiez que le fichier est bien l'export BI standard (onglet « Export »).")
        return

    prev_hot = None
    if f_veille is not None:
        try:
            prev_hot = select_hot_spots(compute_site_metrics(load_export(f_veille)))
        except Exception:
            st.warning("Export de la veille illisible — le marquage 🆕 est désactivé.")
    hot = flag_new_lines(hot, prev_hot)

    jour_label = f"{JOURS_FR[date_rapport.weekday()]} {date_rapport.strftime('%d/%m/%Y')}"

    total = df[df["Niveau"] == "GRAND_TOTAL"]
    c1, c2, c3 = st.columns(3)
    if not total.empty:
        t = total.iloc[0]
        c1.metric("CA réseau", f"{t['CA']:,.0f} FCFA".replace(",", " "),
                  f"{t['CA'] / t['CA N-1'] - 1:+.1%}")
        c2.metric("Taux de marge", f"{t['Marge'] / t['CA']:.1%}")
    c3.metric("Points chauds", len(hot), f"{len(sites[sites['marge_negative']])} marge négative"
              if len(sites[sites["marge_negative"]]) else "aucune marge négative")

    if len(sites[sites["marge_negative"]]):
        for _, row in sites[sites["marge_negative"]].iterrows():
            st.error(f"⛔ {row['RayonL']} — {row['SiteL']} : taux de marge {row['TM_N']:.1%}")

    try:
        data = build_workbook(df, sites, hot, perf, jour_label)
    except Exception as exc:
        st.error(f"Génération impossible : {exc}")
        return

    st.download_button(
        "📥 Télécharger le classeur COPIL",
        data=data,
        file_name=f"COPIL_Journalier_PGC_{date_rapport.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    st.caption("Les formules se recalculent à l'ouverture dans Excel.")


main()
