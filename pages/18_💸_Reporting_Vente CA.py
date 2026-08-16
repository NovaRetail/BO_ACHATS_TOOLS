"""
Module 18 - Reporting Vente CA
================================
Reporting hebdomadaire de performance commerciale (Rayon / Format / Famille
/ Site), comparaison N vs N-1 ISO stricte, priorisation automatique P1/P2,
export Excel prêt pour saisie acheteur (Cause / Commentaire).

Fichier unique et autonome : aucune dépendance à un package externe.
Les seuils métier (P1/P2, drivers, qualité) sont pilotables directement
dans la barre latérale ("Paramètres avancés") - pas besoin de modifier le
code pour les ajuster.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("reporting_vente_ca")

st.set_page_config(page_title="Reporting Vente CA", page_icon="💸", layout="wide")

RESEAU = "RÉSEAU"
FORMATS = ["HYPER", "MARKET", "SUPECO"]
NA_LABEL = "N/D"

DEFAULT_FORMAT_MAPPING = [
    {"pattern": "hyper", "format": "HYPER"},
    {"pattern": "market", "format": "MARKET"},
    {"pattern": "supeco", "format": "SUPECO"},
]
DEFAULT_RAYON_ORDER = ["BOISSONS", "EPICERIE", "PARFUMERIE HYGIENE", "DROGUERIE"]
DEFAULT_RAYON_LABEL_OVERRIDES = {"BOISSON": "BOISSONS"}
CAUSE_OPTIONS = [
    "Rupture", "Promo", "Prix", "Assortiment", "Fournisseur",
    "Saisonnalité", "Mix", "Concurrence", "Disponibilité", "Autre",
]

EXCEL_COLORS = {
    "title_bg": "1B2A4A", "title_font": "FFFFFF", "header_bg": "1B2A4A", "header_font": "FFFFFF",
    "band_bg": "E9ECF2", "input_cell_bg": "FFF6CC",
    "success_bg": "E3F1E1", "success_font": "1E7A34",
    "danger_bg": "FBE4E2", "danger_font": "B3261E",
    "warning_bg": "FDEEDD", "warning_font": "9C5700",
    "p1_bg": "FBE4E2", "p1_font": "B3261E", "p2_bg": "FDEEDD", "p2_font": "9C5700",
    "databar_color": "8FA8C8",
}
EXCEL_NUMBER_FORMATS = {
    "montant": '#,##0,, "M";-#,##0,, "M"',
    "montant_ecart": '+#,##0,, "M";-#,##0,, "M";"0 M"',
    "pct": '+0.0%;-0.0%;"0,0%"',
    "pct_plain": "0.0%",
    "pts": '+0.00" pt";-0.00" pt";"0,00 pt"',
    "qty": "#,##0",
}


# =============================================================================
# 1. CONFIGURATION (seuils métier - alimentés par la sidebar)
# =============================================================================

@dataclass(frozen=True)
class PrioritizationConfig:
    p1_top_n_gap_abs: int = 3
    p1_min_weight_pct: float = 0.10
    p1_significant_evo_pct: float = 0.05
    p1_margin_degradation_pts: float = 2.0
    p2_ca_evo_pct: float = 0.10
    p2_margin_delta_pts: float = 1.0
    p2_promo_weight_delta_pts: float = 3.0
    p2_min_contribution_pct: float = 0.05


@dataclass(frozen=True)
class DriversConfig:
    volume_vs_ca_tolerance_pts: float = 3.0
    price_mix_gap_pts: float = 5.0
    margin_up_pts: float = 1.0
    margin_down_pts: float = -1.0
    promo_up_pts: float = 3.0
    promo_down_pts: float = -3.0


@dataclass(frozen=True)
class QualityConfig:
    ca_ok_tolerance_pct: float = 0.03
    ca_partial_tolerance_pct: float = 0.05


def render_sidebar_config() -> tuple[PrioritizationConfig, DriversConfig, QualityConfig, int]:
    """Affiche les contrôles de seuils métier dans la sidebar et retourne la config construite."""
    with st.sidebar.expander("⚙️ Paramètres avancés (seuils métier)", expanded=False):
        st.caption("Priorisation P1")
        p1_top_n = st.number_input("Top N écart CA absolu -> P1", min_value=1, max_value=10, value=3, step=1)
        p1_weight = st.slider("Poids rayon minimum (%) pour P1 sur évolution", 0, 100, 10) / 100
        p1_evo = st.slider("Évolution CA minimum (%) pour P1", 0, 50, 5) / 100
        p1_margin = st.slider("Dégradation taux de marge (pts) -> P1", 0.0, 10.0, 2.0, step=0.5)

        st.divider()
        st.caption("Priorisation P2")
        p2_ca_evo = st.slider("Évolution CA (%) -> P2", 0, 50, 10) / 100
        p2_margin = st.slider("Δ taux de marge (pts) -> P2", 0.0, 10.0, 1.0, step=0.5)
        p2_promo = st.slider("Δ poids promo (pts) -> P2", 0.0, 20.0, 3.0, step=0.5)
        p2_contrib = st.slider("Contribution minimum (%) -> P2", 0, 50, 5) / 100

        st.divider()
        st.caption("Signaux drivers (Volume / Prix-Mix / Marge / Promo)")
        d_vol_tol = st.slider("Tolérance Qté vs CA (pts) -> 'Volume'", 0.0, 10.0, 3.0, step=0.5)
        d_price_mix = st.slider("Écart Qté/CA (pts) -> 'Prix/Mix'", 0.0, 20.0, 5.0, step=0.5)
        d_margin_up = st.slider("Δ Tx marge (pts) -> 'Marge ↑'", 0.0, 10.0, 1.0, step=0.5)
        d_margin_down = -st.slider("Δ Tx marge (pts) -> 'Marge ↓'", 0.0, 10.0, 1.0, step=0.5)
        d_promo_up = st.slider("Δ Promo (pts) -> 'Promo ↑'", 0.0, 20.0, 3.0, step=0.5)
        d_promo_down = -st.slider("Δ Promo (pts) -> 'Promo ↓'", 0.0, 20.0, 3.0, step=0.5)

        st.divider()
        st.caption("Contrôle qualité")
        q_ok = st.slider("Tolérance réconciliation CA - OK (%)", 0, 20, 3) / 100
        q_partial = st.slider("Tolérance réconciliation CA - Partiel (%)", 0, 30, 5) / 100

        st.divider()
        top_site_famille = st.number_input("Nombre de lignes Top Site x Famille", min_value=5, max_value=50, value=12, step=1)

    prio_cfg = PrioritizationConfig(
        p1_top_n_gap_abs=p1_top_n, p1_min_weight_pct=p1_weight, p1_significant_evo_pct=p1_evo,
        p1_margin_degradation_pts=p1_margin, p2_ca_evo_pct=p2_ca_evo, p2_margin_delta_pts=p2_margin,
        p2_promo_weight_delta_pts=p2_promo, p2_min_contribution_pct=p2_contrib,
    )
    drivers_cfg = DriversConfig(
        volume_vs_ca_tolerance_pts=d_vol_tol, price_mix_gap_pts=d_price_mix,
        margin_up_pts=d_margin_up, margin_down_pts=d_margin_down,
        promo_up_pts=d_promo_up, promo_down_pts=d_promo_down,
    )
    quality_cfg = QualityConfig(ca_ok_tolerance_pct=q_ok, ca_partial_tolerance_pct=q_partial)
    return prio_cfg, drivers_cfg, quality_cfg, int(top_site_famille)


# =============================================================================
# 2. NORMALISATION (Rayon / Site / Format)
# =============================================================================

_CODE_LABEL_RE = re.compile(r"^\s*0*(\d+)\s*-\s*(.+?)\s*$")


@dataclass(frozen=True)
class RayonKey:
    code: int
    raw_label: str


def parse_rayon(value) -> RayonKey | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "total":
        return None
    match = _CODE_LABEL_RE.match(text)
    if not match:
        return RayonKey(code=-1, raw_label=text.upper())
    code_str, label = match.groups()
    return RayonKey(code=int(code_str), raw_label=label.upper())


def clean_famille_label(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = _CODE_LABEL_RE.match(text)
    if match:
        return match.group(2).strip().upper()
    return text.upper()


def parse_site(value) -> tuple[str | None, str]:
    if value is None:
        return None, ""
    text = str(value).strip()
    if not text or text.lower() == "total":
        return None, ""
    if " - " in text:
        code, name = text.split(" - ", 1)
        return code.strip(), name.strip()
    return None, text


def detect_format(site_name: str, mapping: list[dict[str, str]] = DEFAULT_FORMAT_MAPPING) -> str | None:
    name_lower = site_name.lower()
    for rule in mapping:
        pattern = str(rule.get("pattern", "")).lower()
        if pattern and pattern in name_lower:
            return str(rule.get("format", "")).upper()
    return None


# =============================================================================
# 3. LECTURE DES EXPORTS PBI
# =============================================================================

_FILTER_MARKER = "filtres appliqués"
_DATE_RANGE_RE = re.compile(r"Date est le ou après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})")


@dataclass
class ExportPeriod:
    date_start: date | None = None
    date_end: date | None = None
    iso_week: int | None = None
    iso_year: int | None = None

    @property
    def label(self) -> str:
        if self.date_start and self.date_end:
            return f"S{self.iso_week:02d}-{self.iso_year} ({self.date_start:%d/%m/%Y} - {self.date_end:%d/%m/%Y})"
        return "Période non détectée"


def _parse_period(filter_text: str) -> ExportPeriod:
    period = ExportPeriod()
    match = _DATE_RANGE_RE.search(filter_text)
    if match:
        start_str, end_str = match.groups()
        d_start = pd.to_datetime(start_str, format="%d/%m/%Y").date()
        d_end = pd.to_datetime(end_str, format="%d/%m/%Y").date() - pd.Timedelta(days=1)
        iso = d_start.isocalendar()
        period.date_start, period.date_end = d_start, d_end
        period.iso_year, period.iso_week = iso[0], iso[1]
    return period


def read_export_bytes(file_bytes: bytes) -> tuple[pd.DataFrame, ExportPeriod]:
    """Lit un export PBI depuis des bytes : sépare détail / total, parse la période du footer."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    ncols = len(header)
    filter_text = ""
    data_rows: list[tuple] = []

    for row in rows[1:]:
        row = list(row) + [None] * (ncols - len(row))
        first_cell = row[0]
        if isinstance(first_cell, str) and _FILTER_MARKER in first_cell.lower():
            filter_text = first_cell
            continue
        if all(c is None for c in row):
            continue
        if isinstance(first_cell, str) and first_cell.strip().lower() == "total":
            continue  # grand total (redondant avec les sous-totaux, non utilisé)
        data_rows.append(tuple(row))

    df = pd.DataFrame(data_rows, columns=header)
    return df, _parse_period(filter_text)


# =============================================================================
# 4. ENRICHISSEMENT + CLASSIFICATION DES LIGNES (détail vs sous-total)
# =============================================================================

ROW_DETAIL, ROW_RAYON_SUBTOTAL, ROW_FAMILLE_SUBTOTAL, ROW_GRAND_TOTAL = (
    "detail", "rayon_subtotal", "famille_subtotal", "grand_total",
)


def enrich_site_export(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    rayon_keys = out["Rayon"].map(parse_rayon)
    out["rayon_code"] = rayon_keys.map(lambda k: k.code if k else None)
    out["rayon_label"] = rayon_keys.map(lambda k: k.raw_label if k else None)
    site_parsed = out["Site"].map(parse_site)
    out["site_code"] = site_parsed.map(lambda t: t[0])
    out["site_name"] = site_parsed.map(lambda t: t[1])
    out["format"] = out["site_name"].map(lambda n: detect_format(n) if n else None)

    def _classify(row):
        if pd.isna(row["rayon_code"]):
            return ROW_GRAND_TOTAL
        if pd.isna(row["site_code"]):
            return ROW_RAYON_SUBTOTAL
        return ROW_DETAIL

    out["row_type"] = out.apply(_classify, axis=1)
    return out


def enrich_detail_export(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    rayon_keys = out["Rayon"].map(parse_rayon)
    out["rayon_code"] = rayon_keys.map(lambda k: k.code if k else None)
    out["rayon_label"] = rayon_keys.map(lambda k: k.raw_label if k else None)

    famille_raw = out["Famille"]
    is_famille_total = famille_raw.astype(str).str.strip().str.lower() == "total"
    out["famille_label"] = famille_raw.where(~is_famille_total & famille_raw.notna(), None).map(clean_famille_label)
    out["is_famille_total"] = is_famille_total

    site_parsed = out["Site nom long"].map(parse_site)
    out["site_code"] = site_parsed.map(lambda t: t[0])
    out["site_name"] = site_parsed.map(lambda t: t[1])
    out["format"] = out["site_name"].map(lambda n: detect_format(n) if n else None)

    def _classify(row):
        if pd.isna(row["rayon_code"]):
            return ROW_GRAND_TOTAL
        if row["is_famille_total"]:
            return ROW_RAYON_SUBTOTAL
        if pd.isna(row["site_code"]):
            return ROW_FAMILLE_SUBTOTAL
        return ROW_DETAIL

    out["row_type"] = out.apply(_classify, axis=1)
    return out


def detail_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["row_type"] == ROW_DETAIL].copy()


# =============================================================================
# 5. CONTRÔLES QUALITÉ (périmètre sites, réconciliation CA)
# =============================================================================

@dataclass
class SiteCoverage:
    sites_ref: set
    sites_site_export: set

    @property
    def missing(self) -> set:
        return self.sites_ref - self.sites_site_export

    @property
    def coverage_pct(self) -> float:
        n_ref = len(self.sites_ref)
        if not n_ref:
            return 1.0
        return len(self.sites_ref & self.sites_site_export) / n_ref

    @property
    def label(self) -> str:
        return f"{len(self.sites_ref & self.sites_site_export)}/{len(self.sites_ref)}"


@dataclass
class CaReconciliation:
    ca_site: float
    ca_detail: float
    ratio: float
    status: str


def _classify_alignment(ratio: float, cfg: QualityConfig) -> str:
    dev = abs(ratio - 1.0)
    if dev <= cfg.ca_ok_tolerance_pct:
        return "OK"
    if dev <= cfg.ca_partial_tolerance_pct:
        return "PARTIEL"
    return "A VERIFIER"


@dataclass
class QualityReport:
    coverage_global: SiteCoverage
    coverage_by_format: dict[str, SiteCoverage]
    recon_global: CaReconciliation
    recon_by_format: dict[str, CaReconciliation]

    @property
    def has_perimeter_issue(self) -> bool:
        return self.coverage_global.coverage_pct < 1.0

    @property
    def summary_line(self) -> str:
        return (
            f"Export Site = {self.coverage_global.label} magasins | "
            f"Alignement CA sur magasins communs = {self.recon_global.ratio:.1%} ({self.recon_global.status})"
        )


def build_quality_report(site_d, n_d, n1_d, cfg: QualityConfig) -> QualityReport:
    ref_sites = set(n_d["site_code"].dropna()) | set(n1_d["site_code"].dropna())
    site_sites = set(site_d["site_code"].dropna())
    coverage_global = SiteCoverage(sites_ref=ref_sites, sites_site_export=site_sites)

    coverage_by_format = {}
    for fmt in FORMATS:
        ref_fmt = set(n_d.loc[n_d["format"] == fmt, "site_code"].dropna()) | set(
            n1_d.loc[n1_d["format"] == fmt, "site_code"].dropna())
        site_fmt = set(site_d.loc[site_d["format"] == fmt, "site_code"].dropna())
        coverage_by_format[fmt] = SiteCoverage(sites_ref=ref_fmt, sites_site_export=site_fmt)

    def _reconcile(common, s_df, n_df):
        ca_site = pd.to_numeric(s_df.loc[s_df["site_code"].isin(common), "CA"], errors="coerce").sum()
        ca_det = pd.to_numeric(n_df.loc[n_df["site_code"].isin(common), "CA"], errors="coerce").sum()
        ratio = (ca_site / ca_det) if ca_det else 0.0
        status = _classify_alignment(ratio, cfg) if ca_det else "A VERIFIER"
        return CaReconciliation(ca_site=float(ca_site), ca_detail=float(ca_det), ratio=float(ratio), status=status)

    common_global = ref_sites & site_sites
    recon_global = _reconcile(common_global, site_d, n_d)

    recon_by_format = {}
    for fmt in FORMATS:
        common_fmt = coverage_by_format[fmt].sites_ref & coverage_by_format[fmt].sites_site_export
        recon_by_format[fmt] = _reconcile(common_fmt, site_d[site_d["format"] == fmt], n_d[n_d["format"] == fmt])

    return QualityReport(coverage_global, coverage_by_format, recon_global, recon_by_format)


# =============================================================================
# 6. CALCULS MÉTIER (agrégations, N vs N-1 ISO, contribution, taux de marge)
# =============================================================================

DETAIL_METRIC_COLS = {"CA": "CA", "Marge": "Marge", "Qte": "Qté Vente", "CAPromo": "CA Promo"}


def safe_div(num, den):
    if num is None or den is None or pd.isna(num) or pd.isna(den) or den == 0:
        return None
    return num / den


def pct_change(current, previous):
    if current is None or previous is None or pd.isna(current) or pd.isna(previous):
        return None
    return safe_div(current - previous, previous)


def margin_rate(marge, ca):
    return safe_div(marge, ca)


def pts_change(cur_rate, prev_rate):
    if cur_rate is None or prev_rate is None or pd.isna(cur_rate) or pd.isna(prev_rate):
        return None
    return cur_rate - prev_rate


def contribution(gap_item, gap_total, epsilon=1.0):
    if gap_item is None or gap_total is None or pd.isna(gap_item) or pd.isna(gap_total):
        return None
    if abs(gap_total) < epsilon:
        return None
    return gap_item / gap_total


def _sum_metrics(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    work = df.copy()
    for col in DETAIL_METRIC_COLS.values():
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce")
    agg_map = {out: (src, "sum") for out, src in DETAIL_METRIC_COLS.items() if src in work.columns}
    return work.groupby(group_cols, dropna=False).agg(**agg_map).reset_index()


@dataclass
class RayonFormatKpi:
    rayon_code: int
    rayon_label: str
    scope: str
    ca_n: float
    ca_gap_fcfa: float
    ca_evo_pct: float | None
    budget: float | None
    budget_available: bool
    ecart_budget_pct: float | None
    poids_n: float | None
    poids_evo_pts: float | None
    marge_n: float
    marge_evo_pct: float | None
    tx_marge_n: float | None
    tx_marge_evo_pts: float | None
    qte_n: float
    qte_evo_pct: float | None
    poids_promo_n: float | None
    poids_promo_evo_pts: float | None
    debit_n: float | None
    debit_available: bool
    panier_n: float | None
    site_coverage_pct: float


def build_rayon_format_kpis(n_d, n1_d, site_d, coverage_by_scope: dict[str, float]) -> list[RayonFormatKpi]:
    results = []
    rayon_keys = list(n_d[["rayon_code", "rayon_label"]].drop_duplicates().itertuples(index=False))
    scopes = [RESEAU] + FORMATS

    for scope in scopes:
        n_scope = n_d if scope == RESEAU else n_d[n_d["format"] == scope]
        n1_scope = n1_d if scope == RESEAU else n1_d[n1_d["format"] == scope]
        site_scope = site_d if scope == RESEAU else site_d[site_d["format"] == scope]

        n_by_rayon = _sum_metrics(n_scope, ["rayon_code", "rayon_label"]).set_index("rayon_code")
        n1_by_rayon = _sum_metrics(n1_scope, ["rayon_code", "rayon_label"]).set_index("rayon_code")
        budget_by_rayon = (
            site_scope.assign(Budget=pd.to_numeric(site_scope.get("Budget"), errors="coerce"))
            .groupby("rayon_code")["Budget"].sum() if not site_scope.empty else pd.Series(dtype=float)
        )
        debit_by_rayon = (
            site_scope.assign(Debit=pd.to_numeric(site_scope.get("Débit"), errors="coerce"))
            .groupby("rayon_code")["Debit"].sum() if not site_scope.empty else pd.Series(dtype=float)
        )
        ca_total_n = n_by_rayon["CA"].sum()
        ca_total_n1 = n1_by_rayon["CA"].sum()
        coverage_pct = coverage_by_scope.get(scope, 0.0)
        budget_available = coverage_pct >= 1.0

        for rk in rayon_keys:
            code, label = rk.rayon_code, rk.rayon_label
            row_n = n_by_rayon.loc[code] if code in n_by_rayon.index else None
            row_n1 = n1_by_rayon.loc[code] if code in n1_by_rayon.index else None

            ca_n = float(row_n["CA"]) if row_n is not None else 0.0
            ca_n1 = float(row_n1["CA"]) if row_n1 is not None else 0.0
            marge_n = float(row_n["Marge"]) if row_n is not None else 0.0
            marge_n1 = float(row_n1["Marge"]) if row_n1 is not None else 0.0
            qte_n = float(row_n["Qte"]) if row_n is not None else 0.0
            qte_n1 = float(row_n1["Qte"]) if row_n1 is not None else 0.0
            promo_n = float(row_n["CAPromo"]) if row_n is not None else 0.0
            promo_n1 = float(row_n1["CAPromo"]) if row_n1 is not None else 0.0

            budget = float(budget_by_rayon.get(code)) if code in budget_by_rayon.index else None
            debit = float(debit_by_rayon.get(code)) if code in debit_by_rayon.index else None

            tx_marge_n = margin_rate(marge_n, ca_n)
            tx_marge_n1 = margin_rate(marge_n1, ca_n1)
            poids_n = safe_div(ca_n, ca_total_n)
            poids_n1 = safe_div(ca_n1, ca_total_n1)
            promo_pct_n = safe_div(promo_n, ca_n)
            promo_pct_n1 = safe_div(promo_n1, ca_n1)

            results.append(RayonFormatKpi(
                rayon_code=code, rayon_label=label, scope=scope,
                ca_n=ca_n, ca_gap_fcfa=ca_n - ca_n1, ca_evo_pct=pct_change(ca_n, ca_n1),
                budget=budget if budget_available else None, budget_available=budget_available,
                ecart_budget_pct=pct_change(ca_n, budget) if (budget_available and budget) else None,
                poids_n=poids_n, poids_evo_pts=pts_change(poids_n, poids_n1),
                marge_n=marge_n, marge_evo_pct=pct_change(marge_n, marge_n1),
                tx_marge_n=tx_marge_n, tx_marge_evo_pts=pts_change(tx_marge_n, tx_marge_n1),
                qte_n=qte_n, qte_evo_pct=pct_change(qte_n, qte_n1),
                poids_promo_n=promo_pct_n, poids_promo_evo_pts=pts_change(promo_pct_n, promo_pct_n1),
                debit_n=debit, debit_available=budget_available,
                panier_n=safe_div(ca_n, debit) if debit else None,
                site_coverage_pct=coverage_pct,
            ))
    return results


@dataclass
class FamilleKpi:
    rayon_code: int
    rayon_label: str
    famille_label: str
    ca_n: float
    gap_ca_fcfa: float
    evo_pct: float | None
    poids_rayon: float | None
    contribution_pct: float | None
    tx_marge_n: float | None
    delta_tx_marge_pts: float | None
    qte_evo_pct: float | None
    poids_promo_n: float | None
    delta_promo_pts: float | None
    hyper_evo_pct: float | None
    market_evo_pct: float | None
    supeco_evo_pct: float | None
    signal: str = ""
    site_cle: str = ""
    priority: str = ""


def build_famille_kpis(n_d, n1_d) -> list[FamilleKpi]:
    results = []
    group_cols = ["rayon_code", "rayon_label", "famille_label"]
    n_by_f = _sum_metrics(n_d, group_cols).set_index(["rayon_code", "famille_label"])
    n1_by_f = _sum_metrics(n1_d, group_cols).set_index(["rayon_code", "famille_label"])
    n_total = n_d.groupby("rayon_code").apply(lambda g: pd.to_numeric(g["CA"], errors="coerce").sum())
    n1_total = n1_d.groupby("rayon_code").apply(lambda g: pd.to_numeric(g["CA"], errors="coerce").sum())
    all_keys = set(n_by_f.index) | set(n1_by_f.index)

    def _format_evo(rayon_code, famille_label, fmt):
        nf = n_d[(n_d["rayon_code"] == rayon_code) & (n_d["famille_label"] == famille_label) & (n_d["format"] == fmt)]
        n1f = n1_d[(n1_d["rayon_code"] == rayon_code) & (n1_d["famille_label"] == famille_label) & (n1_d["format"] == fmt)]
        ca_n = pd.to_numeric(nf["CA"], errors="coerce").sum() if not nf.empty else 0.0
        ca_n1 = pd.to_numeric(n1f["CA"], errors="coerce").sum() if not n1f.empty else 0.0
        if ca_n == 0 and ca_n1 == 0:
            return None
        return pct_change(ca_n, ca_n1)

    for rayon_code, famille_label in sorted(all_keys, key=lambda k: (k[0], k[1] or "")):
        row_n = n_by_f.loc[(rayon_code, famille_label)] if (rayon_code, famille_label) in n_by_f.index else None
        row_n1 = n1_by_f.loc[(rayon_code, famille_label)] if (rayon_code, famille_label) in n1_by_f.index else None
        rayon_label = row_n["rayon_label"] if row_n is not None else row_n1["rayon_label"]

        ca_n = float(row_n["CA"]) if row_n is not None else 0.0
        ca_n1 = float(row_n1["CA"]) if row_n1 is not None else 0.0
        marge_n = float(row_n["Marge"]) if row_n is not None else 0.0
        marge_n1 = float(row_n1["Marge"]) if row_n1 is not None else 0.0
        qte_n = float(row_n["Qte"]) if row_n is not None else 0.0
        qte_n1 = float(row_n1["Qte"]) if row_n1 is not None else 0.0
        promo_n = float(row_n["CAPromo"]) if row_n is not None else 0.0
        promo_n1 = float(row_n1["CAPromo"]) if row_n1 is not None else 0.0

        gap_ca = ca_n - ca_n1
        total_n = float(n_total.get(rayon_code, 0.0))
        total_n1 = float(n1_total.get(rayon_code, 0.0))
        gap_total = total_n - total_n1

        tx_marge_n = margin_rate(marge_n, ca_n)
        tx_marge_n1 = margin_rate(marge_n1, ca_n1)
        promo_pct_n = safe_div(promo_n, ca_n)
        promo_pct_n1 = safe_div(promo_n1, ca_n1)

        results.append(FamilleKpi(
            rayon_code=rayon_code, rayon_label=rayon_label, famille_label=famille_label or "(non renseigné)",
            ca_n=ca_n, gap_ca_fcfa=gap_ca, evo_pct=pct_change(ca_n, ca_n1),
            poids_rayon=safe_div(ca_n, total_n), contribution_pct=contribution(gap_ca, gap_total),
            tx_marge_n=tx_marge_n, delta_tx_marge_pts=pts_change(tx_marge_n, tx_marge_n1),
            qte_evo_pct=pct_change(qte_n, qte_n1),
            poids_promo_n=promo_pct_n, delta_promo_pts=pts_change(promo_pct_n, promo_pct_n1),
            hyper_evo_pct=_format_evo(rayon_code, famille_label, "HYPER"),
            market_evo_pct=_format_evo(rayon_code, famille_label, "MARKET"),
            supeco_evo_pct=_format_evo(rayon_code, famille_label, "SUPECO"),
        ))
    return results


@dataclass
class SiteFamilleKpi:
    rayon_code: int
    format: str
    site_name: str
    famille_label: str
    gap_ca_fcfa: float
    evo_pct: float | None
    qte_evo_pct: float | None
    delta_tx_marge_pts: float | None
    delta_promo_pts: float | None
    signal: str = ""


def build_site_famille_kpis(n_d, n1_d) -> list[SiteFamilleKpi]:
    group_cols = ["rayon_code", "format", "site_name", "famille_label"]
    n_g = _sum_metrics(n_d, group_cols).set_index(group_cols)
    n1_g = _sum_metrics(n1_d, group_cols).set_index(group_cols)
    all_keys = set(n_g.index) | set(n1_g.index)
    results = []
    for key in all_keys:
        rayon_code, fmt, site_name, famille_label = key
        row_n = n_g.loc[key] if key in n_g.index else None
        row_n1 = n1_g.loc[key] if key in n1_g.index else None
        ca_n = float(row_n["CA"]) if row_n is not None else 0.0
        ca_n1 = float(row_n1["CA"]) if row_n1 is not None else 0.0
        marge_n = float(row_n["Marge"]) if row_n is not None else 0.0
        marge_n1 = float(row_n1["Marge"]) if row_n1 is not None else 0.0
        qte_n = float(row_n["Qte"]) if row_n is not None else 0.0
        qte_n1 = float(row_n1["Qte"]) if row_n1 is not None else 0.0
        promo_n = float(row_n["CAPromo"]) if row_n is not None else 0.0
        promo_n1 = float(row_n1["CAPromo"]) if row_n1 is not None else 0.0
        tx_marge_n = margin_rate(marge_n, ca_n)
        tx_marge_n1 = margin_rate(marge_n1, ca_n1)
        promo_pct_n = safe_div(promo_n, ca_n)
        promo_pct_n1 = safe_div(promo_n1, ca_n1)
        results.append(SiteFamilleKpi(
            rayon_code=rayon_code, format=fmt or "", site_name=site_name or "",
            famille_label=famille_label or "(non renseigné)", gap_ca_fcfa=ca_n - ca_n1,
            evo_pct=pct_change(ca_n, ca_n1), qte_evo_pct=pct_change(qte_n, qte_n1),
            delta_tx_marge_pts=pts_change(tx_marge_n, tx_marge_n1),
            delta_promo_pts=pts_change(promo_pct_n, promo_pct_n1),
        ))
    return results


# =============================================================================
# 7. DRIVERS (Volume / Prix-Mix / Marge / Promo)
# =============================================================================

def _volume_price_mix_signal(ca_evo, qte_evo, cfg: DriversConfig):
    if ca_evo is None or qte_evo is None:
        return None
    gap_pts = (qte_evo - ca_evo) * 100
    if abs(gap_pts) <= cfg.volume_vs_ca_tolerance_pts:
        return "Volume"
    if gap_pts > cfg.price_mix_gap_pts:
        return "Volume ↑ / Prix-Mix ↓"
    if gap_pts < -cfg.price_mix_gap_pts:
        return "Prix/Mix +" if ca_evo > 0 else "Prix/Mix -"
    return None


def _margin_signal(delta_pts, cfg: DriversConfig):
    if delta_pts is None:
        return None
    pts = delta_pts * 100
    if pts <= cfg.margin_down_pts:
        return "Marge ↓"
    if pts >= cfg.margin_up_pts:
        return "Marge ↑"
    return None


def _promo_signal(delta_pts, cfg: DriversConfig):
    if delta_pts is None:
        return None
    pts = delta_pts * 100
    if pts >= cfg.promo_up_pts:
        return "Promo ↑"
    if pts <= cfg.promo_down_pts:
        return "Promo ↓"
    return None


def compute_signal(ca_evo, qte_evo, delta_marge_pts, delta_promo_pts, cfg: DriversConfig) -> str:
    parts = [
        _volume_price_mix_signal(ca_evo, qte_evo, cfg),
        _margin_signal(delta_marge_pts, cfg),
        _promo_signal(delta_promo_pts, cfg),
    ]
    parts = [p for p in parts if p]
    return " | ".join(parts) if parts else "-"


def apply_signals_famille(familles: list[FamilleKpi], cfg: DriversConfig) -> None:
    for f in familles:
        f.signal = compute_signal(f.evo_pct, f.qte_evo_pct, f.delta_tx_marge_pts, f.delta_promo_pts, cfg)


def apply_signals_site_famille(rows: list[SiteFamilleKpi], cfg: DriversConfig) -> None:
    for r in rows:
        r.signal = compute_signal(r.evo_pct, r.qte_evo_pct, r.delta_tx_marge_pts, r.delta_promo_pts, cfg)


# =============================================================================
# 8. PRIORISATION P1 / P2 + SITE CLÉ
# =============================================================================

P1, P2 = "P1", "P2"


def _is_p1(f: FamilleKpi, cfg: PrioritizationConfig, top_n_codes: set) -> bool:
    key = f"{f.rayon_code}|{f.famille_label}"
    if key in top_n_codes:
        return True
    if f.poids_rayon is not None and f.poids_rayon >= cfg.p1_min_weight_pct and f.evo_pct is not None and abs(f.evo_pct) >= cfg.p1_significant_evo_pct:
        return True
    if f.delta_tx_marge_pts is not None and f.delta_tx_marge_pts <= -(cfg.p1_margin_degradation_pts / 100):
        return True
    return False


def _is_p2(f: FamilleKpi, cfg: PrioritizationConfig) -> bool:
    if f.evo_pct is not None and abs(f.evo_pct) >= cfg.p2_ca_evo_pct:
        return True
    if f.delta_tx_marge_pts is not None and abs(f.delta_tx_marge_pts) >= (cfg.p2_margin_delta_pts / 100):
        return True
    if f.delta_promo_pts is not None and abs(f.delta_promo_pts) >= (cfg.p2_promo_weight_delta_pts / 100):
        return True
    if f.contribution_pct is not None and abs(f.contribution_pct) >= cfg.p2_min_contribution_pct:
        return True
    return False


def apply_priorities(familles: list[FamilleKpi], cfg: PrioritizationConfig) -> None:
    by_rayon: dict[int, list[FamilleKpi]] = {}
    for f in familles:
        by_rayon.setdefault(f.rayon_code, []).append(f)
    for rayon_code, group in by_rayon.items():
        top_n_codes = {
            f"{f.rayon_code}|{f.famille_label}"
            for f in sorted(group, key=lambda f: abs(f.gap_ca_fcfa), reverse=True)[: cfg.p1_top_n_gap_abs]
        }
        for f in group:
            if _is_p1(f, cfg, top_n_codes):
                f.priority = P1
            elif _is_p2(f, cfg):
                f.priority = P2
            else:
                f.priority = ""


def sort_familles_by_impact(familles: list[FamilleKpi]) -> list[FamilleKpi]:
    return sorted(familles, key=lambda f: abs(f.gap_ca_fcfa), reverse=True)


def assign_site_cle(familles: list[FamilleKpi], site_famille_rows: list[SiteFamilleKpi]) -> None:
    by_key: dict[tuple, list[SiteFamilleKpi]] = {}
    for row in site_famille_rows:
        by_key.setdefault((row.rayon_code, row.famille_label), []).append(row)
    for f in familles:
        candidates = by_key.get((f.rayon_code, f.famille_label), [])
        if not candidates:
            f.site_cle = "-"
            continue
        best = max(candidates, key=lambda r: r.gap_ca_fcfa) if f.gap_ca_fcfa >= 0 else min(candidates, key=lambda r: r.gap_ca_fcfa)
        sign = "+" if best.gap_ca_fcfa >= 0 else ""
        f.site_cle = f"{best.site_name} ({sign}{best.gap_ca_fcfa / 1_000_000:.1f} M)"


# =============================================================================
# 9. DIAGNOSTIC AUTOMATIQUE PAR RAYON
# =============================================================================

def _fmt_pct_txt(v):
    return "n/d" if v is None else f"{v * 100:+.1f} %"


def _fmt_pts_txt(v):
    return "n/d" if v is None else f"{v * 100:+.1f} pt"


def _fmt_millions_txt(v):
    return f"{v / 1_000_000:+.1f} M"


def build_rayon_diagnostic(rayon_kpi_reseau: RayonFormatKpi, familles: list[FamilleKpi], format_kpis: list[RayonFormatKpi]) -> str:
    sentences = []
    sentences.append(f"CA {_fmt_pct_txt(rayon_kpi_reseau.ca_evo_pct)} vs N-1 ISO ({_fmt_millions_txt(rayon_kpi_reseau.ca_gap_fcfa)} FCFA).")
    sentences.append(
        f"Quantités {_fmt_pct_txt(rayon_kpi_reseau.qte_evo_pct)}, taux de marge {_fmt_pts_txt(rayon_kpi_reseau.tx_marge_evo_pts)} "
        f"et poids promo {_fmt_pts_txt(rayon_kpi_reseau.poids_promo_evo_pts)}."
    )
    growth = sorted([f for f in familles if f.gap_ca_fcfa > 0], key=lambda f: f.gap_ca_fcfa, reverse=True)[:2]
    decline = sorted([f for f in familles if f.gap_ca_fcfa < 0], key=lambda f: f.gap_ca_fcfa)[:2]
    if growth:
        parts = [f"{f.famille_label} ({_fmt_millions_txt(f.gap_ca_fcfa)})" for f in growth]
        sentences.append(f"Croissance portée principalement par {' et '.join(parts)}.")
    if decline:
        parts = [f"{f.famille_label} ({_fmt_millions_txt(f.gap_ca_fcfa)})" for f in decline]
        label = "Le principal frein est" if len(parts) == 1 else "Les principaux freins sont"
        sentences.append(f"{label} {' et '.join(parts)}.")
    fmts = [f for f in format_kpis if f.scope != RESEAU and f.ca_evo_pct is not None]
    if len(fmts) >= 2:
        best, worst = max(fmts, key=lambda f: f.ca_evo_pct), min(fmts, key=lambda f: f.ca_evo_pct)
        if best.scope != worst.scope:
            sentences.append(
                f"{best.scope.capitalize()} est le format le plus dynamique ({_fmt_pct_txt(best.ca_evo_pct)}) "
                f"tandis que {worst.scope.capitalize()} progresse de {_fmt_pct_txt(worst.ca_evo_pct)}."
            )
    return " ".join(sentences)


# =============================================================================
# 10. GÉNÉRATION EXCEL
# =============================================================================

def _argb(hex_color: str) -> str:
    h = hex_color.lstrip("#")
    return "FF" + h if len(h) == 6 else h


class StyleKit:
    def __init__(self):
        c = EXCEL_COLORS
        self.nf = EXCEL_NUMBER_FORMATS
        family = "Aptos Narrow"
        self.font_title = Font(name=family, size=14, bold=True, color=_argb(c["title_font"]))
        self.font_header = Font(name=family, size=10, bold=True, color=_argb(c["header_font"]))
        self.font_body = Font(name=family, size=10)
        self.font_bold = Font(name=family, size=10, bold=True)
        self.font_muted = Font(name=family, size=9, italic=True, color=_argb("6B6B6B"))
        self.fill_title = PatternFill("solid", fgColor=_argb(c["title_bg"]))
        self.fill_header = PatternFill("solid", fgColor=_argb(c["header_bg"]))
        self.fill_band = PatternFill("solid", fgColor=_argb(c["band_bg"]))
        self.fill_input = PatternFill("solid", fgColor=_argb(c["input_cell_bg"]))
        self.fill_success = PatternFill("solid", fgColor=_argb(c["success_bg"]))
        self.fill_danger = PatternFill("solid", fgColor=_argb(c["danger_bg"]))
        self.fill_p1 = PatternFill("solid", fgColor=_argb(c["p1_bg"]))
        self.fill_p2 = PatternFill("solid", fgColor=_argb(c["p2_bg"]))
        self.font_success = Font(name=family, size=10, color=_argb(c["success_font"]))
        self.font_danger = Font(name=family, size=10, color=_argb(c["danger_font"]))
        self.font_p1 = Font(name=family, size=10, bold=True, color=_argb(c["p1_font"]))
        self.font_p2 = Font(name=family, size=10, bold=True, color=_argb(c["p2_font"]))
        self.databar_color = _argb(c["databar_color"])
        thin = Side(style="thin", color="D9D9D9")
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.border_bottom = Border(bottom=Side(style="thin", color="B0B0B0"))

    def evo_style(self, value):
        if value is None:
            return None, self.font_muted
        if value > 0:
            return self.fill_success, self.font_success
        if value < 0:
            return self.fill_danger, self.font_danger
        return None, self.font_body

    def priority_style(self, priority):
        if priority == "P1":
            return self.fill_p1, self.font_p1
        if priority == "P2":
            return self.fill_p2, self.font_p2
        return None, self.font_body


def _set(ws, row, col, value, font=None, fill=None, number_format=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if number_format: cell.number_format = number_format
    if align: cell.alignment = align
    if border: cell.border = border
    return cell


SCOPE_ORDER = [RESEAU, "HYPER", "MARKET", "SUPECO"]
SCOPE_COLS = {RESEAU: (2, 3), "HYPER": (5, 6), "MARKET": (8, 9), "SUPECO": (11, 12)}
BLOCK1_ROWS = [
    ("CA", "ca_n", "ca_evo_pct", "montant", "pct", None),
    ("Budget", "budget", "ecart_budget_pct", "montant", "pct", "budget_available"),
    ("Poids", "poids_n", "poids_evo_pts", "pct_plain", "pts", None),
    ("Marge", "marge_n", "marge_evo_pct", "montant", "pct", None),
    ("Tx Marge", "tx_marge_n", "tx_marge_evo_pts", "pct_plain", "pts", None),
    ("Qté vendue", "qte_n", "qte_evo_pct", "qty", "pct", None),
    ("Poids Promo", "poids_promo_n", "poids_promo_evo_pts", "pct_plain", "pts", None),
    ("Débit", "debit_n", None, "qty", None, "debit_available"),
    ("Panier", "panier_n", None, "qty", None, "debit_available"),
]
FAMILLE_COLUMNS = [
    "Priorité", "Famille", "CA N", "Écart CA", "vs N-1 ISO", "Poids rayon", "Contribution écart",
    "Tx Marge N", "Δ Tx Marge", "Qté vs N-1", "Poids Promo N", "Δ Promo",
    "Hyper vs N-1", "Market vs N-1", "Supeco vs N-1", "Signal Python", "Site clé",
    "Cause acheteur", "Commentaire acheteur",
]
SITE_FAMILLE_COLUMNS = ["Format", "Site", "Famille", "Écart CA", "vs N-1", "Qté vs N-1", "Δ Tx Marge", "Δ Promo", "Signal Python"]


def _write_block1(ws, start_row, kpis_by_scope, coverage_labels, style: StyleKit):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    _set(ws, start_row, 1, "BLOC 1 - SYNTHÈSE : OÙ SE SITUE LA PERFORMANCE ?", font=style.font_header,
         fill=style.fill_header, align=Alignment(horizontal="left", indent=1))
    row = start_row + 2
    for scope in SCOPE_ORDER:
        c1, c2 = SCOPE_COLS[scope]
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        _set(ws, row, c1, scope, font=style.font_bold, fill=style.fill_band, align=Alignment(horizontal="center"))
    row += 1
    for scope in SCOPE_ORDER:
        c1, c2 = SCOPE_COLS[scope]
        _set(ws, row, c1, "N", font=style.font_bold, align=Alignment(horizontal="center"), border=style.border_bottom)
        _set(ws, row, c2, "Écart", font=style.font_bold, align=Alignment(horizontal="center"), border=style.border_bottom)
    row += 1

    for label, val_attr, evo_attr, val_fmt, evo_fmt, star_attr in BLOCK1_ROWS:
        reseau_kpi = kpis_by_scope.get(RESEAU)
        incomplete = bool(star_attr and reseau_kpi is not None and not getattr(reseau_kpi, star_attr, True))
        _set(ws, row, 1, label + ("*" if incomplete else ""), font=style.font_bold)
        for scope in SCOPE_ORDER:
            c1, c2 = SCOPE_COLS[scope]
            kpi = kpis_by_scope.get(scope)
            value = getattr(kpi, val_attr, None) if kpi else None
            evo = getattr(kpi, evo_attr, None) if (kpi and evo_attr) else None
            if value is None:
                _set(ws, row, c1, NA_LABEL, font=style.font_muted, align=Alignment(horizontal="right"))
            else:
                _set(ws, row, c1, value, font=style.font_body, number_format=style.nf.get(val_fmt, "General"),
                     align=Alignment(horizontal="right"))
            if evo_attr:
                if evo is None:
                    _set(ws, row, c2, NA_LABEL, font=style.font_muted, align=Alignment(horizontal="right"))
                else:
                    fill, font = style.evo_style(evo)
                    _set(ws, row, c2, evo, font=font, fill=fill, number_format=style.nf.get(evo_fmt, "General"),
                         align=Alignment(horizontal="right"))
        row += 1

    _set(ws, row, 1, "Couverture Site", font=style.font_bold)
    for scope in SCOPE_ORDER:
        c1, _ = SCOPE_COLS[scope]
        _set(ws, row, c1, coverage_labels.get(scope, NA_LABEL), font=style.font_body, align=Alignment(horizontal="right"))
    return row + 2


def _write_diagnostic(ws, start_row, diagnostic_text, style: StyleKit):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    _set(ws, start_row, 1, "DIAGNOSTIC AUTO PYTHON - PRÉ-ANALYSE CHIFFRÉE", font=style.font_header,
         fill=style.fill_header, align=Alignment(horizontal="left", indent=1))
    ws.merge_cells(start_row=start_row, start_column=10, end_row=start_row, end_column=19)
    _set(ws, start_row, 10, "COMMENTAIRE SYNTHÈSE ACHETEUR", font=style.font_header,
         fill=style.fill_header, align=Alignment(horizontal="left", indent=1))
    body_row = start_row + 1
    ws.merge_cells(start_row=body_row, start_column=1, end_row=body_row + 3, end_column=9)
    _set(ws, body_row, 1, diagnostic_text, font=style.font_body, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.merge_cells(start_row=body_row, start_column=10, end_row=body_row + 3, end_column=19)
    _set(ws, body_row, 10, "", fill=style.fill_input, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    for r in range(body_row, body_row + 4):
        ws.row_dimensions[r].height = 20
    return body_row + 5


def _write_block2(ws, start_row, familles: list[FamilleKpi], style: StyleKit):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=19)
    _set(ws, start_row, 1, "BLOC 2 - FAMILLES TRIÉES PAR IMPACT : QU'EST-CE QUI EXPLIQUE LE RÉSULTAT ?",
         font=style.font_header, fill=style.fill_header, align=Alignment(horizontal="left", indent=1))
    header_row = start_row + 2
    for idx, col_name in enumerate(FAMILLE_COLUMNS, start=1):
        _set(ws, header_row, idx, col_name, font=style.font_header, fill=style.fill_header,
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[header_row].height = 28
    row = header_row + 1
    first_data_row = row

    for f in familles:
        prio_fill, prio_font = style.priority_style(f.priority)
        _set(ws, row, 1, f.priority or "-", font=prio_font, fill=prio_fill, align=Alignment(horizontal="center"))
        _set(ws, row, 2, f.famille_label, font=style.font_body)
        _set(ws, row, 3, f.ca_n, font=style.font_body, number_format=style.nf["montant"])
        _set(ws, row, 4, f.gap_ca_fcfa, font=style.font_body, number_format=style.nf["montant_ecart"])
        evo_fill, evo_font = style.evo_style(f.evo_pct)
        _set(ws, row, 5, f.evo_pct if f.evo_pct is not None else NA_LABEL, font=evo_font, fill=evo_fill,
             number_format=style.nf["pct"] if f.evo_pct is not None else None)
        _set(ws, row, 6, f.poids_rayon if f.poids_rayon is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct_plain"] if f.poids_rayon is not None else None)
        _set(ws, row, 7, f.contribution_pct if f.contribution_pct is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct_plain"] if f.contribution_pct is not None else None)
        _set(ws, row, 8, f.tx_marge_n if f.tx_marge_n is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct_plain"] if f.tx_marge_n is not None else None)
        dtm_fill, dtm_font = style.evo_style(f.delta_tx_marge_pts)
        _set(ws, row, 9, f.delta_tx_marge_pts if f.delta_tx_marge_pts is not None else NA_LABEL, font=dtm_font, fill=dtm_fill,
             number_format=style.nf["pts"] if f.delta_tx_marge_pts is not None else None)
        _set(ws, row, 10, f.qte_evo_pct if f.qte_evo_pct is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct"] if f.qte_evo_pct is not None else None)
        _set(ws, row, 11, f.poids_promo_n if f.poids_promo_n is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct_plain"] if f.poids_promo_n is not None else None)
        _set(ws, row, 12, f.delta_promo_pts if f.delta_promo_pts is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pts"] if f.delta_promo_pts is not None else None)
        for col, val in [(13, f.hyper_evo_pct), (14, f.market_evo_pct), (15, f.supeco_evo_pct)]:
            _set(ws, row, col, val if val is not None else NA_LABEL, font=style.font_body,
                 number_format=style.nf["pct"] if val is not None else None)
        _set(ws, row, 16, f.signal, font=style.font_body)
        _set(ws, row, 17, f.site_cle, font=style.font_body)
        _set(ws, row, 18, "", fill=style.fill_input)
        _set(ws, row, 19, "", fill=style.fill_input)
        for col in range(1, 20):
            ws.cell(row=row, column=col).border = style.border_thin
        row += 1

    last_data_row = row - 1
    if last_data_row >= first_data_row:
        dv = DataValidation(type="list", formula1='"' + ",".join(CAUSE_OPTIONS) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"R{first_data_row}:R{last_data_row}")
        databar_rule = DataBarRule(start_type="min", start_value=None, end_type="max", end_value=None,
                                    color=style.databar_color, showValue=True)
        ws.conditional_formatting.add(f"D{first_data_row}:D{last_data_row}", databar_rule)
        ws.auto_filter.ref = f"A{header_row}:S{last_data_row}"
    ws.freeze_panes = f"C{first_data_row}"
    return last_data_row + 2


def _write_block3(ws, start_row, rows: list[SiteFamilleKpi], style: StyleKit):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    _set(ws, start_row, 1, "BLOC 3 - TOP CONTRIBUTIONS SITE / FAMILLE À EXPLIQUER (GÉNÉRÉ PAR PYTHON)",
         font=style.font_header, fill=style.fill_header, align=Alignment(horizontal="left", indent=1))
    header_row = start_row + 2
    for idx, col_name in enumerate(SITE_FAMILLE_COLUMNS, start=1):
        _set(ws, header_row, idx, col_name, font=style.font_header, fill=style.fill_header,
             align=Alignment(horizontal="center", vertical="center"))
    row = header_row + 1
    for r in rows:
        _set(ws, row, 1, r.format, font=style.font_body)
        _set(ws, row, 2, r.site_name, font=style.font_body)
        _set(ws, row, 3, r.famille_label, font=style.font_body)
        _set(ws, row, 4, r.gap_ca_fcfa, font=style.font_body, number_format=style.nf["montant_ecart"])
        evo_fill, evo_font = style.evo_style(r.evo_pct)
        _set(ws, row, 5, r.evo_pct if r.evo_pct is not None else NA_LABEL, font=evo_font, fill=evo_fill,
             number_format=style.nf["pct"] if r.evo_pct is not None else None)
        _set(ws, row, 6, r.qte_evo_pct if r.qte_evo_pct is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pct"] if r.qte_evo_pct is not None else None)
        _set(ws, row, 7, r.delta_tx_marge_pts if r.delta_tx_marge_pts is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pts"] if r.delta_tx_marge_pts is not None else None)
        _set(ws, row, 8, r.delta_promo_pts if r.delta_promo_pts is not None else NA_LABEL, font=style.font_body,
             number_format=style.nf["pts"] if r.delta_promo_pts is not None else None)
        _set(ws, row, 9, r.signal, font=style.font_body)
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = style.border_thin
        row += 1
    return row + 1


def _set_column_widths(ws, ncols):
    defaults = {1: 16, 2: 26, 3: 13, 4: 13, 5: 11, 6: 11, 7: 11, 8: 11, 9: 11, 10: 11,
                11: 11, 12: 11, 13: 11, 14: 11, 15: 11, 16: 24, 17: 20, 18: 16, 19: 30}
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = defaults.get(col, 12)


def write_rayon_sheet(wb, sheet_name, rayon_label, week_label, quality: QualityReport, coverage_labels,
                       kpis_by_scope, familles, site_famille_rows, diagnostic_text, style: StyleKit):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    _set_column_widths(ws, 19)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    _set(ws, 1, 1, f"REPORTING HEBDOMADAIRE ACHATS - {rayon_label.upper()}", font=style.font_title,
         fill=style.fill_title, align=Alignment(horizontal="center", vertical="center"))
    for c in range(1, 20):
        ws.cell(row=1, column=c).fill = style.fill_title
    ws.row_dimensions[1].height = 22

    _set(ws, 3, 1, "SEMAINE ISO", font=style.font_bold)
    _set(ws, 3, 2, week_label, font=style.font_body)
    _set(ws, 3, 4, "RAYON", font=style.font_bold)
    _set(ws, 3, 5, rayon_label.upper(), font=style.font_body, fill=style.fill_input)
    _set(ws, 3, 7, "ACHETEUR", font=style.font_bold)
    _set(ws, 3, 8, "", fill=style.fill_input)
    _set(ws, 3, 10, "STATUT", font=style.font_bold)
    _set(ws, 3, 11, "À compléter", fill=style.fill_input)
    dv_status = DataValidation(type="list", formula1='"À compléter,En cours,Validé"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("K3")

    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=19)
    _set(ws, 4, 1, quality.summary_line, font=style.font_muted, align=Alignment(horizontal="left", vertical="top", wrap_text=True))

    row = _write_block1(ws, 7, kpis_by_scope, coverage_labels, style)
    row = _write_diagnostic(ws, row, diagnostic_text, style)
    row = _write_block2(ws, row, familles, style)
    _write_block3(ws, row, site_famille_rows, style)
    ws.freeze_panes = "B7"
    return ws


def build_workbook(rayons, week_label, quality, coverage_labels_by_rayon, kpis_by_rayon,
                    familles_by_rayon, site_famille_by_rayon, diagnostics_by_rayon) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    style = StyleKit()
    for sheet_name, rayon_label in rayons:
        write_rayon_sheet(
            wb, sheet_name, rayon_label, week_label, quality,
            coverage_labels_by_rayon.get(rayon_label, {}), kpis_by_rayon.get(rayon_label, {}),
            familles_by_rayon.get(rayon_label, []), site_famille_by_rayon.get(rayon_label, []),
            diagnostics_by_rayon.get(rayon_label, ""), style,
        )
    return wb


# =============================================================================
# 11. INTERFACE STREAMLIT
# =============================================================================

st.markdown(
    """
    <style>
    .kpi-card { background:#FFFFFF; border-radius:14px; padding:16px 20px; border:1px solid #E5E5EA; margin-bottom:8px; }
    .kpi-label { font-size:12px; color:#6E6E73; margin-bottom:4px; }
    .kpi-value { font-size:22px; font-weight:600; color:#1D1D1F; }
    .kpi-evo-pos { font-size:12px; color:#34C759; margin-top:4px; }
    .kpi-evo-neg { font-size:12px; color:#FF3B30; margin-top:4px; }
    .kpi-evo-neutral { font-size:12px; color:#6E6E73; margin-top:4px; }
    .diag-box { background:#F2F2F7; border-radius:14px; padding:16px 20px; font-size:14px; color:#1D1D1F; margin:12px 0 20px 0; }
    .alert-box { background:#FFF3E0; border-left:4px solid #FF9500; border-radius:8px; padding:10px 16px; font-size:13px; color:#6E6E73; margin:12px 0; }
    </style>
    """,
    unsafe_allow_html=True,
)


def _fmt_millions(v):
    return "N/D" if v is None or pd.isna(v) else f"{v / 1_000_000:,.1f} M".replace(",", " ")


def _fmt_pct(v):
    return "N/D" if v is None or pd.isna(v) else f"{v * 100:+.1f} %"


def _fmt_pts(v):
    return "N/D" if v is None or pd.isna(v) else f"{v * 100:+.2f} pt"


def _evo_css(v):
    if v is None or pd.isna(v):
        return "kpi-evo-neutral"
    return "kpi-evo-pos" if v > 0 else ("kpi-evo-neg" if v < 0 else "kpi-evo-neutral")


def kpi_card(col, label, value_str, evo_val, evo_str):
    col.markdown(
        f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value_str}</div>'
        f'<div class="{_evo_css(evo_val)}">{evo_str}</div></div>',
        unsafe_allow_html=True,
    )


@st.cache_data(show_spinner="Lecture des exports...")
def parse_files(site_bytes, current_bytes, previous_bytes):
    """Étape coûteuse (lecture Excel) : mise en cache indépendamment des seuils métier."""
    site_raw, _ = read_export_bytes(site_bytes)
    n_raw, period_n = read_export_bytes(current_bytes)
    n1_raw, period_n1 = read_export_bytes(previous_bytes)

    site_enriched = enrich_site_export(site_raw)
    n_enriched = enrich_detail_export(n_raw)
    n1_enriched = enrich_detail_export(n1_raw)

    return detail_rows(site_enriched), detail_rows(n_enriched), detail_rows(n1_enriched), period_n.label


st.title("Reporting Vente CA")
st.caption("Vue BI exploratoire - la saisie Cause / Commentaire se fait dans le fichier Excel exporté")

with st.sidebar:
    st.subheader("Imports")
    site_file = st.file_uploader("Export DATA_SITE", type=["xlsx"])
    current_file = st.file_uploader("Export DATA_N (semaine courante)", type=["xlsx"])
    previous_file = st.file_uploader("Export DATA_N1_ISO", type=["xlsx"])

prio_cfg, drivers_cfg, quality_cfg, top_site_famille_n = render_sidebar_config()

if not (site_file and current_file and previous_file):
    st.info("Charge les 3 exports PBI dans la barre latérale pour démarrer l'analyse.")
    st.stop()

site_d, n_d, n1_d, period_label = parse_files(site_file.getvalue(), current_file.getvalue(), previous_file.getvalue())

# Recalcul à chaque changement de seuil (rapide, pas besoin de cache) :
quality = build_quality_report(site_d, n_d, n1_d, quality_cfg)
coverage_labels = {RESEAU: quality.coverage_global.label, **{f: c.label for f, c in quality.coverage_by_format.items()}}

rayons = n_d[["rayon_code", "rayon_label"]].drop_duplicates().set_index("rayon_code")["rayon_label"].to_dict()
display_label = {c: DEFAULT_RAYON_LABEL_OVERRIDES.get(l, l) for c, l in rayons.items()}
rayon_labels = sorted(set(display_label.values()))

all_kpis = build_rayon_format_kpis(n_d, n1_d, site_d, {
    RESEAU: quality.coverage_global.coverage_pct,
    **{f: c.coverage_pct for f, c in quality.coverage_by_format.items()},
})
kpis_lookup = {(k.rayon_code, k.scope): k for k in all_kpis}

familles_by_rayon, site_famille_by_rayon, diagnostics_by_rayon = {}, {}, {}
for code, label in display_label.items():
    n_r, n1_r = n_d[n_d["rayon_code"] == code], n1_d[n1_d["rayon_code"] == code]
    familles = build_famille_kpis(n_r, n1_r)
    apply_priorities(familles, prio_cfg)
    apply_signals_famille(familles, drivers_cfg)
    sf_rows = build_site_famille_kpis(n_r, n1_r)
    apply_signals_site_famille(sf_rows, drivers_cfg)
    assign_site_cle(familles, sf_rows)
    familles_sorted = sort_familles_by_impact(familles)
    familles_by_rayon[label] = familles_sorted
    site_famille_by_rayon[label] = sorted(sf_rows, key=lambda r: abs(r.gap_ca_fcfa), reverse=True)[:top_site_famille_n]

    scope_kpis = {s: kpis_lookup.get((code, s)) for s in SCOPE_ORDER}
    if scope_kpis.get(RESEAU) is not None:
        diagnostics_by_rayon[label] = build_rayon_diagnostic(
            scope_kpis[RESEAU], familles_sorted, [k for k in scope_kpis.values() if k is not None]
        )

tab_reseau, tab_rayon, tab_familles, tab_qualite = st.tabs(["Vue réseau", "Vue rayon", "Vue familles", "Contrôles qualité"])

with tab_reseau:
    st.caption(f"Semaine : {period_label}")
    cols = st.columns(len(rayon_labels) or 1)
    for col, label in zip(cols, rayon_labels):
        code = next(c for c, l in display_label.items() if l == label)
        kpi = kpis_lookup.get((code, RESEAU))
        if kpi:
            kpi_card(col, label, _fmt_millions(kpi.ca_n), kpi.ca_evo_pct, _fmt_pct(kpi.ca_evo_pct) + " vs N-1 ISO")
    if quality.has_perimeter_issue:
        st.markdown(f'<div class="alert-box">{quality.summary_line}</div>', unsafe_allow_html=True)

with tab_rayon:
    selected_rayon = st.selectbox("Rayon", rayon_labels, key="rayon_select")
    code = next(c for c, l in display_label.items() if l == selected_rayon)
    scope_cols = st.columns(4)
    for col, scope in zip(scope_cols, SCOPE_ORDER):
        kpi = kpis_lookup.get((code, scope))
        if kpi:
            kpi_card(col, scope.capitalize(), _fmt_millions(kpi.ca_n), kpi.ca_evo_pct, _fmt_pct(kpi.ca_evo_pct) + " vs N-1 ISO")
    diag_text = diagnostics_by_rayon.get(selected_rayon, "")
    if diag_text:
        st.markdown(f'<div class="diag-box">{diag_text}</div>', unsafe_allow_html=True)
    familles = familles_by_rayon.get(selected_rayon, [])
    if familles:
        df = pd.DataFrame([{
            "Priorité": f.priority or "-", "Famille": f.famille_label, "CA N": _fmt_millions(f.ca_n),
            "Écart CA": _fmt_millions(f.gap_ca_fcfa), "vs N-1": _fmt_pct(f.evo_pct),
            "Δ Tx marge": _fmt_pts(f.delta_tx_marge_pts), "Signal": f.signal, "Site clé": f.site_cle,
        } for f in familles])
        st.dataframe(df, use_container_width=True, hide_index=True)
    if quality.has_perimeter_issue:
        st.markdown(f'<div class="alert-box">{quality.summary_line}</div>', unsafe_allow_html=True)

with tab_familles:
    all_familles = []
    for label, familles in familles_by_rayon.items():
        for f in familles:
            all_familles.append({
                "Rayon": label, "Priorité": f.priority or "-", "Famille": f.famille_label,
                "Écart CA (FCFA)": f.gap_ca_fcfa, "vs N-1 (%)": (f.evo_pct or 0) * 100,
                "Δ Tx marge (pt)": (f.delta_tx_marge_pts or 0) * 100, "Signal": f.signal,
            })
    df_all = pd.DataFrame(all_familles)
    if not df_all.empty:
        prio_filter = st.multiselect("Filtrer par priorité", ["P1", "P2", "-"], default=["P1", "P2"])
        filtered = df_all[df_all["Priorité"].isin(prio_filter)] if prio_filter else df_all
        st.dataframe(filtered.sort_values("Écart CA (FCFA)", key=abs, ascending=False), use_container_width=True, hide_index=True)

with tab_qualite:
    st.subheader("Périmètre de sites")
    cov_rows = [{"Scope": RESEAU, "Couverture": quality.coverage_global.label, "%": f"{quality.coverage_global.coverage_pct:.0%}"}]
    for fmt, cov in quality.coverage_by_format.items():
        cov_rows.append({"Scope": fmt, "Couverture": cov.label, "%": f"{cov.coverage_pct:.0%}"})
    st.dataframe(pd.DataFrame(cov_rows), use_container_width=True, hide_index=True)
    if quality.coverage_global.missing:
        st.warning("Sites absents de DATA_SITE (Budget/Débit/Panier indisponibles) : " + ", ".join(sorted(quality.coverage_global.missing)))

    st.subheader("Réconciliation CA (DATA_SITE vs DATA_N, sites communs)")
    recon_rows = [{
        "Scope": RESEAU, "CA DATA_SITE": _fmt_millions(quality.recon_global.ca_site),
        "CA DATA_N": _fmt_millions(quality.recon_global.ca_detail),
        "Alignement": f"{quality.recon_global.ratio:.1%}", "Statut": quality.recon_global.status,
    }]
    for fmt, recon in quality.recon_by_format.items():
        recon_rows.append({
            "Scope": fmt, "CA DATA_SITE": _fmt_millions(recon.ca_site), "CA DATA_N": _fmt_millions(recon.ca_detail),
            "Alignement": f"{recon.ratio:.1%}", "Statut": recon.status,
        })
    st.dataframe(pd.DataFrame(recon_rows), use_container_width=True, hide_index=True)

st.divider()
if st.button("Générer le reporting Excel", type="primary"):
    with st.spinner("Génération du fichier Excel..."):
        rayons_sheet_info = [(f"{i+1:02d}_{label.replace(' ', '_')[:25]}", label) for i, label in enumerate(rayon_labels)]
        kpis_by_rayon = {}
        for label in rayon_labels:
            code = next(c for c, l in display_label.items() if l == label)
            kpis_by_rayon[label] = {s: kpis_lookup[(code, s)] for s in SCOPE_ORDER if (code, s) in kpis_lookup}

        wb = build_workbook(
            rayons=rayons_sheet_info, week_label=period_label, quality=quality,
            coverage_labels_by_rayon={label: coverage_labels for label in rayon_labels},
            kpis_by_rayon=kpis_by_rayon, familles_by_rayon=familles_by_rayon,
            site_famille_by_rayon=site_famille_by_rayon, diagnostics_by_rayon=diagnostics_by_rayon,
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        st.download_button(
            "Télécharger le fichier Excel", data=buffer,
            file_name=f"reporting_achats_{period_label.split(' ')[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
