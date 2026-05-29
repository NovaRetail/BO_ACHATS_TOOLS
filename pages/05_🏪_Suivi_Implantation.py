"""
05_🏪_Suivi_Implantation.py — SmartBuyer Hub
Suivi Implantation Nouvelles Références — v7.0
Source stock : fichiers CSV ERP par magasin (upload direct, 1 à 13 fichiers)
Plus besoin de consolider en amont — le module fusionne lui-même.

Changements v7 vs v6 :
  - Suppression du fichier stock_consolide CSV unique
  - Upload multi-fichiers directs (1 fichier par magasin)
  - Parser unifié : colonnes stables, encodage latin1, séparateur ;
  - Code site et date extraits automatiquement du nom de fichier
  - Logique métier et alertes identiques à v6
"""

import io
import re
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from datetime import date, datetime

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
TODAY      = pd.Timestamp(date.today())
TODAY_STR  = date.today().strftime("%d %b %Y")
TODAY_FILE = date.today().strftime("%Y%m%d")

# Colonnes obligatoires dans chaque fichier stock
STOCK_COLS_REQUIRED = [
    "Site", "Libellé site", "Code article", "Libellé article",
    "Code etat", "Nouveau stock", "Ral", "Pcb",
    "Code marketing", "Nom fourn.", "Libellé rayon", "Libellé famille",
    "Date dernière entrée", "Type saisonnalité",
]

ETAT_ACTIF    = "2"
ETAT_PURGE    = "P"
ETAT_ANOMALIE = {"B", "S", "F", "6", "5", "1"}
ETAT_LABEL    = {
    "B": "Rayon générique", "S": "Suspendu",
    "F": "Fin de vie",      "6": "Déréférencé",
    "5": "Autre",
}

ALERTES = {
    "✅ Implanté"        : "#34C759",
    "🔵 Appro en cours"  : "#007AFF",
    "🛒 Passer commande" : "#FF9500",
    "🚩 Anomalie référ." : "#FFD60A",
}
ACTION_LABEL = {
    "✅ Implanté"        : "—",
    "🔵 Appro en cours"  : "Accélérer livraison",
    "🛒 Passer commande" : "Passer commande fournisseur",
    "🚩 Anomalie référ." : "Vérifier référencement magasin",
}

C = {
    "bg": "#F2F2F7", "surface": "#FFFFFF", "border": "#E5E5EA",
    "text": "#1C1C1E", "muted": "#6D6D72",
    "blue": "#007AFF", "green": "#34C759", "red": "#FF3B30",
    "orange": "#FF9500", "purple": "#AF52DE", "yellow": "#FFD60A",
    "blue_l": "#EFF4FF", "green_l": "#F0FFF4",
    "red_l": "#FFF2F0", "orange_l": "#FFFBEB",
}

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG + CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Implantation · SmartBuyer",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(f"""
<style>
html,body,[class*="css"]{{font-family:-apple-system,BlinkMacSystemFont,"SF Pro Display","Helvetica Neue",Arial,sans-serif!important;background:{C['bg']}!important;}}
.main,section[data-testid="stMain"]{{background:{C['bg']}!important;}}
.block-container{{padding:0 2rem 4rem!important;max-width:1480px;}}
header[data-testid="stHeader"],#MainMenu,footer{{display:none!important;}}

[data-testid="stMetric"]{{background:{C['surface']}!important;border:0.5px solid {C['border']}!important;border-radius:12px!important;padding:16px 18px!important;}}
[data-testid="stMetricLabel"]{{font-size:11px!important;font-weight:500!important;color:{C['muted']}!important;text-transform:uppercase!important;letter-spacing:0.04em!important;}}
[data-testid="stMetricValue"]{{font-size:24px!important;font-weight:600!important;color:{C['text']}!important;}}
[data-testid="stTabs"] button[role="tab"]{{font-size:13px!important;font-weight:500!important;padding:8px 16px!important;color:{C['muted']}!important;border-bottom:2px solid transparent!important;}}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"]{{color:{C['blue']}!important;border-bottom:2px solid {C['blue']}!important;background:transparent!important;}}
[data-testid="stTabs"] [role="tablist"]{{border-bottom:0.5px solid {C['border']}!important;}}
[data-testid="stDataFrame"]{{border:0.5px solid {C['border']}!important;border-radius:10px!important;}}
[data-testid="stDataFrame"] th{{background:{C['bg']}!important;font-size:11px!important;font-weight:600!important;color:{C['muted']}!important;text-transform:uppercase!important;}}
[data-testid="stFileUploader"]{{border:1.5px dashed {C['border']}!important;border-radius:10px!important;background:#F9F9FB!important;}}
.stDownloadButton>button{{background:{C['blue']}!important;color:#fff!important;border:none!important;border-radius:8px!important;font-weight:500!important;padding:10px 24px!important;width:100%!important;}}

.topbar{{background:{C['text']};margin:0 -2rem 28px;padding:16px 28px;display:flex;align-items:center;justify-content:space-between;}}
.topbar-icon{{width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,{C['blue']},{C['purple']});display:flex;align-items:center;justify-content:center;font-size:22px;}}
.topbar-title{{font-size:17px;font-weight:700;color:#fff;letter-spacing:-.01em;}}
.topbar-sub{{font-size:11px;color:#8E8E93;margin-top:2px;}}
.topbar-pill{{background:rgba(255,255,255,.08);color:#8E8E93;border:1px solid rgba(255,255,255,.12);border-radius:8px;padding:4px 14px;font-size:11px;font-weight:600;}}
.topbar-date{{color:{C['blue']};font-size:12px;}}

.sh{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.12em;color:{C['muted']};margin:22px 0 12px;padding-bottom:8px;border-bottom:1px solid {C['border']};}}
.kpi-card{{background:{C['surface']};border:1px solid {C['border']};border-radius:14px;padding:18px 20px 14px;box-shadow:0 1px 3px rgba(0,0,0,.06);position:relative;overflow:hidden;}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:14px 14px 0 0;}}
.kpi-card.green::before{{background:{C['green']};}} .kpi-card.blue::before{{background:{C['blue']};}}
.kpi-card.orange::before{{background:{C['orange']};}} .kpi-card.yellow::before{{background:{C['yellow']};}}
.kpi-label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.10em;color:{C['muted']};margin-bottom:10px;}}
.kpi-value{{font-size:36px;font-weight:800;line-height:1;letter-spacing:-.02em;}}
.kpi-card.green .kpi-value{{color:{C['green']};}} .kpi-card.blue .kpi-value{{color:{C['blue']};}}
.kpi-card.orange .kpi-value{{color:{C['orange']};}} .kpi-card.yellow .kpi-value{{color:#B8860B;}}
.kpi-sub{{font-size:11px;color:{C['muted']};margin-top:4px;}}

.scorecard-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:10px;margin-bottom:22px;}}
.scorecard-card{{background:{C['surface']};border:1px solid {C['border']};border-radius:14px;padding:14px 16px;position:relative;}}
.scorecard-card.ok{{border-color:#6EE7B7;background:{C['green_l']};}}
.scorecard-card.warn{{border-color:#FCD34D;background:{C['orange_l']};}}
.scorecard-card.ko{{border-color:#FECACA;background:{C['red_l']};}}
.scorecard-dot{{width:8px;height:8px;border-radius:50%;position:absolute;top:14px;right:14px;}}
.scorecard-name{{font-size:11px;font-weight:600;color:{C['text']};margin-bottom:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:88%;}}
.scorecard-pct{{font-size:28px;font-weight:800;line-height:1;}}
.scorecard-sub{{font-size:10px;color:{C['muted']};margin-top:3px;}}

.alert-banner{{background:#FFF;border:1px solid #FECACA;border-left:4px solid {C['red']};border-radius:14px;padding:14px 20px;margin-bottom:18px;display:flex;align-items:center;gap:14px;flex-wrap:wrap;}}
.alert-pill{{background:{C['red']};color:#fff;border-radius:6px;padding:4px 12px;font-size:11px;font-weight:700;}}
.alert-item{{display:flex;flex-direction:column;align-items:center;padding:0 14px;border-right:1px solid {C['border']};}}
.alert-item:last-child{{border-right:none;}}
.alert-num{{font-size:24px;font-weight:800;line-height:1;}}
.alert-lbl{{font-size:10px;font-weight:600;color:{C['muted']};text-transform:uppercase;letter-spacing:.06em;margin-top:1px;}}

.info-box{{border-radius:14px;padding:14px 18px;margin-bottom:14px;border:1px solid;font-size:13px;line-height:1.6;}}
.info-box.blue{{background:{C['blue_l']};border-color:#BFDBFE;color:#1D4ED8;}}
.info-box.green{{background:{C['green_l']};border-color:#6EE7B7;color:#065F46;}}
.info-box.orange{{background:{C['orange_l']};border-color:#FCD34D;color:#92400E;}}

.val-box{{background:{C['surface']};border:1px solid {C['border']};border-radius:14px;padding:12px 20px;margin-bottom:18px;display:flex;align-items:center;gap:16px;flex-wrap:wrap;}}
.val-item{{display:flex;flex-direction:column;align-items:center;padding:0 14px;border-right:1px solid {C['border']};}}
.val-item:last-child{{border-right:none;padding-right:0;}}
.val-num{{font-size:20px;font-weight:800;line-height:1;}}
.val-lbl{{font-size:10px;color:{C['muted']};text-transform:uppercase;letter-spacing:.06em;margin-top:2px;}}

section[data-testid="stSidebar"]{{background:#fff!important;border-right:1px solid {C['border']}!important;min-width:270px!important;max-width:270px!important;}}
section[data-testid="stSidebar"] .block-container{{padding:.6rem .8rem 2rem!important;}}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fmt_n(n) -> str:
    try:
        return f"{int(n):,}".replace(",", "\u202f")
    except Exception:
        return str(n)

def color_taux(t: float) -> str:
    if t >= 80: return C["green"]
    if t >= 50: return C["orange"]
    return C["red"]

def scorecard_cls(t: float) -> str:
    if t >= 80: return "ok"
    if t >= 50: return "warn"
    return "ko"

def extract_site_from_filename(filename: str) -> str | None:
    """Extrait le code site (5 chiffres) depuis le nom de fichier."""
    m = re.search(r'_(\d{5})_', filename)
    return m.group(1) if m else None

def extract_date_from_filename(filename: str) -> pd.Timestamp:
    """Extrait la date YYYYMMDD ou DDMM depuis le nom de fichier."""
    m8 = re.search(r'_(\d{8})_', filename)
    if m8:
        for fmt in ('%Y%m%d', '%d%m%Y'):
            try:
                return pd.Timestamp(pd.to_datetime(m8.group(1), format=fmt))
            except Exception:
                continue
    return TODAY

def _sem_sort(s) -> int:
    cleaned = re.sub(r"[Ss]", "", str(s).strip())
    return int(cleaned) if cleaned.isdigit() else 99

def get_alerte(row) -> str:
    etat  = str(row.get("Code etat", "")).strip()
    stock = row.get("Nouveau stock")
    ral   = float(row.get("Ral", 0) or 0)

    if not etat or etat in ("nan", ETAT_PURGE) or etat in ETAT_ANOMALIE:
        return "🚩 Anomalie référ."
    if pd.isna(stock):
        return "🚩 Anomalie référ."
    s = float(stock)
    if s != 0:
        return "✅ Implanté"
    if ral > 0:
        return "🔵 Appro en cours"
    return "🛒 Passer commande"


# ─────────────────────────────────────────────────────────────────────────────
# PARSERS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_t1(file_bytes: bytes, filename: str):
    buf = io.BytesIO(file_bytes)
    try:
        df = pd.read_excel(buf, header=None, dtype=str) \
             if filename.lower().endswith((".xlsx", ".xls")) \
             else pd.read_csv(buf, header=None, sep=None, engine="python",
                              encoding="latin1", dtype=str, on_bad_lines="skip")
    except Exception as e:
        return None, f"Lecture T1 : {e}"

    first = str(df.iloc[0, 0]).strip().replace(".0", "")
    has_header = not first.isdigit()
    if has_header:
        df.columns = df.iloc[0].astype(str).str.strip().str.upper()
        df = df.iloc[1:].reset_index(drop=True)
    else:
        df.columns = ["ARTICLE"] + [f"_COL{i}" for i in range(1, len(df.columns))]

    df.columns = (df.columns.astype(str).str.strip().str.upper()
                  .str.replace("\ufeff", "", regex=False)
                  .str.replace("\xa0", " ", regex=False))

    if "ARTICLE" not in df.columns:
        return None, "Colonne ARTICLE introuvable dans le fichier T1"

    df["SKU"] = (df["ARTICLE"].astype(str).str.strip()
                 .str.replace(r"\.0$", "", regex=True).str.zfill(8).str[:8])
    df = df[df["SKU"].str.match(r"^\d{8}$", na=False)].drop_duplicates("SKU").copy()

    defaults = {
        "LIBELLÉ ARTICLE": "", "LIBELLÉ FOURNISSEUR ORIGINE": "",
        "MODE APPRO": "", "SEMAINE RECEPTION": "", "DATE LIV.": ""
    }
    for col, val in defaults.items():
        if col not in df.columns:
            df[col] = val

    df["SEMAINE RECEPTION"] = df["SEMAINE RECEPTION"].astype(str).str.strip().replace("nan", "")
    df["SEM_NUM"] = df["SEMAINE RECEPTION"].apply(
        lambda s: int(re.sub(r"[Ss]", "", s)) if re.sub(r"[Ss]", "", s).isdigit() else 99
    )
    df["ORIGINE"] = df["MODE APPRO"].apply(
        lambda m: "IM" if "IMPORT" in str(m).upper() else "LO"
    )
    return df, None


@st.cache_data(show_spinner=False)
def parse_stock_file(file_bytes: bytes, filename: str, sku_scope: tuple):
    """
    Parse un seul fichier stock CSV par magasin.
    Format : CSV ; latin1, colonnes stables.
    Le code site est extrait du nom de fichier si absent des données.
    """
    try:
        df = pd.read_csv(
            io.BytesIO(file_bytes),
            sep=";", encoding="latin1", low_memory=False,
            on_bad_lines="skip", dtype=str,
        )
    except Exception as e:
        return None, f"Lecture {filename} : {e}"

    # Vérifier colonnes obligatoires
    missing = [c for c in STOCK_COLS_REQUIRED if c not in df.columns]
    if missing:
        return None, f"{filename} — colonnes manquantes : {', '.join(missing)}"

    # Nettoyage
    for col in ["Libellé site", "Libellé article", "Nom fourn.",
                "Libellé rayon", "Code etat", "Code marketing",
                "Type saisonnalité"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    df["SKU"] = (df["Code article"].astype(str).str.strip()
                 .str.replace(r"\.0$", "", regex=True).str.zfill(8).str[:8])

    # Code site : depuis la colonne ou depuis le nom de fichier
    df["Code site"] = df["Site"].astype(str).str.strip()
    site_from_name  = extract_site_from_filename(filename)
    if site_from_name and (df["Code site"].eq("nan").all() or df["Code site"].isna().all()):
        df["Code site"] = site_from_name

    # Exclure lignes purge + filtrer sur SKU T1
    df = df[df["Code etat"] != ETAT_PURGE].copy()
    if sku_scope:
        df = df[df["SKU"].isin(sku_scope)].copy()

    # Garder uniquement les SKU à 8 chiffres (exclut totaux rayon/famille)
    df = df[df["SKU"].str.match(r"^\d{8}$", na=False)].copy()

    # Numériques
    df["Nouveau stock"] = pd.to_numeric(df["Nouveau stock"], errors="coerce")
    df["Ral"]           = pd.to_numeric(df["Ral"], errors="coerce").fillna(0).astype(int)
    df["Pcb"]           = pd.to_numeric(df["Pcb"], errors="coerce").fillna(1)

    # Flux
    df["Origine"] = df["Code marketing"].apply(
        lambda m: "IM" if str(m).strip().upper() == "IM" else "LO"
    )

    # Date du fichier
    df["_date_fichier"] = extract_date_from_filename(filename)

    return df, None


def consolidate_stock(uploaded_files, sku_scope: tuple):
    """Consolide N fichiers stock (1 par magasin) en un seul DataFrame."""
    frames, errors = [], []
    for f in uploaded_files:
        file_bytes = f.read()
        df_site, err = parse_stock_file(file_bytes, f.name, sku_scope)
        if err:
            errors.append(err)
        elif df_site is not None and not df_site.empty:
            frames.append(df_site)

    if not frames:
        return pd.DataFrame(), errors

    return pd.concat(frames, ignore_index=True), errors


# ─────────────────────────────────────────────────────────────────────────────
# TOPBAR
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="topbar">
  <div style="display:flex;align-items:center;gap:14px;">
    <div class="topbar-icon">📦</div>
    <div>
      <div class="topbar-title">Suivi Implantation · Nouvelles Références</div>
      <div class="topbar-sub">T1 · Stock ERP par magasin · Alertes · Cessions</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:12px;">
    <div class="topbar-date">{TODAY_STR}</div>
    <div class="topbar-pill">v7.0 · SmartBuyer</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📁 Fichiers")
    st.divider()

    st.markdown("**① T1 — Nouvelles Références**")
    t1_file = st.file_uploader("T1", type=["csv", "xlsx", "xls"],
                               key="t1", label_visibility="collapsed")

    st.markdown("**② Fichiers stock par magasin**")
    st.caption("1 à 13 fichiers CSV · 1 fichier = 1 magasin · nommage : *_10206_20260507_*.csv")
    stk_files = st.file_uploader(
        "Stock", type=["csv"], accept_multiple_files=True,
        key="stk", label_visibility="collapsed",
    )


# ─────────────────────────────────────────────────────────────────────────────
# ÉCRAN D'ACCUEIL
# ─────────────────────────────────────────────────────────────────────────────
if not t1_file:
    st.markdown("""
<div class="info-box blue">
  ⬆️ <strong>Étape 1</strong> — Charge le fichier T1 (nouvelles références) dans la sidebar.
</div>""", unsafe_allow_html=True)

    st.markdown("""
<div class="info-box blue" style="margin-top:8px">
  <strong>v7 — Nouveau mode d'import stock</strong><br>
  Plus besoin de consolider les fichiers en amont. Charge directement
  les <strong>extractions ERP par magasin</strong> (1 fichier CSV par site).
  Le module les fusionne automatiquement.<br><br>
  Format attendu : <code>CSV ; latin1</code> · nommage <code>*_CODESITE_DATE_*.csv</code><br>
  Colonnes : Site · Libellé site · Code article · Code etat · Nouveau stock · Ral · Pcb · Code marketing · Type saisonnalité
</div>""", unsafe_allow_html=True)

    with st.expander("📋 Colonnes obligatoires dans chaque fichier stock"):
        for col in STOCK_COLS_REQUIRED:
            st.markdown(f"- `{col}`")
    st.stop()


# ─────────────────────────────────────────────────────────────────────────────
# CHARGEMENT T1
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner("Lecture T1…"):
    t1_df, t1_err = parse_t1(t1_file.read(), t1_file.name)

if t1_err or t1_df is None:
    st.error(f"❌ T1 : {t1_err}")
    st.stop()

if not stk_files:
    st.markdown(
        f'<div class="info-box blue">✅ T1 chargé — <strong>{len(t1_df):,}</strong> références. '
        f'⬆️ <strong>Étape 2</strong> — Charge les fichiers stock (1 par magasin).</div>',
        unsafe_allow_html=True,
    )
    st.stop()


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — FILTRES (après chargement T1)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.divider()
    st.markdown("## 🔍 Filtres")

    orig_sel = st.multiselect("Flux", ["IM", "LO"], default=["IM", "LO"])

    sem_dispo = sorted(
        [s for s in t1_df["SEMAINE RECEPTION"].unique()
         if str(s).strip() not in ("nan", "", "99")],
        key=_sem_sort,
    )
    sem_sel = st.multiselect("Semaine réception", sem_dispo, default=sem_dispo)

    st.divider()
    st.markdown("## 🔄 Cessions")
    min_1pcb = st.toggle("Qté min = 1 PCB", value=True)


# ─────────────────────────────────────────────────────────────────────────────
# CONSOLIDATION STOCK
# ─────────────────────────────────────────────────────────────────────────────
SKU_SCOPE = tuple(sorted(t1_df["SKU"].unique()))

with st.spinner(f"Consolidation de {len(stk_files)} fichier(s) stock…"):
    df_stock, stk_errors = consolidate_stock(stk_files, SKU_SCOPE)

for err in stk_errors:
    st.warning(f"⚠️ {err}")

if df_stock.empty:
    st.error("Aucune donnée stock valide après consolidation.")
    st.stop()

# Informations sur les fichiers chargés
n_sites_loaded = df_stock["Code site"].nunique()
dates_stock    = df_stock["_date_fichier"].unique()
date_stock_str = pd.Timestamp(max(dates_stock)).strftime("%d %b %Y") if len(dates_stock) > 0 else "—"
age_stock      = (TODAY - pd.Timestamp(max(dates_stock))).days if len(dates_stock) > 0 else 0

if age_stock > 7:
    st.warning(f"⚠️ Données stock du **{date_stock_str}** — {age_stock} jours. Recharge des fichiers plus récents.")


# ─────────────────────────────────────────────────────────────────────────────
# RÉFÉRENTIEL MAGASINS
# ─────────────────────────────────────────────────────────────────────────────
site_ref  = (df_stock[["Code site", "Libellé site"]]
             .drop_duplicates("Code site")
             .set_index("Code site")["Libellé site"].to_dict())
all_codes = sorted(site_ref.keys())

# Filtres SKU selon semaines/flux sélectionnés
mask_t1 = t1_df["ORIGINE"].isin(orig_sel)
if sem_sel:
    mask_t1 = mask_t1 & t1_df["SEMAINE RECEPTION"].isin(sem_sel)
t1_scope  = t1_df[mask_t1].copy()
sku_scope = t1_scope["SKU"].unique()

if len(sku_scope) == 0:
    st.warning("Aucun SKU pour les filtres sélectionnés.")
    st.stop()


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION CROISÉE
# ─────────────────────────────────────────────────────────────────────────────
sku_dans_stock = set(df_stock[df_stock["Code site"].isin(all_codes)]["SKU"].unique())
n_sku          = len(sku_scope)
n_sku_trouves  = len([s for s in sku_scope if s in sku_dans_stock])
n_sku_absents  = n_sku - n_sku_trouves
n_mag          = len(all_codes)

st.markdown(f"""
<div class="val-box">
  <div style="font-size:13px;font-weight:700;color:{C['text']};margin-right:4px;">📋 T1 × Stock</div>
  <div class="val-item"><div class="val-num" style="color:{C['blue']}">{fmt_n(n_sku)}</div><div class="val-lbl">SKUs T1</div></div>
  <div class="val-item"><div class="val-num" style="color:{C['green']}">{fmt_n(n_sku_trouves)}</div><div class="val-lbl">Trouvés</div></div>
  <div class="val-item"><div class="val-num" style="color:{C['red'] if n_sku_absents > 0 else C['green']}">{fmt_n(n_sku_absents)}</div><div class="val-lbl">Absents stock</div></div>
  <div class="val-item"><div class="val-num" style="color:{C['muted']}">{n_mag}</div><div class="val-lbl">Magasins</div></div>
  <div class="val-item"><div class="val-num" style="color:{C['muted']}">{len(stk_files)}</div><div class="val-lbl">Fichiers chargés</div></div>
  <div style="margin-left:auto;font-size:11px;color:{C['muted']};">
    Données stock : <strong style="color:{C['blue']}">{date_stock_str}</strong>
    {"&nbsp;·&nbsp;<span style='color:#FF3B30'>⚠️ " + str(age_stock) + "j</span>" if age_stock > 7 else ""}
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCTION DATASET
# ─────────────────────────────────────────────────────────────────────────────
stk_filt = df_stock[
    df_stock["Code site"].isin(all_codes) &
    df_stock["SKU"].isin(sku_scope)
].copy()

# Grille SKU × magasin
grid = pd.DataFrame(
    pd.MultiIndex.from_product(
        [all_codes, sku_scope], names=["Code site", "SKU"]
    ).tolist(), columns=["Code site", "SKU"]
)

KEEP = ["Code site", "SKU", "Nouveau stock", "Ral", "Code etat",
        "Origine", "Libellé article", "Nom fourn.",
        "Libellé rayon", "Libellé famille", "Pcb"]
merged = grid.merge(
    stk_filt[[c for c in KEEP if c in stk_filt.columns]],
    on=["Code site", "SKU"], how="left"
)

# Référentiel T1
t1_ref = t1_scope.set_index("SKU")[[
    "LIBELLÉ ARTICLE", "LIBELLÉ FOURNISSEUR ORIGINE",
    "MODE APPRO", "SEMAINE RECEPTION", "DATE LIV.", "ORIGINE", "SEM_NUM"
]].rename(columns={
    "LIBELLÉ ARTICLE":              "T1_lib",
    "LIBELLÉ FOURNISSEUR ORIGINE":  "Fournisseur T1",
    "MODE APPRO":                   "Mode Appro",
    "SEMAINE RECEPTION":            "Sem. Réception",
    "DATE LIV.":                    "Date Livraison",
    "ORIGINE":                      "Origine_T1",
    "SEM_NUM":                      "SEM_NUM",
})
merged = merged.merge(t1_ref.reset_index(), on="SKU", how="left")

# Libellé et flux
merged["Libellé article"] = merged["Libellé article"].fillna("").astype(str)
merged["Libellé article"] = merged.apply(
    lambda r: r["Libellé article"] if r["Libellé article"] else r.get("T1_lib", ""), axis=1
)
merged["Origine"] = merged.apply(
    lambda r: r.get("Origine") if pd.notna(r.get("Origine")) and str(r.get("Origine")) not in ("nan", "")
    else r.get("Origine_T1", "LO"), axis=1
)
merged.drop(columns=["T1_lib", "Origine_T1"], errors="ignore", inplace=True)
merged["Magasin"]    = merged["Code site"].map(site_ref).fillna(merged["Code site"])
merged["Code etat"]  = merged["Code etat"].fillna("").astype(str)
merged["Ral"]        = pd.to_numeric(merged["Ral"], errors="coerce").fillna(0)
merged["Pcb"]        = pd.to_numeric(merged["Pcb"], errors="coerce").fillna(1)

# Alertes
merged["Alerte"] = merged.apply(get_alerte, axis=1)
merged["Action"] = merged["Alerte"].map(ACTION_LABEL)


# ─────────────────────────────────────────────────────────────────────────────
# MÉTRIQUES
# ─────────────────────────────────────────────────────────────────────────────
merged_actif = merged[merged["Code etat"].str.strip() == ETAT_ACTIF]
n_base_taux  = len(merged_actif)
n_impl       = int((merged["Alerte"] == "✅ Implanté").sum())
n_appro      = int((merged["Alerte"] == "🔵 Appro en cours").sum())
n_cmd        = int((merged["Alerte"] == "🛒 Passer commande").sum())
n_anomalie   = int((merged["Alerte"] == "🚩 Anomalie référ.").sum())
taux_reseau  = int(n_impl / n_base_taux * 100) if n_base_taux else 0
n_sku_im     = int((t1_scope["ORIGINE"] == "IM").sum())
n_sku_lo     = int((t1_scope["ORIGINE"] == "LO").sum())
total_cells  = len(merged)

def taux_mag(mag):
    dm = merged_actif[merged_actif["Magasin"] == mag]
    return int((dm["Alerte"] == "✅ Implanté").sum() / len(dm) * 100) if len(dm) else 0

pivot_mag = (
    merged.groupby(["Magasin", "Alerte"]).size()
    .unstack(fill_value=0)
    .reindex(columns=list(ALERTES.keys()), fill_value=0)
    .reset_index()
)
pivot_mag.columns.name = None
pivot_mag["Taux (%)"] = pivot_mag["Magasin"].apply(taux_mag)

rayon_pivot = pd.DataFrame()
if "Libellé rayon" in merged.columns:
    try:
        rayon_pivot = (
            merged_actif.groupby(["Libellé rayon", "Magasin"])
            .apply(lambda x: int((x["Alerte"] == "✅ Implanté").sum() / len(x) * 100) if len(x) > 0 else 0)
            .reset_index(name="Taux (%)")
            .pivot(index="Libellé rayon", columns="Magasin", values="Taux (%)")
            .fillna(0).astype(int)
        )
    except Exception:
        rayon_pivot = pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# BANNIÈRE ALERTES
# ─────────────────────────────────────────────────────────────────────────────
n_actions = n_appro + n_cmd + n_anomalie
if n_actions > 0:
    st.markdown(f"""
<div class="alert-banner">
  <div class="alert-pill">⚡ ACTIONS</div>
  <div class="alert-item"><div class="alert-num" style="color:{C['blue']}">{fmt_n(n_appro)}</div><div class="alert-lbl">Appro en cours</div></div>
  <div class="alert-item"><div class="alert-num" style="color:{C['orange']}">{fmt_n(n_cmd)}</div><div class="alert-lbl">À commander</div></div>
  <div class="alert-item"><div class="alert-num" style="color:#B8860B">{fmt_n(n_anomalie)}</div><div class="alert-lbl">Anomalies</div></div>
  <div style="margin-left:auto;font-size:11px;color:{C['muted']};">{n_mag} mag · {fmt_n(n_sku)} SKU</div>
</div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# KPI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:20px;">
  <div class="kpi-card green">
    <div class="kpi-label">✅ Implanté</div>
    <div class="kpi-value">{fmt_n(n_impl)}</div>
    <div class="kpi-sub">{int(n_impl/total_cells*100) if total_cells else 0}% du réseau</div>
  </div>
  <div class="kpi-card blue">
    <div class="kpi-label">📊 Taux réseau</div>
    <div class="kpi-value" style="color:{color_taux(taux_reseau)}">{taux_reseau}%</div>
    <div class="kpi-sub">sur {fmt_n(n_base_taux)} refs actives</div>
  </div>
  <div class="kpi-card" style="border-top:3px solid {C['blue']};">
    <div class="kpi-label">🔵 Appro en cours</div>
    <div class="kpi-value" style="color:{C['blue']}">{fmt_n(n_appro)}</div>
    <div class="kpi-sub">RAL &gt; 0 · accélérer</div>
  </div>
  <div class="kpi-card orange">
    <div class="kpi-label">🛒 À commander</div>
    <div class="kpi-value">{fmt_n(n_cmd)}</div>
    <div class="kpi-sub">Stock 0 · RAL 0</div>
  </div>
  <div class="kpi-card yellow">
    <div class="kpi-label">🚩 Anomalies</div>
    <div class="kpi-value">{fmt_n(n_anomalie)}</div>
    <div class="kpi-sub">Référencement à vérifier</div>
  </div>
</div>
<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:20px;">
  <div class="kpi-card" style="border-top:3px solid {C['purple']};">
    <div class="kpi-label">🔀 Flux IM / LO</div>
    <div class="kpi-value" style="font-size:24px;color:{C['purple']}">{n_sku_im} / {n_sku_lo}</div>
    <div class="kpi-sub">Import · Local</div>
  </div>
  <div class="kpi-card" style="border-top:3px solid {C['muted']};">
    <div class="kpi-label">📅 Données stock</div>
    <div class="kpi-value" style="font-size:18px;color:{C['red'] if age_stock>7 else C['blue']}">{date_stock_str}</div>
    <div class="kpi-sub">{"⚠️ " + str(age_stock) + "j — recharger" if age_stock > 7 else str(age_stock) + "j · " + str(len(stk_files)) + " fichier(s)"}</div>
  </div>
  <div class="kpi-card" style="border-top:3px solid {C['muted']};">
    <div class="kpi-label">📋 SKUs analysés</div>
    <div class="kpi-value" style="font-size:24px;color:{C['muted']}">{fmt_n(n_sku)}</div>
    <div class="kpi-sub">{n_mag} magasin(s) · {fmt_n(total_cells)} combinaisons</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SCORECARD MAGASINS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="sh">SCORECARD MAGASINS</div>', unsafe_allow_html=True)
sc_html = '<div class="scorecard-grid">'
for _, row in pivot_mag.sort_values("Taux (%)", ascending=False).iterrows():
    t_   = row["Taux (%)"]
    col  = color_taux(t_)
    impl = int(row.get("✅ Implanté", 0))
    app  = int(row.get("🔵 Appro en cours", 0))
    cmd  = int(row.get("🛒 Passer commande", 0))
    an   = int(row.get("🚩 Anomalie référ.", 0))
    sc_html += f"""
<div class="scorecard-card {scorecard_cls(t_)}">
  <div class="scorecard-dot" style="background:{col}"></div>
  <div class="scorecard-name">{row['Magasin']}</div>
  <div class="scorecard-pct" style="color:{col}">{t_}%</div>
  <div class="scorecard-sub">{impl}✅ {app}🔵 {cmd}🛒 {an}🚩</div>
</div>"""
sc_html += "</div>"
st.markdown(sc_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# ONGLETS
# ─────────────────────────────────────────────────────────────────────────────
tab_copil, tab_reseau, tab_alertes, tab_cessions, tab_export = st.tabs([
    "📋 COPIL", "📊 Vue Réseau", "🚨 Alertes", "🔄 Cessions", "📥 Export"
])


# ══════════ COPIL ══════════════════════════════════════════════════════════════
with tab_copil:
    st.markdown('<div class="sh">SYNTHÈSE — TAUX PAR MAGASIN</div>', unsafe_allow_html=True)
    cols_aff = ["Magasin"] + [c for c in ALERTES if c in pivot_mag.columns] + ["Taux (%)"]
    st.dataframe(
        pivot_mag[cols_aff].sort_values("Taux (%)", ascending=False).reset_index(drop=True)
        .style.format({"Taux (%)": "{}%"}),
        use_container_width=True, hide_index=True,
    )

    if not rayon_pivot.empty:
        st.markdown('<div class="sh">TAUX PAR RAYON × MAGASIN</div>', unsafe_allow_html=True)
        def color_cell(val):
            if val >= 80: return "background-color:#D1FAE5;color:#065F46;font-weight:700"
            if val >= 50: return "background-color:#FEF3C7;color:#92400E;font-weight:700"
            return "background-color:#FEE2E2;color:#991B1B;font-weight:700"
        st.dataframe(rayon_pivot.style.map(color_cell).format("{}%"), use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="sh">🛒 TOP ARTICLES À COMMANDER</div>', unsafe_allow_html=True)
        df_cmd = merged[merged["Alerte"] == "🛒 Passer commande"].groupby("SKU").agg(
            Libellé=("Libellé article", "first"),
            Fournisseur=("Fournisseur T1", "first"),
            Origine=("Origine", "first"),
            Nb_magasins=("Magasin", "count"),
        ).reset_index().sort_values("Nb_magasins", ascending=False).head(10)
        if df_cmd.empty:
            st.success("✅ Aucun article sans commande")
        else:
            st.dataframe(df_cmd.rename(columns={"Nb_magasins": "Mag. sans stock"}),
                         use_container_width=True, hide_index=True)

    with c2:
        st.markdown('<div class="sh">🔵 TOP APPROS À ACCÉLÉRER</div>', unsafe_allow_html=True)
        df_acc = merged[merged["Alerte"] == "🔵 Appro en cours"].groupby("SKU").agg(
            Libellé=("Libellé article", "first"),
            Fournisseur=("Fournisseur T1", "first"),
            Origine=("Origine", "first"),
            Nb_magasins=("Magasin", "count"),
            RAL_total=("Ral", "sum"),
        ).reset_index().sort_values("Nb_magasins", ascending=False).head(10)
        if df_acc.empty:
            st.success("✅ Aucune appro en attente")
        else:
            st.dataframe(df_acc.rename(columns={"Nb_magasins": "Mag. en attente", "RAL_total": "RAL total"}),
                         use_container_width=True, hide_index=True)


# ══════════ VUE RÉSEAU ════════════════════════════════════════════════════════
with tab_reseau:
    c1, c2 = st.columns([3, 2])
    with c1:
        alertes_aff = [a for a in ALERTES if a in pivot_mag.columns and a != "✅ Implanté"]
        mel = pivot_mag.melt(
            id_vars="Magasin",
            value_vars=["✅ Implanté"] + alertes_aff,
            var_name="Alerte", value_name="N"
        )
        fig = px.bar(mel, x="Magasin", y="N", color="Alerte",
                     color_discrete_map=ALERTES, barmode="stack",
                     title="Situation par magasin")
        fig.update_layout(paper_bgcolor=C["surface"], plot_bgcolor=C["surface"],
                          height=400, font=dict(family="Inter", size=12),
                          margin=dict(l=10, r=10, t=44, b=20),
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        labels_d = [a for a in ALERTES if a in pivot_mag.columns]
        vals_d   = [int(pivot_mag[a].sum()) for a in labels_d]
        fig_d = go.Figure(go.Pie(
            labels=labels_d, values=vals_d, hole=0.65,
            marker=dict(colors=[ALERTES[a] for a in labels_d],
                        line=dict(color="#fff", width=3)),
            textinfo="percent", textfont=dict(size=12),
        ))
        fig_d.add_annotation(
            text=f"<b>{taux_reseau}%</b><br>implanté",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=20, color=color_taux(taux_reseau))
        )
        fig_d.update_layout(
            paper_bgcolor=C["surface"], height=400,
            margin=dict(l=10, r=10, t=44, b=20),
            legend=dict(orientation="v", x=1.0, y=0.5, font=dict(size=11)),
            title=dict(text="Répartition réseau", font=dict(size=13, color=C["muted"]))
        )
        st.plotly_chart(fig_d, use_container_width=True)

    if not rayon_pivot.empty:
        st.markdown('<div class="sh">TAUX PAR RAYON × MAGASIN</div>', unsafe_allow_html=True)
        def color_cell(val):
            if val >= 80: return "background-color:#D1FAE5;color:#065F46;font-weight:700"
            if val >= 50: return "background-color:#FEF3C7;color:#92400E;font-weight:700"
            return "background-color:#FEE2E2;color:#991B1B;font-weight:700"
        st.dataframe(rayon_pivot.style.map(color_cell).format("{}%"), use_container_width=True)


# ══════════ ALERTES ══════════════════════════════════════════════════════════
with tab_alertes:
    alertes_dispo = [a for a in ALERTES if a != "✅ Implanté" and (merged["Alerte"] == a).any()]
    alerte_sel    = st.multiselect("Filtrer par alerte", alertes_dispo, default=alertes_dispo,
                                    format_func=lambda a: f"{a} — {ACTION_LABEL[a]}")
    mag_alerte_sel = st.multiselect("Filtrer par magasin", sorted(merged["Magasin"].unique()),
                                     default=sorted(merged["Magasin"].unique()))

    # Ruptures communes réseau
    df_non_impl = merged[merged["Alerte"].isin(["🛒 Passer commande", "🔵 Appro en cours"])]
    sku_counts  = df_non_impl.groupby("SKU")["Magasin"].count()
    sku_rupt    = sku_counts[sku_counts == n_mag].index.tolist()

    if sku_rupt:
        st.markdown(f"""
<div style="background:{C['red']}18;border:1px solid {C['red']}44;border-left:4px solid {C['red']};
            border-radius:10px;padding:10px 16px;margin-bottom:8px;
            display:flex;align-items:center;justify-content:space-between;">
  <div>
    <span style="font-size:14px;font-weight:700;">🚨 Ruptures communes réseau</span>
    <span style="font-size:12px;color:{C['muted']};margin-left:12px;">→ Aucun stock sur tous les magasins</span>
  </div>
  <span style="font-size:22px;font-weight:800;color:{C['red']}">{fmt_n(len(sku_rupt))}</span>
</div>""", unsafe_allow_html=True)
        with st.expander(f"Voir les {len(sku_rupt)} ruptures communes"):
            df_rupt = (merged[merged["SKU"].isin(sku_rupt)]
                       [["SKU", "Libellé article", "Origine", "Fournisseur T1", "Alerte"]]
                       .drop_duplicates("SKU").sort_values("Alerte"))
            st.dataframe(df_rupt, use_container_width=True, hide_index=True)

    # Tableau alertes
    df_al = merged[merged["Alerte"].isin(alerte_sel) & merged["Magasin"].isin(mag_alerte_sel)]
    if df_al.empty:
        st.success("✅ Aucune alerte pour les filtres sélectionnés.")
    else:
        for alerte in alerte_sel:
            df_a = df_al[df_al["Alerte"] == alerte]
            if df_a.empty: continue
            color = ALERTES.get(alerte, C["muted"])
            st.markdown(f"""
<div style="background:{color}18;border:1px solid {color}44;border-left:4px solid {color};
            border-radius:10px;padding:10px 16px;margin-bottom:8px;
            display:flex;align-items:center;justify-content:space-between;">
  <div>
    <span style="font-size:14px;font-weight:700;">{alerte}</span>
    <span style="font-size:12px;color:{C['muted']};margin-left:12px;">→ {ACTION_LABEL.get(alerte,'')}</span>
  </div>
  <span style="font-size:22px;font-weight:800;color:{color}">{fmt_n(len(df_a))}</span>
</div>""", unsafe_allow_html=True)

        COLS_AL = ["Magasin", "SKU", "Libellé article", "Origine", "Code etat",
                   "Nouveau stock", "Ral", "Mode Appro", "Sem. Réception",
                   "Fournisseur T1", "Alerte", "Action"]
        st.dataframe(
            df_al[[c for c in COLS_AL if c in df_al.columns]]
            .sort_values(["Alerte", "Magasin"]).reset_index(drop=True),
            use_container_width=True, hide_index=True,
        )


# ══════════ CESSIONS ══════════════════════════════════════════════════════════
with tab_cessions:
    st.markdown('<div class="sh">MOTEUR CESSIONS INTER-MAGASINS</div>', unsafe_allow_html=True)

    # Suggestions auto pour ruptures totales
    df_non_impl_all = merged[merged["Alerte"].isin(["🛒 Passer commande", "🔵 Appro en cours"])]
    sku_rupt_tot = (df_non_impl_all.groupby("SKU")["Magasin"].count()
                    .pipe(lambda s: s[s == n_mag]).index.tolist())

    if sku_rupt_tot:
        st.markdown(f'<div class="info-box blue">🤖 <strong>{len(sku_rupt_tot)} article(s)</strong> sans stock sur tous les magasins — suggestions automatiques.</div>', unsafe_allow_html=True)
        auto_sugg = []
        for sku in sku_rupt_tot:
            sku_df = df_stock[df_stock["SKU"] == sku].copy()
            sku_df["Nouveau stock"] = pd.to_numeric(sku_df["Nouveau stock"], errors="coerce").fillna(0)
            sku_df["Pcb"] = pd.to_numeric(sku_df.get("Pcb", pd.Series(1, index=sku_df.index)), errors="coerce").fillna(1).clip(lower=1)
            best = sku_df.copy()
            best["Reserve_2pcb"] = (best["Pcb"] * 2).astype(int)
            best = best[best["Nouveau stock"] > best["Reserve_2pcb"]].sort_values("Nouveau stock", ascending=False)
            lib = sku_df["Libellé article"].iloc[0] if len(sku_df) else sku
            if not best.empty:
                b   = best.iloc[0]
                qty = int(b["Nouveau stock"]) - int(b["Reserve_2pcb"])
                auto_sugg.append({
                    "SKU": sku, "Libellé": lib,
                    "Cédant": site_ref.get(b["Code site"], b["Code site"]),
                    "Stock cédant": int(b["Nouveau stock"]),
                    "Qté cessible": qty,
                })
        if auto_sugg:
            st.dataframe(pd.DataFrame(auto_sugg), use_container_width=True, hide_index=True)
        else:
            st.info("Aucun magasin cédant disponible.")

    # Cessions manuelles
    with st.sidebar:
        st.divider()
        mag_labels_all = sorted([site_ref.get(c, c) for c in all_codes])
        mag_detresse   = st.multiselect("Magasins en détresse", mag_labels_all, default=[])
        seuil_det      = st.number_input("Seuil stock (≤)", 0, 50, 0, 1)

    if not mag_detresse:
        st.markdown('<div class="info-box blue">⬅️ Sélectionne des magasins en détresse dans la sidebar.</div>', unsafe_allow_html=True)
    else:
        mag_det_codes = [c for c in all_codes if site_ref.get(c, c) in mag_detresse]
        mag_ced_codes = [c for c in all_codes if c not in mag_det_codes]
        suggestions   = []

        for sku in sku_scope:
            sku_df = df_stock[df_stock["SKU"] == sku].copy()
            if sku_df.empty: continue
            lib = sku_df["Libellé article"].iloc[0] if "Libellé article" in sku_df.columns else sku
            sku_df["Nouveau stock"] = pd.to_numeric(sku_df["Nouveau stock"], errors="coerce").fillna(0)
            sku_df["Pcb"] = pd.to_numeric(sku_df.get("Pcb", pd.Series(1, index=sku_df.index)), errors="coerce").fillna(1).clip(lower=1)

            det_rows = sku_df[sku_df["Code site"].isin(mag_det_codes) & (sku_df["Nouveau stock"] <= seuil_det)]
            if det_rows.empty: continue

            sku_df["Reserve_2pcb"] = (sku_df["Pcb"] * 2).astype(int)
            ced_rows = sku_df[sku_df["Code site"].isin(mag_ced_codes)].copy()
            ced_rows = ced_rows[ced_rows["Nouveau stock"] > ced_rows["Reserve_2pcb"]].sort_values("Nouveau stock", ascending=False)

            for _, dr in det_rows.iterrows():
                if ced_rows.empty:
                    suggestions.append({
                        "SKU": sku, "Libellé": lib,
                        "Magasin détresse": site_ref.get(dr["Code site"], dr["Code site"]),
                        "Stock détresse": int(dr["Nouveau stock"]),
                        "Cédant suggéré": "⚠️ Aucun cédant",
                        "Stock cédant": 0, "Qté cessible": 0, "Faisabilité": "🔴 Impossible",
                    })
                else:
                    best        = ced_rows.iloc[0]
                    reserve_art = int(best["Reserve_2pcb"])
                    qty         = int(best["Nouveau stock"]) - reserve_art
                    suggestions.append({
                        "SKU": sku, "Libellé": lib,
                        "Magasin détresse": site_ref.get(dr["Code site"], dr["Code site"]),
                        "Stock détresse": int(dr["Nouveau stock"]),
                        "Cédant suggéré": site_ref.get(best["Code site"], best["Code site"]),
                        "Stock cédant": int(best["Nouveau stock"]),
                        "Réserve (2 PCB)": reserve_art,
                        "Qté cessible": qty,
                        "Faisabilité": "🟢 Possible" if qty >= 1 else "🟠 Partielle",
                    })

        if not suggestions:
            st.success("✅ Aucune cession nécessaire.")
        else:
            df_all  = pd.DataFrame(suggestions)
            df_cess = df_all[df_all["Faisabilité"] == "🟢 Possible"].copy()
            if min_1pcb and "Réserve (2 PCB)" in df_cess.columns:
                df_cess = df_cess[df_cess["Qté cessible"] >= (df_cess["Réserve (2 PCB)"] / 2).clip(lower=1)]
            df_cess = df_cess.sort_values("Qté cessible", ascending=False).reset_index(drop=True)

            k1, k2 = st.columns(2)
            k1.metric("🟢 Cessions possibles", int((df_all["Faisabilité"] == "🟢 Possible").sum()))
            k2.metric("Articles cessibles",     df_cess["SKU"].nunique() if not df_cess.empty else 0)

            st.dataframe(df_cess, use_container_width=True, hide_index=True)

            buf_c = io.BytesIO()
            with pd.ExcelWriter(buf_c, engine="openpyxl") as w:
                df_cess.to_excel(w, sheet_name="Plan Cessions", index=False)
            buf_c.seek(0)
            st.download_button(f"📥 Plan_Cessions_{TODAY_FILE}.xlsx", data=buf_c,
                               file_name=f"Plan_Cessions_{TODAY_FILE}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════ EXPORT ═════════════════════════════════════════════════════════════
with tab_export:
    from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlignment
    from openpyxl.utils import get_column_letter as gcl

    st.markdown('<div class="info-box blue">3 feuilles : <strong>Synthèse réseau</strong> · <strong>Détail complet</strong> · <strong>Alertes & Actions</strong></div>', unsafe_allow_html=True)

    if st.button("🔨 Générer Export", type="primary"):
        ALERTE_FILLS = {
            "✅ Implanté":        ("D1FAE5", "065F46"),
            "🔵 Appro en cours":  ("DBEAFE", "1D4ED8"),
            "🛒 Passer commande": ("FEF3C7", "92400E"),
            "🚩 Anomalie référ.": ("FFFDE7", "795548"),
        }
        COLS_DET = ["Magasin", "SKU", "Libellé article", "Origine", "Code etat",
                    "Nouveau stock", "Ral", "Mode Appro", "Sem. Réception",
                    "Fournisseur T1", "Alerte", "Action"]

        buf_x = io.BytesIO()
        with pd.ExcelWriter(buf_x, engine="openpyxl") as writer:
            cols_s = ["Magasin"] + [c for c in ALERTES if c in pivot_mag.columns] + ["Taux (%)"]
            pivot_mag[cols_s].sort_values("Taux (%)", ascending=False).to_excel(
                writer, sheet_name="Synthèse Réseau", index=False)
            merged[[c for c in COLS_DET if c in merged.columns]].to_excel(
                writer, sheet_name="Détail Complet", index=False)
            merged[merged["Alerte"] != "✅ Implanté"][[c for c in COLS_DET if c in merged.columns]]\
                .sort_values(["Alerte", "Magasin"]).to_excel(writer, sheet_name="Alertes & Actions", index=False)

            wb = writer.book
            FH = XFill("solid", fgColor="1C1C1E")
            FT = XFont(bold=True, color="FFFFFF", name="Arial", size=11)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for cell in ws[1]:
                    cell.fill = FH; cell.font = FT
                    cell.alignment = XAlignment(horizontal="center")
                for col in ws.columns:
                    ws.column_dimensions[gcl(col[0].column)].width = min(
                        max((len(str(c.value)) for c in col if c.value), default=10) + 4, 50)
                ws.freeze_panes = "A2"
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        v = str(cell.value)
                        if v in ALERTE_FILLS:
                            bg, fg = ALERTE_FILLS[v]
                            cell.fill = XFill("solid", fgColor=bg)
                            cell.font = XFont(color=fg, name="Arial", size=10)

        buf_x.seek(0)
        st.download_button(
            f"📥 Implantation_T1_{TODAY_FILE}.xlsx", data=buf_x,
            file_name=f"Implantation_T1_{TODAY_FILE}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success(f"✅ Export — {fmt_n(len(merged))} lignes · 3 feuilles")


# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    f'<div style="text-align:center;font-size:11px;color:{C["muted"]};margin-top:24px;">'
    f'SmartBuyer · v7.0 · Implantation · {TODAY_STR} · Données : {date_stock_str} · {len(stk_files)} fichier(s) chargé(s)'
    f'</div>',
    unsafe_allow_html=True,
)
