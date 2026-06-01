"""
10_📊_Perf_Hebdo.py — SmartBuyer Hub [v4.0]
Performance commerciale hebdomadaire · Orienté Directeur Achats
Structure 2 vitesses :
  - Brief Directeur : synthèse réseau + scorecard rayons + alertes COPIL
  - Brief Acheteur  : par rayon → sous-famille → article + action recommandée
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Perf Hebdo · SmartBuyer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
html,body,[class*="css"]{font-family:-apple-system,BlinkMacSystemFont,"SF Pro Display","Helvetica Neue",Arial,sans-serif!important;background-color:#F2F2F7;}
.stApp{background:#F2F2F7;}
.main .block-container{padding-top:1.8rem;max-width:1300px;}
[data-testid="stSidebar"]{background:#FFFFFF!important;border-right:0.5px solid #E5E5EA!important;}
[data-testid="stMetric"]{background:#FFFFFF!important;border:0.5px solid #E5E5EA!important;border-radius:12px!important;padding:16px 18px!important;}
[data-testid="stMetricLabel"]{font-size:11px!important;font-weight:500!important;color:#8E8E93!important;text-transform:uppercase!important;letter-spacing:0.04em!important;}
[data-testid="stMetricValue"]{font-size:24px!important;font-weight:600!important;color:#1C1C1E!important;}
[data-testid="stTabs"] button[role="tab"]{font-size:13px!important;font-weight:500!important;padding:8px 16px!important;color:#8E8E93!important;border-bottom:2px solid transparent!important;}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"]{color:#007AFF!important;border-bottom:2px solid #007AFF!important;background:transparent!important;}
[data-testid="stTabs"] [role="tablist"]{border-bottom:0.5px solid #E5E5EA!important;}
[data-testid="stDataFrame"]{border:0.5px solid #E5E5EA!important;border-radius:10px!important;}
[data-testid="stDataFrame"] th{background:#F2F2F7!important;font-size:11px!important;font-weight:600!important;color:#8E8E93!important;text-transform:uppercase!important;}
[data-testid="stFileUploader"]{border:1.5px dashed #D1D1D6!important;border-radius:10px!important;background:#F9F9FB!important;}
.stDownloadButton>button{background:#007AFF!important;color:white!important;border:none!important;border-radius:8px!important;font-weight:500!important;font-size:13px!important;padding:10px 24px!important;width:100%!important;}
hr{border-color:#E5E5EA!important;margin:1rem 0!important;}
.page-title{font-size:28px;font-weight:700;color:#1C1C1E;letter-spacing:-0.03em;margin:0;}
.page-caption{font-size:13px;color:#8E8E93;margin-top:3px;margin-bottom:1.5rem;}
.section-label{font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:0.07em;margin-bottom:10px;}
.alert-card{padding:12px 16px;border-radius:10px;margin-bottom:8px;font-size:13px;line-height:1.5;border-left:3px solid;}
.alert-red{background:#FFF2F2;border-color:#FF3B30;color:#3A0000;}
.alert-amber{background:#FFFBF0;border-color:#FF9500;color:#3A2000;}
.alert-green{background:#F0FFF4;border-color:#34C759;color:#003A10;}
.alert-blue{background:#F0F8FF;border-color:#007AFF;color:#001A3A;}
.rayon-card{background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:14px 16px;border-left:3px solid;}
.sfam-block{background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:10px;margin-bottom:8px;overflow:hidden;}
.sfam-header{padding:10px 14px;border-bottom:0.5px solid #E5E5EA;display:flex;align-items:center;justify-content:space-between;}
.sfam-name{font-size:13px;font-weight:600;color:#1C1C1E;}
.sfam-meta{font-size:12px;color:#8E8E93;}
.art-row{padding:7px 14px 7px 24px;border-bottom:0.5px solid #F2F2F7;display:flex;align-items:center;gap:10px;}
.art-row:last-child{border-bottom:none;}
.art-code{font-size:10px;color:#8E8E93;font-family:monospace;min-width:80px;}
.art-name{font-size:12px;color:#1C1C1E;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.art-marge{font-size:12px;font-weight:600;color:#FF3B30;min-width:120px;text-align:right;}
.art-action{font-size:11px;color:#8E8E93;font-style:italic;min-width:140px;text-align:right;}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:10px;font-weight:600;}
.b-ok{background:#E8F8ED;color:#1A7A3A;}
.b-warn{background:#FFF3E0;color:#B45309;}
.b-bad{background:#FFEAEA;color:#C0392B;}
.divider-label{display:flex;align-items:center;gap:8px;margin:1.25rem 0 .75rem;}
.dl-line{flex:1;height:0.5px;background:#E5E5EA;}
.dl-text{font-size:10px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.06em;white-space:nowrap;}
.kpi-bar{background:#FFFFFF;border-radius:14px;border:0.5px solid #E5E5EA;padding:0.85rem 1.25rem;text-align:center;}
.kpi-bar-val{font-size:20px;font-weight:700;color:#1C1C1E;}
.kpi-bar-label{font-size:11px;color:#8E8E93;margin-top:2px;}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
RAYON_MAP = {
    "00014 - EPICERIE":           "Épicerie",
    "00010 - BOISSONS":           "Boissons",
    "00012 - PARFUMERIE HYGIENE": "DPH",
    "00011 - DROGUERIE":          "DPH",
}
COLORS = {"Épicerie": "#FF9500", "Boissons": "#007AFF", "DPH": "#AF52DE"}
# Seuils de marge par défaut par rayon
DEFAULT_SEUILS = {"Épicerie": 12.0, "Boissons": 14.0, "DPH": 16.0}

def fmt_fcfa(n):
    if pd.isna(n): return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v):
    if pd.isna(v): return "—"
    return f"{v*100:.1f}%" if abs(v) < 10 else f"{v:.1f}%"

def clean_label(s):
    if pd.isna(s): return ""
    m = re.match(r"^\d+ - (.+)$", str(s))
    return m.group(1).strip() if m else str(s).strip()

def get_action(row):
    """Recommandation action acheteur basée sur contexte."""
    pct = row.get("pct_marge", 0) or 0
    pct_promo = row.get("pct_marge_promo", None)
    pct_hp    = row.get("pct_marge_hp", None)
    if pct_promo is not None and pct_hp is not None and not pd.isna(pct_promo) and not pd.isna(pct_hp):
        if (pct_hp - pct_promo) > 0.05:
            return "Arrêter / revoir promo"
    if pct < 0:
        return "Renégocier PA urgent"
    if pct < 0.05:
        return "Vérifier conditions fournisseur"
    return "Surveiller"

# ─── PARSING ──────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_file(file_bytes, filename):
    if filename.endswith(".csv"):
        df = pd.read_csv(BytesIO(file_bytes), encoding="latin-1")
    else:
        df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    df.columns = df.columns.str.strip()

    # Normaliser CA Promo
    if "CA Promo" in df.columns and "CA HT Promo" not in df.columns:
        df = df.rename(columns={"CA Promo": "CA HT Promo"})
    if "Marge Promo" in df.columns and "%Marge Promo" in df.columns:
        pass  # OK

    num_cols = ["CA", "Marge", "%Marge", "CA HT Promo", "Marge Promo", "%Marge Promo",
                "CA Hors Promo", "Marge Hors Promo", "%Marge Hors Promo",
                "%CA Poids Promo", "Qté Vente", "Casse (Valeur)", "Casse (Qté)"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Colonnes optionnelles
    for col in ["CA HT Promo", "Marge Promo", "%Marge Promo", "CA Hors Promo",
                "Marge Hors Promo", "%Marge Hors Promo", "%CA Poids Promo",
                "Casse (Valeur)", "Casse (Qté)", "Sous Famille"]:
        if col not in df.columns:
            df[col] = 0 if col != "Sous Famille" else ""

    # ── Lignes article niveau total réseau ────────────────────────────────────
    if "Site nom long" in df.columns:
        arts = df[
            df["Article"].notna() &
            (df["Article"].astype(str).str.strip() != "Total") &
            (df["Site nom long"].astype(str).str.strip() == "Total")
        ].copy()
    else:
        arts = df[df["Article"].notna() & (df["Article"].astype(str).str.strip() != "Total")].copy()

    arts["art_label"]    = arts["Article"].apply(clean_label)
    arts["art_code"]     = arts["Article"].apply(
        lambda s: str(s).split(" - ", 1)[0].strip() if " - " in str(s) else "")
    arts["rayon_label"]  = arts["Rayon"].apply(
        lambda x: RAYON_MAP.get(str(x).strip(), clean_label(x)))
    arts["sfam_label"]   = arts["Sous Famille"].apply(clean_label)
    arts["fam_label"]    = arts["Famille"].apply(clean_label)

    # Taux de marge calculés
    arts["pct_marge"]       = np.where(arts["CA"] > 0, arts["Marge"] / arts["CA"], np.nan)
    arts["pct_marge_hp"]    = np.where(arts["CA Hors Promo"] > 0,
                                        arts["Marge Hors Promo"] / arts["CA Hors Promo"], np.nan)
    arts["pct_marge_promo"] = np.where(arts["CA HT Promo"] > 0,
                                        arts["Marge Promo"] / arts["CA HT Promo"], np.nan)

    # ── Totaux rayon ──────────────────────────────────────────────────────────
    rayon_tots = arts.groupby("rayon_label", as_index=False).agg(
        CA=("CA", "sum"), Marge=("Marge", "sum"),
        CA_HP=("CA Hors Promo", "sum"), Marge_HP=("Marge Hors Promo", "sum"),
        CA_Promo=("CA HT Promo", "sum"), Marge_Promo=("Marge Promo", "sum"),
        Casse=("Casse (Valeur)", "sum"),
        nb_arts=("art_label", "count"),
        nb_neg=("Marge", lambda x: (x < 0).sum()),
    )
    rayon_tots["pct_marge"]       = rayon_tots["Marge"] / rayon_tots["CA"].replace(0, np.nan)
    rayon_tots["pct_marge_hp"]    = rayon_tots["Marge_HP"] / rayon_tots["CA_HP"].replace(0, np.nan)
    rayon_tots["pct_marge_promo"] = rayon_tots["Marge_Promo"] / rayon_tots["CA_Promo"].replace(0, np.nan)
    rayon_tots["ecart_promo"]     = rayon_tots["pct_marge_hp"] - rayon_tots["pct_marge_promo"]

    # ── Totaux sous-famille ───────────────────────────────────────────────────
    sfam_tots = arts[arts["Marge"] < 0].groupby(
        ["rayon_label", "sfam_label"], as_index=False
    ).agg(
        CA=("CA", "sum"), Marge=("Marge", "sum"),
        CA_HP=("CA Hors Promo", "sum"), Marge_HP=("Marge Hors Promo", "sum"),
        CA_Promo=("CA HT Promo", "sum"), Marge_Promo=("Marge Promo", "sum"),
        nb_arts=("art_label", "count"),
    )
    sfam_tots["pct_marge"]       = sfam_tots["Marge"] / sfam_tots["CA"].replace(0, np.nan)
    sfam_tots["pct_marge_hp"]    = sfam_tots["Marge_HP"] / sfam_tots["CA_HP"].replace(0, np.nan)
    sfam_tots["pct_marge_promo"] = sfam_tots["Marge_Promo"] / sfam_tots["CA_Promo"].replace(0, np.nan)
    sfam_tots["ecart_promo"]     = sfam_tots["pct_marge_hp"] - sfam_tots["pct_marge_promo"]
    sfam_tots = sfam_tots.sort_values("Marge")

    return arts, rayon_tots, sfam_tots


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichier</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Extraction PBI", type=["xlsx", "xls", "csv"],
                                 key="pbi", label_visibility="collapsed")

# ─── PAGE ─────────────────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📊 Performance Hebdomadaire</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Brief Directeur · Brief Acheteur · Alertes jusqu'à l'article</div>", unsafe_allow_html=True)

if not uploaded:
    st.markdown("---")
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ Rapport orienté Directeur Achats</strong><br>
  <strong>Brief Directeur</strong> — synthèse réseau en 30 secondes · scorecard rayons · alertes COPIL<br>
  <strong>Brief Acheteur</strong> — par rayon → sous-famille → article · action recommandée<br><br>
  Charge l'export PBI hebdomadaire dans la sidebar pour lancer l'analyse.
</div>""", unsafe_allow_html=True)
    st.info("⬆️ Charge ton extraction PBI dans la sidebar.")
    st.stop()

# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
with st.spinner("Lecture et analyse…"):
    arts, rayon_tots, sfam_tots = parse_file(uploaded.getvalue(), uploaded.name)

if arts.empty:
    st.error("Impossible de lire les données."); st.stop()

# ─── SEUILS SIDEBAR ───────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Seuils taux de marge (%)</div>", unsafe_allow_html=True)
    seuils = {}
    for rayon in ["Épicerie", "Boissons", "DPH"]:
        if rayon in arts["rayon_label"].unique():
            seuils[rayon] = st.number_input(
                rayon, 0.0, 50.0,
                DEFAULT_SEUILS.get(rayon, 12.0), 0.5,
                key=f"seuil_{rayon}"
            )
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Seuil alerte casse</div>", unsafe_allow_html=True)
    seuil_casse_pct = st.number_input("Casse > % du CA", 0.0, 10.0, 2.0, 0.5, key="seuil_casse")

# ─── CALCULS GLOBAUX ──────────────────────────────────────────────────────────
ca_tot     = arts["CA"].sum()
marge_tot  = arts["Marge"].sum()
pct_tot    = marge_tot / ca_tot if ca_tot > 0 else 0
casse_tot  = arts["Casse (Valeur)"].sum()
nb_neg     = int((arts["Marge"] < 0).sum())
impact_neg = arts[arts["Marge"] < 0]["Marge"].sum()

# ─── BRIEF DIRECTEUR ──────────────────────────────────────────────────────────
st.markdown("<div class='divider-label'><div class='dl-line'></div><div class='dl-text'>Brief Directeur — lecture 30 secondes</div><div class='dl-line'></div></div>", unsafe_allow_html=True)

# KPIs réseau
k1, k2, k3, k4 = st.columns(4)
pct_color = "#34C759" if pct_tot >= 0.12 else "#FF9500" if pct_tot >= 0.10 else "#FF3B30"
for col, val, label, sub in [
    (k1, fmt_fcfa(ca_tot) + " FCFA", "CA réseau", f"{len(arts):,} articles"),
    (k2, f"{pct_tot*100:.1f}%",      "Taux de marge", "seuil réseau 12%"),
    (k3, fmt_fcfa(abs(impact_neg)) + " FCFA", "Impact marges nég.", f"{nb_neg} articles"),
    (k4, fmt_fcfa(abs(casse_tot)) + " FCFA", "Casse réseau", f"{int((arts['Casse (Valeur)'] < 0).sum())} articles"),
]:
    col.markdown(f'<div class="kpi-bar"><div class="kpi-bar-val">{val}</div><div class="kpi-bar-label">{label}</div><div style="font-size:10px;color:#8E8E93">{sub}</div></div>', unsafe_allow_html=True)

st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# ─── SCORECARD RAYONS ─────────────────────────────────────────────────────────
st.markdown("<div class='section-label'>Scorecard rayons — taux de marge vs seuil</div>", unsafe_allow_html=True)

cols_r = st.columns(len(rayon_tots))
for i, (_, row) in enumerate(rayon_tots.iterrows()):
    rayon  = row["rayon_label"]
    color  = COLORS.get(rayon, "#8E8E93")
    seuil  = seuils.get(rayon, 12.0) / 100
    pm     = row["pct_marge"] or 0
    pm_hp  = row["pct_marge_hp"] or 0
    pm_pr  = row["pct_marge_promo"] or 0
    ecart  = row["ecart_promo"] or 0
    delta  = pm - seuil

    if pm >= seuil:           badge_cls, badge_txt = "b-ok",   f"✓ {pm*100:.1f}%"
    elif pm >= seuil * 0.85:  badge_cls, badge_txt = "b-warn", f"⚠ {pm*100:.1f}%"
    else:                     badge_cls, badge_txt = "b-bad",  f"✗ {pm*100:.1f}%"

    bar_w = min(int(pm / seuil * 100), 100) if seuil > 0 else 0
    note_color = "#34C759" if pm >= seuil else "#FF9500" if pm >= seuil * 0.85 else "#FF3B30"
    note = f"{int(row['nb_neg'])} art. marge nég."
    if abs(ecart) > 0.05:
        note += f" · promo −{ecart*100:.1f} pt vs HP"

    with cols_r[i]:
        st.markdown(f"""
<div class='rayon-card' style='border-left-color:{color}'>
  <div style='display:flex;align-items:center;justify-content:space-between;margin-bottom:10px'>
    <span style='font-size:13px;font-weight:600;color:{color}'>{rayon}</span>
    <span class='badge {badge_cls}'>{badge_txt}</span>
  </div>
  <div style='display:grid;grid-template-columns:1fr 1fr;gap:4px 10px;font-size:11px;margin-bottom:8px'>
    <div style='color:#8E8E93'>CA</div><div style='font-weight:600'>{fmt_fcfa(row['CA'])} FCFA</div>
    <div style='color:#8E8E93'>Marge HP</div><div style='font-weight:600;color:{"#34C759" if pm_hp >= seuil else "#FF9500"}'>{pm_hp*100:.1f}%</div>
    <div style='color:#8E8E93'>Marge promo</div><div style='font-weight:600;color:{"#FF3B30" if pm_pr < seuil * 0.5 else "#FF9500"}'>{pm_pr*100:.1f}%</div>
    <div style='color:#8E8E93'>Écart promo</div><div style='font-weight:600;color:{"#FF3B30" if ecart > 0.05 else "#8E8E93"}'>{"−" if ecart > 0 else "+"}{abs(ecart)*100:.1f} pt</div>
  </div>
  <div style='height:3px;background:#E5E5EA;border-radius:2px;margin-bottom:6px'>
    <div style='height:3px;width:{bar_w}%;background:{color};border-radius:2px'></div>
  </div>
  <div style='font-size:11px;color:{note_color}'>{note}</div>
</div>""", unsafe_allow_html=True)

# ─── ALERTES COPIL ────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Alertes COPIL — signaux prioritaires</div>", unsafe_allow_html=True)

alerts_copil = []

# 1. Rayons sous seuil
for _, row in rayon_tots.iterrows():
    rayon = row["rayon_label"]
    seuil = seuils.get(rayon, 12.0) / 100
    pm    = row["pct_marge"] or 0
    if pm < seuil:
        delta = (pm - seuil) * 100
        impact = arts[arts["rayon_label"] == rayon]["Marge"].sum() - \
                 arts[arts["rayon_label"] == rayon]["CA"].sum() * seuil
        cls  = "alert-red" if pm < seuil * 0.85 else "alert-amber"
        ico  = "🔴" if pm < seuil * 0.85 else "⚠️"
        alerts_copil.append((cls, ico,
            f"{rayon} — taux de marge {pm*100:.1f}% vs seuil {seuil*100:.0f}% ({delta:+.1f} pt)",
            f"Impact : {fmt_fcfa(abs(impact))} FCFA de marge non réalisée · {int(row['nb_neg'])} articles en perte",
        ))

# 2. Promos destructrices
for _, row in rayon_tots.iterrows():
    rayon = row["rayon_label"]
    ecart = row["ecart_promo"] or 0
    if ecart > 0.07 and row["CA_Promo"] > 0:
        alerts_copil.append(("alert-amber", "⚠️",
            f"Promo destructrice {rayon} — écart {ecart*100:.1f} pts (HP {row['pct_marge_hp']*100:.1f}% vs promo {row['pct_marge_promo']*100:.1f}%)",
            f"La mécanique promo dégrade la marge de {ecart*100:.1f} pts · revoir conditions avant renouvellement",
        ))

# 3. Casse anormale
pct_casse = abs(casse_tot) / ca_tot * 100 if ca_tot > 0 else 0
if pct_casse > seuil_casse_pct:
    alerts_copil.append(("alert-amber", "⚠️",
        f"Casse réseau {pct_casse:.2f}% du CA — {fmt_fcfa(abs(casse_tot))} FCFA",
        "Revoir DLC, conditions de stockage et politiques de commande sur les articles les plus exposés",
    ))

if not alerts_copil:
    st.markdown("<div class='alert-card alert-green'>✅ Aucune alerte critique cette semaine — tous les rayons au-dessus des seuils.</div>", unsafe_allow_html=True)
else:
    for cls, ico, titre, detail in alerts_copil:
        st.markdown(f"""
<div class='alert-card {cls}'>
  <strong>{ico} {titre}</strong><br>
  <span style='font-size:12px;opacity:.85'>→ {detail}</span>
</div>""", unsafe_allow_html=True)

# ─── BRIEF ACHETEUR ───────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='divider-label'><div class='dl-line'></div><div class='dl-text'>Brief Acheteur — rayon → sous-famille → article → action</div><div class='dl-line'></div></div>", unsafe_allow_html=True)

tab_rayons = [r for r in ["Épicerie", "Boissons", "DPH"] if r in arts["rayon_label"].unique()]
if tab_rayons:
    tabs = st.tabs([f"{COLORS.get(r,'#8E8E93') and r}" for r in tab_rayons])
    for tab, rayon in zip(tabs, tab_rayons):
        with tab:
            seuil = seuils.get(rayon, 12.0) / 100
            arts_r = arts[arts["rayon_label"] == rayon]
            sfam_r = sfam_tots[sfam_tots["rayon_label"] == rayon].copy()

            if sfam_r.empty:
                st.markdown("<div class='alert-card alert-green'>✅ Aucune sous-famille en marge négative sur ce rayon.</div>", unsafe_allow_html=True)
                continue

            color = COLORS.get(rayon, "#8E8E93")
            # Résumé rayon
            row_r = rayon_tots[rayon_tots["rayon_label"] == rayon].iloc[0]
            pm_r  = row_r["pct_marge"] or 0
            st.markdown(f"""
<div class='alert-card {"alert-red" if pm_r < seuil * 0.85 else "alert-amber" if pm_r < seuil else "alert-green"}'>
  <strong>{rayon} — {pm_r*100:.1f}% de marge · seuil {seuil*100:.0f}%</strong>
  · CA {fmt_fcfa(row_r['CA'])} FCFA · {int(row_r['nb_neg'])} articles en marge négative · Impact {fmt_fcfa(abs(arts_r[arts_r['Marge']<0]['Marge'].sum()))} FCFA
</div>""", unsafe_allow_html=True)

            # Filtres inline
            fc1, fc2 = st.columns(2)
            with fc1:
                sel_sfam = st.selectbox("Sous-famille",
                    ["Toutes"] + sorted(sfam_r["sfam_label"].unique().tolist()),
                    key=f"sf_{rayon}")
            with fc2:
                sel_action = st.selectbox("Action",
                    ["Toutes", "Renégocier PA urgent", "Arrêter / revoir promo",
                     "Vérifier conditions fournisseur", "Surveiller"],
                    key=f"ac_{rayon}")

            if sel_sfam != "Toutes":
                sfam_r = sfam_r[sfam_r["sfam_label"] == sel_sfam]

            # Affichage par sous-famille
            for _, sf_row in sfam_r.iterrows():
                sfam = sf_row["sfam_label"]
                arts_sf = arts_r[
                    (arts_r["sfam_label"] == sfam) &
                    (arts_r["Marge"] < 0)
                ].sort_values("Marge").head(10)

                if arts_sf.empty:
                    continue

                # Calculer action dominante
                arts_sf = arts_sf.copy()
                arts_sf["action_rec"] = arts_sf.apply(get_action, axis=1)

                if sel_action != "Toutes" and sel_action not in arts_sf["action_rec"].values:
                    continue

                pm_sf   = sf_row["pct_marge"] or 0
                ecart_sf = sf_row["ecart_promo"] or 0
                badge_cl = "b-bad" if pm_sf < 0 else "b-warn"
                action_badge = "action-promo" if ecart_sf > 0.05 else "action-pa"
                action_label = "Promo à revoir" if ecart_sf > 0.05 else "Vérifier PA"

                st.markdown(f"""
<div class='sfam-block'>
  <div class='sfam-header'>
    <div>
      <span class='sfam-name'>{sfam}</span>
      <span style='font-size:10px;color:#8E8E93;margin-left:8px'>{int(sf_row['nb_arts'])} art. · CA {fmt_fcfa(sf_row['CA'])} FCFA</span>
    </div>
    <div style='display:flex;align-items:center;gap:8px'>
      <span class='badge {badge_cl}'>{pm_sf*100:.1f}%  ·  {fmt_fcfa(sf_row["Marge"])} FCFA</span>
      <span style='font-size:10px;padding:2px 8px;border-radius:20px;background:{"#FFF3E0" if action_badge=="action-promo" else "#FFEAEA"};color:{"#B45309" if action_badge=="action-promo" else "#C0392B"}'>{action_label}</span>
    </div>
  </div>""", unsafe_allow_html=True)

                for _, art in arts_sf.iterrows():
                    if sel_action != "Toutes" and art["action_rec"] != sel_action:
                        continue
                    pm_a   = art["pct_marge"] or 0
                    pm_hp  = art.get("pct_marge_hp") or 0
                    pm_pr  = art.get("pct_marge_promo") or 0
                    detail = ""
                    if pm_hp > 0 and pm_pr < pm_hp - 0.05:
                        detail = f" · HP {pm_hp*100:.1f}% vs promo {pm_pr*100:.1f}%"
                    st.markdown(f"""
<div class='art-row'>
  <span class='art-code'>{art['art_code']}</span>
  <span class='art-name'>{art['art_label']}</span>
  <span class='art-marge'>{pm_a*100:.1f}% · {fmt_fcfa(art['Marge'])} FCFA{detail}</span>
  <span class='art-action'>→ {art['action_rec']}</span>
</div>""", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)

# ─── CLASSEMENTS GLOBAUX ──────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='divider-label'><div class='dl-line'></div><div class='dl-text'>Classements réseau</div><div class='dl-line'></div></div>", unsafe_allow_html=True)

rayon_opts = ["Tous rayons"] + [r for r in ["Épicerie","Boissons","DPH"] if r in arts["rayon_label"].unique()]
rayon_filtre = st.segmented_control("Rayon", rayon_opts, default="Tous rayons", label_visibility="collapsed")
arts_f = arts if rayon_filtre == "Tous rayons" else arts[arts["rayon_label"] == rayon_filtre]

RENAME = {
    "art_label": "Article", "rayon_label": "Rayon", "sfam_label": "Sous-famille",
    "CA": "CA (FCFA)", "Marge": "Marge (FCFA)", "pct_marge": "% Marge",
    "Qté Vente": "Qté vendue", "CA HT Promo": "CA Promo (FCFA)",
    "Marge Promo": "Marge Promo (FCFA)", "%CA Poids Promo": "Poids Promo",
    "Casse (Valeur)": "Casse valeur (FCFA)", "Casse (Qté)": "Casse qté",
    "pct_marge_hp": "% Marge HP", "pct_marge_promo": "% Marge Promo",
}

tab1, tab2, tab3, tab4, tab5 = st.tabs(["🏆 Top CA","💚 Top Marge","🎯 Top Promo","🔴 Marges Nég.","🗑️ Casse"])

def show_df(df, cols, pct_cols=()):
    d = df[cols].copy().rename(columns=RENAME)
    d.index = range(1, len(d)+1)
    for c in pct_cols:
        c2 = RENAME.get(c, c)
        if c2 in d.columns:
            d[c2] = d[c2].apply(lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—")
    for c in ["CA (FCFA)","Marge (FCFA)","CA Promo (FCFA)","Marge Promo (FCFA)","Casse valeur (FCFA)"]:
        if c in d.columns:
            d[c] = d[c].apply(lambda v: fmt_fcfa(v) if pd.notna(v) else "—")
    st.dataframe(d, use_container_width=True, height=380, hide_index=False)

with tab1:
    show_df(arts_f.nlargest(10,"CA"),
            ["art_label","rayon_label","sfam_label","CA","Marge","pct_marge","Qté Vente"],
            pct_cols=["pct_marge"])
with tab2:
    show_df(arts_f.nlargest(10,"Marge"),
            ["art_label","rayon_label","sfam_label","CA","Marge","pct_marge"],
            pct_cols=["pct_marge"])
with tab3:
    df_pr = arts_f[arts_f["CA HT Promo"]>0].nlargest(10,"CA HT Promo")
    if df_pr.empty:
        st.info("Aucun article promotionnel.")
    else:
        show_df(df_pr,
                ["art_label","rayon_label","sfam_label","CA HT Promo","Marge Promo","pct_marge_hp","pct_marge_promo"],
                pct_cols=["pct_marge_hp","pct_marge_promo"])
with tab4:
    df_fl = arts_f[arts_f["Marge"]<0].nsmallest(20,"Marge")
    st.warning(f"⚠️ {len(df_fl)} articles en marge négative · Impact {fmt_fcfa(abs(df_fl['Marge'].sum()))} FCFA")
    show_df(df_fl,
            ["art_label","rayon_label","sfam_label","CA","Marge","pct_marge","pct_marge_hp","pct_marge_promo"],
            pct_cols=["pct_marge","pct_marge_hp","pct_marge_promo"])
with tab5:
    df_cs = arts_f[arts_f["Casse (Valeur)"]<0].nsmallest(10,"Casse (Valeur)")
    if df_cs.empty:
        st.info("Aucune casse enregistrée.")
    else:
        show_df(df_cs,
                ["art_label","rayon_label","sfam_label","Casse (Valeur)","Casse (Qté)"],
                pct_cols=[])

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
st.markdown("---")
with st.expander("📥 Export Excel — Brief Directeur · Brief Acheteur · Classements"):
    if st.button("Générer le fichier Excel", type="primary"):
        with st.spinner("Génération…"):
            wb  = Workbook()
            HDR = "1C3557"; SUB = "2E4B7A"
            C_OK = "D6F0D6"; C_WN = "FEF3CD"; C_KO = "FCE4E4"; C_EVN = "FFFFFF"; C_ODD = "F7F7F7"
            def f(h):  return PatternFill("solid", fgColor=h.replace("#",""))
            def bdr(): s=Side(style="thin",color="DDDDDD"); return Border(left=s,right=s,top=s,bottom=s)
            CTR = Alignment(horizontal="center",vertical="center")
            LFT = Alignment(horizontal="left",  vertical="center")
            RGT = Alignment(horizontal="right",  vertical="center")
            HF  = Font(bold=True,color="FFFFFF",name="Calibri",size=10)
            NF  = Font(name="Calibri",size=10,color="1C1C1E")
            BF  = Font(bold=True,name="Calibri",size=10,color="1C1C1E")

            def hrow(ws, r, headers, widths, bg=SUB):
                for i,(h,w) in enumerate(zip(headers,widths),1):
                    c=ws.cell(row=r,column=i,value=h)
                    c.fill=f(bg);c.font=HF;c.alignment=CTR;c.border=bdr()
                    ws.column_dimensions[get_column_letter(i)].width=w
                ws.row_dimensions[r].height=22
                ws.freeze_panes=f"A{r+1}"

            def drow(ws, r, vals, fmts, is_neg=False):
                bg_h = C_ODD if r%2==0 else C_EVN
                for i,(v,fmt) in enumerate(zip(vals,fmts),1):
                    c=ws.cell(row=r,column=i,value=v)
                    c.fill=f(bg_h);c.border=bdr()
                    if fmt: c.number_format=fmt
                    if isinstance(v,float) and is_neg and v<0:
                        c.font=Font(bold=True,name="Calibri",size=10,color="FF3B30")
                    else: c.font=NF
                    c.alignment=RGT if isinstance(v,(int,float)) else LFT
                ws.row_dimensions[r].height=18

            # ── Onglet 1 : Brief Directeur ────────────────────────────────────
            ws1=wb.active; ws1.title="Brief Directeur"

            # Titre
            ws1.merge_cells("A1:H1")
            ct=ws1["A1"]; ct.value="PERFORMANCE COMMERCIALE HEBDOMADAIRE — BRIEF DIRECTEUR"
            ct.font=Font(bold=True,color="FFFFFF",name="Calibri",size=13)
            ct.fill=f(HDR); ct.alignment=CTR; ws1.row_dimensions[1].height=28

            # KPIs
            kpi_data=[
                ("CA RÉSEAU",f"{ca_tot:,.0f} FCFA","#007AFF"),
                ("TAUX DE MARGE",f"{pct_tot*100:.1f}%","#34C759" if pct_tot>=0.12 else "#FF9500" if pct_tot>=0.10 else "#FF3B30"),
                ("IMPACT MARGES NÉG.",f"{abs(impact_neg):,.0f} FCFA","#FF3B30"),
                ("CASSE RÉSEAU",f"{abs(casse_tot):,.0f} FCFA","#FF9500"),
            ]
            for i,(lbl,val,col) in enumerate(kpi_data,1):
                cl=ws1.cell(row=2,column=i*2-1,value=lbl)
                cl.fill=f(SUB);cl.font=Font(bold=True,color="FFFFFF",name="Calibri",size=9)
                cl.alignment=CTR;cl.border=bdr()
                ws1.merge_cells(start_row=2,start_column=i*2-1,end_row=2,end_column=i*2)
                cv=ws1.cell(row=3,column=i*2-1,value=val)
                cv.fill=f(col.replace("#",""));cv.font=Font(bold=True,name="Calibri",size=13,color="1C1C1E")
                cv.alignment=CTR;cv.border=bdr()
                ws1.merge_cells(start_row=3,start_column=i*2-1,end_row=3,end_column=i*2)
                for ci in [i*2-1,i*2]: ws1.column_dimensions[get_column_letter(ci)].width=20
            ws1.row_dimensions[2].height=18; ws1.row_dimensions[3].height=30; ws1.row_dimensions[4].height=10

            # Scorecard rayons
            ws1.merge_cells("A5:H5")
            c=ws1["A5"]; c.value="  SCORECARD RAYONS — TAUX DE MARGE VS SEUIL"
            c.font=Font(bold=True,color="AABBCC",name="Calibri",size=9,italic=True)
            c.fill=f(HDR); c.alignment=LFT; ws1.row_dimensions[5].height=16

            hrow(ws1,6,["Rayon","CA (FCFA)","Marge (FCFA)","Taux marge","Seuil","Écart","Marge HP","Marge Promo","Écart HP-Promo","Art. marge nég.","Statut"],
                 [16,16,16,12,10,10,12,14,16,16,14])
            for ri,(_, row) in enumerate(rayon_tots.iterrows(),7):
                seuil=seuils.get(row["rayon_label"],12.0)/100
                pm=row["pct_marge"] or 0
                ecart=pm-seuil
                bg_col=C_OK if pm>=seuil else C_WN if pm>=seuil*0.85 else C_KO
                vals=[row["rayon_label"],row["CA"],row["Marge"],pm,seuil,ecart,
                      row["pct_marge_hp"] or 0,row["pct_marge_promo"] or 0,
                      row["ecart_promo"] or 0,int(row["nb_neg"]),
                      "✅ OK" if pm>=seuil else "⚠️ Surveiller" if pm>=seuil*0.85 else "🔴 Action"]
                fmts=[None,"#,##0","#,##0","0.0%","0.0%","+0.0%","0.0%","0.0%","0.0%","#,##0",None]
                for i,(v,fmt) in enumerate(vals,1):
                    c=ws1.cell(row=ri,column=i,value=v)
                    c.fill=f(bg_col) if i in [4,6,11] else f(C_ODD if ri%2==0 else C_EVN)
                    c.font=NF;c.border=bdr()
                    if fmt: c.number_format=fmt
                    c.alignment=RGT if isinstance(v,(int,float)) else LFT
                ws1.row_dimensions[ri].height=20

            # Alertes COPIL
            cur=7+len(rayon_tots)+1
            ws1.row_dimensions[cur].height=10; cur+=1
            ws1.merge_cells(start_row=cur,start_column=1,end_row=cur,end_column=11)
            c=ws1.cell(row=cur,column=1,value="  ALERTES COPIL")
            c.font=Font(bold=True,color="AABBCC",name="Calibri",size=9,italic=True)
            c.fill=f(HDR); c.alignment=LFT; ws1.row_dimensions[cur].height=16; cur+=1
            hrow(ws1,cur,["Niveau","Alerte","Action"],[ 10,60,50],SUB); cur+=1
            for cls,ico,titre,detail in alerts_copil:
                bg_a="FCE4E4" if "red" in cls else "FEF3CD"
                for i,(v,w) in enumerate([(ico,8),(titre,60),(detail,50)],1):
                    c=ws1.cell(row=cur,column=i,value=v)
                    c.fill=f(bg_a);c.font=NF;c.border=bdr();c.alignment=LFT
                    ws1.column_dimensions[get_column_letter(i)].width=w
                ws1.row_dimensions[cur].height=20; cur+=1

            # ── Onglets Brief Acheteur par rayon ──────────────────────────────
            for rayon in tab_rayons:
                color_hex = COLORS.get(rayon,"555555").replace("#","")
                arts_r = arts[arts["rayon_label"]==rayon]
                sfam_r2 = sfam_tots[sfam_tots["rayon_label"]==rayon].copy()
                if sfam_r2.empty: continue

                ws=wb.create_sheet(f"Acheteur {rayon[:8]}")
                ws.merge_cells("A1:J1")
                ct=ws["A1"]; ct.value=f"BRIEF ACHETEUR — {rayon.upper()}"
                ct.font=Font(bold=True,color="FFFFFF",name="Calibri",size=13)
                ct.fill=PatternFill("solid",fgColor=color_hex); ct.alignment=CTR; ws.row_dimensions[1].height=28

                cur2=3
                for _,sf in sfam_r2.iterrows():
                    sfam=sf["sfam_label"]
                    arts_sf2=arts_r[(arts_r["sfam_label"]==sfam)&(arts_r["Marge"]<0)].sort_values("Marge").head(15)
                    if arts_sf2.empty: continue
                    arts_sf2=arts_sf2.copy()
                    arts_sf2["action_rec"]=arts_sf2.apply(get_action,axis=1)

                    # Header sous-famille
                    ws.merge_cells(start_row=cur2,start_column=1,end_row=cur2,end_column=9)
                    c=ws.cell(row=cur2,column=1,value=f"  SOUS-FAMILLE : {sfam.upper()}  ·  {int(sf['nb_arts'])} articles  ·  {sf['pct_marge']*100:.1f}%  ·  {fmt_fcfa(sf['Marge'])} FCFA")
                    c.font=Font(bold=True,color="FFFFFF",name="Calibri",size=10)
                    c.fill=f(SUB); c.alignment=LFT; ws.row_dimensions[cur2].height=22; cur2+=1

                    hrow(ws,cur2,["Code","Article","CA (FCFA)","Marge (FCFA)","% Marge","Marge HP","Marge Promo","Écart Promo","Action"],
                         [12,46,14,14,10,12,14,14,28],SUB)
                    cur2+=1

                    for ri2,(_,art) in enumerate(arts_sf2.iterrows(),cur2):
                        pm_a=art["pct_marge"] or 0
                        pm_hp=art["pct_marge_hp"] or 0
                        pm_pr=art["pct_marge_promo"] or 0
                        ecart_a=(pm_hp-pm_pr) if pm_hp>0 and pm_pr>0 else 0
                        vals2=[art["art_code"],art["art_label"],art["CA"],art["Marge"],
                               pm_a,pm_hp if pm_hp>0 else None,
                               pm_pr if pm_pr>0 else None,
                               ecart_a if ecart_a>0 else None,
                               art["action_rec"]]
                        fmts2=[None,None,"#,##0","#,##0","0.0%","0.0%","0.0%","0.0%",None]
                        for i,(v,fmt) in enumerate(vals2,1):
                            c=ws.cell(row=ri2,column=i,value=v)
                            c.fill=f(C_ODD if ri2%2==0 else C_EVN)
                            c.border=bdr()
                            if fmt: c.number_format=fmt
                            if i in [4,5] and isinstance(v,(int,float)) and v<0:
                                c.font=Font(bold=True,name="Calibri",size=10,color="FF3B30")
                            else: c.font=NF
                            c.alignment=LFT if i in [2,9] else RGT if isinstance(v,(int,float)) else CTR
                        ws.row_dimensions[ri2].height=18
                        cur2+=1
                    cur2+=1

            # ── Onglet classements ─────────────────────────────────────────────
            ws_cl=wb.create_sheet("Classements réseau")
            cur3=1
            for titre,df_cl,cols_cl,fmts_cl in [
                ("TOP 10 CA",arts.nlargest(10,"CA"),
                 ["art_code","art_label","rayon_label","sfam_label","CA","Marge","pct_marge","Qté Vente"],
                 [None,None,None,None,"#,##0","#,##0","0.0%","#,##0"]),
                ("TOP 20 MARGES NÉGATIVES",arts[arts["Marge"]<0].nsmallest(20,"Marge"),
                 ["art_code","art_label","rayon_label","sfam_label","CA","Marge","pct_marge","pct_marge_hp","pct_marge_promo"],
                 [None,None,None,None,"#,##0","#,##0","0.0%","0.0%","0.0%"]),
                ("TOP 10 CASSE",arts[arts["Casse (Valeur)"]<0].nsmallest(10,"Casse (Valeur)"),
                 ["art_code","art_label","rayon_label","sfam_label","Casse (Valeur)","Casse (Qté)"],
                 [None,None,None,None,"#,##0","#,##0"]),
            ]:
                ws_cl.merge_cells(start_row=cur3,start_column=1,end_row=cur3,end_column=len(cols_cl))
                c=ws_cl.cell(row=cur3,column=1,value=f"  {titre}")
                c.font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
                c.fill=f(HDR); c.alignment=LFT; ws_cl.row_dimensions[cur3].height=24; cur3+=1
                labels={"art_code":"Code","art_label":"Article","rayon_label":"Rayon","sfam_label":"Sous-famille",
                        "CA":"CA (FCFA)","Marge":"Marge (FCFA)","pct_marge":"% Marge",
                        "Qté Vente":"Qté vendue","pct_marge_hp":"% Marge HP","pct_marge_promo":"% Marge Promo",
                        "Casse (Valeur)":"Casse (FCFA)","Casse (Qté)":"Casse qté"}
                widths=[12,44,14,22,14,14,10,10,12,14,14,12]
                hrow(ws_cl,cur3,[labels.get(c,c) for c in cols_cl],widths[:len(cols_cl)]); cur3+=1
                if df_cl.empty:
                    ws_cl.cell(row=cur3,column=1,value="Aucune donnée"); cur3+=1
                else:
                    for ri3,(_,row3) in enumerate(df_cl.iterrows(),cur3):
                        for i,(col,fmt) in enumerate(zip(cols_cl,fmts_cl),1):
                            v=row3.get(col)
                            c=ws_cl.cell(row=ri3,column=i,value=v)
                            c.fill=f(C_ODD if ri3%2==0 else C_EVN);c.border=bdr()
                            if fmt: c.number_format=fmt
                            if col in ["Marge","Casse (Valeur)","pct_marge"] and isinstance(v,(int,float)) and v<0:
                                c.font=Font(bold=True,name="Calibri",size=10,color="FF3B30")
                            else: c.font=NF
                            c.alignment=LFT if i<=4 else RGT
                        ws_cl.row_dimensions[ri3].height=18
                    cur3+=len(df_cl)
                cur3+=2

            wb.active=wb["Brief Directeur"]
            buf=BytesIO(); wb.save(buf); buf.seek(0)

        fname=f"PerfHebdo_SmartBuyer_{uploaded.name.rsplit('.',1)[0]}.xlsx"
        st.download_button("⬇️ Télécharger",data=buf,file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(f"""
<div style='text-align:center;color:#C7C7CC;font-size:11px;padding:8px 0'>
  NovaRetail Solutions · SmartBuyer v2.3 · Perf Hebdo · {len(arts):,} articles analysés
</div>""", unsafe_allow_html=True)
