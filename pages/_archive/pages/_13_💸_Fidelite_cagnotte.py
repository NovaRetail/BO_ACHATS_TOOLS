"""
13_🏷️_Fidelite_Cagnotte.py — SmartBuyer Hub
Fidélité Cagnotte · Source : Export PBI standard (même format que Perf Hebdo / Ventes PBI)
Référentiel : Article / Cagnotte / Mois (CSV inchangé)

Migration v2 : suppression du parser CSV custom — le module accepte désormais
le même export PBI Excel que tous les autres modules SmartBuyer.
"""

import re
import numpy as np
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Fidélité Cagnotte · SmartBuyer",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1300px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }
.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.period-badge {
    background: #EAF4FF; border: 1px solid #B8D9FF; border-radius: 10px;
    padding: 12px 20px; margin-bottom: 16px; display: flex; align-items: center; gap: 16px;
}
.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
</style>
""", unsafe_allow_html=True)


# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fmt_xof(n):
    if pd.isna(n) or n is None: return "—"
    return f"{int(round(float(n))):,}".replace(",", " ") + " XOF"

def fmt_num(n):
    if pd.isna(n) or n is None: return "—"
    return f"{float(n):,.0f}".replace(",", " ")

def fmt_pct(v, dec=1):
    if pd.isna(v) or v is None: return "—"
    return f"{float(v):.{dec}f}%"

def short_label(s):
    """Extrait le libellé après 'CODE - '."""
    if pd.isna(s): return ""
    m = re.match(r"^\d[\d\s]*-\s*(.+)$", str(s).strip())
    return m.group(1).strip() if m else str(s).strip()

def extract_code(s):
    """Extrait le code numérique en début de chaine."""
    if pd.isna(s): return None
    m = re.match(r"^(\d+)", str(s).strip())
    try:
        return int(m.group(1)) if m else None
    except (ValueError, AttributeError):
        return None

def extract_periode(df_raw: pd.DataFrame) -> tuple[str, str]:
    """Extrait la plage de dates depuis la ligne de filtre PBI (dernière ligne)."""
    last_col = df_raw.iloc[:, 0].dropna().astype(str)
    if last_col.empty:
        return "—", "—"
    last = last_col.iloc[-1]
    m = re.search(
        r"après le (\d{2}/\d{2}/\d{4}).*?avant le (\d{2}/\d{2}/\d{4})",
        last
    )
    if m:
        return m.group(1), m.group(2)
    return "—", "—"

def date_to_mois(date_str: str) -> str:
    """Convertit 'DD/MM/YYYY' → nom du mois en français."""
    MOIS = {
        1:"Janvier", 2:"Février", 3:"Mars", 4:"Avril", 5:"Mai", 6:"Juin",
        7:"Juillet", 8:"Août", 9:"Septembre", 10:"Octobre", 11:"Novembre", 12:"Décembre"
    }
    try:
        d = pd.to_datetime(date_str, format="%d/%m/%Y")
        return MOIS[d.month]
    except Exception:
        return "—"


# ─── PARSING PBI ──────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_pbi(file_bytes: bytes) -> tuple[pd.DataFrame, str, str, str]:
    """
    Charge un export PBI standard (même format que Perf Hebdo / Ventes PBI).
    Retourne (df_articles, date_debut, date_fin, mois_label).

    Règles de filtrage :
    - Garder uniquement les lignes Article × Site réelles
      (Article non nul, non 'Total', Site non nul, non 'Total')
    - Exclure les lignes de totaux intermédiaires
    - Exclure la ligne de filtres PBI (dernière ligne)
    """
    raw = pd.read_excel(BytesIO(file_bytes))

    # Extraire la période avant tout nettoyage
    date_debut, date_fin = extract_periode(raw)
    mois_label = date_to_mois(date_debut)

    # Colonnes numériques à normaliser
    num_cols = [
        "CA", "Marge", "CA Hors Promo", "Marge Hors Promo",
        "CA Promo", "Marge Promo", "%CA Poids Promo",
        "Qté Vente", "Casse (Valeur)", "Casse (Qté)", "%Marge",
    ]
    for col in num_cols:
        if col in raw.columns:
            raw[col] = pd.to_numeric(raw[col], errors="coerce")

    # Filtrer : lignes article × site réelles
    mask = (
        raw["Article"].notna() &
        (~raw["Article"].astype(str).str.strip().isin(["Total", "nan", ""])) &
        raw["Site nom long"].notna() &
        (~raw["Site nom long"].astype(str).str.strip().isin(["Total", "nan", ""])) &
        # Exclure la ligne de filtres PBI
        (~raw["Article"].astype(str).str.startswith("Filtres", na=False))
    )
    df = raw[mask].copy().reset_index(drop=True)

    # Labels courts
    df["lib_article"]   = df["Article"].apply(short_label)
    df["lib_rayon"]     = df["Rayon"].apply(short_label)
    df["lib_famille"]   = df["Famille"].apply(short_label)
    df["lib_sfam"]      = df["Sous Famille"].apply(short_label) if "Sous Famille" in df.columns else ""
    df["lib_site"]      = df["Site nom long"].apply(short_label)
    df["code_article"]  = df["Article"].apply(extract_code)

    # Assurer l'existence des colonnes optionnelles
    for col in ["CA Promo", "Marge Promo", "CA Hors Promo", "Marge Hors Promo",
                "Casse (Valeur)", "Casse (Qté)", "Qté Vente", "%CA Poids Promo"]:
        if col not in df.columns:
            df[col] = 0.0

    df["Qté Vente"] = df["Qté Vente"].fillna(0)

    return df, date_debut, date_fin, mois_label


@st.cache_data(show_spinner=False)
def parse_fidelite(file_bytes: bytes) -> pd.DataFrame:
    """
    Charge le référentiel fidélité CSV.
    Format attendu : Article (int) | Cagnotte (float) | Mois (texte fr)
    """
    df = pd.read_csv(BytesIO(file_bytes), sep=None, engine="python")
    df.columns = [c.strip() for c in df.columns]

    required = {"Article", "Cagnotte"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Colonnes manquantes dans le référentiel : {', '.join(missing)}")

    df["Article"]  = pd.to_numeric(df["Article"], errors="coerce")
    df["Cagnotte"] = pd.to_numeric(df["Cagnotte"], errors="coerce")
    df = df.dropna(subset=["Article", "Cagnotte"])
    df["Article"] = df["Article"].astype(int)

    if "Mois" not in df.columns:
        df["Mois"] = "—"
    else:
        df["Mois"] = df["Mois"].astype(str).str.strip()

    return df[["Article", "Cagnotte", "Mois"]].copy()


def join_fidelite(
    df_pbi: pd.DataFrame,
    df_fid: pd.DataFrame,
    mois_pbi: str,
) -> pd.DataFrame:
    """
    Joint le PBI avec le référentiel fidélité.
    Priorité : jointure Article × Mois exact, puis fallback Article seul.
    Calcule Total Cagnotte = Cagnotte/unité × Qté Vente.
    """
    # Jointure stricte Article × Mois
    fid_mois = df_fid[df_fid["Mois"] == mois_pbi][["Article", "Cagnotte"]].copy()
    fid_mois = fid_mois.rename(columns={"Cagnotte": "Cagnotte_unit"})

    df = df_pbi.merge(fid_mois, left_on="code_article", right_on="Article", how="left")
    df = df.drop(columns=["Article_y"] if "Article_y" in df.columns else [], errors="ignore")

    # Fallback : si aucun mois ne correspond, prendre la dernière cagnotte connue
    n_matched = df["Cagnotte_unit"].notna().sum()
    if n_matched == 0 and not df_fid.empty:
        # Prendre la cagnotte sans filtre mois (utile si le mois PBI ne matche pas)
        fid_any = (
            df_fid.groupby("Article", as_index=False)["Cagnotte"]
            .last()
            .rename(columns={"Cagnotte": "Cagnotte_unit"})
        )
        df = df_pbi.merge(fid_any, left_on="code_article", right_on="Article", how="left")
        df = df.drop(columns=["Article_y"] if "Article_y" in df.columns else [], errors="ignore")

    df["est_fidelite"]   = df["Cagnotte_unit"].notna()
    df["Total Cagnotte"] = df["Cagnotte_unit"].fillna(0) * df["Qté Vente"].fillna(0)

    return df


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def to_excel(sheets: dict) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        H_FILL = PatternFill("solid", fgColor="1C3557")
        H_FONT = Font(bold=True, color="FFFFFF", size=10)
        CTR    = Alignment(horizontal="center", vertical="center")

        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            for i, col in enumerate(df.columns, 1):
                c = ws.cell(row=1, column=i)
                c.fill = H_FILL; c.font = H_FONT; c.alignment = CTR
                ws.column_dimensions[get_column_letter(i)].width = min(
                    max(len(str(col)) + 4, 12), 40
                )
            ws.freeze_panes = "A2"
    return buf.getvalue()


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichiers</div>", unsafe_allow_html=True)

    st.markdown("**Export PBI ventes** *(même format que Perf Hebdo)*")
    pbi_files = st.file_uploader(
        "PBI", type=["xlsx", "xls"], accept_multiple_files=True,
        key="fid_pbi", label_visibility="collapsed",
        help="Export PBI standard : Rayon / Famille / Article / Site nom long / CA / Marge / CA Promo / Qté Vente"
    )

    st.markdown("**Référentiel Fidélité** *(CSV inchangé)*")
    fid_file = st.file_uploader(
        "Fidelite", type=["csv"], key="fid_ref", label_visibility="collapsed",
        help="Colonnes : Article | Cagnotte | Mois"
    )

    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px'>Filtres</div>", unsafe_allow_html=True)
    # Filtres — peuplés après chargement (déclarés ici pour structure sidebar)
    _sidebar_filters = True


# ─── PAGE PRINCIPALE ──────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>🏷️ Fidélité Cagnotte</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Investissement cagnotte · Performance par article × magasin · Source : Export PBI standard</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL ──────────────────────────────────────────────────────────
if not pbi_files or not fid_file:
    st.markdown("---")
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ Ce module utilise désormais l'export PBI standard</strong><br>
  Le même fichier Excel que tu charges dans <strong>Perf Hebdo</strong> ou <strong>Ventes PBI</strong>.
  Plus de CSV custom — un seul format pour tout SmartBuyer Hub.
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Fichiers attendus</div>", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
<div class='col-required'>
  <div style='font-size:16px'>📊</div>
  <div>
    <div class='col-name'>Export PBI ventes (.xlsx)</div>
    <div class='col-desc'>Même fichier que Perf Hebdo / Ventes PBI. Multi-upload possible (plusieurs semaines).</div>
    <div class='col-desc' style='margin-top:4px;color:#8E8E93'>Colonnes : Rayon · Famille · Sous Famille · Article · Site nom long · CA · Marge · CA Promo · Marge Promo · Qté Vente</div>
  </div>
</div>""", unsafe_allow_html=True)

    with c2:
        st.markdown("""
<div class='col-required'>
  <div style='font-size:16px'>📋</div>
  <div>
    <div class='col-name'>Référentiel Fidélité (.csv)</div>
    <div class='col-desc'>Format inchangé — 3 colonnes.</div>
    <div class='col-desc' style='margin-top:4px;color:#8E8E93'>Article (code entier) · Cagnotte (XOF/unité) · Mois (ex: Mai, Avril)</div>
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
<div class='alert-card alert-green'>
  <strong>✅ Nouveautés v2</strong><br>
  · Plus de parser CSV custom : le module lit directement l'export PBI Excel<br>
  · La période est détectée automatiquement depuis les métadonnées PBI<br>
  · Multi-upload : charge plusieurs semaines pour une analyse consolidée<br>
  · Jointure Article × Mois automatique avec fallback si le mois ne correspond pas
</div>""", unsafe_allow_html=True)

    st.info("⬆️ Charge les fichiers dans la sidebar pour démarrer.")
    st.stop()


# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
errors = []
all_dfs, all_periods = [], []

with st.spinner("Lecture des fichiers PBI…"):
    for f in pbi_files:
        try:
            df_v, d1, d2, mois = parse_pbi(f.read())
            if not df_v.empty:
                df_v["_fichier"] = f.name
                df_v["_periode"] = f"{d1} → {d2}"
                df_v["_mois"]    = mois
                all_dfs.append(df_v)
                all_periods.append({"fichier": f.name, "d1": d1, "d2": d2, "mois": mois})
        except Exception as e:
            errors.append(f"**{f.name}** : {e}")

if errors:
    for err in errors:
        st.warning(err)

if not all_dfs:
    st.error("Aucune donnée valide dans les fichiers PBI chargés.")
    st.stop()

df_pbi_all = pd.concat(all_dfs, ignore_index=True)

try:
    df_fid = parse_fidelite(fid_file.read())
except Exception as e:
    st.error(f"Erreur lecture référentiel fidélité : {e}")
    st.stop()


# ─── FILTRES SIDEBAR ──────────────────────────────────────────────────────────
with st.sidebar:
    mois_dispo = sorted(set(p["mois"] for p in all_periods if p["mois"] != "—"))
    mois_sel   = st.multiselect("Mois", mois_dispo, default=mois_dispo, key="fid_mois_sel")
    if not mois_sel:
        mois_sel = mois_dispo

    rayons_dispo = sorted(df_pbi_all["lib_rayon"].dropna().unique())
    rayon_sel    = st.multiselect("Rayon", rayons_dispo, default=[], key="fid_rayon_sel",
                                   help="Vide = tous")

    fam_src   = df_pbi_all[df_pbi_all["lib_rayon"].isin(rayon_sel)] if rayon_sel else df_pbi_all
    fam_dispo = sorted(fam_src["lib_famille"].dropna().unique())
    fam_sel   = st.multiselect("Famille", fam_dispo, default=[], key="fid_fam_sel",
                                help="Vide = toutes")

    sites_dispo = sorted(df_pbi_all["lib_site"].dropna().unique())
    sites_sel   = st.multiselect("Magasins", sites_dispo, default=[], key="fid_sites_sel",
                                  help="Vide = tous")

    if st.button("↺ Réinitialiser", use_container_width=True):
        for k in ["fid_mois_sel", "fid_rayon_sel", "fid_fam_sel", "fid_sites_sel"]:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()


# ─── FILTRE + JOINTURE ────────────────────────────────────────────────────────
df_filtre = df_pbi_all[df_pbi_all["_mois"].isin(mois_sel)].copy()
if rayon_sel:  df_filtre = df_filtre[df_filtre["lib_rayon"].isin(rayon_sel)]
if fam_sel:    df_filtre = df_filtre[df_filtre["lib_famille"].isin(fam_sel)]
if sites_sel:  df_filtre = df_filtre[df_filtre["lib_site"].isin(sites_sel)]

if df_filtre.empty:
    st.warning("Aucune donnée pour les filtres sélectionnés.")
    st.stop()

# Mois principal pour la jointure (premier mois sélectionné)
mois_join = mois_sel[0] if mois_sel else "—"
df = join_fidelite(df_filtre, df_fid, mois_join)
df_fid_only = df[df["est_fidelite"]].copy()

if df_fid_only.empty:
    st.markdown("""
<div class='alert-card alert-amber'>
  <strong>⚠️ Aucun article fidélité trouvé dans ce fichier PBI</strong><br>
  Vérifie que les codes articles du référentiel correspondent au format PBI (ex: 12001277).
</div>""", unsafe_allow_html=True)
    with st.expander("Diagnostic"):
        st.write("Codes dans le référentiel :", sorted(df_fid["Article"].unique().tolist())[:10])
        st.write("Codes dans le PBI :", sorted([c for c in df["code_article"].dropna().unique().tolist() if c])[:10])
    st.stop()


# ─── PERIOD BADGE ─────────────────────────────────────────────────────────────
periodes_label = " · ".join(sorted(set(p["d1"] + " → " + p["d2"] for p in all_periods)))
mois_label_str = " + ".join(mois_sel)
n_fichiers     = len(pbi_files)
filtres_str    = " · ".join(
    [f"{len(rayon_sel)} rayon(s)"] * bool(rayon_sel) +
    [f"{len(fam_sel)} famille(s)"] * bool(fam_sel) +
    [f"{len(sites_sel)} magasin(s)"] * bool(sites_sel)
) or "Tous rayons · Tous magasins"

st.markdown(f"""
<div class="period-badge">
  <div>
    <div style='font-size:11px;font-weight:600;color:#007AFF;text-transform:uppercase;letter-spacing:.05em'>
      Période · {mois_label_str}
    </div>
    <div style='font-size:15px;font-weight:700;color:#007AFF;margin-top:2px'>{periodes_label}</div>
    <div style='font-size:12px;color:#3A3A3C;margin-top:3px'>{filtres_str}</div>
  </div>
  <div style='margin-left:auto;text-align:right;font-size:12px;color:#3A3A3C'>
    <div>{n_fichiers} fichier(s) PBI · {len(df_fid)} articles en programme</div>
    <div>{len(df_filtre):,} lignes PBI filtrées</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─── KPIs ─────────────────────────────────────────────────────────────────────
budget      = df_fid_only["Total Cagnotte"].sum()
ca_fid      = df_fid_only["CA"].sum()
mg_fid      = df_fid_only["Marge"].sum()
ca_global   = df_filtre["CA"].sum()
poids_ca    = ca_fid / ca_global * 100 if ca_global > 0 else 0
arts_actifs = df_fid_only[df_fid_only["CA"] > 0]["code_article"].nunique()
arts_perim  = df_fid["Article"].nunique()
sites_actif = df_fid_only[df_fid_only["CA"] > 0]["lib_site"].nunique()

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Budget Cagnotte",  fmt_xof(budget))
k2.metric("CA Fidélité",      fmt_xof(ca_fid), f"{poids_ca:.1f}% du CA global")
k3.metric("Marge Fidélité",   fmt_xof(mg_fid))
k4.metric("Articles actifs",  f"{arts_actifs} / {arts_perim}")
k5.metric("Sites couverts",   f"{sites_actif} / {df_filtre['lib_site'].nunique()}")


# ─── ALERTES ──────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Alertes</div>", unsafe_allow_html=True)

# Articles sans ventes
arts_perimetre_ids = set(df_fid["Article"].unique())
arts_avec_vente    = set(
    df_fid_only[df_fid_only["CA"] > 0]["code_article"].dropna().unique()
)
arts_zero = arts_perimetre_ids - arts_avec_vente
n_zero    = len(arts_zero)

# Familles marge négative
fam_marge = df_fid_only.groupby("lib_famille")["Marge"].sum()
fam_neg   = fam_marge[fam_marge < 0]
n_neg     = len(fam_neg)

a1, a2 = st.columns(2)
with a1:
    if n_zero > 0:
        st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>⚠️ {n_zero} article(s) en programme sans aucune vente</strong><br>
  <span style='font-size:12px;opacity:.85'>Budget cagnotte investi à ROI nul — vérifier disponibilité en rayon.</span>
</div>""", unsafe_allow_html=True)
    else:
        st.markdown("<div class='alert-card alert-green'>✅ Tous les articles en programme ont généré des ventes.</div>", unsafe_allow_html=True)

with a2:
    if n_neg > 0:
        noms = ", ".join(fam_neg.index.tolist())
        st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>⚠️ {n_neg} famille(s) en marge négative</strong> : {noms}<br>
  <span style='font-size:12px;opacity:.85'>La cagnotte dégrade la marge sur ces familles.</span>
</div>""", unsafe_allow_html=True)
    else:
        st.markdown("<div class='alert-card alert-green'>✅ Aucune famille en marge négative.</div>", unsafe_allow_html=True)


# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab0, tab1, tab2, tab3, tab4 = st.tabs([
    "🏠 Synthèse",
    "📊 Récap Financier",
    "🔍 Détail Article × Site",
    "📋 Drill-down",
    "📥 Export Excel",
])


# ═══ TAB 0 — SYNTHÈSE ═════════════════════════════════════════════════════════
with tab0:
    c1, c2 = st.columns(2)

    # Top 5 familles par budget cagnotte
    with c1:
        st.markdown("<div class='section-label'>Top familles · Budget cagnotte</div>", unsafe_allow_html=True)
        grp_fam = df_fid_only.groupby("lib_famille", as_index=False).agg(
            CA=("CA","sum"), Marge=("Marge","sum"), Budget=("Total Cagnotte","sum")
        )
        grp_fam_gl = df_filtre.groupby("lib_famille", as_index=False).agg(CA_gl=("CA","sum"))
        grp_fam    = grp_fam.merge(grp_fam_gl, on="lib_famille", how="left")
        grp_fam["Poids %"] = (grp_fam["CA"] / grp_fam["CA_gl"].replace(0, np.nan) * 100).round(1)
        grp_fam = grp_fam.sort_values("Budget", ascending=False).head(5)

        disp_fam = grp_fam.rename(columns={"lib_famille":"Famille","CA":"CA Fidélité","Budget":"Budget Cagnotte"})
        disp_fam["CA Fidélité"]    = disp_fam["CA Fidélité"].apply(fmt_num)
        disp_fam["Marge"]          = disp_fam["Marge"].apply(fmt_num)
        disp_fam["Budget Cagnotte"]= disp_fam["Budget Cagnotte"].apply(fmt_num)
        disp_fam["Poids %"]        = disp_fam["Poids %"].apply(fmt_pct)
        st.dataframe(disp_fam[["Famille","CA Fidélité","Marge","Budget Cagnotte","Poids %"]],
                     use_container_width=True, hide_index=True, height=240)

    # Performance réseau par site
    with c2:
        st.markdown("<div class='section-label'>Performance réseau · Site</div>", unsafe_allow_html=True)
        grp_site = df_fid_only.groupby("lib_site", as_index=False).agg(
            CA=("CA","sum"), Marge=("Marge","sum"),
            Budget=("Total Cagnotte","sum"), Nb_Art=("code_article","nunique")
        )
        grp_site_gl = df_filtre.groupby("lib_site", as_index=False).agg(CA_gl=("CA","sum"))
        grp_site    = grp_site.merge(grp_site_gl, on="lib_site", how="left")
        grp_site["Poids %"] = (grp_site["CA"] / grp_site["CA_gl"].replace(0, np.nan) * 100).round(1)
        grp_site = grp_site.sort_values("CA", ascending=False)

        disp_site = grp_site.rename(columns={"lib_site":"Site","CA":"CA Fidélité","Budget":"Budget","Nb_Art":"Nb Art."})
        disp_site["CA Fidélité"] = disp_site["CA Fidélité"].apply(fmt_num)
        disp_site["Marge"]       = disp_site["Marge"].apply(fmt_num)
        disp_site["Budget"]      = disp_site["Budget"].apply(fmt_num)
        disp_site["Poids %"]     = disp_site["Poids %"].apply(fmt_pct)
        st.dataframe(disp_site[["Site","CA Fidélité","Marge","Budget","Poids %","Nb Art."]],
                     use_container_width=True, hide_index=True, height=240)

    # Articles sans ventes (si applicable)
    if n_zero > 0:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Articles en programme sans ventes</div>", unsafe_allow_html=True)
        arts_zero_df = df_fid[df_fid["Article"].isin(arts_zero)][["Article","Cagnotte","Mois"]].copy()
        arts_zero_df["Cagnotte"] = arts_zero_df["Cagnotte"].apply(fmt_num)
        st.dataframe(arts_zero_df, use_container_width=True, hide_index=True, height=200)


# ═══ TAB 1 — RÉCAP FINANCIER ══════════════════════════════════════════════════
with tab1:
    st.markdown("<div class='section-label'>Site × Rayon × Famille · CA et Marge fidélité vs global</div>", unsafe_allow_html=True)

    grp_gl  = df_filtre.groupby(["lib_site","lib_rayon","lib_famille"], as_index=False).agg(
        CA_gl=("CA","sum"), Mg_gl=("Marge","sum"))
    grp_fid = df_fid_only.groupby(["lib_site","lib_rayon","lib_famille"], as_index=False).agg(
        CA=("CA","sum"), Marge=("Marge","sum"),
        Qte=("Qté Vente","sum"), Budget=("Total Cagnotte","sum"))

    recap = grp_fid.merge(grp_gl, on=["lib_site","lib_rayon","lib_famille"], how="left")
    recap["Poids CA %"]    = (recap["CA"]    / recap["CA_gl"].replace(0, np.nan) * 100).round(1)
    recap["Poids Marge %"] = (recap["Marge"] / recap["Mg_gl"].abs().replace(0, np.nan) * 100).round(1)

    disp_recap = recap.rename(columns={
        "lib_site":"Site","lib_rayon":"Rayon","lib_famille":"Famille",
        "CA":"CA Fidélité","Marge":"Marge Fidélité","CA_gl":"CA Global",
        "Mg_gl":"Marge Globale","Qte":"Qté Vente","Budget":"Budget Cagnotte"
    })
    for col in ["CA Fidélité","Marge Fidélité","CA Global","Marge Globale","Budget Cagnotte"]:
        disp_recap[col] = disp_recap[col].apply(fmt_num)
    disp_recap["Qté Vente"]     = disp_recap["Qté Vente"].apply(lambda x: f"{x:,.0f}".replace(",", " ") if pd.notna(x) else "—")
    disp_recap["Poids CA %"]    = disp_recap["Poids CA %"].apply(fmt_pct)
    disp_recap["Poids Marge %"] = disp_recap["Poids Marge %"].apply(fmt_pct)

    st.dataframe(disp_recap[[
        "Site","Rayon","Famille","CA Fidélité","CA Global","Poids CA %",
        "Marge Fidélité","Marge Globale","Poids Marge %","Qté Vente","Budget Cagnotte"
    ]], use_container_width=True, height=500, hide_index=True)


# ═══ TAB 2 — DÉTAIL ARTICLE × SITE ═══════════════════════════════════════════
with tab2:
    st.markdown("<div class='section-label'>Article × Magasin · Cagnotte unitaire · Total cagnotte</div>", unsafe_allow_html=True)

    grp_art = df_fid_only.groupby(
        ["lib_rayon","lib_famille","lib_article","lib_site"], as_index=False
    ).agg(
        CA=("CA","sum"), Marge=("Marge","sum"), Qte=("Qté Vente","sum"),
        Cagnotte_unit=("Cagnotte_unit","first"), Budget=("Total Cagnotte","sum")
    )
    grp_art = grp_art.sort_values("Budget", ascending=False).reset_index(drop=True)

    disp_art = grp_art.rename(columns={
        "lib_rayon":"Rayon","lib_famille":"Famille","lib_article":"Article",
        "lib_site":"Site","Qte":"Qté Vente","Cagnotte_unit":"Cagnotte/unité","Budget":"Total Cagnotte"
    })
    for col in ["CA","Marge","Total Cagnotte"]:
        disp_art[col] = disp_art[col].apply(fmt_num)
    disp_art["Qté Vente"]      = disp_art["Qté Vente"].apply(lambda x: f"{x:,.0f}".replace(",", " "))
    disp_art["Cagnotte/unité"] = disp_art["Cagnotte/unité"].apply(fmt_num)

    st.dataframe(disp_art, use_container_width=True, height=520, hide_index=True)


# ═══ TAB 3 — DRILL-DOWN ═══════════════════════════════════════════════════════
with tab3:
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        rayon_dd = st.selectbox("Rayon", ["Tous"] + sorted(df_fid_only["lib_rayon"].dropna().unique()), key="dd_r")
    with fc2:
        fam_src_dd = df_fid_only[df_fid_only["lib_rayon"]==rayon_dd] if rayon_dd!="Tous" else df_fid_only
        fam_dd     = st.selectbox("Famille", ["Toutes"] + sorted(fam_src_dd["lib_famille"].dropna().unique()), key="dd_f")
    with fc3:
        site_dd = st.selectbox("Magasin", ["Tous"] + sorted(df_fid_only["lib_site"].dropna().unique()), key="dd_s")

    drill = df_fid_only.copy()
    if rayon_dd != "Tous":   drill = drill[drill["lib_rayon"]    == rayon_dd]
    if fam_dd   != "Toutes": drill = drill[drill["lib_famille"]  == fam_dd]
    if site_dd  != "Tous":   drill = drill[drill["lib_site"]     == site_dd]

    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>"
                f"{len(drill):,} lignes · CA : <strong>{fmt_num(drill['CA'].sum())} XOF</strong> · "
                f"Budget : <strong>{fmt_num(drill['Total Cagnotte'].sum())} XOF</strong></div>",
                unsafe_allow_html=True)

    disp_drill = drill[[
        "_periode","lib_rayon","lib_famille","lib_article","lib_site",
        "CA","Marge","Qté Vente","Cagnotte_unit","Total Cagnotte"
    ]].rename(columns={
        "_periode":"Période","lib_rayon":"Rayon","lib_famille":"Famille",
        "lib_article":"Article","lib_site":"Site",
        "Cagnotte_unit":"Cagnotte/unité"
    }).copy()

    for col in ["CA","Marge","Total Cagnotte"]:
        disp_drill[col] = disp_drill[col].apply(fmt_num)
    disp_drill["Qté Vente"]      = disp_drill["Qté Vente"].apply(lambda x: f"{x:,.0f}".replace(",", " "))
    disp_drill["Cagnotte/unité"] = disp_drill["Cagnotte/unité"].apply(fmt_num)

    st.dataframe(disp_drill, use_container_width=True, height=520, hide_index=True)


# ═══ TAB 4 — EXPORT EXCEL ═════════════════════════════════════════════════════
with tab4:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>📋 Contenu de l'export (5 onglets)</strong><br>
  <strong>1. Récap Financier</strong> — Site × Rayon × Famille · CA / Marge fidélité vs global · Poids % · Budget<br>
  <strong>2. Détail Article × Site</strong> — Cagnotte unitaire · Total cagnotte · CA · Marge par ligne<br>
  <strong>3. Synthèse Familles</strong> — Top familles par budget cagnotte<br>
  <strong>4. Performance Réseau</strong> — Par magasin avec poids CA %<br>
  <strong>5. Alertes</strong> — Articles sans ventes + familles marge négative
</div>""", unsafe_allow_html=True)

    st.caption(f"Période : {periodes_label} · {arts_actifs} articles actifs · {len(df_fid_only):,} lignes")

    if st.button("Générer le fichier Excel", type="primary", key="gen_fid_excel"):
        with st.spinner("Génération…"):

            # Récap financier (valeurs brutes pour Excel)
            recap_xl = recap.rename(columns={
                "lib_site":"Site","lib_rayon":"Rayon","lib_famille":"Famille",
                "CA":"CA Fidélité","Marge":"Marge Fidélité","CA_gl":"CA Global",
                "Mg_gl":"Marge Globale","Qte":"Qté Vente","Budget":"Budget Cagnotte"
            })

            # Détail article × site
            detail_xl = df_fid_only[[
                "_periode","lib_rayon","lib_famille","lib_sfam","lib_article","lib_site",
                "CA","Marge","Qté Vente","Cagnotte_unit","Total Cagnotte"
            ]].rename(columns={
                "_periode":"Période","lib_rayon":"Rayon","lib_famille":"Famille",
                "lib_sfam":"Sous-Famille","lib_article":"Article","lib_site":"Site",
                "Cagnotte_unit":"Cagnotte/unité"
            })

            # Synthèse familles
            synth_fam = grp_fam.rename(columns={
                "lib_famille":"Famille","Budget":"Budget Cagnotte"})

            # Performance réseau
            perf_site = grp_site.rename(columns={
                "lib_site":"Site","Budget":"Budget Cagnotte","Nb_Art":"Nb Articles"})

            # Alertes
            alertes_rows = []
            for art_id in arts_zero:
                fid_rows = df_fid[df_fid["Article"] == art_id]
                cagnotte = fid_rows["Cagnotte"].iloc[0] if not fid_rows.empty else 0
                alertes_rows.append({"Article": art_id, "Cagnotte/unité": cagnotte, "Alerte": "Aucune vente"})
            for fam, mg in fam_neg.items():
                alertes_rows.append({"Famille": fam, "Marge Fidélité": mg, "Alerte": "Marge négative"})
            alertes_xl = pd.DataFrame(alertes_rows) if alertes_rows else pd.DataFrame(
                {"Message": ["Aucune alerte"]})

            excel_bytes = to_excel({
                "Récap Financier":      recap_xl,
                "Détail Article×Site":  detail_xl,
                "Synthèse Familles":    synth_fam,
                "Performance Réseau":   perf_site,
                "Alertes":              alertes_xl,
            })

        st.download_button(
            label=f"⬇️ Télécharger — Fidélité Cagnotte",
            data=excel_bytes,
            file_name=f"SmartBuyer_Fidelite_{mois_label_str.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
