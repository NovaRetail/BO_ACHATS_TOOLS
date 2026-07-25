"""
Revue de performance ADIALEA RCI — script complet
Analyses 1 à 6, sortie Excel multi-onglets, charte visuelle SmartBuyer Hub.

Usage :
    python revue_performance.py <export_article.xlsx> <sortie.xlsx>
"""

import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────── Charte SmartBuyer ────────────────────────────
BLEU = "007AFF"
ROUGE = "FF3B30"
VERT = "34C759"
GRIS_FOND = "F2F2F7"
GRIS_TEXTE = "8E8E93"
BLANC = "FFFFFF"
NOIR = "1C1C1E"
FONT_NAME = "Calibri"  # substitut portable le plus proche de SF Pro/Inter

F_TITRE = Font(name=FONT_NAME, size=14, bold=True, color=BLANC)
F_HEADER = Font(name=FONT_NAME, size=10, bold=True, color=BLANC)
F_SOUSTOTAL = Font(name=FONT_NAME, size=10, bold=True, color=NOIR)
F_NORMAL = Font(name=FONT_NAME, size=10, color=NOIR)
F_NORMAL_GRIS = Font(name=FONT_NAME, size=10, color=GRIS_TEXTE)

FILL_HEADER = PatternFill("solid", fgColor=BLEU)
FILL_TITRE = PatternFill("solid", fgColor=NOIR)
FILL_SOUSTOTAL = PatternFill("solid", fgColor=GRIS_FOND)
FILL_BLANC = PatternFill("solid", fgColor=BLANC)

BORDER_FIN = Border(bottom=Side(style="thin", color="D1D1D6"))

FMT_FCFA = '#,##0" FCFA";-#,##0" FCFA"'
FMT_PCT = "0.0%"
FMT_PTS = '+0.0" pts";-0.0" pts"'


def couleur_evol(valeur):
    if pd.isna(valeur):
        return Font(name=FONT_NAME, size=10, color=GRIS_TEXTE)
    return Font(name=FONT_NAME, size=10, color=(VERT if valeur >= 0 else ROUGE))


# ─────────────────────────── Zoning / Format ───────────────────────────────
ZONE_CD = {
    "10203 - Hyper Yopougon",
    "10604 - Market Cité verte",
    "10601 - Supeco Niangon",
    "10602 - Supeco Terminus 47",
    "10603 - Supeco Toit rouge",
    "10605 - Supeco Aboboté",
}

FORMAT_MAP = {
    "10301 - Hyper Marcory": "Hyper",
    "10202 - Hyper Palmeraie": "Hyper",
    "10203 - Hyper Yopougon": "Hyper",
    "10705 - Market 7 Décembre": "Market",
    "10208 - Market Riviera": "Market",
    "10209 - Market 2 Plateaux": "Market",
    "10604 - Market Cité verte": "Market",
    "10206 - Market Kokoh Mall": "Market",
    "10601 - Supeco Niangon": "Supeco",
    "10602 - Supeco Terminus 47": "Supeco",
    "10603 - Supeco Toit rouge": "Supeco",
    "10605 - Supeco Aboboté": "Supeco",
}
ORDRE_FORMAT = ["Hyper", "Market", "Supeco"]


def zone(site):
    return "CD" if site in ZONE_CD else "AB"


def format_magasin(site):
    return FORMAT_MAP.get(site, "Non mappé")


# ─────────────────────────── Chargement & nettoyage ────────────────────────
def charger_donnees(chemin):
    df = pd.read_excel(chemin, sheet_name="Export")

    mask = df["CA"].notna()
    for col in ["Site nom long", "Rayon", "Famille", "Article"]:
        mask &= df[col].notna() & (df[col].astype(str) != "Total")
    df = df[mask].copy()

    for col in ["Site nom long", "Rayon", "Famille", "Article"]:
        df[col] = df[col].astype(str)

    df["Zone"] = df["Site nom long"].apply(zone)
    df["Format"] = df["Site nom long"].apply(format_magasin)
    df["Marge N-1"] = df["Marge Hors Promo N-1"].fillna(0) + df["Marge Promo N-1"].fillna(0)
    return df


SOMMES = [
    "CA", "CA N-1", "CA Hors Promo", "CA Hors Promo N-1", "CA Promo", "CA Promo N-1",
    "Marge", "Marge N-1", "Marge Hors Promo", "Marge Hors Promo N-1",
    "Marge Promo", "Marge Promo N-1", "Qté Vente", "Qté Vente N-1", "Casse (Valeur)",
]


def agreger(df, group_cols):
    g = df.groupby(group_cols, as_index=False)[SOMMES].sum()

    g["Evol CA valeur"] = g["CA"] - g["CA N-1"]
    g["Evol CA %"] = np.where(g["CA N-1"] != 0, g["Evol CA valeur"] / g["CA N-1"], np.nan)

    g["Evol CA HP valeur"] = g["CA Hors Promo"] - g["CA Hors Promo N-1"]
    g["Evol CA HP %"] = np.where(g["CA Hors Promo N-1"] != 0, g["Evol CA HP valeur"] / g["CA Hors Promo N-1"], np.nan)

    g["Evol CA Promo valeur"] = g["CA Promo"] - g["CA Promo N-1"]
    g["Evol CA Promo %"] = np.where(g["CA Promo N-1"] != 0, g["Evol CA Promo valeur"] / g["CA Promo N-1"], np.nan)

    g["%Poids Promo"] = np.where(g["CA"] != 0, g["CA Promo"] / g["CA"], np.nan)
    g["%Poids Promo N-1"] = np.where(g["CA N-1"] != 0, g["CA Promo N-1"] / g["CA N-1"], np.nan)
    g["Evol %Poids Promo pts"] = g["%Poids Promo"] - g["%Poids Promo N-1"]

    g["Evol Marge valeur"] = g["Marge"] - g["Marge N-1"]
    g["Evol Marge %"] = np.where(g["Marge N-1"] != 0, g["Evol Marge valeur"] / g["Marge N-1"], np.nan)
    g["%Marge"] = np.where(g["CA"] != 0, g["Marge"] / g["CA"], np.nan)
    g["%Marge N-1"] = np.where(g["CA N-1"] != 0, g["Marge N-1"] / g["CA N-1"], np.nan)
    g["Evol %Marge pts"] = g["%Marge"] - g["%Marge N-1"]

    g["Evol Marge HP valeur"] = g["Marge Hors Promo"] - g["Marge Hors Promo N-1"]
    g["%Marge HP"] = np.where(g["CA Hors Promo"] != 0, g["Marge Hors Promo"] / g["CA Hors Promo"], np.nan)
    g["%Marge HP N-1"] = np.where(g["CA Hors Promo N-1"] != 0, g["Marge Hors Promo N-1"] / g["CA Hors Promo N-1"], np.nan)
    g["Evol %Marge HP pts"] = g["%Marge HP"] - g["%Marge HP N-1"]

    g["Evol Marge Promo valeur"] = g["Marge Promo"] - g["Marge Promo N-1"]
    g["%Marge Promo"] = np.where(g["CA Promo"] != 0, g["Marge Promo"] / g["CA Promo"], np.nan)
    g["%Marge Promo N-1"] = np.where(g["CA Promo N-1"] != 0, g["Marge Promo N-1"] / g["CA Promo N-1"], np.nan)
    g["Evol %Marge Promo pts"] = g["%Marge Promo"] - g["%Marge Promo N-1"]

    g["Evol Qté valeur"] = g["Qté Vente"] - g["Qté Vente N-1"]
    g["Evol Qté %"] = np.where(g["Qté Vente N-1"] != 0, g["Evol Qté valeur"] / g["Qté Vente N-1"], np.nan)

    g["%Casse"] = np.where(g["CA"] != 0, g["Casse (Valeur)"] / g["CA"], np.nan)
    return g


# ─────────────────────────── Analyse 1 : Format > Rayon ────────────────────
def analyse1(df):
    a = agreger(df, ["Format", "Rayon"])
    a["ordre"] = a["Format"].map({f: i for i, f in enumerate(ORDRE_FORMAT)})
    return a.sort_values(["ordre", "CA"], ascending=[True, False]).drop(columns="ordre")


# ───────────────────── Analyse 2 : Format > Rayon > Famille ────────────────
def analyse2(df):
    a = agreger(df, ["Format", "Rayon", "Famille"])
    a["ordre"] = a["Format"].map({f: i for i, f in enumerate(ORDRE_FORMAT)})
    return a.sort_values(["ordre", "Rayon", "CA"], ascending=[True, True, False]).drop(columns="ordre")


# ───────────────── Analyse 3 : Magasin > Rayon > Famille (tableau unique) ──
def analyse3(df):
    a = agreger(df, ["Site nom long", "Format", "Zone", "Rayon", "Famille"])
    a["ordre"] = a["Format"].map({f: i for i, f in enumerate(ORDRE_FORMAT)})
    return a.sort_values(
        ["ordre", "Site nom long", "Rayon", "CA"], ascending=[True, True, True, False]
    ).drop(columns="ordre")


# ───────────────── Analyse 4 : Top & Flop 50 Articles (CA / Marge) ────────
def top_flop(df_perimetre, valeur_col, n=50):
    art = df_perimetre.groupby(
        ["Rayon", "Famille", "Article"], as_index=False
    )[["CA", "Marge"]].sum()

    total_famille = art.groupby("Famille")[valeur_col].transform("sum")
    total_rayon = art.groupby("Rayon")[valeur_col].transform("sum")
    art["% dans Famille"] = np.where(total_famille != 0, art[valeur_col] / total_famille, np.nan)
    art["% dans Rayon"] = np.where(total_rayon != 0, art[valeur_col] / total_rayon, np.nan)

    top = art.sort_values(valeur_col, ascending=False).head(n).reset_index(drop=True)
    flop = art.sort_values(valeur_col, ascending=True).head(n).reset_index(drop=True)
    return top, flop


def analyse4(df):
    resultats = {}
    for valeur_col in ["CA", "Marge"]:
        top, flop = top_flop(df, valeur_col)
        resultats[f"Top50_{valeur_col}_Reseau"] = top
        resultats[f"Flop50_{valeur_col}_Reseau"] = flop
        for fmt in ORDRE_FORMAT:
            d_fmt = df[df["Format"] == fmt]
            top_f, flop_f = top_flop(d_fmt, valeur_col)
            resultats[f"Top50_{valeur_col}_{fmt}"] = top_f
            resultats[f"Flop50_{valeur_col}_{fmt}"] = flop_f
    return resultats


# ───────────────── Analyse 5 : Zone CD vs Zone AB ──────────────────────────
def analyse5_comparatif(df):
    agg = agreger(df, ["Zone", "Rayon", "Famille"])
    piv = agg.pivot_table(
        index=["Rayon", "Famille"], columns="Zone",
        values=["CA", "CA N-1", "Evol CA %", "Marge", "%Marge", "Evol %Marge pts"],
    )
    piv.columns = [f"{a} {b}" for a, b in piv.columns]
    piv = piv.reset_index()
    if "%Marge CD" in piv.columns and "%Marge AB" in piv.columns:
        piv["Écart %Marge CD-AB pts"] = piv["%Marge CD"] - piv["%Marge AB"]
    tri_col = "CA CD" if "CA CD" in piv.columns else piv.columns[2]
    return piv.sort_values(tri_col, ascending=False, na_position="last")


def analyse5_positionnement(df):
    agg = agreger(df, ["Zone", "Rayon"])
    total_qte = df["Qté Vente"].sum()
    total_ca = df["CA"].sum()
    agg["%Poids Qté réseau"] = agg["Qté Vente"] / total_qte
    agg["%Poids CA réseau"] = agg["CA"] / total_ca
    agg["Indice pression prix"] = np.where(
        agg["%Poids CA réseau"] != 0, agg["%Poids Qté réseau"] / agg["%Poids CA réseau"], np.nan
    )
    cols = ["Zone", "Rayon", "%Marge", "%Poids Promo", "%Poids Qté réseau",
            "%Poids CA réseau", "Indice pression prix", "%Casse"]
    return agg[cols].sort_values(["Rayon", "Zone"])


def analyse5_contribution(df):
    agg = agreger(df, ["Zone"])
    total_ca, total_marge = agg["CA"].sum(), agg["Marge"].sum()
    agg["Poids CA réseau"] = agg["CA"] / total_ca
    agg["Poids Marge réseau"] = agg["Marge"] / total_marge
    agg["Écart Poids CA-Marge pts"] = agg["Poids CA réseau"] - agg["Poids Marge réseau"]
    return agg[["Zone", "CA", "Marge", "Poids CA réseau", "Poids Marge réseau", "Écart Poids CA-Marge pts"]]


def pareto(df_perimetre, valeur_col):
    d = df_perimetre.groupby("Article", as_index=False)[valeur_col].sum()
    d = d.sort_values(valeur_col, ascending=False).reset_index(drop=True)
    total = d[valeur_col].sum()
    d["cum"] = d[valeur_col].cumsum()
    d["cum_pct"] = d["cum"] / total if total else np.nan
    if total and (d["cum_pct"] >= 0.8).any():
        idx80 = (d["cum_pct"] >= 0.8).idxmax()
        nb80 = idx80 + 1
    else:
        nb80 = np.nan
    pct80 = nb80 / len(d) if len(d) and pd.notna(nb80) else np.nan
    return {"nb_articles": len(d), "nb_pour_80": nb80, "pct_articles_pour_80": pct80}


def analyse5_pareto(df):
    lignes = []
    for z in ["CD", "AB"]:
        dz = df[df["Zone"] == z]
        p_ca = pareto(dz, "CA")
        p_marge = pareto(dz, "Marge")
        lignes.append({"Zone": z, "Nb articles": p_ca["nb_articles"],
                        "% articles pour 80% CA": p_ca["pct_articles_pour_80"],
                        "% articles pour 80% Marge": p_marge["pct_articles_pour_80"]})
    return pd.DataFrame(lignes)


# ───────────────── Analyse 6 : Pareto 20/80 par Format ─────────────────────
def analyse6(df):
    lignes = []
    for fmt in ORDRE_FORMAT:
        d_fmt = df[df["Format"] == fmt]
        p_ca = pareto(d_fmt, "CA")
        p_marge = pareto(d_fmt, "Marge")
        lignes.append({"Format": fmt, "Nb articles": p_ca["nb_articles"],
                        "% articles pour 80% CA": p_ca["pct_articles_pour_80"],
                        "Nb articles pour 80% CA": p_ca["nb_pour_80"],
                        "% articles pour 80% Marge": p_marge["pct_articles_pour_80"],
                        "Nb articles pour 80% Marge": p_marge["nb_pour_80"]})
    return pd.DataFrame(lignes)


# ═══════════════════════════ ÉCRITURE EXCEL ════════════════════════════════
COLONNES_PCT = {
    "Evol CA %", "Evol CA HP %", "Evol CA Promo %", "%Poids Promo", "%Poids Promo N-1",
    "Evol Marge %", "%Marge", "%Marge N-1", "Evol Marge Promo %", "%Marge HP", "%Marge HP N-1",
    "%Marge Promo", "%Marge Promo N-1", "Evol Qté %", "%Casse", "% dans Famille", "% dans Rayon",
    "% articles pour 80% CA", "% articles pour 80% Marge", "Poids CA réseau", "Poids Marge réseau",
    "%Poids Qté réseau", "%Poids CA réseau",
}
COLONNES_PTS = {
    "Evol %Poids Promo pts", "Evol %Marge pts", "Evol %Marge HP pts", "Evol %Marge Promo pts",
    "Écart %Marge CD-AB pts", "Écart Poids CA-Marge pts",
}
COLONNES_FCFA = {
    "CA", "CA N-1", "CA Hors Promo", "CA Hors Promo N-1", "CA Promo", "CA Promo N-1",
    "Marge", "Marge N-1", "Marge Hors Promo", "Marge Hors Promo N-1", "Marge Promo", "Marge Promo N-1",
    "Casse (Valeur)", "Evol CA valeur", "Evol CA HP valeur", "Evol CA Promo valeur",
    "Evol Marge valeur", "Evol Marge HP valeur", "Evol Marge Promo valeur",
    "CA CD", "CA AB", "CA N-1 CD", "CA N-1 AB", "Marge CD", "Marge AB",
}
COLONNES_EVOL_COLOREE = {
    "Evol CA %", "Evol CA valeur", "Evol CA HP %", "Evol CA HP valeur",
    "Evol CA Promo %", "Evol CA Promo valeur", "Evol Marge %", "Evol Marge valeur",
    "Evol %Marge pts", "Evol %Marge HP pts", "Evol %Marge Promo pts",
    "Evol Qté %", "Evol Qté valeur", "Écart Poids CA-Marge pts", "Écart %Marge CD-AB pts",
}


def ecrire_feuille(wb, nom, df_out, titre):
    ws = wb.create_sheet(nom[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"] = titre
    ws["A1"].font = F_TITRE
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = FILL_TITRE
    ws.row_dimensions[1].height = 24

    header_row = 3
    for j, col in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.font = F_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    for i, row in enumerate(df_out.itertuples(index=False), start=header_row + 1):
        for j, col in enumerate(df_out.columns, start=1):
            val = getattr(row, col) if hasattr(row, col) else row[j - 1]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER_FIN

            if col in COLONNES_EVOL_COLOREE and val is not None:
                cell.font = couleur_evol(val)
            else:
                cell.font = F_NORMAL

            if col in COLONNES_FCFA:
                cell.number_format = FMT_FCFA
            elif col in COLONNES_PTS:
                cell.number_format = FMT_PTS
            elif col in COLONNES_PCT:
                cell.number_format = FMT_PCT

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(df_out.columns))}{header_row}"

    for j, col in enumerate(df_out.columns, start=1):
        largeur = max(12, min(28, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(j)].width = largeur

    return ws


def construire_excel(df, chemin_sortie):
    wb = Workbook()
    wb.remove(wb.active)

    ecrire_feuille(wb, "1_Format_Rayon", analyse1(df), "Analyse 1 — Format > Rayon")
    ecrire_feuille(wb, "2_Format_Rayon_Famille", analyse2(df), "Analyse 2 — Format > Rayon > Famille")
    ecrire_feuille(wb, "3_Magasin_Rayon_Famille", analyse3(df), "Analyse 3 — Magasin > Rayon > Famille")

    res4 = analyse4(df)
    for nom, d in res4.items():
        ecrire_feuille(wb, f"4_{nom}", d, f"Analyse 4 — {nom.replace('_', ' ')}")

    ecrire_feuille(wb, "5_Zone_Comparatif", analyse5_comparatif(df), "Analyse 5A — Zone CD vs AB, comparatif")
    ecrire_feuille(wb, "5_Zone_Positionnement", analyse5_positionnement(df), "Analyse 5B — Positionnement par Zone")
    ecrire_feuille(wb, "5_Zone_Contribution", analyse5_contribution(df), "Analyse 5C — Contribution au réseau")
    ecrire_feuille(wb, "5_Zone_Pareto", analyse5_pareto(df), "Analyse 5D — Pareto 20/80 par Zone")

    ecrire_feuille(wb, "6_Pareto_Format", analyse6(df), "Analyse 6 — Pareto 20/80 par Format")

    wb.save(chemin_sortie)


if __name__ == "__main__":
    import io
    import streamlit as st

    st.set_page_config(page_title="Revue perf", page_icon="📊", layout="wide")

    st.markdown(
        f"""
        <style>
        .stApp {{ background-color: #{GRIS_FOND}; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    col_logo, col_titre = st.columns([1, 8])
    with col_logo:
        st.markdown(
            f"""<div style="width:48px; height:48px; border-radius:12px;
            background:#{BLEU}; display:flex; align-items:center;
            justify-content:center; font-size:24px;">📊</div>""",
            unsafe_allow_html=True,
        )
    with col_titre:
        st.title("Revue de performance")

    fichier = st.file_uploader("Export PBI niveau Article (.xlsx)", type=["xlsx"])

    if fichier is not None:
        with st.spinner("Calcul des 6 analyses en cours..."):
            df = charger_donnees(fichier)
            buffer = io.BytesIO()
            construire_excel(df, buffer)
            buffer.seek(0)

        st.success(f"{len(df):,}".replace(",", " ") + " lignes article traitées.")

        st.download_button(
            label="Télécharger la revue de performance (.xlsx)",
            data=buffer,
            file_name="revue_performance_ADIALEA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Charge l'export pour lancer les 6 analyses.")
