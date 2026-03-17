"""
01_📊_Scoring_ABC.py — SmartBuyer Hub
Scoring ABC par Unité de Besoin · 5 règles de recommandation
Charte SmartBuyer v2 — Style Apple / SF Pro
"""

import io
from datetime import date
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Scoring ABC · SmartBuyer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1200px; }

[data-testid="stSidebar"] {
    background: #F2F2F7 !important;
    border-right: 0.5px solid #D1D1D6 !important;
}

[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 0.5px solid #E5E5EA !important;
    border-radius: 12px !important;
    padding: 16px 18px !important;
}
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }

[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }

[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }

[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
[data-testid="baseButton-primary"] { background: #007AFF !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; }

.stDownloadButton > button {
    background: #007AFF !important; color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important;
    padding: 10px 24px !important; width: 100% !important;
}
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.kpi-card { background: #FFFFFF; border: 0.5px solid #E5E5EA; border-radius: 12px; padding: 16px 18px; }

.badge { display: inline-block; padding: 3px 10px; border-radius: 100px; font-size: 11px; font-weight: 600; letter-spacing: 0.02em; }
.badge-new   { background: #F0EEFF; color: #5E35B1; }
.badge-must  { background: #E8F5E9; color: #2E7D32; }
.badge-top   { background: #E3F2FD; color: #1565C0; }
.badge-arb   { background: #FFF8E1; color: #F57F17; }
.badge-out   { background: #FFEBEE; color: #C62828; }

.rule-card { background: #FFFFFF; border: 0.5px solid #E5E5EA; border-radius: 10px; padding: 12px 14px; margin-bottom: 8px; }
.rule-num  { font-size: 10px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 3px; }
.rule-name { font-size: 13px; font-weight: 600; color: #1C1C1E; }
.rule-desc { font-size: 12px; color: #8E8E93; margin-top: 2px; line-height: 1.4; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name     { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc     { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.col-example  { font-size: 11px; color: #8E8E93; font-family: monospace; margin-top: 2px; }

.alert-box { border-radius: 10px; padding: 12px 16px; margin-bottom: 10px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-purple { background: #F5F0FF; border-color: #7C3AED; color: #4C1D95; }
.alert-blue   { background: #EFF6FF; border-color: #007AFF; color: #1E3A5F; }
.alert-red    { background: #FFF2F2; border-color: #FF3B30; color: #7B0000; }
.alert-green  { background: #F0FFF4; border-color: #34C759; color: #1A3A20; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
RECO_COL   = "Recommandation"
RECO_ORDER = ["NOUVEAUTÉ < N MOIS", "INCONTOURNABLE", "TOP PERFORMANCE", "À ARBITRER", "À NETTOYER (OUT)"]

RECO_COLORS = {
    "NOUVEAUTÉ < N MOIS": "#7C3AED",
    "INCONTOURNABLE":     "#059669",
    "TOP PERFORMANCE":    "#007AFF",
    "À ARBITRER":         "#FF9500",
    "À NETTOYER (OUT)":   "#FF3B30",
}
RECO_BADGE = {
    "NOUVEAUTÉ < N MOIS": "badge-new",
    "INCONTOURNABLE":     "badge-must",
    "TOP PERFORMANCE":    "badge-top",
    "À ARBITRER":         "badge-arb",
    "À NETTOYER (OUT)":   "badge-out",
}
RECO_COLORS_XL = {
    "NOUVEAUTÉ < N MOIS": ("7C3AED", "FFFFFF"),
    "INCONTOURNABLE":     ("059669", "FFFFFF"),
    "TOP PERFORMANCE":    ("007AFF", "FFFFFF"),
    "À ARBITRER":         ("FF9500", "FFFFFF"),
    "À NETTOYER (OUT)":   ("FF3B30", "FFFFFF"),
}
ACTIONS = {
    "NOUVEAUTÉ < N MOIS": "Laisser mûrir — réévaluer après la période",
    "INCONTOURNABLE":     "Protéger absolument",
    "TOP PERFORMANCE":    "Développer · Mettre en avant",
    "À ARBITRER":         "Analyser au cas par cas",
    "À NETTOYER (OUT)":   "Supprimer de l'assortiment",
}

REQUIRED_COLUMNS = {
    "UBD":               ("Unité de Besoin", "ex: LAITS EN POUDRE"),
    "Article":           ("Code article", "ex: 14005975"),
    "Prix de Vente":     ("Prix de vente unitaire (FCFA)", "ex: 1250"),
    "Qté":               ("Quantité vendue sur la période", "ex: 4872"),
    "Montant vente HT":  ("Chiffre d'affaires HT (FCFA)", "ex: 6090000"),
    "Montant marge":     ("Marge brute HT (FCFA)", "ex: 1218000"),
    "Date Création":     ("Date de création de l'article", "ex: 15/11/2025"),
}
OPTIONAL_COLUMNS = {
    "Libellé article":    "Nom de l'article",
    "Libellé sous-famille": "Sous-famille",
    "Libellé famille":    "Famille",
    "Libellé rayon":      "Rayon (pour diagnostic par rayon)",
}

# ─── LOGIQUE ABC ──────────────────────────────────────────────────────────────
def abc_class(series):
    total = series.sum()
    if total <= 0:
        return pd.Series(["C"] * len(series), index=series.index)
    si  = series.sort_values(ascending=False).index
    cum = series[si].cumsum() / total
    cls = pd.Series("C", index=series.index)
    cls[si[cum <= 0.50]] = "A"
    cls[si[(cum > 0.50) & (cum <= 0.80)]] = "B"
    return cls

def pricing_tag(series):
    mn, mx = series.min(), series.max()
    tags = pd.Series("", index=series.index)
    if mn == mx:
        tags[:] = "1er Prix / Produit Leader"
        return tags
    tags[series == mn] = "1er Prix"
    tags[series == mx] = "Produit Leader"
    return tags

def compute_reco(row, cutoff):
    if pd.notna(row["Date Création"]) and row["Date Création"] >= cutoff:
        return "NOUVEAUTÉ < N MOIS"
    if row["Pricing"] in ("1er Prix", "Produit Leader", "1er Prix / Produit Leader"):
        return "INCONTOURNABLE"
    if "A" in (row["ABC Qté"], row["ABC Vente"], row["ABC Marge"]):
        return "TOP PERFORMANCE"
    if row["ABC Qté"] == "C" and row["ABC Vente"] == "C" and row["ABC Marge"] == "C":
        return "À NETTOYER (OUT)"
    return "À ARBITRER"

@st.cache_data(show_spinner=False)
def process(file_bytes, analysis_date_str, months):
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()
    missing = set(REQUIRED_COLUMNS.keys()) - set(df.columns)
    if missing:
        raise ValueError(f"Colonnes manquantes : {', '.join(sorted(missing))}")
    for c in ["UBD","Libellé article","Libellé sous-famille","Libellé famille","Libellé rayon"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    df["Date Création"] = pd.to_datetime(df["Date Création"], errors="coerce")
    today  = pd.Timestamp(analysis_date_str)
    cutoff = today - pd.DateOffset(months=months)
    for metric, col in [("ABC Qté","Qté"),("ABC Vente","Montant vente HT"),("ABC Marge","Montant marge")]:
        df[metric] = df.groupby("UBD")[col].transform(abc_class)
    df["Pricing"]        = df.groupby("UBD")["Prix de Vente"].transform(pricing_tag)
    df[RECO_COL]         = df.apply(lambda r: compute_reco(r, cutoff), axis=1)
    df["Âge article (j)"]= (today - df["Date Création"]).dt.days
    return df, today, cutoff

def fmt(n): return f"{int(n):,}".replace(",", " ")

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    # Logo
    st.markdown("""
    <div style='margin-bottom:18px'>
      <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
      <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")

    # Navigation
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Navigation</div>", unsafe_allow_html=True)
    st.page_link("app.py",                                         label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",             label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",                      label="📈  Ventes PBI",         disabled=True)
    st.page_link("pages/03_📦_Detention_Top_CA.py",                label="📦  Détention Top CA",   disabled=True)
    st.page_link("pages/04_💸_Performance_Promo.py",               label="💸  Performance Promo",  disabled=True)
    st.page_link("pages/05_🏪_Suivi_Implantation.py",              label="🏪  Suivi Implantation", disabled=True)
    st.markdown("---")

    # Upload
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Fichier source</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Export Excel (.xlsx)", type=["xlsx"], label_visibility="collapsed")

    st.markdown("---")

    # Paramètres
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Paramètres</div>", unsafe_allow_html=True)
    analysis_date    = st.date_input("Date d'analyse", value=date.today())
    months_threshold = st.slider("Seuil nouveauté (mois)", 1, 12, 4)
    cutoff_display   = pd.Timestamp(str(analysis_date)) - pd.DateOffset(months=months_threshold)
    st.markdown(f"<div style='font-size:12px;color:#7C3AED;background:#F5F0FF;padding:8px 12px;border-radius:8px;border:0.5px solid #DDD6FE;margin-top:6px'>Nouveautés créées après<br><strong>{cutoff_display.strftime('%d/%m/%Y')}</strong></div>", unsafe_allow_html=True)

    st.markdown("---")
    # Règles résumé
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>5 règles</div>", unsafe_allow_html=True)
    rules_sidebar = [
        ("1","NOUVEAUTÉ","#7C3AED","Créé < N mois"),
        ("2","INCONTOURNABLE","#059669","1er Prix ou Leader"),
        ("3","TOP PERFORMANCE","#007AFF","A en Qté/Vente/Marge"),
        ("4","À ARBITRER","#FF9500","Ni A ni CCC"),
        ("5","À NETTOYER","#FF3B30","CCC sur les 3 axes"),
    ]
    for num, name, color, desc in rules_sidebar:
        st.markdown(f"""
        <div style='display:flex;gap:8px;align-items:flex-start;margin-bottom:6px'>
          <div style='width:18px;height:18px;border-radius:50%;background:{color};color:#fff;
                      font-size:10px;font-weight:700;display:flex;align-items:center;
                      justify-content:center;flex-shrink:0;margin-top:1px'>{num}</div>
          <div>
            <div style='font-size:12px;font-weight:600;color:#1C1C1E'>{name}</div>
            <div style='font-size:11px;color:#8E8E93'>{desc}</div>
          </div>
        </div>""", unsafe_allow_html=True)

# ─── PAGE PRINCIPALE ──────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📊 Scoring ABC — Analyse d'Assortiment</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Classement des articles par Unité de Besoin · 5 règles de recommandation · Protection automatique des nouveautés</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL (si pas de fichier) ─────────────────────────────────────
if uploaded is None:
    st.markdown("---")

    # Explication de la fonction
    st.markdown("""
    <div class='alert-box alert-blue'>
      <strong>ℹ️ À quoi sert ce module ?</strong><br>
      Le <strong>Scoring ABC</strong> classe automatiquement chaque article de ton assortiment en 5 catégories d'action, 
      calculées <em>en relatif à son Unité de Besoin</em> (UBD). Un article "faible" en absolu peut être 
      "TOP PERFORMANCE" dans sa famille — c'est cette nuance que le scoring capture.<br><br>
      Résultat : une liste d'actions claires par article — protéger, développer, arbitrer ou supprimer.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Colonnes requises
    st.markdown("<div class='section-label'>Colonnes obligatoires dans ton fichier Excel</div>", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    cols_items = list(REQUIRED_COLUMNS.items())
    for i, (col_name, (desc, example)) in enumerate(cols_items):
        target = c1 if i < 4 else c2
        with target:
            st.markdown(f"""
            <div class='col-required'>
              <div style='font-size:16px;margin-top:1px'>{'📅' if 'Date' in col_name else '🔢' if col_name in ('Qté','Prix de Vente','Montant vente HT','Montant marge') else '🏷️'}</div>
              <div>
                <div class='col-name'>{col_name}</div>
                <div class='col-desc'>{desc}</div>
                <div class='col-example'>{example}</div>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Colonnes optionnelles
    st.markdown("<div class='section-label'>Colonnes optionnelles (enrichissent l'analyse)</div>", unsafe_allow_html=True)
    opt_cols = st.columns(4)
    for i, (col_name, desc) in enumerate(OPTIONAL_COLUMNS.items()):
        with opt_cols[i % 4]:
            st.markdown(f"""
            <div style='background:#F9F9FB;border:0.5px solid #E5E5EA;border-radius:8px;padding:10px 12px;margin-bottom:6px'>
              <div style='font-size:12px;font-weight:600;color:#3A3A3C;font-family:monospace'>{col_name}</div>
              <div style='font-size:11px;color:#8E8E93;margin-top:2px'>{desc}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Les 5 règles
    st.markdown("<div class='section-label'>Les 5 règles de recommandation</div>", unsafe_allow_html=True)
    rules_detail = [
        ("1","NOUVEAUTÉ < N MOIS","#7C3AED","badge-new","Article créé récemment (< seuil paramétrable)", "Protégé automatiquement — ne peut pas être classé À NETTOYER.", "Laisser mûrir, réévaluer après la période"),
        ("2","INCONTOURNABLE","#059669","badge-must","Article au prix le plus bas ou le plus élevé de son UBD", "1er Prix ou Produit Leader dans sa famille.", "Protéger absolument"),
        ("3","TOP PERFORMANCE","#007AFF","badge-top","Au moins un A en Quantité, Vente HT ou Marge", "Performant sur au moins un axe de mesure.", "Développer, mettre en avant"),
        ("4","À ARBITRER","#FF9500","badge-arb","Ni A sur aucun axe, ni CCC sur les trois", "Profil mixte — nécessite une analyse individuelle.", "Analyser au cas par cas"),
        ("5","À NETTOYER (OUT)","#FF3B30","badge-out","C sur les 3 axes : Quantité, Vente ET Marge", "Aucune performance détectée dans son UBD.", "Candidat à la suppression"),
    ]
    r1, r2 = st.columns(2)
    for i, (num, name, color, badge_cls, cond, detail, action) in enumerate(rules_detail):
        with (r1 if i < 3 else r2):
            st.markdown(f"""
            <div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;
                        padding:14px 16px;margin-bottom:10px;border-left:3px solid {color}'>
              <div style='display:flex;align-items:center;gap:8px;margin-bottom:6px'>
                <div style='width:22px;height:22px;border-radius:50%;background:{color};color:#fff;
                            font-size:11px;font-weight:700;display:flex;align-items:center;
                            justify-content:center;flex-shrink:0'>{num}</div>
                <span class='badge {badge_cls}'>{name}</span>
              </div>
              <div style='font-size:13px;font-weight:500;color:#1C1C1E;margin-bottom:3px'>{cond}</div>
              <div style='font-size:12px;color:#6C6C70;margin-bottom:6px'>{detail}</div>
              <div style='font-size:11px;color:{color};font-weight:500'>→ {action}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("⬆️ Charge ton fichier Excel dans la sidebar pour lancer l'analyse.")
    st.stop()

# ─── TRAITEMENT ───────────────────────────────────────────────────────────────
file_bytes = uploaded.read()
with st.spinner("Calcul du scoring ABC…"):
    try:
        df, today, cutoff = process(file_bytes, str(analysis_date), months_threshold)
    except ValueError as e:
        st.error(f"❌ Fichier invalide : {e}")
        with st.expander("Voir les colonnes requises"):
            for col, (desc, ex) in REQUIRED_COLUMNS.items():
                st.markdown(f"- **`{col}`** — {desc} *(ex: {ex})*")
        st.stop()
    except Exception as e:
        st.error(f"❌ Erreur inattendue : {e}")
        st.stop()

total       = len(df)
total_ca    = df["Montant vente HT"].sum()
total_marge = df["Montant marge"].sum()
total_qte   = df["Qté"].sum()

# Remplacer le label générique par le vrai seuil
df[RECO_COL] = df[RECO_COL].str.replace("NOUVEAUTÉ < N MOIS", f"NOUVEAUTÉ < {months_threshold} MOIS", regex=False)
reco_order_real = [r.replace("NOUVEAUTÉ < N MOIS", f"NOUVEAUTÉ < {months_threshold} MOIS") for r in RECO_ORDER]
reco_colors_real = {r.replace("NOUVEAUTÉ < N MOIS", f"NOUVEAUTÉ < {months_threshold} MOIS"): v for r,v in RECO_COLORS.items()}
reco_badge_real  = {r.replace("NOUVEAUTÉ < N MOIS", f"NOUVEAUTÉ < {months_threshold} MOIS"): v for r,v in RECO_BADGE.items()}

g = df.groupby(RECO_COL).agg(
    Nb=("Article","count"), CA=("Montant vente HT","sum"),
    Marge=("Montant marge","sum"), Qte=("Qté","sum")
).reindex(reco_order_real).fillna(0).reset_index()

nb_nouv = int(g.loc[g[RECO_COL].str.startswith("NOUVEAUTÉ"),"Nb"].sum())

# Alerte nouveautés
if nb_nouv > 0:
    st.markdown(f"""
    <div class='alert-box alert-purple'>
      🟣 <strong>{nb_nouv} articles NOUVEAUTÉ</strong> détectés — créés après le <strong>{cutoff.strftime('%d/%m/%Y')}</strong>.
      Ces articles sont protégés et exclus de la liste À NETTOYER.
    </div>""", unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Synthèse", f"🟣 Nouveautés ({nb_nouv})", "🔍 Données", "📈 Visualisations", "⬇️ Export"
])

# ═══ TAB 1 — SYNTHÈSE ═════════════════════════════════════════════════════════
with tab1:
    st.markdown("<div class='section-label'>Indicateurs globaux</div>", unsafe_allow_html=True)
    k1,k2,k3,k4,k5,k6 = st.columns(6)
    k1.metric("Articles actifs",     fmt(total))
    k2.metric("UBD analysées",       str(df["UBD"].nunique()))
    k3.metric(f"Nouveautés < {months_threshold}m", str(nb_nouv))
    k4.metric("CA HT",               f"{total_ca/1e9:.2f} Mds")
    k5.metric("Marge HT",            f"{total_marge/1e9:.2f} Mds")
    k6.metric("Taux de marge",       f"{total_marge/total_ca*100:.1f}%")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Répartition par statut</div>", unsafe_allow_html=True)

    for _, row in g.iterrows():
        reco = row[RECO_COL]
        nb   = int(row["Nb"]); ca = row["CA"]; mg = row["Marge"]
        color = reco_colors_real.get(reco, "#8E8E93")
        badge_cls = reco_badge_real.get(reco, "")
        c1,c2,c3,c4,c5,c6 = st.columns([3,1.5,1.5,1.5,2,3])
        with c1: st.markdown(f"<span class='badge {badge_cls}'>{reco}</span>", unsafe_allow_html=True)
        with c2: st.markdown(f"<div style='font-size:15px;font-weight:600;color:{color};text-align:right'>{fmt(nb)}</div><div style='font-size:11px;color:#8E8E93;text-align:right'>{nb/total*100:.1f}% articles</div>", unsafe_allow_html=True)
        with c3: st.markdown(f"<div style='font-size:13px;color:#1C1C1E;text-align:right'>{ca/total_ca*100:.1f}% CA</div><div style='font-size:11px;color:#8E8E93;text-align:right'>{ca/1e9:.2f} Mds</div>", unsafe_allow_html=True)
        with c4: st.markdown(f"<div style='font-size:13px;color:#1C1C1E;text-align:right'>{mg/ca*100:.1f}% marge</div><div style='font-size:11px;color:#8E8E93;text-align:right'>{mg/total_marge*100:.1f}% total</div>", unsafe_allow_html=True)
        with c5: st.progress(nb/total)
        with c6: st.markdown(f"<div style='font-size:12px;color:#8E8E93'>→ {ACTIONS.get(reco.replace(f'NOUVEAUTÉ < {months_threshold} MOIS','NOUVEAUTÉ < N MOIS'), ACTIONS.get(reco,''))}</div>", unsafe_allow_html=True)
        st.markdown("<hr style='margin:6px 0'>", unsafe_allow_html=True)

    if "Libellé rayon" in df.columns:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Diagnostic par rayon</div>", unsafe_allow_html=True)
        rayons = sorted(df["Libellé rayon"].unique())
        rg = df.groupby("Libellé rayon").agg(Nb=("Article","count"),CA=("Montant vente HT","sum"),Marge=("Montant marge","sum")).reset_index()
        by_rayon = df.groupby(["Libellé rayon",RECO_COL]).size().unstack(fill_value=0)
        cols_r = st.columns(min(len(rayons),4))
        for i, rayon in enumerate(rayons):
            rd = rg[rg["Libellé rayon"]==rayon]
            if rd.empty: continue
            rd = rd.iloc[0]
            n_out  = int(by_rayon.loc[rayon, [c for c in by_rayon.columns if "NETTOYER" in c][0]]) if any("NETTOYER" in c for c in by_rayon.columns) and rayon in by_rayon.index else 0
            n_nouv = int(by_rayon.loc[rayon, [c for c in by_rayon.columns if "NOUVEAUTÉ" in c][0]]) if any("NOUVEAUTÉ" in c for c in by_rayon.columns) and rayon in by_rayon.index else 0
            pct_out = n_out/int(rd["Nb"])*100 if rd["Nb"]>0 else 0
            alert = "🔴" if pct_out>40 else ("🟠" if pct_out>25 else "🟢")
            with cols_r[i%len(cols_r)]:
                st.markdown(f"""
                <div class='kpi-card' style='margin-bottom:10px'>
                  <div style='font-size:11px;font-weight:600;color:#007AFF;text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px'>{rayon}</div>
                  <div style='font-size:18px;font-weight:600;color:#1C1C1E'>{fmt(int(rd['Nb']))} art.</div>
                  <div style='font-size:12px;color:#8E8E93;margin-top:3px'>CA {rd['CA']/1e9:.2f} Mds · Tx {rd['Marge']/rd['CA']*100:.1f}%</div>
                  <div style='font-size:12px;margin-top:5px'>{alert} <strong style='color:#FF3B30'>{pct_out:.0f}%</strong> à nettoyer · 🟣 <strong style='color:#7C3AED'>{n_nouv}</strong> nouveautés</div>
                </div>""", unsafe_allow_html=True)

# ═══ TAB 2 — NOUVEAUTÉS ═══════════════════════════════════════════════════════
with tab2:
    nouv_df = df[df[RECO_COL].str.startswith("NOUVEAUTÉ")].copy()
    nouv_df["Âge (jours)"] = (today - nouv_df["Date Création"]).dt.days

    if len(nouv_df) == 0:
        st.info("Aucune nouveauté détectée avec les paramètres actuels.")
    else:
        st.markdown(f"""
        <div class='alert-box alert-purple'>
          <strong>🟣 {len(nouv_df)} articles protégés</strong> — créés entre le <strong>{cutoff.strftime('%d/%m/%Y')}</strong> et le <strong>{today.strftime('%d/%m/%Y')}</strong>
        </div>""", unsafe_allow_html=True)

        k1,k2,k3 = st.columns(3)
        k1.metric("Âge moyen", f"{nouv_df['Âge (jours)'].mean():.0f} jours")
        k2.metric("Plus récent", f"{int(nouv_df['Âge (jours)'].min())} jours")
        k3.metric("Plus ancien", f"{int(nouv_df['Âge (jours)'].max())} jours")

        st.markdown("<br>", unsafe_allow_html=True)
        display_cols = [c for c in ["Article","Libellé article","UBD","Date Création","Âge (jours)","Prix de Vente","ABC Qté","ABC Vente","ABC Marge","Pricing"] if c in nouv_df.columns]
        st.dataframe(
            nouv_df[display_cols].sort_values("Date Création", ascending=False).reset_index(drop=True),
            use_container_width=True, height=420,
            column_config={
                "Date Création": st.column_config.DateColumn("Date Création", format="DD/MM/YYYY"),
                "Âge (jours)":   st.column_config.NumberColumn("Âge (j)", format="%d j"),
                "Prix de Vente": st.column_config.NumberColumn("Prix (FCFA)", format="%d"),
            }
        )

# ═══ TAB 3 — DONNÉES ══════════════════════════════════════════════════════════
with tab3:
    fc1,fc2,fc3 = st.columns([2,2,2])
    with fc1: filter_reco = st.multiselect("Recommandation", reco_order_real, default=reco_order_real)
    with fc2:
        ub_list = ["Toutes"] + sorted(df["UBD"].unique().tolist())
        filter_ub = st.selectbox("UBD", ub_list)
    with fc3: filter_abc = st.multiselect("Classe ABC Vente", ["A","B","C"], default=["A","B","C"])

    df_view = df.copy()
    if filter_reco: df_view = df_view[df_view[RECO_COL].isin(filter_reco)]
    if filter_ub != "Toutes": df_view = df_view[df_view["UBD"]==filter_ub]
    if filter_abc: df_view = df_view[df_view["ABC Vente"].isin(filter_abc)]

    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>{len(df_view):,} articles affichés</div>", unsafe_allow_html=True)

    display_cols = [c for c in ["Article","Libellé article","UBD","Date Création","Prix de Vente","Qté","Montant vente HT","Montant marge","ABC Qté","ABC Vente","ABC Marge","Pricing",RECO_COL] if c in df_view.columns]
    st.dataframe(
        df_view[display_cols].reset_index(drop=True),
        use_container_width=True, height=520,
        column_config={
            "Date Création":    st.column_config.DateColumn("Date Création", format="DD/MM/YYYY"),
            "Montant vente HT": st.column_config.NumberColumn("CA HT (FCFA)",    format="%d"),
            "Montant marge":    st.column_config.NumberColumn("Marge HT (FCFA)", format="%d"),
            "Prix de Vente":    st.column_config.NumberColumn("Prix (FCFA)",     format="%d"),
            "Qté":              st.column_config.NumberColumn("Qté",             format="%d"),
            RECO_COL:           st.column_config.TextColumn("Recommandation",    width="large"),
        }
    )

# ═══ TAB 4 — VISUALISATIONS ═══════════════════════════════════════════════════
with tab4:
    PL = dict(
        paper_bgcolor="#FFFFFF", plot_bgcolor="#F9F9FB",
        font=dict(family="-apple-system, Helvetica Neue, Arial", color="#3A3A3C", size=12),
        margin=dict(t=20, b=20, l=40, r=20),
    )

    v1, v2 = st.columns(2)
    with v1:
        st.markdown("<div class='section-label'>Articles par recommandation</div>", unsafe_allow_html=True)
        fig = go.Figure(go.Pie(
            labels=g[RECO_COL].tolist(), values=g["Nb"].tolist(), hole=0.58,
            marker_colors=[reco_colors_real.get(r,"#888") for r in g[RECO_COL]],
            marker=dict(line=dict(color="#FFFFFF", width=2)),
            textinfo="label+percent", textfont_size=10,
        ))
        fig.update_layout(**PL, showlegend=False, height=300)
        st.plotly_chart(fig, use_container_width=True)

    with v2:
        st.markdown("<div class='section-label'>Part de CA par recommandation</div>", unsafe_allow_html=True)
        fig2 = go.Figure(go.Bar(
            x=g[RECO_COL].tolist(), y=(g["CA"]/total_ca*100).round(1).tolist(),
            marker_color=[reco_colors_real.get(r,"#888") for r in g[RECO_COL]],
            marker_line_width=0,
            text=[f"{v:.1f}%" for v in (g["CA"]/total_ca*100)],
            textposition="outside",
        ))
        fig2.update_layout(**PL, yaxis=dict(title="% CA", gridcolor="#E5E5EA"), xaxis=dict(tickfont=dict(size=9)), height=300)
        st.plotly_chart(fig2, use_container_width=True)

    v3,v4,v5 = st.columns(3)
    for col_v, abc_col, title in [(v3,"ABC Qté","Quantités"),(v4,"ABC Vente","Vente HT"),(v5,"ABC Marge","Marge HT")]:
        with col_v:
            abc_count = df[abc_col].value_counts().reindex(["A","B","C"], fill_value=0)
            fig3 = go.Figure(go.Bar(
                x=["A","B","C"], y=abc_count.values,
                marker_color=["#34C759","#007AFF","#D1D1D6"],
                marker_line_width=0,
                text=abc_count.values, textposition="outside"
            ))
            fig3.update_layout(**PL, title=dict(text=title, font=dict(size=12,color="#8E8E93")), showlegend=False, height=240, yaxis=dict(gridcolor="#E5E5EA"))
            st.plotly_chart(fig3, use_container_width=True)

    if nb_nouv > 0:
        st.markdown("<div class='section-label'>Calendrier des nouveautés</div>", unsafe_allow_html=True)
        nouv_tl = df[df[RECO_COL].str.startswith("NOUVEAUTÉ")].copy()
        nouv_tl["Mois"] = nouv_tl["Date Création"].dt.to_period("M").astype(str)
        monthly = nouv_tl.groupby("Mois").size().reset_index(name="Nb")
        fig_t = go.Figure(go.Bar(
            x=monthly["Mois"], y=monthly["Nb"],
            marker_color="#7C3AED", marker_line_width=0,
            text=monthly["Nb"], textposition="outside"
        ))
        fig_t.update_layout(**PL, height=240, xaxis=dict(title="Mois de création"), yaxis=dict(title="Nb articles", gridcolor="#E5E5EA"))
        st.plotly_chart(fig_t, use_container_width=True)

# ═══ TAB 5 — EXPORT ═══════════════════════════════════════════════════════════
with tab5:
    st.markdown("<div class='section-label'>Récapitulatif avant export</div>", unsafe_allow_html=True)
    e_cols = st.columns(5)
    for i, row in g.iterrows():
        reco = row[RECO_COL]; nb = int(row["Nb"])
        color = reco_colors_real.get(reco,"#8E8E93")
        short = reco.replace(f"NOUVEAUTÉ < {months_threshold} MOIS","NOUVEAUTÉS").replace("TOP PERFORMANCE","TOP PERF.").replace("À ARBITRER","ARBITRER").replace("À NETTOYER (OUT)","NETTOYER")
        with e_cols[i%5]:
            st.markdown(f"""
            <div class='kpi-card' style='border-top:3px solid {color};text-align:center'>
              <div style='font-size:22px;font-weight:700;color:{color}'>{fmt(nb)}</div>
              <div style='font-size:10px;color:#8E8E93;margin-top:4px;font-weight:600;text-transform:uppercase;letter-spacing:.04em'>{short}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class='alert-box alert-blue'>
      <strong>📋 Contenu de l'export Excel</strong><br>
      · <strong>Données scorées</strong> — tableau complet + 5 colonnes colorées (ABC Qté/Vente/Marge, Pricing, Recommandation)<br>
      · <strong>📊 Synthèse</strong> — KPIs globaux + focus {nb_nouv} nouveauté(s)<br>
      · <strong>📋 Méthodologie</strong> — arbre de décision à 5 règles<br><br>
      Date d'analyse : <strong>{today.strftime('%d/%m/%Y')}</strong> · Seuil nouveauté : après le <strong>{cutoff.strftime('%d/%m/%Y')}</strong>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    def build_excel(df, original_bytes, today, cutoff):
        def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
        def tbdr(color="E5E5EA"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        DARK = "1C3557"; MID = "2C4A6E"; LGREY = "F2F2F7"
        ABC_COLORS  = {"A":("34C759","FFFFFF"),"B":("007AFF","FFFFFF"),"C":("D1D1D6","6C6C70")}
        PRICE_COLORS = {"1er Prix":("FF3B30","FFFFFF"),"Produit Leader":("7C3AED","FFFFFF"),"1er Prix / Produit Leader":("C62828","FFFFFF"),"":("FAFAFA","000000")}

        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb.active

        orig_ncols = sum(1 for c in df.columns if c not in ["ABC Qté","ABC Vente","ABC Marge","Pricing",RECO_COL,"Âge article (j)"])
        new_cols = [("ABC Qté",11),("ABC Vente",11),("ABC Marge",11),("Pricing",24),(RECO_COL,32)]

        for i,(h,w) in enumerate(new_cols):
            col = orig_ncols+1+i
            c = ws.cell(1,col,h)
            c.font=Font(bold=True,color="FFFFFF",size=10); c.fill=fill(DARK)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=tbdr()
            ws.column_dimensions[get_column_letter(col)].width=w

        date_col_idx = next((ci for ci in range(1,ws.max_column+1) if str(ws.cell(1,ci).value or "").strip()=="Date Création"), None)

        for i, row in df.iterrows():
            r = i+2; stripe = LGREY if i%2==0 else "FFFFFF"
            if date_col_idx:
                d = row["Date Création"]
                if pd.notna(d):
                    ws.cell(r,date_col_idx).value = d.to_pydatetime().date()
                    ws.cell(r,date_col_idx).number_format = "DD/MM/YYYY"
            for j,col_name in enumerate(["ABC Qté","ABC Vente","ABC Marge"]):
                col = orig_ncols+1+j; v = row[col_name]
                c = ws.cell(r,col,v)
                bg,fg = ABC_COLORS.get(v,("FFFFFF","000000"))
                c.fill=fill(bg); c.font=Font(bold=True,color=fg,size=10); c.alignment=Alignment(horizontal="center"); c.border=tbdr()
            col=orig_ncols+4; v=row["Pricing"]
            c=ws.cell(r,col,v)
            bg,fg=PRICE_COLORS.get(v,("FAFAFA","000000"))
            c.fill=fill(bg); c.font=Font(bold=bool(v),color=fg,size=9); c.alignment=Alignment(horizontal="center"); c.border=tbdr()
            col=orig_ncols+5; v=row[RECO_COL]
            c=ws.cell(r,col,v)
            rk = v.replace(f"NOUVEAUTÉ < {months_threshold} MOIS","NOUVEAUTÉ < N MOIS")
            bg,fg=RECO_COLORS_XL.get(rk,("FFFFFF","000000"))
            c.fill=fill(bg); c.font=Font(bold=True,color=fg,size=9); c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=tbdr()

        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:{get_column_letter(orig_ncols+5)}{len(df)+1}"

        ws2 = wb.create_sheet("📊 Synthèse")
        ws2.sheet_view.showGridLines=False
        ws2.merge_cells("A1:L1"); ws2["A1"]="SYNTHÈSE SCORING ABC — SMARTBUYER"
        ws2["A1"].font=Font(bold=True,size=15,color="FFFFFF"); ws2["A1"].fill=fill(DARK)
        ws2["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=36

        ws2.merge_cells("A2:L2")
        ws2["A2"]=f"{total:,} articles · {df['UBD'].nunique():,} UBD · CA : {total_ca/1e9:.2f} Mds · Marge : {total_marge/1e9:.2f} Mds · Date : {today.strftime('%d/%m/%Y')}"
        ws2["A2"].font=Font(italic=True,size=10,color="FFFFFF"); ws2["A2"].fill=fill(MID)
        ws2["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[2].height=18

        hdrs=["Statut","Nb Articles","% Articles","CA HT (FCFA)","% CA","Marge HT (FCFA)","% Marge","Tx Marge","Qté vendue","% Qté","Action"]
        widths=[28,13,11,18,10,18,10,10,14,10,30]
        for i,(h,w) in enumerate(zip(hdrs,widths),1):
            c=ws2.cell(4,i,h); c.font=Font(bold=True,color="FFFFFF",size=9); c.fill=fill(DARK)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=tbdr()
            ws2.column_dimensions[get_column_letter(i)].width=w
        ws2.row_dimensions[4].height=20

        for i,row in g.iterrows():
            r=5+i; reco=row[RECO_COL]; nb=row["Nb"]; ca=row["CA"]; mg=row["Marge"]; qt=row["Qte"]
            rk=reco.replace(f"NOUVEAUTÉ < {months_threshold} MOIS","NOUVEAUTÉ < N MOIS")
            bg_r,fg_r=RECO_COLORS_XL.get(rk,("FFFFFF","000000"))
            stripe=LGREY if i%2==0 else "FFFFFF"
            def dc(c,v,bg=stripe,fg="000000",bold=False,fmt=None,ha="center"):
                cell=ws2.cell(r,c,v); cell.font=Font(bold=bold,color=fg,size=9)
                cell.fill=fill(bg); cell.alignment=Alignment(horizontal=ha,vertical="center"); cell.border=tbdr()
                if fmt: cell.number_format=fmt
            dc(1,reco,bg=bg_r,fg=fg_r,bold=True)
            dc(2,int(nb),fmt="#,##0"); dc(3,nb/total,fmt="0.0%"); dc(4,ca,fmt="#,##0",ha="right")
            dc(5,ca/total_ca if total_ca else 0,fmt="0.0%"); dc(6,mg,fmt="#,##0",ha="right")
            dc(7,mg/total_marge if total_marge else 0,fmt="0.0%")
            dc(8,mg/ca if ca else 0,fmt="0.0%"); dc(9,int(qt),fmt="#,##0")
            dc(10,qt/total_qte if total_qte else 0,fmt="0.0%")
            action_key=reco.replace(f"NOUVEAUTÉ < {months_threshold} MOIS","NOUVEAUTÉ < N MOIS")
            dc(11,ACTIONS.get(action_key,ACTIONS.get(reco,"")),bg="FAFAFA",fg="666666",ha="left")
            ws2.row_dimensions[r].height=18

        buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

    with st.spinner("Génération du fichier Excel…"):
        excel_bytes = build_excel(df, file_bytes, today, cutoff)

    st.download_button(
        label=f"⬇️  Télécharger le fichier Excel — {total:,} articles scorés",
        data=excel_bytes,
        file_name=f"SmartBuyer_Scoring_ABC_{today.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(f"Compatible Excel 2016+ · Généré le {today.strftime('%d/%m/%Y')}")
