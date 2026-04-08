"""
SmartBuyer · On Time In Full — v6
──────────────────────────────────
Nouveauté v6 : la liste de surveillance affiche la CLASSE de l'article
  - load_watchlist retourne un dict {code: classe} (ex: {"10001801": "GOLD"})
  - La colonne "Surveillance" affiche la classe (GOLD, SILVER, A, B…) au lieu de ⭐
  - Partout : fond doré Excel + filtre conservés, label = classe
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import date as _date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG PAGE
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="On Time In Full · SmartBuyer",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════════════
# CHARTE GRAPHIQUE SMARTBUYER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1280px; }

[data-testid="stSidebar"] {
    background: #F2F2F7 !important;
    border-right: 0.5px solid #D1D1D6 !important;
}
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 0.5px solid #E5E5EA !important;
    border-radius: 12px !important;
    padding: 16px 18px !important;
}
[data-testid="stMetricLabel"] {
    font-size: 11px !important; font-weight: 500 !important;
    color: #8E8E93 !important; text-transform: uppercase !important;
    letter-spacing: 0.04em !important;
}
[data-testid="stMetricValue"] {
    font-size: 24px !important; font-weight: 600 !important;
    color: #1C1C1E !important; letter-spacing: -0.02em !important;
}
[data-testid="stTabs"] button[role="tab"] {
    font-size: 13px !important; font-weight: 500 !important;
    padding: 8px 16px !important; color: #8E8E93 !important;
    border-bottom: 2px solid transparent !important;
}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    color: #007AFF !important; border-bottom: 2px solid #007AFF !important;
    background: transparent !important;
}
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] {
    border: 0.5px solid #E5E5EA !important; border-radius: 10px !important;
}
[data-testid="stDataFrame"] th {
    background: #F2F2F7 !important; font-size: 11px !important;
    font-weight: 600 !important; color: #8E8E93 !important;
    text-transform: uppercase !important; letter-spacing: 0.04em !important;
}
[data-testid="stFileUploader"] {
    border: 1.5px dashed #D1D1D6 !important;
    border-radius: 10px !important; background: #F9F9FB !important;
}
.stDownloadButton > button {
    background: #007AFF !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 500 !important; font-size: 13px !important;
    padding: 10px 24px !important; width: 100% !important;
}
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label {
    font-size: 11px; font-weight: 600; color: #8E8E93;
    text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px;
}
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.6; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-gold  { background: #FFFBF0; border-color: #FFD60A; color: #3A2D00; }

.kpi-focus {
    background: #EFF6FF; border: 1px solid #B3D9FF;
    border-radius: 12px; padding: 16px 18px;
}
.kpi-focus-label { font-size: 11px; font-weight: 500; color: #007AFF; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 3px; }
.kpi-focus-value { font-size: 24px; font-weight: 700; color: #007AFF; letter-spacing: -0.02em; }
.kpi-focus-sub   { font-size: 12px; color: #0066CC; margin-top: 3px; font-weight: 500; }

.kpi-watch {
    background: #FFFDE7; border: 1px solid #FFD60A;
    border-radius: 12px; padding: 16px 18px;
}
.kpi-watch-label { font-size: 11px; font-weight: 500; color: #B8860B; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 3px; }
.kpi-watch-value { font-size: 24px; font-weight: 700; color: #B8860B; letter-spacing: -0.02em; }
.kpi-watch-sub   { font-size: 12px; color: #8B6914; margin-top: 3px; font-weight: 500; }

.doc-card {
    background: #FFFFFF; border: 0.5px solid #E5E5EA;
    border-radius: 12px; padding: 20px 22px; margin-bottom: 12px;
}
.doc-card-title { font-size: 13px; font-weight: 700; color: #1C1C1E; margin-bottom: 8px; }
.doc-card-body  { font-size: 13px; color: #3A3A3C; line-height: 1.7; }
code { background: #F2F2F7; padding: 2px 6px; border-radius: 4px; font-size: 12px; }

.fiche-header {
    background: #FFFFFF; border: 0.5px solid #E5E5EA;
    border-radius: 14px; padding: 22px 26px; margin-bottom: 16px;
}
.badge-watch {
    display: inline-block;
    background: #FFF9C4; border: 1px solid #FFD60A;
    border-radius: 4px; padding: 1px 6px;
    font-size: 10px; font-weight: 700; color: #7A5F00;
    letter-spacing: 0.03em; vertical-align: middle; margin-left: 4px;
}
.watch-banner {
    background: #FFFDE7; border: 1.5px solid #FFD60A;
    border-radius: 10px; padding: 10px 16px; margin-bottom: 12px;
    font-size: 12px; color: #7A5F00; line-height: 1.6;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES MÉTIER
# ══════════════════════════════════════════════════════════════════════════════
TECHNICAL_SUPPLIERS = {
    "FOURNISSEUR STOCK",
    "FOURNISSEUR PLATEFORME LOCAL",
    "FOURNISSEUR PLATEFORME IMPORT",
}
DATE_EXPECTED_CANDIDATES = ["H Date", "Date livraison", "Date prévue", "Date"]
SEUIL_EXCELLENT  = 97
SEUIL_SURVEILLER = 90

WATCH_GOLD_HEX  = "FFD60A"
WATCH_LIGHT_HEX = "FFFDE7"


# ══════════════════════════════════════════════════════════════════════════════
# CHARGEMENT LISTE SURVEILLANCE — retourne dict {code_str: classe}
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_watchlist(file_bytes: bytes, filename: str) -> dict:
    """
    Charge la liste de surveillance depuis un CSV ou Excel.
    Cherche :
      - une colonne 'code/article/ref/ean' → codes articles
      - une colonne 'classe/class/segment/categ' → libellé de classe (GOLD, A, B…)
    Retourne un dict {code_normalisé: classe_str}.
    Si pas de colonne classe trouvée, la classe = "⭐" par défaut.
    """
    try:
        if filename.lower().endswith((".xlsx", ".xls")):
            wdf = pd.read_excel(BytesIO(file_bytes))
        else:
            wdf = pd.read_csv(BytesIO(file_bytes), sep=None, engine="python")

        wdf.columns = [str(c).strip() for c in wdf.columns]

        # Détection colonne code
        code_col = None
        for col in wdf.columns:
            if any(kw in col.lower() for kw in ["code", "article", "ref", "ean"]):
                code_col = col
                break
        if code_col is None:
            code_col = wdf.columns[0]

        # Détection colonne classe
        classe_col = None
        for col in wdf.columns:
            if col == code_col:
                continue
            if any(kw in col.lower() for kw in ["class", "segment", "categ", "tier", "niveau", "groupe"]):
                classe_col = col
                break
        # Si toujours pas trouvé et qu'il y a une 2e colonne, on la prend
        if classe_col is None and len(wdf.columns) >= 2:
            classe_col = [c for c in wdf.columns if c != code_col][0]

        def norm(v):
            return str(v).strip().replace(".0", "")

        result = {}
        for _, row in wdf.dropna(subset=[code_col]).iterrows():
            code = norm(row[code_col])
            if not code or code == "nan":
                continue
            classe = str(row[classe_col]).strip() if classe_col and pd.notna(row[classe_col]) else "⭐"
            result[code] = classe

        return result

    except Exception as e:
        st.sidebar.error(f"Erreur lecture liste surveillance : {e}")
        return {}


def normalise_code(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip().replace(".0", "")


def watchdict_codes(watchdict: dict) -> set:
    """Retourne l'ensemble des codes de la watchlist."""
    return set(watchdict.keys())


def get_classe(code, watchdict: dict) -> str:
    """Retourne la classe de l'article ou '' s'il n'est pas dans la watchlist."""
    return watchdict.get(normalise_code(code), "")


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS AFFICHAGE
# ══════════════════════════════════════════════════════════════════════════════
def fmt(n) -> str:
    if pd.isna(n) or n is None:
        return "—"
    a = abs(float(n))
    if a >= 1_000_000:
        return f"{n / 1_000_000:.1f} M"
    if a >= 1_000:
        return f"{int(n / 1_000)} K"
    return f"{int(n):,}"


def fmt_pct(v, decimals: int = 1) -> str:
    if pd.isna(v) or v is None:
        return "—"
    return f"{v:.{decimals}f}%"


def score_band(v) -> str:
    if pd.isna(v):
        return "Inconnu"
    if v >= SEUIL_EXCELLENT:
        return "🟢 Excellent"
    if v >= SEUIL_SURVEILLER:
        return "🟠 À surveiller"
    return "🔴 Critique"


def score_color(v) -> str:
    if pd.isna(v):
        return "#8E8E93"
    if v >= SEUIL_EXCELLENT:
        return "#34C759"
    if v >= SEUIL_SURVEILLER:
        return "#FF9500"
    return "#FF3B30"


def safe_div(a, b) -> float:
    return a / b if b not in (0, None) and not pd.isna(b) else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# CHARGEMENT & NETTOYAGE ERP
# ══════════════════════════════════════════════════════════════════════════════
def detect_expected_date_column(df: pd.DataFrame):
    for col in DATE_EXPECTED_CANDIDATES:
        if col in df.columns:
            if pd.to_datetime(df[col], errors="coerce").notna().sum() > 0:
                return col
    return None


@st.cache_data(show_spinner=False)
def load_erp(file_bytes: bytes, filename: str) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(file_bytes), sep=";", low_memory=False)
    df.columns = [str(c).replace("\ufeff", "").strip().rstrip(",.") for c in df.columns]
    df = df.dropna(axis=1, how="all")
    df = df.drop(columns=[c for c in df.columns if c.startswith("Unnamed:")], errors="ignore")
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip().replace({"nan": None, "": None})
    date_cols = [
        "Dt Rec", "Date de commande", "Date", "Date facture",
        "Date comptable du rapprochement", "H Date", "Date livraison", "Date prévue",
    ]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format="%d/%m/%Y", errors="coerce")
    num_cols = [
        "Site", "Département", "N° Cde", "Sit", "Fou", "Famille", "Sous-famille",
        "Code", "Article", "Qté cde", "Poids cde", "Qté rec", "Poids rec",
        "PV", "Px revient", "Prix de vente HT", "Marge ligne", "Taux TVA",
        "EAN", "Poids unitaire", "PCB", "Colis",
    ]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def prepare_dataset(df: pd.DataFrame, exclude_technical: bool = True,
                    cap_sur_receipt: bool = True, watchdict: dict = None):
    if df.empty:
        return df.copy(), {}

    work = df.copy()
    raw_len = len(work)

    def _col(name):
        return work[name] if name in work.columns else pd.Series("Inconnu", index=work.index)

    work["site_label"]    = _col("Libellé site").fillna("Inconnu").astype(str).str.strip()
    supplier_col = next((c for c in work.columns if c.startswith("Nom fourn")), None)
    work["supplier_name"] = (work[supplier_col] if supplier_col else pd.Series("Inconnu", index=work.index)).fillna("Inconnu").astype(str).str.strip()
    work["article_label"] = _col("Libellé article").fillna("Inconnu").astype(str).str.strip()
    work["dept_label"]    = _col("Libellé département").fillna("Inconnu").astype(str).str.strip()
    work["famille_label"] = _col("Libellé famille").fillna("Inconnu").astype(str).str.strip()

    work["qte_cde"] = pd.to_numeric(work.get("Qté cde"), errors="coerce").fillna(0)
    work["qte_rec"] = pd.to_numeric(work.get("Qté rec"), errors="coerce").fillna(0)
    work["pv_ht"]   = pd.to_numeric(work.get("Prix de vente HT"), errors="coerce").fillna(0)
    pv_zero_rows    = int((work["pv_ht"] == 0).sum())

    is_tech     = work["supplier_name"].str.upper().isin(TECHNICAL_SUPPLIERS)
    is_zero_qty = work["qte_cde"] <= 0
    sur_receipt = int((work["qte_rec"] > work["qte_cde"]).sum())
    excl_tech   = int(is_tech.sum())
    excl_zero   = int(is_zero_qty.sum())

    if exclude_technical:
        work = work[~is_tech].copy()
    work = work[work["qte_cde"] > 0].copy()

    work["qte_rec_retained"] = (
        work[["qte_rec", "qte_cde"]].min(axis=1) if cap_sur_receipt else work["qte_rec"]
    )

    expected_col = detect_expected_date_column(work)
    work["date_expected"] = work[expected_col] if expected_col else pd.NaT
    work["date_received"] = work["Dt Rec"] if "Dt Rec" in work.columns else pd.NaT

    missing_exp = int(work["date_expected"].isna().sum())

    work["qty_missing"]       = (work["qte_cde"] - work["qte_rec_retained"]).clip(lower=0)
    work["service_gap_value"] = work["qty_missing"] * work["pv_ht"]
    work["line_fill_rate"]    = np.where(
        work["qte_cde"] > 0, work["qte_rec_retained"] / work["qte_cde"], 0.0
    )
    work["delay_days"] = (work["date_received"] - work["date_expected"]).dt.days
    work["on_time"] = work["date_expected"].isna() | (work["date_received"] <= work["date_expected"])
    work["otif"]    = ((work["qte_rec_retained"] >= work["qte_cde"]) & work["on_time"]).astype(int)
    work["criticality_score"] = work["service_gap_value"] * (1 - work["line_fill_rate"])

    # ── FLAG SURVEILLANCE : colonne "watch_classe" = classe ou "" 
    if watchdict and "Code" in work.columns:
        work["code_str"]    = work["Code"].apply(normalise_code)
        work["watch_classe"] = work["code_str"].map(watchdict).fillna("")
        work["is_watched"]   = work["watch_classe"] != ""
    else:
        work["watch_classe"] = ""
        work["is_watched"]   = False

    quality = {
        "raw_rows":              raw_len,
        "clean_rows":            len(work),
        "excluded_zero_qty":     excl_zero,
        "excluded_technical":    excl_tech,
        "sur_receipt_rows":      sur_receipt,
        "missing_expected_date": missing_exp,
        "all_dates_missing":     (missing_exp == len(work)),
        "pv_zero_rows":          pv_zero_rows,
        "expected_col":          expected_col or "Aucune",
        "usable_rate":           round(safe_div(len(work), raw_len) * 100, 1),
        "watched_in_data":       int(work["is_watched"].sum()),
    }
    return work, quality


# ══════════════════════════════════════════════════════════════════════════════
# KPI GLOBAUX
# ══════════════════════════════════════════════════════════════════════════════
def compute_global_kpis(df: pd.DataFrame) -> dict:
    if df.empty:
        return dict(fill_rate=0, on_time=0, otif=0, score=0,
                    ordered_qty=0, received_qty=0, missing_qty=0,
                    impact_value=0, suppliers=0, articles=0, orders=0, sites=0)
    ordered   = df["qte_cde"].sum()
    received  = df["qte_rec_retained"].sum()
    fill_rate = safe_div(received, ordered) * 100
    on_time   = df["on_time"].mean() * 100
    otif      = df["otif"].mean() * 100
    score     = 0.5 * fill_rate + 0.3 * on_time + 0.2 * otif
    return dict(
        fill_rate=fill_rate, on_time=on_time, otif=otif, score=score,
        ordered_qty=ordered, received_qty=received,
        missing_qty=df["qty_missing"].sum(),
        impact_value=df["service_gap_value"].sum(),
        suppliers=df["supplier_name"].nunique(),
        articles=df["Code"].nunique() if "Code" in df.columns else 0,
        orders=df["N° Cde"].nunique() if "N° Cde" in df.columns else 0,
        sites=df["site_label"].nunique(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# AGRÉGATIONS
# ══════════════════════════════════════════════════════════════════════════════
def _enrich(g: pd.DataFrame) -> pd.DataFrame:
    g["fill_rate"] = np.where(g["qte_cde"] > 0, g["qte_rec"] / g["qte_cde"] * 100, 0.0)
    g["on_time"]  *= 100
    g["otif"]     *= 100
    g["score"]     = 0.5 * g["fill_rate"] + 0.3 * g["on_time"] + 0.2 * g["otif"]
    g["Niveau"]    = g["score"].apply(score_band)
    g["criticality_score"] = g["impact_value"] * (1 - g["fill_rate"] / 100)
    return g


def agg_supplier(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby(["Fou", "supplier_name"], as_index=False).agg(
        qte_cde      =("qte_cde",          "sum"),
        qte_rec      =("qte_rec_retained",  "sum"),
        qty_missing  =("qty_missing",       "sum"),
        impact_value =("service_gap_value", "sum"),
        on_time      =("on_time",           "mean"),
        otif         =("otif",              "mean"),
        orders       =("N° Cde",            "nunique"),
        articles     =("Code",              "nunique"),
        sites        =("site_label",        "nunique"),
    )
    return _enrich(g).sort_values("criticality_score", ascending=False).reset_index(drop=True)


def agg_site(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby(["Site", "site_label"], as_index=False).agg(
        qte_cde      =("qte_cde",          "sum"),
        qte_rec      =("qte_rec_retained",  "sum"),
        qty_missing  =("qty_missing",       "sum"),
        impact_value =("service_gap_value", "sum"),
        on_time      =("on_time",           "mean"),
        otif         =("otif",              "mean"),
        suppliers    =("supplier_name",     "nunique"),
        articles     =("Code",             "nunique"),
    )
    return _enrich(g).sort_values("criticality_score", ascending=False).reset_index(drop=True)


def agg_article(df: pd.DataFrame, watchdict: dict = None) -> pd.DataFrame:
    gcols = [c for c in ["Code", "article_label", "supplier_name"] if c in df.columns]
    g = df.groupby(gcols, as_index=False).agg(
        qte_cde      =("qte_cde",          "sum"),
        qte_rec      =("qte_rec_retained",  "sum"),
        qty_missing  =("qty_missing",       "sum"),
        impact_value =("service_gap_value", "sum"),
        on_time      =("on_time",           "mean"),
        otif         =("otif",              "mean"),
        sites        =("site_label",        "nunique"),
        orders       =("N° Cde",            "nunique"),
    )
    g = _enrich(g).sort_values("criticality_score", ascending=False).reset_index(drop=True)

    # Colonne Classe (vide si hors watchlist)
    if watchdict and "Code" in g.columns:
        g["code_str"]  = g["Code"].apply(normalise_code)
        g["Classe"]    = g["code_str"].map(watchdict).fillna("")
    else:
        g["Classe"] = ""
    return g


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUES PLOTLY
# ══════════════════════════════════════════════════════════════════════════════
_PLOTLY_BASE = dict(
    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
    font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=11),
    margin=dict(t=10, b=10, l=10, r=80),
    xaxis=dict(showgrid=True, gridcolor="#F2F2F7"),
    yaxis=dict(showgrid=False, title=""),
)


def bar_h(data: pd.DataFrame, x_col: str, y_col: str, color: str,
          x_title: str, height: int = 500, fmt_fn=None,
          classe_col: str = None) -> go.Figure:
    top = data.head(15).sort_values(x_col)
    texts = [fmt_fn(v) if fmt_fn else f"{v:,.0f}" for v in top[x_col]]

    # Couleur barre : dorée si article surveillé
    if classe_col and classe_col in top.columns:
        bar_colors = [
            "#FFD60A" if str(c).strip() != "" else color
            for c in top[classe_col]
        ]
    else:
        bar_colors = color

    # Label Y : préfixe avec la classe si surveillé
    y_labels = top[y_col].astype(str).tolist()
    if classe_col and classe_col in top.columns:
        y_labels = [
            f"[{c}] {lbl}" if str(c).strip() != "" else lbl
            for lbl, c in zip(y_labels, top[classe_col])
        ]

    fig = go.Figure(go.Bar(
        x=top[x_col], y=y_labels,
        orientation="h", marker_color=bar_colors, marker_line_width=0,
        text=texts, textposition="outside",
    ))
    fig.update_layout(**{**_PLOTLY_BASE, "height": height,
                         "xaxis": {**_PLOTLY_BASE["xaxis"], "title": x_title}})
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# COMPOSANT KPI ROW
# ══════════════════════════════════════════════════════════════════════════════
def render_kpi_row(kpi: dict, watch_kpi: dict = None):
    if watch_kpi:
        c1, c2, c3, c4, c5, c6 = st.columns(6)
    else:
        c1, c2, c3, c4, c5 = st.columns(5)

    c1.metric("Fill Rate",    fmt_pct(kpi["fill_rate"]))
    c2.metric("On Time",      fmt_pct(kpi["on_time"]))
    c3.metric("OTIF",         fmt_pct(kpi["otif"]))
    c4.metric("Score global", fmt_pct(kpi["score"]))
    with c5:
        st.markdown(f"""
<div class='kpi-focus'>
  <div class='kpi-focus-label'>Volume manquant</div>
  <div class='kpi-focus-value'>{fmt(kpi['missing_qty'])}</div>
  <div class='kpi-focus-sub'>Impact CA proxy : {fmt(kpi['impact_value'])} FCFA</div>
</div>""", unsafe_allow_html=True)

    if watch_kpi:
        with c6:
            st.markdown(f"""
<div class='kpi-watch'>
  <div class='kpi-watch-label'>⭐ Articles surveillés</div>
  <div class='kpi-watch-value'>{fmt(watch_kpi['missing_qty'])}</div>
  <div class='kpi-watch-sub'>Impact : {fmt(watch_kpi['impact_value'])} FCFA · FR {fmt_pct(watch_kpi['fill_rate'])}</div>
</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — helpers couleur classe
# ══════════════════════════════════════════════════════════════════════════════
GOLD_FILL  = PatternFill("solid", fgColor=WATCH_GOLD_HEX)
GOLDL_FILL = PatternFill("solid", fgColor=WATCH_LIGHT_HEX)
SILVER_FILL = PatternFill("solid", fgColor="E8E8E8")
SILVERL_FILL = PatternFill("solid", fgColor="F5F5F5")


def _classe_fills(classe: str):
    """Retourne (fill_ligne, fill_badge) selon la classe."""
    c = str(classe).strip().upper()
    if c == "GOLD":
        return GOLDL_FILL, GOLD_FILL
    elif c == "SILVER":
        return SILVERL_FILL, SILVER_FILL
    elif c in ("A", ""):
        return GOLDL_FILL, GOLD_FILL
    else:
        return SILVERL_FILL, SILVER_FILL


def _xl_write_sheet(ws, title: str, df: pd.DataFrame,
                    hdr_fill=None, hdr_font=None, ctr=None,
                    classe_col_idx: int = None):
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=13, color="1C3557")
    ws.append([])
    headers = list(df.columns)
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(3, i)
        if hdr_fill: c.fill = hdr_fill
        if hdr_font: c.font = hdr_font
        if ctr:      c.alignment = ctr
    for row in df.itertuples(index=False):
        ws.append(list(row))
        if classe_col_idx is not None:
            val = str(list(row)[classe_col_idx - 1]).strip()
            if val != "":
                fill_l, _ = _classe_fills(val)
                n = ws.max_row
                for col_i in range(1, len(headers) + 1):
                    ws.cell(n, col_i).fill = fill_l
    for col_cells in ws.iter_cols(min_row=3, max_row=3):
        letter = get_column_letter(col_cells[0].column)
        val = str(col_cells[0].value or "")
        ws.column_dimensions[letter].width = max(12, min(32, len(val) + 4))


def build_export_excel(df, by_supplier, by_site, by_article, quality,
                       watchdict: dict = None) -> BytesIO:
    wb = Workbook()
    H_FILL = PatternFill("solid", fgColor="1C3557")
    H_FONT = Font(bold=True, color="FFFFFF", size=11)
    CTR    = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "Synthese"
    synthese = pd.DataFrame([
        ["Lignes brutes",                      quality.get("raw_rows", 0)],
        ["Lignes exploitables",                quality.get("clean_rows", 0)],
        ["Taux exploitable %",                 quality.get("usable_rate", 0)],
        ["Date prévue utilisée",               quality.get("expected_col", "Aucune")],
        ["Toutes dates prévues manquantes",     "OUI" if quality.get("all_dates_missing") else "NON"],
        ["Qté cde ≤ 0 exclues",                quality.get("excluded_zero_qty", 0)],
        ["Fournisseurs techniques exclus",      quality.get("excluded_technical", 0)],
        ["Sur-réceptions capées",              quality.get("sur_receipt_rows", 0)],
        ["Dates prévues manquantes",            quality.get("missing_expected_date", 0)],
        ["Lignes PV HT = 0 (impact nul)",      quality.get("pv_zero_rows", 0)],
        ["Articles surveillance matchés",       quality.get("watched_in_data", 0)],
    ], columns=["Indicateur", "Valeur"])
    _xl_write_sheet(ws1, "Synthèse qualité de données", synthese, H_FILL, H_FONT, CTR)

    art_cols = list(by_article.columns)
    classe_idx = (art_cols.index("Classe") + 1) if "Classe" in art_cols else None

    _xl_write_sheet(wb.create_sheet("Par fournisseur"), "Analyse fournisseur", by_supplier, H_FILL, H_FONT, CTR)
    _xl_write_sheet(wb.create_sheet("Par magasin"),     "Analyse magasin",     by_site,     H_FILL, H_FONT, CTR)
    _xl_write_sheet(wb.create_sheet("Par article"),     "Analyse article",     by_article,  H_FILL, H_FONT, CTR,
                    classe_col_idx=classe_idx)

    # Onglet Articles Surveillés
    if watchdict and "is_watched" in df.columns:
        watched_art = by_article[by_article["Classe"] != ""].copy() if "Classe" in by_article.columns else pd.DataFrame()
        if not watched_art.empty:
            ws_w = wb.create_sheet("⭐ Articles surveillés")
            _xl_write_sheet(ws_w, "Articles de la liste surveillance", watched_art, H_FILL, H_FONT, CTR,
                            classe_col_idx=classe_idx)
            for i in range(1, len(watched_art.columns) + 1):
                c = ws_w.cell(3, i)
                c.fill = PatternFill("solid", fgColor=WATCH_GOLD_HEX)
                c.font = Font(bold=True, color="1C1C1E", size=11)

    # Lignes critiques
    detail_cols = [c for c in [
        "date_received", "date_expected", "site_label", "supplier_name",
        "Code", "article_label", "N° Cde", "qte_cde", "qte_rec_retained",
        "qty_missing", "service_gap_value", "on_time", "otif", "delay_days", "watch_classe",
    ] if c in df.columns]
    crit = (df[df["otif"] == 0][detail_cols]
            .sort_values(["is_watched", "qty_missing", "service_gap_value"],
                         ascending=[False, False, False])
            .head(500))
    if "watch_classe" in crit.columns:
        crit = crit.copy().rename(columns={"watch_classe": "Classe"})
    crit_classe_idx = (list(crit.columns).index("Classe") + 1) if "Classe" in crit.columns else None
    _xl_write_sheet(wb.create_sheet("Lignes critiques"), "Lignes non OTIF", crit, H_FILL, H_FONT, CTR,
                    classe_col_idx=crit_classe_idx)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — FICHE FOURNISSEUR
# ══════════════════════════════════════════════════════════════════════════════
def build_fiche_excel(fournisseur: str, df_all: pd.DataFrame, df_bad: pd.DataFrame,
                      site_recap: pd.DataFrame, art_recap: pd.DataFrame,
                      kpis: dict, seuil: int, watchdict: dict = None) -> BytesIO:
    wb = Workbook()
    today_str = _date.today().strftime("%d/%m/%Y")

    H_FILL   = PatternFill("solid", fgColor="1C3557")
    H_FONT   = Font(bold=True, color="FFFFFF", size=10)
    T_FONT   = Font(bold=True, size=13, color="1C3557")
    S_FONT   = Font(bold=True, size=10)
    CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RED_FILL = PatternFill("solid", fgColor="FFF2F2")
    AMB_FILL = PatternFill("solid", fgColor="FFFBF0")
    GRN_FILL = PatternFill("solid", fgColor="F0FFF4")
    THIN     = Border(
        left=Side(style="thin", color="E5E5EA"),
        right=Side(style="thin", color="E5E5EA"),
        top=Side(style="thin", color="E5E5EA"),
        bottom=Side(style="thin", color="E5E5EA"),
    )

    def auto_width(ws, min_w=12, max_w=34):
        for col in ws.iter_cols():
            best = min_w
            for cell in col:
                if cell.value:
                    best = max(best, min(max_w, len(str(cell.value)) + 3))
            ws.column_dimensions[get_column_letter(col[0].column)].width = best

    def write_header_row(ws, row_n, headers, gold=False):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row_n, i, h)
            c.fill = PatternFill("solid", fgColor=WATCH_GOLD_HEX) if gold else H_FILL
            c.font = Font(bold=True, color="1C1C1E" if gold else "FFFFFF", size=10)
            c.alignment = CTR; c.border = THIN

    def kpi_fill(val_str: str):
        try:
            v = float(str(val_str).replace("%", ""))
            return GRN_FILL if v >= SEUIL_EXCELLENT else (AMB_FILL if v >= SEUIL_SURVEILLER else RED_FILL)
        except Exception:
            return None

    def get_code_classe(code) -> str:
        if not watchdict:
            return ""
        return watchdict.get(normalise_code(code), "")

    # ── Onglet 1 : Synthèse KPI
    ws1 = wb.active
    ws1.title = "Synthèse KPI"
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 18
    ws1.cell(1, 1, f"BILAN DE PERFORMANCE — {fournisseur.upper()}").font = Font(bold=True, size=14, color="1C1C1E")
    ws1.cell(2, 1, f"Généré le {today_str}  ·  Période : export ERP  ·  Seuil : Fill Rate < {seuil}%").font = Font(size=10, italic=True, color="8E8E93")
    ws1.merge_cells("A1:D1"); ws1.merge_cells("A2:D2")
    ws1.append([])
    kpi_data = [
        ("Fill Rate",       fmt_pct(kpis["fill_rate"]),    "Part de la qté commandée effectivement livrée"),
        ("On Time",         fmt_pct(kpis["on_time"]),      "Part des livraisons respectant la date prévue"),
        ("OTIF",            fmt_pct(kpis["otif"]),         "Livraisons complètes ET à l'heure"),
        ("Score global",    fmt_pct(kpis["score"]),        "50% × Fill Rate + 30% × On Time + 20% × OTIF"),
        ("Niveau",          score_band(kpis["score"]),     "Évaluation synthétique"),
        ("Vol. manquant",   f"{int(kpis['missing_qty']):,} unités",  "Qté commandée − Qté reçue retenue"),
        ("Impact CA proxy", f"{kpis['impact_value']:,.0f} FCFA",     "Vol. manquant × Prix de vente HT"),
        ("Commandes",       str(kpis["orders"]),           "Nombre de N° de commandes distincts"),
        ("Articles",        str(kpis["articles"]),         "Références concernées"),
        ("Magasins",        str(kpis["sites"]),            "Sites livrés sur la période"),
    ]
    write_header_row(ws1, 4, ["Indicateur", "Valeur", "Commentaire"])
    for i, (label, val, comment) in enumerate(kpi_data, 5):
        ws1.cell(i, 1, label).font = S_FONT; ws1.cell(i, 1).border = THIN
        c_val = ws1.cell(i, 2, val)
        c_val.alignment = CTR; c_val.border = THIN
        if label in ("Fill Rate", "On Time", "OTIF", "Score global"):
            f = kpi_fill(val)
            if f: c_val.fill = f
        ws1.cell(i, 3, comment).alignment = LFT; ws1.cell(i, 3).border = THIN
    auto_width(ws1)

    # ── Onglet 2 : Par magasin
    ws2 = wb.create_sheet("Par magasin")
    ws2.cell(1, 1, f"Livraisons incomplètes par magasin — seuil Fill Rate < {seuil}%").font = T_FONT
    ws2.append([])
    headers2 = ["Magasin", "Fill Rate", "Vol. manquant", "Impact CA proxy (FCFA)", "Cmdes", "Articles"]
    write_header_row(ws2, 3, headers2)
    for _, r in site_recap.sort_values("qty_missing", ascending=False).iterrows():
        ws2.append([r.get("site_label",""), r.get("Fill Rate",""), r.get("Vol. manquant",""),
                    r.get("Impact CA proxy",""), r.get("nb_cdes",""), r.get("nb_articles","")])
        n = ws2.max_row
        for col_i in range(1, 7):
            ws2.cell(n, col_i).border = THIN
            ws2.cell(n, col_i).alignment = CTR if col_i > 1 else LFT
        f = kpi_fill(str(r.get("Fill Rate", ""))) if r.get("Fill Rate") else None
        if f: ws2.cell(n, 2).fill = f
    auto_width(ws2)

    # ── Onglet 3 : Par article (avec colonne Classe)
    ws3 = wb.create_sheet("Par article")
    ws3.cell(1, 1, f"Livraisons incomplètes par article — seuil Fill Rate < {seuil}%").font = T_FONT
    ws3.append([])
    headers3 = ["Classe", "Code article", "Désignation", "Fill Rate", "Vol. manquant",
                "Impact CA proxy (FCFA)", "Magasins", "Cmdes"]
    write_header_row(ws3, 3, headers3)
    for _, r in art_recap.iterrows():
        code   = r.get("Code", "")
        classe = get_code_classe(code)
        row_data = [
            classe,
            code, r.get("article_label",""),
            r.get("Fill Rate",""), r.get("Vol. manquant",""),
            r.get("Impact CA proxy",""), r.get("nb_sites",""), r.get("nb_cdes",""),
        ]
        n = ws3.max_row + 1
        ws3.append(row_data)
        for col_i in range(1, 9):
            ws3.cell(n, col_i).border = THIN
            ws3.cell(n, col_i).alignment = LFT if col_i in (2, 3) else CTR
        if classe:
            fill_l, fill_b = _classe_fills(classe)
            for col_i in range(1, 9):
                ws3.cell(n, col_i).fill = fill_l
            ws3.cell(n, 1).fill = fill_b   # badge classe plus sombre
        else:
            f = kpi_fill(str(r.get("Fill Rate", ""))) if r.get("Fill Rate") else None
            if f: ws3.cell(n, 4).fill = f
    auto_width(ws3)

    # ── Onglet 4 : Détail lignes
    ws4 = wb.create_sheet("Détail lignes")
    ws4.cell(1, 1, f"Détail des livraisons incomplètes — {fournisseur}  (Fill Rate < {seuil}%)").font = T_FONT
    ws4.append([])
    headers4 = [
        "Classe", "N° Commande", "Date réception", "Date prévue", "Magasin",
        "Code article", "Désignation article",
        "Qté commandée", "Qté reçue", "Qté manquante",
        "Fill Rate ligne", "Impact CA proxy (FCFA)",
        "Livré à l'heure", "Retard (jours)",
    ]
    write_header_row(ws4, 3, headers4)

    src_cols = [c for c in [
        "N° Cde", "date_received", "date_expected", "site_label",
        "Code", "article_label", "qte_cde", "qte_rec_retained",
        "qty_missing", "line_fill_rate", "service_gap_value", "on_time", "delay_days",
    ] if c in df_bad.columns]

    for _, row in df_bad[src_cols].iterrows():
        fr     = row.get("line_fill_rate", 0)
        ot     = row.get("on_time", True)
        code   = row.get("Code", "")
        classe = get_code_classe(code)
        n      = ws4.max_row + 1
        ws4.append([
            classe,
            row.get("N° Cde", ""),
            row.get("date_received", pd.NaT),
            row.get("date_expected", pd.NaT),
            row.get("site_label", ""),
            code,
            row.get("article_label", ""),
            int(row.get("qte_cde", 0)),
            int(row.get("qte_rec_retained", 0)),
            int(row.get("qty_missing", 0)),
            f"{fr * 100:.1f}%" if pd.notna(fr) else "—",
            round(row.get("service_gap_value", 0), 0),
            "OUI" if ot else "NON",
            int(row.get("delay_days", 0)) if pd.notna(row.get("delay_days")) else "—",
        ])
        for col_i in range(1, len(headers4) + 1):
            cell = ws4.cell(n, col_i)
            cell.border = THIN
            cell.alignment = LFT if col_i in (5, 7) else CTR
        if classe:
            fill_l, fill_b = _classe_fills(classe)
            for col_i in range(1, len(headers4) + 1):
                ws4.cell(n, col_i).fill = fill_l
            ws4.cell(n, 1).fill = fill_b
        else:
            row_fill = RED_FILL if pd.notna(fr) and fr < 0.90 else None
            if row_fill:
                for col_i in (10, 11, 12):
                    ws4.cell(n, col_i).fill = row_fill
            fr_fill = kpi_fill(f"{fr*100:.1f}%") if pd.notna(fr) else None
            if fr_fill:
                ws4.cell(n, 11).fill = fr_fill

    # ── Onglet 5 : Articles surveillés uniquement
    if watchdict:
        df_bad_watched = df_bad[df_bad["is_watched"]] if "is_watched" in df_bad.columns else df_bad.iloc[0:0]
        if not df_bad_watched.empty:
            ws5 = wb.create_sheet("⭐ Articles surveillés")
            ws5.cell(1, 1, f"Articles de surveillance — {fournisseur}").font = T_FONT
            ws5.cell(2, 1, f"Filtre : Fill Rate < {seuil}% · liste surveillance active").font = Font(size=9, italic=True, color="8E8E93")
            ws5.append([])
            write_header_row(ws5, 4, headers4, gold=True)
            for _, row in df_bad_watched[src_cols].iterrows():
                fr     = row.get("line_fill_rate", 0)
                ot     = row.get("on_time", True)
                code   = row.get("Code", "")
                classe = get_code_classe(code)
                n      = ws5.max_row + 1
                ws5.append([
                    classe,
                    row.get("N° Cde", ""),
                    row.get("date_received", pd.NaT),
                    row.get("date_expected", pd.NaT),
                    row.get("site_label", ""),
                    code,
                    row.get("article_label", ""),
                    int(row.get("qte_cde", 0)),
                    int(row.get("qte_rec_retained", 0)),
                    int(row.get("qty_missing", 0)),
                    f"{fr * 100:.1f}%" if pd.notna(fr) else "—",
                    round(row.get("service_gap_value", 0), 0),
                    "OUI" if ot else "NON",
                    int(row.get("delay_days", 0)) if pd.notna(row.get("delay_days")) else "—",
                ])
                for col_i in range(1, len(headers4) + 1):
                    cell = ws5.cell(n, col_i)
                    cell.border = THIN
                    cell.alignment = LFT if col_i in (5, 7) else CTR
                fill_l, fill_b = _classe_fills(classe)
                for col_i in range(1, len(headers4) + 1):
                    ws5.cell(n, col_i).fill = fill_l
                ws5.cell(n, 1).fill = fill_b
                fr_fill = kpi_fill(f"{fr*100:.1f}%") if pd.notna(fr) else None
                if fr_fill:
                    ws5.cell(n, 11).fill = fr_fill
            auto_width(ws5)
            ws5.freeze_panes = "A5"

    for ws in [ws2, ws3, ws4]:
        ws.freeze_panes = "A4"

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — TOUTES LES FICHES
# ══════════════════════════════════════════════════════════════════════════════
def build_export_all_fiches(df: pd.DataFrame, by_supplier: pd.DataFrame,
                             seuil: int, watchdict: dict = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Toutes livraisons incomplètes"
    today_str = _date.today().strftime("%d/%m/%Y")

    H_FILL    = PatternFill("solid", fgColor="1C3557")
    H_FONT    = Font(bold=True, color="FFFFFF", size=10)
    T_FONT    = Font(bold=True, size=13, color="1C3557")
    GRN_FILL  = PatternFill("solid", fgColor="E8F8EE")
    AMB_FILL  = PatternFill("solid", fgColor="FFF8E8")
    RED_FILL  = PatternFill("solid", fgColor="FFF0F0")
    CTR       = Alignment(horizontal="center", vertical="center")
    LFT       = Alignment(horizontal="left",   vertical="center")
    THIN_SIDE = Side(style="thin", color="E5E5EA")
    THIN      = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    ws.cell(1, 1, "LIVRAISONS INCOMPLÈTES — TOUS FOURNISSEURS").font = T_FONT
    ws.cell(2, 1, (
        f"Généré le {today_str}  ·  "
        f"Seuil : Fill Rate < {seuil}% ou OTIF = 0  ·  "
        f"Trié par criticité fournisseur"
        + ("  ·  Classe = segment 20/80" if watchdict else "")
    )).font = Font(size=9, italic=True, color="8E8E93")
    ws.merge_cells("A1:R1"); ws.merge_cells("A2:R2")
    ws.append([])

    HEADERS = [
        "Classe", "Fournisseur", "Code Fou.", "Score four.", "Niveau",
        "N° Commande", "Date réception", "Date prévue",
        "Magasin", "Code article", "Désignation article",
        "Qté commandée", "Qté reçue", "Qté manquante",
        "Fill Rate ligne", "Impact CA proxy (FCFA)",
        "OTIF", "Retard (j)",
    ]
    ws.append(HEADERS)
    hdr_row = 4
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(hdr_row, i)
        c.value = h; c.fill = H_FILL; c.font = H_FONT
        c.alignment = CTR; c.border = THIN

    sup_meta = (
        by_supplier[["supplier_name", "Fou", "score", "Niveau", "criticality_score"]]
        .copy()
        .rename(columns={"Fou": "fou_code", "criticality_score": "crit_sup"})
    )
    mask = (df["line_fill_rate"] < seuil / 100) | (df["otif"] == 0)
    export_df = df[mask].copy()
    export_df = export_df.merge(sup_meta, on="supplier_name", how="left")
    export_df = export_df.sort_values(
        ["is_watched", "crit_sup", "qty_missing", "service_gap_value"],
        ascending=[False, False, False, False],
    )

    for _, row in export_df.iterrows():
        fr      = row.get("line_fill_rate", np.nan)
        ot      = row.get("otif", 0)
        dly     = row.get("delay_days", np.nan)
        classe  = str(row.get("watch_classe", "")).strip()

        row_data = [
            classe,
            row.get("supplier_name", ""),
            int(row.get("fou_code", 0)) if pd.notna(row.get("fou_code")) else "",
            f"{row.get('score', 0):.1f}%",
            row.get("Niveau", ""),
            row.get("N° Cde", ""),
            row.get("date_received", pd.NaT),
            row.get("date_expected", pd.NaT),
            row.get("site_label", ""),
            row.get("Code", ""),
            row.get("article_label", ""),
            int(row.get("qte_cde", 0)),
            int(row.get("qte_rec_retained", 0)),
            int(row.get("qty_missing", 0)),
            f"{fr * 100:.1f}%" if pd.notna(fr) else "—",
            round(row.get("service_gap_value", 0), 0),
            "OUI" if ot == 1 else "NON",
            int(dly) if pd.notna(dly) else "—",
        ]
        n = ws.max_row + 1
        ws.append(row_data)

        for col_i in range(1, len(HEADERS) + 1):
            cell = ws.cell(n, col_i)
            cell.border = THIN
            cell.alignment = LFT if col_i in (2, 5, 9, 11) else CTR

        if classe:
            fill_l, fill_b = _classe_fills(classe)
            for col_i in range(1, len(HEADERS) + 1):
                ws.cell(n, col_i).fill = fill_l
            ws.cell(n, 1).fill = fill_b
        else:
            if pd.notna(fr):
                fr_pct = fr * 100
                fill_fr = GRN_FILL if fr_pct >= 97 else (AMB_FILL if fr_pct >= 90 else RED_FILL)
                ws.cell(n, 15).fill = fill_fr
            niv = row.get("Niveau", "")
            if "Excellent"    in str(niv): ws.cell(n, 5).fill = GRN_FILL
            elif "surveiller" in str(niv): ws.cell(n, 5).fill = AMB_FILL
            elif "Critique"   in str(niv): ws.cell(n, 5).fill = RED_FILL
            ws.cell(n, 17).fill = GRN_FILL if ot == 1 else RED_FILL

    # Onglet surveillance uniquement
    if watchdict:
        ws_w = wb.create_sheet("⭐ Articles surveillés")
        ws_w.cell(1, 1, "ARTICLES SURVEILLANCE — TOUS FOURNISSEURS").font = T_FONT
        ws_w.merge_cells("A1:R1")
        ws_w.append([])
        ws_w.append(HEADERS)
        watch_hdr_row = 3
        for i, h in enumerate(HEADERS, 1):
            c = ws_w.cell(watch_hdr_row, i)
            c.value = h
            c.fill = PatternFill("solid", fgColor=WATCH_GOLD_HEX)
            c.font = Font(bold=True, color="1C1C1E", size=10)
            c.alignment = CTR; c.border = THIN

        watched_rows = export_df[export_df["is_watched"] == True]
        for _, row in watched_rows.iterrows():
            fr     = row.get("line_fill_rate", np.nan)
            ot     = row.get("otif", 0)
            dly    = row.get("delay_days", np.nan)
            classe = str(row.get("watch_classe", "")).strip()
            ws_w.append([
                classe,
                row.get("supplier_name", ""),
                int(row.get("fou_code", 0)) if pd.notna(row.get("fou_code")) else "",
                f"{row.get('score', 0):.1f}%",
                row.get("Niveau", ""),
                row.get("N° Cde", ""),
                row.get("date_received", pd.NaT),
                row.get("date_expected", pd.NaT),
                row.get("site_label", ""),
                row.get("Code", ""),
                row.get("article_label", ""),
                int(row.get("qte_cde", 0)),
                int(row.get("qte_rec_retained", 0)),
                int(row.get("qty_missing", 0)),
                f"{fr * 100:.1f}%" if pd.notna(fr) else "—",
                round(row.get("service_gap_value", 0), 0),
                "OUI" if ot == 1 else "NON",
                int(dly) if pd.notna(dly) else "—",
            ])
            n = ws_w.max_row
            for col_i in range(1, len(HEADERS) + 1):
                cell = ws_w.cell(n, col_i)
                cell.border = THIN
                cell.alignment = LFT if col_i in (2, 5, 9, 11) else CTR
            fill_l, fill_b = _classe_fills(classe)
            for col_i in range(1, len(HEADERS) + 1):
                ws_w.cell(n, col_i).fill = fill_l
            ws_w.cell(n, 1).fill = fill_b
            if pd.notna(fr):
                fr_pct = fr * 100
                fill_fr = GRN_FILL if fr_pct >= 97 else (AMB_FILL if fr_pct >= 90 else RED_FILL)
                ws_w.cell(n, 15).fill = fill_fr

        ws_w.freeze_panes = "A4"
        col_widths = [10, 28, 10, 11, 14, 14, 16, 14, 22, 13, 32, 13, 11, 14, 14, 22, 8, 11]
        for i, w in enumerate(col_widths, 1):
            ws_w.column_dimensions[get_column_letter(i)].width = w

    col_widths = [10, 28, 10, 11, 14, 14, 16, 14, 22, 13, 32, 13, 11, 14, 14, 22, 8, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{ws.max_row}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[hdr_row].height = 28

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — RÉCAP PRIORISATION FOURNISSEURS
# ══════════════════════════════════════════════════════════════════════════════
def build_recap_fournisseur_excel(df: pd.DataFrame, watchdict: dict = None) -> BytesIO:
    """
    Export récap fournisseur orienté priorisation d'actions.
    1 ligne par fournisseur avec :
      - Refs totales commandées / Refs effectivement livrées (qte_rec > 0) / Taux couverture
      - Fill Rate / On Time / OTIF / Score / Niveau
      - Sites impactés (avec livraison incomplète) / Sites total
      - Vol. manquant / Impact CA proxy
      - Nb refs GOLD / Nb refs SILVER en sous-service (fill rate < 97%)
      - Rang criticité
    Trié par score de criticité décroissant.
    Filtre automatique + freeze + mise en forme conditionnelle couleur.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Récap priorisation fournisseurs"
    today_str = _date.today().strftime("%d/%m/%Y")

    H_FILL    = PatternFill("solid", fgColor="1C3557")
    H_FONT    = Font(bold=True, color="FFFFFF", size=10)
    T_FONT    = Font(bold=True, size=13, color="1C3557")
    GRN_FILL  = PatternFill("solid", fgColor="E8F8EE")
    AMB_FILL  = PatternFill("solid", fgColor="FFF8E8")
    RED_FILL  = PatternFill("solid", fgColor="FFF0F0")
    GOLD_L    = PatternFill("solid", fgColor="FFFDE7")
    CTR       = Alignment(horizontal="center", vertical="center")
    LFT       = Alignment(horizontal="left",   vertical="center")
    THIN_SIDE = Side(style="thin", color="E5E5EA")
    THIN      = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    # ── Titre + sous-titre
    ws.cell(1, 1, "RÉCAP PRIORISATION FOURNISSEURS — OTIF").font = T_FONT
    ws.cell(2, 1,
        f"Généré le {today_str}  ·  "
        "Trié par criticité décroissante  ·  "
        "Réfs livrées = au moins 1 unité reçue sur la période"
        + ("  ·  Cols GOLD/SILVER = nb réfs en sous-service (FR < 97%)" if watchdict else "")
    ).font = Font(size=9, italic=True, color="8E8E93")
    total_cols = 19 if watchdict else 17
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws.append([])

    # ── Construction des colonnes selon présence watchdict
    HEADERS = [
        "Rang",
        "Fournisseur", "Code Fou.",
        "Réfs commandées", "Réfs livrées", "Taux couverture réfs",
        "Fill Rate", "On Time", "OTIF", "Score", "Niveau",
        "Sites impactés", "Sites total",
        "Vol. manquant", "Impact CA proxy (FCFA)",
        "Nb commandes", "Nb lignes",
    ]
    if watchdict:
        HEADERS += ["Réfs GOLD sous-service", "Réfs SILVER sous-service"]

    ws.append(HEADERS)
    hdr_row = 4
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(hdr_row, i)
        c.value = h; c.fill = H_FILL; c.font = H_FONT
        c.alignment = CTR; c.border = THIN
    ws.row_dimensions[hdr_row].height = 30

    # ── Agrégation par fournisseur
    # Refs totales = nb de codes distincts commandés
    # Refs livrées = nb de codes distincts avec au moins 1 unité reçue
    # Sites impactés = sites où fill rate < 100% (au moins une ligne incomplète)
    grp = df.groupby(["Fou", "supplier_name"], as_index=False)

    agg = grp.agg(
        refs_total    =("Code",              "nunique"),
        qte_cde_sum   =("qte_cde",           "sum"),
        qte_rec_sum   =("qte_rec_retained",  "sum"),
        qty_missing   =("qty_missing",        "sum"),
        impact_value  =("service_gap_value",  "sum"),
        on_time_mean  =("on_time",            "mean"),
        otif_mean     =("otif",               "mean"),
        nb_orders     =("N° Cde",             "nunique"),
        nb_lines      =("Code",               "count"),
        sites_total   =("site_label",         "nunique"),
    )

    # Refs livrées : codes avec qte_rec > 0 (au moins 1 unité)
    refs_livrees = (
        df[df["qte_rec_retained"] > 0]
        .groupby(["Fou", "supplier_name"], as_index=False)["Code"]
        .nunique()
        .rename(columns={"Code": "refs_livrees"})
    )
    agg = agg.merge(refs_livrees, on=["Fou", "supplier_name"], how="left")
    agg["refs_livrees"] = agg["refs_livrees"].fillna(0).astype(int)

    # Sites impactés = sites avec au moins une ligne fill rate < 100%
    sites_impactes = (
        df[df["line_fill_rate"] < 1.0]
        .groupby(["Fou", "supplier_name"], as_index=False)["site_label"]
        .nunique()
        .rename(columns={"site_label": "sites_impactes"})
    )
    agg = agg.merge(sites_impactes, on=["Fou", "supplier_name"], how="left")
    agg["sites_impactes"] = agg["sites_impactes"].fillna(0).astype(int)

    # KPIs calculés
    agg["fill_rate"]       = np.where(agg["qte_cde_sum"] > 0, agg["qte_rec_sum"] / agg["qte_cde_sum"] * 100, 0.0)
    agg["on_time_pct"]     = agg["on_time_mean"] * 100
    agg["otif_pct"]        = agg["otif_mean"]    * 100
    agg["score"]           = 0.5 * agg["fill_rate"] + 0.3 * agg["on_time_pct"] + 0.2 * agg["otif_pct"]
    agg["taux_couverture"] = np.where(agg["refs_total"] > 0, agg["refs_livrees"] / agg["refs_total"] * 100, 0.0)
    agg["criticality"]     = agg["impact_value"] * (1 - agg["fill_rate"] / 100)
    agg                    = agg.sort_values("criticality", ascending=False).reset_index(drop=True)
    agg["rang"]            = agg.index + 1

    # Refs 20/80 en sous-service par fournisseur
    if watchdict:
        df_w = df.copy()
        df_w["watch_c"] = df_w["Code"].apply(normalise_code).map(watchdict).fillna("")
        # Agrégation fill rate par (fournisseur, code)
        art_fr = (
            df_w.groupby(["Fou", "supplier_name", "Code", "watch_c"], as_index=False)
            .agg(qte_c=("qte_cde","sum"), qte_r=("qte_rec_retained","sum"))
        )
        art_fr["fr"] = np.where(art_fr["qte_c"] > 0, art_fr["qte_r"] / art_fr["qte_c"] * 100, 0.0)
        art_under = art_fr[art_fr["fr"] < 97]

        gold_under = (
            art_under[art_under["watch_c"] == "GOLD"]
            .groupby(["Fou", "supplier_name"])["Code"].nunique()
            .reset_index().rename(columns={"Code": "refs_gold_under"})
        )
        silver_under = (
            art_under[art_under["watch_c"] == "SILVER"]
            .groupby(["Fou", "supplier_name"])["Code"].nunique()
            .reset_index().rename(columns={"Code": "refs_silver_under"})
        )
        agg = agg.merge(gold_under,   on=["Fou", "supplier_name"], how="left")
        agg = agg.merge(silver_under, on=["Fou", "supplier_name"], how="left")
        agg["refs_gold_under"]   = agg["refs_gold_under"].fillna(0).astype(int)
        agg["refs_silver_under"] = agg["refs_silver_under"].fillna(0).astype(int)

    # ── Écriture des lignes
    def _niveau(score):
        if score >= SEUIL_EXCELLENT:  return "Excellent"
        if score >= SEUIL_SURVEILLER: return "À surveiller"
        return "Critique"

    def _score_fill(score):
        if score >= SEUIL_EXCELLENT:  return GRN_FILL
        if score >= SEUIL_SURVEILLER: return AMB_FILL
        return RED_FILL

    for _, row in agg.iterrows():
        score    = row["score"]
        fr       = row["fill_rate"]
        ot       = row["on_time_pct"]
        otif     = row["otif_pct"]
        taux_cov = row["taux_couverture"]

        data = [
            int(row["rang"]),
            row["supplier_name"],
            int(row["Fou"]) if pd.notna(row["Fou"]) else "",
            int(row["refs_total"]),
            int(row["refs_livrees"]),
            f"{taux_cov:.1f}%",
            f"{fr:.1f}%",
            f"{ot:.1f}%",
            f"{otif:.1f}%",
            f"{score:.1f}%",
            _niveau(score),
            int(row["sites_impactes"]),
            int(row["sites_total"]),
            int(row["qty_missing"]),
            round(row["impact_value"], 0),
            int(row["nb_orders"]),
            int(row["nb_lines"]),
        ]
        if watchdict:
            data += [int(row["refs_gold_under"]), int(row["refs_silver_under"])]

        n = ws.max_row + 1
        ws.append(data)

        for col_i in range(1, len(HEADERS) + 1):
            cell = ws.cell(n, col_i)
            cell.border = THIN
            cell.alignment = LFT if col_i == 2 else CTR

        # Coloration Fill Rate (col 7)
        ws.cell(n, 7).fill  = _score_fill(fr)
        # Coloration OTIF (col 9)
        ws.cell(n, 9).fill  = _score_fill(otif)
        # Coloration Score (col 10)
        ws.cell(n, 10).fill = _score_fill(score)
        # Coloration Niveau (col 11)
        ws.cell(n, 11).fill = _score_fill(score)
        # Taux couverture (col 6)
        ws.cell(n, 6).fill  = _score_fill(taux_cov)

        # GOLD/SILVER : fond doré/gris si > 0
        if watchdict:
            if int(row["refs_gold_under"]) > 0:
                ws.cell(n, 18).fill = PatternFill("solid", fgColor=WATCH_GOLD_HEX)
                ws.cell(n, 18).font = Font(bold=True, color="1C1C1E", size=10)
            if int(row["refs_silver_under"]) > 0:
                ws.cell(n, 19).fill = PatternFill("solid", fgColor="E8E8E8")

    # ── Mise en forme finale
    col_widths = [6, 30, 10, 16, 14, 20, 12, 12, 10, 10, 16, 14, 12, 16, 22, 14, 12]
    if watchdict:
        col_widths += [18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{ws.max_row}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14

    # ── Onglet légende
    ws_leg = wb.create_sheet("Légende")
    ws_leg.cell(1, 1, "LÉGENDE — Récap priorisation fournisseurs").font = Font(bold=True, size=12, color="1C3557")
    leg_data = [
        ("Réfs commandées",         "Nombre de codes articles distincts présents dans au moins 1 commande sur la période"),
        ("Réfs livrées",            "Nombre de codes articles avec au moins 1 unité reçue (qte_rec > 0)"),
        ("Taux couverture réfs",    "Réfs livrées / Réfs commandées × 100 — mesure la largeur de l'assortiment effectivement approvisionné"),
        ("Fill Rate",               "Qté reçue retenue / Qté commandée × 100 — mesure la profondeur du service"),
        ("On Time",                 "% de lignes livrées à la date prévue ou avant"),
        ("OTIF",                    "% de lignes complètes ET à l'heure"),
        ("Score",                   "50% × Fill Rate + 30% × On Time + 20% × OTIF"),
        ("Niveau",                  "Excellent ≥ 97% · À surveiller 90–97% · Critique < 90%"),
        ("Sites impactés",          "Nb de magasins avec au moins une ligne Fill Rate < 100%"),
        ("Vol. manquant",           "Somme (Qté commandée − Qté reçue retenue) pour toutes les lignes"),
        ("Impact CA proxy",         "Vol. manquant × Prix de vente HT — estimation du CA non réalisé"),
        ("Réfs GOLD sous-service",  "Nb de réfs GOLD (watchlist) avec Fill Rate < 97% chez ce fournisseur"),
        ("Réfs SILVER sous-service","Nb de réfs SILVER (watchlist) avec Fill Rate < 97% chez ce fournisseur"),
    ]
    ws_leg.append([])
    ws_leg.append(["Indicateur", "Définition"])
    for i in range(1, 3):
        ws_leg.cell(3, i).fill = H_FILL
        ws_leg.cell(3, i).font = H_FONT
        ws_leg.cell(3, i).alignment = CTR
        ws_leg.cell(3, i).border = THIN
    for label, defn in leg_data:
        ws_leg.append([label, defn])
        n = ws_leg.max_row
        ws_leg.cell(n, 1).font = Font(bold=True, size=10)
        ws_leg.cell(n, 1).border = THIN
        ws_leg.cell(n, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_leg.cell(n, 2).border = THIN
        ws_leg.row_dimensions[n].height = 22
    ws_leg.column_dimensions["A"].width = 26
    ws_leg.column_dimensions["B"].width = 72
    ws_leg.freeze_panes = "A4"

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>📦 SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · On Time In Full</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichier ERP</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Extraction ERP OTIF (CSV ; )", type=["csv"], key="otif")

    st.markdown("---")

    # ── LISTE SURVEILLANCE
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px'>⭐ Liste de surveillance</div>", unsafe_allow_html=True)
    st.caption("CSV ou Excel — colonnes : Code article + Classe (GOLD, SILVER, A, B…)")
    watch_file  = st.file_uploader("Liste articles à surveiller", type=["csv", "xlsx", "xls"], key="watchlist")
    watch_label = st.text_input("Nom de la liste", value="20/80", placeholder="ex : 20/80, Stratégiques…")

    watchdict: dict = {}
    if watch_file is not None:
        watch_bytes = watch_file.read()
        watchdict   = load_watchlist(watch_bytes, watch_file.name)
        if watchdict:
            classes = sorted(set(watchdict.values()))
            st.success(f"✅ {len(watchdict)} codes · Classes : {', '.join(classes)}")
        else:
            st.warning("Aucun code détecté — vérifiez le format du fichier")

    watchlist = watchdict_codes(watchdict)   # set de codes pour compatibilité

    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Règles métier</div>", unsafe_allow_html=True)
    exclude_technical = st.checkbox("Exclure fournisseurs techniques", value=True)
    cap_sur_receipt   = st.checkbox("Caper Qté reçue ≤ Qté commandée", value=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<div class='page-title'>📦 On Time In Full</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Pilotage OTIF fournisseur · Magasin · Article · Fiche fournisseur</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ÉCRAN D'ACCUEIL — MÉTHODOLOGIE
# ══════════════════════════════════════════════════════════════════════════════
if uploaded is None:
    st.markdown("---")
    st.markdown("<div class='section-label'>Méthodologie de calcul</div>", unsafe_allow_html=True)
    docs = [
        ("① Fill Rate — Taux de service quantitatif",
         """<code>Fill Rate = Qté reçue retenue / Qté commandée × 100</code><br>
<strong>Règle cap :</strong> <code>Qté reçue retenue = min(Qté reçue, Qté commandée)</code><br>
<span style='color:#34C759;font-weight:600'>✓ Objectif cible : ≥ 97%</span>"""),
        ("② On Time — Taux de respect du délai",
         """<code>On Time = 1 si Date réception ≤ Date prévue, sinon 0</code><br>
Priorité colonnes : <code>H Date</code> → <code>Date livraison</code> → <code>Date prévue</code><br>
<span style='color:#34C759;font-weight:600'>✓ Objectif cible : ≥ 95%</span>"""),
        ("③ OTIF — On Time In Full",
         """<code>OTIF = 1 si Fill Rate ≥ 100% ET On Time = 1</code><br>
<span style='color:#34C759;font-weight:600'>✓ Objectif cible : ≥ 95%</span>"""),
        ("④ Score global", "<code>Score = 50% × Fill Rate + 30% × On Time + 20% × OTIF</code><br>🟢 ≥ 97% · 🟠 90–97% · 🔴 &lt; 90%"),
        ("⑤ Score de criticité", "<code>Criticité = Impact CA proxy × (1 − Fill Rate)</code>"),
        ("⑥ ⭐ Liste de surveillance — Classe par article",
         """Chargez un fichier CSV/Excel avec <strong>2 colonnes</strong> :<br>
<code>Code article</code> + <code>Classe</code> (ex: GOLD, SILVER, A, B…)<br><br>
La classe apparaît dans <strong>toutes les vues</strong> (tableaux, graphiques, exports Excel).<br>
Dans Excel : fond doré (GOLD) ou argent (SILVER) sur chaque ligne, colonne <em>Classe</em> en première position."""),
    ]
    for title, body in docs:
        st.markdown(f"<div class='doc-card'><div class='doc-card-title'>{title}</div><div class='doc-card-body'>{body}</div></div>", unsafe_allow_html=True)
    st.markdown("""<div class='alert-card alert-amber'><strong>⚠️ Règle temporaire ERP</strong> : date prévue absente → ligne considérée On Time par défaut.</div>""", unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# CHARGEMENT DONNÉES
# ══════════════════════════════════════════════════════════════════════════════
with st.spinner("Lecture et préparation des données…"):
    file_bytes = uploaded.read()
    raw = load_erp(file_bytes, uploaded.name)
    df, quality = prepare_dataset(
        raw,
        exclude_technical=exclude_technical,
        cap_sur_receipt=cap_sur_receipt,
        watchdict=watchdict,
    )

if df.empty:
    st.error("Aucune ligne exploitable après nettoyage. Vérifiez le format du fichier.")
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# FILTRES SIDEBAR (après chargement)
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtres globaux</div>", unsafe_allow_html=True)

    all_sites = sorted([x for x in df["site_label"].dropna().unique()    if x not in ("", "Inconnu")])
    all_sup   = sorted([x for x in df["supplier_name"].dropna().unique()  if x not in ("", "Inconnu")])
    all_depts = sorted([x for x in df["dept_label"].dropna().unique()     if x not in ("", "Inconnu")])

    sel_sites = st.multiselect("Magasin",     all_sites, default=all_sites)
    sel_sup   = st.multiselect("Fournisseur", all_sup,   default=[])
    sel_depts = st.multiselect("Département", all_depts, default=all_depts)
    only_crit  = st.checkbox("Uniquement OTIF = 0", value=False)
    only_watch = st.checkbox(f"Uniquement articles ⭐ {watch_label}", value=False, disabled=(not watchdict))

    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>📋 Fiche Fournisseur</div>", unsafe_allow_html=True)
    fiche_supplier = st.selectbox("Fournisseur à analyser", options=["— Choisir —"] + all_sup, key="fiche_sel")
    seuil_fill = st.slider("Seuil Fill Rate (lignes 'mauvaises')", min_value=50, max_value=100, value=100, step=5, format="%d%%")
    st.caption(f"Lignes retenues dans la fiche : Fill Rate ligne < {seuil_fill}%")

if not sel_sites: sel_sites = all_sites
if not sel_depts: sel_depts = all_depts

view = df[df["site_label"].isin(sel_sites) & df["dept_label"].isin(sel_depts)].copy()
if sel_sup:
    view = view[view["supplier_name"].isin(sel_sup)].copy()
if only_crit:
    view = view[view["otif"] == 0].copy()
if only_watch and watchdict:
    view = view[view["is_watched"]].copy()

if view.empty:
    st.warning("Aucune donnée après filtrage — ajustez les filtres dans la sidebar.")
    st.stop()

kpi         = compute_global_kpis(view)
by_supplier = agg_supplier(view)
by_site     = agg_site(view)
by_article  = agg_article(view, watchdict=watchdict)
watch_kpi   = compute_global_kpis(view[view["is_watched"]]) if watchdict else None


# ══════════════════════════════════════════════════════════════════════════════
# BANDEAU SURVEILLANCE
# ══════════════════════════════════════════════════════════════════════════════
if watchdict:
    n_watched_lines = int(view["is_watched"].sum())
    n_watched_arts  = view[view["is_watched"]]["Code"].nunique() if "Code" in view.columns else 0
    pct_matched     = round(safe_div(n_watched_arts, len(watchdict)) * 100, 1)
    classes         = sorted(set(watchdict.values()))
    st.markdown(f"""
<div class='watch-banner'>
  <strong>⭐ Liste de surveillance active — {watch_label}</strong> &nbsp;·&nbsp;
  {len(watchdict)} codes · Classes : <strong>{', '.join(classes)}</strong> &nbsp;·&nbsp;
  <strong>{n_watched_arts}</strong> articles matchés ({pct_matched}%) &nbsp;·&nbsp;
  <strong>{n_watched_lines:,}</strong> ligne(s) concernée(s)
</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# KPI GLOBAUX
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"<div class='section-label'>{kpi['sites']} magasin(s) · {kpi['suppliers']} fournisseur(s) · {kpi['orders']} commande(s)</div>", unsafe_allow_html=True)
render_kpi_row(kpi, watch_kpi=watch_kpi)


# ══════════════════════════════════════════════════════════════════════════════
# ALERTES
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("<div class='section-label'>Alertes &amp; points d'attention</div>", unsafe_allow_html=True)

if quality.get("all_dates_missing"):
    st.markdown("<div class='alert-card alert-red'><strong>🔴 On Time non significatif</strong> — aucune date prévue dans ce fichier. On Time = 100% artificiel.</div>", unsafe_allow_html=True)
elif quality.get("missing_expected_date", 0) > 0:
    pct = round(quality["missing_expected_date"] / quality["clean_rows"] * 100, 1)
    st.markdown(f"<div class='alert-card alert-amber'><strong>⚠️ Dates prévues partiellement manquantes</strong><br>{quality['missing_expected_date']:,} lignes sans date prévue ({pct}%) → considérées On Time par défaut.</div>", unsafe_allow_html=True)

if not by_supplier.empty:
    top3_sup = by_supplier.head(3)
    st.markdown(f"<div class='alert-card alert-red'><strong>🔴 Fournisseurs les plus critiques</strong><br>{' · '.join(f'<strong>{r.supplier_name}</strong> (score {r.score:.1f}%)' for r in top3_sup.itertuples())}</div>", unsafe_allow_html=True)

if not by_site.empty:
    top3_site = by_site.head(3)
    st.markdown(f"<div class='alert-card alert-amber'><strong>⚠️ Magasins les plus impactés</strong><br>{' · '.join(f'<strong>{r.site_label}</strong> ({int(r.qty_missing):,} unités manquantes)' for r in top3_site.itertuples())}</div>", unsafe_allow_html=True)

if watchdict and watch_kpi and watch_kpi.get("fill_rate", 100) < SEUIL_SURVEILLER:
    st.markdown(f"<div class='alert-card alert-gold'><strong>⭐ Articles surveillance ({watch_label}) — Fill Rate dégradé</strong><br>FR = <strong>{fmt_pct(watch_kpi['fill_rate'])}</strong> · Vol. manquant : <strong>{fmt(watch_kpi['missing_qty'])}</strong> unités · Impact : <strong>{fmt(watch_kpi['impact_value'])}</strong> FCFA</div>", unsafe_allow_html=True)

if quality.get("pv_zero_rows", 0) > 0:
    st.markdown(f"<div class='alert-card alert-amber'><strong>⚠️ Impact CA proxy sous-estimé</strong><br>{quality['pv_zero_rows']:,} ligne(s) avec PV HT = 0.</div>", unsafe_allow_html=True)

_watch_info = f" · ⭐ Lignes matchées : <strong>{quality.get('watched_in_data', 0):,}</strong>" if watchdict else ""
st.markdown(f"<div class='alert-card alert-blue'><strong>ℹ️ Qualité de données</strong> · Date prévue : <strong>{quality['expected_col']}</strong> · Dates manquantes : <strong>{quality['missing_expected_date']:,}</strong> · Sur-réceptions capées : <strong>{quality['sur_receipt_rows']:,}</strong> · Taux exploitable : <strong>{quality['usable_rate']:.1f}%</strong>{_watch_info}</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
n_watched_arts_tab = int((by_article["Classe"] != "").sum()) if "Classe" in by_article.columns else 0
watch_tab_label = f"⭐ {watch_label} ({n_watched_arts_tab})" if watchdict else "⭐ Surveillance"

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    f"🚛 Fournisseurs ({len(by_supplier)})",
    f"🏪 Magasins ({len(by_site)})",
    f"📦 Articles ({len(by_article)})",
    watch_tab_label,
    "🚨 Lignes critiques",
    "🧪 Qualité des données",
    "📋 Fiche Fournisseur",
])


# ── Tab 1 : Fournisseurs
with tab1:
    st.caption("Trié par criticité décroissante = Impact CA proxy × (1 − Fill Rate)")
    st.plotly_chart(bar_h(by_supplier, "criticality_score", "supplier_name", "#FF3B30", "Score de criticité (FCFA-équivalent)"), use_container_width=True)
    d = by_supplier.copy()
    d["Fill Rate"]       = d["fill_rate"].apply(fmt_pct)
    d["On Time"]         = d["on_time"].apply(fmt_pct)
    d["OTIF"]            = d["otif"].apply(fmt_pct)
    d["Score"]           = d["score"].apply(fmt_pct)
    d["Vol. manquant"]   = d["qty_missing"].apply(fmt)
    d["Impact CA proxy"] = d["impact_value"].apply(fmt)
    st.dataframe(d[["supplier_name","Fill Rate","On Time","OTIF","Score","Vol. manquant","Impact CA proxy","orders","articles","sites","Niveau"]].rename(columns={"supplier_name":"Fournisseur","orders":"Cmdes","articles":"Articles","sites":"Magasins"}), use_container_width=True, hide_index=True)


# ── Tab 2 : Magasins
with tab2:
    st.plotly_chart(bar_h(by_site, "qty_missing", "site_label", "#FF9500", "Volume manquant (unités)", fmt_fn=lambda v: f"{int(v):,}"), use_container_width=True)
    d = by_site.copy()
    for src, dst in [("fill_rate","Fill Rate"),("on_time","On Time"),("otif","OTIF"),("score","Score")]:
        d[dst] = d[src].apply(fmt_pct)
    d["Vol. manquant"]   = d["qty_missing"].apply(fmt)
    d["Impact CA proxy"] = d["impact_value"].apply(fmt)
    st.dataframe(d[["site_label","Fill Rate","On Time","OTIF","Score","Vol. manquant","Impact CA proxy","suppliers","articles","Niveau"]].rename(columns={"site_label":"Magasin","suppliers":"Fournisseurs","articles":"Articles"}), use_container_width=True, hide_index=True)


# ── Tab 3 : Articles
with tab3:
    st.markdown("<div class='section-label'>Top 20 articles critiques</div>", unsafe_allow_html=True)
    st.plotly_chart(
        bar_h(by_article.head(20), "qty_missing", "article_label", "#007AFF",
              "Volume manquant (unités)", height=620, fmt_fn=lambda v: f"{int(v):,}",
              classe_col="Classe"),
        use_container_width=True,
    )
    d = by_article.copy()
    for src, dst in [("fill_rate","Fill Rate"),("on_time","On Time"),("otif","OTIF"),("score","Score")]:
        d[dst] = d[src].apply(fmt_pct)
    d["Vol. manquant"]   = d["qty_missing"].apply(fmt)
    d["Impact CA proxy"] = d["impact_value"].apply(fmt)
    show_cols = [c for c in ["Classe","Code","article_label","supplier_name","Fill Rate","On Time","OTIF","Score","Vol. manquant","Impact CA proxy","sites","orders"] if c in d.columns]
    st.dataframe(d[show_cols].rename(columns={"article_label":"Article","supplier_name":"Fournisseur","sites":"Magasins","orders":"Cmdes"}), use_container_width=True, hide_index=True)


# ── Tab 4 : Articles Surveillés
with tab4:
    if not watchdict:
        st.markdown("<div class='alert-card alert-blue'><strong>ℹ️ Aucune liste de surveillance chargée</strong><br>Uploadez un CSV/Excel avec colonnes <code>Code article</code> + <code>Classe</code>.</div>", unsafe_allow_html=True)
    else:
        watched_arts = by_article[by_article["Classe"] != ""].copy() if "Classe" in by_article.columns else pd.DataFrame()
        if watched_arts.empty:
            st.warning(f"Aucun article de la liste {watch_label} trouvé dans les données filtrées.")
        else:
            w_lines = view[view["is_watched"]]
            wk = compute_global_kpis(w_lines)

            # KPIs par classe
            classes_present = sorted(watched_arts["Classe"].unique())
            if len(classes_present) > 1:
                st.markdown(f"<div class='section-label'>KPIs par classe ({', '.join(classes_present)})</div>", unsafe_allow_html=True)
                cls_cols = st.columns(len(classes_present))
                for i, cls in enumerate(classes_present):
                    cls_arts = watched_arts[watched_arts["Classe"] == cls]
                    cls_codes = set(cls_arts["code_str"].tolist()) if "code_str" in cls_arts.columns else set()
                    cls_lines = view[view["code_str"].isin(cls_codes)] if "code_str" in view.columns else view.iloc[0:0]
                    cls_kpi = compute_global_kpis(cls_lines)
                    with cls_cols[i]:
                        st.markdown(f"""
<div class='kpi-watch'>
  <div class='kpi-watch-label'>⭐ {cls} ({len(cls_arts)} réf.)</div>
  <div class='kpi-watch-value'>{fmt_pct(cls_kpi['fill_rate'])}</div>
  <div class='kpi-watch-sub'>FR · OTIF {fmt_pct(cls_kpi['otif'])} · Vol. manquant {fmt(cls_kpi['missing_qty'])}</div>
</div>""", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown(f"<div class='section-label'>KPIs globaux — Articles {watch_label} ({len(watched_arts)} références)</div>", unsafe_allow_html=True)
            wc1, wc2, wc3, wc4 = st.columns(4)
            wc1.metric("Fill Rate ⭐",    fmt_pct(wk["fill_rate"]))
            wc2.metric("On Time ⭐",      fmt_pct(wk["on_time"]))
            wc3.metric("OTIF ⭐",         fmt_pct(wk["otif"]))
            wc4.metric("Score global ⭐", fmt_pct(wk["score"]))
            wc5, wc6 = st.columns(2)
            wc5.metric("Vol. manquant ⭐",   fmt(wk["missing_qty"]))
            wc6.metric("Impact CA proxy ⭐", f"{fmt(wk['impact_value'])} FCFA")

            st.markdown("---")
            st.markdown(f"<div class='section-label'>Top 20 — Articles {watch_label}</div>", unsafe_allow_html=True)
            st.plotly_chart(
                bar_h(watched_arts.head(20), "qty_missing", "article_label", "#FFD60A",
                      "Volume manquant (unités)",
                      height=max(400, min(700, len(watched_arts.head(20)) * 32)),
                      fmt_fn=lambda v: f"{int(v):,}", classe_col="Classe"),
                use_container_width=True,
            )

            st.markdown("---")
            dw = watched_arts.copy()
            for src, dst in [("fill_rate","Fill Rate"),("on_time","On Time"),("otif","OTIF"),("score","Score")]:
                dw[dst] = dw[src].apply(fmt_pct)
            dw["Vol. manquant"]   = dw["qty_missing"].apply(fmt)
            dw["Impact CA proxy"] = dw["impact_value"].apply(fmt)
            show_w = [c for c in ["Classe","Code","article_label","supplier_name","Fill Rate","On Time","OTIF","Score","Vol. manquant","Impact CA proxy","sites","orders"] if c in dw.columns]
            st.dataframe(dw[show_w].rename(columns={"article_label":"Article","supplier_name":"Fournisseur","sites":"Magasins","orders":"Cmdes"}), use_container_width=True, hide_index=True)

            matched_codes = set(watched_arts["code_str"].tolist()) if "code_str" in watched_arts.columns else set()
            not_found = watchlist - matched_codes
            if not_found:
                with st.expander(f"🔍 {len(not_found)} code(s) de la liste non trouvés dans les données"):
                    st.dataframe(pd.DataFrame(sorted(not_found), columns=["Code non trouvé"]), use_container_width=True, hide_index=True)


# ── Tab 5 : Lignes critiques
with tab5:
    crit_lines = view[view["otif"] == 0].sort_values(["is_watched","qty_missing","service_gap_value"], ascending=[False,False,False])
    pct_non_otif = len(crit_lines) / len(view) * 100
    st.caption(f"{len(crit_lines):,} lignes non OTIF sur {len(view):,} ({pct_non_otif:.1f}%) — articles surveillés remontés en tête avec leur classe")
    dcols = [c for c in ["watch_classe","date_received","date_expected","site_label","supplier_name","Code","article_label","N° Cde","qte_cde","qte_rec_retained","qty_missing","service_gap_value","delay_days"] if c in crit_lines.columns]
    dl = crit_lines[dcols].copy()
    if "service_gap_value" in dl.columns:
        dl["service_gap_value"] = dl["service_gap_value"].apply(fmt)
    st.dataframe(
        dl.rename(columns={"watch_classe":"Classe","date_received":"Date réception","date_expected":"Date prévue","site_label":"Magasin","supplier_name":"Fournisseur","article_label":"Article","N° Cde":"N° Commande","qte_cde":"Qté cde","qte_rec_retained":"Qté reçue","qty_missing":"Qté manquante","service_gap_value":"Impact CA proxy","delay_days":"Retard (j)"}),
        use_container_width=True, hide_index=True,
    )


# ── Tab 6 : Qualité données
with tab6:
    q1, q2, q3, q4 = st.columns(4)
    q1.metric("Lignes brutes",       f"{quality['raw_rows']:,}")
    q2.metric("Lignes exploitables", f"{quality['clean_rows']:,}")
    q3.metric("Taux exploitable",    fmt_pct(quality['usable_rate']))
    q4.metric("Date prévue utilisée",quality["expected_col"])
    q5, q6, q7, q8 = st.columns(4)
    q5.metric("Qté cde ≤ 0 exclues",      f"{quality['excluded_zero_qty']:,}")
    q6.metric("Fournisseurs tech. exclus", f"{quality['excluded_technical']:,}")
    q7.metric("Dates prévues manquantes",  f"{quality['missing_expected_date']:,}")
    q8.metric("Sur-réceptions capées",     f"{quality['sur_receipt_rows']:,}")
    if watchdict:
        st.metric(f"⭐ Lignes articles {watch_label} matchées", f"{quality.get('watched_in_data', 0):,}")
    if quality.get("all_dates_missing"):
        st.error("🔴 Aucune date prévue — On Time = 100% artificiel.")
    else:
        st.warning("Règle temporaire : date prévue absente → ligne considérée On Time.")
    if quality.get("pv_zero_rows", 0) > 0:
        st.warning(f"⚠️ {quality['pv_zero_rows']:,} ligne(s) avec PV HT = 0.")


# ── Tab 7 : Fiche Fournisseur
with tab7:
    if fiche_supplier == "— Choisir —":
        st.markdown("<div class='alert-card alert-blue'><strong>ℹ️ Comment utiliser la Fiche Fournisseur ?</strong><br><br>1. Sélectionnez un fournisseur dans la sidebar<br>2. Ajustez le seuil Fill Rate<br>3. Téléchargez la fiche Excel (colonne Classe incluse si liste chargée)</div>", unsafe_allow_html=True)
    else:
        fiche_df  = view[view["supplier_name"] == fiche_supplier].copy()
        fiche_bad = fiche_df[fiche_df["line_fill_rate"] < seuil_fill / 100].sort_values(["is_watched","qty_missing","service_gap_value"], ascending=[False,False,False])
        fkpi      = compute_global_kpis(fiche_df)
        col_txt   = score_color(fkpi["score"])

        fou_code = "—"
        if "Fou" in fiche_df.columns and not fiche_df["Fou"].isna().all():
            fou_code = str(int(fiche_df["Fou"].dropna().iloc[0]))

        n_fiche_watched = int(fiche_df["is_watched"].sum()) if "is_watched" in fiche_df.columns else 0
        watch_badge = f"&nbsp;·&nbsp; <span class='badge-watch'>⭐ {n_fiche_watched} lignes {watch_label}</span>" if watchdict and n_fiche_watched > 0 else ""

        st.markdown(f"""
<div class='fiche-header'>
  <div style='display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px'>
    <div>
      <div style='font-size:20px;font-weight:700;color:#1C1C1E'>{fiche_supplier}</div>
      <div style='font-size:12px;color:#8E8E93;margin-top:4px'>
        Code : <strong>{fou_code}</strong> · {fkpi['orders']} commande(s) · {fkpi['articles']} article(s) · {fkpi['sites']} magasin(s){watch_badge}
      </div>
    </div>
    <div style='background:{col_txt}1A;border:1.5px solid {col_txt};border-radius:8px;padding:8px 16px;font-size:13px;font-weight:700;color:{col_txt}'>
      {score_band(fkpi["score"])} · Score {fkpi["score"]:.1f}%
    </div>
  </div>
</div>""", unsafe_allow_html=True)

        render_kpi_row(fkpi)
        st.markdown("---")

        st.markdown("<div class='section-label'>Livraisons incomplètes par magasin</div>", unsafe_allow_html=True)
        site_recap = fiche_bad.groupby("site_label", as_index=False).agg(nb_cdes=("N° Cde","nunique"), nb_articles=("Code","nunique"), qte_cde=("qte_cde","sum"), qte_rec=("qte_rec_retained","sum"), qty_missing=("qty_missing","sum"), impact_value=("service_gap_value","sum")).sort_values("qty_missing", ascending=False)
        site_recap["Fill Rate"]       = (site_recap["qte_rec"] / site_recap["qte_cde"] * 100).apply(fmt_pct)
        site_recap["Vol. manquant"]   = site_recap["qty_missing"].apply(fmt)
        site_recap["Impact CA proxy"] = site_recap["impact_value"].apply(fmt)
        st.dataframe(site_recap[["site_label","Fill Rate","Vol. manquant","Impact CA proxy","nb_cdes","nb_articles"]].rename(columns={"site_label":"Magasin","nb_cdes":"Cmdes","nb_articles":"Articles"}), use_container_width=True, hide_index=True)

        st.markdown("<div class='section-label' style='margin-top:18px'>Livraisons incomplètes par article</div>", unsafe_allow_html=True)
        art_recap = fiche_bad.groupby(["Code","article_label"], as_index=False).agg(nb_sites=("site_label","nunique"), nb_cdes=("N° Cde","nunique"), qte_cde=("qte_cde","sum"), qte_rec=("qte_rec_retained","sum"), qty_missing=("qty_missing","sum"), impact_value=("service_gap_value","sum")).sort_values("qty_missing", ascending=False)
        art_recap["Fill Rate"]       = (art_recap["qte_rec"] / art_recap["qte_cde"] * 100).apply(fmt_pct)
        art_recap["Vol. manquant"]   = art_recap["qty_missing"].apply(fmt)
        art_recap["Impact CA proxy"] = art_recap["impact_value"].apply(fmt)
        if watchdict:
            art_recap["Classe"] = art_recap["Code"].apply(lambda c: get_classe(c, watchdict))
            cols_art = ["Classe","Code","article_label","Fill Rate","Vol. manquant","Impact CA proxy","nb_sites","nb_cdes"]
        else:
            cols_art = ["Code","article_label","Fill Rate","Vol. manquant","Impact CA proxy","nb_sites","nb_cdes"]
        st.dataframe(art_recap[cols_art].rename(columns={"article_label":"Article","nb_sites":"Magasins","nb_cdes":"Cmdes"}), use_container_width=True, hide_index=True)

        st.markdown("<div class='section-label' style='margin-top:18px'>Détail des livraisons incomplètes</div>", unsafe_allow_html=True)
        st.caption(f"{len(fiche_bad):,} ligne(s) avec Fill Rate < {seuil_fill}% sur {len(fiche_df):,} total" + (f" · articles {watch_label} remontés en tête" if watchdict else ""))
        dcols_f = [c for c in ["watch_classe","N° Cde","date_received","date_expected","site_label","Code","article_label","qte_cde","qte_rec_retained","qty_missing","line_fill_rate","service_gap_value","on_time","delay_days"] if c in fiche_bad.columns]
        df_disp = fiche_bad[dcols_f].copy()
        if "line_fill_rate" in df_disp.columns:
            df_disp["line_fill_rate"] = df_disp["line_fill_rate"].apply(lambda v: fmt_pct(v * 100))
        if "service_gap_value" in df_disp.columns:
            df_disp["service_gap_value"] = df_disp["service_gap_value"].apply(fmt)
        if "on_time" in df_disp.columns:
            df_disp["on_time"] = df_disp["on_time"].map({True:"✅",False:"❌",1:"✅",0:"❌"}).fillna("—")
        st.dataframe(df_disp.rename(columns={"watch_classe":"Classe","N° Cde":"N° Commande","date_received":"Date réception","date_expected":"Date prévue","site_label":"Magasin","article_label":"Article","qte_cde":"Qté cde","qte_rec_retained":"Qté reçue","qty_missing":"Qté manquante","line_fill_rate":"Fill Rate ligne","service_gap_value":"Impact CA proxy","on_time":"À l'heure","delay_days":"Retard (j)"}), use_container_width=True, hide_index=True)

        st.markdown("---")
        if st.button("📥 Générer la fiche Excel fournisseur", type="primary", key="btn_fiche"):
            with st.spinner("Génération de la fiche…"):
                buf_fiche = build_fiche_excel(fournisseur=fiche_supplier, df_all=fiche_df, df_bad=fiche_bad, site_recap=site_recap, art_recap=art_recap, kpis=fkpi, seuil=seuil_fill, watchdict=watchdict)
            safe_name = fiche_supplier.strip().replace(" ","_").replace("/","-")[:40]
            st.download_button(label=f"⬇️ Télécharger la fiche — {fiche_supplier}", data=buf_fiche, file_name=f"SmartBuyer_OTIF_{safe_name}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_fiche")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTS EXCEL GLOBAUX
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("<div class='section-label'>Exports Excel</div>", unsafe_allow_html=True)
col_exp1, col_exp2, col_exp3 = st.columns(3)

with col_exp1:
    _recap_mention = "<br>⭐ Colonnes GOLD/SILVER incluses" if watchdict else ""
    st.markdown(f"""
<div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px 18px;margin-bottom:8px'>
  <div style='font-size:13px;font-weight:700;color:#1C1C1E;margin-bottom:4px'>🎯 Récap priorisation fournisseurs</div>
  <div style='font-size:12px;color:#8E8E93;line-height:1.5'>
    1 ligne par fournisseur · Réfs commandées / livrées / taux couverture<br>
    Fill Rate · OTIF · Sites impactés · Impact CA proxy<br>
    Trié par criticité · Filtre automatique · Légende incluse{_recap_mention}
  </div>
</div>""", unsafe_allow_html=True)
    if st.button("Générer le Récap fournisseurs", type="primary", key="btn_recap"):
        with st.spinner("Génération…"):
            buf_recap = build_recap_fournisseur_excel(view, watchdict=watchdict)
        st.download_button(
            "⬇️ Télécharger — Récap fournisseurs",
            data=buf_recap,
            file_name="SmartBuyer_OTIF_Recap_Fournisseurs.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_recap",
        )

with col_exp2:
    watch_mention = f"<br>⭐ Onglet dédié {watch_label} · colonne Classe incluse" if watchdict else ""
    st.markdown(f"<div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px 18px;margin-bottom:8px'><div style='font-size:13px;font-weight:700;color:#1C1C1E;margin-bottom:4px'>📋 Toutes les fiches fournisseurs</div><div style='font-size:12px;color:#8E8E93;line-height:1.5'>1 onglet · toutes livraisons incomplètes · tous fournisseurs<br>Filtre automatique · trié par criticité{watch_mention}</div></div>", unsafe_allow_html=True)
    if st.button("Générer l'export Fiches", type="primary", key="btn_all_fiches"):
        with st.spinner("Génération…"):
            buf_all = build_export_all_fiches(view, by_supplier, seuil_fill, watchdict=watchdict)
        st.download_button("⬇️ Télécharger — Toutes les fiches", data=buf_all, file_name="SmartBuyer_OTIF_Fiches_Fournisseurs.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_all_fiches")

with col_exp3:
    st.markdown(f"<div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px 18px;margin-bottom:8px'><div style='font-size:13px;font-weight:700;color:#1C1C1E;margin-bottom:4px'>📊 Synthèse analytique globale</div><div style='font-size:12px;color:#8E8E93;line-height:1.5'>Multi-onglets · Par fournisseur · Par magasin · Par article<br>Lignes critiques · Qualité des données{'<br>⭐ Onglet Articles surveillés + colonne Classe' if watchdict else ''}</div></div>", unsafe_allow_html=True)
    if st.button("Générer l'export Synthèse", type="primary", key="btn_global"):
        with st.spinner("Génération…"):
            buf_global = build_export_excel(view, by_supplier, by_site, by_article, quality, watchdict=watchdict)
        st.download_button("⬇️ Télécharger — Synthèse globale", data=buf_global, file_name="SmartBuyer_OTIF_Synthese.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_global")
