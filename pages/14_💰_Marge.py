"""
14_💰_Marge.py — Module Marge · SmartBuyer Hub
Analyse de rentabilité multi-rayon sur export PowerBI.
Double vue : Executive (DG) / Opérationnelle (acheteur).
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG & CHARTE
# ============================================================
st.set_page_config(page_title="Marge", page_icon="💰", layout="wide")

BLUE = "#007AFF"
GREEN = "#34C759"
RED = "#FF3B30"
AMBER = "#FF9500"
DARK = "#1C1C1E"
GREY = "#8E8E93"
BG = "#F2F2F7"

# ============================================================
# 🎯 CIBLES DE MARGE PAR RAYON  —  À METTRE À JOUR ICI
# ------------------------------------------------------------
# Les directives marge changent régulièrement. Quand la direction
# communique de nouveaux objectifs, modifier UNIQUEMENT ce bloc.
# Clé = libellé exact du rayon (tel qu'il apparaît dans l'export PBI).
# Dernière mise à jour : juin 2026
# ============================================================
CIBLES_DEFAUT = {
    "BOISSONS": 19.5,
    "DROGUERIE": 25.0,
    "PARFUMERIE HYGIENE": 29.0,
    "EPICERIE": 16.0,
}
CIBLE_FALLBACK = 23.5  # rayon non listé ci-dessus

st.markdown(f"""
<style>
.stApp {{ background-color: {BG}; }}
.block-container {{ padding-top: 2rem; max-width: 1200px; }}
.card {{ background:#fff; border-radius:14px; padding:18px 20px; margin-bottom:12px;
        box-shadow:0 1px 3px rgba(0,0,0,0.06); }}
.kpi-card {{ background:#fff; border-radius:12px; padding:14px 16px;
            box-shadow:0 1px 3px rgba(0,0,0,0.06); }}
.kpi-label {{ font-size:12px; color:{GREY}; margin-bottom:4px; }}
.kpi-value {{ font-size:24px; font-weight:600; color:{DARK}; }}
.kpi-sub {{ font-size:12px; color:{GREY}; margin-top:2px; }}
.headline {{ font-size:26px; font-weight:600; color:{DARK}; }}
.so-what {{ background:{DARK}; color:#fff; border-radius:14px; padding:18px 22px;
           font-size:15px; line-height:1.7; }}
.badge {{ display:inline-block; padding:2px 10px; border-radius:6px;
         font-size:11px; font-weight:600; }}
h1, h2, h3 {{ color:{DARK}; }}
</style>
""", unsafe_allow_html=True)

# ============================================================
# MOTEUR DE CALCUL
# ============================================================

EXPECTED_COLS = ['Rayon', 'Famille', 'Article', 'Site nom long', 'CA', 'CA N-1',
                 'Marge', 'CA Promo', 'Marge Promo', 'CA Hors Promo',
                 'Marge Hors Promo', 'Casse (Valeur)']

@st.cache_data(show_spinner=False)
def load_pbi(file_bytes):
    """Charge l'export PBI et prépare les lignes article par site."""
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [str(c).lstrip('\ufeff') for c in df.columns]
    art = df[df['Site nom long'].notna()
             & (df['Site nom long'] != 'Total')
             & df['Article'].notna()].copy()
    art['Site'] = art['Site nom long'].astype(str).str.split(' - ').str[-1].str.strip()
    art['RayonLib'] = art['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    art['FamLib'] = art['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    art['ArtCode'] = art['Article'].astype(str).str.split(' - ').str[0].str.strip()
    art['ArtLib'] = art['Article'].astype(str).str.split(' - ').str[-1].str.strip()
    return art

def safe_rate(num, den):
    """Taux sécurisé (zéro si dénominateur nul)."""
    den = np.where(den == 0, np.nan, den)
    return np.where(np.isnan(den), 0.0, num / den * 100)

def agg_site(df):
    """Agrège par site. Exclut les sites à CA nul (magasins fermés/sans activité)."""
    g = df.groupby('Site').agg(
        CA=('CA', 'sum'), CA_N1=('CA N-1', 'sum'), Marge=('Marge', 'sum'),
        CA_Promo=('CA Promo', 'sum'), Marge_Promo=('Marge Promo', 'sum'),
        CA_HP=('CA Hors Promo', 'sum'), Marge_HP=('Marge Hors Promo', 'sum'),
        Casse=('Casse (Valeur)', 'sum'), Nb=('Article', 'count')
    ).reset_index()
    g = g[g['CA'] > 0].copy()  # écarter les sites sans activité sur la période
    g['TxMarge'] = safe_rate(g['Marge'], g['CA'])
    g['EvoCA'] = safe_rate(g['CA'] - g['CA_N1'], g['CA_N1'])
    g['PdsPromo'] = safe_rate(g['CA_Promo'], g['CA'])
    g['TxMargePromo'] = safe_rate(g['Marge_Promo'], g['CA_Promo'])
    g['TxMargeHP'] = safe_rate(g['Marge_HP'], g['CA_HP'])
    return g.sort_values('CA', ascending=False)

def alert_level(tx, cible, crit=5.0, moyen=2.0):
    """Niveau d'alerte selon écart à la cible."""
    ecart = cible - tx
    if ecart > crit: return 'critique'
    if ecart > moyen: return 'moyen'
    return 'bon'

def bennet(df, site_ref, exclure_ref=True):
    """Décomposition Bennet symétrique : site_ref vs reste du réseau."""
    site_df = df[df['Site'] == site_ref]
    if exclure_ref:
        res_df = df[df['Site'] != site_ref]
    else:
        res_df = df
    fc = site_df.groupby('FamLib').agg(CA_c=('CA', 'sum'), M_c=('Marge', 'sum'))
    fr = res_df.groupby('FamLib').agg(CA_r=('CA', 'sum'), M_r=('Marge', 'sum'))
    b = fc.join(fr, how='outer').fillna(0)
    CA_c, CA_r = b['CA_c'].sum(), b['CA_r'].sum()
    if CA_c == 0 or CA_r == 0:
        return None, 0, 0, 0, 0
    b['tx_c'] = safe_rate(b['M_c'].values, b['CA_c'].values)
    b['tx_r'] = safe_rate(b['M_r'].values, b['CA_r'].values)
    b['w_c'] = b['CA_c'] / CA_c
    b['w_r'] = b['CA_r'] / CA_r
    b['w_moy'] = (b['w_c'] + b['w_r']) / 2
    b['t_moy'] = (b['tx_c'] + b['tx_r']) / 2
    b['effet_mix'] = (b['w_c'] - b['w_r']) * b['t_moy']
    b['effet_taux'] = b['w_moy'] * (b['tx_c'] - b['tx_r'])
    tx_c_tot = b['M_c'].sum() / CA_c * 100
    tx_r_tot = b['M_r'].sum() / CA_r * 100
    return (b.sort_values('CA_c', ascending=False),
            tx_c_tot, tx_r_tot,
            b['effet_mix'].sum(), b['effet_taux'].sum())

def ecart_inter_sites(df, min_sites=3, min_ca=50000, tx_lo=-50, tx_hi=60):
    """Articles à fort écart de taux de marge entre sites."""
    v = df.copy()
    v['TxMarge'] = safe_rate(v['CA'].values * 0 + v['Marge'].values, v['CA'].values)
    v['TxMarge'] = v['Marge'] / v['CA'].replace(0, np.nan) * 100
    v = v[(v['CA'] > min_ca) & v['TxMarge'].notna() & np.isfinite(v['TxMarge'])]
    v = v[(v['TxMarge'] > tx_lo) & (v['TxMarge'] < tx_hi)]
    rows = []
    for art, sub in v.groupby('ArtLib'):
        if sub['Site'].nunique() < min_sites:
            continue
        imin, imax = sub['TxMarge'].idxmin(), sub['TxMarge'].idxmax()
        ecart = sub.loc[imax, 'TxMarge'] - sub.loc[imin, 'TxMarge']
        ca_pire = sub.loc[imin, 'CA']
        rows.append({
            'Article': art, 'Famille': sub['FamLib'].iloc[0],
            'Nb sites': sub['Site'].nunique(),
            'Tx min %': round(sub.loc[imin, 'TxMarge'], 1),
            'Tx max %': round(sub.loc[imax, 'TxMarge'], 1),
            'Écart (pts)': round(ecart, 1),
            'Site faible': sub.loc[imin, 'Site'],
            'Site fort': sub.loc[imax, 'Site'],
            'CA total': round(sub['CA'].sum()),
            'Perte estimée': round(ca_pire * ecart / 100),
        })
    res = pd.DataFrame(rows)
    if len(res):
        res = res.sort_values('Écart (pts)', ascending=False).reset_index(drop=True)
    return res

def detect_destructeurs(df, sites_agg, cible_par_rayon):
    """Moteur de détection des destructeurs de rentabilité (4 règles).
    Impact = perte de marge en FCFA (écart de taux × CA local), trié par enjeu réel."""
    out = []
    fam_res = df.groupby('FamLib').apply(
        lambda x: x['Marge'].sum() / x['CA'].sum() * 100 if x['CA'].sum() else 0
    )
    # Règle 1+2 : Famille × Site déviant (effet taux + promo)
    for (site, fam), sub in df.groupby(['Site', 'FamLib']):
        ca = sub['CA'].sum()
        if ca < 100000:
            continue
        tx = sub['Marge'].sum() / ca * 100 if ca else 0
        tx_ref = fam_res.get(fam, 0)
        ecart = tx - tx_ref
        ca_promo = sub['CA Promo'].sum()
        m_promo = sub['Marge Promo'].sum()
        pds_promo = ca_promo / ca * 100 if ca else 0
        tx_promo = m_promo / ca_promo * 100 if ca_promo else 0
        if ecart < -3:
            perte = ecart * ca / 100  # marge en FCFA perdue vs référence réseau
            tags = ['Effet taux']
            if pds_promo > 20 and tx_promo < 10:
                tags.append('Promo excessive')
            if ecart < -5:
                tags.append('Prix achat à vérifier')
            out.append({
                'nom': f"{fam} · {site}", 'type': 'Famille × Site',
                'perte': round(perte), 'tx_site': round(tx, 1),
                'tx_ref': round(tx_ref, 1), 'ecart': round(ecart, 1),
                'pds_promo': round(pds_promo, 1), 'ca': round(ca),
                'tags': tags,
            })
    # Règle 4 : articles à marge négative significative
    neg = df[(df['Marge'] < 0) & (df['CA'] > 30000)]
    for _, row in neg.nlargest(5, 'CA').iterrows():
        out.append({
            'nom': f"{row['ArtLib']} · {row['Site']}", 'type': 'Article négatif',
            'perte': round(row['Marge']),
            'tx_site': round(row['Marge'] / row['CA'] * 100, 1) if row['CA'] else 0,
            'tx_ref': None, 'ecart': None, 'pds_promo': None,
            'ca': round(row['CA']), 'tags': ['Marge négative'],
        })
    out = sorted(out, key=lambda x: x['perte'])  # perte la plus négative en tête
    return out[:15]

# ============================================================
# EXPORT EXCEL (règles reviewer : Lisez-moi → Synthèse board → détail)
# ============================================================

def build_excel(ecarts_df, rayon_label):
    wb = Workbook()
    thin = Side(style='thin', color='D9D9D9')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    DARKX = '1C1C1E'; REDX = 'C94040'; GREYX = '8E8E93'

    # Feuille 1 : Lisez-moi
    ws = wb.active; ws.title = 'Lisez-moi'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 72
    r = 2
    ws.cell(r, 2, f'ANALYSE DES ÉCARTS DE MARGE INTER-SITES — {rayon_label}').font = Font(bold=True, size=15, color=DARKX)
    r += 2
    ws.cell(r, 2, 'À quoi sert ce fichier ?').font = Font(bold=True, size=12, color='1C6FB8')
    r += 1
    ws.cell(r, 3, "Il identifie les articles vendus dans plusieurs magasins mais avec des taux de marge très différents d'un site à l'autre. Un même produit rentable dans un magasin et déficitaire dans un autre signale un problème de prix d'achat ou de prix de vente localisé.").alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 50
    r += 2
    defs = [
        ('Écart (pts)', "Différence entre le taux de marge le plus élevé et le plus bas du même article entre sites."),
        ('Site faible / fort', "Magasin où l'article a la marge la plus basse / la plus haute."),
        ('Perte estimée', "Marge récupérable si le site faible atteignait le taux du site fort. À prioriser."),
    ]
    ws.cell(r, 2, 'Comment lire').font = Font(bold=True, size=12, color='1C6FB8'); r += 1
    for lab, d in defs:
        ws.cell(r, 2, lab).font = Font(bold=True, size=10, color=DARKX)
        ws.cell(r, 2).alignment = Alignment(vertical='top')
        ws.cell(r, 3, d).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 36; r += 1
    r += 1
    ws.cell(r, 2, 'Filtrage').font = Font(bold=True, size=12, color='1C6FB8'); r += 1
    ws.cell(r, 3, "Articles présents sur ≥3 sites, CA > 50 000 FCFA par site, taux borné entre -50% et +60% (au-delà : anomalies ou liquidations, pas des problèmes de pilotage).").alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 44

    # Feuille 2 : Synthèse board
    wsb = wb.create_sheet('Synthèse board')
    wsb.sheet_view.showGridLines = False
    wsb.column_dimensions['A'].width = 3
    for col, w in zip('BCDEF', [34, 13, 18, 18, 16]):
        wsb.column_dimensions[col].width = w
    rb = 2
    wsb.cell(rb, 2, f'SYNTHÈSE — ÉCARTS DE MARGE INTER-SITES · {rayon_label}').font = Font(bold=True, size=14, color=DARKX)
    rb += 2
    total_perte = ecarts_df.sort_values('Écart (pts)', ascending=False).head(50)['Perte estimée'].sum() if len(ecarts_df) else 0
    nb_crit = len(ecarts_df[ecarts_df['Écart (pts)'] > 40]) if len(ecarts_df) else 0
    wsb.cell(rb, 2, 'Articles analysés').font = Font(size=9, color=GREYX)
    wsb.cell(rb, 4, 'Écarts critiques').font = Font(size=9, color=GREYX)
    wsb.cell(rb, 5, 'Marge récupérable*').font = Font(size=9, color=GREYX)
    rb += 1
    wsb.cell(rb, 2, len(ecarts_df)).font = Font(bold=True, size=18, color=DARKX)
    wsb.cell(rb, 4, nb_crit).font = Font(bold=True, size=18, color=REDX)
    wsb.cell(rb, 5, f'{total_perte/1e6:.1f} M').font = Font(bold=True, size=18, color=DARKX)
    rb += 2
    heads = ['Article', 'Écart (pts)', 'Site faible', 'Site fort', 'Récupérable']
    for c, h in enumerate(heads, 2):
        cell = wsb.cell(rb, c, h); cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=DARKX)
        cell.alignment = Alignment(horizontal='center'); cell.border = bd
    rb += 1
    top15 = ecarts_df.sort_values('Perte estimée', ascending=False).head(15)
    for _, row in top15.iterrows():
        vals = [row['Article'], row['Écart (pts)'], row['Site faible'], row['Site fort'], row['Perte estimée']]
        for c, v in enumerate(vals, 2):
            cell = wsb.cell(rb, c, v); cell.border = bd; cell.font = Font(size=10)
            if c == 6:
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
                cell.font = Font(size=10, bold=True, color=REDX)
        rb += 1
    rb += 1
    wsb.cell(rb, 2, '* Si le site faible atteignait le taux du site fort. Détail complet en feuille suivante.').font = Font(size=9, italic=True, color=GREYX)

    # Feuille 3 : détail (couleur uniquement sur Écart)
    ws3 = wb.create_sheet('Écarts inter-sites')
    ws3.sheet_view.showGridLines = False
    cols = list(ecarts_df.columns)
    for c, h in enumerate(cols, 1):
        cell = ws3.cell(1, c, h); cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=DARKX)
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = bd
    ws3.row_dimensions[1].height = 28
    ecart_col = cols.index('Écart (pts)') + 1
    for i, (_, row) in enumerate(ecarts_df.iterrows(), 2):
        for c, h in enumerate(cols, 1):
            cell = ws3.cell(i, c, row[h]); cell.border = bd; cell.font = Font(size=10)
            if h in ('CA total', 'Perte estimée'):
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
        ec = row['Écart (pts)']
        ecell = ws3.cell(i, ecart_col)
        if ec > 40:
            ecell.fill = PatternFill('solid', fgColor='FCEBEB')
            ecell.font = Font(size=10, bold=True, color='A32D2D')
    widths = [34, 18, 9, 9, 9, 11, 18, 18, 13, 14]
    for c, w in enumerate(widths[:len(cols)], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(ecarts_df)+1}'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ============================================================
# INTERFACE
# ============================================================

st.title("💰 Module Marge")
st.caption("Analyse de rentabilité multi-rayon · export PowerBI")

up = st.file_uploader("Déposez votre export PowerBI (.xlsx)", type=['xlsx'])
if up is None:
    st.info("⬆️ Importez un export PBI pour démarrer l'analyse.")
    st.stop()

art = load_pbi(up.getvalue())
rayons = sorted(art['RayonLib'].dropna().unique().tolist())

# --- Barre de paramètres ---
with st.container():
    c1, c2, c3 = st.columns([2, 2, 1.2])
    with c1:
        sel_rayons = st.multiselect("Rayon(s)", rayons, default=rayons)
    with c2:
        vue = st.radio("Vue", ["Executive", "Opérationnelle"], horizontal=True)
    with c3:
        st.caption("Cible de marge par rayon")

df = art[art['RayonLib'].isin(sel_rayons)].copy()
if df.empty:
    st.warning("Aucune donnée pour ce filtre.")
    st.stop()

# --- Cibles différenciées par rayon ---
def cible_defaut(rayon):
    """Récupère la cible directive du rayon (matching tolérant casse/accents)."""
    key = str(rayon).strip().upper()
    if key in CIBLES_DEFAUT:
        return CIBLES_DEFAUT[key]
    # matching partiel (ex. 'PARFUMERIE HYGIENE' vs 'PARFUMERIE-HYGIENE')
    for k, v in CIBLES_DEFAUT.items():
        if k.replace(' ', '').replace('-', '') == key.replace(' ', '').replace('-', ''):
            return v
    return CIBLE_FALLBACK

with st.expander("⚙️ Cibles de marge par rayon (directives)", expanded=False):
    st.caption("Valeurs par défaut = dernières directives connues. Modifiables ici pour une simulation ponctuelle ; "
               "pour un changement durable, mettre à jour le bloc CIBLES_DEFAUT en haut du fichier.")
    cible_par_rayon = {}
    cols = st.columns(min(len(sel_rayons), 4))
    for i, ray in enumerate(sel_rayons):
        with cols[i % len(cols)]:
            cible_par_rayon[ray] = st.number_input(
                f"{ray}", min_value=0.0, max_value=60.0,
                value=float(cible_defaut(ray)), step=0.5, key=f"cible_{ray}")

# Cible pondérée (pour KPI globaux multi-rayon)
ca_par_rayon = df.groupby('RayonLib')['CA'].sum()
cible_glob = sum(cible_par_rayon.get(r, CIBLE_FALLBACK) * ca_par_rayon.get(r, 0) for r in sel_rayons) / ca_par_rayon.sum()

sites = agg_site(df)
# Restreindre l'analyse aux sites actifs (CA > 0) — cohérence Bennet / destructeurs / écarts
sites_actifs = sites['Site'].tolist()
df = df[df['Site'].isin(sites_actifs)].copy()
CA_tot = df['CA'].sum()
M_tot = df['Marge'].sum()
TX_tot = M_tot / CA_tot * 100 if CA_tot else 0
casse_tot = df['Casse (Valeur)'].sum()
ca_promo_tot = df['CA Promo'].sum()
m_promo_tot = df['Marge Promo'].sum()
pds_promo_tot = ca_promo_tot / CA_tot * 100 if CA_tot else 0
tx_promo_tot = m_promo_tot / ca_promo_tot * 100 if ca_promo_tot else 0

# Niveau d'alerte par site (cible = cible du rayon dominant du site)
def site_cible(site_name):
    sub = df[df['Site'] == site_name]
    dom = sub.groupby('RayonLib')['CA'].sum().idxmax()
    return cible_par_rayon.get(dom, cible_glob)

sites['cible'] = sites['Site'].apply(site_cible)
sites['niveau'] = sites.apply(lambda r: alert_level(r['TxMarge'], r['cible']), axis=1)
nb_alerte = (sites['niveau'] != 'bon').sum()
nb_crit = (sites['niveau'] == 'critique').sum()

# ============================================================
# VUE EXECUTIVE
# ============================================================

def fmt_m(x):
    """Formate en M FCFA."""
    return f"{x/1e6:,.0f} M".replace(",", " ")

def color_for(ecart):
    if ecart > 5: return RED
    if ecart > 2: return AMBER
    return GREEN

if vue == "Executive":
    ecart_glob = cible_glob - TX_tot
    col_head = RED if ecart_glob > 2 else GREEN
    st.markdown(
        f"<div class='headline'>Marge à <span style='color:{col_head}'>{TX_tot:.1f}%</span> "
        f"— objectif {cible_glob:.1f}%</div>"
        f"<div style='color:{GREY};font-size:14px;margin-top:4px'>"
        f"Écart de {ecart_glob:+.1f} pts · {nb_alerte} site(s) sous objectif dont {nb_crit} critique(s)</div>",
        unsafe_allow_html=True)
    st.write("")

    # Score + KPI
    cscore, ckpi = st.columns([1, 2])
    with cscore:
        st.markdown(
            f"<div class='card' style='text-align:center;padding:28px 18px'>"
            f"<div style='font-size:52px;font-weight:600;color:{RED if nb_alerte else GREEN};line-height:1'>{nb_alerte}</div>"
            f"<div style='font-size:16px;color:{GREY}'>/ {len(sites)} sites</div>"
            f"<div style='font-size:13px;color:{GREY};margin-top:8px'>sous objectif de marge</div>"
            f"</div>", unsafe_allow_html=True)
    with ckpi:
        k1, k2, k3 = st.columns(3)
        for col, lab, val, sub, c in [
            (k1, "CA réseau", fmt_m(CA_tot), f"{safe_rate(np.array([CA_tot - sites['CA_N1'].sum()]), np.array([sites['CA_N1'].sum()]))[0]:+.1f}% vs N-1", DARK),
            (k2, "Marge réseau", fmt_m(M_tot), f"Tx {TX_tot:.1f}%", DARK),
            (k3, "Casse réseau", fmt_m(casse_tot), f"{casse_tot/CA_tot*100:.2f}% du CA", RED)]:
            with col:
                st.markdown(
                    f"<div class='kpi-card'><div class='kpi-label'>{lab}</div>"
                    f"<div class='kpi-value' style='color:{c}'>{val}</div>"
                    f"<div class='kpi-sub'>{sub}</div></div>", unsafe_allow_html=True)

    st.write("")
    # 3 signaux
    st.markdown("##### Signaux à arbitrer")
    worst = sites[sites['niveau'] == 'critique'].nsmallest(1, 'TxMarge')
    sig = st.columns(3)
    if len(worst):
        w = worst.iloc[0]
        with sig[0]:
            st.markdown(f"<div class='card'><div style='font-size:12px;color:{GREY}'>Site critique</div>"
                        f"<div style='font-size:15px;font-weight:600'>{w['Site']}</div>"
                        f"<div style='font-size:20px;font-weight:600;color:{RED}'>{w['TxMarge']:.1f}%</div></div>",
                        unsafe_allow_html=True)
    with sig[1]:
        worst_evo = sites.nsmallest(1, 'EvoCA').iloc[0]
        st.markdown(f"<div class='card'><div style='font-size:12px;color:{GREY}'>Plus forte baisse CA</div>"
                    f"<div style='font-size:15px;font-weight:600'>{worst_evo['Site']}</div>"
                    f"<div style='font-size:20px;font-weight:600;color:{RED}'>{worst_evo['EvoCA']:.1f}%</div></div>",
                    unsafe_allow_html=True)
    with sig[2]:
        st.markdown(f"<div class='card'><div style='font-size:12px;color:{GREY}'>Marge promo réseau</div>"
                    f"<div style='font-size:15px;font-weight:600'>vs {tx_promo_tot:.1f}% hors promo</div>"
                    f"<div style='font-size:20px;font-weight:600;color:{AMBER}'>{tx_promo_tot:.1f}%</div></div>",
                    unsafe_allow_html=True)

    st.write("")
    # Classement épuré
    st.markdown("##### Classement sites — taux de marge")
    rank = sites.sort_values('TxMarge', ascending=False)[['Site', 'CA', 'TxMarge', 'EvoCA', 'niveau']]
    disp = rank.copy()
    disp['CA'] = disp['CA'].apply(fmt_m)
    disp['TxMarge'] = disp['TxMarge'].apply(lambda x: f"{x:.1f}%")
    disp['EvoCA'] = disp['EvoCA'].apply(lambda x: f"{x:+.1f}%")
    disp = disp.rename(columns={'TxMarge': 'Tx marge', 'EvoCA': 'vs N-1', 'niveau': 'Niveau'})
    st.dataframe(disp[['Site', 'CA', 'Tx marge', 'vs N-1', 'Niveau']],
                 hide_index=True, use_container_width=True)

    # So what auto
    st.write("")
    b, txc, txr, emix, etaux = bennet(df, worst.iloc[0]['Site']) if len(worst) else (None, 0, 0, 0, 0)
    if len(worst) and b is not None:
        part_taux = abs(etaux) / (abs(emix) + abs(etaux)) * 100 if (abs(emix) + abs(etaux)) else 0
        cause = "taux de marge" if part_taux > 50 else "mix produit"
        wname = worst.iloc[0]['Site']
        sowhat = (f"<strong>Constat :</strong> la marge {', '.join(sel_rayons) if len(sel_rayons)<=2 else 'globale'} "
                  f"est à {TX_tot:.1f}%, soit {ecart_glob:+.1f} pts vs objectif. "
                  f"<strong>Cause :</strong> l'écart vient à {part_taux:.0f}% d'un problème de {cause}, "
                  f"concentré sur {wname} ({worst.iloc[0]['TxMarge']:.1f}%). "
                  f"<strong>Action :</strong> prioriser {wname} — voir vue opérationnelle pour le détail des leviers.")
    else:
        sowhat = (f"<strong>Constat :</strong> marge à {TX_tot:.1f}%, {ecart_glob:+.1f} pts vs objectif. "
                  f"<strong>Action :</strong> {nb_alerte} site(s) à surveiller, voir vue opérationnelle.")
    st.markdown(f"<div class='so-what'>{sowhat}</div>", unsafe_allow_html=True)

# ============================================================
# VUE OPÉRATIONNELLE
# ============================================================
else:
    t1, t2, t3, t4 = st.tabs(["🏪 Diagnostic site", "⚠️ Destructeurs",
                              "⚖️ Bennet Mix/Taux", "🔍 Drill-down article"])

    # --- Onglet 1 : Diagnostic site ---
    with t1:
        k = st.columns(4)
        for col, lab, val, sub, c in [
            (k[0], "CA réseau", fmt_m(CA_tot), "", DARK),
            (k[1], "Tx marge", f"{TX_tot:.1f}%", f"cible {cible_glob:.1f}%", RED if cible_glob - TX_tot > 2 else GREEN),
            (k[2], "Sites alerte", f"{nb_alerte}/{len(sites)}", f"{nb_crit} critique(s)", AMBER),
            (k[3], "Casse", fmt_m(casse_tot), f"{casse_tot/CA_tot*100:.2f}%", RED)]:
            with col:
                st.markdown(f"<div class='kpi-card'><div class='kpi-label'>{lab}</div>"
                            f"<div class='kpi-value' style='color:{c}'>{val}</div>"
                            f"<div class='kpi-sub'>{sub}</div></div>", unsafe_allow_html=True)
        st.write("")
        order = {'critique': 0, 'moyen': 1, 'bon': 2}
        sites_sorted = sites.sort_values(by='niveau', key=lambda s: s.map(order))
        for _, r in sites_sorted.iterrows():
            ec = r['cible'] - r['TxMarge']
            c = color_for(ec)
            lab = {'critique': 'Critique', 'moyen': 'Moyen', 'bon': 'Bon'}[r['niveau']]
            pct = min(r['TxMarge'] / r['cible'] * 100, 100) if r['cible'] else 0
            st.markdown(
                f"<div class='card' style='border-left:4px solid {c}'>"
                f"<div style='display:flex;justify-content:space-between'>"
                f"<div><span class='badge' style='background:{c}22;color:{c}'>{lab}</span> "
                f"<span style='font-weight:600;margin-left:8px'>{r['Site']}</span></div>"
                f"<div style='color:{GREY};font-size:13px'>{fmt_m(r['CA'])} · promo {r['PdsPromo']:.0f}%</div></div>"
                f"<div style='display:flex;gap:24px;margin-top:8px'>"
                f"<div><span style='font-size:11px;color:{GREY}'>Tx marge</span><br>"
                f"<span style='font-size:18px;font-weight:600;color:{c}'>{r['TxMarge']:.1f}%</span></div>"
                f"<div><span style='font-size:11px;color:{GREY}'>vs N-1</span><br>"
                f"<span style='font-size:18px;font-weight:600;color:{RED if r['EvoCA']<0 else GREEN}'>{r['EvoCA']:+.1f}%</span></div>"
                f"<div style='flex:1'><span style='font-size:11px;color:{GREY}'>vers cible {r['cible']:.1f}%</span>"
                f"<div style='background:#eee;border-radius:3px;height:6px;margin-top:6px'>"
                f"<div style='background:{c};width:{pct:.0f}%;height:6px;border-radius:3px'></div></div></div>"
                f"</div></div>", unsafe_allow_html=True)

    # --- Onglet 2 : Destructeurs ---
    with t2:
        dest = detect_destructeurs(df, sites, cible_par_rayon)
        if not dest:
            st.success("Aucun destructeur majeur détecté sur ce périmètre.")
        for i, d in enumerate(dest, 1):
            tags_html = " ".join(f"<span class='badge' style='background:{RED}22;color:{RED}'>{t}</span>" for t in d['tags'])
            extra = ""
            if d['ecart'] is not None:
                extra = (f"<div style='display:flex;gap:20px;margin-top:8px;font-size:13px'>"
                         f"<div><span style='color:{GREY};font-size:11px'>Tx site</span><br><b style='color:{RED}'>{d['tx_site']}%</b></div>"
                         f"<div><span style='color:{GREY};font-size:11px'>Tx réseau</span><br><b>{d['tx_ref']}%</b></div>"
                         f"<div><span style='color:{GREY};font-size:11px'>Écart</span><br><b style='color:{RED}'>{d['ecart']} pts</b></div>"
                         f"<div><span style='color:{GREY};font-size:11px'>Poids promo</span><br><b>{d['pds_promo']}%</b></div>"
                         f"<div><span style='color:{GREY};font-size:11px'>CA</span><br><b>{fmt_m(d['ca'])}</b></div></div>")
            st.markdown(
                f"<div class='card' style='border-left:4px solid {RED}'>"
                f"<div style='display:flex;justify-content:space-between'>"
                f"<div><span style='background:{RED}22;color:{RED};border-radius:5px;padding:2px 8px;font-weight:600;font-size:12px'>{i}</span> "
                f"<span style='font-weight:600;margin-left:8px'>{d['nom']}</span>"
                f"<div style='color:{GREY};font-size:11px;margin-top:2px'>{d['type']}</div></div>"
                f"<div style='text-align:right'><div style='font-size:18px;font-weight:600;color:{RED}'>{fmt_m(d['perte'])}</div>"
                f"<div style='font-size:10px;color:{GREY}'>marge perdue vs réseau</div></div></div>"
                f"{extra}<div style='margin-top:8px'>{tags_html}</div></div>", unsafe_allow_html=True)

    # --- Onglet 3 : Bennet ---
    with t3:
        site_ref = st.selectbox("Site à analyser", sites.sort_values('TxMarge')['Site'].tolist())
        b, txc, txr, emix, etaux = bennet(df, site_ref)
        if b is None:
            st.warning("Données insuffisantes.")
        else:
            ec = txc - txr
            part_taux = abs(etaux) / (abs(emix) + abs(etaux)) * 100 if (abs(emix) + abs(etaux)) else 0
            k = st.columns(3)
            for col, lab, val, c in [
                (k[0], f"Tx {site_ref}", f"{txc:.1f}%", RED if txc < txr else GREEN),
                (k[1], "Tx réseau", f"{txr:.1f}%", DARK),
                (k[2], "Écart", f"{ec:+.1f} pts", RED if ec < 0 else GREEN)]:
                with col:
                    st.markdown(f"<div class='kpi-card'><div class='kpi-label'>{lab}</div>"
                                f"<div class='kpi-value' style='color:{c}'>{val}</div></div>", unsafe_allow_html=True)
            st.write("")
            cc = st.columns(2)
            with cc[0]:
                st.markdown(f"<div class='card'><div style='font-size:12px;color:{GREY}'>Effet Mix</div>"
                            f"<div style='font-size:26px;font-weight:600;color:{color_for(-emix)}'>{emix:+.2f} pt</div>"
                            f"<div style='font-size:12px;color:{GREY}'>structure d'assortiment</div></div>", unsafe_allow_html=True)
            with cc[1]:
                st.markdown(f"<div class='card'><div style='font-size:12px;color:{GREY}'>Effet Taux</div>"
                            f"<div style='font-size:26px;font-weight:600;color:{color_for(-etaux)}'>{etaux:+.2f} pts</div>"
                            f"<div style='font-size:12px;color:{GREY}'>marge sur mêmes familles</div></div>", unsafe_allow_html=True)
            cause = "taux de marge" if part_taux > 50 else "mix produit"
            st.markdown(f"<div class='so-what' style='margin-top:12px'>L'écart de {ec:+.1f} pts est à "
                        f"<b style='color:{AMBER}'>{part_taux:.0f}%</b> un problème de {cause}.</div>",
                        unsafe_allow_html=True)
            st.write("")
            bd = b[['tx_c', 'tx_r', 'w_c', 'w_r', 'effet_mix', 'effet_taux']].copy()
            bd.columns = ['Tx site %', 'Tx réseau %', 'Poids site', 'Poids réseau', 'Effet mix', 'Effet taux']
            bd = bd.round(2)
            st.dataframe(bd, use_container_width=True)

    # --- Onglet 4 : Drill-down + écarts inter-sites ---
    with t4:
        sub_t = st.radio("Analyse", ["Écarts inter-sites", "Top contributeurs", "Marges négatives"], horizontal=True)
        if sub_t == "Écarts inter-sites":
            ecarts = ecart_inter_sites(df)
            if len(ecarts):
                tri = st.radio("Trier par", ["Écart (pts)", "Perte estimée"], horizontal=True)
                ecarts_show = ecarts.sort_values(tri, ascending=False)
                st.dataframe(ecarts_show, hide_index=True, use_container_width=True)
                rayon_lbl = " + ".join(sel_rayons) if len(sel_rayons) <= 2 else "Multi-rayon"
                xls = build_excel(ecarts, rayon_lbl)
                st.download_button("📥 Export Excel (board + détail)", xls,
                                   file_name=f"Analyse_Ecart_Marge_{rayon_lbl.replace(' ', '_')}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Pas assez d'articles multi-sites pour cette analyse.")
        elif sub_t == "Top contributeurs":
            top = df.nlargest(30, 'CA')[['ArtLib', 'FamLib', 'Site', 'CA', 'Marge']].copy()
            top['Tx %'] = (top['Marge'] / top['CA'] * 100).round(1)
            st.dataframe(top.rename(columns={'ArtLib': 'Article', 'FamLib': 'Famille'}),
                         hide_index=True, use_container_width=True)
        else:
            neg = df[df['Marge'] < 0][['ArtLib', 'FamLib', 'Site', 'CA', 'Marge']].copy()
            neg['Tx %'] = (neg['Marge'] / neg['CA'] * 100).round(1)
            neg = neg.sort_values('Marge')
            st.dataframe(neg.rename(columns={'ArtLib': 'Article', 'FamLib': 'Famille'}),
                         hide_index=True, use_container_width=True)
            st.caption(f"{len(neg)} article(s) en marge négative · perte cumulée {fmt_m(neg['Marge'].sum())}")
