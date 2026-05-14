"""
11_📊_Rentabilite.py — SmartBuyer Hub
Suivi Déviation Marge · Pilotage acheteurs · Cibles vs N-1
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Rentabilité · SmartBuyer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1300px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-purple{ background: #F5F0FF; border-color: #AF52DE; color: #1A0033; }

.acheteur-card { border-radius: 12px; padding: 14px 16px; margin-bottom: 6px; border: 0.5px solid; }
.card-boissons { background: #E3F0FF; border-color: #B3D9FF; }
.card-epicerie { background: #F0FFF4; border-color: #A8E6BF; }
.card-dph      { background: #F5F0FF; border-color: #D9B3FF; }

.badge-ok     { display:inline-block; padding:2px 8px; border-radius:6px; font-size:11px; font-weight:600; background:#D5F5E3; color:#145A32; }
.badge-warn   { display:inline-block; padding:2px 8px; border-radius:6px; font-size:11px; font-weight:600; background:#FEF9C3; color:#854D0E; }
.badge-alert  { display:inline-block; padding:2px 8px; border-radius:6px; font-size:11px; font-weight:600; background:#FEE2E2; color:#991B1B; }
.badge-na     { display:inline-block; padding:2px 8px; border-radius:6px; font-size:11px; font-weight:600; background:#F3F4F6; color:#6B7280; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
ACHETEURS = {
    'BOISSONS':           'Acheteur Boissons',
    'EPICERIE':           'Acheteur Épicerie',
    'DROGUERIE':          'Acheteur DPH',
    'PARFUMERIE HYGIENE': 'Acheteur DPH',
}
PLANCHERS = {
    'Produit d appel': 0.10, 'Valeur ajoutee': 0.25,
    'PH Droguerie':    0.22, 'Coeur de gamme': 0.18,
}
SEG_LABELS = {
    'Produit d appel': "Produit d'appel",
    'Valeur ajoutee':  'Valeur ajoutée',
    'PH Droguerie':    'PH / Droguerie',
    'Coeur de gamme':  'Cœur de gamme',
}
PRODUITS_APPEL = ['RIZ LONG','HUILES','LAITS','EAUX PLATES','EAUX GAZEUSES',
                  'SUCRES ET LEVURES','FARINES CEREALES','LEGUMES SEC',
                  'PATES LONGUES','PATES COURTES','SEMOULES COUSCOUS',
                  'BOUILLON AIDE CULINAIRE','SAUCES FROIDES']
VALEUR_AJOUTEE = ['BIO','CHIPS','SOINS DU CORPS','SOINS DU VISAGE',
                  'PANSEMENT ET COMPLEMENTS','PRODUITS DU MONDE',
                  'SELS ET EPICES','SNACKING','DERMO COSMETIQUE','DIET','MAQUILLAGE']
ORDRE_RAYONS  = ['BOISSONS','EPICERIE','DROGUERIE','PARFUMERIE HYGIENE']
TOLERANCE     = 0.015
COLS_REQUIRED = ['Rayon','Sous Famille','CA','Marge','%Marge','CA N-1','%Vs N-1.1']

# ─── HELPERS FORMAT ────────────────────────────────────────────────────────────
def fp(v, sign=True):
    try:
        if pd.isna(v): return '—'
        return f"{v:+.1%}" if sign else f"{v:.1%}"
    except: return '—'

def fk(v):
    try:
        if pd.isna(v) or v == 0: return '—'
        a = abs(v)
        if a >= 1_000_000: return f"{v/1_000_000:+.1f} M"
        return f"{v/1000:+,.0f} K"
    except: return '—'

def badge_statut(s):
    if '✅' in str(s): return '<span class="badge-ok">✅ OK</span>'
    if '🟡' in str(s): return '<span class="badge-warn">🟡 Vigilance</span>'
    if '🔴' in str(s): return '<span class="badge-alert">🔴 Action</span>'
    return '<span class="badge-na">⚪ N/A</span>'

def cs(v):
    if '✅' in str(v): return 'background:#D5F5E3;color:#145A32;font-weight:600'
    if '🟡' in str(v): return 'background:#FEF9C3;color:#854D0E;font-weight:600'
    if '🔴' in str(v): return 'background:#FEE2E2;color:#991B1B;font-weight:600'
    return ''

def cd(v):
    try:
        x = float(str(v).replace('%','').replace('+','').replace(' K','')
                        .replace(' M','').replace(',','').replace('—','').strip())
        if x >= -1.5: return 'color:#145A32;font-weight:600'
        if x >= -3.0: return 'color:#854D0E;font-weight:600'
        return 'color:#991B1B;font-weight:600'
    except: return ''

# ─── FONCTIONS VECTORISÉES ────────────────────────────────────────────────────
def _segment_vec(sf_s, ray_s):
    sf  = sf_s.str.upper().fillna('')
    ray = ray_s.str.upper().fillna('')
    seg = pd.Series('Coeur de gamme', index=sf.index)
    seg[ray.isin(['DROGUERIE','PARFUMERIE HYGIENE'])] = 'PH Droguerie'
    mask_a = sf.apply(lambda s: any(p in s for p in PRODUITS_APPEL))
    seg[mask_a & (seg == 'Coeur de gamme')] = 'Produit d appel'
    mask_v = sf.apply(lambda s: any(v in s for v in VALEUR_AJOUTEE))
    seg[mask_v & (seg == 'Coeur de gamme')] = 'Valeur ajoutee'
    return seg

def _statut_vec(dev):
    s = pd.Series('⚪ N/A', index=dev.index)
    nn = dev.notna()
    s[nn & (dev >= -TOLERANCE)]                         = '✅ OK'
    s[nn & (dev < -TOLERANCE) & (dev >= -TOLERANCE*2)]  = '🟡 Vigilance'
    s[nn & (dev < -TOLERANCE*2)]                        = '🔴 Action requise'
    return s

def _que_faire_vec(df):
    tx  = df['%Marge']
    dev = df['Dev_N1_pts']
    seg = df['Segment']
    ca_med = df.groupby('Rayon_court')['CA'].transform('median').fillna(1)
    gros   = df['CA'] > ca_med * 1.5
    c = pd.Series('✅ RAS — surveiller la tendance', index=df.index)
    c[dev.notna() & (dev < -TOLERANCE*2)]              = '🔍 Optimiser mix promo / conditions fournisseur'
    c[dev.notna() & (dev < -0.05)]                     = '📋 Révision conditions achat + audit promos'
    c[dev.notna() & (dev < -0.05) & gros]              = '📞 Volume élevé + marge en chute — renégociation urgente'
    c[dev.notna() & (dev < -0.10)]                     = '📞 Convocation fournisseur — analyse prix achat vs marché'
    c[seg == 'Produit d appel']                        = '📋 Négocier remise arrière ou ristourne volume'
    c[(seg=='Produit d appel') & dev.notna() & (dev < -0.05)] = "📞 Produit d'appel sous pression — remise arrière urgente"
    c[tx.notna() & (tx < 0)]                           = '🚨 Marge négative — bloquer la promo immédiatement'
    return c

def _impact_score_vec(df):
    ca_med = df.groupby('Rayon_court')['CA'].transform('median').replace(0, 1)
    poids  = (df['CA'] / ca_med).clip(0.5, 3.0)
    return (df['Dev_N1_FCFA'].abs() * poids).round(0)

def _commentaires_auto(df):
    rouge = df[df['Statut'] == '🔴 Action requise'].nlargest(6, 'Impact_Score')
    out = []
    for _, r in rouge.iterrows():
        sf=r['SF_court']; ray=r['Rayon_court']; tx=r['%Marge']
        dev=r['Dev_N1_pts']; fcfa=r['Dev_N1_FCFA']; seg=r['Segment']; vol=r['CA']
        if pd.notna(tx) and tx < 0:
            out.append(('red', f"🚨 <strong>{sf}</strong> ({ray}) — marge négative à {tx:.1%}. Arrêt immédiat des promos déficitaires."))
        elif pd.notna(dev) and dev < -0.10:
            out.append(('red', f"🔴 <strong>{sf}</strong> ({ray}) — effondrement de {dev:+.1%} vs N-1 ({fcfa:+,.0f} FCFA). Convocation fournisseur urgente."))
        elif seg == 'Produit d appel':
            out.append(('amber', f"🟠 <strong>{sf}</strong> ({ray}) — produit d'appel à {tx:.1%} ({dev:+.1%} vs N-1). Négocier remise arrière."))
        else:
            vtxt = f", volume {vol/1e6:.1f}M FCFA" if vol > 5e6 else ""
            out.append(('amber', f"🟠 <strong>{sf}</strong> ({ray}{vtxt}) — {dev:+.1%} vs N-1 ({fcfa:+,.0f} FCFA). Réviser conditions achat."))
    return out

def _detect_periode(df_raw):
    col_a = df_raw.iloc[:, 0].dropna().astype(str)
    last  = col_a.iloc[-1] if len(col_a) else ''
    dates = re.findall(r'\d{2}/\d{2}/\d{4}', last)
    if len(dates) >= 2:   return f"{dates[0]} → {dates[1]}"
    elif len(dates) == 1: return dates[0]
    return 'Période inconnue'

def _validate(df, filename):
    missing = [c for c in COLS_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"**{filename}** — colonnes manquantes : `{'`, `'.join(missing)}`")

# ─── CHARGEMENT RÉFÉRENTIEL ───────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_referentiel(override_bytes=None):
    if override_bytes:
        try:
            ref = pd.read_excel(BytesIO(override_bytes))
            if 'Cible' in ref.columns: return ref
        except Exception: pass
    repo_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'referentiel_cibles.csv')
    if os.path.exists(repo_path): return pd.read_csv(repo_path)
    return None

# ─── CHARGEMENT EXTRACTION — lecture unique, tout vectorisé ──────────────────
@st.cache_data(show_spinner=False)
def load_extraction(file_bytes: bytes, filename: str, ref_bytes=None):
    raw      = BytesIO(file_bytes)
    df_raw   = pd.read_excel(raw, header=None)
    periode  = _detect_periode(df_raw)
    raw.seek(0)
    df = pd.read_excel(raw)
    _validate(df, filename)

    df = df[
        (df['Sous Famille'] != 'Total') &
        df['Rayon'].str.startswith('000', na=False) &
        ~df['Rayon'].str.contains('CIGARETTE', na=False) &
        df['CA'].notna() & (df['CA'] > 0)
    ].copy()

    df['SF_court']    = df['Sous Famille'].str.extract(r'\d+ - (.+)')[0]
    df['Rayon_court'] = df['Rayon'].str.extract(r'- (.+)')[0]
    df['Acheteur']    = df['Rayon_court'].map(ACHETEURS)
    df['Segment']     = _segment_vec(df['SF_court'], df['Rayon_court'])
    df['Plancher']    = df['Segment'].map(PLANCHERS)

    valid_n1   = df['%Vs N-1.1'].notna() & (df['%Vs N-1.1'] != -1) & (df['CA N-1'] > 0)
    df['Marge_N1'] = np.where(valid_n1, df['Marge'] / (1 + df['%Vs N-1.1']), np.nan)
    df['CA_N1']    = df['CA N-1']
    df['Tx_N1']    = np.where(valid_n1 & (df['CA_N1'] > 0), df['Marge_N1'] / df['CA_N1'], np.nan)

    ref = load_referentiel(ref_bytes)
    if ref is not None and 'Cible' in ref.columns:
        df = df.merge(
            ref[['Rayon','Famille','Cible','Plancher']].rename(
                columns={'Rayon':'Rayon_court','Famille':'SF_court',
                         'Cible':'Cible_ref','Plancher':'Plancher_ref'}),
            on=['Rayon_court','SF_court'], how='left')
        df['Cible']    = df['Cible_ref'].fillna(df['Plancher'])
        df['Plancher'] = df['Plancher_ref'].fillna(df['Plancher'])
    else:
        df['Cible'] = np.where(
            df['Tx_N1'].notna(),
            np.maximum(df['Tx_N1'] * 1.02, df['Plancher']),
            df['Plancher'])

    df['Source_cible']   = np.where(df['Tx_N1'].notna(), 'N-1 × 1,02', 'Plancher segment (nouveauté)')
    df['Dev_N1_pts']     = df['%Marge'] - df['Tx_N1']
    df['Dev_N1_FCFA']    = df['Dev_N1_pts'] * df['CA']
    df['Dev_Cible_pts']  = df['%Marge'] - df['Cible']
    df['Dev_Cible_FCFA'] = df['Dev_Cible_pts'] * df['CA']
    df['Statut']         = _statut_vec(df['Dev_N1_pts'])
    df['Impact_Score']   = _impact_score_vec(df)
    df['Que_faire']      = _que_faire_vec(df)

    _ord = {'🔴 Action requise':0,'🟡 Vigilance':1,'✅ OK':2,'⚪ N/A':3}
    df['_ord_statut'] = df['Statut'].map(_ord)
    df['_ord_rayon']  = df['Rayon_court'].map({r:i for i,r in enumerate(ORDRE_RAYONS)})
    df['Periode']     = periode
    df['Fichier']     = filename
    return df

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(df_all, periodes):
    wb = Workbook()
    C_HDR='1B2A4A'; C_SUB='2E4B7A'; C_WH='FFFFFF'; C_DK='1A1A2E'

    def xfill(h): return PatternFill('solid', fgColor=h)
    def xbdr():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def xctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def xrgt(): return Alignment(horizontal='right',  vertical='center')
    def xlft(w=False): return Alignment(horizontal='left', vertical='center', wrap_text=w)

    def title_block(ws, txt, span=15):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(row=1, column=1, value=txt)
        c.font = Font('Calibri', size=13, bold=True, color=C_WH)
        c.fill = xfill(C_HDR); c.alignment = xctr()
        ws.row_dimensions[1].height = 30

    def write_header_row(ws, row_num, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths)):
            c = ws.cell(row=row_num, column=i+1, value=h)
            c.font = Font('Calibri', size=9, bold=True, color=C_WH)
            c.fill = xfill(C_SUB); c.alignment = xctr(); c.border = xbdr()
            ws.column_dimensions[get_column_letter(i+1)].width = w
        ws.row_dimensions[row_num].height = 28

    for i_p, periode in enumerate(periodes):
        df = df_all[df_all['Periode'] == periode].copy()
        df = df.sort_values(['_ord_statut','Impact_Score'], ascending=[True, False])
        safe = periode.replace('/','').replace('→','_').replace(' ','')[:28]
        ws   = wb.active if i_p == 0 else wb.create_sheet(safe)
        ws.title = safe; ws.sheet_view.showGridLines = False

        title_block(ws, f'SUIVI RENTABILITÉ — DÉVIATION MARGE vs N-1 · {periode}', span=15)

        # Sous-titre
        ws.merge_cells('A2:O2')
        c2 = ws.cell(row=2, column=1,
            value=f'  Cible = MAX(Taux N-1 × 1,02 ; Plancher segment) · Tolérance ±1,5 pt · Trié par score d\'impact financier')
        c2.font = Font('Calibri', size=9, italic=True, color='AABBCC')
        c2.fill = xfill(C_HDR); c2.alignment = xlft()
        ws.row_dimensions[2].height = 16

        HDRS = ['Rayon','Famille','Segment','Source cible','Acheteur',
                'CA (FCFA)','Taux actuel','Taux N-1','Cible','Plancher',
                'Dév. N-1 (pts)','Dév. Cible (pts)','Marge Δ FCFA','Score Impact',
                'Statut','Que faire ?']
        WIDTHS = [20,30,16,18,18,14,12,12,12,12,14,14,16,13,16,46]
        write_header_row(ws, 3, HDRS, WIDTHS)

        C_R='FFD6D6'; C_O='FFF3CC'; C_G='D6F5D6'; C_L='F7F7F7'; C_W='FFFFFF'
        for i, (_, r) in enumerate(df.iterrows(), 4):
            stat = r.get('Statut','')
            bg = C_R if '🔴' in str(stat) else (C_O if '🟡' in str(stat)
                 else (C_G if '✅' in str(stat) else (C_L if i%2==0 else C_W)))
            vals = [
                r.get('Rayon_court',''),
                r.get('SF_court',''),
                SEG_LABELS.get(r.get('Segment',''), r.get('Segment','')),
                r.get('Source_cible',''),
                r.get('Acheteur',''),
                r.get('CA', None),
                r.get('%Marge', None),
                r.get('Tx_N1', None),
                r.get('Cible', None),
                r.get('Plancher', None),
                r.get('Dev_N1_pts', None),
                r.get('Dev_Cible_pts', None),
                r.get('Dev_N1_FCFA', None),
                r.get('Impact_Score', None),
                stat,
                r.get('Que_faire',''),
            ]
            FMTS = [None,None,None,None,None,'#,##0','0.0%','0.0%','0.0%','0.0%',
                    '+0.0%;-0.0%;-','+0.0%;-0.0%;-','+#,##0;-#,##0;-','#,##0',None,None]
            for j, (v, f) in enumerate(zip(vals, FMTS), 1):
                c = ws.cell(row=i, column=j, value=v)
                c.fill = xfill(bg); c.border = xbdr()
                c.font = Font('Calibri', size=10 if j < 15 else 9, color=C_DK)
                if f: c.number_format = f
                c.alignment = xctr() if j in (7,8,9,10,11,12,15) else xrgt() if j in (6,13,14) else xlft(w=(j==16))
            ws.row_dimensions[i].height = 16

        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS))}{3+len(df)}'

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Navigation</div>", unsafe_allow_html=True)
    st.page_link("app.py",                                      label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",          label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",                   label="📈  Ventes PBI")
    st.page_link("pages/03_📦_Detention_Top_CA.py",             label="📦  Détention Top CA")
    st.page_link("pages/04_💸_Performance_Promo.py",            label="💸  Performance Promo")
    st.page_link("pages/05_🏪_Suivi_Implantation.py",           label="🏪  Suivi Implantation")
    st.page_link("pages/06_💸_Marges_Negatives.py",             label="💸  Marges Négatives")
    st.page_link("pages/09_📦_OTIF.py",                         label="📦  OTIF")
    st.page_link("pages/10_📉_OOS_Ruptures.py",                 label="📉  OOS Ruptures")
    st.page_link("pages/11_📊_Rentabilite.py",                  label="📊  Rentabilité")
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichiers</div>", unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Extraction(s) PBI", type=['xlsx'],
        accept_multiple_files=True,
        help=f"Colonnes requises : {', '.join(COLS_REQUIRED)}"
    )
    ref_override = st.file_uploader(
        "Référentiel cibles (optionnel)", type=['xlsx'],
        help="Laissez vide → référentiel embarqué. Uploadez pour écraser."
    )
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtres</div>", unsafe_allow_html=True)
    seuil_fcfa = st.slider("Impact min (K FCFA)", 0, 2000, 0, 50,
                           help="Masquer les familles sous ce seuil de marge perdue")
    st.markdown("---")
    st.markdown("""
<div style='font-size:11px;color:#8E8E93;line-height:2.2'>
  <span style='color:#FF3B30'>●</span> &lt; −3 pts vs N-1 : Action requise<br>
  <span style='color:#FF9500'>●</span> −1,5 à −3 pts : Vigilance<br>
  <span style='color:#34C759'>●</span> &gt; −1,5 pt : OK
</div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.caption("NovaRetail Solutions · SmartBuyer v2.2")

# ─── PAGE VIDE ────────────────────────────────────────────────────────────────
if not uploaded_files:
    st.markdown("<div class='page-title'>📊 Rentabilité — Suivi Déviation Marge</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-caption'>Pilotage hebdomadaire des taux de marge · Déviation vs N-1 · Cibles acheteurs · Score d'impact financier</div>", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Ce module pilote la <strong>déviation du taux de marge vs N-1</strong> famille par famille, en croisant trois dimensions :
  acheteur, rayon et magasin. Il répond à la question : <em>où ça coince, pourquoi, et que faire ?</em>
  <br><br>
  <strong>1. Vue Réseau</strong> — KPIs globaux, synthèse par rayon, top familles en alerte triées par impact financier<br>
  <strong>2. Vue Acheteur</strong> — briefing personnalisé, familles prioritaires, colonne "Que faire ?"<br>
  <strong>3. Vue Magasin</strong> — comparaison site par site (nécessite extraction avec dimension Site)<br>
  <strong>4. Tendance</strong> — évolution multi-périodes, détection des dégradations persistantes
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Indicateurs calculés automatiquement</div>", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    indics = [
        ("📉", "Déviation vs N-1", "#FF3B30",
         "Écart entre le taux de marge actuel et le taux de la même période N-1",
         "Δ = Tx marge actuel − Tx marge N-1",
         "Signal principal : la marge progresse ou régresse par rapport à l'année dernière."),
        ("🎯", "Déviation vs Cible", "#007AFF",
         "Écart entre le taux réalisé et la cible fixée avec la direction",
         "Cible = MAX(Taux N-1 × 1,02 ; Plancher segment)",
         "Mesure l'atteinte de l'objectif de progression défini en début d'exercice."),
        ("💰", "Score d'impact financier", "#FF9500",
         "Marge perdue pondérée par le poids volume de la famille",
         "Score = |Marge Δ FCFA| × Poids volume (0,5→3,0)",
         "Priorise les actions : une petite déviation sur le Riz > grosse déviation sur une petite famille."),
        ("🏷️", "Segmentation automatique", "#34C759",
         "Classification de chaque famille selon sa nature commerciale",
         "Appel (10%) · Cœur (18%) · Val. aj. (25%) · PH/Drog (22%)",
         "Applique le plancher P&L adapté : évite de pénaliser les produits d'appel structurellement bas."),
        ("🆕", "Gestion des nouveautés", "#AF52DE",
         "Familles sans historique N-1 traitées séparément",
         "Si Tx N-1 = N/A → Cible = Plancher segment uniquement",
         "Pas de faux écart sur les articles lancés après la période N-1."),
    ]
    for i, (ico, titre, color, desc, formule, interp) in enumerate(indics):
        with (c1 if i % 2 == 0 else c2):
            st.markdown(f"""
<div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;
            padding:16px;border-left:3px solid {color};margin-bottom:10px'>
  <div style='display:flex;align-items:center;gap:8px;margin-bottom:8px'>
    <span style='font-size:18px'>{ico}</span>
    <span style='font-size:14px;font-weight:600;color:#1C1C1E'>{titre}</span>
  </div>
  <div style='font-size:12px;color:#3A3A3C;margin-bottom:4px'>{desc}</div>
  <div style='font-size:11px;color:{color};font-family:monospace;background:#F9F9FB;
              padding:4px 8px;border-radius:6px;margin-bottom:6px'>{formule}</div>
  <div style='font-size:11px;color:#8E8E93;font-style:italic'>{interp}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Fichier attendu</div>", unsafe_allow_html=True)
    st.markdown(f"""
<div style='background:#F0F8FF;border:0.5px solid #B3D9FF;border-radius:10px;padding:12px 16px;margin-bottom:8px'>
  <div style='font-size:13px;font-weight:600;color:#0066CC;font-family:monospace;margin-bottom:4px'>Export PBI ventes — format hebdomadaire ou journalier</div>
  <div style='font-size:12px;color:#3A3A3C'>Axes : Rayon / Sous Famille · Colonnes obligatoires : {', '.join(COLS_REQUIRED)}</div>
  <div style='font-size:12px;color:#8E8E93;margin-top:4px'>La période est détectée automatiquement depuis la dernière cellule de la colonne A.</div>
</div>""", unsafe_allow_html=True)
    st.info("⬆️ Charge le(s) fichier(s) extraction PBI dans la sidebar pour lancer l'analyse.")
    st.stop()

# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
ref_bytes = ref_override.read() if ref_override else None
all_dfs, errors = [], []

for f in uploaded_files:
    raw = f.read()
    try:    all_dfs.append(load_extraction(raw, f.name, ref_bytes))
    except ValueError as e: errors.append(str(e))
    except Exception as e:  errors.append(f"Erreur inattendue **{f.name}** : {e}")

for err in errors:
    st.sidebar.error(err)

if not all_dfs:
    st.error("Aucun fichier valide chargé. Vérifiez les colonnes obligatoires.")
    st.stop()

df_all = pd.concat(all_dfs, ignore_index=True)
periodes_dispo = sorted(df_all['Periode'].unique(), reverse=True)

with st.sidebar:
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px'>Périodes chargées</div>", unsafe_allow_html=True)
    for p in periodes_dispo:
        st.markdown(f"<span style='background:#E3F0FF;color:#185FA5;border-radius:20px;padding:3px 10px;font-size:11px;font-weight:500;display:inline-block;margin:2px 0'>{p}</span>", unsafe_allow_html=True)
    st.markdown("")
    periode_sel = st.selectbox("Période active", periodes_dispo, label_visibility='collapsed')

df = df_all[df_all['Periode'] == periode_sel].copy()
if seuil_fcfa > 0:
    df = df[(df['Dev_N1_FCFA'].abs() >= seuil_fcfa * 1000) | df['Dev_N1_FCFA'].isna()]

n_nouv = (df_all[df_all['Periode']==periode_sel]['Source_cible'].str.contains('nouveauté', na=False)).sum()

# ─── HEADER PAGE ──────────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📊 Rentabilité — Suivi Déviation Marge</div>", unsafe_allow_html=True)
st.markdown(f"<div class='page-caption'>Période active : <strong>{periode_sel}</strong> · {len(periodes_dispo)} période(s) chargée(s) · {n_nouv} nouveauté(s) sur cible plancher uniquement</div>", unsafe_allow_html=True)

# ─── KPIs GLOBAUX ─────────────────────────────────────────────────────────────
ca_t  = df['CA'].sum(); mg_t = df['Marge'].sum(); tx_t = mg_t/ca_t if ca_t>0 else 0
mn1_t = df['Marge_N1'].sum(); cn1_t = df['CA_N1'].sum(); tx_n1 = mn1_t/cn1_t if cn1_t>0 else 0
dev_t = tx_t - tx_n1
n_r   = (df['Statut']=='🔴 Action requise').sum()
n_o   = (df['Statut']=='🟡 Vigilance').sum()
n_v   = (df['Statut']=='✅ OK').sum()
cib_t = (df['Cible']*df['CA']).sum()/ca_t if ca_t>0 else 0

st.markdown(f"<div class='section-label'>{len(df)} famille(s) · {(df['Rayon_court'].nunique())} rayon(s) · taux cible moyen {cib_t:.1%}</div>", unsafe_allow_html=True)
k1,k2,k3,k4,k5,k6 = st.columns(6)
k1.metric("Taux Actuel",   f"{tx_t:.1%}",  fp(dev_t))
k2.metric("Taux N-1",      f"{tx_n1:.1%}")
k3.metric("Marge Δ FCFA",  fk(df['Dev_N1_FCFA'].sum()))
k4.metric("Cible Réseau",  f"{cib_t:.1%}")
k5.metric("🔴 Action",     f"{n_r}",  f"sur {len(df)} familles")
k6.metric("✅ OK",         f"{n_v}")

# ─── ALERTES ──────────────────────────────────────────────────────────────────
st.markdown("---")
commentaires = _commentaires_auto(df)
if commentaires:
    st.markdown("<div class='section-label'>Signaux critiques — Points de blocage prioritaires</div>", unsafe_allow_html=True)
    for cls, txt in commentaires:
        st.markdown(f"<div class='alert-card alert-{'red' if cls=='red' else 'amber'}'>{txt}</div>", unsafe_allow_html=True)

if n_v == len(df):
    st.markdown("<div class='alert-card alert-green'>✅ <strong>Aucune alerte rouge cette période.</strong> Tous les rayons sont dans la tolérance vs N-1.</div>", unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab1, tab2, tab3, tab4 = st.tabs(["📊 Réseau", "👤 Acheteur", "🏪 Magasin", "📥 Export Excel"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — RÉSEAU
# ══════════════════════════════════════════════════════════════════════════════
with tab1:

    # Synthèse par rayon — cards comme module 06
    st.markdown("<div class='section-label'>Performance par rayon</div>", unsafe_allow_html=True)
    rayon_cols = st.columns(len(ORDRE_RAYONS))
    RAYON_STYLES = {
        'BOISSONS':           ('#154360','#E3F0FF','#B3D9FF'),
        'EPICERIE':           ('#145A32','#F0FFF4','#A8E6BF'),
        'DROGUERIE':          ('#6E2F8A','#F5F0FF','#D9B3FF'),
        'PARFUMERIE HYGIENE': ('#7D1435','#FCE4EC','#F5A7B8'),
    }
    for i, rayon in enumerate(ORDRE_RAYONS):
        sub = df[df['Rayon_court'] == rayon]
        if len(sub) == 0: continue
        fc, bg, brd = RAYON_STYLES.get(rayon, ('#3A3A3C','#F9F9FB','#CCCCCC'))
        ca_r  = sub['CA'].sum(); mg_r = sub['Marge'].sum(); tx_r = mg_r/ca_r if ca_r>0 else 0
        mn1_r = sub['Marge_N1'].sum(); cn1_r = sub['CA_N1'].sum()
        tn1_r = mn1_r/cn1_r if cn1_r>0 else 0
        cib_r = (sub['Cible']*sub['CA']).sum()/ca_r if ca_r>0 else 0
        dev_r = tx_r - tn1_r
        n_r_r = (sub['Statut']=='🔴 Action requise').sum()
        with rayon_cols[i]:
            st.markdown(f"""
<div style='background:{bg};border:1px solid {brd};border-radius:12px;padding:16px;margin-bottom:8px'>
  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:10px'>
    <span style='font-size:13px;font-weight:700;color:{fc}'>{rayon.title()}</span>
    <span style='font-size:11px;color:#8E8E93'>{ACHETEURS.get(rayon,"—").replace("Acheteur ","")}</span>
  </div>
  <div style='font-size:26px;font-weight:700;color:{fc};letter-spacing:-0.02em'>{fp(tx_r,False)}</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:2px'>Taux de marge</div>
  <hr style='margin:10px 0;border-color:{brd}'>
  <div style='display:grid;grid-template-columns:1fr 1fr;gap:6px;font-size:12px'>
    <div><span style='color:#8E8E93'>N-1</span><br><strong>{fp(tn1_r,False)}</strong></div>
    <div><span style='color:#8E8E93'>Cible</span><br><strong>{fp(cib_r,False)}</strong></div>
    <div><span style='color:#8E8E93'>Dév. N-1</span><br><strong style='color:{"#34C759" if dev_r>=-TOLERANCE else "#FF9500" if dev_r>=-TOLERANCE*2 else "#FF3B30"}'>{fp(dev_r)}</strong></div>
    <div><span style='color:#8E8E93'>🔴 Alertes</span><br><strong>{n_r_r}</strong></div>
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Récapitulatif par rayon — Réalisé vs N-1 vs Cible</div>", unsafe_allow_html=True)

    rows_r = []
    for rayon in ORDRE_RAYONS:
        sub = df[df['Rayon_court']==rayon]
        if len(sub)==0: continue
        ca_r=sub['CA'].sum(); mg_r=sub['Marge'].sum(); tx_r=mg_r/ca_r if ca_r>0 else 0
        mn1_r=sub['Marge_N1'].sum(); cn1_r=sub['CA_N1'].sum(); tn1_r=mn1_r/cn1_r if cn1_r>0 else 0
        cib_r=(sub['Cible']*sub['CA']).sum()/ca_r if ca_r>0 else 0
        rows_r.append({
            'Rayon': rayon.title(),'Acheteur': ACHETEURS.get(rayon,'—'),
            'CA (FCFA)': f"{ca_r:,.0f}",'Poids CA': fp(ca_r/ca_t,False),
            'Taux actuel': fp(tx_r,False),'Taux N-1': fp(tn1_r,False),
            'Dév. vs N-1': fp(tx_r-tn1_r),'Cible': fp(cib_r,False),
            'Dév. vs Cible': fp(tx_r-cib_r),
            'Marge Δ FCFA': fk(sub['Dev_N1_FCFA'].sum()),
            '🔴': (sub['Statut']=='🔴 Action requise').sum(),
            '🟡': (sub['Statut']=='🟡 Vigilance').sum(),
            '✅': (sub['Statut']=='✅ OK').sum(),
        })
    st.dataframe(pd.DataFrame(rows_r), use_container_width=True, hide_index=True,
                 column_config={'Rayon': st.column_config.TextColumn('Rayon', width='medium'),
                                'Acheteur': st.column_config.TextColumn('Acheteur', width='medium')})

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"<div class='section-label'>Familles en action requise — triées par score d'impact financier</div>", unsafe_allow_html=True)

    rouge = df[df['Statut']=='🔴 Action requise'].sort_values('Impact_Score', ascending=False)
    if len(rouge) == 0:
        st.markdown("<div class='alert-card alert-green'>✅ Aucune famille en alerte rouge cette période.</div>", unsafe_allow_html=True)
    else:
        disp = rouge[['Rayon_court','SF_court','Segment','Acheteur','CA',
                      '%Marge','Tx_N1','Dev_N1_pts','Dev_N1_FCFA','Impact_Score',
                      'Cible','Source_cible','Dev_Cible_pts','Statut','Que_faire']].copy()
        disp.columns = ['Rayon','Famille','Segment','Acheteur','CA (FCFA)',
                        'Taux act.','Taux N-1','Dév. N-1','Marge Δ FCFA','Score Impact',
                        'Cible','Source cible','Dév. Cible','Statut','Que faire ?']
        disp['CA (FCFA)']=disp['CA (FCFA)'].apply(lambda x: f"{x:,.0f}")
        disp['Score Impact']=disp['Score Impact'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '—')
        for col in ['Taux act.','Taux N-1','Cible']: disp[col]=disp[col].apply(lambda x: fp(x,False))
        for col in ['Dév. N-1','Dév. Cible']:         disp[col]=disp[col].apply(fp)
        disp['Marge Δ FCFA']=disp['Marge Δ FCFA'].apply(fk)
        disp['Segment']=disp['Segment'].apply(lambda x: SEG_LABELS.get(x,x))
        st.dataframe(
            disp.style.map(cs, subset=['Statut']).map(cd, subset=['Dév. N-1','Dév. Cible']),
            use_container_width=True, hide_index=True, height=420,
            column_config={
                'Famille':    st.column_config.TextColumn('Famille',    width='large'),
                'Que faire ?':st.column_config.TextColumn('Que faire ?',width='large'),
            }
        )

    with st.expander(f"🟡 Familles en vigilance ({n_o})"):
        orange = df[df['Statut']=='🟡 Vigilance'].sort_values('Impact_Score',ascending=False)
        d2 = orange[['Rayon_court','SF_court','Segment','Acheteur',
                     '%Marge','Tx_N1','Dev_N1_pts','Dev_N1_FCFA','Impact_Score','Que_faire']].copy()
        d2.columns=['Rayon','Famille','Segment','Acheteur',
                    'Taux act.','Taux N-1','Dév. N-1','Marge Δ FCFA','Score Impact','Que faire ?']
        for col in ['Taux act.','Taux N-1']: d2[col]=d2[col].apply(lambda x: fp(x,False))
        d2['Dév. N-1']=d2['Dév. N-1'].apply(fp)
        d2['Marge Δ FCFA']=d2['Marge Δ FCFA'].apply(fk)
        d2['Score Impact']=d2['Score Impact'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '—')
        d2['Segment']=d2['Segment'].apply(lambda x: SEG_LABELS.get(x,x))
        st.dataframe(d2.style.map(cd, subset=['Dév. N-1']),
                     use_container_width=True, hide_index=True)
    st.caption("Score Impact = marge perdue (FCFA) × poids volume · priorise les grosses familles en dérive")

    # Tendance multi-périodes si plusieurs fichiers
    if len(periodes_dispo) > 1:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Évolution globale — comparaison périodes chargées</div>", unsafe_allow_html=True)
        rows_t = []
        for p in sorted(periodes_dispo):
            dp=df_all[df_all['Periode']==p]; ca_p=dp['CA'].sum(); mg_p=dp['Marge'].sum()
            tx_p=mg_p/ca_p if ca_p>0 else 0
            mn1_p=dp['Marge_N1'].sum(); cn1_p=dp['CA_N1'].sum(); tn1_p=mn1_p/cn1_p if cn1_p>0 else 0
            rows_t.append({'Période':p,'Taux réalisé':fp(tx_p,False),'Taux N-1':fp(tn1_p,False),
                           'Dév. vs N-1':fp(tx_p-tn1_p),'Marge Δ FCFA':fk(dp['Dev_N1_FCFA'].sum()),
                           '🔴':(dp['Statut']=='🔴 Action requise').sum(),
                           '🟡':(dp['Statut']=='🟡 Vigilance').sum(),
                           '✅':(dp['Statut']=='✅ OK').sum()})
        st.dataframe(pd.DataFrame(rows_t).style.map(cd, subset=['Dév. vs N-1']),
                     use_container_width=True, hide_index=True)
        st.caption("Chargez d'autres périodes dans la sidebar pour enrichir cette vue tendance.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ACHETEUR
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    acheteurs = sorted(df['Acheteur'].dropna().unique())
    ach_sel   = st.selectbox("Acheteur", acheteurs, key='ach_sel')
    df_ach    = df[df['Acheteur']==ach_sel].sort_values(['_ord_statut','Impact_Score'],ascending=[True,False])

    ca_a  = df_ach['CA'].sum(); mg_a=df_ach['Marge'].sum(); tx_a=mg_a/ca_a if ca_a>0 else 0
    mn1_a = df_ach['Marge_N1'].sum(); cn1_a=df_ach['CA_N1'].sum(); tn1_a=mn1_a/cn1_a if cn1_a>0 else 0
    cib_a = (df_ach['Cible']*df_ach['CA']).sum()/ca_a if ca_a>0 else 0
    n_r_a = (df_ach['Statut']=='🔴 Action requise').sum()

    st.markdown(f"<div class='section-label'>{ach_sel} · {periode_sel}</div>", unsafe_allow_html=True)
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("Taux réalisé", f"{tx_a:.1%}", fp(tx_a-tn1_a))
    k2.metric("Taux N-1",     f"{tn1_a:.1%}")
    k3.metric("Marge Δ",      fk(df_ach['Dev_N1_FCFA'].sum()))
    k4.metric("Cible moy.",   f"{cib_a:.1%}")
    k5.metric("🔴 Alertes",   f"{n_r_a}", f"sur {len(df_ach)} familles")

    st.markdown("")
    if n_r_a > 0:
        top3 = df_ach[df_ach['Statut']=='🔴 Action requise'].nlargest(3,'Impact_Score')['SF_court'].tolist()
        st.markdown(f"""
<div class='alert-card alert-amber'>
  🎯 <strong>Briefing {ach_sel} — semaine {periode_sel}</strong><br>
  {n_r_a} famille(s) en alerte rouge · Priorités immédiates : <strong>{' · '.join(top3)}</strong><br>
  <span style='font-size:12px;opacity:.85'>→ Tableau trié par score d'impact · Colonne "Que faire ?" = action à mener cette semaine.</span>
</div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
<div class='alert-card alert-green'>
  ✅ <strong>{ach_sel}</strong> — Aucune alerte rouge cette période. Surveiller les familles en vigilance ci-dessous.
</div>""", unsafe_allow_html=True)

    st.markdown("<div class='section-label'>Toutes les familles — triées par priorité et impact financier</div>", unsafe_allow_html=True)
    da = df_ach[['Rayon_court','SF_court','Segment','CA','%Marge','Tx_N1',
                 'Dev_N1_pts','Dev_N1_FCFA','Impact_Score','Cible','Plancher',
                 'Source_cible','Dev_Cible_pts','Statut','Que_faire']].copy()
    da.columns=['Rayon','Famille','Segment','CA (FCFA)','Taux act.','Taux N-1',
                'Dév. N-1','Marge Δ FCFA','Score Impact','Cible','Plancher',
                'Source cible','Dév. Cible','Statut','Que faire ?']
    da['CA (FCFA)']=da['CA (FCFA)'].apply(lambda x: f"{x:,.0f}")
    da['Score Impact']=da['Score Impact'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '—')
    for col in ['Taux act.','Taux N-1','Cible','Plancher']: da[col]=da[col].apply(lambda x: fp(x,False))
    for col in ['Dév. N-1','Dév. Cible']: da[col]=da[col].apply(fp)
    da['Marge Δ FCFA']=da['Marge Δ FCFA'].apply(fk)
    da['Segment']=da['Segment'].apply(lambda x: SEG_LABELS.get(x,x))
    st.dataframe(
        da.style.map(cs,subset=['Statut']).map(cd,subset=['Dév. N-1','Dév. Cible']),
        use_container_width=True, hide_index=True, height=560,
        column_config={
            'Famille':    st.column_config.TextColumn('Famille',    width='large'),
            'Que faire ?':st.column_config.TextColumn('Que faire ?',width='large'),
        }
    )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — MAGASIN
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    has_site = 'Site nom long' in df.columns and df['Site nom long'].notna().any()
    if not has_site:
        st.markdown("""
<div class='alert-card alert-blue'>
  ℹ️ <strong>Extraction au niveau réseau — dimension magasin non disponible.</strong><br>
  Pour activer cette vue, relancez l'extraction PBI en ajoutant l'axe <strong>Site nom long</strong> en plus de Rayon et Sous Famille.
  Le format et la détection de période restent identiques.
</div>""", unsafe_allow_html=True)
    else:
        sites = sorted([s for s in df['Site nom long'].dropna().unique() if s not in ['Total','']])
        cf1,cf2,cf3 = st.columns(3)
        with cf1: site_sel = st.selectbox("Magasin",['Tous']+sites,key='site_sel')
        with cf2: rayon_f  = st.selectbox("Rayon",  ['Tous']+[r.title() for r in ORDRE_RAYONS],key='rayon_f')
        with cf3: stat_f   = st.selectbox("Statut", ['Tous','🔴 Action requise','🟡 Vigilance','✅ OK'],key='stat_f')

        df_mag = df.copy()
        if site_sel != 'Tous': df_mag = df_mag[df_mag['Site nom long']==site_sel]
        if rayon_f  != 'Tous': df_mag = df_mag[df_mag['Rayon_court']==rayon_f.upper()]
        if stat_f   != 'Tous': df_mag = df_mag[df_mag['Statut']==stat_f]

        if site_sel == 'Tous':
            st.markdown("<div class='section-label'>Palmarès magasins — classé par taux de marge décroissant</div>", unsafe_allow_html=True)
            rows_s = []
            for site in sites:
                sub_s=df[df['Site nom long']==site]; ca_s=sub_s['CA'].sum()
                if ca_s==0: continue
                mg_s=sub_s['Marge'].sum(); tx_s=mg_s/ca_s
                mn1_s=sub_s['Marge_N1'].sum(); cn1_s=sub_s['CA_N1'].sum()
                tn1_s=mn1_s/cn1_s if cn1_s>0 else 0
                dev_s=tx_s-tn1_s
                rows_s.append({
                    'Magasin':site,'CA (FCFA)':f"{ca_s:,.0f}",
                    'Taux act.':fp(tx_s,False),'Taux N-1':fp(tn1_s,False),
                    'Dév. vs N-1':fp(dev_s),'Marge Δ FCFA':fk(sub_s['Dev_N1_FCFA'].sum()),
                    '🔴':(sub_s['Statut']=='🔴 Action requise').sum(),
                    '🟡':(sub_s['Statut']=='🟡 Vigilance').sum(),
                    '✅':(sub_s['Statut']=='✅ OK').sum(),
                    'Statut':'🔴 Action requise' if dev_s<-TOLERANCE*2 else ('🟡 Vigilance' if dev_s<-TOLERANCE else '✅ OK'),
                })
            df_sites = pd.DataFrame(rows_s).sort_values('Taux act.', ascending=False).reset_index(drop=True)
            df_sites.insert(0,'Rang',range(1,len(df_sites)+1))
            st.dataframe(
                df_sites.style.map(cs,subset=['Statut']).map(cd,subset=['Dév. vs N-1']),
                use_container_width=True, hide_index=True,
                column_config={'Magasin': st.column_config.TextColumn('Magasin', width='medium')}
            )
        else:
            st.markdown(f"<div class='section-label'>{site_sel} — familles triées par priorité</div>", unsafe_allow_html=True)
            dm = df_mag.sort_values(['_ord_statut','Impact_Score'],ascending=[True,False])
            d3=dm[['Rayon_court','SF_court','Segment','Acheteur','%Marge','Tx_N1',
                   'Dev_N1_pts','Dev_N1_FCFA','Impact_Score','Statut','Que_faire']].copy()
            d3.columns=['Rayon','Famille','Segment','Acheteur','Taux act.','Taux N-1',
                        'Dév. N-1','Marge Δ FCFA','Score Impact','Statut','Que faire ?']
            for col in ['Taux act.','Taux N-1']: d3[col]=d3[col].apply(lambda x: fp(x,False))
            d3['Dév. N-1']=d3['Dév. N-1'].apply(fp)
            d3['Marge Δ FCFA']=d3['Marge Δ FCFA'].apply(fk)
            d3['Score Impact']=d3['Score Impact'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '—')
            d3['Segment']=d3['Segment'].apply(lambda x: SEG_LABELS.get(x,x))
            st.dataframe(
                d3.style.map(cs,subset=['Statut']).map(cd,subset=['Dév. N-1']),
                use_container_width=True, hide_index=True, height=520,
                column_config={'Que faire ?': st.column_config.TextColumn('Que faire ?', width='large')}
            )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("<div class='section-label'>Export Excel — Rapport complet</div>", unsafe_allow_html=True)
    st.markdown(f"""
<div class='alert-card alert-blue'>
  <strong>📋 Contenu du fichier exporté :</strong><br>
  <strong>Un onglet par période chargée</strong> — toutes les familles avec : Rayon · Famille · Segment · Source cible · Acheteur · CA · Taux actuel · Taux N-1 · Cible · Plancher · Dév. N-1 · Dév. Cible · Marge Δ FCFA · Score Impact · Statut · Que faire ?<br>
  Mise en couleur automatique : 🔴 rouge · 🟡 orange · ✅ vert · Tri par score d'impact décroissant
</div>""", unsafe_allow_html=True)
    st.caption(f"Périmètre : {len(periodes_dispo)} période(s) · {len(df)} famille(s) active(s) après filtre seuil · {periode_sel}")

    if len(periodes_dispo) > 1:
        st.markdown("<div class='section-label'>Dégradations persistantes — rouge sur toutes les périodes</div>", unsafe_allow_html=True)
        rouge_sets={p:set(df_all[df_all['Periode']==p][df_all[df_all['Periode']==p]['Statut']=='🔴 Action requise']
                   .apply(lambda r: f"{r['Rayon_court']}|{r['SF_court']}",axis=1)) for p in periodes_dispo}
        persistants=set.intersection(*rouge_sets.values()) if rouge_sets else set()
        if persistants:
            st.markdown(f"""
<div class='alert-card alert-red'>
  ⚠️ <strong>{len(persistants)} famille(s)</strong> en rouge sur toutes les {len(periodes_dispo)} périodes chargées — problèmes structurels à inscrire en ordre du jour fournisseur.
</div>""", unsafe_allow_html=True)
            rows_p=[]
            for key in sorted(persistants):
                rayon,sf=key.split('|',1)
                sub=df[(df['Rayon_court']==rayon)&(df['SF_court']==sf)]
                if len(sub):
                    r0=sub.iloc[0]
                    rows_p.append({'Rayon':rayon,'Famille':sf,'Acheteur':r0.get('Acheteur','—'),
                                   'Taux actuel':fp(r0.get('%Marge'),False),
                                   'Score Impact':f"{r0.get('Impact_Score',0):,.0f}",
                                   'Que faire ?':r0.get('Que_faire','—')})
            st.dataframe(pd.DataFrame(rows_p).sort_values('Score Impact',ascending=False),
                         use_container_width=True, hide_index=True,
                         column_config={'Que faire ?': st.column_config.TextColumn('Que faire ?', width='large')})
        else:
            st.markdown("<div class='alert-card alert-green'>✅ Aucune famille en rouge sur toutes les périodes.</div>", unsafe_allow_html=True)

    st.markdown("")
    if st.button("Générer le fichier Excel", type="primary", key="gen_excel"):
        with st.spinner("Génération du rapport…"):
            buf = export_excel(df_all, periodes_dispo)
        st.download_button(
            label="⬇️ Télécharger le rapport Excel",
            data=buf,
            file_name=f"SmartBuyer_Rentabilite_{periode_sel.replace('/','').replace('→','_').replace(' ','')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
