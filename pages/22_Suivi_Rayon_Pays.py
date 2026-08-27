"""
22_📈_Suivi_Rayon_Pays.py — Module Suivi Rayon vs Pays · SmartBuyer Hub
Compare la performance d'un rayon PGC à la Tendance Pays (tout le réseau)
et au total PGC, avec une liste de magasins à regarder priorisée.
Aucune persistance externe : jusqu'à MAX_HISTORIQUES exports PBI sont
chargés dans la barre latérale à chaque session (pas de Google Sheets).
"""

import datetime as _dt
import re
import io

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# ============================================================
# CONFIG & CHARTE (Apple clair — identique aux autres modules)
# ============================================================
st.set_page_config(page_title="Suivi Rayon vs Pays · SmartBuyer", page_icon="📈", layout="wide", initial_sidebar_state="expanded")

BLUE = "#007AFF"
GREEN = "#34C759"
RED = "#FF3B30"
AMBER = "#FF9500"
DARK = "#1D1D1F"
GREY = "#86868B"
BG = "#F2F2F7"

MAX_HISTORIQUES = 6

RAYONS = [
    "Tous PGC",
    "010 - BOISSON",
    "011 - DROGUERIE",
    "012 - PARFUMERIE HYGIENE",
    "014 - EPICERIE",
]

# Ordre exact des colonnes d'un export PBI (Département -> Volume_VsN1_pct).
EXPORT_COLUMNS = [
    "Departement", "Rayon", "Site", "CA_N1", "Budget", "CA", "Poids",
    "VsN1_pct", "VsBgt_pct", "Marge_N1", "Marge", "TauxMarge_N1", "TauxMarge",
    "TauxMarge_VsN1", "Debit_N1", "Debit", "Debit_VsN1_pct",
    "Panier_N1", "Panier", "Panier_VsN1_pct", "PanierQte_N1", "PanierQte",
    "PanierQte_VsN1_pct", "Volume_N1", "Volume", "Volume_VsN1_pct",
]

SEUILS_FORMAT = {"Hyper": 0.02, "Market": 0.03, "Supeco": 0.04}
SEUIL_HISTORIQUE_SUFFISANT = 4

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1350px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
.stButton > button[kind="primary"] { background: #007AFF !important; border: none !important; border-radius: 8px !important; font-weight: 600 !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; background: #FFFFFF; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-purple{ background: #F5F0FF; border-color: #AF52DE; color: #1A0033; }

.format-card { border-radius: 12px; padding: 14px 16px; margin-bottom: 6px; border: 0.5px solid; }
.format-hyper  { background: #EFF6FF; border-color: #B3D9FF; }
.format-market { background: #F0FFF4; border-color: #A8E6BF; }
.format-supeco { background: #F5F0FF; border-color: #D9B3FF; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 600; }
.badge-hyper  { background: #154360; color: #FFFFFF; }
.badge-market { background: #145A32; color: #FFFFFF; }
.badge-supeco { background: #6E2F8A; color: #FFFFFF; }
.badge-red    { background: #FF3B30; color: #FFFFFF; }
.badge-green  { background: #34C759; color: #FFFFFF; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.card { background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px;margin-bottom:10px; }
.small-muted { font-size:12px;color:#8E8E93; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# HELPERS DE FORMAT (convention SmartBuyer)
# ============================================================
def fmt(n):
    if n is None or (isinstance(n, float) and (pd.isna(n) or not np.isfinite(n))):
        return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if v is None or pd.isna(v): return "—"
    return f"{v:.{dec}f}%"

def fmt_delta(v):
    if v is None or pd.isna(v): return "—"
    return f"{v:+.1f} pts"

# ============================================================
# PARSING D'UN EXPORT PBI
# ============================================================
def detect_format(site) -> str:
    if pd.isna(site):
        return ""
    s = str(site)
    if "Hyper" in s: return "Hyper"
    if "Market" in s: return "Market"
    if "Supeco" in s: return "Supeco"
    return ""

def detect_row_type(row) -> str:
    if pd.isna(row["Site"]) or row["Site"] == "":
        return "Dept Total"
    if row["Site"] == "Total":
        return "Rayon Total"
    if row["Format"]:
        return "Site"
    return "Autre"

def parse_pbi_export(uploaded_file) -> pd.DataFrame:
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=0)
    raw = raw.iloc[:, : len(EXPORT_COLUMNS)]
    raw.columns = EXPORT_COLUMNS
    raw["Departement_rempli"] = raw["Departement"].ffill()
    raw["Rayon_rempli"] = raw["Rayon"].ffill()
    raw["Format"] = raw["Site"].apply(detect_format)
    raw["RowType"] = raw.apply(detect_row_type, axis=1)
    return raw

def extract_filtres(uploaded_file) -> str | None:
    """Cherche une ligne 'Filtres appliqués' dans l'export, comme le fait
    extract_periode() du module Reporting Ventes — purement informatif ici,
    affiché dans un expander (notre export ne contient pas de date absolue,
    contrairement au module Article)."""
    try:
        uploaded_file.seek(0)
        raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
        mask = raw[0].astype(str).str.startswith("Filtres", na=False)
        if mask.any():
            idx = mask.idxmax()
            lines = raw.iloc[idx:, 0].dropna().astype(str).tolist()
            return "\n".join(lines)[:500]
    except Exception:
        pass
    return None

# ============================================================
# DETECTION DE DATE (repli : nom de fichier, sinon aujourd'hui)
# ============================================================
DATE_PATTERN = re.compile(r"(\d{4})[-_](\d{2})[-_](\d{2})")

def guess_date_from_filename(filename: str) -> _dt.date:
    match = DATE_PATTERN.search(filename)
    if match:
        try:
            return _dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            pass
    return _dt.date.today()

# ============================================================
# CONSTRUCTION DU JOURNAL EN MEMOIRE
# ============================================================
def build_journal_from_uploads(files_with_dates: list) -> pd.DataFrame:
    frames = []
    for uploaded_file, chosen_date in files_with_dates:
        try:
            uploaded_file.seek(0)
            df = parse_pbi_export(uploaded_file)
        except Exception as exc:
            st.sidebar.error(f"{uploaded_file.name} : échec de lecture ({exc})")
            continue
        df["Date"] = chosen_date
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

# ============================================================
# CALCULS PARTAGES
# ============================================================
def pgc_and_pays_refs(journal: pd.DataFrame, target_date: _dt.date):
    day = journal[journal["Date"] == target_date]
    pgc = day[(day["Departement"] == "01 - PGC") & (day["Rayon"] == "Total")]
    pays = day[day["Departement"] == "Total"]
    return (pgc.iloc[0] if len(pgc) else None, pays.iloc[0] if len(pays) else None)

def rayon_total_row(journal: pd.DataFrame, target_date: _dt.date, rayon: str):
    day = journal[journal["Date"] == target_date]
    if rayon == "Tous PGC":
        row = day[(day["Departement"] == "01 - PGC") & (day["Rayon"] == "Total")]
    else:
        row = day[(day["Rayon"] == rayon) & (day["Site"] == "Total")]
    return row.iloc[0] if len(row) else None

def site_table(journal: pd.DataFrame, target_date: _dt.date, rayon: str) -> pd.DataFrame:
    day = journal[(journal["Date"] == target_date) & (journal["RowType"] == "Site")]
    total_pgc_ca = day["CA"].sum()
    if rayon != "Tous PGC":
        day = day[day["Rayon"] == rayon]
    if day.empty:
        return pd.DataFrame()

    agg = day.groupby("Site").agg(
        Format=("Format", "first"),
        CA_N1=("CA_N1", "sum"), CA=("CA", "sum"),
        Marge_N1=("Marge_N1", "sum"), Marge=("Marge", "sum"),
        Debit_N1=("Debit_N1", "sum"), Debit=("Debit", "sum"),
        Volume_N1=("Volume_N1", "sum"), Volume=("Volume", "sum"),
    ).reset_index()

    pgc_row, pays_row = pgc_and_pays_refs(journal, target_date)
    agg["CA_vs_N1"] = agg["CA"] / agg["CA_N1"] - 1
    agg["Ecart_Pays"] = agg["CA_vs_N1"] - pays_row["VsN1_pct"]
    agg["Ecart_PGC"] = agg["CA_vs_N1"] - pgc_row["VsN1_pct"]
    agg["Poids_CA"] = agg["CA"] / total_pgc_ca
    agg["Debit_vs_N1"] = agg["Debit"] / agg["Debit_N1"] - 1
    agg["Volume_vs_N1"] = agg["Volume"] / agg["Volume_N1"] - 1
    agg["Panier_vs_N1"] = (agg["CA"] / agg["Debit"]) / (agg["CA_N1"] / agg["Debit_N1"]) - 1
    agg["PanierQte_vs_N1"] = (agg["Volume"] / agg["Debit"]) / (agg["Volume_N1"] / agg["Debit_N1"]) - 1
    agg["TauxMarge"] = agg["Marge"] / agg["CA"]
    agg["TauxMarge_N1"] = agg["Marge_N1"] / agg["CA_N1"]
    agg["DeltaMarge"] = agg["TauxMarge"] - agg["TauxMarge_N1"]

    def flag(r):
        seuil = SEUILS_FORMAT.get(r["Format"], 0.03)
        worst = min(r["Ecart_Pays"], r["Ecart_PGC"])
        if worst < -seuil: return "Rouge"
        if worst < -seuil * 0.6: return "Orange"
        return "Vert"
    agg["Flag"] = agg.apply(flag, axis=1)

    def contact(r):
        if r["Flag"] == "Vert":
            return pd.Series(["", ""])
        if r["Debit_vs_N1"] < -0.05 and abs(r["Panier_vs_N1"]) < 0.03:
            return pd.Series(["Magasin", "Trafic en forte baisse"])
        if r["Volume_vs_N1"] < 0 and r["PanierQte_vs_N1"] < -0.03 and abs(r["Debit_vs_N1"]) < 0.03:
            return pd.Series(["Supply", "Volume/rupture, trafic stable (piste à vérifier)"])
        if r["DeltaMarge"] < -0.02 and r["CA_vs_N1"] > 0:
            return pd.Series(["Achat", "Marge brute en recul, CA en hausse"])
        cause = (
            f"Trafic {r['Debit_vs_N1']:+.0%}, panier {r['Panier_vs_N1']:+.0%}, "
            f"volume {r['Volume_vs_N1']:+.0%}, marge brute {r['DeltaMarge']*100:+.1f} pt "
            f"— pas de cause dominante, plusieurs facteurs à la fois"
        )
        return pd.Series(["À investiguer", cause])

    agg[["Qui_contacter", "Causes"]] = agg.apply(contact, axis=1)
    agg["Score"] = agg[["Ecart_Pays", "Ecart_PGC"]].min(axis=1).clip(upper=0).abs() * agg["Poids_CA"]
    agg = agg.sort_values("Score", ascending=False).reset_index(drop=True)
    agg.insert(0, "Priorite", range(1, len(agg) + 1))
    return agg

def build_headline(rayon_label, ca_n1_pct, ecart_pays, ecart_pgc, top_row):
    """Ligne de synthèse façon build_headline() du module Reporting Ventes."""
    line1 = (
        f"{rayon_label} — CA vs N-1 <b>{ca_n1_pct*100:+.1f}%</b> &nbsp;·&nbsp; "
        f"Écart vs Pays <b>{fmt_delta(ecart_pays*100)}</b> &nbsp;·&nbsp; "
        f"Écart vs PGC <b>{fmt_delta(ecart_pgc*100)}</b>"
    )
    line2 = ""
    if top_row is not None:
        line2 = (
            f"📌 magasin prioritaire : <b>{top_row['Site']}</b> "
            f"({fmt_delta(top_row['Ecart_Pays']*100)} vs Pays, "
            f"{top_row['Poids_CA']*100:.1f}% du CA pays)"
        )
    return line1, line2

# ============================================================
# EXPORT EXCEL — mise en forme conditionnelle façon build_excel_full
# ============================================================
def build_excel_mes_magasins(table: pd.DataFrame, rayon: str, d: _dt.date) -> bytes:
    from openpyxl import Workbook
    ARIAL = "Arial"
    thin = Side(style="thin", color="FFD1D1D6")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    BLUE_H, WHITE_H = "FF007AFF", "FFFFFFFF"

    export = table[[
        "Priorite", "Site", "Format", "CA_vs_N1", "Ecart_Pays", "Ecart_PGC",
        "Poids_CA", "Qui_contacter", "Causes",
    ]].rename(columns={
        "CA_vs_N1": "CA vs N-1", "Ecart_Pays": "Écart vs Pays", "Ecart_PGC": "Écart vs PGC",
        "Poids_CA": "Poids CA pays", "Qui_contacter": "Piste à vérifier",
    })
    export["Plan d'actions"] = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Mes magasins"
    ws.sheet_view.showGridLines = False

    headers = list(export.columns)
    for j, lbl in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=lbl)
        c.font = Font(name=ARIAL, bold=True, size=10, color=WHITE_H)
        c.fill = PatternFill("solid", fgColor=BLUE_H)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pct_cols = {"CA vs N-1", "Écart vs Pays", "Écart vs PGC"}
    for i, (_, row_) in enumerate(export.iterrows(), start=2):
        for j, col in enumerate(headers, start=1):
            v = row_[col]
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name=ARIAL, size=10)
            c.border = box
            c.alignment = Alignment(horizontal="left" if col in ("Site", "Causes", "Plan d'actions") else "center")
            if col in pct_cols:
                c.number_format = '+0.0%;-0.0%'
            elif col == "Poids CA pays":
                c.number_format = "0.0%"

    last_row = len(export) + 1
    for col_letter in ("E", "F"):  # Écart vs Pays, Écart vs PGC
        rng = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFFD6D4")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFD7F5DE")))

    widths = {"A": 8, "B": 26, "C": 10, "D": 12, "E": 13, "F": 13, "G": 13, "H": 14, "I": 50, "J": 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ============================================================
# INTERFACE
# ============================================================
st.markdown("<div class='page-title'>📈 Suivi Rayon vs Pays</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Comparaison de mon rayon à la Tendance Pays et au PGC · magasins à regarder priorisés · jusqu'à 6 historiques chargés en sidebar, aucune sauvegarde externe</div>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div class='section-label'>Import fichiers</div>", unsafe_allow_html=True)
    files = st.file_uploader(f"Exports PBI, jusqu'à {MAX_HISTORIQUES} (.xlsx)", type=["xlsx"],
                              accept_multiple_files=True, key="up_exports")
    st.caption("La date de chaque export est devinée depuis son nom de fichier "
               "(ex: export_2026-08-19.xlsx), à confirmer ci-dessous.")

    files_with_dates = []
    if files:
        if len(files) > MAX_HISTORIQUES:
            st.warning(f"{len(files)} fichiers sélectionnés — seuls les {MAX_HISTORIQUES} premiers sont pris en compte.")
            files = files[:MAX_HISTORIQUES]
        for f in files:
            guessed = guess_date_from_filename(f.name)
            confirmed = st.date_input(f.name, value=guessed, key=f"date_{f.name}")
            files_with_dates.append((f, confirmed))

    st.markdown("---")
    st.caption("SmartBuyer Hub · Module Suivi Rayon vs Pays")

if not files:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Compare la performance d'un rayon PGC à la <strong>Tendance Pays</strong> (tout le réseau)
  et au <strong>total PGC</strong>, et priorise les magasins à regarder — sans connexion
  externe, à partir de tes exports PBI chargés dans la barre latérale.
</div>
""", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<div class='section-label'>Contenu du module</div>", unsafe_allow_html=True)
        st.markdown("""
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>📅 Ma semaine</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    CA vs N-1, CA vs budget, marge brute, écart vs Tendance Pays et vs PGC — sur la
    date la plus récente chargée.
  </div>
</div>
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>📈 Ma tendance</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Courbe hebdomadaire du rayon vs la moyenne PGC, sur tout l'historique chargé
    (jusqu'à 6 semaines).
  </div>
</div>
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>🎯 Mes magasins à regarder</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Sites triés par priorité (ampleur de l'écart × poids CA), avec piste à
    vérifier et export Excel prêt pour COPIL.
  </div>
</div>
""", unsafe_allow_html=True)

    with c2:
        st.markdown("<div class='section-label'>Seuils de flag par format</div>", unsafe_allow_html=True)
        st.markdown(f"""
<div class='format-card format-hyper'>
  <span class='badge badge-hyper'>Hyper</span>
  <div class='small-muted' style='margin-top:6px'>Seuil : {SEUILS_FORMAT['Hyper']*100:.0f} pts d'écart</div>
</div>
<div class='format-card format-market'>
  <span class='badge badge-market'>Market</span>
  <div class='small-muted' style='margin-top:6px'>Seuil : {SEUILS_FORMAT['Market']*100:.0f} pts d'écart</div>
</div>
<div class='format-card format-supeco'>
  <span class='badge badge-supeco'>Supeco</span>
  <div class='small-muted' style='margin-top:6px'>Seuil : {SEUILS_FORMAT['Supeco']*100:.0f} pts d'écart</div>
</div>
""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Fonctionnement</div>", unsafe_allow_html=True)
        st.markdown(f"""
<div class='alert-card alert-green'>
  <strong>1.</strong> Charge jusqu'à {MAX_HISTORIQUES} exports PBI dans la sidebar.<br>
  <strong>2.</strong> Confirme la date devinée pour chaque fichier.<br>
  <strong>3.</strong> Consulte Ma semaine, Ma tendance et Mes magasins à regarder.<br>
  <strong>4.</strong> Télécharge la liste priorisée en Excel pour le COPIL.
</div>
""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Colonnes attendues dans l'export</div>", unsafe_allow_html=True)
    cols_expected = [
        ("Département / Rayon / Site", "Hiérarchie — obligatoire"),
        ("CA / CA N-1 / Budget", "Chiffre d'affaires, période, N-1, budget"),
        ("Marge / Taux de Marge", "Marge brute et taux, période et N-1"),
        ("Débit / Panier / Volume", "Trafic, panier moyen, volume vendu"),
    ]
    col_left, col_right = st.columns(2)
    for i, (name, desc) in enumerate(cols_expected):
        target = col_left if i % 2 == 0 else col_right
        with target:
            st.markdown(f"""
<div class='col-required'>
  <div style='font-size:16px'>▪️</div>
  <div>
    <div class='col-name'>{name}</div>
    <div class='col-desc'>{desc}</div>
  </div>
</div>
""", unsafe_allow_html=True)

    st.info("⬅️ Charge tes exports PBI dans la barre latérale pour démarrer.")
    st.stop()

journal = build_journal_from_uploads(files_with_dates)
if journal.empty:
    st.error("Aucun export n'a pu être lu — vérifie le format des fichiers chargés.")
    st.stop()

with st.expander("🔎 Périmètre détecté dans le premier fichier"):
    filtres = extract_filtres(files_with_dates[0][0])
    st.code(filtres if filtres else "Aucune ligne 'Filtres appliqués' trouvée dans ce fichier.", language=None)

tab1, tab2, tab3 = st.tabs(["📅 Ma semaine", "📈 Ma tendance", "🎯 Mes magasins à regarder"])

# ---------------- TAB 1 : MA SEMAINE ----------------
with tab1:
    rayon = st.selectbox("Rayon", RAYONS, key="week_rayon")
    d = journal["Date"].max()
    st.caption(f"Semaine du {d.strftime('%d/%m/%Y')}")

    row = rayon_total_row(journal, d, rayon)
    pgc_row, pays_row = pgc_and_pays_refs(journal, d)
    if row is None or pgc_row is None or pays_row is None:
        st.warning("Données manquantes pour ce rayon à cette date.")
    else:
        ca_n1, ca_bgt, marge = row["VsN1_pct"], row["VsBgt_pct"], row["TauxMarge"]
        ecart_pays = ca_n1 - pays_row["VsN1_pct"]
        ecart_pgc = ca_n1 - pgc_row["VsN1_pct"]

        table_preview = site_table(journal, d, rayon)
        top_row = table_preview.iloc[0] if not table_preview.empty and table_preview.iloc[0]["Flag"] != "" else None
        line1, line2 = build_headline(rayon, ca_n1, ecart_pays, ecart_pgc, top_row)
        st.markdown(f"<div class='alert-card alert-blue'><strong>{line1}</strong>"
                    + (f"<br>{line2}" if line2 else "") + "</div>", unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("CA vs N-1", f"{ca_n1*100:+.1f}%")
        c2.metric("CA vs budget", f"{ca_bgt*100:+.1f}%")
        c3.metric("Marge brute (taux)", fmt_pct(marge*100))
        c4, c5 = st.columns(2)
        c4.metric("Écart vs Tendance Pays", fmt_delta(ecart_pays*100))
        c5.metric("Écart vs PGC", fmt_delta(ecart_pgc*100))
        st.caption(
            f"Tendance Pays (tout réseau) : {pays_row['VsN1_pct']*100:+.1f}% vs N-1  ·  "
            f"PGC (total département) : {pgc_row['VsN1_pct']*100:+.1f}% vs N-1"
        )

# ---------------- TAB 2 : MA TENDANCE ----------------
with tab2:
    rayon_t = st.selectbox("Rayon", RAYONS, key="trend_rayon")
    dates = sorted(journal["Date"].unique())
    values, pgc_values = [], []
    for dt_ in dates:
        r = rayon_total_row(journal, dt_, rayon_t)
        pgc_r, _ = pgc_and_pays_refs(journal, dt_)
        values.append(r["VsN1_pct"] if r is not None else np.nan)
        pgc_values.append(pgc_r["VsN1_pct"] if pgc_r is not None else np.nan)
    trend_df = pd.DataFrame({"Mon rayon": values, "PGC (référence)": pgc_values},
                             index=pd.Index(dates, name="Date"))
    st.line_chart(trend_df)
    n = len(dates)
    if n < SEUIL_HISTORIQUE_SUFFISANT:
        st.caption(f"{n} semaine(s) chargée(s) — seuils encore provisoires. "
                   f"Cible : {SEUIL_HISTORIQUE_SUFFISANT} à {MAX_HISTORIQUES} semaines "
                   f"(le maximum chargeable d'un coup) pour un minimum de recul.")
    else:
        recent = pd.Series(values).dropna()
        st.caption(f"Calibrable sur {len(recent)} semaines — moyenne {recent.mean()*100:+.1f}%, "
                   f"écart-type {recent.std()*100:.1f} pts. Reste indicatif avec {MAX_HISTORIQUES} "
                   f"semaines maximum (13 et plus serait plus robuste statistiquement).")

# ---------------- TAB 3 : MES MAGASINS A REGARDER ----------------
with tab3:
    rayon_s = st.selectbox("Rayon", RAYONS, key="store_rayon")
    d3 = journal["Date"].max()
    table = site_table(journal, d3, rayon_s)
    if table.empty:
        st.info("Pas de données site pour ce rayon à cette date.")
    else:
        n_rouge = int((table["Flag"] == "Rouge").sum())
        n_orange = int((table["Flag"] == "Orange").sum())
        n_vert = int((table["Flag"] == "Vert").sum())
        m1, m2, m3 = st.columns(3)
        m1.metric("🔴 Rouge", n_rouge, delta_color="off")
        m2.metric("🟠 Orange", n_orange, delta_color="off")
        m3.metric("🟢 Vert", n_vert, delta_color="off")

        display = table[[
            "Priorite", "Site", "Format", "CA_vs_N1", "Ecart_Pays", "Ecart_PGC",
            "Poids_CA", "Qui_contacter", "Causes",
        ]].rename(columns={
            "CA_vs_N1": "CA vs N-1", "Ecart_Pays": "Écart vs Pays", "Ecart_PGC": "Écart vs PGC",
            "Poids_CA": "Poids CA pays", "Qui_contacter": "Piste à vérifier",
        })
        disp = display.copy()
        for c in ("CA vs N-1", "Écart vs Pays", "Écart vs PGC"):
            disp[c] = disp[c].map(lambda v: fmt_pct(v*100))
        disp["Poids CA pays"] = disp["Poids CA pays"].map(lambda v: fmt_pct(v*100))
        st.dataframe(disp, use_container_width=True, hide_index=True)
        st.caption("Piste à vérifier = hypothèse reconstruite à partir du volume et du panier, "
                   "pas une mesure de rupture réelle — à confirmer sur le terrain.")

        st.markdown("<div class='section-label'>Export</div>", unsafe_allow_html=True)
        xls = build_excel_mes_magasins(table, rayon_s, d3)
        filename = f"Mes magasins - {rayon_s.replace(' ', '_')} - {d3.strftime('%Y%m%d')}.xlsx"
        st.download_button(f"📥 Télécharger {filename}", xls, file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
