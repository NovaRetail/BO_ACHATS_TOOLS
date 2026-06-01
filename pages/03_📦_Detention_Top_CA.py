"""
03_📦_Detention_Top_CA.py — SmartBuyer Hub
Taux de détention Top CA · GOLD / SILVER · Articles Permanents
Source : Export PBI stock pivot (article × site) + Liste Top CA CSV
v3.1 — Corrections mapping TYPE · libellés · noms sites · export exécutif · couleurs management visuel
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Détention Top CA · SmartBuyer",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1300px; }
[data-testid="stSidebar"] { background: #FFFFFF !important; border-right: 0.5px solid #E5E5EA !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }
.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-gray  { background: #F9F9FB; border-color: #C7C7CC; color: #3A3A3C; }
.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.sc-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 10px; margin-bottom: 16px; }
.sc-card { background: #FFFFFF; border: 0.5px solid #E5E5EA; border-radius: 12px; padding: 14px 16px; position: relative; }
.sc-card.ok   { border-color: #6EE7B7; background: #F0FFF4; }
.sc-card.warn { border-color: #FCD34D; background: #FFFBF0; }
.sc-card.ko   { border-color: #FECACA; background: #FFF2F2; }
.sc-dot { width: 8px; height: 8px; border-radius: 50%; position: absolute; top: 14px; right: 14px; }
.sc-name { font-size: 11px; font-weight: 600; color: #1C1C1E; margin-bottom: 10px; max-width: 85%; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.sc-seg { display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px; font-size: 11px; }
.sc-seg-label { color: #8E8E93; display: flex; align-items: center; gap: 4px; }
.sc-seg-dot { width: 6px; height: 6px; border-radius: 50%; flex-shrink: 0; }
.sc-seg-right { display: flex; align-items: center; gap: 6px; }
.sc-seg-pct { font-weight: 600; }
.sc-seg-count { font-size: 10px; color: #8E8E93; }
.bar-track { height: 3px; background: #E5E5EA; border-radius: 2px; margin-bottom: 6px; }
.bar-fill { height: 3px; border-radius: 2px; }
.sc-divider { height: 0.5px; background: #E5E5EA; margin: 8px 0; }
.sc-total-row { display: flex; align-items: center; justify-content: space-between; }
.sc-total-label { font-size: 10px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: .04em; }
.sc-total-pct { font-size: 20px; font-weight: 700; }
.green-txt { color: #34C759; } .amber-txt { color: #FF9500; } .red-txt { color: #FF3B30; }
</style>
""", unsafe_allow_html=True)

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def norm_code(s):
    return s.astype(str).str.strip().str.replace(r"\.0$", "", regex=True).str.zfill(8)

def color_taux(v, cible):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "#8E8E93"
    return "#34C759" if v >= cible else "#FF9500" if v >= cible - 10 else "#FF3B30"

def cls_taux(v, cible):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "warn"
    return "ok" if v >= cible else "warn" if v >= cible - 10 else "ko"

def fmt_pct(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "—"
    return f"{v:.1f}%"

# ─── PARSERS ──────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_topca(byt, fname):
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(BytesIO(byt), sep=";", encoding=enc, dtype=str)
            df.columns = df.columns.str.strip().str.upper()
            # Colonnes
            code_col = next((c for c in df.columns if "CODE" in c), df.columns[0])
            type_col = next((c for c in df.columns if "TYPE" in c), None)
            lib_col  = next((c for c in df.columns if "LIB" in c), None)
            out = pd.DataFrame()
            out["code"] = norm_code(df[code_col])
            out["type"] = df[type_col].str.strip().str.upper() if type_col else "GOLD"
            out["lib"]  = df[lib_col].astype(str).str.strip() if lib_col else ""
            out = out[out["code"].str.match(r"^\d{8}$", na=False)].drop_duplicates("code")
            if len(out) > 0:
                return out
        except Exception:
            continue
    return pd.DataFrame(columns=["code", "type", "lib"])


@st.cache_data(show_spinner=False)
def load_pbi(byt, fname):
    """
    Parse l'export PBI pivot : Site nom court | 10202 - Palmeraie | ... | Total
    Retourne df articles, liste colonnes sites, dict {col → libellé exact PBI}
    """
    df_raw = pd.read_excel(BytesIO(byt), dtype=str)

    # Colonnes sites (tout sauf 'Site nom court' et 'Total')
    site_cols = [c for c in df_raw.columns if c not in ["Site nom court", "Total"]]

    # Noms de sites : libellé exact après " - " (ex: "10202 - Palmeraie" → "Palmeraie")
    sites_info = {}
    for col in site_cols:
        if " - " in str(col):
            sites_info[col] = str(col).split(" - ", 1)[1].strip()
        else:
            sites_info[col] = str(col).strip()

    # Lignes articles valides
    mask = (
        df_raw["Site nom court"].notna() &
        (~df_raw["Site nom court"].astype(str).str.strip().isin(
            ["Article", "Total", "nan", ""])) &
        (~df_raw["Site nom court"].astype(str).str.startswith("Filtres"))
    )
    df = df_raw[mask].copy().reset_index(drop=True)

    # Code article (avant " - ")
    df["code"] = df["Site nom court"].apply(
        lambda s: str(s).split(" - ", 1)[0].strip().zfill(8)
        if " - " in str(s) else str(s).strip().zfill(8)
    )
    # Libellé PBI (après " - ")
    df["libelle_pbi"] = df["Site nom court"].apply(
        lambda s: str(s).split(" - ", 1)[1].strip()
        if " - " in str(s) else str(s).strip()
    )

    # Convertir stocks
    for col in site_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df, site_cols, sites_info


def compute(df_pbi, site_cols, sites_info, df_topca, type_filtre, seuil_reappro):
    # Filtrer Top CA par type
    topca_f = df_topca[df_topca["type"] == type_filtre].copy() \
              if type_filtre != "Tous" else df_topca.copy()

    top_codes = set(topca_f["code"].unique())
    top_meta  = topca_f.drop_duplicates("code").set_index("code")[["type", "lib"]]

    # Mapping directs — correction bug
    type_map = top_meta["type"].to_dict()
    lib_map  = top_meta["lib"].to_dict()

    # Filtrer PBI
    df = df_pbi[df_pbi["code"].isin(top_codes)].copy()

    # Melt article × site × stock
    df_long = df.melt(
        id_vars=["code", "libelle_pbi"],
        value_vars=site_cols,
        var_name="site_col",
        value_name="stock"
    )
    df_long["site"]      = df_long["site_col"].map(sites_info)
    df_long["type"]      = df_long["code"].map(type_map).fillna("?")
    df_long["lib_topca"] = df_long["code"].map(lib_map).fillna("")

    # Libellé : priorité Top CA, sinon PBI
    df_long["libelle"] = df_long.apply(
        lambda r: r["lib_topca"] if r["lib_topca"] else r["libelle_pbi"], axis=1
    )

    # Calculs
    df_long["detenu"]         = df_long["stock"].notna() & (df_long["stock"] > 0)
    df_long["reappro_urgent"] = (
        df_long["stock"].notna() &
        (df_long["stock"] > 0) &
        (df_long["stock"] < seuil_reappro)
    )

    # Absents PBI
    absents = topca_f[~topca_f["code"].isin(df_pbi["code"].unique())].copy()

    return df_long, absents, top_codes, top_meta


def build_alertes(df_long, site_cols, seuil_reappro):
    # Agréger par article
    agg = df_long.groupby(["code", "libelle", "type"]).agg(
        nb_sites_detenu=("detenu", "sum"),
        nb_sites_total=("detenu", "count"),
        stock_total=("stock", lambda x: x.fillna(0).sum()),
    ).reset_index()
    agg["nb_sites_detenu"] = agg["nb_sites_detenu"].astype(int)

    # 1. Ruptures nettes
    ruptures = agg[agg["nb_sites_detenu"] == 0].sort_values(
        ["type", "libelle"]).reset_index(drop=True)

    # 2. Réappro urgent
    reappro = df_long[df_long["reappro_urgent"]][
        ["code", "libelle", "type", "site", "stock"]
    ].sort_values(["type", "stock"]).reset_index(drop=True)

    # 3. Ruptures partielles
    partielles = agg[
        (agg["nb_sites_detenu"] > 0) &
        (agg["nb_sites_detenu"] / agg["nb_sites_total"] < 0.5)
    ].copy()

    def sites_manquants(code):
        rows = df_long[(df_long["code"] == code) & (~df_long["detenu"])]
        return ", ".join(sorted(rows["site"].dropna().unique().tolist()))

    if not partielles.empty:
        partielles["sites_manquants"] = partielles["code"].apply(sites_manquants)
        partielles["taux_detenu"] = (
            partielles["nb_sites_detenu"] / partielles["nb_sites_total"] * 100
        ).round(1)
        partielles = partielles.sort_values(["type", "nb_sites_detenu"]).reset_index(drop=True)

    return ruptures, reappro, partielles


def compute_taux(df_long, types_seg, cible):
    rows = []
    for site in sorted(df_long["site"].dropna().unique()):
        s = df_long[df_long["site"] == site]
        row = {"site": site}
        for t in types_seg:
            st = s[s["type"] == t]
            n  = len(st); d = int(st["detenu"].sum())
            row[f"n_{t}"]    = n
            row[f"det_{t}"]  = d
            row[f"taux_{t}"] = round(d / n * 100, 1) if n > 0 else None
        n_tot = len(s); d_tot = int(s["detenu"].sum())
        row["n_total"]    = n_tot
        row["det_total"]  = d_tot
        row["taux_total"] = round(d_tot / n_tot * 100, 1) if n_tot > 0 else None
        rows.append(row)
    return pd.DataFrame(rows)


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def gen_excel(taux_df, types_seg, ruptures, reappro, partielles,
              absents, type_filtre, cible, seuil_reappro):
    wb = Workbook()

    # Couleurs management visuel
    C_HDR  = "1C3557"   # Entête foncé
    C_SUB  = "2E4B7A"   # Sous-entête
    C_OK   = "D6F0D6"   # Vert — taux ≥ cible
    C_WARN = "FEF3CD"   # Orange — vigilance
    C_KO   = "FCE4E4"   # Rouge — action
    C_GOLD = "FFFBF0"   # Doré — lignes GOLD
    C_ODD  = "F7F7F7"   # Alternance grise
    C_EVN  = "FFFFFF"   # Blanc

    def f(h): return PatternFill("solid", fgColor=h)
    def bdr():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LFT = Alignment(horizontal="left",   vertical="center")
    HF  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    BF  = Font(bold=True, name="Calibri", size=10, color="1C1C1E")
    NF  = Font(name="Calibri", size=10, color="1C1C1E")

    def header_row(ws, headers, widths, bg=C_HDR):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = f(bg); c.font = HF; c.alignment = CTR; c.border = bdr()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def color_cell_taux(c, val):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return
        pct = val  # déjà en %
        c.fill = f(C_OK) if pct >= cible else f(C_WARN) if pct >= cible - 10 else f(C_KO)
        c.font = Font(bold=True, name="Calibri", size=10,
                      color="1C1C1E")

    def row_fill(ri, is_gold):
        if is_gold: return f(C_GOLD)
        return f(C_ODD) if ri % 2 == 0 else f(C_EVN)

    # ── Feuille 1 : Synthèse réseau ──────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Synthèse réseau"

    # ── Calculs zone exécutive ───────────────────────────────────────────────
    taux_moy   = taux_df["taux_total"].mean()
    n_rupt     = len(ruptures)
    n_reappro  = len(reappro)
    n_part     = len(partielles)
    n_absents  = len(absents)
    n_sites    = len(taux_df)
    sites_ko   = int((taux_df["taux_total"] < cible).sum())
    n_cols_tot = 4 + len(types_seg) * 3 + 1  # nb colonnes du tableau

    # Nb colonnes pour les merges (au moins 6)
    span = max(n_cols_tot, 6)

    # ── Ligne 1 : Titre ───────────────────────────────────────────────────────
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ct = ws1.cell(row=1, column=1)
    ct.value = f"SYNTHÈSE DÉTENTION TOP CA — {type_filtre} · Cible : {cible}%"
    ct.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    ct.fill = f(C_HDR); ct.alignment = CTR
    ws1.row_dimensions[1].height = 30

    # ── Ligne 2 : Label section KPIs ─────────────────────────────────────────
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    cl2 = ws1.cell(row=2, column=1, value="  INDICATEURS RÉSEAU")
    cl2.font = Font(bold=True, color="AABBCC", name="Calibri", size=9, italic=True)
    cl2.fill = f(C_HDR); cl2.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 16

    # ── Lignes 3-4 : KPIs réseau (2 rangées : labels + valeurs) ──────────────
    kpi_labels = ["Taux réseau moyen", "Sites sous cible",
                  "Ruptures nettes", "Réappro urgents",
                  "Ruptures partielles", "Absents PBI"]
    kpi_values = [
        f"{taux_moy:.1f}%" if taux_moy else "—",
        f"{sites_ko} / {n_sites}",
        str(n_rupt), str(n_reappro), str(n_part), str(n_absents)
    ]
    kpi_colors = [
        C_OK if (taux_moy or 0) >= cible else C_WARN if (taux_moy or 0) >= cible - 10 else C_KO,
        C_KO if sites_ko > 0 else C_OK,
        C_KO if n_rupt > 0 else C_OK,
        C_WARN if n_reappro > 0 else C_OK,
        C_WARN if n_part > 0 else C_OK,
        C_WARN if n_absents > 0 else C_OK,
    ]
    for i, (lbl, val, col) in enumerate(zip(kpi_labels, kpi_values, kpi_colors), 1):
        cl = ws1.cell(row=3, column=i, value=lbl)
        cl.fill = f(C_SUB); cl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cl.alignment = CTR; cl.border = bdr()
        cv = ws1.cell(row=4, column=i, value=val)
        cv.fill = f(col); cv.font = Font(bold=True, name="Calibri", size=13, color="1C1C1E")
        cv.alignment = CTR; cv.border = bdr()
        ws1.column_dimensions[get_column_letter(i)].width = 22
    ws1.row_dimensions[3].height = 18; ws1.row_dimensions[4].height = 30

    # ── Ligne 5 : Label section répartition refs ──────────────────────────────
    ws1.merge_cells(start_row=5, start_column=1, end_row=5, end_column=span)
    cl5 = ws1.cell(row=5, column=1, value="  RÉPARTITION DES RÉFÉRENCES · TAUX MOYEN RÉSEAU")
    cl5.font = Font(bold=True, color="AABBCC", name="Calibri", size=9, italic=True)
    cl5.fill = f(C_HDR); cl5.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[5].height = 16

    # ── Lignes 6-7 : Répartition par type (labels + valeurs) ─────────────────
    # 3 colonnes par type : Nb réf | Détenus réseau | Taux moyen
    ref_headers = []
    ref_values  = []
    ref_colors  = []
    for t in types_seg:
        n_refs_t   = int(taux_df[f"n_{t}"].iloc[0]) if f"n_{t}" in taux_df.columns and len(taux_df) > 0 else 0
        taux_t_moy = taux_df[f"taux_{t}"].mean() if f"taux_{t}" in taux_df.columns else None
        det_t_tot  = int(taux_df[f"det_{t}"].sum()) if f"det_{t}" in taux_df.columns else 0
        ecart_t    = (taux_t_moy or 0) - cible
        ecart_str  = f"{ecart_t:+.1f} pt vs cible"
        col_t      = C_OK if (taux_t_moy or 0) >= cible else C_WARN if (taux_t_moy or 0) >= cible - 10 else C_KO
        ref_headers += [f"Réf {t}", f"Détenus {t} (total réseau)", f"Taux moy. {t}"]
        ref_values  += [str(n_refs_t), str(det_t_tot), f"{taux_t_moy:.1f}%" if taux_t_moy else "—"]
        ref_colors  += [C_GOLD if t == "GOLD" else "E8E8F0", C_GOLD if t == "GOLD" else "E8E8F0", col_t]
    # Ajouter total
    n_refs_tot = int(taux_df["n_total"].iloc[0]) if len(taux_df) > 0 else 0
    det_tot    = int(taux_df["det_total"].sum())
    ref_headers += ["Réf TOTAL", "Détenus (total réseau)", "Taux moy. TOTAL"]
    ref_values  += [str(n_refs_tot), str(det_tot), f"{taux_moy:.1f}%" if taux_moy else "—"]
    col_tot     = C_OK if (taux_moy or 0) >= cible else C_WARN if (taux_moy or 0) >= cible - 10 else C_KO
    ref_colors  += ["DBEAFE", "DBEAFE", col_tot]

    for i, (lbl, val, col) in enumerate(zip(ref_headers, ref_values, ref_colors), 1):
        cl = ws1.cell(row=6, column=i, value=lbl)
        cl.fill = f(C_SUB); cl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cl.alignment = CTR; cl.border = bdr()
        cv = ws1.cell(row=7, column=i, value=val)
        cv.fill = f(col); cv.font = Font(bold=True, name="Calibri", size=12, color="1C1C1E")
        cv.alignment = CTR; cv.border = bdr()
    ws1.row_dimensions[6].height = 18; ws1.row_dimensions[7].height = 28

    # ── Ligne 8 : Séparateur ─────────────────────────────────────────────────
    ws1.row_dimensions[8].height = 8

    # ── Ligne 9 : Entête tableau ──────────────────────────────────────────────
    headers = ["Magasin", "Réf total", "Détenus", "Taux total %"]
    widths  = [22, 12, 12, 14]
    for t in types_seg:
        headers += [f"Réf {t}", f"Détenus {t}", f"Taux {t} %"]
        widths  += [12, 12, 14]
    headers.append("Statut"); widths.append(14)

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws1.cell(row=9, column=i, value=h)
        c.fill = f(C_SUB); c.font = HF; c.alignment = CTR; c.border = bdr()
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[9].height = 22
    ws1.freeze_panes = "A10"

    # Données
    for ri, (_, row) in enumerate(taux_df.sort_values("taux_total").iterrows(), 10):
        tv = row["taux_total"]
        bg = f(C_ODD) if ri % 2 == 0 else f(C_EVN)

        vals = [row["site"], row["n_total"], row["det_total"], tv]
        for t in types_seg:
            vals += [row.get(f"n_{t}", 0), row.get(f"det_{t}", 0), row.get(f"taux_{t}")]
        statut = "✅ OK" if (tv or 0) >= cible else "⚠️ Surveiller" if (tv or 0) >= cible - 10 else "🔴 Action"
        vals.append(statut)

        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.fill = bg; c.font = NF; c.border = bdr()
            c.alignment = CTR if isinstance(val, (int, float)) else LFT
            if isinstance(val, float) and val is not None:
                c.number_format = "0.0"
            # Colorier les colonnes taux
            taux_cols = [4] + [4 + 3 + (j * 3) for j in range(len(types_seg))]
            if ci in taux_cols and val is not None:
                color_cell_taux(c, val)
        ws1.row_dimensions[ri].height = 18

    # ── Feuille 2 : Ruptures nettes ───────────────────────────────────────────
    ws2 = wb.create_sheet("Ruptures nettes")
    header_row(ws2,
        ["Code article", "Libellé article", "Type",
         "Sites détenus", "Sites total", "Action"],
        [14, 42, 10, 14, 12, 28])

    for ri, (_, row) in enumerate(ruptures.iterrows(), 2):
        is_gold = row["type"] == "GOLD"
        bg = row_fill(ri, is_gold)
        vals = [row["code"], row["libelle"], row["type"],
                int(row["nb_sites_detenu"]), int(row["nb_sites_total"]),
                "Commander en urgence"]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.fill = bg; c.font = NF; c.border = bdr()
            c.alignment = CTR if ci in [4, 5] else LFT
        ws2.row_dimensions[ri].height = 18

    # ── Feuille 3 : Réappro urgent ────────────────────────────────────────────
    ws3 = wb.create_sheet("Réappro urgent")
    header_row(ws3,
        ["Code article", "Libellé article", "Type",
         "Magasin", "Stock actuel", f"Seuil (< {seuil_reappro})", "Action"],
        [14, 42, 10, 22, 14, 14, 26])

    for ri, (_, row) in enumerate(reappro.iterrows(), 2):
        is_gold = row["type"] == "GOLD"
        bg = row_fill(ri, is_gold)
        stock_val = int(row["stock"]) if pd.notna(row["stock"]) else 0
        # Colorer stock selon urgence
        vals = [row["code"], row["libelle"], row["type"],
                row["site"], stock_val, f"< {seuil_reappro}", "Réassort immédiat"]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill = bg; c.font = NF; c.border = bdr()
            c.alignment = CTR if ci in [5] else LFT
            # Stock : rouge si = 1, orange si = 2
            if ci == 5 and isinstance(val, int):
                if val == 1:   c.fill = f(C_KO);  c.font = Font(bold=True, name="Calibri", size=10, color="1C1C1E")
                elif val == 2: c.fill = f(C_WARN); c.font = Font(bold=True, name="Calibri", size=10, color="1C1C1E")
        ws3.row_dimensions[ri].height = 18

    # ── Feuille 4 : Ruptures partielles ───────────────────────────────────────
    ws4 = wb.create_sheet("Ruptures partielles")
    header_row(ws4,
        ["Code article", "Libellé article", "Type",
         "Sites détenus", "Sites total", "Taux %", "Sites manquants"],
        [14, 42, 10, 14, 12, 10, 60])

    if not partielles.empty:
        for ri, (_, row) in enumerate(partielles.iterrows(), 2):
            is_gold = row["type"] == "GOLD"
            bg = row_fill(ri, is_gold)
            tv = row.get("taux_detenu")
            vals = [row["code"], row["libelle"], row["type"],
                    int(row["nb_sites_detenu"]), int(row["nb_sites_total"]),
                    tv, row.get("sites_manquants", "")]
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(row=ri, column=ci, value=val)
                c.fill = bg; c.font = NF; c.border = bdr()
                c.alignment = CTR if ci in [4, 5, 6] else LFT
                if ci == 6 and val is not None:
                    c.number_format = "0.0"
                    color_cell_taux(c, val)
            ws4.row_dimensions[ri].height = 18

    # ── Feuille 5 : Absents PBI ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Absents PBI")
    header_row(ws5,
        ["Code article", "Libellé", "Type", "Vérification"],
        [14, 42, 10, 40])

    for ri, (_, row) in enumerate(absents.iterrows(), 2):
        is_gold = row["type"] == "GOLD"
        bg = row_fill(ri, is_gold)
        vals = [row["code"], row["lib"], row["type"],
                "Vérifier déréférencement ou code article"]
        for ci, val in enumerate(vals, 1):
            c = ws5.cell(row=ri, column=ci, value=val)
            c.fill = bg; c.font = NF; c.border = bdr()
            c.alignment = LFT
        ws5.row_dimensions[ri].height = 18

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichiers</div>", unsafe_allow_html=True)
    st.markdown("**Liste Top CA** *(CSV · CODE ARTICLE · TYPE)*")
    f_topca = st.file_uploader("Top CA", type=["csv"], key="topca", label_visibility="collapsed")
    st.markdown("**Export PBI stock** *(Excel pivot article × site)*")
    f_pbi   = st.file_uploader("PBI",    type=["xlsx", "xls"], key="pbi", label_visibility="collapsed")
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Paramètres</div>", unsafe_allow_html=True)
    cible         = st.slider("Cible taux de détention (%)", 70, 100, 85, 1)
    seuil_reappro = st.number_input("Seuil réappro urgent (stock <)", 1, 50, 3, 1)


# ─── PAGE ─────────────────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📦 Détention Top CA</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Articles permanents · GOLD / SILVER · Taux de détention · Ruptures · Réappro urgent</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL ──────────────────────────────────────────────────────────
if not f_topca or not f_pbi:
    st.markdown("---")
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Vérifie la présence en magasin des articles Top CA et calcule le taux de détention
  par type (GOLD / SILVER) et par site. Détecte les ruptures et anticipe les réapprovisionnements urgents.
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Fichiers attendus</div>", unsafe_allow_html=True)
    cf1, cf2 = st.columns(2)
    with cf1:
        st.markdown("""
<div class='col-required'><div style='font-size:16px'>📋</div>
<div><div class='col-name'>Liste Top CA (.csv)</div>
<div class='col-desc'>Colonnes : CODE ARTICLE · LIBELLÉ ARTICLE · TYPE</div>
<div class='col-desc' style='color:#8E8E93;font-size:11px;margin-top:2px'>TYPE flexible : GOLD / SILVER / toute valeur</div>
</div></div>""", unsafe_allow_html=True)
    with cf2:
        st.markdown("""
<div class='col-required'><div style='font-size:16px'>📊</div>
<div><div class='col-name'>Export PBI stock (.xlsx)</div>
<div class='col-desc'>Pivot article × site · colonne "Site nom court"</div>
<div class='col-desc' style='color:#8E8E93;font-size:11px;margin-top:2px'>Valeur = stock · NaN = absent du site</div>
</div></div>""", unsafe_allow_html=True)
    st.info("⬆️ Charge les deux fichiers dans la sidebar pour lancer l'analyse.")
    st.stop()

# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
with st.spinner("Lecture des fichiers…"):
    df_topca              = load_topca(f_topca.read(), f_topca.name)
    df_pbi, site_cols, sites_info = load_pbi(f_pbi.read(), f_pbi.name)

if df_topca.empty: st.error("Liste Top CA vide ou non reconnue."); st.stop()
if df_pbi.empty:   st.error("Export PBI vide ou non reconnu.");    st.stop()

# ─── FILTRE TYPE SIDEBAR ──────────────────────────────────────────────────────
types_dispo_all = sorted(df_topca["type"].unique())
with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtre type</div>", unsafe_allow_html=True)
    type_filtre = st.radio("Type", ["Tous"] + types_dispo_all,
                            horizontal=False, label_visibility="collapsed")

# ─── CALCUL ───────────────────────────────────────────────────────────────────
with st.spinner("Calcul des taux de détention…"):
    df_long, absents, top_codes, top_meta = compute(
        df_pbi, site_cols, sites_info, df_topca, type_filtre, seuil_reappro)
    ruptures, reappro, partielles = build_alertes(df_long, site_cols, seuil_reappro)
    types_seg = sorted(df_long[df_long["type"] != "?"]["type"].unique())
    taux_df   = compute_taux(df_long, types_seg, cible)

# Métriques
n_sites       = len(site_cols)
n_refs        = df_long["code"].nunique()
taux_moy      = taux_df["taux_total"].mean()
n_rupt        = len(ruptures)
n_reappro_al  = len(reappro)
n_partielles  = len(partielles)
n_absents     = len(absents)

taux_par_type = {t: taux_df[f"taux_{t}"].mean()
                 for t in types_seg if f"taux_{t}" in taux_df.columns}

# ─── KPIs ─────────────────────────────────────────────────────────────────────
st.markdown(f"<div class='section-label'>{n_sites} sites · {n_refs} références · {type_filtre}</div>",
            unsafe_allow_html=True)

kpi_cols = st.columns(2 + len(types_seg))
kpi_cols[0].metric("Réf analysées", str(n_refs),
                    f"{n_absents} absents PBI" if n_absents else "")
kpi_cols[1].metric("Taux réseau",   fmt_pct(taux_moy),
                    f"{(taux_moy or 0) - cible:+.1f} pt vs cible" if taux_moy else "")
for i, t in enumerate(types_seg):
    tv = taux_par_type.get(t)
    kpi_cols[2 + i].metric(f"Taux {t}", fmt_pct(tv),
                            f"{(tv or 0) - cible:+.1f} pt" if tv else "")

# ─── ALERTES ──────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Alertes</div>", unsafe_allow_html=True)

if n_rupt > 0:
    nb_gold_r = int((ruptures["type"] == "GOLD").sum())
    st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>🔴 {n_rupt} rupture(s) nette(s)</strong> — stock = 0 sur tous les sites
  · {nb_gold_r} GOLD · {n_rupt - nb_gold_r} autres<br>
  <span style='font-size:12px;opacity:.85'>→ Commander en urgence.</span>
</div>""", unsafe_allow_html=True)

if n_reappro_al > 0:
    nb_gold_ra = int((reappro["type"] == "GOLD").sum())
    st.markdown(f"""
<div class='alert-card alert-blue'>
  <strong>🔵 {n_reappro_al} ligne(s) en réappro urgent</strong>
  — 0 &lt; stock &lt; {seuil_reappro} · {nb_gold_ra} GOLD<br>
  <span style='font-size:12px;opacity:.85'>→ Réassort immédiat avant épuisement.</span>
</div>""", unsafe_allow_html=True)

if n_partielles > 0:
    nb_gold_p = int((partielles["type"] == "GOLD").sum())
    st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>🟡 {n_partielles} rupture(s) partielle(s)</strong>
  — détenu sur &lt; 50% des sites · {nb_gold_p} GOLD<br>
  <span style='font-size:12px;opacity:.85'>→ Redistribuer ou approvisionner les sites manquants.</span>
</div>""", unsafe_allow_html=True)

if n_absents > 0:
    st.markdown(f"""
<div class='alert-card alert-gray'>
  <strong>⚪ {n_absents} article(s) Top CA absents du PBI</strong><br>
  <span style='font-size:12px;opacity:.85'>→ Vérifier déréférencement ou code article.</span>
</div>""", unsafe_allow_html=True)

if n_rupt == 0 and n_reappro_al == 0 and n_partielles == 0:
    st.markdown("<div class='alert-card alert-green'>✅ Aucune alerte — tous les articles au-dessus des seuils.</div>",
                unsafe_allow_html=True)

# ─── SCORECARD ────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Scorecard magasins</div>", unsafe_allow_html=True)

TYPE_DOT_COLORS = {"GOLD": "#FF9500", "SILVER": "#8E8E93"}
def type_dot(t): return TYPE_DOT_COLORS.get(t, "#007AFF")

sc_html = '<div class="sc-grid">'
for _, row in taux_df.sort_values("taux_total").iterrows():
    tv  = row["taux_total"]
    cls = cls_taux(tv, cible)
    dot_c = "#34C759" if cls == "ok" else "#FF9500" if cls == "warn" else "#FF3B30"
    txt_c = "green-txt" if cls == "ok" else "amber-txt" if cls == "warn" else "red-txt"

    segs = ""
    for t in types_seg:
        tv_t  = row.get(f"taux_{t}")
        det_t = int(row.get(f"det_{t}", 0))
        n_t   = int(row.get(f"n_{t}", 0))
        pct_t = tv_t if (tv_t and not np.isnan(tv_t)) else 0
        tc    = type_dot(t)
        tc_txt = "green-txt" if pct_t >= cible else "amber-txt" if pct_t >= cible - 10 else "red-txt"
        segs += f"""
<div class='sc-seg'>
  <div class='sc-seg-label'>
    <div class='sc-seg-dot' style='background:{tc}'></div>{t}
  </div>
  <div class='sc-seg-right'>
    <span class='sc-seg-pct {tc_txt}'>{fmt_pct(tv_t)}</span>
    <span class='sc-seg-count'>{det_t}/{n_t}</span>
  </div>
</div>
<div class='bar-track'><div class='bar-fill' style='width:{min(pct_t,100):.0f}%;background:{tc}'></div></div>"""

    sc_html += f"""
<div class='sc-card {cls}'>
  <div class='sc-dot' style='background:{dot_c}'></div>
  <div class='sc-name'>{row['site']}</div>
  {segs}
  <div class='sc-divider'></div>
  <div class='sc-total-row'>
    <span class='sc-total-label'>Total</span>
    <span class='sc-total-pct {txt_c}'>{fmt_pct(tv)}</span>
  </div>
  <div class='bar-track'><div class='bar-fill' style='width:{min(tv or 0,100):.0f}%;background:{dot_c}'></div></div>
</div>"""
sc_html += "</div>"
st.markdown(sc_html, unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Synthèse réseau",
    "🚨 Alertes détaillées",
    "🔍 Détail article × site",
    "🚫 Absents PBI",
    "📥 Export Excel",
])

# ══ TAB 1 ═════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown("<div class='section-label'>Taux de détention par magasin et par type</div>",
                unsafe_allow_html=True)
    disp1 = taux_df[["site", "n_total", "det_total", "taux_total"]].copy()
    disp1.columns = ["Magasin", "Réf", "Détenus", "Taux total %"]
    for t in types_seg:
        disp1[f"Réf {t}"]     = taux_df[f"n_{t}"].astype(int)
        disp1[f"Détenus {t}"] = taux_df[f"det_{t}"].astype(int)
        disp1[f"Taux {t} %"]  = taux_df[f"taux_{t}"].apply(fmt_pct)
    disp1["Taux total %"] = disp1["Taux total %"].apply(fmt_pct)
    disp1["Statut"] = taux_df["taux_total"].apply(
        lambda v: "🟢 OK" if (v or 0) >= cible
        else "🟡 Surveiller" if (v or 0) >= cible - 10 else "🔴 Action")
    st.dataframe(disp1.sort_values("Taux total %"), use_container_width=True, hide_index=True)

# ══ TAB 2 ═════════════════════════════════════════════════════════════════════
with tab2:
    al_sel = st.multiselect("Afficher",
        ["🔴 Ruptures nettes", "🔵 Réappro urgent",
         "🟡 Ruptures partielles", "⚪ Absents PBI"],
        default=["🔴 Ruptures nettes", "🔵 Réappro urgent",
                 "🟡 Ruptures partielles", "⚪ Absents PBI"])

    if "🔴 Ruptures nettes" in al_sel and not ruptures.empty:
        st.markdown(f"<div class='section-label'>🔴 Ruptures nettes — {len(ruptures)} article(s)</div>",
                    unsafe_allow_html=True)
        d = ruptures[["code", "libelle", "type", "nb_sites_detenu", "nb_sites_total"]].copy()
        d.columns = ["Code", "Libellé", "Type", "Sites détenus", "Sites total"]
        st.dataframe(d, use_container_width=True, hide_index=True)

    if "🔵 Réappro urgent" in al_sel and not reappro.empty:
        st.markdown(f"<div class='section-label'>🔵 Réappro urgent — {len(reappro)} ligne(s)</div>",
                    unsafe_allow_html=True)
        d = reappro[["code", "libelle", "type", "site", "stock"]].copy()
        d.columns = ["Code", "Libellé", "Type", "Magasin", "Stock actuel"]
        d["Stock actuel"] = d["Stock actuel"].apply(
            lambda x: int(x) if pd.notna(x) else 0)
        st.dataframe(d, use_container_width=True, hide_index=True)

    if "🟡 Ruptures partielles" in al_sel and not partielles.empty:
        st.markdown(f"<div class='section-label'>🟡 Ruptures partielles — {len(partielles)} article(s)</div>",
                    unsafe_allow_html=True)
        d = partielles[["code", "libelle", "type",
                         "nb_sites_detenu", "nb_sites_total",
                         "taux_detenu", "sites_manquants"]].copy()
        d.columns = ["Code", "Libellé", "Type",
                     "Sites détenus", "Sites total", "Taux %", "Sites manquants"]
        d["Taux %"] = d["Taux %"].apply(fmt_pct)
        st.dataframe(d, use_container_width=True, hide_index=True,
            column_config={"Sites manquants": st.column_config.TextColumn(
                "Sites manquants", width="large")})

    if "⚪ Absents PBI" in al_sel and not absents.empty:
        st.markdown(f"<div class='section-label'>⚪ Absents PBI — {len(absents)} article(s)</div>",
                    unsafe_allow_html=True)
        d = absents[["code", "lib", "type"]].copy()
        d.columns = ["Code", "Libellé", "Type"]
        st.dataframe(d, use_container_width=True, hide_index=True)

# ══ TAB 3 ═════════════════════════════════════════════════════════════════════
with tab3:
    fc1, fc2, fc3 = st.columns(3)
    with fc1: sel_type_d = st.selectbox("Type",    ["Tous"] + types_seg,   key="d_t")
    with fc2: sel_site_d = st.selectbox("Magasin", ["Tous"] + sorted(
        df_long["site"].dropna().unique().tolist()), key="d_s")
    with fc3: sel_al_d   = st.selectbox("Statut",
        ["Tous", "Détenu", "Non détenu", "Réappro urgent"], key="d_a")

    df_det = df_long.copy()
    if sel_type_d != "Tous":   df_det = df_det[df_det["type"] == sel_type_d]
    if sel_site_d != "Tous":   df_det = df_det[df_det["site"] == sel_site_d]
    if sel_al_d == "Détenu":            df_det = df_det[df_det["detenu"]]
    elif sel_al_d == "Non détenu":      df_det = df_det[~df_det["detenu"]]
    elif sel_al_d == "Réappro urgent":  df_det = df_det[df_det["reappro_urgent"]]

    disp = df_det[["code", "libelle", "type", "site", "stock", "detenu", "reappro_urgent"]].copy()
    disp.columns = ["Code", "Libellé", "Type", "Magasin", "Stock", "Détenu", "Réappro"]
    disp["Stock"]   = disp["Stock"].apply(lambda x: int(x) if pd.notna(x) else "—")
    disp["Détenu"]  = disp["Détenu"].map({True: "✅", False: "❌"})
    disp["Réappro"] = disp["Réappro"].map({True: "🔵", False: ""})
    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>{len(disp):,} lignes</div>",
                unsafe_allow_html=True)
    st.dataframe(disp, use_container_width=True, hide_index=True)

# ══ TAB 4 ═════════════════════════════════════════════════════════════════════
with tab4:
    if absents.empty:
        st.markdown("<div class='alert-card alert-green'>✅ Tous les articles Top CA sont présents dans l'export PBI.</div>",
                    unsafe_allow_html=True)
    else:
        st.warning(f"⚠️ {len(absents)} référence(s) Top CA absentes du fichier PBI.")
        d = absents[["code", "lib", "type"]].copy()
        d.columns = ["Code article", "Libellé", "Type"]
        d["Vérification"] = "Vérifier déréférencement ou code article"
        st.dataframe(d, use_container_width=True, hide_index=True)

# ══ TAB 5 ═════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>📋 5 feuilles Excel · Code couleur management visuel</strong><br>
  <strong>1. Synthèse réseau</strong> — zone exécutive KPIs + tableau taux par magasin et type<br>
  <strong>2. Ruptures nettes</strong> — stock = 0 sur tous les sites · GOLD en doré<br>
  <strong>3. Réappro urgent</strong> — stock rouge (1 unité) · orange (2 unités)<br>
  <strong>4. Ruptures partielles</strong> — sites manquants listés · taux coloré<br>
  <strong>5. Absents PBI</strong> — articles Top CA introuvables
</div>""", unsafe_allow_html=True)
    st.caption(f"Périmètre : {n_refs} réf · {n_sites} sites · {type_filtre} · seuil réappro : {seuil_reappro}")

    if st.button("Générer l'export Excel", type="primary"):
        with st.spinner("Génération…"):
            buf = gen_excel(taux_df, types_seg, ruptures, reappro, partielles,
                            absents, type_filtre, cible, seuil_reappro)
        st.download_button("⬇️ Télécharger", data=buf,
            file_name=f"SmartBuyer_Detention_{type_filtre}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(f"""
<div style='text-align:center;color:#C7C7CC;font-size:11px;padding:8px 0'>
  NovaRetail Solutions · SmartBuyer v2.3 · Détention Top CA · {n_refs} réf · {n_sites} sites
</div>""", unsafe_allow_html=True)
