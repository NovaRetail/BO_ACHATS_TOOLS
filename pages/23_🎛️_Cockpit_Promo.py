"""
23_🎛️_Cockpit_Promo.py — Module Cockpit Promo · SmartBuyer Hub
Balaie un export PROMO et sort les alertes Disponibilité (stock ≤0, avec/sans
réappro RAL) et Marge (vente à perte, marge faible, PMP manquant), avec un
flag combiné "Disponibilité + Marge" pour les cas cumulés prioritaires.
Colonne Action : message actionnable par ligne (Rupture -> Passer commande,
RAL -> Accélérer livraison, etc.).
Visuel : cartes HTML façon app moderne (pas de grille type tableur) pour la
liste des alertes et la synthèse par magasin.
Export Excel "Cockpit Promo - AAAAMMJJ.xlsx" — valeurs calculées en Python,
aucune formule dans le classeur.
"""

import datetime as _dt
import html as _html

import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CONFIG & CHARTE (Apple clair — identique aux autres modules SmartBuyer)
# ============================================================
st.set_page_config(page_title="Cockpit Promo · SmartBuyer", page_icon="🎛️", layout="wide", initial_sidebar_state="expanded")

BLUE = "#007AFF"
GREEN = "#34C759"
RED = "#FF3B30"
AMBER = "#FF9500"
DARK = "#1D1D1F"
GREY = "#86868B"
BG = "#F2F2F7"

MARGE_FAIBLE_DEFAUT = 10  # % — modifiable dans la sidebar
VALID_SITES = {"0010301", "0010202", "0010203"}  # à ajuster si le périmètre magasins change

REQUIRED_COLS = ["Code site", "Rayon", "Libellé rayon", "Libellé article", "Code article",
                  "DPR", "PV Promo", "Taux TVA", "PMP", "Stock", "RAL", "Marge en cours", "Four."]
OPTIONAL_COLS = ["Quantité vendue", "Montant vente HT", "Montant achat"]

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1350px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; gap: 4px !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; background: #FFFFFF; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }

.format-card { border-radius: 12px; padding: 14px 16px; margin-bottom: 6px; border: 0.5px solid; }
.format-hyper  { background: #EFF6FF; border-color: #B3D9FF; }
.format-market { background: #FFFBF0; border-color: #FFD9A0; }
.format-supeco { background: #FFF2F2; border-color: #FFB3B3; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 600; }
.badge-hyper  { background: #154360; color: #FFFFFF; }
.badge-red    { background: #FF3B30; color: #FFFFFF; }
.badge-amber  { background: #FF9500; color: #FFFFFF; }
.badge-green  { background: #34C759; color: #FFFFFF; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-optional { background: #F9F9FB; border-color: #D1D1D6; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.card { background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px;margin-bottom:10px; }
.small-muted { font-size:12px;color:#8E8E93; }

/* ---- Cartes de liste "app moderne" (remplace le tableau type tableur) ---- */
.alist { background:#FFFFFF; border-radius:14px; border:0.5px solid #E5E5EA; overflow:hidden; margin-bottom:14px; }
.alist-head { display:grid; grid-template-columns: 64px 1.6fr 1fr 64px 90px 70px 240px; gap:0;
              padding:10px 16px; font-size:10.5px; font-weight:600; color:#8E8E93;
              text-transform:uppercase; letter-spacing:.05em; background:#FAFAFC; border-bottom:0.5px solid #E5E5EA; }
.alist-body { max-height:440px; overflow-y:auto; }
.alist-row { display:grid; grid-template-columns: 64px 1.6fr 1fr 64px 90px 70px 240px; gap:0;
             padding:10px 16px; font-size:13px; border-bottom:0.5px solid #F2F2F7; align-items:center; }
.alist-row:last-child { border-bottom:none; }
.alist-empty { padding:28px; text-align:center; font-size:13px; color:#8E8E93; }
.pill { display:inline-block; padding:3px 10px; border-radius:8px; font-size:11px; font-weight:600; white-space:nowrap; }

/* ---- Cartes de synthèse par magasin ---- */
.site-grid { display:grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap:12px; margin-bottom:16px; }
.site-card { background:#FFFFFF; border-radius:14px; border:0.5px solid #E5E5EA; padding:16px 18px; }
.site-card-head { display:flex; align-items:center; justify-content:space-between; margin-bottom:12px; }
.site-card-title { font-size:15px; font-weight:600; color:#1C1C1E; }
.site-stat-row { display:flex; align-items:center; justify-content:space-between; padding:5px 0; font-size:12.5px; color:#3A3A3C; border-bottom:0.5px solid #F5F5F7; }
.site-stat-row:last-child { border-bottom:none; }
.site-stat-val { font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# HELPERS DE FORMAT (convention SmartBuyer)
# ============================================================
def fmt(n):
    if n is None or (isinstance(n, float) and (pd.isna(n) or not np.isfinite(n))):
        return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if v is None or pd.isna(v): return "—"
    return f"{v:.{dec}f}%"

def read_csv_robust(file) -> pd.DataFrame:
    """Lecture CSV avec repli d'encodage UTF-8 -> CP1252 -> Latin-1 (convention SmartBuyer)."""
    raw = file.read()
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(raw), sep=";", encoding=enc, dtype=str)
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("Impossible de décoder le fichier (UTF-8 / CP1252 / Latin-1 ont échoué).")

def to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s.astype(str).str.strip(), errors="coerce")

# ============================================================
# CALCUL DES ALERTES
# ============================================================
@st.cache_data(show_spinner=False)
def compute_alerts(file_bytes, seuil_marge_faible):
    df = read_csv_robust(io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]
    df = df[df["Code site"].isin(VALID_SITES)].copy()

    for c in ["DPR", "PV Promo", "Taux TVA", "PMP", "Stock", "RAL", "Marge en cours",
              "Quantité vendue", "Montant vente HT", "Montant achat", "Four."]:
        df[c] = to_num(df[c]) if c in df.columns else np.nan

    df["Site"] = df["Code site"].str.lstrip("0").astype(int)
    df["Rayon_aff"] = df["Libellé rayon"].str.strip()
    df["Article_aff"] = df["Libellé article"].str.strip()
    df["Code_article"] = df["Code article"].astype(str).str.lstrip("0")
    df["RAL_int"] = df["RAL"].fillna(0).astype(int)

    df["stat_stock"] = np.where(
        (df["Stock"] <= 0) & (df["RAL_int"] <= 0), "Rupture sans réappro",
        np.where((df["Stock"] <= 0) & (df["RAL_int"] > 0), "Réappro en cours", ""))
    df["stat_marge"] = np.where(
        df["Marge en cours"] < 0, "Vente à perte",
        np.where(df["Marge en cours"] < seuil_marge_faible, "Marge faible", ""))
    df["pmp_manquant"] = df["PMP"].isna()
    df["flag_dispo"] = df["stat_stock"] != ""
    df["flag_marge"] = (df["stat_marge"] != "") | df["pmp_manquant"]
    df["type_alerte"] = np.where(
        df["flag_dispo"] & df["flag_marge"], "Disponibilité + Marge",
        np.where(df["flag_dispo"], "Disponibilité",
                 np.where(df["flag_marge"], "Marge", "")))
    df["marge_realisee"] = df["Montant vente HT"] - df["Montant achat"]
    df["pmp_effectif"] = df["PMP"].fillna(df["DPR"])

    # ---- Colonne Action : message actionnable, pas un statut descriptif ----
    df["action_stock"] = np.where(
        df["stat_stock"] == "Rupture sans réappro", "Rupture - Passer commande",
        np.where(df["stat_stock"] == "Réappro en cours",
                 "RAL - Accélérer livraison (" + df["RAL_int"].astype(str) + ")", ""))
    df["action_marge"] = np.where(
        df["stat_marge"] != "", df["stat_marge"],
        np.where(df["pmp_manquant"], "PMP manquant", ""))
    df["action"] = np.where(
        df["type_alerte"] == "Disponibilité + Marge",
        df["action_stock"] + " + " + df["action_marge"],
        np.where(df["type_alerte"] == "Disponibilité", df["action_stock"], df["action_marge"]))

    return df[df["type_alerte"] != ""].copy()

def synthese_par_site(alerts):
    rows = []
    for site in sorted(alerts["Site"].unique()):
        sub = alerts[alerts["Site"] == site]
        rows.append({
            "Site": site,
            "Rupture sans réappro": int((sub["stat_stock"] == "Rupture sans réappro").sum()),
            "Réappro en cours": int((sub["stat_stock"] == "Réappro en cours").sum()),
            "Vente à perte": int((sub["stat_marge"] == "Vente à perte").sum()),
            "Marge faible": int((sub["stat_marge"] == "Marge faible").sum()),
            "PMP manquant": int((sub["action_marge"] == "PMP manquant").sum()),
            "Dispo + Marge": int((sub["type_alerte"] == "Disponibilité + Marge").sum()),
        })
    return pd.DataFrame(rows)

def build_headline(alerts, synth):
    n = len(alerts)
    n_cumul = int((alerts["type_alerte"] == "Disponibilité + Marge").sum())
    n_perte = int((alerts["stat_marge"] == "Vente à perte").sum())
    pire_site = synth.loc[synth["Dispo + Marge"].idxmax()] if len(synth) else None
    line1 = (f"{n} lignes en alerte &nbsp;·&nbsp; "
             f"{n_cumul} en cumul Disponibilité + Marge (prioritaires) &nbsp;·&nbsp; "
             f"{n_perte} ventes à perte")
    line2 = ""
    if pire_site is not None and pire_site["Dispo + Marge"] > 0:
        line2 = f"📌 Site à prioriser : <b>{int(pire_site['Site'])}</b> ({int(pire_site['Dispo + Marge'])} alertes cumulées)"
    return line1, line2

# ============================================================
# RENDU HTML "APP MODERNE" — remplace le tableau type tableur
# ============================================================
def _marge_pill(v, seuil):
    if pd.isna(v):
        return "<span style='color:#8E8E93;'>—</span>"
    if v < 0:
        color = "#C0392B"
    elif v < seuil:
        color = "#B36B00"
    else:
        color = "#1E7B3C"
    return f"<span style='color:{color}; font-weight:600;'>{v:.1f}%</span>"

def _action_pill(action):
    a = action or ""
    if "Rupture" in a or "Vente à perte" in a:
        bg, fg = "#FDECEA", "#C0392B"
    elif "RAL" in a or "Marge faible" in a:
        bg, fg = "#FFF6E5", "#B36B00"
    elif "PMP manquant" in a:
        bg, fg = "#F0F0F2", "#5A5A5E"
    else:
        bg, fg = "#F0F0F2", "#5A5A5E"
    return f"<span class='pill' style='background:{bg}; color:{fg};'>{_html.escape(a)}</span>"

def render_alert_list(d, seuil, max_height=440):
    if d.empty:
        st.markdown("<div class='alist'><div class='alist-empty'>Aucune ligne pour ce filtre.</div></div>", unsafe_allow_html=True)
        return
    rows_html = []
    for _, row_ in d.iterrows():
        stock = row_["Stock"]
        stock_html = (f"<span style='color:#C0392B; font-weight:600;'>{int(stock)}</span>"
                      if pd.notna(stock) and stock <= 0 else f"{fmt(stock)}")
        ral = row_["RAL"]
        ral_html = fmt(ral) if pd.notna(ral) and ral else "<span style='color:#C7C7CC;'>—</span>"
        rows_html.append(f"""
<div class="alist-row">
  <div style="color:#8E8E93;">{int(row_['Site'])}</div>
  <div>
    <div style="font-weight:500; color:#1C1C1E;">{_html.escape(row_['Article_aff'])}</div>
    <div style="font-size:11px; color:#8E8E93; margin-top:1px;">{_html.escape(row_['Rayon_aff'])}</div>
  </div>
  <div style="color:#8E8E93; font-size:12px;">{_html.escape(str(row_['Code_article']))}</div>
  <div>{stock_html}</div>
  <div style="color:#8E8E93;">{ral_html}</div>
  <div>{_marge_pill(row_['Marge en cours'], seuil)}</div>
  <div>{_action_pill(row_['action'])}</div>
</div>""")
    body = "".join(rows_html)
    st.markdown(f"""
<div class="alist">
  <div class="alist-head">
    <div>Site</div><div>Article</div><div>Code</div><div>Stock</div><div>RAL</div><div>Marge</div><div>Action</div>
  </div>
  <div class="alist-body" style="max-height:{max_height}px;">{body}</div>
</div>
""", unsafe_allow_html=True)

def render_site_synthesis(synth):
    cards = []
    for _, r in synth.iterrows():
        cards.append(f"""
<div class="site-card">
  <div class="site-card-head">
    <div class="site-card-title">Site {int(r['Site'])}</div>
    <span class="badge {'badge-red' if r['Dispo + Marge'] > 0 else 'badge-green'}">{int(r['Dispo + Marge'])} cumul</span>
  </div>
  <div class="site-stat-row"><span>Rupture sans réappro</span><span class="site-stat-val" style="color:#C0392B;">{int(r['Rupture sans réappro'])}</span></div>
  <div class="site-stat-row"><span>Réappro en cours</span><span class="site-stat-val" style="color:#B36B00;">{int(r['Réappro en cours'])}</span></div>
  <div class="site-stat-row"><span>Vente à perte</span><span class="site-stat-val" style="color:#C0392B;">{int(r['Vente à perte'])}</span></div>
  <div class="site-stat-row"><span>Marge faible</span><span class="site-stat-val" style="color:#B36B00;">{int(r['Marge faible'])}</span></div>
  <div class="site-stat-row"><span>PMP manquant</span><span class="site-stat-val" style="color:#8E8E93;">{int(r['PMP manquant'])}</span></div>
</div>""")
    st.markdown(f"<div class='site-grid'>{''.join(cards)}</div>", unsafe_allow_html=True)

# ============================================================
# EXPORT EXCEL — valeurs calculées en Python, aucune formule
# Couleur réservée à ce qui a un vrai impact financier (Cumul, Marge) ;
# Disponibilité reste neutre (c'est un sujet opérationnel, pas une perte).
# ============================================================
def build_excel(alerts, synth):
    """Une seule liste d'alertes (triée Cumul -> Disponibilité -> Marge), avec une
    colonne Type en couleur de texte plutôt que 3 blocs de section répétés."""
    BLUE_H = "FF007AFF"; NEUTRAL_H = "FF48484A"
    WHITE_H = "FFFFFFFF"; LGREY_H = "FFF7F7F9"; ARIAL = "Arial"
    TYPE_COLOR = {"Disponibilité + Marge": "FFC0392B", "Disponibilité": "FF5F5E5A", "Marge": "FFB36B00"}
    TYPE_LABEL = {"Disponibilité + Marge": "Cumul", "Disponibilité": "Disponibilité", "Marge": "Marge"}
    thin = Side(style="thin", color="FFE0E0E2")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    QTY = "#,##0"; PCT = "0.0%"; ACC = "#,##0"

    def section_bar(ws, row, ncols, text, color=NEUTRAL_H):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name=ARIAL, bold=True, size=11, color=WHITE_H)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def header_row(ws, row, labels):
        for i, lbl in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lbl)
            c.font = Font(name=ARIAL, bold=True, size=10, color=WHITE_H)
            c.fill = PatternFill("solid", fgColor=BLUE_H)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_row(ws, row, values, zebra=False, left_cols=()):
        fillc = LGREY_H if zebra else "FFFFFFFF"
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name=ARIAL, size=10)
            c.border = box
            c.fill = PatternFill("solid", fgColor=fillc)
            c.alignment = Alignment(horizontal="left" if i in left_cols else "center")

    def autosize(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    wb = Workbook()
    ws = wb.active
    ws.title = "Cockpit Promo"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:K1")
    ws["A1"] = "COCKPIT PROMO — ALERTES DISPONIBILITÉ & MARGE"
    ws["A1"].font = Font(name=ARIAL, bold=True, size=14, color=WHITE_H)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_H)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # ---- Bandeau compact : date + KPI sur une seule ligne ----
    n_cumul = int((alerts["type_alerte"] == "Disponibilité + Marge").sum())
    n_dispo = int((alerts["type_alerte"] == "Disponibilité").sum())
    n_marge = int((alerts["type_alerte"] == "Marge").sum())
    ws.merge_cells("A2:K2")
    kpi_txt = (f"Généré le {_dt.date.today().strftime('%d/%m/%Y')}   ·   "
               f"{len(alerts)} alertes   ·   {n_cumul} cumul   ·   {n_dispo} disponibilité   ·   {n_marge} marge")
    ws["A2"] = kpi_txt
    ws["A2"].font = Font(name=ARIAL, size=10, color="FF48484A")
    ws["A2"].fill = PatternFill("solid", fgColor=LGREY_H)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    r = 4
    section_bar(ws, r, 11, "SYNTHÈSE PAR MAGASIN"); r += 1
    header_row(ws, r, list(synth.columns)); r += 1
    for i, (_, row_) in enumerate(synth.iterrows()):
        data_row(ws, r, list(row_.values), zebra=(i % 2 == 1), left_cols=(1,))
        for col in range(2, len(synth.columns) + 1):
            ws.cell(row=r, column=col).number_format = QTY
        r += 1
    r += 1

    # ---- Liste unique, triée Cumul -> Disponibilité -> Marge, puis par marge croissante ----
    section_bar(ws, r, 11, f"DÉTAIL DES ALERTES — {len(alerts)} lignes"); r += 1
    labels = ["Site", "Rayon", "Article", "Code article", "Stock", "RAL",
               "PMP", "PV Promo", "Marge %", "Type", "Action", "Fournisseur"]
    header_row(ws, r, labels); r += 1
    table_start = r

    priority = {"Disponibilité + Marge": 0, "Disponibilité": 1, "Marge": 2}
    sorted_alerts = alerts.assign(_p=alerts["type_alerte"].map(priority)).sort_values(
        ["_p", "Marge en cours"])

    for i, (_, row_) in enumerate(sorted_alerts.iterrows()):
        vals = [row_["Site"], row_["Rayon_aff"], row_["Article_aff"], row_["Code_article"],
                row_["Stock"], (row_["RAL"] if pd.notna(row_["RAL"]) else 0),
                row_["pmp_effectif"], row_["PV Promo"],
                (row_["Marge en cours"] / 100 if pd.notna(row_["Marge en cours"]) else None),
                TYPE_LABEL[row_["type_alerte"]], row_["action"],
                (int(row_["Four."]) if pd.notna(row_["Four."]) else None)]
        data_row(ws, r, vals, zebra=(i % 2 == 1), left_cols=(2, 3, 11))
        ws.cell(row=r, column=9).number_format = PCT
        for c in (5, 6, 7, 8):
            ws.cell(row=r, column=c).number_format = ACC
        type_cell = ws.cell(row=r, column=10)
        type_cell.font = Font(name=ARIAL, size=10, bold=True, color=TYPE_COLOR[row_["type_alerte"]])
        r += 1

    autosize(ws, {'A': 8, 'B': 20, 'C': 32, 'D': 13, 'E': 8, 'F': 7, 'G': 9, 'H': 10,
                   'I': 9, 'J': 15, 'K': 30, 'L': 13})
    ws.freeze_panes = f"A{table_start}"
    ws.auto_filter.ref = f"A{table_start - 1}:L{r - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ============================================================
# INTERFACE
# ============================================================
st.markdown("<div class='page-title'>🎛️ Cockpit Promo</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Alertes Disponibilité & Marge sur l'export PROMO · flag combiné pour les cas cumulés · synthèse par magasin</div>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichier</div>", unsafe_allow_html=True)
    up = st.file_uploader("Export PROMO (.csv)", type=["csv"], key="up_promo")
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Paramètres</div>", unsafe_allow_html=True)
    seuil = st.slider("Seuil marge faible (%)", 0, 30, MARGE_FAIBLE_DEFAUT, step=1)
    st.markdown("---")
    st.caption("SmartBuyer Hub · Module Cockpit Promo")

if up is None:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Ce module balaie un export PROMO et sort les lignes à risque de rupture et/ou de marge
  négative, avec un flag d'action par ligne et une synthèse par magasin.
  Un fichier à charger dans la barre latérale.
</div>
""", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<div class='section-label'>Contenu du module</div>", unsafe_allow_html=True)
        st.markdown("""
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>📦 Disponibilité</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Stock ≤ 0 : "Rupture - Passer commande" (RAL = 0) ou "RAL - Accélérer livraison" (RAL &gt; 0).
  </div>
</div>
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>📉 Marge</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Marge % calculée depuis PMP, PV Promo et Taux TVA : vente à perte, marge faible
    (seuil réglable), ou PMP manquant.
  </div>
</div>
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>🚨 Cumul prioritaire</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Colonne Action combinée pour les articles qui cumulent les deux — à traiter en premier.
  </div>
</div>
""", unsafe_allow_html=True)
    with c2:
        st.markdown("<div class='section-label'>Types d'alerte</div>", unsafe_allow_html=True)
        st.markdown("""
<div class='format-card format-hyper'>
  <span class='badge badge-hyper'>Disponibilité</span>
  <div class='small-muted' style='margin-top:6px'>Rupture sans réappro ou réappro en cours</div>
</div>
<div class='format-card format-market'>
  <span class='badge badge-amber'>Marge faible</span>
  <div class='small-muted' style='margin-top:6px'>Sous le seuil réglable (10% par défaut)</div>
</div>
<div class='format-card format-supeco'>
  <span class='badge badge-red'>Vente à perte / Cumul</span>
  <div class='small-muted' style='margin-top:6px'>Marge négative, seule ou combinée à une rupture</div>
</div>
""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Fonctionnement</div>", unsafe_allow_html=True)
        st.markdown("""
<div class='alert-card alert-green'>
  <strong>1.</strong> Charge l'export PROMO dans la sidebar.<br>
  <strong>2.</strong> Ajuste si besoin le seuil de marge faible.<br>
  <strong>3.</strong> Consulte le cockpit et la synthèse par magasin.<br>
  <strong>4.</strong> Télécharge l'Excel (valeurs figées, prêt à diffuser).
</div>
""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Colonnes attendues</div>", unsafe_allow_html=True)
    col_left, col_right = st.columns(2)
    with col_left:
        for c in REQUIRED_COLS:
            st.markdown(f"""
<div class='col-required'>
  <div style='font-size:16px'>▪️</div>
  <div class='col-name'>{c}</div>
</div>""", unsafe_allow_html=True)
    with col_right:
        st.caption("Optionnelles (marge réalisée)")
        for c in OPTIONAL_COLS:
            st.markdown(f"""
<div class='col-required col-optional'>
  <div style='font-size:16px'>▪️</div>
  <div class='col-name'>{c}</div>
</div>""", unsafe_allow_html=True)

    st.info("⬅️ Charge ton export PROMO dans la barre latérale pour démarrer.")
    st.stop()

# ============================================================
# TRAITEMENT
# ============================================================
try:
    alerts = compute_alerts(up.getvalue(), seuil)
except Exception as e:
    st.error(f"Lecture du fichier impossible : {e}")
    st.stop()

if alerts.empty:
    st.success("Aucune alerte détectée sur ce périmètre — rien à traiter aujourd'hui.")
    st.stop()

synth = synthese_par_site(alerts)
line1, line2 = build_headline(alerts, synth)
st.markdown(
    f"<div class='alert-card alert-blue'><strong>{line1}</strong>"
    + (f"<br>{line2}" if line2 else "")
    + "</div>", unsafe_allow_html=True)

st.markdown("<div class='section-label'>Vue d'ensemble</div>", unsafe_allow_html=True)
n_dispo = int((alerts["type_alerte"] == "Disponibilité").sum())
n_marge = int((alerts["type_alerte"] == "Marge").sum())
n_cumul = int((alerts["type_alerte"] == "Disponibilité + Marge").sum())
c1, c2, c3, c4 = st.columns(4)
c1.metric("Total alertes", len(alerts))
c2.metric("Disponibilité", n_dispo)
c3.metric("Marge", n_marge)
c4.metric("Cumul (critique)", n_cumul, delta_color="off")

st.markdown("<div class='section-label'>Synthèse par magasin</div>", unsafe_allow_html=True)
render_site_synthesis(synth)

st.markdown("<div class='section-label'>Détail des alertes</div>", unsafe_allow_html=True)

tab_cumul, tab_dispo, tab_marge = st.tabs([
    f"🚨 Cumul ({n_cumul})",
    f"📦 Disponibilité ({n_dispo})",
    f"📉 Marge ({n_marge})",
])
with tab_cumul:
    st.caption("Articles qui cumulent une rupture et une marge dégradée — à traiter en premier.")
    render_alert_list(alerts[alerts["type_alerte"] == "Disponibilité + Marge"].sort_values("Marge en cours"), seuil)

with tab_dispo:
    st.caption("Stock ≤ 0, triés du plus critique au moins critique.")
    render_alert_list(alerts[alerts["type_alerte"] == "Disponibilité"].sort_values("Stock"), seuil)

with tab_marge:
    st.caption("Marge négative, marge faible ou PMP manquant, triés du pire au moins grave.")
    render_alert_list(alerts[alerts["type_alerte"] == "Marge"].sort_values("Marge en cours"), seuil)

st.markdown("<div class='section-label'>Export</div>", unsafe_allow_html=True)
xls = build_excel(alerts, synth)
today_str = _dt.date.today().strftime("%Y%m%d")
export_filename = f"Cockpit Promo - {today_str}.xlsx"
st.download_button(f"📥 Télécharger {export_filename}", xls, file_name=export_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
st.caption("Fichier à une feuille : synthèse par magasin puis une liste unique triée Cumul → Disponibilité → Marge, filtre automatique activé — valeurs figées, aucune formule.")
