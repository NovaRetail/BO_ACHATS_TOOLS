"""
17_💸_Reporting_Sous_Familles.py — Module Reporting Sous Familles · SmartBuyer Hub
Variations Réseau + Site à la maille Sous Famille, alerte marge (rentabilité),
un onglet par Rayon + Synthèse (KPI, alertes/rayon, Top 10 critiques par rayon).
Export Excel nommé "Reporting Sous Familles - AAAAMMJJ.xlsx".
"""

import datetime as _dt
import re
import io

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ============================================================
# CONFIG & CHARTE (Apple clair — identique aux autres modules SmartBuyer)
# ============================================================
st.set_page_config(page_title="Reporting Sous Familles · SmartBuyer", page_icon="💸",
                    layout="wide", initial_sidebar_state="expanded")

BLUE = "#007AFF"
GREEN = "#34C759"
RED = "#FF3B30"
AMBER = "#FF9500"
DARK = "#1D1D1F"
GREY = "#86868B"
BG = "#F2F2F7"

SITE_ORDER = ['10301 - Hyper Marcory', '10202 - Hyper Palmeraie', '10203 - Hyper Yopougon',
              '10705 - Market 7 Décembre', '10208 - Market Riviera', '10206 - Market Kokoh Mall',
              '10604 - Market Cité verte', '10601 - Supeco Niangon', '10602 - Supeco Terminus 47',
              '10603 - Supeco Toit rouge']
SITE_SHORT = {s: s.split(' - ', 1)[1] for s in SITE_ORDER}

RAYON_LABELS = {
    '00010 - BOISSONS': 'Boissons',
    '00011 - DROGUERIE': 'Droguerie',
    '00012 - PARFUMERIE HYGIENE': 'Parfumerie Hygiène',
    '00014 - EPICERIE': 'Epicerie',
}

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1350px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
.stButton > button[kind="primary"] { background: #007AFF !important; border: none !important; border-radius: 8px !important; font-weight: 600 !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; background: #FFFFFF; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-purple{ background: #F5F0FF; border-color: #AF52DE; color: #1A0033; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 600; }
.badge-hyper  { background: #154360; color: #FFFFFF; }
.badge-market { background: #145A32; color: #FFFFFF; }
.badge-supeco { background: #6E2F8A; color: #FFFFFF; }
.badge-red    { background: #FF3B30; color: #FFFFFF; }
.badge-amber  { background: #FF9500; color: #FFFFFF; }
.badge-green  { background: #34C759; color: #FFFFFF; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.card { background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;padding:16px;margin-bottom:10px; }
.small-muted { font-size:12px;color:#8E8E93; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# HELPERS DE FORMAT (convention SmartBuyer)
# ============================================================
def fmt(n):
    if n is None or (isinstance(n, float) and (pd.isna(n) or not np.isfinite(n))):
        return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if v is None or pd.isna(v) or not np.isfinite(v):
        return "—"
    return f"{v*100:.{dec}f}%"

def fmt_delta(v):
    if v is None or pd.isna(v) or not np.isfinite(v):
        return "—"
    return f"{v*100:+.2f} pts"

def extract_periode(perimetre_text):
    """Extrait 'JJ/MM/AAAA → JJ/MM/AAAA' depuis la dernière cellule de la 1ère colonne
    de l'export PBI (texte de filtre)."""
    if not perimetre_text:
        return "Voir périmètre ci-dessous"
    m = re.search(r"après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})", str(perimetre_text))
    if m:
        return f"{m.group(1)} → {m.group(2)}"
    return str(perimetre_text).replace("\n", " ")[:120]

# ============================================================
# CHARGEMENT & AGRÉGATION — export Rayon → Famille → Sous Famille → Article × Site
# ============================================================
@st.cache_data(show_spinner=False)
def load_and_aggregate(file_bytes, seuil_critique, seuil_surveiller):
    raw = pd.read_excel(io.BytesIO(file_bytes))
    raw.columns = [str(c).lstrip('\ufeff').strip() for c in raw.columns]

    perimetre = None
    note_rows = raw[raw['Rayon'].astype(str).str.contains('Filtres appliqués', na=False)]
    if not note_rows.empty:
        perimetre = str(note_rows.iloc[0]['Rayon'])

    df = raw[~raw['Rayon'].astype(str).str.contains('Filtres appliqués', na=False)].copy()
    df['Rayon'] = df['Rayon'].ffill()
    df['Famille'] = df['Famille'].ffill()
    df['Sous Famille'] = df['Sous Famille'].ffill()

    # Marge N-1 non fournie directement -> dérivée de Marge et du taux de croissance
    marge_growth_col = '%Vs N-1.1' if '%Vs N-1.1' in df.columns else None
    if marge_growth_col:
        df['Marge N-1'] = df['Marge'] / (1 + df[marge_growth_col])
        df.loc[df[marge_growth_col] == -1, 'Marge N-1'] = np.nan
    else:
        df['Marge N-1'] = np.nan

    # ---- RESEAU : lignes Article='Total' et Site vide -> total Sous Famille direct ----
    reseau_raw = df[(df['Article'] == 'Total') & (df['Site nom long'].isna())].copy()
    reseau_raw = reseau_raw[reseau_raw['Sous Famille'] != 'Total']
    res_agg = reseau_raw.groupby(['Rayon', 'Famille', 'Sous Famille'], as_index=False).agg(
        CA=('CA', 'sum'), CA_N1=('CA N-1', 'sum'),
        Marge=('Marge', 'sum'), Marge_N1=('Marge N-1', 'sum'),
        Qte=('Qté Vente', 'sum'), Qte_N1=('Qté Vente N-1', 'sum'))
    res_agg['VarCA_Reseau'] = res_agg['CA'] / res_agg['CA_N1'] - 1
    res_agg['MargePct'] = res_agg['Marge'] / res_agg['CA']
    res_agg['MargePct_N1'] = res_agg['Marge_N1'] / res_agg['CA_N1']
    res_agg['EvoMarge_Reseau'] = res_agg['MargePct'] - res_agg['MargePct_N1']
    res_agg['VarQte_Reseau'] = res_agg['Qte'] / res_agg['Qte_N1'] - 1

    # ---- SITE : lignes de détail (Article réel x Site réel), agrégées à la Sous Famille ----
    site_leaf = df[(df['Article'].notna()) & (df['Article'] != 'Total') &
                   (df['Site nom long'].notna()) & (df['Site nom long'] != 'Total')].copy()
    site_agg = site_leaf.groupby(['Rayon', 'Famille', 'Sous Famille', 'Site nom long'], as_index=False).agg(
        CA=('CA', 'sum'), CA_N1=('CA N-1', 'sum'),
        Marge=('Marge', 'sum'), Marge_N1=('Marge N-1', 'sum'),
        Qte=('Qté Vente', 'sum'), Qte_N1=('Qté Vente N-1', 'sum'))
    site_agg['VarCA_Site'] = site_agg['CA'] / site_agg['CA_N1'] - 1
    site_agg['MargePct_Site'] = site_agg['Marge'] / site_agg['CA']
    site_agg['MargePct_Site_N1'] = site_agg['Marge_N1'] / site_agg['CA_N1']
    site_agg['EvoMarge_Site'] = site_agg['MargePct_Site'] - site_agg['MargePct_Site_N1']
    site_agg['VarQte_Site'] = site_agg['Qte'] / site_agg['Qte_N1'] - 1

    # ---- ALERTE (niveau Réseau, axée rentabilité) ----
    def alerte(row):
        evo = row['EvoMarge_Reseau']
        if pd.isna(evo) or np.isinf(row['VarCA_Reseau']):
            return "🔴 Activité nouvelle/disparue - à vérifier"
        ca_up = row['VarCA_Reseau'] > 0
        marge_down = row['Marge'] < row['Marge_N1']
        if evo < -seuil_critique or row['MargePct'] < 0:
            return "🔴 Marge en chute"
        if (evo <= -seuil_surveiller) or (ca_up and marge_down):
            return "🟠 Marge à surveiller"
        return "🟢 Marge saine"

    res_agg['Alerte'] = res_agg.apply(alerte, axis=1)

    return res_agg, site_agg, perimetre

def severity_rank(alerte):
    if alerte.startswith('🔴'):
        return 0
    if alerte.startswith('🟠'):
        return 1
    return 2

# ============================================================
# EXPORT EXCEL — 1 onglet Synthèse + 1 onglet par Rayon
# ============================================================
def build_excel_sous_familles(res_agg, site_agg, perimetre, periode):
    ARIAL = "Arial"
    header_fill = PatternFill("solid", fgColor="FF007AFF")
    header_font = Font(color="FFFFFFFF", bold=True, name=ARIAL, size=10)
    sub_header_fill = PatternFill("solid", fgColor="FFD6E9FF")
    site_header_fill = PatternFill("solid", fgColor="FFE8E8ED")
    thin = Side(style="thin", color="FFC7C7CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_badge = PatternFill("solid", fgColor="FFFFE5E3")
    orange_badge = PatternFill("solid", fgColor="FFFFF1DA")
    green_badge = PatternFill("solid", fgColor="FFE5F8EA")

    def badge_fill(alerte):
        if alerte.startswith('🔴'):
            return red_badge
        if alerte.startswith('🟠'):
            return orange_badge
        return green_badge

    site_lookup = site_agg.set_index(['Rayon', 'Famille', 'Sous Famille', 'Site nom long'])

    def write_rayon_sheet(wb, rayon_code, base_all):
        label = RAYON_LABELS[rayon_code]
        ws = wb.create_sheet(label[:31])
        base = base_all[base_all['Rayon'] == rayon_code].sort_values(
            by='Alerte', key=lambda s: s.map(severity_rank)).reset_index(drop=True)

        ws.merge_cells('A1:A2'); ws['A1'] = "Famille"
        ws.merge_cells('B1:B2'); ws['B1'] = "Sous Famille"
        ws.merge_cells('C1:E1')
        ws['C2'] = "Var %CA"; ws['D2'] = "Evo Marge (pts)"; ws['E2'] = "Var %Qté"
        ws.merge_cells('F1:F2'); ws['F1'] = "Alerte"
        ws.cell(row=1, column=3, value="Réseau")

        col = 7
        site_cols = {}
        for s in SITE_ORDER:
            site_cols[s] = col
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
            ws.cell(row=1, column=col, value=SITE_SHORT[s])
            ws.cell(row=2, column=col, value="Var %CA")
            ws.cell(row=2, column=col + 1, value="Evo Marge (pts)")
            ws.cell(row=2, column=col + 2, value="Var %Qté")
            col += 3
        max_col = col - 1

        for r in (1, 2):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if r == 1:
                    cell.fill = header_fill if c <= 6 else site_header_fill
                    cell.font = header_font if c <= 6 else Font(bold=True, name=ARIAL, size=10)
                else:
                    cell.fill = sub_header_fill if c <= 6 else PatternFill("solid", fgColor="FFF2F2F2")
                    cell.font = Font(bold=True, name=ARIAL, size=9)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 30

        r = 3
        for _, row in base.iterrows():
            ws.cell(row=r, column=1, value=row['Famille'])
            ws.cell(row=r, column=2, value=row['Sous Famille'])
            v_ca = row['VarCA_Reseau']
            ws.cell(row=r, column=3, value=None if pd.isna(v_ca) or np.isinf(v_ca) else v_ca)
            ws.cell(row=r, column=4, value=None if pd.isna(row['EvoMarge_Reseau']) else row['EvoMarge_Reseau'])
            v_qt = row['VarQte_Reseau']
            ws.cell(row=r, column=5, value=None if pd.isna(v_qt) or np.isinf(v_qt) else v_qt)
            alerte_cell = ws.cell(row=r, column=6, value=row['Alerte'])
            alerte_cell.fill = badge_fill(row['Alerte'])
            alerte_cell.font = Font(name=ARIAL, size=9, bold=True)

            for s in SITE_ORDER:
                c = site_cols[s]
                try:
                    srow = site_lookup.loc[(row['Rayon'], row['Famille'], row['Sous Famille'], s)]
                    for off, key in enumerate(['VarCA_Site', 'EvoMarge_Site', 'VarQte_Site']):
                        val = srow[key]
                        ws.cell(row=r, column=c + off, value=None if pd.isna(val) or np.isinf(val) else val)
                except KeyError:
                    pass

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c not in (1, 2, 6):
                    cell.font = Font(name=ARIAL, size=9)
            ws.cell(row=r, column=3).number_format = "0.0%"
            ws.cell(row=r, column=4).number_format = '0.00"pts"'
            ws.cell(row=r, column=5).number_format = "0.0%"
            for s in SITE_ORDER:
                c = site_cols[s]
                ws.cell(row=r, column=c).number_format = "0.0%"
                ws.cell(row=r, column=c + 1).number_format = '0.00"pts"'
                ws.cell(row=r, column=c + 2).number_format = "0.0%"
            r += 1

        last_row = r - 1
        ws.freeze_panes = "C3"
        ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{last_row}"
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 26
        for c in range(3, max_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

        if last_row >= 3:
            var_cols = [3, 4, 5] + [site_cols[s] + off for s in SITE_ORDER for off in (0, 1, 2)]
            for c in var_cols:
                letter = get_column_letter(c)
                rng = f"{letter}3:{letter}{last_row}"
                rule = ColorScaleRule(start_type='min', start_color='FFF7C1C1',
                                       mid_type='num', mid_value=0, mid_color='FFFFFFFF',
                                       end_type='max', end_color='FFC0DD97')
                ws.conditional_formatting.add(rng, rule)
        return base

    wb = Workbook()
    wb.remove(wb.active)

    rayon_bases = {}
    for rayon_code in RAYON_LABELS:
        if rayon_code in res_agg['Rayon'].unique():
            rayon_bases[rayon_code] = write_rayon_sheet(wb, rayon_code, res_agg)

    # ---- SYNTHESE (créée puis déplacée en premier) ----
    ws2 = wb.create_sheet("Synthèse")
    wb.move_sheet("Synthèse", offset=-len(rayon_bases))
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    ws2['A1'] = "SYNTHÈSE MARGE — REPORTING SOUS FAMILLES"
    ws2['A1'].font = Font(name=ARIAL, bold=True, size=14, color="FFFFFFFF")
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 26
    ws2['A2'] = "Période :"; ws2['A2'].font = Font(name=ARIAL, bold=True, size=10)
    ws2['B2'] = periode if periode else "Voir périmètre"; ws2['B2'].font = Font(name=ARIAL, size=10, color="FF0000FF", bold=True)

    kpi = [
        ("Sous familles analysées", len(res_agg), "FFF2F2F7", "FF1C1C1E"),
        ("Marge en chute", int((res_agg['Alerte'] == '🔴 Marge en chute').sum()), "FFFFE5E3", "FFFF3B30"),
        ("Activité à vérifier", int(res_agg['Alerte'].str.startswith('🔴 Activité').sum()), "FFFFE5E3", "FFFF3B30"),
        ("À surveiller", int((res_agg['Alerte'] == '🟠 Marge à surveiller').sum()), "FFFFF1DA", "FFFF9500"),
        ("Marge saine", int((res_agg['Alerte'] == '🟢 Marge saine').sum()), "FFE5F8EA", "FF34C759"),
    ]
    col0 = 1
    for label, val, bgc, fgc in kpi:
        c = get_column_letter(col0)
        ws2[f"{c}4"] = label; ws2[f"{c}4"].font = Font(size=9, color="FF6E6E73", name=ARIAL)
        ws2[f"{c}5"] = val; ws2[f"{c}5"].font = Font(size=20, bold=True, name=ARIAL, color=fgc)
        for rr in (4, 5):
            ws2[f"{c}{rr}"].fill = PatternFill("solid", fgColor=bgc)
            ws2[f"{c}{rr}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[c].width = 18
        col0 += 1
    ws2.row_dimensions[4].height = 18
    ws2.row_dimensions[5].height = 30

    ws2['A8'] = "Alertes par Rayon"; ws2['A8'].font = Font(size=12, bold=True, name=ARIAL)
    headers = ["Rayon", "🔴 Critique", "🟠 À surveiller", "🟢 Saine", "Total"]
    for i, h in enumerate(headers):
        cell = ws2.cell(row=9, column=1 + i, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF", name=ARIAL, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rr = 10
    for rayon_code, label in RAYON_LABELS.items():
        if rayon_code not in rayon_bases:
            continue
        sub = rayon_bases[rayon_code]
        crit = int(sub['Alerte'].str.startswith('🔴').sum())
        surv = int((sub['Alerte'] == '🟠 Marge à surveiller').sum())
        saine = int((sub['Alerte'] == '🟢 Marge saine').sum())
        for i, v in enumerate([label, crit, surv, saine, len(sub)]):
            cell = ws2.cell(row=rr, column=1 + i, value=v)
            cell.font = Font(name=ARIAL, size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if i > 0 else "left")
        rr += 1
    for i in range(5):
        ws2.column_dimensions[get_column_letter(1 + i)].width = 22 if i == 0 else 14

    row_cursor = rr + 2
    for rayon_code, label in RAYON_LABELS.items():
        if rayon_code not in rayon_bases:
            continue
        sub = rayon_bases[rayon_code]
        top10 = sub[sub['Alerte'].str.startswith('🔴')].sort_values('CA', ascending=False).head(10)
        if top10.empty:
            continue
        ws2.cell(row=row_cursor, column=1, value=f"Top 10 critiques — {label}").font = Font(size=12, bold=True, name=ARIAL)
        hr = row_cursor + 1
        headers2 = ["Sous Famille", "CA", "Var %CA Réseau", "Evo Marge (pts)", "Alerte"]
        for i, h in enumerate(headers2):
            cell = ws2.cell(row=hr, column=1 + i, value=h)
            cell.font = Font(bold=True, color="FFFFFFFF", name=ARIAL, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        dr = hr + 1
        for _, row in top10.iterrows():
            vals = [row['Sous Famille'], row['CA'], row['VarCA_Reseau'], row['EvoMarge_Reseau'], row['Alerte']]
            for i, v in enumerate(vals):
                cell = ws2.cell(row=dr, column=1 + i, value=v)
                cell.font = Font(name=ARIAL, size=9)
                cell.border = border
                cell.fill = red_badge
            ws2.cell(row=dr, column=2).number_format = "#,##0"
            ws2.cell(row=dr, column=3).number_format = "0.0%"
            ws2.cell(row=dr, column=4).number_format = '0.00"pts"'
            dr += 1
        row_cursor = dr + 2

    ws2.column_dimensions['B'].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ============================================================
# STYLING ÉCRAN — dégradé rouge/blanc/vert sur les colonnes de variation
# ============================================================
def style_variation_table(disp, var_cols):
    fmt_map = {}
    for c in var_cols:
        fmt_map[c] = (lambda v: fmt_pct(v)) if 'Var %' in c else (lambda v: fmt_delta(v))
    styler = disp.style.background_gradient(subset=var_cols, cmap="RdYlGn", vmin=-0.15, vmax=0.15)
    return styler.format({c: f for c, f in fmt_map.items()})

# ============================================================
# INTERFACE
# ============================================================
st.markdown("<div class='page-title'>💸 Reporting Sous Familles — Variations Réseau/Site</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Alerte marge (rentabilité) à la maille Sous Famille · Réseau + détail par site · un onglet par rayon</div>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichier</div>", unsafe_allow_html=True)
    up = st.file_uploader("Export Article × Site (.xlsx)", type=['xlsx'], key="up_sf")
    st.caption("Export PBI à la maille Rayon → Famille → Sous Famille → Article × Site.")
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Seuils d'alerte marge</div>", unsafe_allow_html=True)
    seuil_critique = st.slider("Seuil critique 🔴 (perte de pts)", 1.0, 10.0, 3.0, step=0.5) / 100
    seuil_surveiller = st.slider("Seuil à surveiller 🟠 (perte de pts)", 0.1, 5.0, 1.0, step=0.1) / 100
    st.markdown("---")
    st.caption("SmartBuyer Hub · Module Reporting Sous Familles")

if up is None:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Ce module calcule les variations CA / Marge / Quantité à la maille <strong>Sous Famille</strong>,
  au niveau <strong>Réseau</strong> et par <strong>Site</strong>, avec une alerte marge (axée rentabilité)
  calculée uniquement au niveau Réseau.
</div>
""", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<div class='section-label'>Contenu du module</div>", unsafe_allow_html=True)
        st.markdown("""
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>📊 Synthèse</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    KPI globaux, compteurs d'alertes par rayon, Top 10 sous-familles critiques par rayon (poids CA).
  </div>
</div>
<div class='card'>
  <div style='font-size:14px;font-weight:700;color:#1C1C1E;margin-bottom:8px'>🗂️ Un onglet par Rayon</div>
  <div style='font-size:12px;color:#3A3A3C;line-height:1.5'>
    Var %CA / Evo Marge (pts) / Var %Qté au niveau Réseau puis par site, alertes propres au rayon,
    triées par sévérité, avec dégradé conditionnel rouge → blanc → vert.
  </div>
</div>
""", unsafe_allow_html=True)
    with c2:
        st.markdown("<div class='section-label'>Règle d'alerte (rentabilité)</div>", unsafe_allow_html=True)
        st.markdown(f"""
<div class='alert-card alert-red'><strong>🔴 Marge en chute</strong> — perte de marge &gt; seuil critique, marge devenue négative, ou activité nouvelle/disparue (CA ou CA N-1 = 0)</div>
<div class='alert-card alert-amber'><strong>🟠 Marge à surveiller</strong> — perte de marge modérée, ou dilution (CA en hausse, marge en baisse)</div>
<div class='alert-card alert-green'><strong>🟢 Marge saine</strong> — marge stable ou en amélioration</div>
""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Colonnes attendues</div>", unsafe_allow_html=True)
        for name, desc in [
            ("Rayon / Famille / Sous Famille / Article", "Hiérarchie — obligatoire"),
            ("Site nom long", "Dimension site — obligatoire"),
            ("CA / CA N-1", "Chiffre d'affaires"),
            ("Marge / %Vs N-1.1", "Marge et évolution du taux"),
            ("Qté Vente / Qté Vente N-1", "Quantités vendues"),
        ]:
            st.markdown(f"""
<div class='col-required'>
  <div style='font-size:16px'>▪️</div>
  <div><div class='col-name'>{name}</div><div class='col-desc'>{desc}</div></div>
</div>
""", unsafe_allow_html=True)
    st.info("⬅️ Charge ton export dans la barre latérale pour démarrer.")
    st.stop()

res_agg, site_agg, perimetre = load_and_aggregate(up.getvalue(), seuil_critique, seuil_surveiller)
periode = extract_periode(perimetre)

if perimetre:
    with st.expander("🔎 Périmètre détecté dans le fichier"):
        st.code(perimetre, language=None)

rayons_disponibles = [r for r in RAYON_LABELS if r in res_agg['Rayon'].unique()]
tab_labels = ["📊 Synthèse"] + [f"🗂️ {RAYON_LABELS[r]}" for r in rayons_disponibles]
tabs = st.tabs(tab_labels)

# ---------------- TAB SYNTHESE ----------------
with tabs[0]:
    st.markdown("<div class='section-label'>Vue d'ensemble</div>", unsafe_allow_html=True)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Sous familles", len(res_agg))
    c2.metric("🔴 Marge en chute", int((res_agg['Alerte'] == '🔴 Marge en chute').sum()))
    c3.metric("🔴 Activité à vérifier", int(res_agg['Alerte'].str.startswith('🔴 Activité').sum()))
    c4.metric("🟠 À surveiller", int((res_agg['Alerte'] == '🟠 Marge à surveiller').sum()))
    c5.metric("🟢 Marge saine", int((res_agg['Alerte'] == '🟢 Marge saine').sum()))

    st.markdown("<div class='section-label'>Alertes par rayon</div>", unsafe_allow_html=True)
    rows = []
    for r in rayons_disponibles:
        sub = res_agg[res_agg['Rayon'] == r]
        rows.append({
            "Rayon": RAYON_LABELS[r],
            "🔴 Critique": int(sub['Alerte'].str.startswith('🔴').sum()),
            "🟠 À surveiller": int((sub['Alerte'] == '🟠 Marge à surveiller').sum()),
            "🟢 Saine": int((sub['Alerte'] == '🟢 Marge saine').sum()),
            "Total": len(sub),
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("<div class='section-label'>Top 10 critiques par rayon (poids CA)</div>", unsafe_allow_html=True)
    for r in rayons_disponibles:
        sub = res_agg[res_agg['Rayon'] == r]
        top10 = sub[sub['Alerte'].str.startswith('🔴')].sort_values('CA', ascending=False).head(10)
        st.markdown(f"**{RAYON_LABELS[r]}**")
        if top10.empty:
            st.caption("Aucune alerte critique sur ce rayon.")
            continue
        disp = top10[['Sous Famille', 'CA', 'VarCA_Reseau', 'EvoMarge_Reseau', 'Alerte']].copy()
        disp['CA'] = disp['CA'].map(fmt)
        disp['VarCA_Reseau'] = disp['VarCA_Reseau'].map(lambda v: fmt_pct(v))
        disp['EvoMarge_Reseau'] = disp['EvoMarge_Reseau'].map(fmt_delta)
        disp = disp.rename(columns={'VarCA_Reseau': 'Var %CA Réseau', 'EvoMarge_Reseau': 'Evo Marge (pts)'})
        st.dataframe(disp, use_container_width=True, hide_index=True)

    st.markdown("<div class='section-label'>Export</div>", unsafe_allow_html=True)
    xls = build_excel_sous_familles(res_agg, site_agg, perimetre, periode)
    today_str = _dt.date.today().strftime("%Y%m%d")
    export_filename = f"Reporting Sous Familles - {today_str}.xlsx"
    st.download_button(f"📥 Télécharger {export_filename}", xls, file_name=export_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.caption(f"Période détectée : {periode} · {len(rayons_disponibles) + 1} feuilles (Synthèse + {len(rayons_disponibles)} rayon(s))")

# ---------------- TABS PAR RAYON ----------------
site_lookup = site_agg.set_index(['Rayon', 'Famille', 'Sous Famille', 'Site nom long'])

for i, rayon_code in enumerate(rayons_disponibles, start=1):
    with tabs[i]:
        base = res_agg[res_agg['Rayon'] == rayon_code].sort_values(
            by='Alerte', key=lambda s: s.map(severity_rank)).reset_index(drop=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("🔴 Critique", int(base['Alerte'].str.startswith('🔴').sum()))
        c2.metric("🟠 À surveiller", int((base['Alerte'] == '🟠 Marge à surveiller').sum()))
        c3.metric("🟢 Saine", int((base['Alerte'] == '🟢 Marge saine').sum()))

        st.markdown("<div class='section-label'>Niveau Réseau</div>", unsafe_allow_html=True)
        disp = base[['Famille', 'Sous Famille', 'VarCA_Reseau', 'EvoMarge_Reseau', 'VarQte_Reseau', 'Alerte']].copy()
        disp = disp.rename(columns={'VarCA_Reseau': 'Var %CA', 'EvoMarge_Reseau': 'Evo Marge (pts)', 'VarQte_Reseau': 'Var %Qté'})
        styler = style_variation_table(disp, ['Var %CA', 'Evo Marge (pts)', 'Var %Qté'])
        st.dataframe(styler, use_container_width=True, hide_index=True, height=420)

        st.markdown("<div class='section-label'>Détail par site — sélection d'une sous famille</div>", unsafe_allow_html=True)
        sf_options = base['Sous Famille'].tolist()
        if sf_options:
            sf_sel = st.selectbox("Sous famille", sf_options, key=f"sf_sel_{rayon_code}")
            rows_site = []
            row_ref = base[base['Sous Famille'] == sf_sel].iloc[0]
            for s in SITE_ORDER:
                try:
                    srow = site_lookup.loc[(rayon_code, row_ref['Famille'], sf_sel, s)]
                    rows_site.append({
                        "Site": SITE_SHORT[s],
                        "Var %CA": srow['VarCA_Site'],
                        "Evo Marge (pts)": srow['EvoMarge_Site'],
                        "Var %Qté": srow['VarQte_Site'],
                    })
                except KeyError:
                    rows_site.append({"Site": SITE_SHORT[s], "Var %CA": np.nan, "Evo Marge (pts)": np.nan, "Var %Qté": np.nan})
            disp_site = pd.DataFrame(rows_site)
            styler_site = style_variation_table(disp_site, ['Var %CA', 'Evo Marge (pts)', 'Var %Qté'])
            st.dataframe(styler_site, use_container_width=True, hide_index=True)
