"""
02_📈_Ventes_PBI.py — SmartBuyer Hub
Suivi ventes hebdomadaires PBI · Charte SmartBuyer v2
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Ventes PBI · SmartBuyer",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1200px; }

[data-testid="stSidebar"] {
    background: #F2F2F7 !important;
    border-right: 0.5px solid #D1D1D6 !important;
}
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 0.5px solid #E5E5EA !important;
    border-radius: 12px !important;
    padding: 16px 18px !important;
}
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stMetricDelta"] { font-size: 12px !important; }

[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }

[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }

[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
[data-testid="baseButton-primary"] { background: #007AFF !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; }

.stDownloadButton > button {
    background: #007AFF !important; color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important;
    padding: 10px 24px !important; width: 100% !important;
}
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }

.alert-card { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red    { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber  { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green  { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue   { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name     { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc     { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.col-example  { font-size: 11px; color: #8E8E93; font-family: monospace; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

# ─── PALETTE ─────────────────────────────────────────────────────────────────
COLORS = {
    "EPICERIE":           "#378ADD",
    "BOISSONS":           "#1D9E75",
    "DROGUERIE":          "#D85A30",
    "PARFUMERIE HYGIENE": "#D4537E",
}
COLORS_FADE = {
    "EPICERIE":           "rgba(55,138,221,0.25)",
    "BOISSONS":           "rgba(29,158,117,0.25)",
    "DROGUERIE":          "rgba(216,90,48,0.25)",
    "PARFUMERIE HYGIENE": "rgba(212,83,126,0.25)",
}
COLOR_DEFAULT      = "#8E8E93"
COLOR_DEFAULT_FADE = "rgba(142,142,147,0.25)"
RED   = "#FF3B30"
GREEN = "#34C759"
AMBER = "#FF9500"

PL_BG   = "rgba(0,0,0,0)"
PL_FONT = dict(family="-apple-system, Helvetica Neue, Arial", color="#3A3A3C", size=12)
PL_GRID = "#F2F2F7"

REQUIRED_COLS = {
    "Rayon":          ("Rayon de l'article",                    "ex: 00014 - EPICERIE"),
    "Famille":        ("Famille de l'article",                  "ex: 00140 - PRODUITS PETIT DEJEUNER"),
    "Article":        ("Code et libellé article",               "ex: 14006584 - 1KG SUCRE SACHET"),
    "Site nom long":  ("Nom du magasin",                        "ex: 10301 - Hyper Marcory"),
    "CA":             ("Chiffre d'affaires HT (FCFA)",          "ex: 9654673"),
    "Marge":          ("Marge brute HT (FCFA)",                 "ex: 883423"),
    "Qté Vente":      ("Quantité vendue sur la période",        "ex: 14034"),
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def fmt_fcfa(n):
    if pd.isna(n): return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def evol_pct(curr, prev):
    if prev and prev != 0: return (curr / prev - 1) * 100
    return None

def delta_str(val):
    if val is None: return ""
    return f"{'+' if val >= 0 else ''}{val:.1f}% vs réf"

def alert_html(level, title, action):
    cls = {"🔴":"alert-red","🟡":"alert-amber","🟢":"alert-green"}.get(level,"alert-blue")
    ico = {"🔴":"⚠️","🟡":"⚠️","🟢":"✅"}.get(level,"ℹ️")
    return f"""<div class="alert-card {cls}"><strong>{ico} {title}</strong><br>
    <span style="font-size:12px;opacity:.85">→ {action}</span></div>"""

# ─── PARSING ─────────────────────────────────────────────────────────────────
def extract_periode(b):
    try:
        df_raw = pd.read_excel(b, header=None, dtype=str)
        last = str(df_raw.iloc[-1, 0]) if not df_raw.empty else ""
        m = re.search(r"après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})", last)
        if m: return f"{m.group(1)} → {m.group(2)}"
    except: pass
    return "Période inconnue"

def split_last(val):
    if pd.isna(val): return ""
    s = str(val).strip(); p = s.split(" - ", 1)
    return p[-1].strip() if len(p) > 1 else s

def split_code(val):
    if pd.isna(val): return ""
    return str(val).strip().split(" - ", 1)[0].strip()

@st.cache_data(show_spinner=False)
def parse_file(file_bytes, filename):
    periode = extract_periode(BytesIO(file_bytes))
    df = pd.read_excel(BytesIO(file_bytes), dtype=str)
    col_map = {}
    for col in df.columns:
        c = col.lower().replace(" ","").replace("_","")
        if "rayon"    in c: col_map["rayon"]   = col
        if "famille"  in c and "sous" not in c: col_map.setdefault("famille", col)
        if "sousfamille" in c or ("sous" in c and "famille" in c): col_map["sfam"] = col
        if "article"  in c: col_map["article"] = col
        if "sitenom"  in c or "site" in c: col_map.setdefault("site", col)
        if c in ("ca","caht") or (c.startswith("ca") and "promo" not in c and "hors" not in c and len(c)<=4):
            col_map.setdefault("ca", col)
        if c in ("marge","marge€"): col_map.setdefault("marge", col)
        if "cahors"    in c: col_map.setdefault("ca_hp",    col)
        if "cahtpromo" in c or ("caht" in c and "promo" in c): col_map.setdefault("ca_promo", col)
        if "qtévente"  in c or "qtevente" in c or c in ("qté","qte","quantite"): col_map.setdefault("qte", col)
        if "casse"     in c and "valeur" in c: col_map.setdefault("casse_v", col)
    cols = list(df.columns)
    defaults = ["rayon","famille","sfam","article","site","ca","marge",None,"ca_hp",None,None,"ca_promo",None,None,None,"qte","casse_v"]
    for i, key in enumerate(defaults):
        if key and key not in col_map and i < len(cols): col_map[key] = cols[i]
    def gcol(k): return col_map.get(k)
    records = []
    for _, row in df.iterrows():
        art    = row.get(gcol("article"), None) if gcol("article") else None
        site   = row.get(gcol("site"),    None) if gcol("site")    else None
        ca_raw = row.get(gcol("ca"),      None) if gcol("ca")      else None
        if pd.isna(art) or pd.isna(site): continue
        if str(site).strip() in ("Total","nan",""): continue
        try: ca_val = float(str(ca_raw).replace(" ","").replace(",","."))
        except: continue
        def sf(key):
            v = row.get(gcol(key), None) if gcol(key) else None
            try: return float(str(v).replace(" ","").replace(",","."))
            except: return 0.0
        records.append({
            "rayon":    split_last(row.get(gcol("rayon"),"")).upper(),
            "famille":  split_last(row.get(gcol("famille"),"")),
            "sfam":     split_last(row.get(gcol("sfam"),"")),
            "art_code": split_code(art),
            "art_lib":  split_last(art),
            "site":     split_last(site),
            "ca": ca_val, "marge": sf("marge"), "ca_hp": sf("ca_hp"),
            "ca_promo": sf("ca_promo"), "qte": sf("qte"), "casse_v": sf("casse_v"),
        })
    return pd.DataFrame(records), periode

def agg(df, keys):
    return df.groupby(keys).agg(
        ca=("ca","sum"), marge=("marge","sum"),
        ca_hp=("ca_hp","sum"), ca_promo=("ca_promo","sum"),
        qte=("qte","sum"), casse_v=("casse_v","sum")
    ).reset_index()

def merge_periods(curr, prev, keys):
    m = curr.merge(prev[keys+["ca","marge","qte"]], on=keys, how="outer", suffixes=("","_p")).fillna(0)
    m["evol_ca"] = m.apply(lambda r: evol_pct(r["ca"],    r["ca_p"]),    axis=1)
    m["evol_mg"] = m.apply(lambda r: evol_pct(r["marge"], r["marge_p"]), axis=1)
    m["tx_marge"]= m["marge"] / m["ca"].replace(0, np.nan) * 100
    return m

# ─── EXPORT EXCEL ────────────────────────────────────────────────────────────
def gen_excel(df_curr, df_prev, p_curr, p_prev):
    wb = Workbook()
    RED_F = PatternFill("solid", fgColor="FCE4E4")
    AMB_F = PatternFill("solid", fgColor="FEF3CD")
    GRN_F = PatternFill("solid", fgColor="D6F0D6")
    HDR_F = PatternFill("solid", fgColor="1C3557")
    HDR_T = Font(bold=True, color="FFFFFF", size=11)
    CTR   = Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, data_rows, headers, title):
        ws.append([title]); ws.cell(1,1).font = Font(bold=True, size=13)
        ws.append([]); ws.append(headers)
        for i,h in enumerate(headers,1):
            c = ws.cell(3,i); c.fill=HDR_F; c.font=HDR_T; c.alignment=CTR
        for row in data_rows: ws.append(row)
        NEU_F = PatternFill("solid", fgColor="FFFFFF")
        ec = [i+1 for i,h in enumerate(headers) if "évol" in h.lower()]
        for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in r:
                if cell.column in ec and isinstance(cell.value,(int,float)):
                    cell.fill = RED_F if cell.value<=-10 else AMB_F if cell.value<=-5 else GRN_F if cell.value>=5 else NEU_F
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(col[0].value or ""))+4,12)

    ws1 = wb.active; ws1.title = "Synthèse"
    m = merge_periods(agg(df_curr,["rayon"]), agg(df_prev,["rayon"]), ["rayon"])
    write_sheet(ws1,
        [[r["rayon"],int(r["ca"]),int(r["ca_p"]),
          round(r["evol_ca"],1) if r["evol_ca"] is not None else None,
          int(r["marge"]),round(r["tx_marge"],1),
          round(r["evol_mg"],1) if r["evol_mg"] is not None else None,
          int(r["qte"]),int(r["qte_p"])] for _,r in m.iterrows()],
        ["Rayon","CA S actuelle","CA réf","Évol CA %","Marge","Tx Marge %","Évol Marge %","Qté","Qté réf"],
        f"Synthèse par rayon — {p_curr} vs {p_prev}")

    mf = merge_periods(agg(df_curr,["rayon","famille"]), agg(df_prev,["rayon","famille"]), ["rayon","famille"])
    for rayon in sorted(df_curr["rayon"].unique()):
        ws = wb.create_sheet(title=rayon[:30])
        sub = mf[mf["rayon"]==rayon].sort_values("ca",ascending=False)
        write_sheet(ws,
            [[r["famille"],int(r["ca"]),int(r["ca_p"]),
              round(r["evol_ca"],1) if r["evol_ca"] is not None else None,
              int(r["marge"]),round(r["tx_marge"],1),int(r["qte"]),int(r["qte_p"])]
             for _,r in sub.iterrows()],
            ["Famille","CA S actuelle","CA réf","Évol CA %","Marge","Tx Marge %","Qté","Qté réf"],
            f"{rayon} — {p_curr} vs {p_prev}")

    ws_top = wb.create_sheet("Top 30 articles")
    mt = merge_periods(
        agg(df_curr,["art_code","art_lib","rayon","famille"]),
        agg(df_prev,["art_code","art_lib","rayon","famille"]),
        ["art_code","art_lib","rayon","famille"]
    ).sort_values("ca",ascending=False).head(30)
    write_sheet(ws_top,
        [[r["art_code"],r["art_lib"],r["rayon"],r["famille"],int(r["ca"]),int(r["ca_p"]),
          round(r["evol_ca"],1) if r["evol_ca"] is not None else None,
          int(r["marge"]),round(r["tx_marge"],1),int(r["qte"])] for _,r in mt.iterrows()],
        ["Code","Libellé","Rayon","Famille","CA S","CA réf","Évol CA %","Marge","Tx Marge %","Qté"],
        f"Top 30 articles — {p_curr}")

    ws_flop = wb.create_sheet("Réf à 0 vente")
    arts_curr = set(df_curr[df_curr["qte"]>0]["art_code"])
    arts_prev = df_prev.groupby("art_code").agg(ca_p=("ca","sum"),qte_p=("qte","sum")).reset_index()
    zero = arts_prev[(~arts_prev["art_code"].isin(arts_curr))&(arts_prev["ca_p"]>0)]
    zero = zero.merge(df_prev[["art_code","art_lib","rayon","famille"]].drop_duplicates("art_code"),on="art_code",how="left")
    write_sheet(ws_flop,
        [[r["art_code"],r["art_lib"],r["rayon"],r["famille"],int(r["ca_p"]),int(r["qte_p"]),"Vérifier rupture / déréf"]
         for _,r in zero.sort_values("ca_p",ascending=False).iterrows()],
        ["Code","Libellé","Rayon","Famille","CA réf","Qté réf","Action"],
        f"Références sans vente — {p_curr}")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Navigation</div>", unsafe_allow_html=True)
    st.page_link("app.py",                                  label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",      label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",               label="📈  Ventes PBI")
    st.page_link("pages/03_📦_Detention_Top_CA.py",         label="📦  Détention Top CA",   disabled=True)
    st.page_link("pages/04_💸_Performance_Promo.py",        label="💸  Performance Promo",  disabled=True)
    st.page_link("pages/05_🏪_Suivi_Implantation.py",       label="🏪  Suivi Implantation", disabled=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichiers</div>", unsafe_allow_html=True)
    f_curr = st.file_uploader("Semaine en cours (S actuelle)", type=["xlsx","xls"], key="curr")
    f_prev = st.file_uploader("Semaine de référence (S-1)",   type=["xlsx","xls"], key="prev")

# ─── PAGE PRINCIPALE ──────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📈 Suivi Ventes PBI</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Analyse hebdomadaire · EPICERIE · BOISSONS · DROGUERIE · PARFUMERIE HYGIENE</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL ─────────────────────────────────────────────────────────
if not f_curr or not f_prev:
    st.markdown("---")

    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Le <strong>Suivi Ventes PBI</strong> compare les performances hebdomadaires de vente entre deux périodes.
  Il analyse le CA, la marge et les quantités par rayon, magasin et famille — avec alertes automatiques
  sur les reculs significatifs et identification des références sans vente.
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Colonnes attendues dans le fichier PBI</div>", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    for i, (col_name, (desc, example)) in enumerate(REQUIRED_COLS.items()):
        with (c1 if i < 4 else c2):
            st.markdown(f"""
<div class='col-required'>
  <div style='font-size:16px;margin-top:1px'>{'📅' if 'Date' in col_name else '🔢' if col_name in ('CA','Marge','Qté Vente') else '🏷️'}</div>
  <div>
    <div class='col-name'>{col_name}</div>
    <div class='col-desc'>{desc}</div>
    <div class='col-example'>{example}</div>
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>📌 Format du fichier</strong><br>
  Export PBI standard avec hiérarchie <strong>Rayon → Famille → Sous-Famille → Article → Site</strong>.<br>
  La <strong>période est lue automatiquement</strong> depuis la ligne de filtre en bas du fichier.<br>
  Les lignes de sous-totaux (<em>Total</em>) sont ignorées automatiquement.
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("⬆️ Charge les deux exports PBI dans la sidebar pour lancer l'analyse.")
    st.stop()

# ─── PARSING ─────────────────────────────────────────────────────────────────
with st.spinner("Lecture des fichiers…"):
    bytes_curr = f_curr.read()
    bytes_prev = f_prev.read()
    df_curr, p_curr = parse_file(bytes_curr, f_curr.name)
    df_prev, p_prev = parse_file(bytes_prev, f_prev.name)

if df_curr.empty or df_prev.empty:
    st.error("Impossible de lire les données. Vérifier le format des fichiers.")
    st.stop()

# ─── FILTRES SIDEBAR ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtres</div>", unsafe_allow_html=True)
    sel_rayon = st.multiselect("Rayon",   sorted(set(df_curr["rayon"])   | set(df_prev["rayon"])),   default=sorted(set(df_curr["rayon"])   | set(df_prev["rayon"])))
    sel_site  = st.multiselect("Magasin", sorted(set(df_curr["site"])    | set(df_prev["site"])),    default=sorted(set(df_curr["site"])    | set(df_prev["site"])))
    sel_fam   = st.multiselect("Famille", sorted(set(df_curr["famille"]) | set(df_prev["famille"])), default=sorted(set(df_curr["famille"]) | set(df_prev["famille"])))
    st.markdown("---")
    st.caption(f"**S actuelle :** {p_curr}")
    st.caption(f"**Référence :**  {p_prev}")

dfc = df_curr[df_curr["rayon"].isin(sel_rayon) & df_curr["site"].isin(sel_site) & df_curr["famille"].isin(sel_fam)]
dfp = df_prev[df_prev["rayon"].isin(sel_rayon) & df_prev["site"].isin(sel_site) & df_prev["famille"].isin(sel_fam)]

# ─── KPIs ────────────────────────────────────────────────────────────────────
ca_c  = dfc["ca"].sum();    ca_p  = dfp["ca"].sum()
mg_c  = dfc["marge"].sum(); mg_p  = dfp["marge"].sum()
tx_c  = mg_c/ca_c*100 if ca_c else 0
tx_p  = mg_p/ca_p*100 if ca_p else 0
poids_promo = dfc["ca_promo"].sum()/ca_c*100 if ca_c else 0
e_ca  = evol_pct(ca_c, ca_p)
e_mg  = evol_pct(mg_c, mg_p)

st.markdown(f"<div class='section-label'>{p_curr} · vs {p_prev}</div>", unsafe_allow_html=True)
k1,k2,k3,k4 = st.columns(4)
k1.metric("CA HT",         fmt_fcfa(ca_c),  delta_str(e_ca))
k2.metric("Marge",         fmt_fcfa(mg_c),  delta_str(e_mg))
k3.metric("Taux de marge", f"{tx_c:.1f}%",  f"{tx_c-tx_p:+.1f} pt vs réf" if tx_p else "")
k4.metric("Poids promo",   f"{poids_promo:.1f}%", "")

# ─── ALERTES ─────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Alertes &amp; actions prioritaires</div>", unsafe_allow_html=True)

alerts = []
if e_ca is not None:
    if e_ca <= -15: alerts.append(("🔴", f"CA global {e_ca:.1f}% vs réf", "Analyser la part conjoncturelle vs structurelle. Reprioriser les commandes urgentes."))
    elif e_ca <= -5: alerts.append(("🟡", f"CA global {e_ca:.1f}% vs réf", "Identifier les rayons moteurs du recul. Relancer les acheteurs concernés."))

by_ray = merge_periods(agg(dfc,["rayon"]), agg(dfp,["rayon"]), ["rayon"])
for _, r in by_ray.iterrows():
    e = r["evol_ca"]
    if e is not None and e <= -20: alerts.append(("🔴", f"{r['rayon']} : {e:.1f}% vs réf", f"Identifier les familles impactées. Vérifier ruptures et commandes sur {r['rayon']}."))
    elif e is not None and e <= -10: alerts.append(("🟡", f"{r['rayon']} : {e:.1f}% vs réf", "Analyser les sous-familles en recul et ajuster les réassorts."))

by_fam = merge_periods(agg(dfc,["rayon","famille"]), agg(dfp,["rayon","famille"]), ["rayon","famille"])
for _, r in by_fam[by_fam["ca_p"]>2_000_000].iterrows():
    if r["evol_ca"] is not None and r["evol_ca"] <= -30:
        alerts.append(("🔴", f"{r['famille']} ({r['rayon']}) : {r['evol_ca']:.1f}%", "Famille à fort impact CA. Vérifier assortiment et disponibilité."))

by_site = merge_periods(agg(dfc,["site"]), agg(dfp,["site"]), ["site"])
for _, r in by_site[by_site["evol_ca"]>5].sort_values("evol_ca",ascending=False).head(2).iterrows():
    alerts.append(("🟢", f"{r['site']} : +{r['evol_ca']:.1f}% vs réf", "Point de vente en progression. Identifier les leviers transposables."))

arts_zero = set(dfp[dfp["qte"]>0]["art_code"]) - set(dfc[dfc["qte"]>0]["art_code"])
if arts_zero: alerts.append(("🟡", f"{len(arts_zero)} références sans vente (avaient vendu en réf)", "Vérifier rupture de stock ou déréférencement non planifié."))

if not alerts:
    st.success("✅ Aucune alerte critique — situation stable vs référence.")
else:
    st.markdown("".join(alert_html(*a) for a in alerts), unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 CA par rayon", "🏪 Par magasin", "📉 Familles en recul",
    "🏆 Top 30 articles", "⚠️ Réf à 0 vente"
])

# ═══ TAB 1 ════════════════════════════════════════════════════════════════════
with tab1:
    by_ray_s = by_ray.sort_values("ca", ascending=False)
    fig = go.Figure()
    fig.add_bar(
        x=by_ray_s["rayon"], y=by_ray_s["ca"], name="S actuelle",
        marker_color=[COLORS.get(r, COLOR_DEFAULT) for r in by_ray_s["rayon"]],
        text=[fmt_fcfa(v) for v in by_ray_s["ca"]], textposition="outside",
        marker_line_width=0,
    )
    fig.add_bar(
        x=by_ray_s["rayon"], y=by_ray_s["ca_p"], name="Référence",
        marker_color=[COLORS_FADE.get(r, COLOR_DEFAULT_FADE) for r in by_ray_s["rayon"]],
        marker_line_width=0,
    )
    fig.update_layout(
        plot_bgcolor=PL_BG, paper_bgcolor=PL_BG, font=PL_FONT,
        barmode="group", height=370, margin=dict(t=16,b=16,l=8,r=8),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=12)),
        yaxis=dict(showgrid=True, gridcolor=PL_GRID, title=""),
        xaxis=dict(showgrid=False, title=""),
    )
    st.plotly_chart(fig, use_container_width=True)

    disp = by_ray_s[["rayon","ca","ca_p","evol_ca","marge","tx_marge","qte","qte_p"]].copy()
    disp.columns = ["Rayon","CA S actuelle","CA réf","Évol CA %","Marge","Tx Marge %","Qté","Qté réf"]
    for c in ["CA S actuelle","CA réf","Marge"]: disp[c] = disp[c].apply(fmt_fcfa)
    disp["Évol CA %"]  = disp["Évol CA %"].apply(lambda x: f"{x:+.1f}%" if x is not None else "—")
    disp["Tx Marge %"] = disp["Tx Marge %"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
    disp["Qté"]        = disp["Qté"].apply(lambda x: f"{int(x):,}")
    disp["Qté réf"]    = disp["Qté réf"].apply(lambda x: f"{int(x):,}")
    st.dataframe(disp, use_container_width=True, hide_index=True)

# ═══ TAB 2 ════════════════════════════════════════════════════════════════════
with tab2:
    by_site_s = by_site.sort_values("evol_ca")
    fig2 = go.Figure(go.Bar(
        x=by_site_s["evol_ca"].round(1), y=by_site_s["site"], orientation="h",
        marker_color=[GREEN if v >= 0 else (RED if v <= -15 else AMBER) for v in by_site_s["evol_ca"].fillna(0)],
        text=[f"{v:+.1f}%" if v is not None else "—" for v in by_site_s["evol_ca"]],
        textposition="outside", marker_line_width=0,
    ))
    fig2.update_layout(
        plot_bgcolor=PL_BG, paper_bgcolor=PL_BG, font=PL_FONT,
        height=380, margin=dict(t=10,b=10,l=10,r=80),
        xaxis=dict(title="Évolution CA %", ticksuffix="%", showgrid=True, gridcolor=PL_GRID),
        yaxis=dict(showgrid=False, title=""),
    )
    fig2.add_vline(x=0, line_width=1, line_color="#E5E5EA")
    st.plotly_chart(fig2, use_container_width=True)

    disp2 = by_site_s[["site","ca","ca_p","evol_ca","marge","tx_marge"]].copy()
    disp2.columns = ["Magasin","CA S actuelle","CA réf","Évol CA %","Marge","Tx Marge %"]
    for c in ["CA S actuelle","CA réf","Marge"]: disp2[c] = disp2[c].apply(fmt_fcfa)
    disp2["Évol CA %"]  = disp2["Évol CA %"].apply(lambda x: f"{x:+.1f}%" if x is not None else "—")
    disp2["Tx Marge %"] = disp2["Tx Marge %"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
    st.dataframe(disp2, use_container_width=True, hide_index=True)

# ═══ TAB 3 ════════════════════════════════════════════════════════════════════
with tab3:
    col_sl, _ = st.columns([1,3])
    with col_sl:
        seuil = st.slider("Seuil recul (%)", -50, -5, -15, key="seuil_fam")
    fam_recul = by_fam[by_fam["evol_ca"].notna() & (by_fam["evol_ca"]<=seuil) & (by_fam["ca_p"]>=500_000)].sort_values("evol_ca")
    if fam_recul.empty:
        st.info(f"Aucune famille avec un recul ≥ {abs(seuil)}%.")
    else:
        fig3 = px.bar(fam_recul.head(20), x="evol_ca", y="famille",
                      orientation="h", color="rayon", color_discrete_map=COLORS,
                      labels={"evol_ca":"Évol CA %","famille":""})
        fig3.update_layout(
            plot_bgcolor=PL_BG, paper_bgcolor=PL_BG, font=PL_FONT,
            height=max(340, len(fam_recul.head(20))*36+80),
            margin=dict(t=16,b=16,l=8,r=8),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=12)),
            xaxis=dict(ticksuffix="%", showgrid=True, gridcolor=PL_GRID),
            yaxis=dict(showgrid=False, title=""),
        )
        fig3.add_vline(x=0, line_width=1, line_color="#E5E5EA")
        st.plotly_chart(fig3, use_container_width=True)

        disp3 = fam_recul[["rayon","famille","ca","ca_p","evol_ca","marge","tx_marge","qte"]].copy()
        disp3.columns = ["Rayon","Famille","CA S","CA réf","Évol %","Marge","Tx mg %","Qté"]
        for c in ["CA S","CA réf","Marge"]: disp3[c] = disp3[c].apply(fmt_fcfa)
        disp3["Évol %"]  = disp3["Évol %"].apply(lambda x: f"{x:+.1f}%")
        disp3["Tx mg %"] = disp3["Tx mg %"].apply(lambda x: f"{x:.1f}%")
        st.dataframe(disp3, use_container_width=True, hide_index=True)

# ═══ TAB 4 ════════════════════════════════════════════════════════════════════
with tab4:
    by_art = merge_periods(
        agg(dfc,["art_code","art_lib","rayon","famille"]),
        agg(dfp,["art_code","art_lib","rayon","famille"]),
        ["art_code","art_lib","rayon","famille"]
    ).sort_values("ca",ascending=False).head(30)
    disp4 = by_art[["art_code","art_lib","rayon","famille","ca","ca_p","evol_ca","marge","tx_marge","qte"]].copy()
    disp4.columns = ["Code","Libellé","Rayon","Famille","CA S","CA réf","Évol %","Marge","Tx mg %","Qté"]
    for c in ["CA S","CA réf","Marge"]: disp4[c] = disp4[c].apply(fmt_fcfa)
    disp4["Évol %"]  = disp4["Évol %"].apply(lambda x: f"{x:+.1f}%" if x is not None else "—")
    disp4["Tx mg %"] = disp4["Tx mg %"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
    disp4["Qté"]     = disp4["Qté"].apply(lambda x: f"{int(x):,}")
    st.dataframe(disp4, use_container_width=True, hide_index=True)

# ═══ TAB 5 ════════════════════════════════════════════════════════════════════
with tab5:
    arts_prev_grp = dfp.groupby(["art_code","art_lib","rayon","famille"]).agg(
        ca_p=("ca","sum"), qte_p=("qte","sum")
    ).reset_index()
    zero_df = arts_prev_grp[
        (~arts_prev_grp["art_code"].isin(set(dfc[dfc["qte"]>0]["art_code"]))) &
        (arts_prev_grp["ca_p"] > 0)
    ].sort_values("ca_p", ascending=False)

    if zero_df.empty:
        st.success("✅ Toutes les références actives ont vendu en S actuelle.")
    else:
        st.warning(f"⚠️ {len(zero_df)} références sans vente en S actuelle (CA réf > 0)")
        disp5 = zero_df[["art_code","art_lib","rayon","famille","ca_p","qte_p"]].copy()
        disp5.columns = ["Code","Libellé","Rayon","Famille","CA réf","Qté réf"]
        disp5["CA réf"]  = disp5["CA réf"].apply(fmt_fcfa)
        disp5["Qté réf"] = disp5["Qté réf"].apply(lambda x: f"{int(x):,}")
        disp5["Action"]  = "Vérifier rupture / déréf"
        st.dataframe(disp5, use_container_width=True, hide_index=True)

# ─── EXPORT ──────────────────────────────────────────────────────────────────
st.markdown("---")
with st.expander("📥 Export Excel — Synthèse · 4 rayons · Top 30 · Réf à 0 vente"):
    st.caption(f"S actuelle : {p_curr} · Référence : {p_prev}")
    if st.button("Générer le fichier Excel", type="primary"):
        with st.spinner("Génération en cours…"):
            buf = gen_excel(dfc, dfp, p_curr, p_prev)
        fname = f"SmartBuyer_Ventes_{p_curr.replace('/','').replace(' ','').replace('→','_vs_')}.xlsx"
        st.download_button(
            "⬇️ Télécharger", data=buf, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
