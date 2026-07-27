"""
generate_reporting_ventes.py — SmartBuyer Hub
Génère "Reporting Ventes - AAAAMMJJ.xlsx" à partir d'un export PBI
(Rayon → Famille → Sous Famille → Article [× Site]).

Reprend telle quelle la logique métier et les onglets du module
15_📋_COPIL_Hebdo.py :
  - Dashboard COPIL (vue réseau, perf par rayon, top/flop familles,
    marge négative par site si la colonne Site est présente)
  - Destructeurs Performeurs (réseau entier)
  - DP - <Rayon> (un onglet par rayon détecté dans l'export)
  - Marges Négatives par Site (uniquement si une colonne Site/Magasin
    est détectée dans l'export)

La période affichée dans le fichier est extraite automatiquement de
la dernière cellule de la première colonne de l'export PBI
(ligne "Filtres appliqués : Date est le ou après le JJ/MM/AAAA
et est avant le JJ/MM/AAAA...").

Usage :
    python generate_reporting_ventes.py chemin/vers/export_pbi.xlsx [dossier_sortie]

Si aucun argument n'est fourni, le script cherche un .xlsx dans le
dossier courant.
"""

import io
import re
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Colonnes candidates pour la dimension Site dans l'export PBI (auto-détection)
SITE_CANDIDATES = ["Site", "Magasin", "Code Site", "Libellé Site", "Nom Site",
                   "Site de vente", "Etablissement", "Établissement", "Store"]

# ============================================================
# 🎯 CIBLES DE MARGE PAR RAYON
# ============================================================
CIBLES_DEFAUT = {
    "BOISSONS": 19.5,
    "DROGUERIE": 25.0,
    "PARFUMERIE HYGIENE": 29.0,
    "EPICERIE": 16.0,
}
CIBLE_FALLBACK = 23.5

# Colonnes candidates pour la dimension Site dans l'export PBI (auto-détection)
SITE_CANDIDATES = ["Site", "Magasin", "Code Site", "Libellé Site", "Nom Site",
                   "Site de vente", "Etablissement", "Établissement", "Store"]
def fmt(n):
    if n is None or (isinstance(n, float) and (pd.isna(n) or not np.isfinite(n))):
        return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if v is None or pd.isna(v): return "—"
    return f"{v:.{dec}f}%"

def fmt_delta(v):
    if v is None or pd.isna(v): return "—"
    return f"{v:+.1f} pts"

def rayon_key(libelle):
    s = str(libelle).upper()
    for k in CIBLES_DEFAUT:
        if k in s:
            return k
    return None

def _wavg(values, weights):
    """Moyenne pondérée robuste (NaN-safe). Renvoie NaN si aucun poids valide."""
    v = pd.Series(values, dtype=float)
    w = pd.Series(weights, dtype=float)
    m = v.notna() & w.notna() & (w > 0)
    if not m.any() or w[m].sum() == 0:
        return np.nan
    return float((v[m] * w[m]).sum() / w[m].sum())

# ============================================================
# CHARGEMENT — EXPORT ARTICLE UNIQUE (Rayon → Famille → Sous Famille → Article [× Site])
# ============================================================
def load_export(file_bytes):
    raw = pd.read_excel(io.BytesIO(file_bytes))
    raw.columns = [str(c).lstrip('\ufeff').strip() for c in raw.columns]

    perimetre = None
    note_rows = raw[raw['Rayon'].astype(str).str.startswith('Filtres', na=False)]
    if not note_rows.empty:
        perimetre = str(note_rows.iloc[0]['Rayon'])

    df = raw[raw['Rayon'].notna()].copy()
    df = df[~df['Rayon'].astype(str).str.startswith('Filtres')]
    if 'Article' not in df.columns:
        df['Article'] = np.nan

    # Auto-détection de la colonne Site (export à la maille Article × Site)
    site_col = None
    for cand in SITE_CANDIDATES:
        if cand in df.columns:
            site_col = cand
            break
    if site_col is None:
        for c in df.columns:
            cu = str(c).upper()
            if 'SITE' in cu or 'MAGASIN' in cu:
                site_col = c
                break
    return df, perimetre, site_col

# ============================================================
# AGRÉGATS RÉSEAU (robustes mono/multi-site)
# ============================================================
def kpis_globaux_rayon(df):
    g = df[df['Rayon'] == 'Total']
    if g.empty:
        return None
    # Si l'export contient un Total par site, on somme (mono-site : identique).
    ca = g['CA'].sum()
    ca_n1 = g['CA N-1'].sum()
    marge = g['Marge'].sum()
    evol_marge_pct = g.get('%Vs N-1.1', pd.Series(np.nan, index=g.index))
    with np.errstate(divide='ignore', invalid='ignore'):
        marge_n1_rows = (g['Marge'] / (1 + evol_marge_pct)).replace([np.inf, -np.inf], np.nan)
    marge_n1 = marge_n1_rows.sum() if marge_n1_rows.notna().any() else np.nan
    qte = g.get('Qté Vente', pd.Series(np.nan, index=g.index)).sum()
    qte_n1 = g.get('Qté Vente N-1', pd.Series(np.nan, index=g.index)).sum()
    poids_promo = _wavg(g.get('%CA Poids Promo', pd.Series(np.nan, index=g.index)), g['CA'])
    casse = g.get('Casse (Valeur)', pd.Series(np.nan, index=g.index)).sum()
    return {
        'ca': ca, 'ca_n1': ca_n1, 'evol_ca': ca/ca_n1 - 1 if ca_n1 else np.nan,
        'marge': marge, 'marge_n1': marge_n1,
        'tx_marge': marge/ca*100 if ca else np.nan,
        'tx_marge_n1': marge_n1/ca_n1*100 if ca_n1 and pd.notna(marge_n1) else np.nan,
        'qte': qte, 'qte_n1': qte_n1,
        'poids_promo': poids_promo * 100 if pd.notna(poids_promo) else np.nan,
        'casse': casse,
    }

def perf_par_rayon(df, cibles):
    sub = df[(df['Famille'] == 'Total') & (df['Rayon'] != 'Total')].copy()
    sub['Rayon_aff'] = sub['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    if 'Qté Vente' not in sub.columns: sub['Qté Vente'] = np.nan
    if 'Qté Vente N-1' not in sub.columns: sub['Qté Vente N-1'] = np.nan
    agg = sub.groupby('Rayon_aff', as_index=False).agg(
        CA=('CA', 'sum'), CA_N1=('CA N-1', 'sum'), Marge=('Marge', 'sum'),
        Qte=('Qté Vente', 'sum'), Qte_N1=('Qté Vente N-1', 'sum'))
    rows = []
    for _, r in agg.iterrows():
        key = rayon_key(r['Rayon_aff'])
        cible = cibles.get(key, CIBLE_FALLBACK) if key else CIBLE_FALLBACK
        tx = r['Marge']/r['CA']*100 if r['CA'] else np.nan
        rows.append({
            'Rayon': r['Rayon_aff'],
            'CA': r['CA'],
            'Évol CA %': (r['CA']/r['CA_N1']-1)*100 if r['CA_N1'] else np.nan,
            'Évol Qté %': (r['Qte']/r['Qte_N1']-1)*100 if r['Qte_N1'] else np.nan,
            'Taux Marge %': tx, 'Objectif %': cible,
            'Écart (pts)': tx - cible if pd.notna(tx) else np.nan,
        })
    return pd.DataFrame(rows).sort_values('CA', ascending=False)

def family_metrics(df):
    """Métriques niveau Famille, agrégées réseau (robuste aux exports multi-site)."""
    sub = df[(df['Sous Famille'] == 'Total') & (df['Famille'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    sub['Rayon_aff'] = sub['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    sub['Famille_aff'] = sub['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    for c in ['CA', 'CA N-1', 'Marge']:
        sub[c] = sub[c].fillna(0)
    for c in ['Qté Vente', 'Qté Vente N-1', 'Casse (Valeur)', '%CA Poids Promo',
              '%Marge Promo', '%Marge Hors Promo', '%Vs N-1.1']:
        if c not in sub.columns:
            sub[c] = np.nan

    # Marge N-1 par ligne (ratio non sommable, à calculer AVANT agrégation)
    with np.errstate(divide='ignore', invalid='ignore'):
        sub['_marge_n1'] = (sub['Marge'] / (1 + sub['%Vs N-1.1'])).replace([np.inf, -np.inf], np.nan)
    # CA promo par ligne (pour re-pondérer les % promo après agrégation)
    sub['_ca_promo'] = sub['%CA Poids Promo'] * sub['CA']
    sub['_ca_hp'] = sub['CA'] - sub['_ca_promo'].fillna(0)
    sub['_mp_w'] = sub['%Marge Promo'] * sub['_ca_promo']
    sub['_mhp_w'] = sub['%Marge Hors Promo'] * sub['_ca_hp']

    agg = sub.groupby(['Rayon_aff', 'Famille_aff'], as_index=False).agg(**{
        'CA': ('CA', 'sum'), 'CA N-1': ('CA N-1', 'sum'), 'Marge': ('Marge', 'sum'),
        '_marge_n1': ('_marge_n1', 'sum'),
        'Qté Vente': ('Qté Vente', 'sum'), 'Qté Vente N-1': ('Qté Vente N-1', 'sum'),
        'Casse (Valeur)': ('Casse (Valeur)', 'sum'),
        '_ca_promo': ('_ca_promo', 'sum'), '_ca_hp': ('_ca_hp', 'sum'),
        '_mp_w': ('_mp_w', 'sum'), '_mhp_w': ('_mhp_w', 'sum'),
    })
    agg['Perte CA'] = agg['CA'] - agg['CA N-1']
    agg['Évol CA %'] = np.where(agg['CA N-1'] > 0, (agg['CA']/agg['CA N-1']-1)*100, np.nan)
    agg['Tx Marge %'] = np.where(agg['CA'] > 0, agg['Marge']/agg['CA']*100, np.nan)
    agg['Tx Marge N-1 %'] = np.where(agg['CA N-1'] > 0, agg['_marge_n1']/agg['CA N-1']*100, np.nan)
    agg['Écart Tx Marge (pts)'] = agg['Tx Marge %'] - agg['Tx Marge N-1 %']
    agg['Évol Qté %'] = np.where(agg['Qté Vente N-1'] > 0, (agg['Qté Vente']/agg['Qté Vente N-1']-1)*100, np.nan)
    with np.errstate(divide='ignore', invalid='ignore'):
        agg['%CA Poids Promo'] = np.where(agg['CA'] > 0, agg['_ca_promo']/agg['CA'], np.nan)
        agg['%Marge Promo'] = np.where(agg['_ca_promo'] > 0, agg['_mp_w']/agg['_ca_promo'], np.nan)
        agg['%Marge Hors Promo'] = np.where(agg['_ca_hp'] > 0, agg['_mhp_w']/agg['_ca_hp'], np.nan)
        agg['%Casse (Valeur)'] = np.where(agg['CA'] > 0, agg['Casse (Valeur)']/agg['CA'], np.nan)
    return agg.drop(columns=['_marge_n1', '_ca_promo', '_ca_hp', '_mp_w', '_mhp_w'])

def top_flop_table(sub, metric, n, mode, cols, ca_floor=0, directional=True):
    base = sub[sub['CA'] > ca_floor] if ca_floor else sub
    base = base[base[metric].notna()]
    if directional:
        base = base[base[metric] > 0] if mode == 'top' else base[base[metric] < 0]
    out = base.nlargest(n, metric) if mode == 'top' else base.nsmallest(n, metric)
    return out[['Rayon_aff','Famille_aff'] + cols].reset_index(drop=True)

# ============================================================
# VUE ARTICLE — 🔧 FIX DOUBLONS
# L'export PBI est à la maille Article × Site : un même article vendu dans
# plusieurs magasins produit plusieurs lignes. On agrège au niveau réseau
# AVANT tout classement, sinon les Top/Flop contiennent des doublons.
# ============================================================
def prep_articles(df):
    art = df[df['Article'].notna() & (df['Article'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    art['Rayon_aff'] = art['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    art['Famille_aff'] = art['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    art['SousFamille_aff'] = art['Sous Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    art['Article_aff'] = art['Article'].astype(str)
    for c in ['CA', 'CA N-1', 'Marge', 'Qté Vente', 'Qté Vente N-1']:
        if c not in art.columns: art[c] = 0
        art[c] = art[c].fillna(0)

    # Marge N-1 recalculée AVANT agrégation (le % N-1 est un ratio par ligne)
    evol_marge = art.get('%Vs N-1.1', pd.Series(np.nan, index=art.index))
    with np.errstate(divide='ignore', invalid='ignore'):
        art['Marge N-1 (calc)'] = (art['Marge'] / (1 + evol_marge)).replace([np.inf, -np.inf], np.nan)

    # 🔧 Agrégation réseau : 1 article = 1 ligne, tous sites confondus
    group_cols = ['Rayon_aff', 'Famille_aff', 'SousFamille_aff', 'Article_aff']
    sum_cols = ['CA', 'CA N-1', 'Marge', 'Marge N-1 (calc)', 'Qté Vente', 'Qté Vente N-1']
    art = art.groupby(group_cols, as_index=False)[sum_cols].sum(min_count=1)
    art['Marge N-1 (calc)'] = art['Marge N-1 (calc)'].where(art['Marge N-1 (calc)'].notna(), np.nan)
    for c in ['CA', 'CA N-1', 'Marge', 'Qté Vente', 'Qté Vente N-1']:
        art[c] = art[c].fillna(0)

    art['Tx Marge %'] = np.where(art['CA'] > 0, art['Marge']/art['CA']*100, np.nan)
    art['Tx Marge N-1 % (calc)'] = np.where(art['CA N-1'] > 0, art['Marge N-1 (calc)']/art['CA N-1']*100, np.nan)
    art['Écart Tx Marge (pts)'] = art['Tx Marge %'] - art['Tx Marge N-1 % (calc)']
    art['Gain Marge (FCFA)'] = art['Marge'] - art['Marge N-1 (calc)']
    art['Variation Qté'] = art['Qté Vente'] - art['Qté Vente N-1']
    art['Perte CA (FCFA)'] = art['CA'] - art['CA N-1']
    return art

def destructeurs_performeurs(art, n=15, seuil_ca=100_000):
    res = {}
    res['A_marge_neg'] = art[art['Marge'] < 0].nsmallest(n, 'Marge')
    pos = art[art['Marge'] >= 0].copy()
    deg = pos[pos['Écart Tx Marge (pts)'].notna() & (pos['Écart Tx Marge (pts)'] < 0)]
    res['B_degrad_marge'] = deg.nsmallest(n, 'Écart Tx Marge (pts)')
    mat = art[art['CA'] > seuil_ca]
    gain = mat[mat['Gain Marge (FCFA)'].notna() & (mat['Gain Marge (FCFA)'] > 0)]
    res['C_perf_gain_marge'] = gain.nlargest(n, 'Gain Marge (FCFA)')
    mat_n1 = mat[mat['CA N-1'] > 0].assign(_evol=lambda d: d['CA']/d['CA N-1']-1)
    mat_n1_pos = mat_n1[mat_n1['_evol'] > 0]
    res['D_croissance_ca'] = mat_n1_pos.nlargest(n, '_evol')
    baisse = art[art['Perte CA (FCFA)'] < 0]
    res['E_baisse_ca'] = baisse.nsmallest(n, 'Perte CA (FCFA)')
    hausse_q = art[art['Variation Qté'] > 0]
    res['F_hausse_qte'] = hausse_q.nlargest(n, 'Variation Qté')
    baisse_q = art[art['Variation Qté'] < 0]
    res['G_baisse_qte'] = baisse_q.nsmallest(n, 'Variation Qté')
    return res

# ============================================================
# 🆕 MARGE NÉGATIVE PAR SITE
# Exploite la dimension Site de l'export (celle qui créait les doublons).
# ============================================================
def marge_negative_par_site(df, site_col):
    """Synthèse par site : CA site, nb articles en marge négative, marge négative
    cumulée, poids vs CA site. Renvoie None si la colonne Site est absente."""
    if not site_col or site_col not in df.columns:
        return None
    art = df[df['Article'].notna() & (df['Article'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    art = art[art[site_col].notna() & (art[site_col].astype(str) != 'Total')]
    if art.empty:
        return None
    art['CA'] = art['CA'].fillna(0)
    art['Marge'] = art['Marge'].fillna(0)

    ca_site = art.groupby(site_col, as_index=False)['CA'].sum().rename(columns={'CA': 'CA Site'})
    neg = art[art['Marge'] < 0]
    g = neg.groupby(site_col, as_index=False).agg(
        **{'Nb Articles Nég.': ('Marge', 'count'), 'Marge Négative (FCFA)': ('Marge', 'sum')})
    out = ca_site.merge(g, on=site_col, how='left')
    out['Nb Articles Nég.'] = out['Nb Articles Nég.'].fillna(0).astype(int)
    out['Marge Négative (FCFA)'] = out['Marge Négative (FCFA)'].fillna(0)
    out['% Marge Nég / CA Site'] = np.where(out['CA Site'] > 0,
                                            out['Marge Négative (FCFA)']/out['CA Site']*100, np.nan)
    out = out.rename(columns={site_col: 'Site'})
    return out.sort_values('Marge Négative (FCFA)').reset_index(drop=True)

def _site_format(site):
    """Format magasin dérivé du libellé site (Hyper / Market / Supeco)."""
    s = str(site).upper()
    if 'HYPER' in s: return 'Hyper'
    if 'MARKET' in s: return 'Market'
    if 'SUPECO' in s: return 'Supeco'
    return '—'

def detail_marge_neg_site(df, site_col):
    """Détail EXHAUSTIF des articles en marge négative, ligne article × site
    (format du récap de référence 'Marges Négatives par Site').
    Colonnes : Code Art. / Article / Rayon / Famille / Magasin / Format /
    CA / Marge / Tx Marge % / Qté — trié par libellé article."""
    if not site_col or site_col not in df.columns:
        return None
    art = df[df['Article'].notna() & (df['Article'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    art = art[art[site_col].notna() & (art[site_col].astype(str) != 'Total')]
    art['CA'] = art['CA'].fillna(0)
    art['Marge'] = art['Marge'].fillna(0)
    art = art[art['Marge'] < 0].copy()
    if art.empty:
        return None
    art['Rayon_aff'] = art['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    art['Famille_aff'] = art['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    # Séparation "code - libellé" de l'export PBI
    art_str = art['Article'].astype(str)
    has_sep = art_str.str.contains(' - ')
    art['Code Art.'] = np.where(has_sep, art_str.str.split(' - ').str[0].str.strip(), '')
    art['Article_aff'] = np.where(has_sep,
                                  art_str.str.split(' - ', n=1).str[1].str.strip(), art_str)
    art['Site'] = art[site_col].astype(str).str.split(' - ').str[-1].str.strip()
    art['Format'] = art['Site'].map(_site_format)
    art['Tx Marge %'] = np.where(art['CA'] > 0, art['Marge']/art['CA']*100, np.nan)
    qte = art.get('Qté Vente', pd.Series(np.nan, index=art.index))
    art['Qté'] = qte.fillna(0)
    return art.sort_values('Article_aff')[
        ['Code Art.', 'Article_aff', 'Rayon_aff', 'Famille_aff', 'Site', 'Format',
         'CA', 'Marge', 'Tx Marge %', 'Qté']].reset_index(drop=True)

DISPLAY_COLS = ['Rayon_aff','Famille_aff','SousFamille_aff','Article_aff']
DISPLAY_RENAME = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article'}

def build_excel_full(k, perf, fam, n_top, art_res, perimetre, mns=None, mns_detail=None,
                      art_full=None, seuil_ca=100_000):
    BLUE_H = "FF007AFF"; DARK_H = "FF1D1D1F"; RED_H = "FFFF3B30"; GREEN_H = "FF34C759"
    WHITE_H = "FFFFFFFF"; LGREY_H = "FFF2F2F7"; ARIAL = "Arial"
    thin = Side(style="thin", color="FFD1D1D6")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    ACC = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
    ACC_SIGNED = '_-* +#,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
    QTY = "#,##0"
    PTS = '+0.00" pts";-0.00" pts"'
    PCT = "0.0%"
    PCT2 = "0.00%"

    def fmt_value(kind, v):
        if kind == 'pct100' and v is not None and pd.notna(v):
            return v / 100
        return v

    def fmt_code(kind):
        return {'amount': ACC, 'amount_signed': ACC_SIGNED, 'qty': QTY,
                'pct100': PCT, 'pct1': PCT, 'pts': PTS}.get(kind, "General")

    def section_bar(ws, row, ncols, text, color=DARK_H):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name=ARIAL, bold=True, size=11, color=WHITE_H)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def header_row(ws, row, labels):
        for i, lbl in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lbl)
            c.font = Font(name=ARIAL, bold=True, size=10, color=WHITE_H)
            c.fill = PatternFill("solid", fgColor=BLUE_H)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_row(ws, row, values, zebra=False, left_cols=()):
        fillc = LGREY_H if zebra else "FFFFFFFF"
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name=ARIAL, size=10)
            c.border = box
            c.fill = PatternFill("solid", fgColor=fillc)
            c.alignment = Alignment(horizontal="left" if i in left_cols else "center")

    def autosize(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    wb = Workbook()

    # ============== FEUILLE 1 : DASHBOARD COPIL ==============
    ws = wb.active
    ws.title = "Dashboard COPIL"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "DASHBOARD COPIL HEBDO — PGC (Épicerie · Boissons · Droguerie · Parfumerie Hygiène)"
    ws["A1"].font = Font(name=ARIAL, bold=True, size=14, color=WHITE_H)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_H)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws["A2"] = "Période :"; ws["A2"].font = Font(name=ARIAL, bold=True, size=10)
    ws["B2"] = "Voir export source"; ws["B2"].font = Font(name=ARIAL, size=10, color="FF0000FF", bold=True)
    if perimetre:
        ws["A3"] = "Périmètre :"; ws["A3"].font = Font(name=ARIAL, size=9, italic=True, color="FF86868B")
        ws["B3"] = perimetre.replace("\n", " ")[:200]
        ws["B3"].font = Font(name=ARIAL, size=9, italic=True, color="FF86868B")

    r = 5
    section_bar(ws, r, 8, "1.  VUE D'ENSEMBLE RÉSEAU"); r += 1
    header_row(ws, r, ["Indicateur", "Cette semaine", "N-1", "Évolution"]); r += 1
    evo_tx = k['tx_marge'] - k['tx_marge_n1'] if pd.notna(k['tx_marge_n1']) else None
    evol_qte = (k['qte']/k['qte_n1']-1) if k['qte_n1'] else None
    pct_casse = k['casse']/k['ca'] if k['ca'] else None
    kpi_rows = [
        ("CA (FCFA)",          k['ca'],            k['ca_n1'],   k['evol_ca'], 'amount', 'pct1'),
        ("Marge (FCFA)",       k['marge'],         k['marge_n1'],None,         'amount', None),
        ("Taux de marge",      k['tx_marge']/100,  (k['tx_marge_n1']/100 if pd.notna(k['tx_marge_n1']) else None), evo_tx, 'pct1', 'pts'),
        ("Qté vendue",         k['qte'],           k['qte_n1'],  evol_qte,     'qty',    'pct1'),
        ("Poids Promo (% CA)", k['poids_promo']/100 if pd.notna(k['poids_promo']) else None, None, None, 'pct1', None),
        ("Casse (FCFA)",       k['casse'],         None,         pct_casse,    'amount', 'pct1'),
    ]
    r0kpi = r
    for i, (label, v, n1, evo, kind_v, kind_evo) in enumerate(kpi_rows):
        zebra = i % 2 == 1
        fillc = LGREY_H if zebra else "FFFFFFFF"
        for col in range(1, 5):
            cc = ws.cell(row=r, column=col)
            cc.fill = PatternFill("solid", fgColor=fillc); cc.border = box; cc.font = Font(name=ARIAL, size=10)
        ws.cell(row=r, column=1, value=label).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=r, column=2, value=v); ws.cell(row=r, column=2).number_format = fmt_code(kind_v)
        if n1 is not None and pd.notna(n1):
            ws.cell(row=r, column=3, value=n1); ws.cell(row=r, column=3).number_format = fmt_code(kind_v)
        if evo is not None and pd.notna(evo):
            ws.cell(row=r, column=4, value=evo)
            ws.cell(row=r, column=4).number_format = fmt_code(kind_evo)
        r += 1
    ws.conditional_formatting.add(f"D{r0kpi}:D{r-1}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFFD6D4")))
    ws.conditional_formatting.add(f"D{r0kpi}:D{r-1}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFD7F5DE")))
    r += 1

    section_bar(ws, r, 8, "2.  PERFORMANCE PAR RAYON VS OBJECTIFS MARGE (MÉTI)"); r += 1
    header_row(ws, r, ["Rayon", "CA (FCFA)", "Évol CA", "Évol Qté", "Taux Marge", "Objectif Méti", "Écart (pts)"]); r += 1
    r0r = r
    for i, (_, row_) in enumerate(perf.iterrows()):
        data_row(ws, r, [row_['Rayon'], row_['CA'],
                          (row_['Évol CA %']/100 if pd.notna(row_['Évol CA %']) else None),
                          (row_['Évol Qté %']/100 if pd.notna(row_['Évol Qté %']) else None),
                          (row_['Taux Marge %']/100 if pd.notna(row_['Taux Marge %']) else None),
                          row_['Objectif %']/100, row_['Écart (pts)']], zebra=(i%2==1), left_cols=(1,))
        for col, fmt_ in [(2,ACC),(3,PCT),(4,PCT),(5,PCT2),(6,PCT),(7,PTS)]:
            ws.cell(row=r, column=col).number_format = fmt_
        r += 1
    ws.conditional_formatting.add(f"G{r0r}:G{r-1}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFFD6D4")))
    ws.conditional_formatting.add(f"G{r0r}:G{r-1}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFD7F5DE")))
    r += 1

    # ---- 🆕 3. MARGE NÉGATIVE PAR SITE (synthèse) ----
    if mns is not None and not mns.empty:
        section_bar(ws, r, 8, "3.  MARGE NÉGATIVE PAR SITE", color=RED_H); r += 1
        header_row(ws, r, ["Site", "CA Site (FCFA)", "Nb Articles Nég.", "Marge Négative (FCFA)", "% vs CA Site"]); r += 1
        r0s = r
        for i, (_, row_) in enumerate(mns.iterrows()):
            data_row(ws, r, [row_['Site'], row_['CA Site'], row_['Nb Articles Nég.'],
                              row_['Marge Négative (FCFA)'],
                              (row_['% Marge Nég / CA Site']/100 if pd.notna(row_['% Marge Nég / CA Site']) else None)],
                      zebra=(i % 2 == 1), left_cols=(1,))
            for col, fmt_ in [(2, ACC), (3, QTY), (4, ACC), (5, PCT2)]:
                ws.cell(row=r, column=col).number_format = fmt_
            r += 1
        ws.conditional_formatting.add(f"D{r0s}:D{r-1}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFFD6D4")))
        r += 1

    section_bar(ws, r, 8, "4.  DÉTAIL PAR FAMILLE — TOUTES FAMILLES (sans objectif)"); r += 1
    header_row(ws, r, ["Rayon", "Famille", "CA", "Évol CA %", "Marge", "Taux Marge %", "Qté Vente", "Évol Qté %"]); r += 1
    fam_sorted = fam.sort_values(['Rayon_aff', 'CA'], ascending=[True, False])
    for i, (_, row_) in enumerate(fam_sorted.iterrows()):
        data_row(ws, r, [row_['Rayon_aff'], row_['Famille_aff'], row_['CA'],
                          (row_['Évol CA %']/100 if pd.notna(row_['Évol CA %']) else None), row_['Marge'],
                          (row_['Tx Marge %']/100 if pd.notna(row_['Tx Marge %']) else None), row_['Qté Vente'],
                          (row_['Évol Qté %']/100 if pd.notna(row_['Évol Qté %']) else None)], zebra=(i%2==1), left_cols=(1,2))
        for col, fmt_ in [(3,ACC),(4,PCT),(5,ACC),(6,PCT),(7,QTY),(8,PCT)]:
            ws.cell(row=r, column=col).number_format = fmt_
        r += 1
    r += 1

    def top_section(title, dframe, cols_map, kinds, color=RED_H):
        nonlocal r
        section_bar(ws, r, 8, title, color=color); r += 1
        header_row(ws, r, ["Rang"] + list(cols_map.values())); r += 1
        keys = list(cols_map.keys())
        for i, (_, row_) in enumerate(dframe.iterrows()):
            vals = [i+1] + [fmt_value(kinds.get(c, 'amount'), row_[c]) for c in keys]
            data_row(ws, r, vals, zebra=(i%2==1), left_cols=(2,3))
            for j, ck in enumerate(keys, start=2):
                ws.cell(row=r, column=j).number_format = fmt_code(kinds.get(ck, 'amount'))
            r += 1
        r += 1

    top_section(f"5.  FLOP {n_top} — PLUS FORTE BAISSE DE CA (par Famille)",
                top_flop_table(fam, 'Perte CA', n_top, 'flop', ['CA','CA N-1','Évol CA %','Perte CA','Tx Marge %']),
                {'Rayon_aff':'Rayon','Famille_aff':'Famille','CA':'CA (FCFA)','CA N-1':'CA N-1','Évol CA %':'Évol %','Perte CA':'Perte (FCFA)','Tx Marge %':'Taux Marge'},
                {'CA':'amount','CA N-1':'amount','Évol CA %':'pct100','Perte CA':'amount','Tx Marge %':'pct100'})

    top_section(f"6.  TOP {n_top} — MEILLEUR GAIN DE CA (par Famille)",
                top_flop_table(fam, 'Perte CA', n_top, 'top', ['CA','CA N-1','Évol CA %','Perte CA','Tx Marge %']),
                {'Rayon_aff':'Rayon','Famille_aff':'Famille','CA':'CA (FCFA)','CA N-1':'CA N-1','Évol CA %':'Évol %','Perte CA':'Gain (FCFA)','Tx Marge %':'Taux Marge'},
                {'CA':'amount','CA N-1':'amount','Évol CA %':'pct100','Perte CA':'amount','Tx Marge %':'pct100'},
                color=GREEN_H)

    top_section(f"7.  FLOP {n_top} — CASSE EN VALEUR (par Famille)",
                top_familles_for_excel(fam, n_top, 'casse'),
                {'Rayon_aff':'Rayon','Famille_aff':'Famille','CA':'CA (FCFA)','Casse (Valeur)':'Casse (FCFA)','%Casse (Valeur)':'% Casse'},
                {'CA':'amount','Casse (Valeur)':'amount','%Casse (Valeur)':'pct1'})

    top_section(f"8.  TOP {n_top} — POIDS PROMO LE PLUS ÉLEVÉ (Famille, CA > 1M)",
                top_familles_for_excel(fam, n_top, 'promo'),
                {'Rayon_aff':'Rayon','Famille_aff':'Famille','CA':'CA (FCFA)','%CA Poids Promo':'Poids Promo','%Marge Promo':'Tx M. Promo','%Marge Hors Promo':'Tx M. HP'},
                {'CA':'amount','%CA Poids Promo':'pct1','%Marge Promo':'pct1','%Marge Hors Promo':'pct1'},
                color="FFFF9500")

    autosize(ws, {'A':22,'B':26,'C':16,'D':13,'E':16,'F':13,'G':13,'H':13})
    ws.freeze_panes = "A6"

    # ============== FEUILLE 2 : DESTRUCTEURS & PERFORMEURS (réseau) + 🆕 1 feuille par Rayon ==============
    def fill_destructeurs_sheet(ws_target, header_text, res, n_top_local):
        """Remplit une feuille Destructeurs & Performeurs (sections A→G) pour
        n'importe quel périmètre (réseau entier ou un seul Rayon)."""
        ws_target.sheet_view.showGridLines = False
        ws_target.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        c1 = ws_target.cell(row=1, column=1, value=header_text)
        c1.font = Font(name=ARIAL, bold=True, size=14, color=WHITE_H)
        c1.fill = PatternFill("solid", fgColor=BLUE_H)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_target.row_dimensions[1].height = 26

        r_local = 3
        def art_section(title, dframe, cols_map, kinds, color):
            nonlocal r_local
            ws_target.merge_cells(start_row=r_local, start_column=1, end_row=r_local, end_column=9)
            c = ws_target.cell(row=r_local, column=1, value=title)
            c.font = Font(name=ARIAL, bold=True, size=11, color=WHITE_H)
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_target.row_dimensions[r_local].height = 20
            r_local += 1
            for i, lbl in enumerate(["Rang"] + list(cols_map.values()), start=1):
                cell = ws_target.cell(row=r_local, column=i, value=lbl)
                cell.font = Font(name=ARIAL, bold=True, size=10, color=WHITE_H)
                cell.fill = PatternFill("solid", fgColor=BLUE_H)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            r_local += 1
            keys = list(cols_map.keys())
            for i, (_, row_) in enumerate(dframe.iterrows()):
                vals = [i+1] + [fmt_value(kinds.get(c, 'amount'), row_.get(c, None)) for c in keys]
                fillc = LGREY_H if i % 2 == 1 else "FFFFFFFF"
                for j, v in enumerate(vals, start=1):
                    cell = ws_target.cell(row=r_local, column=j, value=v)
                    cell.font = Font(name=ARIAL, size=10)
                    cell.border = box
                    cell.fill = PatternFill("solid", fgColor=fillc)
                    cell.alignment = Alignment(horizontal="left" if j in (2,3,4,5) else "center")
                    if j >= 6:
                        cell.number_format = fmt_code(kinds.get(keys[j-2], 'amount'))
                r_local += 1
            r_local += 1

        cm_neg = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','CA':'CA','Marge':'Marge','Tx Marge %':'Taux Marge'}
        art_section(f"A.  FLOP {n_top_local} — ARTICLES EN MARGE NÉGATIVE", res['A_marge_neg'], cm_neg,
                    {'CA':'amount','Marge':'amount','Tx Marge %':'pct100'}, RED_H)

        cm_deg = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','CA':'CA','Tx Marge %':'Taux Marge','Écart Tx Marge (pts)':'Écart pts'}
        art_section(f"B.  FLOP {n_top_local} — DÉGRADATION DU TAUX DE MARGE (marge encore positive)", res['B_degrad_marge'], cm_deg,
                    {'CA':'amount','Tx Marge %':'pct100','Écart Tx Marge (pts)':'pts'}, RED_H)

        cm_gain = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','CA':'CA','Gain Marge (FCFA)':'Gain Marge','Tx Marge %':'Taux Marge'}
        art_section(f"C.  TOP {n_top_local} — PERFORMEURS : GAIN DE MARGE EN VALEUR", res['C_perf_gain_marge'], cm_gain,
                    {'CA':'amount','Gain Marge (FCFA)':'amount','Tx Marge %':'pct100'}, GREEN_H)

        d4_local = res['D_croissance_ca'].copy()
        if not d4_local.empty:
            d4_local['Évol CA %'] = (d4_local['CA']/d4_local['CA N-1']-1)*100
        cm_croi = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','CA':'CA','CA N-1':'CA N-1','Évol CA %':'Évol %','Tx Marge %':'Taux Marge'}
        art_section(f"D.  TOP {n_top_local} — PLUS FORTE CROISSANCE DE CA", d4_local, cm_croi,
                    {'CA':'amount','CA N-1':'amount','Évol CA %':'pct100','Tx Marge %':'pct100'}, GREEN_H)

        cm_baisse = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','CA':'CA','CA N-1':'CA N-1','Perte CA (FCFA)':'Perte (FCFA)'}
        art_section(f"E.  FLOP {n_top_local} — PLUS FORTE BAISSE DE CA", res['E_baisse_ca'], cm_baisse,
                    {'CA':'amount','CA N-1':'amount','Perte CA (FCFA)':'amount'}, RED_H)

        cm_qte = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article','Qté Vente':'Qté Vente','Qté Vente N-1':'Qté N-1','Variation Qté':'Variation','CA':'CA'}
        kinds_qte = {'Qté Vente':'qty','Qté Vente N-1':'qty','Variation Qté':'amount_signed','CA':'amount'}
        art_section(f"F.  TOP {n_top_local} — PLUS FORTE HAUSSE DE QUANTITÉ VENDUE", res['F_hausse_qte'], cm_qte, kinds_qte, GREEN_H)
        art_section(f"G.  FLOP {n_top_local} — PLUS FORTE BAISSE DE QUANTITÉ VENDUE", res['G_baisse_qte'], cm_qte, kinds_qte, RED_H)

        autosize(ws_target, {'A':7,'B':20,'C':24,'D':22,'E':38,'F':13,'G':13,'H':13,'I':13})

    ws2 = wb.create_sheet("Destructeurs Performeurs")
    fill_destructeurs_sheet(ws2, "DESTRUCTEURS & PERFORMEURS — NIVEAU ARTICLE (agrégé réseau, tous sites)",
                             art_res, n_top)

    # ---- 🆕 Une feuille dédiée par Rayon, même structure A→G, filtrée sur ce seul rayon ----
    if art_full is not None and not art_full.empty:
        def _safe_sheet_name(base, used):
            invalid = '[]:*?/\\'
            name = "".join(ch for ch in base if ch not in invalid).strip()[:31]
            if not name:
                name = "Rayon"
            candidate = name
            i = 2
            while candidate in used:
                suffix = f" {i}"
                candidate = (name[:31 - len(suffix)] + suffix)
                i += 1
            used.add(candidate)
            return candidate

        used_names = {"Dashboard COPIL", "Destructeurs Performeurs", "Marges Négatives par Site"}
        rayons = sorted(rr for rr in art_full['Rayon_aff'].dropna().unique() if str(rr).strip())
        for rayon in rayons:
            art_r = art_full[art_full['Rayon_aff'] == rayon]
            if art_r.empty:
                continue
            res_r = destructeurs_performeurs(art_r, n=n_top, seuil_ca=seuil_ca)
            sheet_name = _safe_sheet_name(f"DP - {rayon}", used_names)
            ws_r = wb.create_sheet(sheet_name)
            fill_destructeurs_sheet(ws_r, f"DESTRUCTEURS & PERFORMEURS — {rayon.upper()}", res_r, n_top)

    # ============== 🆕 FEUILLE 3 : MARGES NÉGATIVES PAR SITE ==============
    # Format aligné sur le récap de référence : bandeau KPI + tableau plat
    # EXHAUSTIF (toutes les lignes article × site en marge négative).
    if mns_detail is not None and not mns_detail.empty:
        NAVY = "FF1B2A4A"; NAVY2 = "FF2E4B7A"; INK = "FF1A1A2E"
        AMBER_F = "FFFFF3E0"; REDT_F = "FFFDECEA"; RED_TX = "FFC0392B"
        CAL = "Calibri"

        ws3 = wb.create_sheet("Marges Négatives par Site")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:K1")
        ws3["A1"] = "ARTICLES À MARGE NÉGATIVE — DÉTAIL PAR MAGASIN"
        ws3["A1"].font = Font(name=CAL, bold=True, size=13, color=WHITE_H)
        ws3["A1"].fill = PatternFill("solid", fgColor=NAVY)
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 24

        ws3.merge_cells("A2:K2")
        per_txt = (perimetre.replace("\n", " ")[:150] if perimetre else "Voir export source")
        ws3["A2"] = f"  Période : {per_txt}"
        ws3["A2"].font = Font(name=CAL, size=9, color="FFAABBCC")
        ws3["A2"].fill = PatternFill("solid", fgColor=NAVY)
        ws3["A2"].alignment = Alignment(horizontal="left", vertical="center")

        # ---- Bandeau KPI (lignes 4-5) ----
        pertes = mns_detail['Marge'].sum()
        kpi_data = [
            ("Lignes article × site", f"{len(mns_detail):,}"),
            ("Articles distincts", f"{mns_detail['Code Art.'].replace('', np.nan).fillna(mns_detail['Article_aff']).nunique():,}"),
            ("Magasins touchés", f"{mns_detail['Site'].nunique():,}"),
            ("Pertes nettes (FCFA)", f"{abs(pertes):,.0f}"),
            ("CA exposé (FCFA)", f"{mns_detail['CA'].sum():,.0f}"),
        ]
        for j, (lbl, val) in enumerate(kpi_data, start=1):
            c4 = ws3.cell(row=4, column=j, value=lbl)
            c4.font = Font(name=CAL, bold=True, size=9, color=WHITE_H)
            c4.fill = PatternFill("solid", fgColor=NAVY2)
            c4.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c5 = ws3.cell(row=5, column=j, value=val)
            c5.font = Font(name=CAL, bold=True, size=12, color=INK)
            c5.fill = PatternFill("solid", fgColor=WHITE_H)
            c5.border = box
            c5.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[4].height = 22
        ws3.row_dimensions[5].height = 20

        # ---- En-tête tableau (ligne 7) ----
        headers3 = ["#", "Code Art.", "Article", "Rayon", "Famille", "Magasin",
                    "Format", "CA (FCFA)", "Marge (FCFA)", "Tx Marge", "Qté"]
        for j, lbl in enumerate(headers3, start=1):
            c = ws3.cell(row=7, column=j, value=lbl)
            c.font = Font(name=CAL, bold=True, size=10, color=WHITE_H)
            c.fill = PatternFill("solid", fgColor=NAVY2)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # ---- Données (exhaustif, trié par article) ----
        r3 = 8
        for i, (_, row_) in enumerate(mns_detail.iterrows(), start=1):
            tx = row_['Tx Marge %']
            severe = pd.notna(tx) and tx < -5.0
            fillc = REDT_F if severe else AMBER_F
            code = row_['Code Art.']
            vals = [i,
                    int(code) if str(code).isdigit() else (code or None),
                    row_['Article_aff'], row_['Rayon_aff'], row_['Famille_aff'],
                    row_['Site'], row_['Format'],
                    row_['CA'], row_['Marge'],
                    (tx/100 if pd.notna(tx) else None), row_['Qté']]
            for j, v in enumerate(vals, start=1):
                cell = ws3.cell(row=r3, column=j, value=v)
                cell.fill = PatternFill("solid", fgColor=fillc)
                cell.border = box
                if j == 10:  # Tx Marge en rouge gras
                    cell.font = Font(name=CAL, size=10, bold=True, color=RED_TX)
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="center")
                elif j in (8, 9, 11):
                    cell.font = Font(name=CAL, size=10, color=INK)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif j == 1:
                    cell.font = Font(name=CAL, size=10, color=INK)
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.font = Font(name=CAL, size=10, color=INK)
                    cell.alignment = Alignment(horizontal="left")
            r3 += 1

        autosize(ws3, {'A':5,'B':12,'C':38,'D':18,'E':24,'F':28,'G':10,'H':14,'I':14,'J':10,'K':8})
        ws3.freeze_panes = "A8"
        ws3.auto_filter.ref = f"A7:K{r3-1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def top_familles_for_excel(fam, n, by):
    if by == 'casse':
        sub = fam[fam['Casse (Valeur)'].notna()]
        out = sub.nsmallest(n, 'Casse (Valeur)')[['Rayon_aff','Famille_aff','CA','Casse (Valeur)','%Casse (Valeur)']]
    elif by == 'promo':
        mat = fam[fam['CA'] > 1_000_000]
        out = mat.nlargest(n, '%CA Poids Promo')[['Rayon_aff','Famille_aff','CA','%CA Poids Promo','%Marge Promo','%Marge Hors Promo']]
    return out.reset_index(drop=True)

# ============================================================
# 🆕 RENDU DESTRUCTEURS & PERFORMEURS — réutilisable par périmètre
# (Réseau entier OU un seul Rayon, chacun dans son onglet)
# ============================================================

# ============================================================
# 📅 EXTRACTION DE LA PÉRIODE — dernière cellule de la 1ère colonne
# ============================================================
def extract_periode(perimetre_text):
    """Extrait 'JJ/MM/AAAA → JJ/MM/AAAA' depuis le texte de filtre PBI.
    Retourne le texte brut tronqué si le pattern n'est pas trouvé."""
    if not perimetre_text:
        return None
    m = re.search(r"après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})",
                  str(perimetre_text))
    if m:
        return f"{m.group(1)} → {m.group(2)}"
    return str(perimetre_text).replace("\n", " ")[:120]


# ============================================================
# 🚀 GÉNÉRATION DU RAPPORT
# ============================================================
def generate_reporting_ventes(input_path, output_dir=".", n_top=15, seuil_ca=100_000):
    input_path = Path(input_path)
    with open(input_path, "rb") as f:
        file_bytes = f.read()

    df, perimetre, site_col = load_export(file_bytes)
    periode = extract_periode(perimetre)

    art = prep_articles(df)
    mns = marge_negative_par_site(df, site_col)
    mns_detail = detail_marge_neg_site(df, site_col)

    k = kpis_globaux_rayon(df)
    if k is None:
        raise ValueError("Ligne de total réseau ('Total') introuvable dans l'export — vérifiez le fichier.")

    perf = perf_par_rayon(df, CIBLES_DEFAUT)
    fam = family_metrics(df)
    art_res = destructeurs_performeurs(art, n=n_top, seuil_ca=seuil_ca)

    xls_bytes = build_excel_full(k, perf, fam, n_top, art_res, perimetre, mns, mns_detail,
                                  art_full=art, seuil_ca=seuil_ca)

    # Injecte la période lisible en A2/B2 de l'onglet Dashboard COPIL
    # (à la place du placeholder "Voir export source")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xls_bytes))
    ws = wb["Dashboard COPIL"]
    ws["B2"] = periode or "Voir périmètre ci-dessous"
    buf = io.BytesIO()
    wb.save(buf)
    xls_bytes = buf.getvalue()

    today_str = date.today().strftime("%Y%m%d")
    out_name = f"Reporting Ventes - {today_str}.xlsx"
    out_path = Path(output_dir) / out_name
    with open(out_path, "wb") as f:
        f.write(xls_bytes)

    rayons = sorted(art["Rayon_aff"].dropna().unique().tolist())
    print(f"✅ Fichier généré : {out_path}")
    print(f"   Période détectée : {periode}")
    print(f"   Rayons (onglets DP) : {', '.join(rayons)}")
    print(f"   Marges Négatives par Site : {'incluse' if (mns_detail is not None and not mns_detail.empty) else 'absente (pas de colonne Site dans l\'export)'}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        candidates = list(Path(".").glob("*.xlsx"))
        if not candidates:
            print("Usage : python generate_reporting_ventes.py chemin/vers/export_pbi.xlsx [dossier_sortie]")
            sys.exit(1)
        input_file = candidates[0]
        print(f"Aucun fichier fourni — utilisation de {input_file}")
    else:
        input_file = sys.argv[1]

    out_dir = sys.argv[2] if len(sys.argv) > 2 else "."
    generate_reporting_ventes(input_file, out_dir)
