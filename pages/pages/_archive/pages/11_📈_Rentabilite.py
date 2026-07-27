"""
11_📊_Rentabilite.py — SmartBuyer Hub
Cockpit Direction · Briefing Acheteur · Analyse Approfondie
Architecture : 1 question par zone · 1 réponse immédiate · 1 action
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Rentabilité · SmartBuyer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1300px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }
.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.6; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.verdict-box {
    background: #FFFFFF; border-radius: 14px; padding: 18px 22px;
    border: 0.5px solid #E5E5EA; margin-bottom: 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.verdict-title { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 6px; }
.verdict-text  { font-size: 15px; font-weight: 500; color: #1C1C1E; line-height: 1.5; }
.action-card {
    background: #FFFFFF; border-radius: 12px; padding: 14px 18px;
    border: 0.5px solid #E5E5EA; margin-bottom: 8px;
    border-left: 4px solid #FF3B30;
}
.action-card.amber { border-left-color: #FF9500; }
.action-card.green { border-left-color: #34C759; }
.action-num  { font-size: 11px; font-weight: 700; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.05em; }
.action-fam  { font-size: 15px; font-weight: 700; color: #1C1C1E; margin: 2px 0; }
.action-fcfa { font-size: 13px; font-weight: 600; color: #FF3B30; }
.action-fcfa.amber { color: #FF9500; }
.action-what { font-size: 13px; color: #3A3A3C; margin-top: 4px; }
.sante-rouge { background:#FEE2E2; color:#991B1B; border-radius:8px; padding:2px 10px; font-size:12px; font-weight:700; display:inline-block; }
.sante-orange{ background:#FEF9C3; color:#854D0E; border-radius:8px; padding:2px 10px; font-size:12px; font-weight:700; display:inline-block; }
.sante-vert  { background:#D5F5E3; color:#145A32; border-radius:8px; padding:2px 10px; font-size:12px; font-weight:700; display:inline-block; }
.casse-badge { background:#F5F0FF; color:#6B21A8; border-radius:8px; padding:2px 8px; font-size:11px; font-weight:600; display:inline-block; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
ACHETEURS = {
    'BOISSONS':           'Acheteur Boissons',
    'EPICERIE':           'Acheteur Épicerie',
    'DROGUERIE':          'Acheteur DPH',
    'PARFUMERIE HYGIENE': 'Acheteur DPH',
}
PLANCHERS = {
    'Produit d appel': 0.10, 'Valeur ajoutee': 0.25,
    'PH Droguerie':    0.22, 'Coeur de gamme': 0.18,
}
SEG_LABELS = {
    'Produit d appel': "Produit d'appel",
    'Valeur ajoutee':  'Valeur ajoutée',
    'PH Droguerie':    'PH / Droguerie',
    'Coeur de gamme':  'Cœur de gamme',
}
PRODUITS_APPEL = ['RIZ LONG','HUILES','LAITS','EAUX PLATES','EAUX GAZEUSES',
                  'SUCRES ET LEVURES','FARINES CEREALES','LEGUMES SEC',
                  'PATES LONGUES','PATES COURTES','SEMOULES COUSCOUS',
                  'BOUILLON AIDE CULINAIRE','SAUCES FROIDES']
VALEUR_AJOUTEE = ['BIO','CHIPS','SOINS DU CORPS','SOINS DU VISAGE',
                  'PANSEMENT ET COMPLEMENTS','PRODUITS DU MONDE',
                  'SELS ET EPICES','SNACKING','DERMO COSMETIQUE','DIET','MAQUILLAGE']
ORDRE_RAYONS  = ['BOISSONS','EPICERIE','DROGUERIE','PARFUMERIE HYGIENE']
TOLERANCE     = 0.015
COLS_REQUIRED = ['Rayon','Sous Famille','CA','Marge','%Marge','CA N-1','%Vs N-1.1']

# ─── HELPERS FORMAT ───────────────────────────────────────────────────────────
def fp(v, sign=True):
    try:
        if pd.isna(v): return '—'
        return f"{v:+.1%}" if sign else f"{v:.1%}"
    except: return '—'

def fk(v):
    try:
        if pd.isna(v) or v == 0: return '—'
        a = abs(v)
        if a >= 1_000_000: return f"{v/1_000_000:+.2f} M FCFA"
        return f"{v/1000:+,.0f} K FCFA"
    except: return '—'

def fk_abs(v):
    try:
        if pd.isna(v) or v == 0: return '—'
        a = abs(v)
        if a >= 1_000_000: return f"{a/1_000_000:.1f} M FCFA"
        return f"{a/1000:,.0f} K FCFA"
    except: return '—'

def cs(v):
    if '✅' in str(v) or 'OK' == str(v).strip(): return 'background:#D5F5E3;color:#145A32;font-weight:600'
    if '🟡' in str(v) or 'Vigi' in str(v):       return 'background:#FEF9C3;color:#854D0E;font-weight:600'
    if '🔴' in str(v) or 'Action' in str(v):      return 'background:#FEE2E2;color:#991B1B;font-weight:600'
    return ''

def cd(v):
    try:
        x = float(str(v).replace('%','').replace('+','').replace(' K FCFA','')
                        .replace(' M FCFA','').replace(',','').replace('—','').strip())
        if x >= -1.5: return 'color:#145A32;font-weight:600'
        if x >= -3.0: return 'color:#854D0E;font-weight:600'
        return 'color:#991B1B;font-weight:600'
    except: return ''

# ─── MOTEUR DE CALCUL — fonctions vectorisées ─────────────────────────────────
def _segment_vec(sf_s, ray_s):
    sf  = sf_s.str.upper().fillna('')
    ray = ray_s.str.upper().fillna('')
    seg = pd.Series('Coeur de gamme', index=sf.index)
    seg[ray.isin(['DROGUERIE','PARFUMERIE HYGIENE'])] = 'PH Droguerie'
    mask_a = sf.apply(lambda s: any(p in s for p in PRODUITS_APPEL))
    seg[mask_a & (seg == 'Coeur de gamme')] = 'Produit d appel'
    mask_v = sf.apply(lambda s: any(v in s for v in VALEUR_AJOUTEE))
    seg[mask_v & (seg == 'Coeur de gamme')] = 'Valeur ajoutee'
    return seg

def _statut_vec(dev):
    s = pd.Series('N/A', index=dev.index)
    nn = dev.notna()
    s[nn & (dev >= -TOLERANCE)]                        = 'OK'
    s[nn & (dev < -TOLERANCE) & (dev >= -TOLERANCE*2)] = 'Vigilance'
    s[nn & (dev < -TOLERANCE*2)]                       = 'Action'
    return s

def _cause_vec(df):
    """Cause de la déviation : Promo / Conditions achat / Effet mix / Produit appel / OK"""
    dev = df['Dev_N1_pts']
    seg = df['Segment']
    has_promo = 'CA Promo' in df.columns and 'Marge Promo' in df.columns
    if has_promo:
        pp = df['CA Promo'].fillna(0) / df['CA'].replace(0, 1)
        tp = df['Marge Promo'].fillna(0) / df['CA Promo'].replace(0, 1)
        promo_def = (pp > 0.20) & (tp < 0.05) & dev.notna() & (dev < -0.02)
    else:
        promo_def = pd.Series(False, index=df.index)
    alerte_mix = df.get('Alerte_Mix', pd.Series(False, index=df.index))
    c = pd.Series('A vérifier', index=df.index)
    c[dev.notna() & (dev >= -TOLERANCE)]                        = 'OK'
    c[promo_def]                                                 = 'Promo déficitaire'
    c[alerte_mix & ~promo_def]                                   = 'Effet mix défavorable'
    c[dev.notna() & (dev < -0.05) & ~promo_def & ~alerte_mix]   = 'Conditions achat'
    c[dev.notna() & (dev < -0.10) & ~promo_def]                 = 'Chute sévère PA'
    c[seg == 'Produit d appel']                                  = "Prod. d'appel — remise arrière"
    c[df['%Marge'].notna() & (df['%Marge'] < 0)]                = 'Marge négative'
    return c

def _action_vec(df):
    """1 action claire par famille — courte, directe, actionnable"""
    tx  = df['%Marge']
    dev = df['Dev_N1_pts']
    seg = df['Segment']
    ca_med = df.groupby('Rayon_court')['CA'].transform('median').fillna(1)
    gros   = df['CA'] > ca_med * 1.5
    has_promo = 'CA Promo' in df.columns and 'Marge Promo' in df.columns
    if has_promo:
        pp       = df['CA Promo'].fillna(0) / df['CA'].replace(0, 1)
        tp       = df['Marge Promo'].fillna(0) / df['CA Promo'].replace(0, 1)
        promo_def = (pp > 0.20) & (tp < 0.05)
    else:
        promo_def = pd.Series(False, index=df.index)

    a = pd.Series('Maintenir les conditions', index=df.index)
    # Marge négative — priorité absolue
    a[tx.notna() & (tx < 0)]                                    = 'Stopper toute promo — vérifier PA'
    # Promo déficitaire
    a[promo_def & dev.notna() & (dev < -0.02)]                  = 'Suspendre la promo en cours'
    # Produit d'appel sous pression
    a[(seg == 'Produit d appel') & dev.notna() & (dev < -0.05)] = 'Négocier remise arrière fournisseur'
    # Gros volume + chute sévère
    a[dev.notna() & (dev < -0.10) & gros & ~promo_def]          = 'Convoquer le fournisseur cette semaine'
    # Gros volume + dérive
    a[dev.notna() & (dev < -0.05) & gros & ~promo_def]          = 'Renégocier conditions — volume en jeu'
    # Dérive standard
    a[dev.notna() & (dev < -0.05) & ~promo_def & ~gros]         = 'Revoir tarif achat fournisseur'
    # Vigilance
    a[dev.notna() & (dev < -TOLERANCE) & (dev >= -TOLERANCE*2) & ~promo_def] = 'Surveiller — confirmer semaine prochaine'
    return a

def _score_sante_vec(df):
    """
    Score de santé 0-100 par famille.
    100 = parfait. 0 = catastrophique.
    Composantes : déviation N-1 (50%) + atteinte cible (30%) + impact casse (20%)
    """
    # Composante 1 : déviation N-1 (50 pts max)
    dev = df['Dev_N1_pts'].fillna(0)
    # +5pts = 50, 0pt = 40, -3pts = 20, -10pts = 0
    score_dev = np.clip(40 + dev * (30 / 0.05), 0, 50)

    # Composante 2 : atteinte cible (30 pts max)
    dev_cible = df['Dev_Cible_pts'].fillna(0)
    score_cible = np.clip(30 + dev_cible * (20 / 0.05), 0, 30)

    # Composante 3 : impact casse (20 pts max)
    tx_casse = df.get('Tx_Casse_Fam', pd.Series(0, index=df.index)).fillna(0)
    # 0% casse = 20pts, 1% casse = 10pts, 5% casse = 0pts
    score_casse = np.clip(20 - tx_casse * (20 / 0.05), 0, 20)

    total = (score_dev + score_cible + score_casse).round(0).astype(int)
    return total.clip(0, 100)

def _verdict_auto(df, periode):
    """Génère la phrase de verdict réseau pour la direction."""
    ca   = df['CA'].sum()
    mg   = df['Marge'].sum()
    tx   = mg / ca if ca > 0 else 0
    mn1  = df['Marge_N1'].sum()
    cn1  = df['CA_N1'].sum()
    tn1  = mn1 / cn1 if cn1 > 0 else 0
    dev  = tx - tn1
    perdu = df['Dev_N1_FCFA'].sum()
    n_action = (df['Statut'] == 'Action').sum()

    # Rayon le plus dégradé
    worst = None
    worst_dev = 0
    for r in ORDRE_RAYONS:
        sub = df[df['Rayon_court'] == r]
        if len(sub) == 0: continue
        ca_r = sub['CA'].sum(); mg_r = sub['Marge'].sum()
        mn1_r = sub['Marge_N1'].sum(); cn1_r = sub['CA_N1'].sum()
        tx_r  = mg_r / ca_r if ca_r > 0 else 0
        tn1_r = mn1_r / cn1_r if cn1_r > 0 else 0
        d_r   = tx_r - tn1_r
        if d_r < worst_dev:
            worst_dev = d_r
            worst = r

    # Top famille à risque
    top = df[df['Statut'] == 'Action'].nlargest(1, 'Impact_Score')
    top_txt = ''
    if len(top):
        r0 = top.iloc[0]
        top_txt = f" Sujet principal : {r0['SF_court']} ({fk_abs(r0['Dev_N1_FCFA'])} perdus)."

    if dev >= 0:
        etat = f"en progression de {dev:+.1%} vs N-1"
    elif dev >= -0.01:
        etat = f"stable à {tx:.1%} (−{abs(dev):.1%} vs N-1)"
    else:
        etat = f"en recul de {abs(dev):.1%} vs N-1"

    perdu_txt = f" Marge perdue : {fk_abs(perdu)}." if perdu < 0 else ""
    worst_txt = f" {worst.title()} décroche." if worst else ""
    action_txt = f" {n_action} famille(s) nécessitent une action cette semaine." if n_action > 0 else " Aucune alerte critique."

    return f"Réseau {etat} à {tx:.1%}.{perdu_txt}{worst_txt}{action_txt}{top_txt}"

def _top3_actions(df_ach):
    """Retourne les 3 actions prioritaires pour un acheteur avec contexte complet."""
    rouge = df_ach[df_ach['Statut'] == 'Action'].nlargest(3, 'Impact_Score')
    actions = []
    for _, r in rouge.iterrows():
        cause = str(r.get('Cause', ''))
        action = str(r.get('Action', ''))
        perdu  = r.get('Dev_N1_FCFA', 0)
        # Icône selon la cause
        if 'négatif' in cause.lower():  icone = 'rouge'
        elif 'promo'  in cause.lower(): icone = 'amber'
        else:                           icone = 'rouge'
        actions.append({
            'famille': r['SF_court'],
            'perdu':   fk_abs(perdu),
            'action':  action,
            'cause':   cause,
            'icone':   icone,
            'site':    r.get('Site nom long', '') if 'Site nom long' in r.index else '',
        })
    return actions

# ─── CHARGEMENT RÉFÉRENTIEL ───────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_referentiel(override_bytes=None):
    if override_bytes:
        try:
            ref = pd.read_excel(BytesIO(override_bytes))
            if 'Cible' in ref.columns: return ref
        except Exception: pass
    repo_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'referentiel_cibles.csv')
    if os.path.exists(repo_path): return pd.read_csv(repo_path)
    return None

# ─── CHARGEMENT EXTRACTION ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_extraction(file_bytes: bytes, filename: str, ref_bytes=None):
    raw    = BytesIO(file_bytes)
    df_raw = pd.read_excel(raw, header=None)
    # Détection période depuis dernière cellule col A (lecture pandas uniquement)
    col_a  = df_raw.iloc[:, 0].dropna().astype(str)
    last   = col_a.iloc[-1] if len(col_a) else ''
    dates  = re.findall(r'\d{2}/\d{2}/\d{4}', last)
    if len(dates) >= 2:   periode = f"{dates[0]} → {dates[1]}"
    elif len(dates) == 1: periode = dates[0]
    else:                 periode = 'Période inconnue'

    raw.seek(0)
    df = pd.read_excel(raw)

    # Validation colonnes obligatoires
    missing = [c for c in COLS_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"**{filename}** — colonnes manquantes : `{'`, `'.join(missing)}`")

    # Nettoyage lignes parasites PBI
    mask = (
        df['Sous Famille'].notna() &
        (df['Sous Famille'].astype(str).str.strip() != 'Total') &
        (~df['Sous Famille'].astype(str).str.startswith('Filtres', na=False)) &
        df['Rayon'].notna() &
        df['Rayon'].str.startswith('000', na=False) &
        ~df['Rayon'].str.contains('CIGARETTE', na=False) &
        ~df.get('Site nom long', pd.Series('', index=df.index)).astype(str).isin(['Total','']) &
        ~df.get('Famille',       pd.Series('', index=df.index)).astype(str).isin(['Total']) &
        df['CA'].notna() & (df['CA'] > 0)
    )
    df = df[mask].copy()

    df['SF_court']    = df['Sous Famille'].str.extract(r'\d+ - (.+)')[0]
    df['Rayon_court'] = df['Rayon'].str.extract(r'- (.+)')[0]
    df['Acheteur']    = df['Rayon_court'].map(ACHETEURS)
    df['Segment']     = _segment_vec(df['SF_court'], df['Rayon_court'])
    df['Plancher']    = df['Segment'].map(PLANCHERS)

    # N-1 vectorisé
    valid_n1       = df['%Vs N-1.1'].notna() & (df['%Vs N-1.1'] != -1) & (df['CA N-1'] > 0)
    df['Marge_N1'] = np.where(valid_n1, df['Marge'] / (1 + df['%Vs N-1.1']), np.nan)
    df['CA_N1']    = df['CA N-1']
    df['Tx_N1']    = np.where(valid_n1 & (df['CA_N1'] > 0), df['Marge_N1'] / df['CA_N1'], np.nan)

    # Cibles
    ref = load_referentiel(ref_bytes)
    if ref is not None and 'Cible' in ref.columns:
        df = df.merge(
            ref[['Rayon','Famille','Cible','Plancher']].rename(
                columns={'Rayon':'Rayon_court','Famille':'SF_court',
                         'Cible':'Cible_ref','Plancher':'Plancher_ref'}),
            on=['Rayon_court','SF_court'], how='left')
        df['Cible']    = df['Cible_ref'].fillna(df['Plancher'])
        df['Plancher'] = df['Plancher_ref'].fillna(df['Plancher'])
    else:
        cible_base    = np.where(df['Tx_N1'].notna(),
                                  np.maximum(df['Tx_N1'] * 1.02, df['Plancher']),
                                  df['Plancher'])
        plafond_appel = df['Plancher'] + 0.02
        df['Cible']   = np.where(df['Segment'] == 'Produit d appel',
                                  np.minimum(cible_base, plafond_appel), cible_base)

    # Déviations
    df = df[df['%Marge'].between(-0.4, 0.8)].copy()
    df['Dev_N1_pts']     = df['%Marge'] - df['Tx_N1']
    df['Dev_N1_FCFA']    = df['Dev_N1_pts'] * df['CA']
    df['Dev_Cible_pts']  = df['%Marge'] - df['Cible']
    df['Dev_Cible_FCFA'] = df['Dev_Cible_pts'] * df['CA']

    # Casse
    if 'Casse (Valeur)' in df.columns:
        df['Casse_val']    = df['Casse (Valeur)'].fillna(0)
        df['Tx_Casse_Fam'] = df['Casse_val'].abs() / df['CA'].replace(0, 1)
        df['Marge_Nette']  = (df['Marge'] + df['Casse_val']) / df['CA'].replace(0, 1)
        df['Alerte_Casse'] = df['Tx_Casse_Fam'] > 0.005
    else:
        df['Casse_val']    = 0.0
        df['Tx_Casse_Fam'] = 0.0
        df['Marge_Nette']  = df['%Marge']
        df['Alerte_Casse'] = False

    # Alerte mix (CA progresse mais marge décroche)
    _evo_ca = df['%Vs N-1'].dropna()
    _factor = 100 if _evo_ca.abs().median() > 1 else 1
    df['_evo_norm'] = np.where(df['%Vs N-1'].notna(), df['%Vs N-1'] / _factor, np.nan)
    df['Alerte_Mix'] = (
        df['_evo_norm'].notna() & (df['_evo_norm'] > 0.15) &
        df['Dev_N1_pts'].notna() & (df['Dev_N1_pts'] < -0.02)
    )
    df.drop(columns=['_evo_norm'], inplace=True)

    # Scores et diagnostics
    df['Statut']   = _statut_vec(df['Dev_N1_pts'])
    df['Cause']    = _cause_vec(df)
    df['Action']   = _action_vec(df)

    # Score Impact (pour tri interne)
    ca_med = df.groupby('Rayon_court')['CA'].transform('median').replace(0, 1)
    df['Impact_Score'] = (df['Dev_N1_FCFA'].abs() * (df['CA'] / ca_med).clip(0.5, 3.0)).round(0)

    # Score santé 0-100
    df['Score_Sante'] = _score_sante_vec(df)

    # Remise nécessaire pour atteindre la cible
    df['Remise_Necessaire'] = np.where(df['Dev_Cible_pts'] < 0,
                                        df['Dev_Cible_pts'].abs() * df['CA'], 0)

    # Tri
    _ord = {'Action': 0, 'Vigilance': 1, 'OK': 2, 'N/A': 3}
    df['_ord_statut'] = df['Statut'].map(_ord)
    df['_ord_rayon']  = df['Rayon_court'].map({r: i for i, r in enumerate(ORDRE_RAYONS)})
    df['Periode']     = periode
    df['Fichier']     = filename
    return df

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(df_all, periodes):
    wb  = Workbook()
    C_HDR = '1B2A4A'; C_SUB = '2E4B7A'; C_WH = 'FFFFFF'; C_DK = '1A1A2E'

    def xfill(h): return PatternFill('solid', fgColor=h)
    def xbdr():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def xctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def xrgt(): return Alignment(horizontal='right',  vertical='center')
    def xlft(w=False): return Alignment(horizontal='left', vertical='center', wrap_text=w)

    # ── Onglets par période ────────────────────────────────────────────────────
    for i_p, periode in enumerate(periodes):
        df = df_all[df_all['Periode'] == periode].copy()
        df = df.sort_values(['_ord_statut', 'Impact_Score'], ascending=[True, False])
        safe = periode.replace('/', '').replace('→', '_').replace(' ', '')[:28]
        ws   = wb.active if i_p == 0 else wb.create_sheet(safe)
        ws.title = safe; ws.sheet_view.showGridLines = False

        # Titre
        has_s = 'Site nom long' in df.columns and df['Site nom long'].notna().any()
        HDRS  = ['Rayon', 'Famille']
        if has_s: HDRS.append('Magasin')
        HDRS += ['Segment', 'Acheteur', 'CA (FCFA)', 'Marge brute (FCFA)',
                 'Taux brut', 'Tx marge nette (casse)', 'Taux N-1', 'Cible',
                 'Marge perdue FCFA', 'Score santé', 'Cause', 'Statut', 'Action']
        WS = [20, 30]
        if has_s: WS.append(24)
        WS += [16, 18, 14, 14, 11, 14, 11, 11, 16, 11, 22, 14, 38]

        ws.merge_cells(f'A1:{get_column_letter(len(HDRS))}1')
        t = ws.cell(row=1, column=1, value=f'SUIVI RENTABILITE — {periode}')
        t.font = Font('Calibri', size=13, bold=True, color=C_WH)
        t.fill = xfill(C_HDR); t.alignment = xctr()
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f'A2:{get_column_letter(len(HDRS))}2')
        t2 = ws.cell(row=2, column=1,
            value=f'  Cible = MAX(N-1 x 1,02 ; Plancher segment) · Score sante = déviation + cible + casse · Trie par marge perdue')
        t2.font = Font('Calibri', size=9, italic=True, color='AABBCC')
        t2.fill = xfill(C_HDR); t2.alignment = xlft()
        ws.row_dimensions[2].height = 16

        for j, (h, w) in enumerate(zip(HDRS, WS), 1):
            c = ws.cell(row=3, column=j, value=h)
            c.font = Font('Calibri', size=9, bold=True, color=C_WH)
            c.fill = xfill(C_SUB); c.alignment = xctr(); c.border = xbdr()
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[3].height = 28

        C_R='FFD6D6'; C_O='FFF3CC'; C_G='D6F5D6'; C_L='F7F7F7'; C_W='FFFFFF'
        for i, (_, r) in enumerate(df.iterrows(), 4):
            stat = r.get('Statut', '')
            bg = C_R if stat=='Action' else (C_O if stat=='Vigilance' else (C_G if stat=='OK' else (C_L if i%2==0 else C_W)))
            vals = [r.get('Rayon_court',''), r.get('SF_court','')]
            if has_s: vals.append(r.get('Site nom long','—') if 'Site nom long' in r.index else '—')
            vals += [
                SEG_LABELS.get(r.get('Segment',''), r.get('Segment','')),
                r.get('Acheteur',''),
                r.get('CA', None),
                r.get('Marge', None),
                r.get('%Marge', None),
                r.get('Marge_Nette', None),
                r.get('Tx_N1', None),
                r.get('Cible', None),
                r.get('Dev_N1_FCFA', None),
                r.get('Score_Sante', None),
                r.get('Cause', ''),
                stat,
                r.get('Action', ''),
            ]
            _off  = 1 if has_s else 0
            FMTS  = [None, None]
            if has_s: FMTS.append(None)
            FMTS += [None, None, '#,##0', '#,##0', '0.0%', '0.0%', '0.0%', '0.0%',
                     '+#,##0;-#,##0;-', '0', None, None, None]
            for j, (v, f) in enumerate(zip(vals, FMTS), 1):
                c = ws.cell(row=i, column=j, value=v)
                c.fill = xfill(bg); c.border = xbdr()
                c.font = Font('Calibri', size=9, color=C_DK)
                c.alignment = xctr() if j in range(7+_off, 13+_off) else xrgt() if j in (5+_off, 6+_off, 11+_off) else xlft(w=(j==len(HDRS)))
                if f: c.number_format = f
            ws.row_dimensions[i].height = 16

        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS))}{3+len(df)}'

    # ── Onglet Plan de Négociation ─────────────────────────────────────────────
    ws_neg = wb.create_sheet("Plan de Negociation")
    ws_neg.sheet_view.showGridLines = False
    periode_latest = sorted(periodes)[-1]
    df_neg = df_all[(df_all['Periode']==periode_latest) & (df_all['Statut']=='Action')].copy()
    df_neg = df_neg.sort_values('Remise_Necessaire', ascending=False)

    NEG_H = ['Rayon','Famille','Magasin','Acheteur','CA (FCFA)','Marge brute (FCFA)',
             'Taux actuel','Cible','Marge perdue FCFA','Remise nécessaire FCFA','Cause','Action']
    NEG_W = [20,30,24,20,14,14,11,11,16,18,22,40]

    ws_neg.merge_cells(f'A1:{get_column_letter(len(NEG_H))}1')
    n1 = ws_neg.cell(row=1,column=1,value='PLAN DE NEGOCIATION — FAMILLES EN ACTION REQUISE')
    n1.font=Font('Calibri',size=13,bold=True,color=C_WH); n1.fill=xfill(C_HDR); n1.alignment=xctr()
    ws_neg.row_dimensions[1].height=30
    ws_neg.merge_cells(f'A2:{get_column_letter(len(NEG_H))}2')
    n2 = ws_neg.cell(row=2,column=1,
        value=f'  Periode : {periode_latest}  —  Trie par remise necessaire decroissante  —  Utiliser en preparation RDV fournisseurs')
    n2.font=Font('Calibri',size=9,italic=True,color='AABBCC'); n2.fill=xfill(C_HDR); n2.alignment=xlft()
    ws_neg.row_dimensions[2].height=16

    for j,(h,w) in enumerate(zip(NEG_H,NEG_W),1):
        c=ws_neg.cell(row=3,column=j,value=h)
        c.font=Font('Calibri',size=9,bold=True,color=C_WH)
        c.fill=xfill('CC2200' if j==10 else C_SUB)
        c.alignment=xctr(); c.border=xbdr()
        ws_neg.column_dimensions[get_column_letter(j)].width=w
    ws_neg.row_dimensions[3].height=28

    for i,(_, r) in enumerate(df_neg.iterrows(),4):
        bg='F7F7F7' if i%2==0 else 'FFFFFF'
        has_s_neg = 'Site nom long' in r.index
        nv=[r.get('Rayon_court',''), r.get('SF_court',''),
            r.get('Site nom long','—') if has_s_neg else '—',
            r.get('Acheteur',''), r.get('CA',None), r.get('Marge',None),
            r.get('%Marge',None), r.get('Cible',None),
            r.get('Dev_N1_FCFA',None), r.get('Remise_Necessaire',None),
            r.get('Cause',''), r.get('Action','')]
        NF=[None,None,None,None,'#,##0','#,##0','0.0%','0.0%','+#,##0;-#,##0;-','#,##0',None,None]
        for j,(v,f) in enumerate(zip(nv,NF),1):
            c=ws_neg.cell(row=i,column=j,value=v)
            c.fill=xfill('FFE8E8' if j==10 and (v or 0)>500000 else bg)
            c.font=Font('Calibri',size=9,bold=(j==10 and (v or 0)>500000),
                        color='CC2200' if j==10 and (v or 0)>500000 else '1A1A2E')
            c.alignment=Alignment(horizontal='right' if j in(5,6,9,10) else 'center' if j in(7,8) else 'left',
                                  vertical='center',wrap_text=(j==12))
            c.border=xbdr()
            if f: c.number_format=f
        ws_neg.row_dimensions[i].height=18

    # Total remise
    r_tot=len(df_neg)+4
    ws_neg.merge_cells(f'A{r_tot}:I{r_tot}')
    ct=ws_neg.cell(row=r_tot,column=1,value='TOTAL REMISE NECESSAIRE POUR ATTEINDRE LES CIBLES')
    ct.font=Font('Calibri',size=10,bold=True,color=C_WH); ct.fill=xfill(C_HDR); ct.alignment=xlft()
    cv=ws_neg.cell(row=r_tot,column=10,value=df_neg['Remise_Necessaire'].sum())
    cv.font=Font('Calibri',size=11,bold=True,color=C_WH); cv.fill=xfill('CC2200')
    cv.number_format='#,##0'; cv.alignment=xctr(); cv.border=xbdr()
    ws_neg.row_dimensions[r_tot].height=24
    ws_neg.freeze_panes='A4'
    ws_neg.auto_filter.ref=f'A3:{get_column_letter(len(NEG_H))}{r_tot-1}'

    # ── Onglet Synthèse Magasins ───────────────────────────────────────────────
    has_site_all = 'Site nom long' in df_all.columns and df_all['Site nom long'].notna().any()
    if has_site_all:
        ws_mag = wb.create_sheet("Synthese Magasins")
        ws_mag.sheet_view.showGridLines = False
        MAG_H = ['Magasin','CA (FCFA)','Marge (FCFA)','Taux actuel','Taux N-1',
                 'Deviation N-1','Marge perdue FCFA','Score sante moy.','Statut',
                 'Fam. en Action','Fam. Vigilance','Famille la plus impactee','Action magasin']
        MAG_W = [28,14,14,11,11,14,16,14,12,12,14,38,44]

        ws_mag.merge_cells(f'A1:{get_column_letter(len(MAG_H))}1')
        cm=ws_mag.cell(row=1,column=1,value='SYNTHESE PAR MAGASIN — DEVIATION MARGE vs N-1')
        cm.font=Font('Calibri',size=13,bold=True,color=C_WH); cm.fill=xfill(C_HDR); cm.alignment=xctr()
        ws_mag.row_dimensions[1].height=30
        ws_mag.merge_cells(f'A2:{get_column_letter(len(MAG_H))}2')
        cm2=ws_mag.cell(row=2,column=1,value=f'  Periodes : {", ".join(periodes)}  —  Classe par deviation croissante')
        cm2.font=Font('Calibri',size=9,italic=True,color='AABBCC'); cm2.fill=xfill(C_HDR); cm2.alignment=xlft()
        ws_mag.row_dimensions[2].height=16

        for j,(h,w) in enumerate(zip(MAG_H,MAG_W),1):
            c=ws_mag.cell(row=3,column=j,value=h)
            c.font=Font('Calibri',size=9,bold=True,color=C_WH)
            c.fill=xfill(C_SUB); c.alignment=xctr(); c.border=xbdr()
            ws_mag.column_dimensions[get_column_letter(j)].width=w
        ws_mag.row_dimensions[3].height=28

        df_sites = df_all[df_all['Periode'].isin(periodes)].copy()
        df_sites = df_sites[df_sites['Site nom long'].notna() &
                            ~df_sites['Site nom long'].isin(['Total','']) &
                            (df_sites['CA']>0)].copy()
        site_rows=[]
        for site in sorted(df_sites['Site nom long'].unique()):
            sub=df_sites[df_sites['Site nom long']==site]
            ca_s=sub['CA'].sum(); mg_s=sub['Marge'].sum()
            tx_s=mg_s/ca_s if ca_s>0 else 0
            mn1_s=sub['Marge_N1'].sum(); cn1_s=sub['CA_N1'].sum()
            tn1_s=mn1_s/cn1_s if cn1_s>0 else 0
            dev_s=tx_s-tn1_s
            n_act=(sub['Statut']=='Action').sum()
            n_vig=(sub['Statut']=='Vigilance').sum()
            top_f=sub.nlargest(1,'Impact_Score')
            top_fam=top_f['SF_court'].values[0] if len(top_f) else '—'
            top_dev=top_f['Dev_N1_pts'].values[0] if len(top_f) else None
            score_moy=sub['Score_Sante'].mean() if 'Score_Sante' in sub.columns else 50
            if dev_s < -0.030:   stat_s='Action'
            elif dev_s < -0.015: stat_s='Vigilance'
            else:                stat_s='OK'
            if n_act>=5:      action_m=f'{n_act} alertes — reunion acheteur + revue fournisseurs'
            elif n_act>=2:    action_m=f'{n_act} alertes — priorite {top_fam} ({top_dev:+.1f} pts)' if pd.notna(top_dev) else f'{n_act} alertes'
            elif n_act==1:    action_m=f'Traiter {top_fam} cette semaine'
            elif n_vig>=3:    action_m=f'{n_vig} en vigilance — surveiller la tendance'
            else:             action_m='RAS — maintenir les conditions'
            site_rows.append({'site':site,'ca':ca_s,'mg':mg_s,'tx':tx_s,'tn1':tn1_s,
                              'dev':dev_s,'dmg':sub['Dev_N1_FCFA'].sum(),
                              'score':round(score_moy),'stat':stat_s,
                              'n_act':n_act,'n_vig':n_vig,'top':top_fam,'action':action_m})
        site_rows.sort(key=lambda x: x['dev'])

        STAT_C={'Action':('FFD6D6','991B1B'),'Vigilance':('FFF3CC','854D0E'),'OK':('D6F5D6','145A32')}
        for i,row in enumerate(site_rows,4):
            bg='F7F7F7' if i%2==0 else 'FFFFFF'
            s_bg,s_fg=STAT_C.get(row['stat'],('FFFFFF','1A1A2E'))
            data=[(row['site'],bg,'1A1A2E',True,'left',None),
                  (row['ca'],bg,'1A1A2E',False,'right','#,##0'),
                  (row['mg'],bg,'1A1A2E',False,'right','#,##0'),
                  (row['tx'],bg,'1A1A2E',False,'center','0.0%'),
                  (row['tn1'],bg,'1A1A2E',False,'center','0.0%'),
                  (row['dev'],bg,s_fg,True,'center','+0.0%;-0.0%;-'),
                  (row['dmg'],bg,s_fg,True,'right','+#,##0;-#,##0;-'),
                  (row['score'],bg,'1A1A2E',False,'center','0'),
                  (row['stat'],s_bg,s_fg,True,'center',None),
                  (row['n_act'],'FFD6D6' if row['n_act']>0 else bg,'991B1B' if row['n_act']>0 else '1A1A2E',True,'center',None),
                  (row['n_vig'],'FFF3CC' if row['n_vig']>0 else bg,'854D0E' if row['n_vig']>0 else '1A1A2E',True,'center',None),
                  (row['top'],bg,'1A1A2E',False,'left',None),
                  (row['action'],bg,'1A1A2E',False,'left',None)]
            for j,(v,bg_c,fg,bold,align,fmt) in enumerate(data,1):
                c=ws_mag.cell(row=i,column=j,value=v)
                c.fill=xfill(bg_c); c.font=Font('Calibri',size=9,bold=bold,color=fg)
                c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=(j>=12))
                c.border=xbdr()
                if fmt: c.number_format=fmt
            ws_mag.row_dimensions[i].height=20

        # Ligne totale
        r_t=len(site_rows)+4
        ws_mag.merge_cells(f'A{r_t}:A{r_t}')
        tt=ws_mag.cell(row=r_t,column=1,value='TOTAL RESEAU')
        tt.font=Font('Calibri',size=10,bold=True,color=C_WH); tt.fill=xfill(C_HDR)
        tt.alignment=xlft(); tt.border=xbdr()
        totals=[(df_sites['CA'].sum(),'right','#,##0'),(df_sites['Marge'].sum(),'right','#,##0'),
                (df_sites['Marge'].sum()/df_sites['CA'].sum() if df_sites['CA'].sum()>0 else 0,'center','0.0%'),
                (df_sites['Marge_N1'].sum()/df_sites['CA_N1'].sum() if df_sites['CA_N1'].sum()>0 else 0,'center','0.0%'),
                (df_sites['Marge'].sum()/df_sites['CA'].sum()-(df_sites['Marge_N1'].sum()/df_sites['CA_N1'].sum() if df_sites['CA_N1'].sum()>0 else 0),'center','+0.0%;-0.0%;-'),
                (df_sites['Dev_N1_FCFA'].sum(),'right','+#,##0;-#,##0;-'),
                (df_sites['Score_Sante'].mean() if 'Score_Sante' in df_sites.columns else 50,'center','0')]
        for j,(v,align,fmt) in enumerate(totals,2):
            c=ws_mag.cell(row=r_t,column=j,value=v)
            c.font=Font('Calibri',size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB)
            c.alignment=Alignment(horizontal=align,vertical='center'); c.border=xbdr()
            c.number_format=fmt
        ws_mag.row_dimensions[r_t].height=22
        ws_mag.freeze_panes='A4'
        ws_mag.auto_filter.ref=f'A3:{get_column_letter(len(MAG_H))}{r_t-1}'

    # ── Onglet Lexique ─────────────────────────────────────────────────────────
    ws_lex = wb.create_sheet("Guide de lecture")
    ws_lex.sheet_view.showGridLines = False
    ws_lex.column_dimensions['A'].width = 24
    ws_lex.column_dimensions['B'].width = 48
    ws_lex.column_dimensions['C'].width = 52

    ws_lex.merge_cells('A1:C1')
    tl=ws_lex.cell(row=1,column=1,value='GUIDE DE LECTURE — MODULE RENTABILITE SMARTBUYER v2.3')
    tl.font=Font('Calibri',size=13,bold=True,color=C_WH); tl.fill=xfill(C_HDR); tl.alignment=xctr()
    ws_lex.row_dimensions[1].height=30

    def lex_sec(ws,row,txt):
        ws.merge_cells(f'A{row}:C{row}')
        c=ws.cell(row=row,column=1,value=txt)
        c.font=Font('Calibri',size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft()
        ws.row_dimensions[row].height=22

    def lex_row(ws,row,vals,bg='FFFFFF',h=42):
        for j,v in enumerate(vals,1):
            c=ws.cell(row=row,column=j,value=v)
            c.font=Font('Calibri',size=9,color='1A1A2E',bold=(j==1))
            c.fill=PatternFill('solid',fgColor=bg)
            c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            c.border=xbdr()
        ws.row_dimensions[row].height=h

    lex_sec(ws_lex,2,'  1. COLONNES CLES')
    for j,h in enumerate(['Colonne','Ce que ca mesure','Comment lire'],1):
        c=ws_lex.cell(row=3,column=j,value=h)
        c.font=Font('Calibri',size=9,bold=True,color='3A3A3C')
        c.fill=PatternFill('solid',fgColor='F2F2F7'); c.alignment=xctr(); c.border=xbdr()
    ws_lex.row_dimensions[3].height=20

    COLS=[
        ("Taux brut","Taux de marge brute realise. Formule : Marge / CA x 100",
         "Ex 18,5% = sur 100 FCFA vendus, 18,5 FCFA restent en marge avant charges"),
        ("Tx marge nette (casse)","Taux apres deduction de la casse. Mesure l'impact operationnel.",
         "Ex 18,5% brut -> 17,0% net = la casse greve 1,5 pt de marge. Levier magasin, pas acheteur."),
        ("Marge perdue FCFA","Marge absolue perdue vs N-1 sur la periode. Clé pour prioriser.",
         "Ex -920 K FCFA = si rien ne change cette semaine, 920 K de moins vs l'annee derniere."),
        ("Score sante (0-100)","Score unique : déviation N-1 (50%) + atteinte cible (30%) + impact casse (20%).",
         "100 = parfait. 70+ = satisfaisant. 50-70 = vigilance. <50 = action requise."),
        ("Cause","Diagnostic automatique de la cause de la deviation.",
         "Promo deficitaire / Conditions achat / Effet mix / Produit appel / Chute severe PA"),
        ("Action","1 action concrete recommandee pour cette famille cette semaine.",
         "Toujours 1 phrase directe et actionnable. Pas d'analyse supplementaire necessaire."),
        ("Remise necessaire FCFA","Montant a recuperer en conditions achat pour atteindre la cible.",
         "Ex 400 K FCFA = argument chiffre a presenter au fournisseur en negociation."),
    ]
    for i,(col,mesure,lecture) in enumerate(COLS):
        lex_row(ws_lex,4+i,[col,mesure,lecture],bg='FFFFFF' if i%2==0 else 'F7F7F7')

    r2=4+len(COLS)+1
    lex_sec(ws_lex,r2,'  2. STATUTS ET SCORE SANTE')
    for j,h in enumerate(['Statut','Seuil','Exemple + action'],1):
        c=ws_lex.cell(row=r2+1,column=j,value=h)
        c.font=Font('Calibri',size=9,bold=True,color='3A3A3C')
        c.fill=PatternFill('solid',fgColor='F2F2F7'); c.alignment=xctr(); c.border=xbdr()
    ws_lex.row_dimensions[r2+1].height=20

    STATS=[
        ('OK','D8F5D8','145A32','Taux actuel >= Cible - 1,5 pt et score > 65',
         'Cafe Soluble 21,8% / cible 21,0% / score 72 -> RAS. Surveiller la semaine suivante.'),
        ('Vigilance','FEF9C3','854D0E','Taux entre Cible - 1,5 pt et Cible - 3 pts / score 50-65',
         'Biscuits 19,2% / cible 21,5% -> Surveiller. Pas d urgence mais a confirmer cette semaine.'),
        ('Action','FFD6D6','991B1B','Taux < Cible - 3 pts OU score < 50',
         'Huile 8,1% / cible 12,0% / score 32 -> Revoir PA fournisseur. Action cette semaine.'),
    ]
    for i,(stat,bg,fg,seuil,ex) in enumerate(STATS):
        r=r2+2+i
        for j,v in enumerate([stat,seuil,ex],1):
            c=ws_lex.cell(row=r,column=j,value=v)
            c.fill=PatternFill('solid',fgColor=bg)
            c.font=Font('Calibri',size=9,bold=(j==1),color=fg if j==1 else '1A1A2E')
            c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            c.border=xbdr()
        ws_lex.row_dimensions[r].height=52

    r3=r2+2+len(STATS)+1
    lex_sec(ws_lex,r3,'  3. SEGMENTATION — PLANCHERS DE MARGE')
    SEGS=[
        ("Produit d appel","FFF3E0","B25000","10%","Riz, Huile, Lait, Eau, Sucre, Farine, Pates, Semoules — prix contraints par le marche"),
        ("Coeur de gamme","E3F2FD","0D47A1","18%","PGC standard Epicerie et Boissons — objectif atteignable via negociation fournisseur"),
        ("Valeur ajoutee","E8F5E9","1B5E20","25%","BIO, Chips, Snacking, Cosmetique — moins de concurrence prix"),
        ("PH / Droguerie","FCE4EC","880E4F","22%","Parfumerie, Hygiene, Droguerie — structurellement plus riche"),
    ]
    for j,h in enumerate(['Segment','Plancher','Familles concernees'],1):
        c=ws_lex.cell(row=r3+1,column=j,value=h)
        c.font=Font('Calibri',size=9,bold=True,color='3A3A3C')
        c.fill=PatternFill('solid',fgColor='F2F2F7'); c.alignment=xctr(); c.border=xbdr()
    for i,(seg,bg,fg,plancher,fam) in enumerate(SEGS):
        r=r3+2+i
        for j,v in enumerate([seg,plancher,fam],1):
            c=ws_lex.cell(row=r,column=j,value=v)
            c.fill=PatternFill('solid',fgColor=bg)
            c.font=Font('Calibri',size=9,bold=(j in(1,2)),color=fg if j in(1,2) else '1A1A2E')
            c.alignment=Alignment(horizontal='center' if j==2 else 'left',vertical='center',wrap_text=True)
            c.border=xbdr()
        ws_lex.row_dimensions[r].height=46

    r_note=r3+2+len(SEGS)+1
    ws_lex.merge_cells(f'A{r_note}:C{r_note}')
    cn=ws_lex.cell(row=r_note,column=1,
        value="SmartBuyer Hub v2.3 — NovaRetail Solutions — Cibles revisables chaque debut d'exercice via le referentiel embarque.")
    cn.font=Font('Calibri',size=8,italic=True,color='8E8E93')
    cn.fill=PatternFill('solid',fgColor='F9F9FB')
    cn.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws_lex.row_dimensions[r_note].height=20

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Navigation</div>", unsafe_allow_html=True)
    st.page_link("pages/home.py",                               label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",          label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",                   label="📈  Ventes PBI")
    st.page_link("pages/03_📦_Detention_Top_CA.py",             label="📦  Détention Top CA")
    st.page_link("pages/04_💸_Performance_Promo.py",            label="💸  Performance Promo")
    st.page_link("pages/05_🏪_Suivi_Implantation.py",           label="🏪  Suivi Implantation")
    st.page_link("pages/06_💸_Marges_Negatives.py",             label="💸  Marges Négatives")
    st.page_link("pages/07_📈_OTIF.py",                         label="📈  OTIF")
    st.page_link("pages/08_📦_OOS.py",                          label="📦  OOS Ruptures")
    st.page_link("pages/09_✅_Tasks_Trackers.py",               label="✅  Tasks Tracker")
    st.page_link("pages/10_📊_Perf_Hebdo.py",                   label="📊  Perf Hebdo")
    st.page_link("pages/11_📊_Rentabilite.py",                  label="📊  Rentabilité")
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import</div>", unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Extraction(s) PBI", type=['xlsx'], accept_multiple_files=True,
        help=f"Semaine ou journée — période détectée automatiquement.\nColonnes : {', '.join(COLS_REQUIRED)}"
    )
    ref_override = st.file_uploader("Référentiel cibles (optionnel)", type=['xlsx'],
                                     help="Laissez vide → référentiel embarqué.")
    st.markdown("---")
    st.caption("NovaRetail Solutions · SmartBuyer v2.3")

# ─── ÉCRAN VIDE ───────────────────────────────────────────────────────────────
if not uploaded_files:
    st.markdown("<div class='page-title'>📊 Rentabilité</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-caption'>Cockpit Direction · Briefing Acheteur · Analyse Approfondie</div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>Chargez une extraction PBI</strong> dans la sidebar pour démarrer.<br>
  Fonctionne avec une extraction journalière ou hebdomadaire. La période est détectée automatiquement.
</div>""", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("""
<div class='verdict-box'>
  <div class='verdict-title'>Zone 1 — Cockpit Direction</div>
  <div style='font-size:13px;color:#3A3A3C'>Le réseau est-il en bonne santé ?<br>
  4 KPIs · Tableau par rayon · Phrase de verdict · Tendance</div>
</div>""", unsafe_allow_html=True)
    with c2:
        st.markdown("""
<div class='verdict-box'>
  <div class='verdict-title'>Zone 2 — Briefing Acheteur</div>
  <div style='font-size:13px;color:#3A3A3C'>Que faire cette semaine ?<br>
  3 actions prioritaires · Tableau condensé · Casse séparée</div>
</div>""", unsafe_allow_html=True)
    with c3:
        st.markdown("""
<div class='verdict-box'>
  <div class='verdict-title'>Zone 3 — Analyse</div>
  <div style='font-size:13px;color:#3A3A3C'>Pourquoi ? Sur quel magasin ?<br>
  Fiche famille · Vue magasin · Vue promo</div>
</div>""", unsafe_allow_html=True)
    st.stop()

# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
ref_bytes  = ref_override.read() if ref_override else None
all_dfs, errors = [], []
for f in uploaded_files:
    raw = f.read()
    try:    all_dfs.append(load_extraction(raw, f.name, ref_bytes))
    except ValueError as e: errors.append(str(e))
    except Exception  as e: errors.append(f"Erreur **{f.name}** : {e}")
for err in errors: st.sidebar.error(err)
if not all_dfs:
    st.error("Aucun fichier valide. Vérifiez les colonnes obligatoires.")
    st.stop()

df_all = pd.concat(all_dfs, ignore_index=True)
periodes_dispo = sorted(df_all['Periode'].unique(), reverse=True)

with st.sidebar:
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px'>Périodes chargées</div>", unsafe_allow_html=True)
    for p in periodes_dispo:
        st.markdown(f"<span style='background:#E3F0FF;color:#185FA5;border-radius:20px;padding:3px 10px;font-size:11px;font-weight:500;display:inline-block;margin:2px 0'>{p}</span>", unsafe_allow_html=True)
    st.markdown("")
    periode_sel = st.selectbox("Période active", periodes_dispo, label_visibility='collapsed')

df = df_all[df_all['Periode'] == periode_sel].copy()
df = df.sort_values(['_ord_statut', 'Impact_Score'], ascending=[True, False])

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📊 Rentabilité</div>", unsafe_allow_html=True)
label_periode = "Journée" if "→" not in periode_sel else "Semaine"
st.markdown(f"<div class='page-caption'>{label_periode} : <strong>{periode_sel}</strong> · {len(periodes_dispo)} période(s) chargée(s)</div>", unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs([
    "📊 Cockpit Direction",
    "👤 Briefing Acheteur",
    "🔍 Analyse",
    "📥 Export Excel",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — COCKPIT DIRECTION
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    ca_t  = df['CA'].sum(); mg_t  = df['Marge'].sum(); tx_t = mg_t/ca_t if ca_t>0 else 0
    mn1_t = df['Marge_N1'].sum(); cn1_t = df['CA_N1'].sum()
    tx_n1 = mn1_t/cn1_t if cn1_t>0 else 0
    dev_t = tx_t - tx_n1
    cib_t = (df['Cible']*df['CA']).sum()/ca_t if ca_t>0 else 0
    perdu = df['Dev_N1_FCFA'].sum()
    n_act = (df['Statut']=='Action').sum()
    score_moy = df['Score_Sante'].mean() if 'Score_Sante' in df.columns else 50

    # Tendance si multi-périodes
    tendance_txt = ""
    if len(periodes_dispo) >= 2:
        p_prev  = sorted(periodes_dispo)[periodes_dispo.index(periode_sel)-1] if periodes_dispo.index(periode_sel) > 0 else None
        if p_prev:
            dp = df_all[df_all['Periode']==p_prev]
            ca_p=dp['CA'].sum(); mg_p=dp['Marge'].sum()
            tx_p=mg_p/ca_p if ca_p>0 else 0
            mn1_p=dp['Marge_N1'].sum(); cn1_p=dp['CA_N1'].sum()
            tn1_p=mn1_p/cn1_p if cn1_p>0 else 0
            dev_p=tx_p-tn1_p
            if dev_t > dev_p + 0.005:   tendance_txt = "Amélioration vs période précédente"
            elif dev_t < dev_p - 0.005: tendance_txt = "Dégradation vs période précédente"
            else:                        tendance_txt = "Stable vs période précédente"

    # Verdict automatique
    verdict = _verdict_auto(df, periode_sel)

    # Phrase verdict
    st.markdown(f"""
<div class='verdict-box'>
  <div class='verdict-title'>Verdict {label_periode.lower()} — {periode_sel}</div>
  <div class='verdict-text'>{verdict}</div>
  {f"<div style='font-size:12px;color:#8E8E93;margin-top:6px'>{tendance_txt}</div>" if tendance_txt else ""}
</div>""", unsafe_allow_html=True)

    # 4 KPIs
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Taux de marge", f"{tx_t:.1%}", fp(dev_t))
    k2.metric("Marge perdue", fk(perdu), f"vs {tx_n1:.1%} N-1")
    k3.metric(f"Score santé réseau", f"{score_moy:.0f}/100",
              "Bon" if score_moy >= 70 else ("Vigilance" if score_moy >= 50 else "Critique"))
    k4.metric("Familles en action", str(n_act), f"/ {len(df)} familles")

    st.markdown("---")

    # Tableau par rayon — 1 ligne par rayon, 6 colonnes max
    st.markdown("<div class='section-label'>État par rayon — vue direction</div>", unsafe_allow_html=True)
    RAYON_COLORS = {'BOISSONS':'#E3F0FF','EPICERIE':'#EBF5E0','DROGUERIE':'#FFF8E1','PARFUMERIE HYGIENE':'#FCE4EC'}
    RAYON_ACHETEUR = {'BOISSONS':'Acheteur Boissons','EPICERIE':'Acheteur Épicerie','DROGUERIE':'Acheteur DPH','PARFUMERIE HYGIENE':'Acheteur DPH'}

    rows_dir = []
    for rayon in ORDRE_RAYONS:
        sub = df[df['Rayon_court']==rayon]
        if len(sub)==0: continue
        ca_r=sub['CA'].sum(); mg_r=sub['Marge'].sum(); tx_r=mg_r/ca_r if ca_r>0 else 0
        mn1_r=sub['Marge_N1'].sum(); cn1_r=sub['CA_N1'].sum(); tn1_r=mn1_r/cn1_r if cn1_r>0 else 0
        dev_r=tx_r-tn1_r
        n_r_act=(sub['Statut']=='Action').sum()
        score_r=sub['Score_Sante'].mean() if 'Score_Sante' in sub.columns else 50
        # Action principale du rayon
        top_act=sub[sub['Statut']=='Action'].nlargest(1,'Impact_Score')
        action_dir='RAS — maintenir les conditions'
        if len(top_act):
            r0=top_act.iloc[0]
            action_dir = f"{r0['SF_court']} : {r0['Action']}"

        # Santé badge
        if score_r >= 70:   sante = f"<span class='sante-vert'>{score_r:.0f}/100</span>"
        elif score_r >= 50: sante = f"<span class='sante-orange'>{score_r:.0f}/100</span>"
        else:               sante = f"<span class='sante-rouge'>{score_r:.0f}/100</span>"

        rows_dir.append({
            'Rayon': rayon.title(),
            'Acheteur': RAYON_ACHETEUR.get(rayon,'—').replace('Acheteur ',''),
            'Taux': fp(tx_r, False),
            'vs N-1': fp(dev_r),
            'Marge perdue': fk(sub['Dev_N1_FCFA'].sum()),
            'Alertes': n_r_act,
            'Sante': sante,
            'Action semaine': action_dir,
        })

    df_dir = pd.DataFrame(rows_dir)
    st.dataframe(
        df_dir.drop(columns=['Sante']).style.map(cd, subset=['vs N-1']),
        use_container_width=True, hide_index=True, height=210,
        column_config={
            'Action semaine': st.column_config.TextColumn('Action semaine', width='large'),
        }
    )
    # Score santé en HTML
    sante_html = " &nbsp;·&nbsp; ".join([f"<strong>{r['Rayon']}</strong> {r['Sante']}" for r in rows_dir])
    st.markdown(f"<div style='font-size:12px;margin-top:4px'>Score santé : {sante_html}</div>", unsafe_allow_html=True)

    # Alertes casse réseau
    fam_casse = df[df.get('Alerte_Casse', pd.Series(False, index=df.index))==True] if 'Alerte_Casse' in df.columns else pd.DataFrame()
    if len(fam_casse) > 0:
        st.markdown("---")
        st.markdown("<div class='section-label'>Casse anormale — levier opérationnel magasin</div>", unsafe_allow_html=True)
        top_c = fam_casse.nlargest(5,'Tx_Casse_Fam')[['Rayon_court','SF_court','%Marge','Tx_Casse_Fam','Marge_Nette','Casse_val']]
        dc = top_c.copy()
        dc.columns=['Rayon','Famille','Tx brut','% Casse CA','Tx net casse','Casse FCFA']
        dc['Tx brut']    = dc['Tx brut'].apply(lambda x: fp(x,False))
        dc['% Casse CA'] = dc['% Casse CA'].apply(lambda x: f"{x:.2%}")
        dc['Tx net casse']= dc['Tx net casse'].apply(lambda x: fp(x,False))
        dc['Casse FCFA'] = dc['Casse FCFA'].apply(lambda x: f"{abs(x):,.0f}")
        st.dataframe(dc, use_container_width=True, hide_index=True)
        st.caption("La casse est un levier opérationnel (magasin), pas un levier achat. Ces familles ne déclenchent pas d'alerte marge — elles remontent ici séparément.")

    # Tendance multi-périodes
    if len(periodes_dispo) > 1:
        st.markdown("---")
        st.markdown("<div class='section-label'>Tendance — évolution période par période</div>", unsafe_allow_html=True)
        rows_t=[]
        for p in sorted(periodes_dispo):
            dp=df_all[df_all['Periode']==p]; ca_p=dp['CA'].sum(); mg_p=dp['Marge'].sum()
            tx_p=mg_p/ca_p if ca_p>0 else 0
            mn1_p=dp['Marge_N1'].sum(); cn1_p=dp['CA_N1'].sum(); tn1_p=mn1_p/cn1_p if cn1_p>0 else 0
            sc_p=dp['Score_Sante'].mean() if 'Score_Sante' in dp.columns else 50
            rows_t.append({'Période':p,'Taux':fp(tx_p,False),'vs N-1':fp(tx_p-tn1_p),
                           'Marge Δ':fk(dp['Dev_N1_FCFA'].sum()),
                           'Score santé':f"{sc_p:.0f}/100",
                           'Alertes':(dp['Statut']=='Action').sum()})
        st.dataframe(pd.DataFrame(rows_t).style.map(cd,subset=['vs N-1']),
                     use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — BRIEFING ACHETEUR
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    acheteurs = sorted(df['Acheteur'].dropna().unique())
    ach_sel   = st.selectbox("Acheteur", acheteurs, key='ach_sel')
    df_ach    = df[df['Acheteur']==ach_sel].sort_values(['_ord_statut','Impact_Score'],ascending=[True,False])

    ca_a  = df_ach['CA'].sum(); mg_a=df_ach['Marge'].sum(); tx_a=mg_a/ca_a if ca_a>0 else 0
    mn1_a = df_ach['Marge_N1'].sum(); cn1_a=df_ach['CA_N1'].sum(); tn1_a=mn1_a/cn1_a if cn1_a>0 else 0
    cib_a = (df_ach['Cible']*df_ach['CA']).sum()/ca_a if ca_a>0 else 0
    n_r_a = (df_ach['Statut']=='Action').sum()
    score_a = df_ach['Score_Sante'].mean() if 'Score_Sante' in df_ach.columns else 50
    n_nouv  = (df_ach['Tx_N1'].isna()).sum()

    # KPIs acheteur
    k1,k2,k3,k4 = st.columns(4)
    k1.metric("Taux réalisé", f"{tx_a:.1%}", fp(tx_a-tn1_a))
    k2.metric("Marge perdue", fk(df_ach['Dev_N1_FCFA'].sum()), f"N-1 : {tn1_a:.1%}")
    k3.metric("Score santé", f"{score_a:.0f}/100", f"Cible : {cib_a:.1%}")
    k4.metric("Familles en action", str(n_r_a), f"/ {len(df_ach)} familles")

    # Info base de comparaison
    n1_src_txt = f"Cible basée sur N-1 × 1,02"
    if n_nouv > 0:
        n1_src_txt += f" · {n_nouv} nouveauté(s) comparée(s) sur plancher segment uniquement"
    st.caption(n1_src_txt)

    st.markdown("---")

    # BLOC 3 ACTIONS — toujours visible, pas dans un expander
    if n_r_a > 0:
        actions = _top3_actions(df_ach)
        marge_tot = abs(df_ach[df_ach['Statut']=='Action']['Dev_N1_FCFA'].sum())
        st.markdown(f"<div class='section-label'>Tes {min(3, len(actions))} actions cette {label_periode.lower()} — {fk_abs(marge_tot)} en jeu</div>", unsafe_allow_html=True)

        for i, act in enumerate(actions, 1):
            site_txt = f" · {act['site']}" if act['site'] and act['site'] not in ['—',''] else ""
            color_cls = act['icone']
            st.markdown(f"""
<div class='action-card {color_cls}'>
  <div class='action-num'>Action {i}{site_txt}</div>
  <div class='action-fam'>{act['famille']}</div>
  <div class='action-fcfa {color_cls}'>{act['perdu']} perdus</div>
  <div class='action-what'>{act['action']} · <em style='color:#8E8E93'>{act['cause']}</em></div>
</div>""", unsafe_allow_html=True)

        # Toutes les alertes en dessous
        if n_r_a > 3:
            st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin:8px 0'>+ {n_r_a-3} autre(s) famille(s) en action — voir le tableau ci-dessous</div>", unsafe_allow_html=True)
    else:
        st.markdown(f"""
<div class='alert-card alert-green'>
  Aucune alerte rouge cette {label_periode.lower()} pour {ach_sel}. Score santé : {score_a:.0f}/100.
  Surveiller les familles en vigilance ci-dessous.
</div>""", unsafe_allow_html=True)

    # Alerte casse acheteur
    if 'Alerte_Casse' in df_ach.columns:
        fam_c_ach = df_ach[df_ach['Alerte_Casse']==True]
        if len(fam_c_ach) > 0:
            noms_c = ' · '.join(fam_c_ach.nlargest(3,'Tx_Casse_Fam')['SF_court'].tolist())
            st.markdown(f"""
<div class='alert-card alert-purple' style='margin-top:8px'>
  Casse anormale (> 0,5% CA) sur : <strong>{noms_c}</strong> — sujet opérationnel à remonter au responsable magasin.
</div>""", unsafe_allow_html=True)

    st.markdown("---")

    # TABLEAU CONDENSÉ — 5 colonnes essentielles
    fc1, fc2 = st.columns([2, 2])
    with fc1:
        filtre_act = st.checkbox("Actions uniquement", value=False, key='filtre_act')
    with fc2:
        masque_appel = st.checkbox("Masquer produits d'appel OK", value=True, key='masque_appel',
                                    help="Masque les produits d'appel au-dessus de leur plancher.")

    df_view = df_ach.copy()
    if filtre_act:
        df_view = df_view[df_view['Statut']=='Action']
    if masque_appel:
        df_view = df_view[~((df_view['Segment']=='Produit d appel') & (df_view['%Marge']>=df_view['Plancher']))]

    n_aff = len(df_view); n_cach = len(df_ach)-n_aff
    lbl = f"{n_aff} famille(s)" + (f" · {n_cach} masquées" if n_cach>0 and masque_appel else "")
    st.markdown(f"<div class='section-label'>{lbl}</div>", unsafe_allow_html=True)

    has_site_a = 'Site nom long' in df_view.columns and df_view['Site nom long'].notna().any()

    # Vue condensée : Famille · Santé · Marge perdue · Cause · Action
    cv = ['SF_court']
    if has_site_a: cv.append('Site nom long')
    cv += ['Score_Sante','%Marge','Dev_N1_FCFA','Cause','Statut','Action']
    cn = ['Famille']
    if has_site_a: cn.append('Magasin')
    cn += ['Santé','Taux actuel','Marge perdue','Cause','Statut','Action']

    dv = df_view[[c for c in cv if c in df_view.columns]].copy()
    dv.columns = cn[:len(dv.columns)]
    if 'Taux actuel' in dv.columns: dv['Taux actuel'] = dv['Taux actuel'].apply(lambda x: fp(x,False))
    if 'Marge perdue'in dv.columns: dv['Marge perdue'] = dv['Marge perdue'].apply(fk)
    if 'Santé'       in dv.columns: dv['Santé']       = dv['Santé'].apply(lambda x: f"{x:.0f}/100" if pd.notna(x) else '—')

    def color_sante(v):
        try:
            n = float(str(v).replace('/100',''))
            if n >= 70: return 'color:#145A32;font-weight:600'
            if n >= 50: return 'color:#854D0E;font-weight:600'
            return 'color:#991B1B;font-weight:600'
        except: return ''

    def color_cause_v(v):
        v = str(v).lower()
        if 'négative' in v or 'negative' in v: return 'color:#CC0000;font-weight:700'
        if 'promo' in v:    return 'color:#B25000;font-weight:600'
        if 'sévère' in v or 'severe' in v: return 'color:#991B1B;font-weight:600'
        if 'mix'    in v:   return 'color:#6B21A8;font-weight:600'
        if 'ok'     in v:   return 'color:#145A32'
        return 'color:#3A3A3C'

    s_dv = dv.style
    if 'Statut'     in dv.columns: s_dv = s_dv.map(cs,            subset=['Statut'])
    if 'Santé'      in dv.columns: s_dv = s_dv.map(color_sante,   subset=['Santé'])
    if 'Cause'      in dv.columns: s_dv = s_dv.map(color_cause_v, subset=['Cause'])

    st.dataframe(s_dv, use_container_width=True, hide_index=True, height=480,
        column_config={
            'Famille': st.column_config.TextColumn('Famille', width='large'),
            'Action':  st.column_config.TextColumn('Action',  width='large'),
            'Cause':   st.column_config.TextColumn('Cause',   width='medium'),
        })

    # Détail analytique complet en expander
    with st.expander("Détail analytique complet"):
        cols_d = ['Rayon_court','SF_court','Segment']
        noms_d = ['Rayon','Famille','Segment']
        if has_site_a: cols_d.insert(2,'Site nom long'); noms_d.insert(2,'Magasin')
        extra = [('CA','CA'),('Marge','Marge brute'),('%Marge','Tx brut'),
                 ('Marge_Nette','Tx net casse'),('Tx_N1','Tx N-1'),('Cible','Cible'),
                 ('Dev_N1_pts','Dév. N-1'),('Dev_N1_FCFA','Marge Δ'),
                 ('Tx_Casse_Fam','% Casse'),('Remise_Necessaire','Remise nec.'),('Statut','Statut')]
        for ec,en in extra:
            if ec in df_view.columns: cols_d.append(ec); noms_d.append(en)
        dd = df_view[cols_d].copy(); dd.columns=noms_d
        for c in ['Tx brut','Tx net casse','Tx N-1','Cible']:
            if c in dd.columns: dd[c]=dd[c].apply(lambda x: fp(x,False))
        if 'Dév. N-1' in dd.columns: dd['Dév. N-1']=dd['Dév. N-1'].apply(fp)
        if '% Casse'  in dd.columns: dd['% Casse']=dd['% Casse'].apply(lambda x: f"{x:.2%}" if pd.notna(x) else '—')
        if 'CA'       in dd.columns: dd['CA']=dd['CA'].apply(lambda x: f"{x:,.0f}")
        if 'Marge brute' in dd.columns: dd['Marge brute']=dd['Marge brute'].apply(lambda x: f"{x:,.0f}")
        if 'Marge Δ'  in dd.columns: dd['Marge Δ']=dd['Marge Δ'].apply(fk)
        if 'Remise nec.' in dd.columns: dd['Remise nec.']=dd['Remise nec.'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) and x>0 else '—')
        if 'Segment'  in dd.columns: dd['Segment']=dd['Segment'].apply(lambda x: SEG_LABELS.get(x,x))
        s_dd = dd.style.map(cs,subset=['Statut']) if 'Statut' in dd.columns else dd
        st.dataframe(s_dd, use_container_width=True, hide_index=True)
    st.caption("Vue principale = 5 colonnes pour décider. Détail analytique complet dans l'expander.")

    # Tendance acheteur si multi-périodes
    if len(periodes_dispo) >= 2:
        p_prev = sorted(periodes_dispo)[max(0, periodes_dispo.index(periode_sel)-1)]
        if p_prev != periode_sel:
            df_prv = df_all[(df_all['Periode']==p_prev)&(df_all['Acheteur']==ach_sel)][['SF_court','Statut']].copy()
            df_prv.columns=['SF_court','Statut_prev']
            df_td  = df_view[['SF_court','Statut']].merge(df_prv,on='SF_court',how='left')
            n_dg   = ((df_td['Statut']=='Action')&(df_td['Statut_prev']!='Action')).sum()
            n_am   = ((df_td['Statut']!='Action')&(df_td['Statut_prev']=='Action')).sum()
            if n_dg > 0 or n_am > 0:
                parts=[]
                if n_dg>0: parts.append(f"{n_dg} famille(s) passée(s) en alerte")
                if n_am>0: parts.append(f"{n_am} famille(s) résolue(s)")
                st.markdown(
                    f"<div class='alert-card alert-blue' style='margin-top:8px'>Tendance vs {p_prev} : {' · '.join(parts)}</div>",
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — ANALYSE APPROFONDIE
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    sous_tab1, sous_tab2, sous_tab3 = st.tabs(["🔎 Fiche famille","🏪 Vue magasin","💸 Vue promo"])

    # ── Fiche famille ─────────────────────────────────────────────────────────
    with sous_tab1:
        st.markdown("<div class='section-label'>Analyse détaillée d'une famille</div>", unsafe_allow_html=True)
        familles_dispo = sorted(df['SF_court'].dropna().unique())
        fam_sel = st.selectbox("Sélectionner une famille", familles_dispo, key='fam_sel')
        df_fam  = df[df['SF_court']==fam_sel]

        if len(df_fam)==0:
            st.info("Famille non trouvée dans la période active.")
        else:
            r0 = df_fam.iloc[0]
            # En-tête fiche
            col_left, col_right = st.columns([2,1])
            with col_left:
                st.markdown(f"<div style='font-size:20px;font-weight:700;color:#1C1C1E'>{fam_sel}</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-size:13px;color:#8E8E93'>{r0['Rayon_court'].title()} · {SEG_LABELS.get(r0['Segment'],r0['Segment'])} · {r0.get('Acheteur','—')}</div>", unsafe_allow_html=True)
            with col_right:
                score = r0.get('Score_Sante', 50)
                color = '#34C759' if score>=70 else ('#FF9500' if score>=50 else '#FF3B30')
                st.markdown(f"<div style='text-align:right;font-size:36px;font-weight:700;color:{color}'>{score:.0f}<span style='font-size:16px;color:#8E8E93'>/100</span></div>", unsafe_allow_html=True)
                st.markdown(f"<div style='text-align:right;font-size:12px;color:#8E8E93'>Score de santé</div>", unsafe_allow_html=True)

            st.markdown("")
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Taux brut",     fp(r0.get('%Marge'), False), fp(r0.get('Dev_N1_pts')))
            m2.metric("Taux N-1",      fp(r0.get('Tx_N1'), False))
            m3.metric("Cible",         fp(r0.get('Cible'), False))
            m4.metric("Marge perdue",  fk(r0.get('Dev_N1_FCFA',0)))

            # Casse
            tx_casse = r0.get('Tx_Casse_Fam', 0)
            if pd.notna(tx_casse) and tx_casse > 0.003:
                st.markdown(f"""
<div class='alert-card alert-purple' style='margin-top:8px'>
  Casse : {tx_casse:.2%} du CA · Taux net après casse : {fp(r0.get('Marge_Nette'), False)}
  {"· <strong>Anormale — remonter au magasin</strong>" if tx_casse > 0.005 else ""}
</div>""", unsafe_allow_html=True)

            # Diagnostic et action
            st.markdown(f"""
<div class='verdict-box' style='margin-top:12px'>
  <div class='verdict-title'>Diagnostic</div>
  <div style='font-size:14px;font-weight:600;color:#1C1C1E'>{r0.get('Cause','—')}</div>
  <div style='font-size:13px;color:#3A3A3C;margin-top:6px'>{r0.get('Action','—')}</div>
</div>""", unsafe_allow_html=True)

            # Décomposition HP vs Promo si disponible
            if 'Marge Hors Promo' in r0.index and 'CA Hors Promo' in r0.index:
                st.markdown("<div class='section-label' style='margin-top:16px'>Décomposition fond de rayon vs promo</div>", unsafe_allow_html=True)
                ca_hp=r0.get('CA Hors Promo',0); mg_hp=r0.get('Marge Hors Promo',0)
                ca_pr=r0.get('CA Promo',0) if 'CA Promo' in r0.index else 0
                mg_pr=r0.get('Marge Promo',0) if 'Marge Promo' in r0.index else 0
                tx_hp=mg_hp/ca_hp if ca_hp and ca_hp>0 else None
                tx_pr=mg_pr/ca_pr if ca_pr and ca_pr>0 else None
                c1d,c2d = st.columns(2)
                c1d.metric("Fond de rayon", fp(tx_hp,False) if tx_hp is not None else '—',
                           f"CA : {fk_abs(ca_hp) if ca_hp else '—'}")
                c2d.metric("Sous promo", fp(tx_pr,False) if tx_pr is not None else '—',
                           f"CA : {fk_abs(ca_pr) if ca_pr else '—'}")
                if tx_hp and tx_pr:
                    effet = tx_pr - tx_hp
                    if effet < -0.03:
                        st.markdown(f"<div class='alert-card alert-amber'>La promo dégrade la marge de {effet:+.1%}. Revoir la mécanique promotionnelle.</div>", unsafe_allow_html=True)

            # Multi-sites si disponible
            if 'Site nom long' in df_fam.columns and df_fam['Site nom long'].notna().any():
                st.markdown("<div class='section-label' style='margin-top:16px'>Par magasin</div>", unsafe_allow_html=True)
                df_sites_fam = df_fam[df_fam['Site nom long'].notna() & ~df_fam['Site nom long'].isin(['Total',''])]
                if len(df_sites_fam) > 0:
                    ds = df_sites_fam[['Site nom long','%Marge','Dev_N1_pts','Dev_N1_FCFA','Score_Sante','Statut']].copy()
                    ds.columns=['Magasin','Taux','Dév. N-1','Marge Δ','Santé','Statut']
                    ds['Taux']  = ds['Taux'].apply(lambda x: fp(x,False))
                    ds['Dév. N-1']=ds['Dév. N-1'].apply(fp)
                    ds['Marge Δ']=ds['Marge Δ'].apply(fk)
                    ds['Santé']=ds['Santé'].apply(lambda x: f"{x:.0f}/100" if pd.notna(x) else '—')
                    ds=ds.sort_values('Marge Δ')
                    st.dataframe(ds.style.map(cs,subset=['Statut']).map(cd,subset=['Dév. N-1']),
                                 use_container_width=True, hide_index=True)

    # ── Vue magasin ───────────────────────────────────────────────────────────
    with sous_tab2:
        has_site = 'Site nom long' in df.columns and df['Site nom long'].notna().any()
        if not has_site:
            st.markdown("<div class='alert-card alert-blue'>Extraction au niveau réseau — ajoutez la dimension <strong>Site nom long</strong> dans PBI pour activer cette vue.</div>", unsafe_allow_html=True)
        else:
            sites = sorted([s for s in df['Site nom long'].dropna().unique() if s not in ['Total','']])
            cf1,cf2 = st.columns([2,2])
            with cf1: site_sel = st.selectbox("Magasin",['Tous']+sites,key='site_sel2')
            with cf2: rayon_f  = st.selectbox("Rayon",  ['Tous']+[r.title() for r in ORDRE_RAYONS],key='rayon_f2')

            df_mag = df.copy()
            if site_sel != 'Tous': df_mag=df_mag[df_mag['Site nom long']==site_sel]
            if rayon_f  != 'Tous': df_mag=df_mag[df_mag['Rayon_court']==rayon_f.upper()]

            if site_sel == 'Tous':
                st.markdown("<div class='section-label'>Palmarès magasins — score de santé moyen</div>", unsafe_allow_html=True)
                rows_s=[]
                for site in sites:
                    sub_s=df[df['Site nom long']==site]; ca_s=sub_s['CA'].sum()
                    if ca_s==0: continue
                    mg_s=sub_s['Marge'].sum(); tx_s=mg_s/ca_s
                    mn1_s=sub_s['Marge_N1'].sum(); cn1_s=sub_s['CA_N1'].sum(); tn1_s=mn1_s/cn1_s if cn1_s>0 else 0
                    sc_s=sub_s['Score_Sante'].mean() if 'Score_Sante' in sub_s.columns else 50
                    n_act_s=(sub_s['Statut']=='Action').sum()
                    rows_s.append({'Magasin':site,'CA':f"{ca_s:,.0f}",'Taux':fp(tx_s,False),
                                   'vs N-1':fp(tx_s-tn1_s),'Marge Δ':fk(sub_s['Dev_N1_FCFA'].sum()),
                                   'Score santé':f"{sc_s:.0f}/100",'Alertes':n_act_s})
                df_s=pd.DataFrame(rows_s).sort_values('Score santé')
                st.dataframe(df_s.style.map(cd,subset=['vs N-1']),use_container_width=True,hide_index=True)
            else:
                st.markdown(f"<div class='section-label'>{site_sel} — familles par priorité</div>", unsafe_allow_html=True)
                dm=df_mag.sort_values(['_ord_statut','Impact_Score'],ascending=[True,False])
                d3=dm[['SF_court','Score_Sante','%Marge','Dev_N1_pts','Dev_N1_FCFA','Cause','Statut','Action']].copy()
                d3.columns=['Famille','Santé','Taux','Dév. N-1','Marge Δ','Cause','Statut','Action']
                d3['Taux']=d3['Taux'].apply(lambda x: fp(x,False))
                d3['Dév. N-1']=d3['Dév. N-1'].apply(fp)
                d3['Marge Δ']=d3['Marge Δ'].apply(fk)
                d3['Santé']=d3['Santé'].apply(lambda x: f"{x:.0f}/100" if pd.notna(x) else '—')
                st.dataframe(d3.style.map(cs,subset=['Statut']).map(cd,subset=['Dév. N-1']),
                             use_container_width=True,hide_index=True,height=520,
                             column_config={'Action':st.column_config.TextColumn('Action',width='large')})

    # ── Vue promo ─────────────────────────────────────────────────────────────
    with sous_tab3:
        has_promo = 'CA Promo' in df.columns and 'Marge Promo' in df.columns
        if not has_promo:
            st.markdown("<div class='alert-card alert-blue'>Les colonnes CA Promo et Marge Promo ne sont pas dans cette extraction.</div>", unsafe_allow_html=True)
        else:
            df_promo = df[df['CA Promo'].notna() & (df['CA Promo']>0)].copy()
            df_promo['Tx_Promo']  = df_promo['Marge Promo']/df_promo['CA Promo']
            df_promo['Poids_Promo']= df_promo['CA Promo']/df_promo['CA']
            df_promo['Tx_HP']     = df_promo['Marge Hors Promo'].fillna(0)/df_promo['CA Hors Promo'].replace(0,1)
            df_promo['Effet_Promo']= df_promo['Tx_Promo'] - df_promo['Tx_HP']
            df_promo['Statut_Promo'] = df_promo['Tx_Promo'].apply(
                lambda x: 'Rentable' if x>=0.15 else ('Limite' if x>=0.05 else 'Deficitaire'))

            # KPIs promo
            ca_promo_tot = df_promo['CA Promo'].sum()
            ca_tot_p     = df['CA'].sum()
            poids_promo  = ca_promo_tot/ca_tot_p if ca_tot_p>0 else 0
            tx_promo_g   = df_promo['Marge Promo'].sum()/df_promo['CA Promo'].sum() if df_promo['CA Promo'].sum()>0 else 0
            tx_hp_g      = df_promo['Marge Hors Promo'].sum()/df_promo['CA Hors Promo'].sum() if df_promo['CA Hors Promo'].sum()>0 else 0
            n_deficit    = (df_promo['Statut_Promo']=='Deficitaire').sum()

            pk1,pk2,pk3,pk4 = st.columns(4)
            pk1.metric("Poids promo / CA", fp(poids_promo,False))
            pk2.metric("Tx marge promo",   fp(tx_promo_g,False), fp(tx_promo_g-tx_hp_g))
            pk3.metric("Tx marge hors promo", fp(tx_hp_g,False))
            pk4.metric("Promos déficitaires",  str(n_deficit), f"/ {len(df_promo)} familles en promo")

            st.markdown("<div class='section-label'>Toutes les familles en promo — triées par effet marge</div>", unsafe_allow_html=True)
            dp=df_promo[['Rayon_court','SF_court','CA Promo','Tx_HP','Tx_Promo','Effet_Promo','Poids_Promo','Statut_Promo']].copy()
            dp.columns=['Rayon','Famille','CA Promo','Tx HP','Tx Promo','Effet promo','% CA en promo','Statut promo']
            dp['CA Promo']=dp['CA Promo'].apply(lambda x: f"{x:,.0f}")
            for c in ['Tx HP','Tx Promo','Effet promo','% CA en promo']: dp[c]=dp[c].apply(lambda x: fp(x,False if c!='Effet promo' else True))
            dp=dp.sort_values('Effet promo')

            def cs_promo(v):
                if 'Deficitaire' in str(v): return 'background:#FEE2E2;color:#991B1B;font-weight:600'
                if 'Limite'      in str(v): return 'background:#FEF9C3;color:#854D0E;font-weight:600'
                return 'background:#D5F5E3;color:#145A32;font-weight:600'

            st.dataframe(dp.style.map(cs_promo,subset=['Statut promo']).map(cd,subset=['Effet promo']),
                         use_container_width=True,hide_index=True)
            st.caption("Effet promo = Tx marge promo − Tx marge hors promo. Négatif = la promo dégrade la marge vs fond de rayon.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("<div class='page-title' style='font-size:20px'>Export Excel</div>", unsafe_allow_html=True)
    st.markdown(f"""
<div class='alert-card alert-blue'>
  <strong>Contenu du fichier :</strong><br>
  Un onglet par période chargée · Colonnes : Famille · Magasin · Taux brut · Tx net casse · Marge perdue · Score santé · Cause · Action<br>
  Onglet <strong>Plan de Négociation</strong> — familles en action avec remise nécessaire en FCFA<br>
  Onglet <strong>Synthèse Magasins</strong> — si extraction multi-sites disponible<br>
  Onglet <strong>Guide de lecture</strong> — définitions et exemples pour diffusion équipe
</div>""", unsafe_allow_html=True)

    # Dégradations persistantes si multi-périodes
    if len(periodes_dispo) > 1:
        rouge_sets = {p: set(df_all[df_all['Periode']==p][df_all[df_all['Periode']==p]['Statut']=='Action']
                     .apply(lambda r: f"{r['Rayon_court']}|{r['SF_court']}", axis=1))
                     for p in periodes_dispo}
        persistants = set.intersection(*rouge_sets.values()) if rouge_sets else set()
        if persistants:
            st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>{len(persistants)} famille(s)</strong> en alerte sur <strong>toutes</strong> les {len(periodes_dispo)} périodes chargées — problèmes structurels à inscrire à l'ordre du jour fournisseur.
</div>""", unsafe_allow_html=True)
            rows_p=[]
            for key in sorted(persistants):
                rayon,sf=key.split('|',1)
                sub=df[(df['Rayon_court']==rayon)&(df['SF_court']==sf)]
                if len(sub):
                    r0=sub.iloc[0]
                    rp={'Rayon':rayon,'Famille':sf,'Acheteur':r0.get('Acheteur','—')}
                    if 'Site nom long' in r0.index and pd.notna(r0.get('Site nom long')):
                        rp['Magasin']=r0['Site nom long']
                    rp.update({'Taux':fp(r0.get('%Marge'),False),
                               'Score santé':f"{r0.get('Score_Sante',0):.0f}/100",
                               'Action':r0.get('Action','—')})
                    rows_p.append(rp)
            st.dataframe(pd.DataFrame(rows_p).sort_values('Score santé'),
                         use_container_width=True,hide_index=True,
                         column_config={'Action':st.column_config.TextColumn('Action',width='large')})

    st.markdown("")
    if st.button("Générer le fichier Excel", type="primary", key="gen_excel"):
        with st.spinner("Génération du rapport…"):
            buf = export_excel(df_all, periodes_dispo)
        st.download_button(
            label="Télécharger le rapport Excel",
            data=buf,
            file_name=f"SmartBuyer_Rentabilite_{periode_sel.replace('/','').replace('→','_').replace(' ','')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
