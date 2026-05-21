"""
SmartBuyer Hub - Module Fidelite Cagnotte
=====================================
Suivi hebdomadaire des cagnottes article x magasin a partir d'extractions Power BI.

Fonctionnalites:
- Multi-upload de fichiers PBI (.xlsx)
- 1 fichier CSV liste articles (Article, Cagnotte, Mois)
- Detection auto des dates / semaine / mois dans le bloc "Filtres appliques"
- Filtrage par mois propre a chaque fichier
- Empilage multi-semaines
- Export Excel 2 onglets : Recap detaille + Recap Financier
- Calcul des poids %CA et %Marge a la FAMILLE (Epicerie, Boissons, DPH)
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CHARTE SMARTBUYER HUB
# ============================================================
COULEUR_BLEU = "007AFF"
COULEUR_GRIS_FOND = "F2F2F7"
COULEUR_JAUNE_TOTAL = "FFF9E6"
COULEUR_BLEU_CLAIR = "E6F2FF"
COULEUR_TEXTE = "1C1C1E"
COULEUR_ROUGE = "D70015"
COULEUR_BORDURE = "E5E5EA"
POLICE = "Calibri"

MOIS_FR = {
    1: "Janvier", 2: "Fevrier", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Aout",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Decembre",
}

# Normalisation pour comparaison avec la colonne Mois du CSV (qui peut avoir accents)
MOIS_NORMALIZE = {
    "janvier": "Janvier", "fevrier": "Fevrier",
    "mars": "Mars", "avril": "Avril", "mai": "Mai", "juin": "Juin",
    "juillet": "Juillet", "aout": "Aout",
    "septembre": "Septembre", "octobre": "Octobre", "novembre": "Novembre",
    "decembre": "Decembre",
}


def normaliser_mois(s: str) -> str:
    """Normalise un libelle mois (gere accents et casse)."""
    if not isinstance(s, str):
        return ""
    s_clean = (s.strip().lower()
               .replace("é", "e").replace("è", "e").replace("ê", "e")
               .replace("û", "u").replace("à", "a").replace("ç", "c"))
    return MOIS_NORMALIZE.get(s_clean, s.strip().capitalize())


def normaliser_famille(s: str) -> str:
    """Normalise un libelle famille (Epicerie / Boissons / DPH) pour comparaison."""
    if not isinstance(s, str):
        return ""
    return (s.strip()
            .replace("é", "e").replace("è", "e").replace("ê", "e")
            .replace("û", "u").replace("à", "a").replace("ç", "c")
            .upper())


# ============================================================
# PARSING PBI
# ============================================================
def extraire_periode_pbi(ws) -> dict:
    """
    Cherche dans toutes les cellules du fichier le bloc 'Filtres appliques'
    et extrait Date Debut / Date Fin.
    Pattern attendu: "Date est le ou apres le DD/MM/YYYY et est avant le DD/MM/YYYY"
    """
    pattern = re.compile(
        r"apr[eè]s\s+le\s+(\d{2}/\d{2}/\d{4})\s+et\s+est\s+avant\s+le\s+(\d{2}/\d{2}/\d{4})",
        re.IGNORECASE,
    )
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Filtres appliqu" in cell.value:
                m = pattern.search(cell.value)
                if m:
                    d_deb = datetime.strptime(m.group(1), "%d/%m/%Y").date()
                    d_fin_exclusive = datetime.strptime(m.group(2), "%d/%m/%Y").date()
                    # Le filtre PBI est "avant le" donc la vraie fin est j-1
                    d_fin = pd.Timestamp(d_fin_exclusive) - pd.Timedelta(days=1)
                    d_fin = d_fin.date()
                    return {
                        "date_debut": d_deb,
                        "date_fin": d_fin,
                        "semaine": f"S{d_fin.isocalendar().week:02d}",
                        "mois": f"{MOIS_FR[d_fin.month]} {d_fin.year}",
                        "mois_court": MOIS_FR[d_fin.month],
                    }
    return None


def _get(row, idx):
    """Acces securise a une cellule du tuple : None si l'index depasse."""
    return row[idx] if idx < len(row) else None


def parser_pbi(fichier) -> dict:
    """
    Parse un fichier PBI et retourne :
    - periode : dict avec dates / semaine / mois
    - lignes : DataFrame avec les lignes article x magasin (incluant Famille)
    - totaux_pgc : dict {CA, Marge, Qte} global PGC
    - totaux_famille : dict {famille_normalisee: {CA, Marge, Qte}}

    Structure PBI :
    - Colonne A : Famille (Epicerie, Boissons, DPH, ...) -- niveau le plus haut
    - Colonne B : Rayon
    - Colonne C : Sous Famille
    - Colonne D : Article (code + libelle)
    - Colonne E : Site nom long
    - Colonne F : CA
    - Colonne I : Marge
    - Colonne AC (index 28) : Qte Vente

    Logique de hierarchie :
    - Ligne famille : colonne A = nom famille, B='Total' (sous-total famille)
    - Ligne rayon : colonne B = nom rayon, C='Total' (sous-total rayon)
    - Ligne sous-famille : colonne C = nom sous-famille, D vide
    - Ligne article : colonne D = "code - libelle", E='Total'
    - Ligne magasin : colonne D vide (heritage), E = "code - nom magasin"
    - Ligne grand total : ligne au tout debut avec A='Total' (toutes colonnes vides
      sauf F, I, AC) -> totaux PGC globaux
    """
    wb = load_workbook(fichier, data_only=True)
    ws = wb.active

    periode = extraire_periode_pbi(ws)
    if periode is None:
        return {
            "periode": None, "lignes": pd.DataFrame(),
            "totaux_pgc": None, "totaux_famille": {},
            "erreur": "Bloc 'Filtres appliques' introuvable"
        }

    lignes = []
    totaux_pgc = None
    totaux_famille = {}  # {nom_famille_normalise: {CA, Marge, Qte, libelle_original}}

    famille_courante = None
    rayon_courant = None
    sous_famille_courante = None
    article_code_courant = None
    article_libelle_courant = None

    pattern_article = re.compile(r"^(\d{6,9})\s*[-–]?\s*(.*)$")
    pattern_site = re.compile(r"^(\d{4,6})\s*[-–]\s*(.+)$")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        famille = _get(row, 0)
        rayon = _get(row, 1)
        sous_famille = _get(row, 2)
        article = _get(row, 3)
        site = _get(row, 4)
        ca = _get(row, 5)
        marge = _get(row, 8)
        qte = _get(row, 28)

        # --- Detection ligne grand TOTAL PGC ---
        # Famille = "Total" sur la premiere ligne agregee
        if totaux_pgc is None and isinstance(famille, str) and famille.strip() == "Total":
            totaux_pgc = {
                "CA": ca if isinstance(ca, (int, float)) else 0,
                "Marge": marge if isinstance(marge, (int, float)) else 0,
                "Qte": qte if isinstance(qte, (int, float)) else 0,
            }
            continue

        # --- Detection ligne sous-total FAMILLE ---
        # Cas typique : Famille = nom (Epicerie/Boissons/DPH), Rayon='Total', le reste vide
        # On capture les valeurs CA/Marge/Qte de cette ligne
        if (isinstance(famille, str) and famille.strip()
                and famille.strip() != "Total"
                and isinstance(rayon, str) and rayon.strip() == "Total"
                and (sous_famille is None or sous_famille == "" or sous_famille == "Total")
                and (article is None or article == "" or article == "Total")):
            fam_norm = normaliser_famille(famille)
            if fam_norm and fam_norm not in totaux_famille:
                totaux_famille[fam_norm] = {
                    "CA": ca if isinstance(ca, (int, float)) else 0,
                    "Marge": marge if isinstance(marge, (int, float)) else 0,
                    "Qte": qte if isinstance(qte, (int, float)) else 0,
                    "libelle": famille.strip(),
                }
            # On met aussi a jour le contexte famille courant
            famille_courante = famille.strip()
            continue

        # --- Maj contexte Famille (sans ligne Total) ---
        if (isinstance(famille, str) and famille.strip()
                and famille.strip() != "Total"):
            famille_courante = famille.strip()

        # --- Maj contexte Rayon ---
        if isinstance(rayon, str) and rayon.strip() and rayon.strip() != "Total":
            rayon_courant = rayon.strip()

        # --- Maj contexte Sous Famille ---
        if (isinstance(sous_famille, str) and sous_famille.strip()
                and sous_famille.strip() != "Total"):
            sous_famille_courante = sous_famille.strip()

        # --- Detection ligne article (D rempli, E = Total ou vide) ---
        if isinstance(article, str) and article.strip() and article.strip() != "Total":
            m = pattern_article.match(article.strip())
            if m and (site == "Total" or site is None or site == ""):
                article_code_courant = m.group(1)
                article_libelle_courant = article.strip()
                continue

        # --- Detection ligne magasin (E = code-nom) ---
        if isinstance(site, str) and article_code_courant:
            m_site = pattern_site.match(site.strip())
            if m_site:
                code_site = m_site.group(1)
                nom_magasin = m_site.group(2).strip()
                lignes.append({
                    "Famille": famille_courante or "",
                    "Famille_norm": normaliser_famille(famille_courante or ""),
                    "Rayon": rayon_courant or "",
                    "Sous Famille": sous_famille_courante or "",
                    "Code Article": article_code_courant,
                    "Article": article_libelle_courant,
                    "Code Site": code_site,
                    "Magasin": nom_magasin,
                    "CA": ca if isinstance(ca, (int, float)) else 0,
                    "Marge": marge if isinstance(marge, (int, float)) else 0,
                    "Qte": qte if isinstance(qte, (int, float)) else 0,
                })

    df = pd.DataFrame(lignes)
    return {
        "periode": periode,
        "lignes": df,
        "totaux_pgc": totaux_pgc,
        "totaux_famille": totaux_famille,
        "erreur": None,
    }


# ============================================================
# CONSTRUCTION DU RECAP
# ============================================================
def lire_liste_csv(fichier) -> pd.DataFrame:
    """Lit le CSV liste articles (separateur ; UTF-8)."""
    try:
        df = pd.read_csv(fichier, sep=";", dtype={"Article": str, "Cagnotte": float, "Mois": str})
    except Exception:
        fichier.seek(0)
        df = pd.read_csv(fichier, sep=",", dtype={"Article": str, "Cagnotte": float, "Mois": str})
    df["Article"] = df["Article"].astype(str).str.strip()
    df["Mois_norm"] = df["Mois"].apply(normaliser_mois)
    return df


def construire_recap_fichier(parsing: dict, liste_df: pd.DataFrame) -> dict:
    """
    Applique le filtre articles de la liste sur le mois du PBI,
    et construit les lignes du recap article x magasin.
    """
    periode = parsing["periode"]
    df_pbi = parsing["lignes"]
    totaux_pgc = parsing.get("totaux_pgc")
    totaux_famille = parsing.get("totaux_famille", {})

    if periode is None or df_pbi.empty:
        return {
            "lignes_recap": pd.DataFrame(),
            "articles_attendus": [], "articles_trouves": [],
            "totaux_pgc": totaux_pgc,
            "totaux_famille": totaux_famille,
            "semaine": periode["semaine"] if periode else None,
            "mois": periode["mois"] if periode else None,
        }

    mois_court = periode["mois_court"]

    # Filtre liste articles sur le mois
    articles_mois = liste_df[liste_df["Mois_norm"] == mois_court].copy()
    articles_codes = articles_mois["Article"].tolist()

    articles_pbi = df_pbi["Code Article"].unique().tolist()
    articles_trouves = [a for a in articles_codes if a in articles_pbi]

    df_filtre = df_pbi[df_pbi["Code Article"].isin(articles_codes)].copy()

    if df_filtre.empty:
        return {
            "lignes_recap": pd.DataFrame(),
            "articles_attendus": articles_codes,
            "articles_trouves": [],
            "totaux_pgc": totaux_pgc,
            "totaux_famille": totaux_famille,
            "semaine": periode["semaine"],
            "mois": periode["mois"],
        }

    # Cagnotte unitaire et budget
    cag_map = dict(zip(articles_mois["Article"], articles_mois["Cagnotte"]))
    df_filtre["Cagnotte"] = df_filtre["Code Article"].map(cag_map)
    df_filtre["Budget Cagnotte"] = df_filtre["Cagnotte"] * df_filtre["Qte"]

    # Colonnes temporelles
    df_filtre["Date Debut"] = periode["date_debut"]
    df_filtre["Date Fin"] = periode["date_fin"]
    df_filtre["Semaine"] = periode["semaine"]
    df_filtre["Mois"] = periode["mois"]

    df_filtre = df_filtre[[
        "Date Debut", "Date Fin", "Semaine", "Mois",
        "Famille", "Famille_norm", "Rayon", "Sous Famille",
        "Code Article", "Article",
        "Code Site", "Magasin",
        "CA", "Marge", "Qte", "Cagnotte", "Budget Cagnotte",
    ]]

    return {
        "lignes_recap": df_filtre,
        "articles_attendus": articles_codes,
        "articles_trouves": articles_trouves,
        "totaux_pgc": totaux_pgc,
        "totaux_famille": totaux_famille,
        "semaine": periode["semaine"],
        "mois": periode["mois"],
    }


# ============================================================
# CALCUL RECAP FINANCIER (poids a la FAMILLE)
# ============================================================
def construire_recap_financier(df_recap_all: pd.DataFrame,
                                totaux_famille_par_semaine: dict) -> pd.DataFrame:
    """
    Agrege le recap detaille par Semaine x Article.
    totaux_famille_par_semaine : dict {semaine: {famille_norm: {CA, Marge, Qte, libelle}}}
    Ajoute les colonnes %CA et %Marge calcules a la FAMILLE de l'article
    (denominateur = total CA / Marge de la famille pour la semaine concernee).
    """
    cols = [
        "Semaine", "Mois", "Famille", "Code Article", "Article",
        "Nb magasins", "Cagnotte/u", "Qte", "Budget cagnotte",
        "CA", "Marge", "CA Famille", "Marge Famille", "%CA", "%Marge",
    ]
    if df_recap_all.empty:
        return pd.DataFrame(columns=cols)

    grouped = df_recap_all.groupby(
        ["Semaine", "Mois", "Famille", "Famille_norm", "Code Article", "Article"],
        as_index=False
    ).agg(
        Nb_magasins=("Code Site", "nunique"),
        Cagnotte_mag=("Cagnotte", "first"),
        CA=("CA", "sum"),
        Marge=("Marge", "sum"),
        Qte=("Qte", "sum"),
        Budget_cagnotte=("Budget Cagnotte", "sum"),
    )
    grouped = grouped.rename(columns={
        "Nb_magasins": "Nb magasins",
        "Cagnotte_mag": "Cagnotte/u",
        "Budget_cagnotte": "Budget cagnotte",
    })

    # Recuperation du total famille pour chaque ligne (semaine + famille_norm)
    def _ca_famille(row):
        sem_tot = totaux_famille_par_semaine.get(row["Semaine"], {})
        fam_tot = sem_tot.get(row["Famille_norm"], {})
        return fam_tot.get("CA", 0)

    def _marge_famille(row):
        sem_tot = totaux_famille_par_semaine.get(row["Semaine"], {})
        fam_tot = sem_tot.get(row["Famille_norm"], {})
        return fam_tot.get("Marge", 0)

    grouped["CA Famille"] = grouped.apply(_ca_famille, axis=1)
    grouped["Marge Famille"] = grouped.apply(_marge_famille, axis=1)

    # Poids a la famille (gestion division par zero)
    grouped["%CA"] = grouped.apply(
        lambda r: (r["CA"] / r["CA Famille"]) if r["CA Famille"] else 0, axis=1
    )
    grouped["%Marge"] = grouped.apply(
        lambda r: (r["Marge"] / r["Marge Famille"]) if r["Marge Famille"] else 0, axis=1
    )

    grouped = grouped[cols]
    return grouped


# ============================================================
# EXPORT EXCEL
# ============================================================
def style_header(cell, couleur_fond=None):
    cell.font = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=couleur_fond or COULEUR_BLEU)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=COULEUR_BORDURE),
        right=Side(style="thin", color=COULEUR_BORDURE),
        top=Side(style="thin", color=COULEUR_BORDURE),
        bottom=Side(style="thin", color=COULEUR_BORDURE),
    )


def style_cell(cell, fond=None, bold=False, fmt=None, align="left"):
    cell.font = Font(name=POLICE, size=10, bold=bold, color=COULEUR_TEXTE)
    if fond:
        cell.fill = PatternFill("solid", fgColor=fond)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    cell.border = Border(
        left=Side(style="thin", color=COULEUR_BORDURE),
        right=Side(style="thin", color=COULEUR_BORDURE),
        top=Side(style="thin", color=COULEUR_BORDURE),
        bottom=Side(style="thin", color=COULEUR_BORDURE),
    )


def exporter_excel(df_recap: pd.DataFrame, df_financier: pd.DataFrame,
                   kpis: dict, totaux_famille_cumul: dict) -> BytesIO:
    """Construit le fichier Excel 2 onglets : Recap + Recap Financier."""
    wb = Workbook()

    # ===== ONGLET 1 : RECAP =====
    ws = wb.active
    ws.title = "Recap"

    headers = [
        "Date Debut", "Date Fin", "Semaine", "Mois",
        "Famille", "Rayon", "Sous Famille",
        "Code Article", "Article",
        "Code Site", "Magasin",
        "CA", "Marge", "Qte", "Cagnotte", "Budget Cagnotte",
    ]

    widths = [12, 12, 9, 14, 14, 16, 18, 13, 36, 10, 24, 13, 13, 10, 11, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header(cell)
    ws.row_dimensions[1].height = 28

    if not df_recap.empty:
        df_sorted = df_recap.sort_values(
            by=["Famille", "Code Article", "Magasin", "Semaine"]
        ).reset_index(drop=True)

        row_excel = 2
        for _, row in df_sorted.iterrows():
            values = [
                row["Date Debut"], row["Date Fin"], row["Semaine"], row["Mois"],
                row["Famille"], row["Rayon"], row["Sous Famille"],
                row["Code Article"], row["Article"],
                row["Code Site"], row["Magasin"],
                row["CA"], row["Marge"], row["Qte"], row["Cagnotte"], row["Budget Cagnotte"],
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_excel, column=col_idx, value=val)
                if col_idx in (1, 2):
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, fmt="DD/MM/YYYY", align="center")
                elif col_idx == 3:
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center")
                elif col_idx == 4:
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center")
                elif col_idx == 5:  # Famille
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center", bold=True)
                elif col_idx in (12, 13):  # CA, Marge
                    style_cell(cell, fmt="#,##0;[Red]-#,##0", align="right")
                elif col_idx == 14:  # Qte
                    style_cell(cell, fmt="#,##0.0", align="right")
                elif col_idx == 15:  # Cagnotte unitaire
                    style_cell(cell, fmt="#,##0", align="right")
                elif col_idx == 16:  # Budget Cagnotte
                    style_cell(cell, fond=COULEUR_JAUNE_TOTAL, fmt="#,##0", align="right", bold=True)
                else:
                    style_cell(cell, align="left")
            row_excel += 1

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ===== ONGLET 2 : RECAP FINANCIER =====
    ws2 = wb.create_sheet("Recap Financier")

    # Largeurs
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 2
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["E"].width = 18

    # Titre
    ws2["A1"] = "RECAP FINANCIER · FIDELITE CAGNOTTE"
    ws2["A1"].font = Font(name=POLICE, size=14, bold=True, color=COULEUR_BLEU)
    ws2.merge_cells("A1:E1")

    # --- Bloc KPIs globaux ---
    ws2["A3"] = "Indicateurs globaux"
    ws2["A3"].font = Font(name=POLICE, size=11, bold=True, color=COULEUR_TEXTE)

    kpi_items = [
        ("Budget cagnotte total", kpis.get("budget_total", 0), "#,##0"),
        ("CA articles fidelite", kpis.get("ca_total", 0), "#,##0"),
        ("Marge articles fidelite", kpis.get("marge_total", 0), "#,##0;[Red]-#,##0"),
        ("Qte articles fidelite", kpis.get("qte_total", 0), "#,##0.0"),
        ("CA total PGC (reference)", kpis.get("ca_pgc_total", 0), "#,##0"),
        ("Marge totale PGC (reference)", kpis.get("marge_pgc_total", 0), "#,##0;[Red]-#,##0"),
        ("% CA fidelite / PGC", kpis.get("pct_ca_pgc", 0), "0.00%"),
        ("% Marge fidelite / PGC", kpis.get("pct_marge_pgc", 0), "0.00%"),
    ]

    for i, (lbl, val, fmt) in enumerate(kpi_items):
        r = 4 + i
        ws2.cell(row=r, column=1, value=lbl).font = Font(name=POLICE, size=10, color=COULEUR_TEXTE)
        c = ws2.cell(row=r, column=2, value=val)
        if lbl.startswith("%"):
            c.font = Font(name=POLICE, size=11, bold=True, color=COULEUR_BLEU)
            c.fill = PatternFill("solid", fgColor=COULEUR_JAUNE_TOTAL)
        else:
            c.font = Font(name=POLICE, size=11, bold=True, color=COULEUR_BLEU)
            c.fill = PatternFill("solid", fgColor=COULEUR_GRIS_FOND)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")

    # --- Bloc Poids par famille ---
    row_fam_title = 14
    ws2.cell(row=row_fam_title, column=1,
             value="Poids fidelite par FAMILLE").font = Font(
                 name=POLICE, size=11, bold=True, color=COULEUR_TEXTE)

    headers_fam = ["Famille", "CA Fidelite", "CA Famille", "%CA", "Marge Fidelite", "Marge Famille", "%Marge"]
    widths_fam = [22, 16, 16, 10, 16, 16, 10]
    for i, w in enumerate(widths_fam, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row_fam_header = row_fam_title + 1
    for col_idx, h in enumerate(headers_fam, start=1):
        cell = ws2.cell(row=row_fam_header, column=col_idx, value=h)
        style_header(cell)
    ws2.row_dimensions[row_fam_header].height = 28

    # Agregation par famille depuis df_financier
    row_fam = row_fam_header + 1
    if not df_financier.empty:
        # On somme les CA/Marge des articles par famille, et on prend CA Famille / Marge Famille
        # (max car identique pour toutes les lignes d'une meme famille et semaine, mais on cumule sur les semaines)
        # Pour le cumul multi-semaines, on additionne les totaux famille de chaque semaine
        fam_articles = df_financier.groupby("Famille", as_index=False).agg(
            CA_fid=("CA", "sum"),
            Marge_fid=("Marge", "sum"),
        )
        # Cumul des totaux famille sur toutes les semaines
        fam_totaux = {}  # {libelle_famille: {CA, Marge}}
        for sem, dico_fam in totaux_famille_cumul.items():
            for fam_norm, tot in dico_fam.items():
                lib = tot.get("libelle", fam_norm)
                if lib not in fam_totaux:
                    fam_totaux[lib] = {"CA": 0, "Marge": 0}
                fam_totaux[lib]["CA"] += tot.get("CA", 0)
                fam_totaux[lib]["Marge"] += tot.get("Marge", 0)

        for _, r in fam_articles.iterrows():
            fam = r["Famille"]
            ca_fid = r["CA_fid"]
            marge_fid = r["Marge_fid"]
            ca_fam = fam_totaux.get(fam, {}).get("CA", 0)
            marge_fam = fam_totaux.get(fam, {}).get("Marge", 0)
            pct_ca = (ca_fid / ca_fam) if ca_fam else 0
            pct_marge = (marge_fid / marge_fam) if marge_fam else 0

            ws2.cell(row=row_fam, column=1, value=fam)
            ws2.cell(row=row_fam, column=2, value=ca_fid)
            ws2.cell(row=row_fam, column=3, value=ca_fam)
            ws2.cell(row=row_fam, column=4, value=pct_ca)
            ws2.cell(row=row_fam, column=5, value=marge_fid)
            ws2.cell(row=row_fam, column=6, value=marge_fam)
            ws2.cell(row=row_fam, column=7, value=pct_marge)

            for col_idx in range(1, 8):
                cell = ws2.cell(row=row_fam, column=col_idx)
                if col_idx == 1:
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, bold=True, align="left")
                elif col_idx in (4, 7):
                    style_cell(cell, fond=COULEUR_JAUNE_TOTAL, fmt="0.00%", align="right", bold=True)
                else:
                    style_cell(cell, fmt="#,##0;[Red]-#,##0", align="right")
            row_fam += 1

    # --- Tableau par article x semaine ---
    row_titre = row_fam + 2
    ws2.cell(row=row_titre, column=1,
             value="Detail par article et semaine").font = Font(
                 name=POLICE, size=11, bold=True, color=COULEUR_TEXTE)

    headers_fin = [
        "Semaine", "Mois", "Famille", "Code Article", "Article",
        "Nb magasins", "Cagnotte/u", "Qte", "Budget cagnotte",
        "CA", "Marge", "CA Famille", "Marge Famille", "%CA", "%Marge",
    ]
    widths_fin = [10, 14, 14, 13, 36, 12, 12, 11, 16, 14, 14, 15, 15, 9, 9]
    for i, w in enumerate(widths_fin, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = max(
            ws2.column_dimensions[get_column_letter(i)].width or 0, w
        )

    row_header = row_titre + 1
    for col_idx, h in enumerate(headers_fin, start=1):
        cell = ws2.cell(row=row_header, column=col_idx, value=h)
        style_header(cell)
    ws2.row_dimensions[row_header].height = 28

    row_data = row_header + 1
    if not df_financier.empty:
        df_fin_sorted = df_financier.sort_values(
            by=["Semaine", "Famille", "Code Article"]
        ).reset_index(drop=True)
        for _, r in df_fin_sorted.iterrows():
            values = [
                r["Semaine"], r["Mois"], r["Famille"], r["Code Article"], r["Article"],
                r["Nb magasins"], r["Cagnotte/u"], r["Qte"], r["Budget cagnotte"],
                r["CA"], r["Marge"], r["CA Famille"], r["Marge Famille"], r["%CA"], r["%Marge"],
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws2.cell(row=row_data, column=col_idx, value=val)
                if col_idx == 1:
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center")
                elif col_idx == 2:
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center")
                elif col_idx == 3:  # Famille
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, align="center", bold=True)
                elif col_idx == 6:  # Nb magasins
                    style_cell(cell, fmt="#,##0", align="center")
                elif col_idx == 7:  # Cagnotte/u
                    style_cell(cell, fmt="#,##0", align="right")
                elif col_idx == 8:  # Qte
                    style_cell(cell, fmt="#,##0.0", align="right")
                elif col_idx == 9:  # Budget cagnotte
                    style_cell(cell, fond=COULEUR_JAUNE_TOTAL, fmt="#,##0", align="right", bold=True)
                elif col_idx in (10, 11):  # CA, Marge article
                    style_cell(cell, fmt="#,##0;[Red]-#,##0", align="right")
                elif col_idx in (12, 13):  # CA Famille, Marge Famille
                    style_cell(cell, fond=COULEUR_GRIS_FOND, fmt="#,##0;[Red]-#,##0", align="right")
                elif col_idx in (14, 15):  # %CA, %Marge a la famille
                    style_cell(cell, fond=COULEUR_BLEU_CLAIR, fmt="0.00%", align="right", bold=True)
                else:
                    style_cell(cell, align="left")
            row_data += 1

        # Ligne TOTAL
        row_total = row_data
        ws2.cell(row=row_total, column=1, value="TOTAL")
        ws2.cell(row=row_total, column=5, value=f"{len(df_fin_sorted)} ligne(s)")
        # Sommes
        ws2.cell(row=row_total, column=6, value=f"=SUM(F{row_header+1}:F{row_data-1})")
        ws2.cell(row=row_total, column=8, value=f"=SUM(H{row_header+1}:H{row_data-1})")
        ws2.cell(row=row_total, column=9, value=f"=SUM(I{row_header+1}:I{row_data-1})")
        ws2.cell(row=row_total, column=10, value=f"=SUM(J{row_header+1}:J{row_data-1})")
        ws2.cell(row=row_total, column=11, value=f"=SUM(K{row_header+1}:K{row_data-1})")
        # %CA et %Marge total = poids global articles fidelite vs PGC total
        ws2.cell(row=row_total, column=14, value=kpis.get("pct_ca_pgc", 0))
        ws2.cell(row=row_total, column=15, value=kpis.get("pct_marge_pgc", 0))

        for col_idx in range(1, 16):
            cell = ws2.cell(row=row_total, column=col_idx)
            if col_idx == 1:
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, align="center")
            elif col_idx == 6:
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, fmt="#,##0", align="center")
            elif col_idx == 8:
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, fmt="#,##0.0", align="right")
            elif col_idx == 9:
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, fmt="#,##0", align="right")
            elif col_idx in (10, 11):
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, fmt="#,##0;[Red]-#,##0", align="right")
            elif col_idx in (14, 15):
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, fmt="0.00%", align="right")
            else:
                style_cell(cell, fond=COULEUR_JAUNE_TOTAL, bold=True, align="left")

    ws2.freeze_panes = f"A{row_header+1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# INTERFACE STREAMLIT
# ============================================================
def render_fidelite_cagnotte():
    """Module a appeler depuis SmartBuyer Hub."""
    st.markdown("""
    <style>
        .stApp { font-family: -apple-system, 'SF Pro Display', Calibri, sans-serif; }
        .header-cagnotte {
            display: flex; align-items: center; gap: 12px;
            padding: 16px 0; margin-bottom: 20px;
            border-bottom: 1px solid #E5E5EA;
        }
        .header-cagnotte .icone {
            width: 40px; height: 40px; border-radius: 10px;
            background: #007AFF; display: flex; align-items: center;
            justify-content: center; color: white; font-size: 20px;
        }
        .header-cagnotte .titre { font-size: 20px; font-weight: 600; color: #1C1C1E; margin: 0; }
        .header-cagnotte .sous-titre { font-size: 13px; color: #8E8E93; margin: 0; }
        .bandeau-periode {
            background: #E6F2FF; border-radius: 10px;
            padding: 12px 16px; margin: 16px 0;
        }
        .bandeau-periode .lbl { font-size: 11px; color: #0040A0; letter-spacing: 0.5px; }
        .bandeau-periode .val { font-size: 14px; font-weight: 600; color: #007AFF; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="header-cagnotte">
        <div class="icone">$</div>
        <div>
            <p class="titre">Fidelite Cagnotte</p>
            <p class="sous-titre">Suivi hebdomadaire · Investissement vs Performance (poids a la famille)</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # === ETAPE 1 : Upload ===
    st.markdown("**1 · Charger les fichiers**")
    col1, col2 = st.columns(2)
    with col1:
        fichiers_pbi = st.file_uploader(
            "Extractions PBI (xlsx, plusieurs possibles)",
            type=["xlsx"],
            accept_multiple_files=True,
            key="fidelite_cagnotte_files",
        )
    with col2:
        fichier_liste = st.file_uploader(
            "Liste articles (csv)",
            type=["csv"],
            key="fidelite_cagnotte_liste",
        )

    if not fichiers_pbi or not fichier_liste:
        st.info("Charge au moins un fichier PBI et le CSV liste pour demarrer.")
        return

    # === ETAPE 2 : Parsing ===
    liste_df = lire_liste_csv(fichier_liste)

    parsings = []
    semaines_vues = {}

    for f in fichiers_pbi:
        p = parser_pbi(f)
        if p["periode"] is None:
            st.error(f"`{f.name}` : impossible de detecter la periode. {p.get('erreur', '')}")
            continue
        sem = p["periode"]["semaine"]
        if sem in semaines_vues:
            st.warning(f"Doublon detecte : la semaine **{sem}** est presente dans `{semaines_vues[sem]}` et `{f.name}`. Les deux seront empilees.")
        else:
            semaines_vues[sem] = f.name
        parsings.append({"fichier": f.name, "parsing": p})

    if not parsings:
        st.error("Aucun fichier PBI exploitable.")
        return

    # Affichage periodes detectees + familles trouvees
    st.markdown("**2 · Periodes detectees**")
    for item in parsings:
        per = item["parsing"]["periode"]
        fams = list(item["parsing"].get("totaux_famille", {}).keys())
        fams_str = ", ".join(fams) if fams else "aucune famille detectee"
        st.markdown(f"""
        <div style="background: #F2F2F7; border-radius: 8px; padding: 8px 12px; margin-bottom: 6px; font-size: 13px;">
            <strong>{item['fichier']}</strong> · {per['date_debut'].strftime('%d/%m/%Y')} → {per['date_fin'].strftime('%d/%m/%Y')} · {per['semaine']} · {per['mois']}
            <br><span style="color: #8E8E93; font-size: 12px;">Familles : {fams_str}</span>
        </div>
        """, unsafe_allow_html=True)

    # === ETAPE 3 : Construction recap ===
    recap_parts = []
    articles_manquants_global = {}
    totaux_pgc_par_semaine = {}
    totaux_famille_par_semaine = {}  # {semaine: {famille_norm: {CA, Marge, Qte, libelle}}}

    for item in parsings:
        res = construire_recap_fichier(item["parsing"], liste_df)
        if not res["lignes_recap"].empty:
            recap_parts.append(res["lignes_recap"])
        manquants = set(res["articles_attendus"]) - set(res["articles_trouves"])
        if manquants:
            sem = item["parsing"]["periode"]["semaine"]
            articles_manquants_global[f"{item['fichier']} ({sem})"] = list(manquants)

        sem = res.get("semaine")
        tot_pgc = res.get("totaux_pgc")
        tot_fam = res.get("totaux_famille", {})

        if sem and tot_pgc:
            if sem in totaux_pgc_par_semaine:
                # Cumul si meme semaine chargee 2x
                totaux_pgc_par_semaine[sem] = {
                    "CA": totaux_pgc_par_semaine[sem]["CA"] + tot_pgc.get("CA", 0),
                    "Marge": totaux_pgc_par_semaine[sem]["Marge"] + tot_pgc.get("Marge", 0),
                    "Qte": totaux_pgc_par_semaine[sem]["Qte"] + tot_pgc.get("Qte", 0),
                }
            else:
                totaux_pgc_par_semaine[sem] = tot_pgc

        if sem and tot_fam:
            if sem not in totaux_famille_par_semaine:
                totaux_famille_par_semaine[sem] = {}
            for fam_norm, vals in tot_fam.items():
                if fam_norm in totaux_famille_par_semaine[sem]:
                    totaux_famille_par_semaine[sem][fam_norm]["CA"] += vals.get("CA", 0)
                    totaux_famille_par_semaine[sem][fam_norm]["Marge"] += vals.get("Marge", 0)
                    totaux_famille_par_semaine[sem][fam_norm]["Qte"] += vals.get("Qte", 0)
                else:
                    totaux_famille_par_semaine[sem][fam_norm] = dict(vals)

    df_recap_all = pd.concat(recap_parts, ignore_index=True) if recap_parts else pd.DataFrame()

    if articles_manquants_global:
        with st.expander(f"⚠ Articles non trouves dans certaines extractions ({sum(len(v) for v in articles_manquants_global.values())})"):
            for fic, arts in articles_manquants_global.items():
                st.markdown(f"**{fic}** : {', '.join(arts)}")

    if df_recap_all.empty:
        st.warning("Aucun article de la liste n'a ete trouve dans les extractions PBI pour les mois concernes.")
        return

    # Verification que les familles sont bien detectees
    if not any(totaux_famille_par_semaine.values()):
        st.warning("⚠ Aucun total par famille detecte dans les extractions PBI. Les poids %CA/%Marge seront a 0. Verifie la structure du fichier (colonne A = Famille, ligne avec Rayon='Total' attendue).")

    # === ETAPE 4 : KPIs et recap financier ===
    df_financier = construire_recap_financier(df_recap_all, totaux_famille_par_semaine)

    ca_pgc_total = sum(v.get("CA", 0) for v in totaux_pgc_par_semaine.values())
    marge_pgc_total = sum(v.get("Marge", 0) for v in totaux_pgc_par_semaine.values())

    ca_fid = df_financier["CA"].sum()
    marge_fid = df_financier["Marge"].sum()

    kpis = {
        "budget_total": df_financier["Budget cagnotte"].sum(),
        "ca_total": ca_fid,
        "marge_total": marge_fid,
        "qte_total": df_financier["Qte"].sum(),
        "ca_pgc_total": ca_pgc_total,
        "marge_pgc_total": marge_pgc_total,
        "pct_ca_pgc": (ca_fid / ca_pgc_total) if ca_pgc_total else 0,
        "pct_marge_pgc": (marge_fid / marge_pgc_total) if marge_pgc_total else 0,
    }

    st.markdown("**3 · Synthese globale**")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Budget cagnotte", f"{kpis['budget_total']:,.0f}".replace(",", " "))
    c2.metric("CA articles fidelite", f"{kpis['ca_total']:,.0f}".replace(",", " "))
    c3.metric("Marge articles fidelite", f"{kpis['marge_total']:,.0f}".replace(",", " "))
    c4.metric("Qte vendue", f"{kpis['qte_total']:,.1f}".replace(",", " "))

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("CA total PGC", f"{kpis['ca_pgc_total']:,.0f}".replace(",", " "))
    c6.metric("Marge totale PGC", f"{kpis['marge_pgc_total']:,.0f}".replace(",", " "))
    c7.metric("% CA fid. / PGC", f"{kpis['pct_ca_pgc']*100:.2f}%")
    c8.metric("% Marge fid. / PGC", f"{kpis['pct_marge_pgc']*100:.2f}%")

    # === Tableau poids par famille (affichage Streamlit) ===
    st.markdown("**4 · Poids fidelite par FAMILLE**")
    # Cumul des totaux famille sur toutes les semaines
    fam_totaux_cumul = {}
    for sem, dico_fam in totaux_famille_par_semaine.items():
        for fam_norm, tot in dico_fam.items():
            lib = tot.get("libelle", fam_norm)
            if lib not in fam_totaux_cumul:
                fam_totaux_cumul[lib] = {"CA": 0, "Marge": 0}
            fam_totaux_cumul[lib]["CA"] += tot.get("CA", 0)
            fam_totaux_cumul[lib]["Marge"] += tot.get("Marge", 0)

    if not df_financier.empty and fam_totaux_cumul:
        fam_articles = df_financier.groupby("Famille", as_index=False).agg(
            CA_fid=("CA", "sum"),
            Marge_fid=("Marge", "sum"),
        )
        rows_fam = []
        for _, r in fam_articles.iterrows():
            fam = r["Famille"]
            ca_fam = fam_totaux_cumul.get(fam, {}).get("CA", 0)
            marge_fam = fam_totaux_cumul.get(fam, {}).get("Marge", 0)
            rows_fam.append({
                "Famille": fam,
                "CA Fidelite": r["CA_fid"],
                "CA Famille": ca_fam,
                "%CA": (r["CA_fid"] / ca_fam) if ca_fam else 0,
                "Marge Fidelite": r["Marge_fid"],
                "Marge Famille": marge_fam,
                "%Marge": (r["Marge_fid"] / marge_fam) if marge_fam else 0,
            })
        df_fam_display = pd.DataFrame(rows_fam)
        st.dataframe(
            df_fam_display.style.format({
                "CA Fidelite": "{:,.0f}",
                "CA Famille": "{:,.0f}",
                "%CA": "{:.2%}",
                "Marge Fidelite": "{:,.0f}",
                "Marge Famille": "{:,.0f}",
                "%Marge": "{:.2%}",
            }),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Pas de totaux famille disponibles.")

    # === ETAPE 5 : Apercu ===
    st.markdown("**5 · Apercu Recap Financier (poids a la famille)**")
    st.dataframe(df_financier, use_container_width=True, hide_index=True)

    with st.expander("Voir le Recap detaille (Article x Magasin)"):
        df_aff = df_recap_all.drop(columns=["Famille_norm"], errors="ignore")
        st.dataframe(df_aff, use_container_width=True, hide_index=True)

    # === ETAPE 6 : Export ===
    st.markdown("**6 · Telecharger**")
    buf = exporter_excel(df_recap_all, df_financier, kpis, totaux_famille_par_semaine)

    mois_dedans = df_recap_all["Mois"].unique()
    if len(mois_dedans) == 1:
        nom_fichier = f"Fidelite_Cagnotte_{mois_dedans[0].replace(' ', '_')}.xlsx"
    else:
        nom_fichier = f"Fidelite_Cagnotte_multi_mois.xlsx"

    st.download_button(
        label=f"📥 Telecharger {nom_fichier}",
        data=buf,
        file_name=nom_fichier,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# POINT D'ENTREE
# ============================================================
if __name__ == "__main__":
    st.set_page_config(page_title="Fidelite Cagnotte", layout="wide")
    render_fidelite_cagnotte()
