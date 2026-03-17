"""
04_💸_Performance_Promo.py — SmartBuyer Hub
Performance promotionnelle · Article × Magasin · Charte SmartBuyer v2
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Performance Promo · SmartBuyer",
    page_icon="💸",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1200px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stMetricDelta"] { font-size: 12px !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.col-ex   { font-size: 11px; color: #8E8E93; font-family: monospace; margin-top: 2px; }

/* KPI Promo highlight */
.kpi-promo { background: #EFF6FF; border: 1px solid #B3D9FF; border-radius: 12px; padding: 16px 18px; }
.kpi-promo-label { font-size: 11px; font-weight: 500; color: #007AFF; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 3px; }
.kpi-promo-value { font-size: 24px; font-weight: 700; color: #007AFF; letter-spacing: -0.02em; }
.kpi-promo-sub   { font-size: 12px; color: #0066CC; margin-top: 3px; font-weight: 500; }
</style>
""", unsafe_allow_html=True)

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fmt(n):
    if pd.isna(n) or n is None: return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v):
    if pd.isna(v) or v is None: return "—"
    return f"{v:.1f}%"

def statut_marge(tx):
    if pd.isna(tx) or tx is None: return ("🔴 Négatif",   "#FF3B30", "#FFF2F2")
    if tx < 0:                     return ("🔴 Négatif",   "#FF3B30", "#FFF2F2")
    if tx < 3:                     return ("⚠️ Faible",    "#FF9500", "#FFFBF0")
    if tx < 8:                     return ("⚡ Moyen",     "#FF9500", "#FFFBF0")
    return                               ("✅ Bon",        "#34C759", "#F0FFF4")

def extract_periode(df_raw):
    try:
        last = str(df_raw.iloc[-1, 0])
        m = re.search(r"après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})", last)
        if m:
            from datetime import datetime
            d1 = datetime.strptime(m.group(1), "%d/%m/%Y")
            d2 = datetime.strptime(m.group(2), "%d/%m/%Y")
            nb_jours = (d2 - d1).days
            return f"{m.group(1)} → {m.group(2)}", nb_jours
    except: pass
    return "Période inconnue", 7

# ─── PARSING ──────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_liste_promo(byt, fname):
    for enc in ("utf-8-sig","utf-8","latin-1","cp1252"):
        try:
            raw = byt.decode(enc, errors="strict")
            sep = ";" if raw.count(";") > raw.count(",") else ","
            df  = pd.read_csv(BytesIO(byt), sep=sep, encoding=enc, header=None, dtype=str)
            codes = set(df.iloc[:,0].str.strip().dropna().unique())
            if codes: return codes
        except: continue
    return set()

@st.cache_data(show_spinner=False)
def load_pbi(byt, fname):
    ext = fname.lower().rsplit(".",1)[-1]
    if ext in ("xlsx","xls"):
        df = pd.read_excel(BytesIO(byt), dtype=str)
    else:
        for enc in ("utf-8-sig","utf-8","latin-1"):
            try:
                df = pd.read_csv(BytesIO(byt), sep=";", encoding=enc, dtype=str)
                break
            except: continue

    periode, nb_jours = extract_periode(df)

    # Colonnes numériques
    num_cols = ["CA","CA Hors Promo","CA HT Promo","Marge","Marge Hors Promo",
                "Marge Promo","%CA Poids Promo","%Marge","%Marge Promo","Qté Vente"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Extraire code et libellé
    df["code_art"]  = df["Article"].apply(
        lambda s: str(s).split(" - ",1)[0].strip() if (pd.notna(s) and " - " in str(s)) else None)
    df["lib_art"]   = df["Article"].apply(
        lambda s: str(s).split(" - ",1)[-1].strip() if (pd.notna(s) and " - " in str(s)) else None)
    df["lib_site"]  = df["Site nom long"].apply(
        lambda s: str(s).split(" - ",1)[-1].strip() if (pd.notna(s) and " - " in str(s)) else None)
    df["lib_rayon"] = df["Rayon"].apply(
        lambda s: str(s).split(" - ",1)[-1].strip() if (pd.notna(s) and " - " in str(s)) else None)
    df["lib_fam"]   = df["Famille"].apply(
        lambda s: str(s).split(" - ",1)[-1].strip() if (pd.notna(s) and " - " in str(s)) else None)

    return df, periode, nb_jours

def build_dataset(df, promo_codes):
    """Filtre les lignes article × magasin des articles promo."""
    mask = (
        df["code_art"].isin(promo_codes) &
        df["lib_art"].notna() &
        (df["lib_art"] != "Total") &
        df["lib_site"].notna()
    )
    return df[mask].copy()

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def gen_excel(by_art, by_site, df_mx, periode):
    wb    = Workbook()
    HDR_F = PatternFill("solid", fgColor="1C3557")
    HDR_T = Font(bold=True, color="FFFFFF", size=11)
    RED_F = PatternFill("solid", fgColor="FCE4E4")
    AMB_F = PatternFill("solid", fgColor="FEF3CD")
    GRN_F = PatternFill("solid", fgColor="D6F0D6")
    BLU_F = PatternFill("solid", fgColor="DBEAFE")
    NEU_F = PatternFill("solid", fgColor="FFFFFF")
    CTR   = Alignment(horizontal="center", vertical="center")

    def write_ws(ws, headers, rows, title):
        ws.append([title]); ws.cell(1,1).font = Font(bold=True, size=13)
        ws.append([]); ws.append(headers)
        for i,h in enumerate(headers,1):
            c = ws.cell(3,i); c.fill=HDR_F; c.font=HDR_T; c.alignment=CTR
        for row in rows: ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                len(str(ws.cell(3, col[0].column).value or ""))+4, 12)

    # Onglet 1 — Par article
    ws1 = wb.active; ws1.title = "Par article"
    rows1 = [[
        r["code_art"], r["lib_art"],
        round(r["ca"]/1e6,2), round(r["ca_promo"]/1e6,2),
        round(r["poids_promo"],1),
        round(r["ca_hp"]/1e6,2),
        round(r["tx_mg_promo"],1) if not pd.isna(r["tx_mg_promo"]) else "",
        round(r["delta_mg"],1) if not pd.isna(r["delta_mg"]) else "",
        round(r["efficacite"]/1e3,1),
        "Oui" if r["dependance"] else "Non",
        int(r["n_sites"]), int(r["qte"]),
        r["statut"]
    ] for _,r in by_art.iterrows()]
    write_ws(ws1,
        ["Code","Libellé","CA total (M)","CA Promo (M)","Poids promo %",
         "CA Hors Promo (M)","Tx marge promo %","Écart mg promo vs HP (pts)",
         "Efficacité (K/j)","Dépendance 100%","Sites","Qté","Statut marge"],
        rows1, f"Performance promo par article — {periode}")
    for r in ws1.iter_rows(min_row=4, max_row=ws1.max_row):
        s = str(r[12].value or "")
        r[12].fill = RED_F if "Négatif" in s or "Faible" in s else AMB_F if "Moyen" in s else GRN_F if "Bon" in s else NEU_F
        v = r[7].value
        if isinstance(v, (int,float)):
            r[7].fill = RED_F if v < 0 else AMB_F if v < -1 else NEU_F
        d = r[9].value
        if d == "Oui": r[9].fill = AMB_F

    # Onglet 2 — Par magasin
    ws2 = wb.create_sheet("Par magasin")
    rows2 = [[
        r["lib_site"],
        round(r["ca"]/1e6,2), round(r["ca_promo"]/1e6,2),
        round(r["poids_promo"],1), round(r["ca_hp"]/1e6,2),
        round(r["marge"]/1e6,2), round(r["mg_promo"]/1e6,2),
        round(r["tx_mg_promo"],1) if not pd.isna(r["tx_mg_promo"]) else "",
        int(r["qte"])
    ] for _,r in by_site.iterrows()]
    write_ws(ws2,
        ["Magasin","CA total (M)","CA Promo (M)","Poids promo %",
         "CA Hors Promo (M)","Marge (M)","Marge Promo (M)","Tx marge promo %","Qté"],
        rows2, f"Performance promo par magasin — {periode}")
    for r in ws2.iter_rows(min_row=4, max_row=ws2.max_row):
        v = r[7].value
        if isinstance(v,(int,float)):
            r[7].fill = RED_F if v < 3 else AMB_F if v < 6 else GRN_F

    # Onglet 3 — Article × Magasin
    ws3 = wb.create_sheet("Article × Magasin")
    cols_mx = ["code_art","lib_art","lib_site","ca","ca_promo","poids_promo",
               "ca_hp","marge","mg_promo","tx_mg_promo","qte"]
    rows3 = [[
        r["code_art"], r["lib_art"], r["lib_site"],
        round(r["ca"]/1e6,2), round(r["ca_promo"]/1e6,2),
        round(r["poids_promo"],1), round(r["ca_hp"]/1e6,2),
        round(r["marge"]/1e6,2),
        round(r["mg_promo"]/1e6,2),
        round(r["tx_mg_promo"],1) if not pd.isna(r["tx_mg_promo"]) else "",
        int(r["qte"])
    ] for _,r in df_mx.iterrows()]
    write_ws(ws3,
        ["Code","Article","Magasin","CA (M)","CA Promo (M)","Poids %",
         "CA HP (M)","Marge (M)","Marge Promo (M)","Tx mg promo %","Qté"],
        rows3, f"Détail article × magasin — {periode}")
    for r in ws3.iter_rows(min_row=4, max_row=ws3.max_row):
        v = r[9].value
        if isinstance(v,(int,float)):
            r[9].fill = RED_F if v < 0 else AMB_F if v < 3 else NEU_F

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Navigation</div>", unsafe_allow_html=True)
    st.page_link("app.py",                                  label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",      label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",               label="📈  Ventes PBI")
    st.page_link("pages/03_📦_Detention_Top_CA.py",         label="📦  Détention Top CA")
    st.page_link("pages/04_💸_Performance_Promo.py",        label="💸  Performance Promo")
    st.page_link("pages/05_🏪_Suivi_Implantation.py",       label="🏪  Suivi Implantation", disabled=True)
    st.markdown("---")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichiers</div>", unsafe_allow_html=True)
    f_liste = st.file_uploader("Liste articles promo (CSV)", type=["csv","xlsx"], key="liste")
    f_pbi   = st.file_uploader("Export PBI ventes (Excel)", type=["xlsx","xls","csv"], key="pbi")

# ─── PAGE PRINCIPALE ──────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>💸 Performance Promotion</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Analyse des ventes promotionnelles · Article × Magasin · Taux de marge · Efficacité · Dépendance</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL ──────────────────────────────────────────────────────────
if not f_liste or not f_pbi:
    st.markdown("---")

    # Description du module
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Ce module mesure la <strong>performance des articles mis en promotion</strong> sur une période donnée.
  Il croise la liste des articles promos avec l'export PBI des ventes pour répondre à 4 questions clés :
  <br><br>
  <strong>1. Le CA promo est-il significatif ?</strong> — Poids du CA promo dans le CA total de l'article<br>
  <strong>2. La promo est-elle rentable ?</strong> — Taux de marge promo vs taux de marge hors promo<br>
  <strong>3. L'article est-il dépendant de la promo ?</strong> — Articles vendus à 100% en promo<br>
  <strong>4. La promo est-elle efficace dans le temps ?</strong> — CA promo par jour de promotion
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Les 4 indicateurs
    st.markdown("<div class='section-label'>Les 4 indicateurs clés</div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    indics = [
        ("📊","Poids promo","#007AFF",
         "Part du CA réalisée en mode promotionnel",
         "CA Promo ÷ CA total × 100",
         "Un poids > 80% signale une forte dépendance à la promotion."),
        ("📉","Écart marge","#FF3B30",
         "Delta entre taux de marge promo et hors promo",
         "Tx marge promo − Tx marge hors promo",
         "Un écart négatif = la promo coûte de la marge. À surveiller avant renouvellement."),
        ("⚡","Dépendance promo","#FF9500",
         "Article vendu exclusivement en mode promo",
         "CA Hors Promo = 0 sur la période",
         "Signal que l'article ne se vend pas sans promotion — risque de dévalorisation."),
        ("🚀","Efficacité promo","#34C759",
         "CA promo généré par jour de promotion",
         "CA Promo ÷ Nb jours de la période",
         "Permet de comparer des promos de durées différentes sur une base commune."),
    ]
    for i, (ico, titre, color, desc, formule, interp) in enumerate(indics):
        with (c1 if i%2==0 else c2):
            st.markdown(f"""
<div style='background:#FFFFFF;border:0.5px solid #E5E5EA;border-radius:12px;
            padding:16px;border-left:3px solid {color};margin-bottom:10px'>
  <div style='display:flex;align-items:center;gap:8px;margin-bottom:8px'>
    <span style='font-size:18px'>{ico}</span>
    <span style='font-size:14px;font-weight:600;color:#1C1C1E'>{titre}</span>
  </div>
  <div style='font-size:12px;color:#3A3A3C;margin-bottom:4px'>{desc}</div>
  <div style='font-size:11px;color:{color};font-family:monospace;background:#F9F9FB;
              padding:4px 8px;border-radius:6px;margin-bottom:6px'>{formule}</div>
  <div style='font-size:11px;color:#8E8E93;font-style:italic'>{interp}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Statuts marge
    st.markdown("<div class='section-label'>Grille de lecture — Statut marge promo</div>", unsafe_allow_html=True)
    sc1, sc2, sc3, sc4 = st.columns(4)
    for col_s, ico, label, seuil, color, bg in [
        (sc1,"✅","Bon","Tx marge promo ≥ 8%","#34C759","#F0FFF4"),
        (sc2,"⚡","Moyen","3% ≤ Tx marge < 8%","#FF9500","#FFFBF0"),
        (sc3,"⚠️","Faible","0% ≤ Tx marge < 3%","#FF9500","#FFFBF0"),
        (sc4,"🔴","Négatif","Tx marge promo < 0%","#FF3B30","#FFF2F2"),
    ]:
        with col_s:
            st.markdown(f"""
<div style='background:{bg};border:0.5px solid {color};border-radius:10px;
            padding:12px;text-align:center'>
  <div style='font-size:20px'>{ico}</div>
  <div style='font-size:13px;font-weight:600;color:{color};margin-top:4px'>{label}</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:3px'>{seuil}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Fichiers attendus
    st.markdown("<div class='section-label'>Fichiers attendus</div>", unsafe_allow_html=True)
    cf1, cf2 = st.columns(2)
    with cf1:
        st.markdown("""
<div class='col-required'><div style='font-size:16px'>📋</div>
<div><div class='col-name'>Liste articles promo</div>
<div class='col-desc'>CSV · 1 colonne · codes articles sans en-tête</div>
<div class='col-ex'>ex: 14006617 / 10002101 / 14005975…</div></div></div>""", unsafe_allow_html=True)
    with cf2:
        st.markdown("""
<div class='col-required'><div style='font-size:16px'>📊</div>
<div><div class='col-name'>Export PBI ventes</div>
<div class='col-desc'>Excel · colonnes CA, CA HT Promo, Marge Promo…</div>
<div class='col-ex'>Export standard PBI avec CA Promo et CA Hors Promo</div></div></div>""", unsafe_allow_html=True)

    st.info("⬆️ Charge les deux fichiers dans la sidebar pour lancer l'analyse.")
    st.stop()

# ─── TRAITEMENT ───────────────────────────────────────────────────────────────
with st.spinner("Lecture des fichiers…"):
    promo_codes   = load_liste_promo(f_liste.read(), f_liste.name)
    pbi_bytes     = f_pbi.read()
    df_pbi, periode, nb_jours = load_pbi(pbi_bytes, f_pbi.name)

if not promo_codes:
    st.error("Liste promo vide ou illisible."); st.stop()
if df_pbi.empty:
    st.error("Export PBI vide ou illisible."); st.stop()

df_raw = build_dataset(df_pbi, promo_codes)

if df_raw.empty:
    st.error("Aucun article de la liste promo trouvé dans l'export PBI."); st.stop()

# Filtre sidebar
with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtres</div>", unsafe_allow_html=True)
    rayons_dispo = sorted(df_raw["lib_rayon"].dropna().unique())
    sel_rayon    = st.multiselect("Rayon", rayons_dispo, default=rayons_dispo)
    sites_dispo  = sorted(df_raw["lib_site"].dropna().unique())
    sel_site     = st.multiselect("Magasin", sites_dispo, default=sites_dispo)
    st.markdown("---")
    st.caption(f"**Période :** {periode}")
    st.caption(f"**Durée :** {nb_jours} jours")

df = df_raw[
    df_raw["lib_rayon"].isin(sel_rayon) &
    df_raw["lib_site"].isin(sel_site)
].copy()

# ─── AGRÉGATIONS ──────────────────────────────────────────────────────────────
# Par article
by_art = df.groupby(["code_art","lib_art"], as_index=False).agg(
    ca=("CA","sum"), ca_promo=("CA HT Promo","sum"),
    ca_hp=("CA Hors Promo","sum"), marge=("Marge","sum"),
    mg_promo=("Marge Promo","sum"), mg_hp=("Marge Hors Promo","sum"),
    qte=("Qté Vente","sum"), n_sites=("lib_site","nunique")
)
by_art["poids_promo"]  = np.where(by_art["ca"]>0, by_art["ca_promo"]/by_art["ca"]*100, 0)
by_art["tx_mg_promo"]  = np.where(by_art["ca_promo"]>0, by_art["mg_promo"]/by_art["ca_promo"]*100, np.nan)
by_art["tx_mg_hp"]     = np.where(by_art["ca_hp"]>0,    by_art["mg_hp"]/by_art["ca_hp"]*100, np.nan)
by_art["tx_mg_total"]  = np.where(by_art["ca"]>0,       by_art["marge"]/by_art["ca"]*100, np.nan)
by_art["delta_mg"]     = by_art["tx_mg_promo"] - by_art["tx_mg_hp"]
by_art["efficacite"]   = by_art["ca_promo"] / nb_jours  # FCFA par jour
by_art["dependance"]   = by_art["ca_hp"] <= 0           # 100% en promo
by_art["statut"]       = by_art["tx_mg_promo"].apply(lambda x: statut_marge(x)[0])
by_art = by_art.sort_values("ca_promo", ascending=False)

# Par magasin
by_site = df.groupby("lib_site", as_index=False).agg(
    ca=("CA","sum"), ca_promo=("CA HT Promo","sum"),
    ca_hp=("CA Hors Promo","sum"), marge=("Marge","sum"),
    mg_promo=("Marge Promo","sum"), qte=("Qté Vente","sum")
)
by_site["poids_promo"]  = np.where(by_site["ca"]>0, by_site["ca_promo"]/by_site["ca"]*100, 0)
by_site["tx_mg_promo"]  = np.where(by_site["ca_promo"]>0, by_site["mg_promo"]/by_site["ca_promo"]*100, np.nan)
by_site = by_site.sort_values("ca_promo", ascending=False)

# Article × Magasin
df_mx = df.groupby(["code_art","lib_art","lib_site"], as_index=False).agg(
    ca=("CA","sum"), ca_promo=("CA HT Promo","sum"),
    ca_hp=("CA Hors Promo","sum"), marge=("Marge","sum"),
    mg_promo=("Marge Promo","sum"), qte=("Qté Vente","sum")
)
df_mx["poids_promo"] = np.where(df_mx["ca"]>0, df_mx["ca_promo"]/df_mx["ca"]*100, 0)
df_mx["tx_mg_promo"] = np.where(df_mx["ca_promo"]>0, df_mx["mg_promo"]/df_mx["ca_promo"]*100, np.nan)
df_mx = df_mx.sort_values("ca_promo", ascending=False)

# KPIs globaux
ca_total   = df["CA"].sum()
ca_promo   = df["CA HT Promo"].sum()
ca_hp      = df["CA Hors Promo"].sum()
mg_total   = df["Marge"].sum()
mg_promo   = df["Marge Promo"].sum()
mg_hp      = df["Marge Hors Promo"].sum()
qte_total  = df["Qté Vente"].sum()
tx_mg_tot  = mg_total/ca_total*100 if ca_total>0 else 0
tx_mg_pro  = mg_promo/ca_promo*100 if ca_promo>0 else 0
tx_mg_hp_  = mg_hp/ca_hp*100 if ca_hp>0 else 0
delta_mg   = tx_mg_pro - tx_mg_hp_
poids_prom = ca_promo/ca_total*100 if ca_total>0 else 0
efficacite = ca_promo/nb_jours if nb_jours>0 else 0
n_dep      = int(by_art["dependance"].sum())
n_neg      = int((by_art["tx_mg_promo"]<0).sum())

# ─── KPIs ─────────────────────────────────────────────────────────────────────
st.markdown(f"<div class='section-label'>{len(promo_codes)} articles promo · {len(sel_site)} magasin(s) · {periode}</div>", unsafe_allow_html=True)

k1,k2,k3,k4,k5 = st.columns(5)
k1.metric("CA articles promo",   fmt(ca_total),  "FCFA période")
k2.metric("Qté vendue",          f"{int(qte_total):,}", "unités")
k3.metric("Taux marge global",   fmt_pct(tx_mg_tot),   f"promo {fmt_pct(tx_mg_pro)}")
k4.metric("Écart marge",         f"{delta_mg:+.1f} pts", "promo vs hors promo")
with k5:
    st.markdown(f"""
<div class='kpi-promo'>
  <div class='kpi-promo-label'>CA Promo</div>
  <div class='kpi-promo-value'>{fmt(ca_promo)}</div>
  <div class='kpi-promo-sub'>{poids_prom:.1f}% du CA total</div>
</div>""", unsafe_allow_html=True)

# ─── ALERTES ──────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Alertes &amp; points d'attention</div>", unsafe_allow_html=True)

if n_neg > 0:
    arts_neg = by_art[by_art["tx_mg_promo"]<0]["lib_art"].tolist()
    st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>🔴 {n_neg} article(s) en marge négative en mode promo</strong> :
  {", ".join(arts_neg[:3])}{"…" if len(arts_neg)>3 else ""}<br>
  <span style='font-size:12px;opacity:.85'>→ Vérifier les conditions d'achat avant tout renouvellement de ces promos.</span>
</div>""", unsafe_allow_html=True)

if delta_mg < -2:
    st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>⚠️ La promo dégrade la marge de {abs(delta_mg):.1f} pts</strong>
  — Taux hors promo {fmt_pct(tx_mg_hp_)} vs promo {fmt_pct(tx_mg_pro)}<br>
  <span style='font-size:12px;opacity:.85'>→ Analyser article par article l'impact réel des remises accordées.</span>
</div>""", unsafe_allow_html=True)

if n_dep > 0:
    arts_dep = by_art[by_art["dependance"]]["lib_art"].tolist()
    st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>⚡ {n_dep} article(s) vendus à 100% en mode promo</strong> :
  {", ".join([a[:30] for a in arts_dep[:3]])}{"…" if len(arts_dep)>3 else ""}<br>
  <span style='font-size:12px;opacity:.85'>→ Ces articles ne tournent pas hors promo — risque de dévalorisation de la marque.</span>
</div>""", unsafe_allow_html=True)

if n_neg==0 and delta_mg>=-2:
    st.success(f"✅ Performance promo globalement saine · {fmt(ca_promo)} FCFA · {poids_prom:.1f}% de poids promo")

# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab1, tab2, tab3, tab4 = st.tabs([
    f"📋 Par article ({len(by_art)})",
    f"🏪 Par magasin ({len(by_site)})",
    "🔢 Article × Magasin",
    "📈 Pareto & Efficacité",
])

# ═══ TAB 1 — PAR ARTICLE ══════════════════════════════════════════════════════
with tab1:
    st.caption("Trié par CA Promo décroissant · Statut marge basé sur le taux de marge en mode promotionnel")

    # Préparer affichage
    disp1 = by_art[[
        "lib_art","ca","ca_promo","poids_promo","ca_hp",
        "tx_mg_promo","tx_mg_hp","delta_mg","efficacite",
        "dependance","n_sites","qte","statut"
    ]].copy()

    disp1["CA total"]        = disp1["ca"].apply(fmt)
    disp1["CA Promo"]        = disp1["ca_promo"].apply(fmt)
    disp1["Poids promo %"]   = disp1["poids_promo"].apply(lambda x: f"{x:.0f}%")
    disp1["CA Hors Promo"]   = disp1["ca_hp"].apply(fmt)
    disp1["Tx mg promo %"]   = disp1["tx_mg_promo"].apply(fmt_pct)
    disp1["Tx mg HP %"]      = disp1["tx_mg_hp"].apply(fmt_pct)
    disp1["Écart mg (pts)"]  = disp1["delta_mg"].apply(lambda x: f"{x:+.1f}" if not pd.isna(x) else "—")
    disp1["Efficacité /j"]   = disp1["efficacite"].apply(lambda x: fmt(x) + " /j")
    disp1["Dépendance"]      = disp1["dependance"].apply(lambda x: "⚡ 100% promo" if x else "")
    disp1["Sites"]           = disp1["n_sites"].apply(int)
    disp1["Qté"]             = disp1["qte"].apply(lambda x: f"{int(x):,}")

    final1 = disp1.rename(columns={"lib_art":"Article","statut":"Statut marge"})[
        ["Article","CA total","CA Promo","Poids promo %","CA Hors Promo",
         "Tx mg promo %","Tx mg HP %","Écart mg (pts)","Efficacité /j",
         "Dépendance","Sites","Qté","Statut marge"]
    ]
    st.dataframe(final1, use_container_width=True, hide_index=True,
                 column_config={
                     "Article":       st.column_config.TextColumn("Article",      width="large"),
                     "Poids promo %": st.column_config.TextColumn("Poids promo",  width="small"),
                     "Sites":         st.column_config.NumberColumn("Sites",       format="%d"),
                 })

    st.markdown("""
<div style='font-size:11px;color:#8E8E93;margin-top:8px;font-style:italic'>
  ✅ Bon ≥ 8% · ⚡ Moyen 3–8% · ⚠️ Faible 0–3% · 🔴 Négatif &lt; 0% · Efficacité = CA promo ÷ nb jours
</div>""", unsafe_allow_html=True)

# ═══ TAB 2 — PAR MAGASIN ══════════════════════════════════════════════════════
with tab2:
    disp2 = by_site.copy()
    disp2["CA total"]      = disp2["ca"].apply(fmt)
    disp2["CA Promo"]      = disp2["ca_promo"].apply(fmt)
    disp2["Poids promo %"] = disp2["poids_promo"].apply(lambda x: f"{x:.0f}%")
    disp2["CA Hors Promo"] = disp2["ca_hp"].apply(fmt)
    disp2["Marge"]         = disp2["marge"].apply(fmt)
    disp2["Marge Promo"]   = disp2["mg_promo"].apply(fmt)
    disp2["Tx mg promo %"] = disp2["tx_mg_promo"].apply(fmt_pct)
    disp2["Qté"]           = disp2["qte"].apply(lambda x: f"{int(x):,}")

    st.dataframe(
        disp2.rename(columns={"lib_site":"Magasin"})[
            ["Magasin","CA total","CA Promo","Poids promo %","CA Hors Promo",
             "Marge","Marge Promo","Tx mg promo %","Qté"]
        ],
        use_container_width=True, hide_index=True
    )

    # Mini graphique poids promo par magasin
    try:
        import plotly.graph_objects as go
        s = by_site.sort_values("poids_promo")
        fig = go.Figure(go.Bar(
            x=s["poids_promo"].tolist(), y=s["lib_site"].tolist(),
            orientation="h",
            marker_color=[
                "#34C759" if v<70 else "#FF9500" if v<90 else "#FF3B30"
                for v in s["poids_promo"]
            ],
            marker_line_width=0,
            text=[f"{v:.0f}%" for v in s["poids_promo"]],
            textposition="outside",
        ))
        fig.update_layout(
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=11),
            height=max(240, len(by_site)*44+60),
            margin=dict(t=10,b=10,l=10,r=60),
            xaxis=dict(title="Poids promo (%)", ticksuffix="%",
                       showgrid=True, gridcolor="#F2F2F7", range=[0,115]),
            yaxis=dict(showgrid=False, title=""),
        )
        fig.add_vline(x=80, line_width=1.5, line_dash="dot", line_color="#007AFF",
                      annotation_text=" 80%", annotation_font=dict(color="#007AFF", size=10))
        st.plotly_chart(fig, use_container_width=True)
    except ImportError:
        pass

# ═══ TAB 3 — ARTICLE × MAGASIN ════════════════════════════════════════════════
with tab3:
    # Filtres inline
    fc1, fc2 = st.columns(2)
    with fc1:
        sel_art_mx = st.selectbox("Article", ["Tous"]+sorted(df_mx["lib_art"].unique()))
    with fc2:
        sel_site_mx = st.selectbox("Magasin ", ["Tous"]+sorted(df_mx["lib_site"].unique()))

    df_view = df_mx.copy()
    if sel_art_mx  != "Tous": df_view = df_view[df_view["lib_art"]==sel_art_mx]
    if sel_site_mx != "Tous": df_view = df_view[df_view["lib_site"]==sel_site_mx]

    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>{len(df_view)} lignes affichées</div>", unsafe_allow_html=True)

    disp3 = df_view.copy()
    disp3["CA total"]      = disp3["ca"].apply(fmt)
    disp3["CA Promo"]      = disp3["ca_promo"].apply(fmt)
    disp3["Poids %"]       = disp3["poids_promo"].apply(lambda x: f"{x:.0f}%")
    disp3["CA HP"]         = disp3["ca_hp"].apply(fmt)
    disp3["Marge Promo"]   = disp3["mg_promo"].apply(fmt)
    disp3["Tx mg promo %"] = disp3["tx_mg_promo"].apply(fmt_pct)
    disp3["Qté"]           = disp3["qte"].apply(lambda x: f"{int(x):,}")

    st.dataframe(
        disp3.rename(columns={"lib_art":"Article","lib_site":"Magasin"})[
            ["Article","Magasin","CA total","CA Promo","Poids %","CA HP",
             "Marge Promo","Tx mg promo %","Qté"]
        ],
        use_container_width=True, hide_index=True
    )

# ═══ TAB 4 — PARETO & EFFICACITÉ ══════════════════════════════════════════════
with tab4:
    c1, c2 = st.columns(2)

    with c1:
        st.markdown("<div class='section-label'>Pareto — Contribution au CA promo</div>", unsafe_allow_html=True)
        try:
            import plotly.graph_objects as go
            pa = by_art.sort_values("ca_promo", ascending=False).copy()
            pa["cum_pct"] = pa["ca_promo"].cumsum() / pa["ca_promo"].sum() * 100
            pa["lib_short"] = pa["lib_art"].str[:22]

            fig_p = go.Figure()
            fig_p.add_bar(
                x=pa["lib_short"].tolist(),
                y=(pa["ca_promo"]/1e6).tolist(),
                name="CA Promo (M FCFA)",
                marker_color="#007AFF", marker_line_width=0,
            )
            fig_p.add_scatter(
                x=pa["lib_short"].tolist(),
                y=pa["cum_pct"].tolist(),
                name="Cumul %", yaxis="y2",
                line=dict(color="#FF9500", width=2),
                mode="lines+markers", marker=dict(size=5),
            )
            fig_p.add_hline(y=80, line_dash="dot", line_color="#FF3B30",
                           line_width=1, yref="y2",
                           annotation_text="80%", annotation_font=dict(color="#FF3B30",size=9))
            fig_p.update_layout(
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=10),
                height=340, margin=dict(t=10,b=60,l=10,r=50),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=10)),
                xaxis=dict(showgrid=False, tickangle=-35, tickfont=dict(size=9)),
                yaxis=dict(title="CA Promo (M)", showgrid=True, gridcolor="#F2F2F7"),
                yaxis2=dict(title="Cumul %", overlaying="y", side="right",
                            range=[0,110], ticksuffix="%"),
            )
            st.plotly_chart(fig_p, use_container_width=True)
            # Trouver le seuil 80%
            n80 = int((pa["cum_pct"]<=80).sum()) + 1
            st.caption(f"→ {n80} article(s) concentrent 80% du CA promo")
        except ImportError:
            st.dataframe(by_art[["lib_art","ca_promo"]].head(10), use_container_width=True)

    with c2:
        st.markdown("<div class='section-label'>Efficacité promo — CA promo par jour</div>", unsafe_allow_html=True)
        try:
            eff = by_art.sort_values("efficacite", ascending=True).copy()
            eff["lib_short"] = eff["lib_art"].str[:22]
            fig_e = go.Figure(go.Bar(
                x=(eff["efficacite"]/1e3).tolist(),
                y=eff["lib_short"].tolist(),
                orientation="h",
                marker_color=[
                    "#34C759" if v>=500_000/nb_jours
                    else "#FF9500" if v>=200_000/nb_jours
                    else "#FF3B30"
                    for v in eff["efficacite"]
                ],
                marker_line_width=0,
                text=[(f"{v/1e3:.0f}K") for v in eff["efficacite"]],
                textposition="outside",
            ))
            fig_e.update_layout(
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=10),
                height=340, margin=dict(t=10,b=10,l=10,r=60),
                xaxis=dict(title=f"CA promo / jour (K FCFA)",
                           showgrid=True, gridcolor="#F2F2F7"),
                yaxis=dict(showgrid=False, title="", tickfont=dict(size=9)),
            )
            st.plotly_chart(fig_e, use_container_width=True)
            st.caption(f"Durée de la période : {nb_jours} jours · Efficacité = CA Promo ÷ {nb_jours}j")
        except ImportError:
            pass

    # Dépendance promo
    dep_arts = by_art[by_art["dependance"]]
    if not dep_arts.empty:
        st.markdown("<div class='section-label' style='margin-top:16px'>Articles à dépendance totale (100% en promo)</div>", unsafe_allow_html=True)
        st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>⚡ {len(dep_arts)} article(s) sans vente hors promo sur la période</strong><br>
  Ces articles génèrent <strong>{fmt(dep_arts['ca_promo'].sum())}</strong> FCFA
  uniquement en mode promotionnel.
  <span style='font-size:12px;opacity:.85;display:block;margin-top:4px'>
  → Surveiller : si l'article ne se vend pas hors promo, la promotion crée une habitude d'achat à prix réduit.
  </span>
</div>""", unsafe_allow_html=True)
        dep_disp = dep_arts[["lib_art","ca_promo","tx_mg_promo","n_sites","qte"]].copy()
        dep_disp.columns = ["Article","CA Promo","Tx marge promo %","Sites","Qté"]
        dep_disp["CA Promo"]        = dep_disp["CA Promo"].apply(fmt)
        dep_disp["Tx marge promo %"]= dep_disp["Tx marge promo %"].apply(fmt_pct)
        dep_disp["Qté"]             = dep_disp["Qté"].apply(lambda x: f"{int(x):,}")
        st.dataframe(dep_disp, use_container_width=True, hide_index=True)

# ─── EXPORT ───────────────────────────────────────────────────────────────────
st.markdown("---")
with st.expander("📥 Export Excel — Par article · Par magasin · Article × Magasin"):
    st.caption(f"3 onglets · {len(by_art)} articles · {len(by_site)} magasins · {periode}")
    if st.button("Générer le fichier Excel", type="primary"):
        with st.spinner("Génération…"):
            buf = gen_excel(by_art, by_site, df_mx, periode)
        st.download_button(
            "⬇️ Télécharger",
            data=buf,
            file_name=f"SmartBuyer_Performance_Promo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
