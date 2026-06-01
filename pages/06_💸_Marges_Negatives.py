"""
06_💸_Marges_Negatives.py — SmartBuyer Hub
Diagnostic Rentabilité Réseau · Flop 100 · Analyse par format et rayon
v2.1 — Ajout code article dans Flop 100 et Analyse Casse
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Marges Négatives · SmartBuyer",
    page_icon="💸",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1300px; }
[data-testid="stSidebar"] { background: #F2F2F7 !important; border-right: 0.5px solid #D1D1D6 !important; }
[data-testid="stMetric"] { background: #FFFFFF !important; border: 0.5px solid #E5E5EA !important; border-radius: 12px !important; padding: 16px 18px !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 600 !important; color: #1C1C1E !important; letter-spacing: -0.02em !important; }
[data-testid="stTabs"] button[role="tab"] { font-size: 13px !important; font-weight: 500 !important; padding: 8px 16px !important; color: #8E8E93 !important; border-radius: 0 !important; border-bottom: 2px solid transparent !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #007AFF !important; border-bottom: 2px solid #007AFF !important; background: transparent !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom: 0.5px solid #E5E5EA !important; }
[data-testid="stDataFrame"] { border: 0.5px solid #E5E5EA !important; border-radius: 10px !important; }
[data-testid="stDataFrame"] th { background: #F2F2F7 !important; font-size: 11px !important; font-weight: 600 !important; color: #8E8E93 !important; text-transform: uppercase !important; letter-spacing: 0.04em !important; }
[data-testid="stFileUploader"] { border: 1.5px dashed #D1D1D6 !important; border-radius: 10px !important; background: #F9F9FB !important; }
.stDownloadButton > button { background: #007AFF !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 500 !important; font-size: 13px !important; padding: 10px 24px !important; width: 100% !important; }
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }
.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E; letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px; margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 10px; }
.alert-card  { padding: 12px 16px; border-radius: 10px; margin-bottom: 8px; font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-red   { background: #FFF2F2; border-color: #FF3B30; color: #3A0000; }
.alert-amber { background: #FFFBF0; border-color: #FF9500; color: #3A2000; }
.alert-green { background: #F0FFF4; border-color: #34C759; color: #003A10; }
.alert-blue  { background: #F0F8FF; border-color: #007AFF; color: #001A3A; }
.alert-purple{ background: #F5F0FF; border-color: #AF52DE; color: #1A0033; }
.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 10px; }
.col-name { font-size: 13px; font-weight: 600; color: #0066CC; font-family: monospace; }
.col-desc { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
</style>
""", unsafe_allow_html=True)

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fmt(n):
    if pd.isna(n) or n is None: return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if pd.isna(v) or v is None: return "—"
    return f"{v:.{dec}f}%"

def fmt_delta(v):
    if pd.isna(v) or v is None: return "—"
    return f"{v:+.1f} pts"

def get_format(site_name):
    s = str(site_name)
    if "Supeco" in s: return "Supeco"
    if "Hyper"  in s: return "Hyper"
    return "Market"

def short_name(s):
    s = str(s)
    return s.split(" - ", 1)[-1].strip() if " - " in s else s

def extract_code(s):
    if pd.isna(s): return None
    return str(s).split(" - ", 1)[0].strip() if " - " in str(s) else None

def extract_periode(df_raw):
    try:
        for val in df_raw.iloc[:, 0].astype(str):
            m = re.search(r"après le (\d{2}/\d{2}/\d{4}) et est avant le (\d{2}/\d{2}/\d{4})", val)
            if m:
                from datetime import datetime
                d1 = datetime.strptime(m.group(1), "%d/%m/%Y")
                d2 = datetime.strptime(m.group(2), "%d/%m/%Y")
                nb = (d2 - d1).days
                return f"{m.group(1)} → {m.group(2)}", nb
    except: pass
    return "Période inconnue", 1

# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_data(byt, fname):
    ext = fname.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        df = pd.read_excel(BytesIO(byt), dtype=str)
    else:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                df = pd.read_csv(BytesIO(byt), sep=";", encoding=enc, dtype=str)
                break
            except: continue

    periode, nb_jours = extract_periode(df)

    if "CA Promo" in df.columns and "CA HT Promo" not in df.columns:
        df = df.rename(columns={"CA Promo": "CA HT Promo"})

    num_cols = ["CA", "Marge", "CA Hors Promo", "Marge Hors Promo",
                "CA HT Promo", "Marge Promo", "Qté Vente",
                "Casse (Valeur)", "Casse (Qté)", "%Marge", "%CA Poids Promo"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["lib_art"]   = df["Article"].apply(lambda s: short_name(s) if pd.notna(s) else None)
    df["code_art"]  = df["Article"].apply(extract_code)
    df["lib_site"]  = df["Site nom long"].apply(lambda s: short_name(s) if pd.notna(s) else None)
    df["lib_rayon"] = df["Rayon"].apply(lambda s: short_name(s) if pd.notna(s) else None)
    df["lib_fam"]   = df["Famille"].apply(lambda s: short_name(s) if pd.notna(s) else None)
    df["format"]    = df["Site nom long"].apply(lambda s: get_format(s) if pd.notna(s) else None)

    df_clean = df[
        df["lib_art"].notna()   & (df["lib_art"]   != "Total") &
        df["lib_site"].notna()  & (df["lib_site"]  != "Total") &
        df["lib_rayon"].notna() &
        ~df["Rayon"].astype(str).str.startswith("Filtres") &
        (df["lib_rayon"] != "Total") &
        df["lib_fam"].notna()   & (df["lib_fam"]   != "Total")
    ].copy()

    return df_clean, periode, nb_jours

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:18px'>
  <div style='font-size:20px;font-weight:700;color:#1C1C1E;letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:1px'>Hub analytique · Équipe Achats</div>
</div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Import fichier</div>", unsafe_allow_html=True)
    f_pbi = st.file_uploader("Export PBI ventes (Excel)", type=["xlsx", "xls", "csv"], key="pbi_marge")

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>💸 Diagnostic Rentabilité Réseau</div>", unsafe_allow_html=True)
st.markdown("<div class='page-caption'>Analyse des marges · Flop 100 destructeurs · Décomposition par format (Hyper / Market / Supeco) · Fuites de valeur</div>", unsafe_allow_html=True)

# ─── ÉCRAN D'ACCUEIL ──────────────────────────────────────────────────────────
if not f_pbi:
    st.markdown("---")
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>ℹ️ À quoi sert ce module ?</strong><br>
  Diagnostic complet de la rentabilité réseau à partir d'un export PBI ventes.
  Identifie précisément où se perdent les marges : par rayon, par format de magasin, par article.<br><br>
  <strong>1. Vue réseau globale</strong> — KPIs, synthèse par format et par rayon, palmarès magasins<br>
  <strong>2. Matrice rayon × magasin</strong> — Taux de marge croisé pour repérer les combinaisons critiques<br>
  <strong>3. Flop 100</strong> — Articles destructeurs avec code article, impact par site (Hyper / Market / Supeco)<br>
  <strong>4. Analyse des fuites</strong> — Effet promo, casse, familles sous seuil de rentabilité
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Fichier attendu</div>", unsafe_allow_html=True)
    st.markdown("""
<div class='col-required'><div style='font-size:16px'>📊</div>
<div><div class='col-name'>Export PBI ventes réseau</div>
<div class='col-desc'>Excel · Axes : Rayon / Famille / Article / Site nom long · Colonnes : CA, Marge, CA Promo, Marge Promo, CA Hors Promo, Marge Hors Promo, Qté Vente, Casse (Valeur)</div>
</div></div>""", unsafe_allow_html=True)
    st.info("⬆️ Charge le fichier export PBI dans la sidebar pour lancer le diagnostic.")
    st.stop()

# ─── CHARGEMENT & CALCULS ─────────────────────────────────────────────────────
with st.spinner("Lecture et analyse des données…"):
    df, periode, nb_jours = load_data(f_pbi.read(), f_pbi.name)

if df.empty:
    st.error("Fichier vide ou colonnes non reconnues.")
    st.stop()

with st.sidebar:
    st.markdown("---")
    st.markdown("<div style='font-size:11px;font-weight:600;color:#8E8E93;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>Filtres</div>", unsafe_allow_html=True)
    formats_dispo = sorted(df["format"].dropna().unique())
    sel_format = st.multiselect("Format magasin", formats_dispo, default=formats_dispo)
    rayons_dispo = sorted(df["lib_rayon"].dropna().unique())
    sel_rayon = st.multiselect("Rayon", rayons_dispo, default=rayons_dispo)
    sites_dispo = sorted(df[df["format"].isin(sel_format)]["lib_site"].dropna().unique())
    sel_site = st.multiselect("Magasin", sites_dispo, default=sites_dispo)
    st.markdown("---")
    st.caption(f"**Période :** {periode}")
    st.caption(f"**Durée :** {nb_jours} jour(s)")

df_f = df[
    df["format"].isin(sel_format) &
    df["lib_rayon"].isin(sel_rayon) &
    df["lib_site"].isin(sel_site)
].copy()

if df_f.empty:
    st.warning("Aucune donnée pour la sélection en cours.")
    st.stop()

# ── Agrégations ───────────────────────────────────────────────────────────────
nb_sites_actifs = df_f[df_f["CA"] > 0]["lib_site"].nunique()

def agg_base(df_in, by):
    g = df_in.groupby(by).agg(
        CA=("CA","sum"), Marge=("Marge","sum"),
        CA_Promo=("CA HT Promo","sum"), CA_HP=("CA Hors Promo","sum"),
        Marge_HP=("Marge Hors Promo","sum"), Marge_Promo=("Marge Promo","sum"),
        Casse=("Casse (Valeur)","sum")
    ).reset_index()
    g["TxMarge"]       = (g["Marge"] / g["CA"] * 100).where(g["CA"] > 0)
    g["PdsPromo"]      = (g["CA_Promo"] / g["CA"] * 100).where(g["CA"] > 0)
    g["TxMarge_HP"]    = (g["Marge_HP"] / g["CA_HP"] * 100).where(g["CA_HP"] > 0)
    g["TxMarge_Promo"] = (g["Marge_Promo"] / g["CA_Promo"] * 100).where(g["CA_Promo"] > 0)
    g["TxCasse"]       = (g["Casse"].abs() / g["CA"] * 100).where(g["CA"] > 0)
    return g

agg_fmt  = agg_base(df_f, "format").sort_values("TxMarge", ascending=False)
agg_rax  = agg_base(df_f, "lib_rayon")
agg_rax["PoidsCA"] = (agg_rax["CA"] / agg_rax["CA"].sum() * 100)
agg_rax  = agg_rax.sort_values("TxMarge")

agg_site = df_f.groupby(["lib_site","format"]).agg(
    CA=("CA","sum"), Marge=("Marge","sum"),
    CA_Promo=("CA HT Promo","sum"), Casse=("Casse (Valeur)","sum")
).reset_index()
agg_site["TxMarge"]  = (agg_site["Marge"] / agg_site["CA"] * 100).where(agg_site["CA"] > 0)
agg_site["PdsPromo"] = (agg_site["CA_Promo"] / agg_site["CA"] * 100).where(agg_site["CA"] > 0)
agg_site["TxCasse"]  = (agg_site["Casse"].abs() / agg_site["CA"] * 100).where(agg_site["CA"] > 0)
agg_site = agg_site[agg_site["CA"] > 0].sort_values("TxMarge", ascending=False).reset_index(drop=True)
moy_marge_site = agg_site["TxMarge"].mean()

mat = df_f.groupby(["lib_site","lib_rayon"]).agg(CA=("CA","sum"), Marge=("Marge","sum")).reset_index()
mat["TxMarge"] = (mat["Marge"] / mat["CA"] * 100).where(mat["CA"] > 0)
mat_pivot = mat.pivot_table(index="lib_rayon", columns="lib_site", values="TxMarge").round(1)

art_site = df_f[df_f["CA"] > 0].groupby(["Article","lib_site","format"]).agg(
    CA=("CA","sum"), Marge=("Marge","sum"), Qte=("Qté Vente","sum")
).reset_index()
art_site["TxMarge_site"] = (art_site["Marge"] / art_site["CA"] * 100).where(art_site["CA"] > 0)
art_site["lib_court"]    = art_site["lib_site"]

# ── agg_art avec code_art ─────────────────────────────────────────────────────
agg_art = df_f.groupby(["Article", "lib_art", "code_art", "lib_rayon", "lib_fam"]).agg(
    CA=("CA","sum"), Marge=("Marge","sum"),
    CA_Promo=("CA HT Promo","sum"), Qte=("Qté Vente","sum")
).reset_index()
agg_art["TxMarge"]  = (agg_art["Marge"] / agg_art["CA"] * 100).where(agg_art["CA"] > 0)
agg_art["PdsPromo"] = (agg_art["CA_Promo"] / agg_art["CA"] * 100).where(agg_art["CA"] > 0)

flop100 = agg_art[agg_art["CA"] > 5000].nsmallest(100, "TxMarge").copy().reset_index(drop=True)
flop100["Rang"] = range(1, len(flop100) + 1)

def build_bloc(article_full, fmt_name):
    rows = art_site[(art_site["Article"] == article_full) & (art_site["format"] == fmt_name)].sort_values("TxMarge_site")
    if rows.empty: return "—"
    parts = []
    for _, r in rows.iterrows():
        tm = r["TxMarge_site"]; qty = r["Qte"]
        if pd.notna(tm):
            parts.append(f"{r['lib_court']}: {tm:.1f}% | Qty: {int(qty):,}" if pd.notna(qty) else f"{r['lib_court']}: {tm:.1f}%")
    return "  |  ".join(parts) if parts else "—"

flop100["Bloc_Hyper"]  = flop100["Article"].apply(lambda a: build_bloc(a, "Hyper"))
flop100["Bloc_Market"] = flop100["Article"].apply(lambda a: build_bloc(a, "Market"))
flop100["Bloc_Supeco"] = flop100["Article"].apply(lambda a: build_bloc(a, "Supeco"))

# KPIs
ca_total    = df_f["CA"].sum()
marge_total = df_f["Marge"].sum()
ca_promo    = df_f["CA HT Promo"].sum()
ca_hp       = df_f["CA Hors Promo"].sum()
m_promo     = df_f["Marge Promo"].sum()
m_hp        = df_f["Marge Hors Promo"].sum()
casse_total = df_f["Casse (Valeur)"].sum()
tx_marge    = marge_total / ca_total * 100  if ca_total > 0 else 0
tx_m_promo  = m_promo / ca_promo * 100      if ca_promo > 0 else 0
tx_m_hp     = m_hp / ca_hp * 100            if ca_hp > 0    else 0
poids_promo = ca_promo / ca_total * 100     if ca_total > 0 else 0
delta_hp_p  = tx_m_hp - tx_m_promo
tx_casse    = abs(casse_total) / ca_total * 100 if ca_total > 0 else 0
nb_art_neg  = int((agg_art["TxMarge"] < 0).sum())
nb_flop_neg = int((flop100["TxMarge"] < 0).sum())

# ─── KPIs GLOBAUX ─────────────────────────────────────────────────────────────
st.markdown(f"<div class='section-label'>{nb_sites_actifs} magasin(s) actifs · {len(sel_rayon)} rayon(s) · {periode}</div>", unsafe_allow_html=True)
k1,k2,k3,k4,k5,k6 = st.columns(6)
k1.metric("CA Réseau",     fmt(ca_total),            "FCFA")
k2.metric("Marge Brute",   fmt(marge_total),          "FCFA")
k3.metric("Taux de Marge", fmt_pct(tx_marge),        f"HP {fmt_pct(tx_m_hp)}")
k4.metric("Effet Promo",   f"−{delta_hp_p:.1f} pts", f"promo {fmt_pct(tx_m_promo)} vs HP {fmt_pct(tx_m_hp)}")
k5.metric("Poids Promo",   fmt_pct(poids_promo),      fmt(ca_promo) + " FCFA")
k6.metric("Casse Réseau",  fmt_pct(tx_casse, dec=2),  fmt(abs(casse_total)) + " FCFA")

# ─── ALERTES ──────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-label'>Signaux critiques réseau</div>", unsafe_allow_html=True)

supeco_row = agg_fmt[agg_fmt["format"] == "Supeco"]
hyper_row  = agg_fmt[agg_fmt["format"] == "Hyper"]
if not supeco_row.empty and not hyper_row.empty:
    tm_sup = supeco_row["TxMarge"].values[0]
    tm_hyp = hyper_row["TxMarge"].values[0]
    if tm_sup < 10:
        st.markdown(f"""
<div class='alert-card alert-purple'>
  <strong>🏪 Format Supeco : taux de marge {tm_sup:.1f}%</strong>
  — écart de {tm_hyp - tm_sup:.1f} pts vs Hypers ({tm_hyp:.1f}%)<br>
  CA concerné : <strong>{fmt(supeco_row['CA'].values[0])} FCFA</strong>
  · Poids promo Supeco : {supeco_row['PdsPromo'].values[0]:.1f}%
</div>""", unsafe_allow_html=True)

if nb_art_neg > 0:
    st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>🔴 {nb_art_neg} article(s) à marge négative</strong> · {nb_flop_neg} dans le Flop 100<br>
  <span style='font-size:12px;opacity:.85'>→ Vérification PA / PV / mécanique promo urgente.</span>
</div>""", unsafe_allow_html=True)

if delta_hp_p > 5:
    st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>⚠️ La promotion dégrade la marge de {delta_hp_p:.1f} pts</strong>
  — HP : {fmt_pct(tx_m_hp)} · Promo : {fmt_pct(tx_m_promo)}
</div>""", unsafe_allow_html=True)

sites_casse = agg_site[agg_site["TxCasse"] > 1].sort_values("TxCasse", ascending=False)
if not sites_casse.empty:
    noms = ", ".join([f"{r['lib_site']} ({r['TxCasse']:.1f}%)" for _, r in sites_casse.iterrows()])
    st.markdown(f"""
<div class='alert-card alert-amber'>
  <strong>🗑️ Taux de casse anormal sur {len(sites_casse)} site(s)</strong> : {noms}
</div>""", unsafe_allow_html=True)

agg_fam = df_f.groupby("lib_fam").agg(CA=("CA","sum"), Marge=("Marge","sum")).reset_index()
agg_fam["TxMarge"] = (agg_fam["Marge"] / agg_fam["CA"] * 100).where(agg_fam["CA"] > 0)
fam_sous_seuil = agg_fam[(agg_fam["TxMarge"] < 8) & (agg_fam["CA"] > 500_000)]
if not fam_sous_seuil.empty:
    noms_fam = ", ".join([f"{r['lib_fam']} ({r['TxMarge']:.1f}%)" for _, r in fam_sous_seuil.iterrows()])
    st.markdown(f"""
<div class='alert-card alert-blue'>
  <strong>📦 {len(fam_sous_seuil)} famille(s) sous 8% de marge</strong> (CA > 500K) : {noms_fam}
</div>""", unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────
st.markdown("---")
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Synthèse Réseau",
    "🔢 Matrice Rayon × Magasin",
    f"💣 Flop {min(100, len(flop100))}",
    "🗑️ Analyse Casse",
    "📥 Export Excel",
])

# ══ TAB 1 — SYNTHÈSE RÉSEAU ═══════════════════════════════════════════════════
with tab1:
    st.markdown("<div class='section-label'>Performance par format de magasin</div>", unsafe_allow_html=True)
    fmt_cols = st.columns(len(agg_fmt))
    fmt_colors = {"Hyper":("#154360","#EFF6FF","#B3D9FF"), "Market":("#145A32","#F0FFF4","#A8E6BF"), "Supeco":("#6E2F8A","#F5F0FF","#D9B3FF")}
    for i, (_, row) in enumerate(agg_fmt.iterrows()):
        fc, bg, border = fmt_colors.get(row["format"], ("#3A3A3C","#F9F9FB","#CCCCCC"))
        with fmt_cols[i]:
            st.markdown(f"""
<div style='background:{bg};border:1px solid {border};border-radius:12px;padding:16px;margin-bottom:8px'>
  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:10px'>
    <span style='font-size:15px;font-weight:700;color:{fc}'>{row["format"]}</span>
    <span style='font-size:11px;color:#8E8E93'>{fmt(row["CA"])} FCFA</span>
  </div>
  <div style='font-size:26px;font-weight:700;color:{fc}'>{fmt_pct(row["TxMarge"])}</div>
  <div style='font-size:11px;color:#8E8E93;margin-top:2px'>Taux de marge</div>
  <hr style='margin:10px 0;border-color:{border}'>
  <div style='display:grid;grid-template-columns:1fr 1fr;gap:6px;font-size:12px'>
    <div><span style='color:#8E8E93'>Promo</span><br><strong>{fmt_pct(row.get("TxMarge_Promo"))}</strong></div>
    <div><span style='color:#8E8E93'>Hors promo</span><br><strong>{fmt_pct(row.get("TxMarge_HP"))}</strong></div>
    <div><span style='color:#8E8E93'>Pds promo</span><br><strong>{fmt_pct(row["PdsPromo"])}</strong></div>
    <div><span style='color:#8E8E93'>Casse</span><br><strong>{fmt_pct(row["TxCasse"], dec=2)}</strong></div>
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Récapitulatif par rayon</div>", unsafe_allow_html=True)
    disp_rax = agg_rax.copy()
    disp_rax["Rayon"]          = disp_rax["lib_rayon"]
    disp_rax["CA (FCFA)"]      = disp_rax["CA"].apply(fmt)
    disp_rax["Poids CA"]       = disp_rax["PoidsCA"].apply(fmt_pct)
    disp_rax["Marge (FCFA)"]   = disp_rax["Marge"].apply(fmt)
    disp_rax["Tx Marge"]       = disp_rax["TxMarge"].apply(fmt_pct)
    disp_rax["Tx Marge HP"]    = disp_rax["TxMarge_HP"].apply(fmt_pct)
    disp_rax["Tx Marge Promo"] = disp_rax["TxMarge_Promo"].apply(fmt_pct)
    disp_rax["Pds Promo"]      = disp_rax["PdsPromo"].apply(fmt_pct)
    disp_rax["Tx Casse"]       = disp_rax["TxCasse"].apply(lambda x: fmt_pct(x, dec=2))
    disp_rax["Écart HP−Promo"] = (disp_rax["TxMarge_HP"] - disp_rax["TxMarge_Promo"]).apply(
        lambda x: fmt_delta(x) if pd.notna(x) else "—")
    st.dataframe(
        disp_rax[["Rayon","CA (FCFA)","Poids CA","Marge (FCFA)","Tx Marge",
                  "Tx Marge HP","Tx Marge Promo","Écart HP−Promo","Pds Promo","Tx Casse"]],
        use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Palmarès magasins</div>", unsafe_allow_html=True)
    disp_site = agg_site.copy()
    disp_site["Rang"]      = range(1, len(disp_site)+1)
    disp_site["Magasin"]   = disp_site["lib_site"]
    disp_site["Format"]    = disp_site["format"]
    disp_site["CA (FCFA)"] = disp_site["CA"].apply(fmt)
    disp_site["Marge"]     = disp_site["Marge"].apply(fmt)
    disp_site["Tx Marge"]  = disp_site["TxMarge"].apply(fmt_pct)
    disp_site["Pds Promo"] = disp_site["PdsPromo"].apply(fmt_pct)
    disp_site["Tx Casse"]  = disp_site["TxCasse"].apply(lambda x: fmt_pct(x, dec=2))
    disp_site["Δ vs Moy."] = (disp_site["TxMarge"] - moy_marge_site).apply(fmt_delta)
    st.dataframe(
        disp_site[["Rang","Magasin","Format","CA (FCFA)","Marge","Tx Marge","Pds Promo","Tx Casse","Δ vs Moy."]],
        use_container_width=True, hide_index=True)

    try:
        import plotly.graph_objects as go
        s = agg_site.sort_values("TxMarge")
        colors_bar = [{"Hyper":"#154360","Market":"#145A32"}.get(r["format"],"#6E2F8A") for _,r in s.iterrows()]
        fig = go.Figure(go.Bar(x=s["TxMarge"].tolist(), y=s["lib_site"].tolist(), orientation="h",
            marker_color=colors_bar, marker_line_width=0,
            text=[f"{v:.1f}%" for v in s["TxMarge"]], textposition="outside"))
        fig.add_vline(x=moy_marge_site, line_width=1.5, line_dash="dot", line_color="#FF9500",
            annotation_text=f" Moy. {moy_marge_site:.1f}%", annotation_font=dict(color="#FF9500", size=10))
        fig.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=11),
            height=max(280, len(agg_site)*40+60), margin=dict(t=10,b=10,l=10,r=70),
            xaxis=dict(title="Taux de marge (%)", ticksuffix="%", showgrid=True, gridcolor="#F2F2F7",
                       range=[0, max(s["TxMarge"])*1.25]),
            yaxis=dict(showgrid=False, title=""))
        st.plotly_chart(fig, use_container_width=True)
        st.caption("🔵 Hyper  ·  🟢 Market  ·  🟣 Supeco  ·  Ligne pointillée = moyenne réseau")
    except ImportError:
        pass

# ══ TAB 2 — MATRICE ══════════════════════════════════════════════════════════
with tab2:
    st.markdown("<div class='section-label'>Taux de marge (%) par combinaison Rayon × Magasin</div>", unsafe_allow_html=True)
    if not mat_pivot.empty:
        mat_display = mat_pivot.copy()
        for col in mat_display.columns:
            mat_display[col] = mat_display[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
        st.dataframe(mat_display, use_container_width=True)
        try:
            import plotly.graph_objects as go
            z = mat_pivot.values.tolist()
            text_z = [[f"{v:.1f}%" if pd.notna(v) else "—" for v in row] for row in z]
            fig_h = go.Figure(go.Heatmap(
                z=z, x=mat_pivot.columns.tolist(), y=mat_pivot.index.tolist(),
                text=text_z, texttemplate="%{text}", textfont=dict(size=12),
                colorscale=[[0,"#C0392B"],[0.25,"#E74C3C"],[0.45,"#F39C12"],
                            [0.60,"#F0E68C"],[0.75,"#A8D5A2"],[1,"#27AE60"]],
                showscale=True, colorbar=dict(title="Tx Marge %", ticksuffix="%", len=0.8),
                zmin=0, zmax=30))
            fig_h.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=11),
                height=max(250, len(mat_pivot.index)*80+80),
                margin=dict(t=20,b=60,l=20,r=20),
                xaxis=dict(tickangle=-35, tickfont=dict(size=10)),
                yaxis=dict(tickfont=dict(size=11)))
            st.plotly_chart(fig_h, use_container_width=True)
        except ImportError:
            pass
    else:
        st.info("Pas de données suffisantes pour construire la matrice.")

# ══ TAB 3 — FLOP 100 ══════════════════════════════════════════════════════════
with tab3:
    st.markdown(f"""
<div class='alert-card alert-red'>
  <strong>💣 {nb_flop_neg} article(s) à marge négative dans le Flop {len(flop100)}</strong>
  · Pertes : <strong>{fmt(flop100[flop100['Marge']<0]['Marge'].sum())} FCFA</strong>
  · CA : <strong>{fmt(flop100['CA'].sum())} FCFA</strong>
</div>""", unsafe_allow_html=True)

    fc1, fc2, fc3 = st.columns(3)
    with fc1: filtre_rayon_f = st.selectbox("Rayon", ["Tous"]+sorted(flop100["lib_rayon"].dropna().unique().tolist()), key="f100_rayon")
    with fc2: filtre_marge   = st.selectbox("Statut marge", ["Tous","Négatif uniquement","< 3%","< 8%"], key="f100_marge")
    with fc3: filtre_promo   = st.selectbox("Promo", ["Tous","100% sous promo","Hors promo uniquement"], key="f100_promo")

    df_flop = flop100.copy()
    if filtre_rayon_f != "Tous": df_flop = df_flop[df_flop["lib_rayon"] == filtre_rayon_f]
    if filtre_marge == "Négatif uniquement": df_flop = df_flop[df_flop["TxMarge"] < 0]
    elif filtre_marge == "< 3%": df_flop = df_flop[df_flop["TxMarge"] < 3]
    elif filtre_marge == "< 8%": df_flop = df_flop[df_flop["TxMarge"] < 8]
    if filtre_promo == "100% sous promo": df_flop = df_flop[df_flop["PdsPromo"] >= 99.9]
    elif filtre_promo == "Hors promo uniquement": df_flop = df_flop[df_flop["PdsPromo"].fillna(0) < 1]

    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>{len(df_flop)} article(s) affichés</div>", unsafe_allow_html=True)

    disp_flop = df_flop.copy()
    disp_flop["#"]           = disp_flop["Rang"]
    disp_flop["Code Art."]   = disp_flop["code_art"]   # ← AJOUT
    disp_flop["Article"]     = disp_flop["lib_art"]
    disp_flop["Rayon"]       = disp_flop["lib_rayon"]
    disp_flop["Famille"]     = disp_flop["lib_fam"]
    disp_flop["CA (FCFA)"]   = disp_flop["CA"].apply(fmt)
    disp_flop["Marge (FCFA)"]= disp_flop["Marge"].apply(fmt)
    disp_flop["Tx Marge"]    = disp_flop["TxMarge"].apply(fmt_pct)
    disp_flop["Pds Promo"]   = disp_flop["PdsPromo"].apply(lambda x: fmt_pct(x) if pd.notna(x) else "—")
    disp_flop["Qté"]         = disp_flop["Qte"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
    disp_flop["🔵 HYPER"]    = disp_flop["Bloc_Hyper"]
    disp_flop["🟢 MARKET"]   = disp_flop["Bloc_Market"]
    disp_flop["🟣 SUPECO"]   = disp_flop["Bloc_Supeco"]

    st.dataframe(
        disp_flop[["#","Code Art.","Article","Rayon","Famille","CA (FCFA)","Marge (FCFA)",
                   "Tx Marge","Pds Promo","Qté","🔵 HYPER","🟢 MARKET","🟣 SUPECO"]],
        use_container_width=True, hide_index=True,
        column_config={
            "#":          st.column_config.NumberColumn("#",         width=40),
            "Code Art.":  st.column_config.TextColumn("Code Art.",   width=100),
            "Article":    st.column_config.TextColumn("Article",     width="large"),
            "Rayon":      st.column_config.TextColumn("Rayon",       width="medium"),
            "Famille":    st.column_config.TextColumn("Famille",     width="medium"),
            "🔵 HYPER":   st.column_config.TextColumn("🔵 HYPER",    width="large"),
            "🟢 MARKET":  st.column_config.TextColumn("🟢 MARKET",   width="large"),
            "🟣 SUPECO":  st.column_config.TextColumn("🟣 SUPECO",   width="large"),
        }
    )
    st.caption("Code Art. = code ERP article · Blocs magasins triés du taux de marge le plus bas")

# ══ TAB 4 — ANALYSE CASSE ═════════════════════════════════════════════════════
with tab4:
    casse_col_v = "Casse (Valeur)"
    casse_col_q = "Casse (Qté)"
    has_qty = casse_col_q in df_f.columns

    agg_casse_site = df_f.groupby(["lib_site","format"]).agg(
        CA=("CA","sum"), Casse_V=(casse_col_v,"sum"),
        **({} if not has_qty else {"Casse_Q": (casse_col_q,"sum")})
    ).reset_index()
    agg_casse_site["TxCasse"] = (agg_casse_site["Casse_V"].abs() / agg_casse_site["CA"] * 100).where(agg_casse_site["CA"] > 0)
    agg_casse_site = agg_casse_site[agg_casse_site["CA"] > 0].sort_values("TxCasse", ascending=False).reset_index(drop=True)

    agg_casse_rax = df_f.groupby("lib_rayon").agg(CA=("CA","sum"), Casse_V=(casse_col_v,"sum")).reset_index()
    agg_casse_rax["TxCasse"] = (agg_casse_rax["Casse_V"].abs() / agg_casse_rax["CA"] * 100).where(agg_casse_rax["CA"] > 0)
    agg_casse_rax = agg_casse_rax.sort_values("TxCasse", ascending=False)

    # ── agg_casse_art avec code_art ───────────────────────────────────────────
    agg_casse_art = df_f.groupby(["Article","lib_art","code_art","lib_rayon","lib_fam"]).agg(
        CA=("CA","sum"), Casse_V=(casse_col_v,"sum"),
        **({} if not has_qty else {"Casse_Q": (casse_col_q,"sum")})
    ).reset_index()
    agg_casse_art["TxCasse"] = (agg_casse_art["Casse_V"].abs() / agg_casse_art["CA"] * 100).where(agg_casse_art["CA"] > 0)
    top30_casse = agg_casse_art[agg_casse_art["Casse_V"].abs() > 0].nlargest(30, "Casse_V").reset_index(drop=True)
    top30_casse["Rang"] = range(1, len(top30_casse)+1)

    casse_reseau      = df_f[casse_col_v].sum()
    tx_casse_reseau   = abs(casse_reseau) / ca_total * 100 if ca_total > 0 else 0
    nb_sites_alerte   = int((agg_casse_site["TxCasse"] > 1).sum())
    moy_tx_casse_site = agg_casse_site["TxCasse"].mean()

    st.markdown("<div class='section-label'>Vue globale casse réseau</div>", unsafe_allow_html=True)
    ck1,ck2,ck3,ck4 = st.columns(4)
    ck1.metric("Casse Réseau",      fmt(abs(casse_reseau))+" FCFA", f"{tx_casse_reseau:.2f}% du CA")
    ck2.metric("Sites > 1% casse",  str(nb_sites_alerte),           f"sur {len(agg_casse_site)} actifs")
    ck3.metric("Tx casse moyen",    fmt_pct(moy_tx_casse_site,dec=2),"moyenne par site")
    ck4.metric("Top rayon casse",
               agg_casse_rax.iloc[0]["lib_rayon"] if not agg_casse_rax.empty else "—",
               fmt_pct(agg_casse_rax.iloc[0]["TxCasse"],dec=2) if not agg_casse_rax.empty else "")

    if nb_sites_alerte > 0:
        noms_alerte = ", ".join([f"{r['lib_site']} ({r['TxCasse']:.1f}%)"
            for _, r in agg_casse_site[agg_casse_site["TxCasse"] > 1].iterrows()])
        st.markdown(f"<div class='alert-card alert-amber'>🗑️ <strong>{nb_sites_alerte} site(s) > 1% casse</strong> : {noms_alerte}</div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("<div class='section-label'>Classement des magasins</div>", unsafe_allow_html=True)
    disp_cs = agg_casse_site.copy()
    disp_cs["Rang"]         = range(1, len(disp_cs)+1)
    disp_cs["Magasin"]      = disp_cs["lib_site"]
    disp_cs["Format"]       = disp_cs["format"]
    disp_cs["CA (FCFA)"]    = disp_cs["CA"].apply(fmt)
    disp_cs["Casse (FCFA)"] = disp_cs["Casse_V"].abs().apply(fmt)
    if has_qty:
        disp_cs["Casse (Qté)"] = disp_cs["Casse_Q"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
    disp_cs["Tx Casse"]     = disp_cs["TxCasse"].apply(lambda x: fmt_pct(x, dec=2))
    disp_cs["Δ vs Moy."]    = (disp_cs["TxCasse"] - moy_tx_casse_site).apply(fmt_delta)
    cols_cs = ["Rang","Magasin","Format","CA (FCFA)","Casse (FCFA)"]
    if has_qty: cols_cs.append("Casse (Qté)")
    cols_cs += ["Tx Casse","Δ vs Moy."]
    st.dataframe(disp_cs[cols_cs], use_container_width=True, hide_index=True)

    try:
        import plotly.graph_objects as go
        sc = agg_casse_site.sort_values("TxCasse")
        bar_colors = [{"Hyper":"#154360","Market":"#145A32"}.get(r["format"],"#6E2F8A") for _,r in sc.iterrows()]
        fig_c = go.Figure(go.Bar(x=sc["TxCasse"].tolist(), y=sc["lib_site"].tolist(), orientation="h",
            marker_color=bar_colors, marker_line_width=0,
            text=[f"{v:.2f}%" for v in sc["TxCasse"]], textposition="outside"))
        fig_c.add_vline(x=1.0, line_width=1.5, line_dash="dot", line_color="#FF3B30",
            annotation_text=" Seuil 1%", annotation_font=dict(color="#FF3B30", size=10))
        fig_c.add_vline(x=moy_tx_casse_site, line_width=1.5, line_dash="dot", line_color="#FF9500",
            annotation_text=f" Moy. {moy_tx_casse_site:.2f}%", annotation_font=dict(color="#FF9500", size=10))
        fig_c.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="-apple-system, Helvetica Neue", color="#3A3A3C", size=11),
            height=max(280, len(agg_casse_site)*40+60), margin=dict(t=10,b=10,l=10,r=80),
            xaxis=dict(title="Taux de casse (%)", ticksuffix="%", showgrid=True, gridcolor="#F2F2F7"),
            yaxis=dict(showgrid=False, title=""))
        st.plotly_chart(fig_c, use_container_width=True)
    except ImportError:
        pass

    st.markdown("---")
    st.markdown("<div class='section-label'>Casse par rayon</div>", unsafe_allow_html=True)
    disp_cr = agg_casse_rax.copy()
    disp_cr["Rayon"]          = disp_cr["lib_rayon"]
    disp_cr["CA (FCFA)"]      = disp_cr["CA"].apply(fmt)
    disp_cr["Casse (FCFA)"]   = disp_cr["Casse_V"].abs().apply(fmt)
    disp_cr["Tx Casse"]       = disp_cr["TxCasse"].apply(lambda x: fmt_pct(x, dec=2))
    disp_cr["% Casse Réseau"] = (disp_cr["Casse_V"].abs() / abs(casse_reseau) * 100).apply(fmt_pct) if casse_reseau != 0 else "—"
    st.dataframe(disp_cr[["Rayon","CA (FCFA)","Casse (FCFA)","Tx Casse","% Casse Réseau"]],
        use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("<div class='section-label'>Top 30 articles — valeur de casse la plus élevée</div>", unsafe_allow_html=True)

    c30f1, c30f2 = st.columns(2)
    with c30f1: filtre_rax_c  = st.selectbox("Rayon", ["Tous"]+sorted(top30_casse["lib_rayon"].dropna().unique().tolist()), key="c_rayon")
    with c30f2: filtre_seuil_c = st.selectbox("Seuil Tx Casse", ["Tous","> 1%","> 2%","> 5%"], key="c_seuil")

    df_top30 = top30_casse.copy()
    if filtre_rax_c  != "Tous": df_top30 = df_top30[df_top30["lib_rayon"] == filtre_rax_c]
    if filtre_seuil_c == "> 1%": df_top30 = df_top30[df_top30["TxCasse"] > 1]
    elif filtre_seuil_c == "> 2%": df_top30 = df_top30[df_top30["TxCasse"] > 2]
    elif filtre_seuil_c == "> 5%": df_top30 = df_top30[df_top30["TxCasse"] > 5]

    disp_t30 = df_top30.copy()
    disp_t30["#"]            = disp_t30["Rang"]
    disp_t30["Code Art."]    = disp_t30["code_art"]   # ← AJOUT
    disp_t30["Article"]      = disp_t30["lib_art"]
    disp_t30["Rayon"]        = disp_t30["lib_rayon"]
    disp_t30["Famille"]      = disp_t30["lib_fam"]
    disp_t30["CA (FCFA)"]    = disp_t30["CA"].apply(fmt)
    disp_t30["Casse (FCFA)"] = disp_t30["Casse_V"].abs().apply(fmt)
    if has_qty:
        disp_t30["Casse (Qté)"] = disp_t30["Casse_Q"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
    disp_t30["Tx Casse"]     = disp_t30["TxCasse"].apply(lambda x: fmt_pct(x, dec=2))

    cols_t30 = ["#","Code Art.","Article","Rayon","Famille","CA (FCFA)","Casse (FCFA)"]
    if has_qty: cols_t30.append("Casse (Qté)")
    cols_t30.append("Tx Casse")

    st.markdown(f"<div style='font-size:12px;color:#8E8E93;margin-bottom:8px'>{len(disp_t30)} article(s)</div>", unsafe_allow_html=True)
    st.dataframe(disp_t30[cols_t30], use_container_width=True, hide_index=True,
        column_config={
            "#":         st.column_config.NumberColumn("#",        width=40),
            "Code Art.": st.column_config.TextColumn("Code Art.",  width=100),
            "Article":   st.column_config.TextColumn("Article",    width="large"),
        })

# ══ TAB 5 — EXPORT EXCEL ══════════════════════════════════════════════════════
with tab5:
    st.markdown("""
<div class='alert-card alert-blue'>
  <strong>📋 Contenu de l'export (5 onglets)</strong><br>
  <strong>Synthèse Réseau</strong> · <strong>Récap Rayon</strong> · <strong>Matrice Marge</strong> ·
  <strong>Flop 100</strong> (avec Code Art.) · <strong>Analyse Casse</strong> (avec Code Art.)
</div>""", unsafe_allow_html=True)
    st.caption(f"Périmètre : {len(sel_site)} magasin(s) · {len(sel_rayon)} rayon(s) · {periode}")

    if st.button("Générer le fichier Excel", type="primary", key="gen_excel"):
        with st.spinner("Génération du rapport…"):
            wb_exp = Workbook()
            C_HDR="1B2A4A"; C_SUB="2E4B7A"; C_WH="FFFFFF"; C_DK="1A1A2E"
            C_HYP="154360"; C_MKT="145A32"; C_SUP="6E2F8A"

            def xfill(h): return PatternFill("solid", fgColor=h)
            def xbdr():
                s = Side(style="thin", color="CCCCCC")
                return Border(left=s, right=s, top=s, bottom=s)
            def xctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
            def xrgt(): return Alignment(horizontal="right",  vertical="center")
            def xlft(w=False): return Alignment(horizontal="left", vertical="center", wrap_text=w)

            def write_header_row(ws, row_num, headers, widths, bg=C_SUB):
                for i,(h,w) in enumerate(zip(headers,widths)):
                    c = ws.cell(row=row_num, column=i+1, value=h)
                    c.font=Font("Calibri",size=10,bold=True,color=C_WH)
                    c.fill=xfill(bg); c.alignment=xctr(); c.border=xbdr()
                    ws.column_dimensions[get_column_letter(i+1)].width = w
                ws.row_dimensions[row_num].height = 24

            def title_block(ws, txt, span=10):
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
                c = ws.cell(row=1,column=1,value=txt)
                c.font=Font("Calibri",size=13,bold=True,color=C_WH); c.fill=xfill(C_HDR); c.alignment=xctr()
                ws.row_dimensions[1].height=30
                ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=span)
                c2=ws.cell(row=2,column=1,value=f"  Période : {periode}  ·  {nb_jours} jour(s)")
                c2.font=Font("Calibri",size=9,italic=True,color="AABBCC"); c2.fill=xfill(C_HDR); c2.alignment=xlft()
                ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=6

            # ── Onglet 1 : Synthèse
            ws1 = wb_exp.active; ws1.title = "Synthèse Réseau"
            title_block(ws1, "DIAGNOSTIC RENTABILITÉ RÉSEAU — SYNTHÈSE", span=8)
            r=4
            kpi_data = [
                ("CA Réseau (FCFA)",f"{ca_total:,.0f}"),("Marge Brute (FCFA)",f"{marge_total:,.0f}"),
                ("Taux de Marge",f"{tx_marge:.1f}%"),("Taux Marge HP",f"{tx_m_hp:.1f}%"),
                ("Taux Marge Promo",f"{tx_m_promo:.1f}%"),("Effet Promo (pts)",f"−{delta_hp_p:.1f}"),
                ("Poids Promo",f"{poids_promo:.1f}%"),("Taux Casse",f"{tx_casse:.2f}%"),
            ]
            for ci,(lbl,val) in enumerate(kpi_data):
                c=ws1.cell(row=r,column=ci+1,value=lbl); c.font=Font("Calibri",size=9,bold=True,color=C_WH)
                c.fill=xfill(C_SUB); c.alignment=xctr(); c.border=xbdr()
                c2=ws1.cell(row=r+1,column=ci+1,value=val); c2.font=Font("Calibri",size=12,bold=True,color=C_DK)
                c2.fill=xfill("FFFFFF"); c2.alignment=xctr(); c2.border=xbdr()
                ws1.column_dimensions[get_column_letter(ci+1)].width=18
            ws1.row_dimensions[r].height=20; ws1.row_dimensions[r+1].height=28; r+=3

            ws1.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            c=ws1.cell(row=r,column=1,value="  PERFORMANCE PAR FORMAT")
            c.font=Font("Calibri",size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft(); ws1.row_dimensions[r].height=22; r+=1
            write_header_row(ws1,r,["Format","CA (FCFA)","Marge (FCFA)","Tx Marge","Tx Marge HP","Tx Marge Promo","Pds Promo","Tx Casse"],[14,18,18,12,14,16,12,12]); r+=1
            for _,(fd) in agg_fmt.iterrows():
                bg_f={"Hyper":"D6EAF8","Market":"D5F5E3","Supeco":"E8DAEF"}.get(fd["format"],"FFFFFF")
                for ci,(v,fmt3) in enumerate([(fd["format"],None),(fd["CA"],"#,##0"),(fd["Marge"],"#,##0"),
                    (fd["TxMarge"]/100 if pd.notna(fd["TxMarge"]) else None,"0.0%"),
                    (fd.get("TxMarge_HP",None)/100 if pd.notna(fd.get("TxMarge_HP")) else None,"0.0%"),
                    (fd.get("TxMarge_Promo",None)/100 if pd.notna(fd.get("TxMarge_Promo")) else None,"0.0%"),
                    (fd["PdsPromo"]/100 if pd.notna(fd["PdsPromo"]) else None,"0.0%"),
                    (fd["TxCasse"]/100 if pd.notna(fd["TxCasse"]) else None,"0.0%")]):
                    c=ws1.cell(row=r,column=ci+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg_f); c.border=xbdr()
                    if fmt3: c.number_format=fmt3
                    c.alignment=xrgt() if ci in [1,2] else xctr()
                ws1.row_dimensions[r].height=20; r+=1

            r+=1
            ws1.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            c=ws1.cell(row=r,column=1,value="  PALMARÈS MAGASINS")
            c.font=Font("Calibri",size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft(); ws1.row_dimensions[r].height=22; r+=1
            write_header_row(ws1,r,["Rang","Magasin","Format","CA (FCFA)","Marge (FCFA)","Tx Marge","Pds Promo","Tx Casse"],[5,24,10,18,18,12,12,12]); r+=1
            for ri3,(_,sd) in enumerate(agg_site.iterrows()):
                bg_s="F7F7F7" if ri3%2==0 else "FFFFFF"
                for ci3,(v,f3) in enumerate([(ri3+1,None),(sd["lib_site"],None),(sd["format"],None),
                    (sd["CA"],"#,##0"),(sd["Marge"],"#,##0"),
                    (sd["TxMarge"]/100 if pd.notna(sd["TxMarge"]) else None,"0.0%"),
                    (sd["PdsPromo"]/100 if pd.notna(sd["PdsPromo"]) else None,"0.0%"),
                    (sd["TxCasse"]/100 if pd.notna(sd["TxCasse"]) else None,"0.0%")]):
                    c=ws1.cell(row=r,column=ci3+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg_s); c.border=xbdr()
                    if f3: c.number_format=f3
                    c.alignment=xctr() if ci3==0 else xrgt() if ci3 in [3,4] else xctr()
                ws1.row_dimensions[r].height=20; r+=1
            ws1.freeze_panes="A4"

            # ── Onglet 2 : Récap rayon
            ws2=wb_exp.create_sheet("Récap Rayon")
            title_block(ws2,"RÉCAPITULATIF PAR RAYON",span=10)
            write_header_row(ws2,4,["Rayon","CA (FCFA)","Poids CA","Marge (FCFA)","Tx Marge","Tx Marge HP","Tx Marge Promo","Écart HP−Promo","Pds Promo","Tx Casse"],[22,16,10,16,12,14,16,16,12,12])
            for ri4,(_,rd4) in enumerate(agg_rax.iterrows()):
                r4=ri4+5; bg4="F7F7F7" if ri4%2==0 else "FFFFFF"
                ecart4=rd4["TxMarge_HP"]-rd4["TxMarge_Promo"] if pd.notna(rd4.get("TxMarge_Promo")) else None
                for ci4,(v,f4) in enumerate([(rd4["lib_rayon"],None),(rd4["CA"],"#,##0"),
                    (rd4["PoidsCA"]/100,"0.0%"),(rd4["Marge"],"#,##0"),
                    (rd4["TxMarge"]/100 if pd.notna(rd4["TxMarge"]) else None,"0.0%"),
                    (rd4["TxMarge_HP"]/100 if pd.notna(rd4.get("TxMarge_HP")) else None,"0.0%"),
                    (rd4["TxMarge_Promo"]/100 if pd.notna(rd4.get("TxMarge_Promo")) else None,"0.0%"),
                    (ecart4/100 if ecart4 is not None and pd.notna(ecart4) else None,"0.0%"),
                    (rd4["PdsPromo"]/100 if pd.notna(rd4["PdsPromo"]) else None,"0.0%"),
                    (rd4["TxCasse"]/100 if pd.notna(rd4["TxCasse"]) else None,"0.0%")]):
                    c=ws2.cell(row=r4,column=ci4+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg4); c.border=xbdr()
                    if f4: c.number_format=f4
                    c.alignment=xrgt() if ci4 in [1,3] else xctr()
                ws2.row_dimensions[r4].height=20
            ws2.freeze_panes="A5"

            # ── Onglet 3 : Matrice
            ws3=wb_exp.create_sheet("Matrice Marge")
            title_block(ws3,"MATRICE TAUX DE MARGE — RAYON × MAGASIN",span=len(mat_pivot.columns)+2)
            c=ws3.cell(row=4,column=1,value="Rayon"); c.font=Font("Calibri",size=10,bold=True,color=C_WH)
            c.fill=xfill(C_SUB); c.alignment=xctr(); c.border=xbdr(); ws3.column_dimensions["A"].width=22
            for ci5,col_name in enumerate(mat_pivot.columns):
                c=ws3.cell(row=4,column=ci5+2,value=short_name(col_name))
                c.font=Font("Calibri",size=9,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xctr(); c.border=xbdr()
                ws3.column_dimensions[get_column_letter(ci5+2)].width=16
            ws3.row_dimensions[4].height=36
            for ri5,rayon_idx in enumerate(mat_pivot.index):
                r5=ri5+5
                c0=ws3.cell(row=r5,column=1,value=rayon_idx); c0.font=Font("Calibri",size=10,bold=True,color=C_DK)
                c0.fill=xfill("F7F7F7"); c0.alignment=xlft(); c0.border=xbdr()
                for ci5,col_name in enumerate(mat_pivot.columns):
                    v5=mat_pivot.loc[rayon_idx,col_name]
                    c=ws3.cell(row=r5,column=ci5+2)
                    if pd.notna(v5): c.value=v5/100; c.number_format="0.0%"
                    c.font=Font("Calibri",size=11,bold=True,color=C_DK); c.alignment=xctr(); c.border=xbdr(); c.fill=xfill("FFFFFF")
                ws3.row_dimensions[r5].height=28
            ws3.freeze_panes="B5"

            # ── Onglet 4 : Flop 100 avec Code Art.
            ws4=wb_exp.create_sheet("Flop 100")
            title_block(ws4,f"FLOP {len(flop100)} — DESTRUCTEURS DE MARGE",span=13)
            hdrs4=["#","Code Art.","Article","Rayon","Famille","CA (FCFA)","Marge (FCFA)","Tx Marge","Pds Promo","Qté",
                   "🔵 HYPER","🟢 MARKET","🟣 SUPECO"]
            wdths4=[5,12,38,16,24,13,13,10,10,8,46,50,54]
            bloc_bg={10:C_HYP,11:C_MKT,12:C_SUP}
            for ci6,(h,w) in enumerate(zip(hdrs4,wdths4)):
                bg6=bloc_bg.get(ci6,C_SUB)
                c=ws4.cell(row=4,column=ci6+1,value=h); c.font=Font("Calibri",size=9,bold=True,color=C_WH)
                c.fill=xfill(bg6); c.alignment=xctr(); c.border=xbdr()
                ws4.column_dimensions[get_column_letter(ci6+1)].width=w
            ws4.row_dimensions[4].height=28
            bloc_fill={10:"D6EAF8",11:"D5F5E3",12:"E8DAEF"}
            for ri6,(_,rd6) in enumerate(flop100.iterrows()):
                r6=ri6+5; bg6="F7F7F7" if ri6%2==0 else "FFFFFF"
                tm6=rd6["TxMarge"]; pp6=rd6["PdsPromo"]
                vals6=[rd6["Rang"],rd6["code_art"],rd6["lib_art"],rd6["lib_rayon"],rd6["lib_fam"],
                       rd6["CA"],rd6["Marge"],
                       tm6/100 if pd.notna(tm6) else None,
                       pp6/100 if pd.notna(pp6) else None,
                       int(rd6["Qte"]) if pd.notna(rd6["Qte"]) else None,
                       rd6["Bloc_Hyper"],rd6["Bloc_Market"],rd6["Bloc_Supeco"]]
                fmts6=[None,None,None,None,None,"#,##0","#,##0","0.0%","0.0%","#,##0",None,None,None]
                for ci6,(v,f6) in enumerate(zip(vals6,fmts6)):
                    c=ws4.cell(row=r6,column=ci6+1,value=v)
                    cell_bg=bloc_fill.get(ci6,bg6) if (ci6>=10 and v and v!="—") else bg6
                    c.font=Font("Calibri",size=10 if ci6<10 else 9,color=C_DK)
                    c.fill=xfill(cell_bg); c.border=xbdr()
                    if f6: c.number_format=f6
                    if ci6==0: c.font=Font("Calibri",size=10,bold=True,color=C_DK); c.alignment=xctr()
                    elif ci6 in [5,6]: c.alignment=xrgt()
                    elif ci6 in [7,8,9]: c.alignment=xctr()
                    elif ci6>=10: c.alignment=xlft(w=True)
                    else: c.alignment=xlft(w=(ci6 in [2,4]))
                ws4.row_dimensions[r6].height=30
            ws4.freeze_panes="A5"

            # ── Onglet 5 : Analyse Casse avec Code Art.
            ws5=wb_exp.create_sheet("Analyse Casse")
            title_block(ws5,"ANALYSE CASSE RÉSEAU",span=9)
            kpi_casse=[("Casse Réseau (FCFA)",f"{abs(casse_reseau):,.0f}"),
                       ("Tx Casse Réseau",f"{tx_casse_reseau:.2f}%"),
                       ("Sites > 1% casse",str(nb_sites_alerte)),
                       ("Tx Casse Moyen",f"{moy_tx_casse_site:.2f}%")]
            for ci_k,(lbl,val) in enumerate(kpi_casse):
                ck=ws5.cell(row=4,column=ci_k+1,value=lbl); ck.font=Font("Calibri",size=9,bold=True,color=C_WH)
                ck.fill=xfill(C_SUB); ck.alignment=xctr(); ck.border=xbdr()
                ck2=ws5.cell(row=5,column=ci_k+1,value=val); ck2.font=Font("Calibri",size=12,bold=True,color=C_DK)
                ck2.fill=xfill("FFFFFF"); ck2.alignment=xctr(); ck2.border=xbdr()
                ws5.column_dimensions[get_column_letter(ci_k+1)].width=22
            ws5.row_dimensions[4].height=20; ws5.row_dimensions[5].height=28

            r5s=7
            ws5.merge_cells(start_row=r5s,start_column=1,end_row=r5s,end_column=7)
            c=ws5.cell(row=r5s,column=1,value="  CLASSEMENT MAGASINS")
            c.font=Font("Calibri",size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft(); ws5.row_dimensions[r5s].height=22; r5s+=1
            hdrs_cs=["Rang","Magasin","Format","CA (FCFA)","Casse (FCFA)","Casse (Qté)","Tx Casse"]
            write_header_row(ws5,r5s,hdrs_cs,[5,28,10,16,16,14,12]); r5s+=1
            for ri_s,(_,sd) in enumerate(agg_casse_site.iterrows()):
                bg_s="FFF8F0" if sd["TxCasse"]>1 else ("F7F7F7" if ri_s%2==0 else "FFFFFF")
                casse_q_val=int(sd["Casse_Q"]) if has_qty and pd.notna(sd.get("Casse_Q")) else None
                for ci_s,(v,f_s) in enumerate([(ri_s+1,None),(sd["lib_site"],None),(sd["format"],None),
                    (sd["CA"],"#,##0"),(abs(sd["Casse_V"]),"#,##0"),(casse_q_val,"#,##0"),
                    (sd["TxCasse"]/100 if pd.notna(sd["TxCasse"]) else None,"0.00%")]):
                    c=ws5.cell(row=r5s,column=ci_s+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg_s); c.border=xbdr()
                    if f_s: c.number_format=f_s
                    c.alignment=xctr() if ci_s==0 else xrgt() if ci_s in [3,4,5] else xctr()
                ws5.row_dimensions[r5s].height=20; r5s+=1

            r5s+=1
            ws5.merge_cells(start_row=r5s,start_column=1,end_row=r5s,end_column=5)
            c=ws5.cell(row=r5s,column=1,value="  CASSE PAR RAYON")
            c.font=Font("Calibri",size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft(); ws5.row_dimensions[r5s].height=22; r5s+=1
            write_header_row(ws5,r5s,["Rayon","CA (FCFA)","Casse (FCFA)","Tx Casse","% Casse Réseau"],[22,16,16,12,16]); r5s+=1
            for ri_r,(_,rd) in enumerate(agg_casse_rax.iterrows()):
                bg_r="F7F7F7" if ri_r%2==0 else "FFFFFF"
                pct_res=abs(rd["Casse_V"])/abs(casse_reseau) if casse_reseau!=0 else 0
                for ci_r,(v,f_r) in enumerate([(rd["lib_rayon"],None),(rd["CA"],"#,##0"),
                    (abs(rd["Casse_V"]),"#,##0"),
                    (rd["TxCasse"]/100 if pd.notna(rd["TxCasse"]) else None,"0.00%"),(pct_res,"0.0%")]):
                    c=ws5.cell(row=r5s,column=ci_r+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg_r); c.border=xbdr()
                    if f_r: c.number_format=f_r
                    c.alignment=xlft() if ci_r==0 else xrgt() if ci_r in [1,2] else xctr()
                ws5.row_dimensions[r5s].height=20; r5s+=1

            r5s+=1
            ws5.merge_cells(start_row=r5s,start_column=1,end_row=r5s,end_column=9)
            c=ws5.cell(row=r5s,column=1,value="  TOP 30 ARTICLES — valeur de casse la plus élevée")
            c.font=Font("Calibri",size=10,bold=True,color=C_WH); c.fill=xfill(C_SUB); c.alignment=xlft(); ws5.row_dimensions[r5s].height=22; r5s+=1
            hdrs_a=["#","Code Art.","Article","Rayon","Famille","CA (FCFA)","Casse (FCFA)","Casse (Qté)","Tx Casse"]
            wdths_a=[5,12,38,18,24,16,16,14,12]
            write_header_row(ws5,r5s,hdrs_a,wdths_a); r5s+=1
            for ri_a,(_,ra) in enumerate(top30_casse.iterrows()):
                bg_a="F7F7F7" if ri_a%2==0 else "FFFFFF"
                casse_q_a=int(ra["Casse_Q"]) if has_qty and pd.notna(ra.get("Casse_Q")) else None
                for ci_a,(v,f_a) in enumerate([(ra["Rang"],None),(ra["code_art"],None),(ra["lib_art"],None),
                    (ra["lib_rayon"],None),(ra["lib_fam"],None),
                    (ra["CA"],"#,##0"),(abs(ra["Casse_V"]),"#,##0"),(casse_q_a,"#,##0"),
                    (ra["TxCasse"]/100 if pd.notna(ra["TxCasse"]) else None,"0.00%")]):
                    c=ws5.cell(row=r5s,column=ci_a+1,value=v); c.font=Font("Calibri",size=10,color=C_DK)
                    c.fill=xfill(bg_a); c.border=xbdr()
                    if f_a: c.number_format=f_a
                    if ci_a==0: c.alignment=xctr()
                    elif ci_a in [5,6,7]: c.alignment=xrgt()
                    elif ci_a==8: c.alignment=xctr()
                    else: c.alignment=xlft(w=(ci_a in [2,4]))
                ws5.row_dimensions[r5s].height=20; r5s+=1
            ws5.freeze_panes="A4"

            buf=BytesIO(); wb_exp.save(buf); buf.seek(0)

        st.download_button(
            label="⬇️ Télécharger le rapport Excel",
            data=buf,
            file_name=f"SmartBuyer_Diagnostic_Reseau_{periode.replace('/','').replace(' ','_').replace('→','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
