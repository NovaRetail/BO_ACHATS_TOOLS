"""
08_📦_Ruptures_Stock.py — SmartBuyer Hub
Détection de ruptures · Plan d'action Commander / Cession
Charte SmartBuyer v2 — Style Apple / SF Pro
"""

import io
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Ruptures de Stock · SmartBuyer",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CHARTE SMARTBUYER ────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "SF Pro Text", "Helvetica Neue", Arial, sans-serif !important;
    background-color: #F2F2F7;
}
.stApp { background: #F2F2F7; }
.main .block-container { padding-top: 1.8rem; max-width: 1280px; }

[data-testid="stSidebar"] {
    background: #F2F2F7 !important;
    border-right: 0.5px solid #D1D1D6 !important;
}
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 0.5px solid #E5E5EA !important;
    border-radius: 12px !important;
    padding: 16px 18px !important;
}
[data-testid="stMetricLabel"] {
    font-size: 11px !important; font-weight: 500 !important;
    color: #8E8E93 !important; text-transform: uppercase !important;
    letter-spacing: 0.04em !important;
}
[data-testid="stMetricValue"] {
    font-size: 24px !important; font-weight: 600 !important;
    color: #1C1C1E !important; letter-spacing: -0.02em !important;
}
[data-testid="stTabs"] button[role="tab"] {
    font-size: 13px !important; font-weight: 500 !important;
    padding: 8px 16px !important; color: #8E8E93 !important;
    border-radius: 0 !important; border-bottom: 2px solid transparent !important;
}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    color: #007AFF !important; border-bottom: 2px solid #007AFF !important;
    background: transparent !important;
}
[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 0.5px solid #E5E5EA !important;
}
[data-testid="stDataFrame"] {
    border: 0.5px solid #E5E5EA !important; border-radius: 10px !important;
}
[data-testid="stDataFrame"] th {
    background: #F2F2F7 !important; font-size: 11px !important;
    font-weight: 600 !important; color: #8E8E93 !important;
    text-transform: uppercase !important; letter-spacing: 0.04em !important;
}
[data-testid="stFileUploader"] {
    border: 1.5px dashed #D1D1D6 !important;
    border-radius: 10px !important; background: #F9F9FB !important;
}
.stDownloadButton > button {
    background: #007AFF !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 500 !important; font-size: 13px !important;
    padding: 10px 24px !important; width: 100% !important;
}
hr { border-color: #E5E5EA !important; margin: 1rem 0 !important; }

.page-title   { font-size: 28px; font-weight: 700; color: #1C1C1E;
                letter-spacing: -0.03em; margin: 0; }
.page-caption { font-size: 13px; color: #8E8E93; margin-top: 3px;
                margin-bottom: 1.5rem; }
.section-label { font-size: 11px; font-weight: 600; color: #8E8E93;
                 text-transform: uppercase; letter-spacing: 0.07em;
                 margin-bottom: 10px; }
.kpi-card { background: #FFFFFF; border: 0.5px solid #E5E5EA;
            border-radius: 12px; padding: 16px 18px; }

.alert-box { border-radius: 10px; padding: 12px 16px; margin-bottom: 10px;
             font-size: 13px; line-height: 1.5; border-left: 3px solid; }
.alert-blue   { background: #EFF6FF; border-color: #007AFF; color: #1E3A5F; }
.alert-red    { background: #FFF2F2; border-color: #FF3B30; color: #7B0000; }
.alert-green  { background: #F0FFF4; border-color: #34C759; color: #1A3A20; }
.alert-orange { background: #FFF8EC; border-color: #FF9500; color: #7A4500; }

.badge { display: inline-block; padding: 3px 10px; border-radius: 100px;
         font-size: 11px; font-weight: 600; letter-spacing: 0.02em; }
.badge-cmd  { background: #FDECEA; color: #C0392B; }
.badge-cess { background: #E8F8F0; color: #196F3D; }

.col-required { background: #F0F8FF; border: 0.5px solid #B3D9FF;
                border-radius: 8px; padding: 10px 14px; margin-bottom: 6px;
                display: flex; align-items: flex-start; gap: 10px; }
.col-name    { font-size: 13px; font-weight: 600; color: #0066CC;
               font-family: monospace; }
.col-desc    { font-size: 12px; color: #3A3A3C; margin-top: 1px; }
.col-example { font-size: 11px; color: #8E8E93; font-family: monospace;
               margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
PL_BG   = "rgba(0,0,0,0)"
PL_FONT = "#3A3A3C"
PL_GRID = "#E5E5EA"
COLOR_CMD  = "#FF3B30"
COLOR_CESS = "#34C759"
COLOR_WARN = "#FF9500"
COLOR_INFO = "#007AFF"

# ─── PARSING DYNAMIQUE ────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_data(file_bytes: bytes):
    """
    Parsing dynamique : détecte automatiquement les colonnes magasins
    via le pattern 'CODE - NOM' dans la ligne 0 de l'export.
    Compatible avec tous les rayons (DPH, Épicerie, Boissons…).
    """
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0,
                           header=None, engine='openpyxl')

    row0 = df_raw.iloc[0].tolist()

    # Détecter colonnes magasins (pattern: chiffres - texte) et Total
    mag_labels = []   # "10202 - Palmeraie"
    mag_names  = []   # "Palmeraie"
    mag_idx    = []   # indices colonnes
    total_idx  = None

    for i, val in enumerate(row0):
        if not isinstance(val, str):
            continue
        val = val.strip()
        if val == 'Total':
            total_idx = i
        elif ' - ' in val:
            code, nom = val.split(' - ', 1)
            if code.strip().isdigit():
                mag_labels.append(val)
                mag_names.append(nom.strip())
                mag_idx.append(i)

    if not mag_names:
        raise ValueError(
            "Impossible de détecter les colonnes magasins. "
            "Vérifiez que la ligne d'en-tête contient des codes au format 'CODE - NOM'."
        )

    # Construire le DataFrame articles
    data = df_raw.iloc[2:].copy()
    base_cols = ['Rayon', 'Famille', 'Sous_Famille', 'Article']
    col_names = base_cols + mag_names
    if total_idx is not None:
        col_names += ['Total']
        usecols = [0, 1, 2, 3] + mag_idx + [total_idx]
    else:
        usecols = [0, 1, 2, 3] + mag_idx

    data = data.iloc[:, usecols].copy()
    data.columns = col_names

    # Garder uniquement les lignes articles
    arts = data[
        data['Article'].notna() &
        data['Article'].astype(str).str.contains(' - ', regex=False)
    ].copy().reset_index(drop=True)

    # Conversion numérique stocks
    for col in mag_names:
        arts[col] = pd.to_numeric(arts[col], errors='coerce')

    # Labels courts
    for col in ['Rayon', 'Famille', 'Sous_Famille']:
        arts[col] = arts[col].astype(str).apply(
            lambda x: x.split(' - ', 1)[-1].strip() if ' - ' in x else x
        )
    arts['Code_Article'] = (arts['Article'].astype(str)
                            .str.split(' - ').str[0].str.strip())
    arts['Libelle']      = arts['Article'].astype(str).apply(
        lambda x: ' - '.join(x.split(' - ')[1:]).strip()
    )
    return arts, mag_names


@st.cache_data(show_spinner=False)
def compute_plan(file_bytes: bytes, seuil: int):
    arts, mag_names = load_data(file_bytes)

    def is_rupt(v):
        return pd.isna(v) or v <= 0

    rows = []
    for _, r in arts.iterrows():
        rups   = [m for m in mag_names if is_rupt(r[m])]
        dispos = {m: int(r[m]) for m in mag_names
                  if not is_rupt(r[m]) and r[m] > seuil}
        if not rups:
            continue
        if dispos:
            max_mag   = max(dispos, key=dispos.get)
            action    = 'VOIR CESSION'
            site_ref  = max_mag
            stock_ref = dispos[max_mag]
        else:
            action    = 'COMMANDER'
            site_ref  = '—'
            stock_ref = 0
        rows.append({
            'Rayon':            r['Rayon'],
            'Famille':          r['Famille'],
            'Sous_Famille':     r['Sous_Famille'],
            'Code_Article':     r['Code_Article'],
            'Libellé':          r['Libelle'],
            'Nb_Ruptures':      len(rups),
            'Magasins_Rupture': ', '.join(rups),
            'Action':           action,
            'Site_Donneur':     site_ref,
            'Stock_Disponible': stock_ref if action == 'VOIR CESSION' else None,
        })
    plan = pd.DataFrame(rows).sort_values(
        ['Nb_Ruptures', 'Action'], ascending=[False, True]
    ).reset_index(drop=True)
    return plan, arts, mag_names


@st.cache_data(show_spinner=False)
def to_excel_bytes(plan: pd.DataFrame, seuil: int) -> bytes:
    def bd():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def fill(h):
        return PatternFill('solid', fgColor=h)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Plan Action'

    ws.merge_cells('A1:I1')
    ws['A1'] = f"PLAN D'ACTION RUPTURES — Seuil cession : {seuil} unités"
    ws['A1'].font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = fill('1C3557')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    hdrs   = ['Rayon', 'Famille', 'Code Article', 'Libellé', 'Nb Rupt.',
              'Magasins en Rupture', 'Action', 'Site Donneur', 'Stock Dispo']
    widths = [18, 26, 12, 44, 9, 56, 16, 18, 11]
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font      = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        c.fill      = fill('2C4A6E')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        c.border    = bd()
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[3].height = 22

    LGREY = 'F2F2F7'
    for i, row in plan.iterrows():
        r      = i + 4
        is_cmd = row['Action'] == 'COMMANDER'
        bg     = (LGREY if i % 2 == 0 else 'FFFFFF')
        stock_val = (int(row['Stock_Disponible'])
                     if pd.notna(row['Stock_Disponible']) and row['Stock_Disponible']
                     else '')
        vals = [row['Rayon'], row['Famille'], row['Code_Article'], row['Libellé'],
                row['Nb_Ruptures'], row['Magasins_Rupture'],
                '🔴 COMMANDER' if is_cmd else '🟢 VOIR CESSION',
                row['Site_Donneur'], stock_val]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(
                name='Calibri', size=9, bold=(j == 7),
                color=('C0392B' if (j == 7 and is_cmd) else
                       '196F3D' if (j in [7, 8, 9] and not is_cmd) else '333333')
            )
            c.fill = fill(
                'FDECEA' if (j == 7 and is_cmd) else
                'EAF7EC' if (j == 7 and not is_cmd) else bg
            )
            c.alignment = Alignment(
                horizontal='center' if j in [3, 5, 7, 8, 9] else 'left',
                vertical='center', wrap_text=(j == 6)
            )
            c.border = bd()
        ws.row_dimensions[r].height = 15

    ws.freeze_panes   = 'A4'
    ws.auto_filter.ref = f'A3:I{len(plan) + 3}'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fmt(n):
    return f"{int(n):,}".replace(",", " ")


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='margin-bottom:18px'>
      <div style='font-size:20px;font-weight:700;color:#1C1C1E;
                  letter-spacing:-0.02em'>🛍️ SmartBuyer</div>
      <div style='font-size:11px;color:#8E8E93;margin-top:1px'>
        Hub analytique · Équipe Achats</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown(
        "<div style='font-size:11px;font-weight:600;color:#8E8E93;"
        "text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>"
        "Navigation</div>", unsafe_allow_html=True
    )
    st.page_link("app.py",                                       label="🏠  Accueil")
    st.page_link("pages/01_📊_Analyse_Scoring_ABC.py",           label="📊  Scoring ABC")
    st.page_link("pages/02_📈_Ventes_PBI.py",                    label="📈  Ventes PBI",         disabled=True)
    st.page_link("pages/03_📦_Detention_Top_CA.py",              label="📦  Détention Top CA",   disabled=True)
    st.page_link("pages/04_💸_Performance_Promo.py",             label="💸  Performance Promo",  disabled=True)
    st.page_link("pages/05_🏪_Suivi_Implantation.py",            label="🏪  Suivi Implantation", disabled=True)
    st.page_link("pages/08_📦_OOS.py",                           label="📦  Ruptures Stock")
    st.markdown("---")

    st.markdown(
        "<div style='font-size:11px;font-weight:600;color:#8E8E93;"
        "text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>"
        "Fichier source</div>", unsafe_allow_html=True
    )
    uploaded = st.file_uploader(
        "Export stock (.xlsx)", type=["xlsx"], label_visibility="collapsed"
    )
    st.markdown("---")

    st.markdown(
        "<div style='font-size:11px;font-weight:600;color:#8E8E93;"
        "text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>"
        "Paramètres</div>", unsafe_allow_html=True
    )
    seuil = st.slider(
        "Seuil minimum cession (unités)", min_value=0, max_value=500,
        value=100, step=10,
        help="Le site donneur doit avoir STRICTEMENT PLUS que ce stock pour qu'une cession soit proposée"
    )
    st.markdown(
        f"<div style='font-size:12px;color:#007AFF;background:#EFF6FF;"
        f"padding:8px 12px;border-radius:8px;border:0.5px solid #BFDBFE;"
        f"margin-top:6px'>Cession si stock donneur<br>"
        f"<strong>&gt; {seuil} unités</strong></div>",
        unsafe_allow_html=True
    )
    st.markdown("---")

    # Légende actions
    st.markdown(
        "<div style='font-size:11px;font-weight:600;color:#8E8E93;"
        "text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px'>"
        "Légende actions</div>", unsafe_allow_html=True
    )
    for icon, label, bg, fg, desc in [
        ("🔴", "COMMANDER",    "#FDECEA", "#C0392B",
         "Rupture réseau — aucun stock disponible"),
        ("🟢", "VOIR CESSION", "#E8F8F0", "#196F3D",
         f"Stock donneur > {seuil} u. — transfert possible"),
    ]:
        st.markdown(
            f"<div style='display:flex;gap:8px;align-items:flex-start;"
            f"margin-bottom:8px'>"
            f"<div style='background:{bg};color:{fg};border-radius:6px;"
            f"padding:3px 8px;font-size:11px;font-weight:700;"
            f"flex-shrink:0'>{icon} {label}</div>"
            f"<div style='font-size:11px;color:#8E8E93;margin-top:3px'>"
            f"{desc}</div></div>",
            unsafe_allow_html=True
        )


# ─── PAGE PRINCIPALE ──────────────────────────────────────────────────────────
st.markdown("<div class='page-title'>📦 Ruptures de Stock</div>",
            unsafe_allow_html=True)
st.markdown(
    "<div class='page-caption'>Détection des ruptures par magasin · "
    "Plan d'action Commander / Cession · Seuil de sécurité paramétrable</div>",
    unsafe_allow_html=True
)

# ─── ÉCRAN D'ACCUEIL ──────────────────────────────────────────────────────────
if uploaded is None:
    st.markdown("---")

    st.markdown("""
    <div class='alert-box alert-blue'>
      <strong>ℹ️ À quoi sert ce module ?</strong><br>
      Le module <strong>Ruptures de Stock</strong> analyse ton export ERP article par article
      et magasin par magasin pour détecter chaque cellule vide ou à zéro.<br><br>
      Pour chaque référence en rupture, il propose automatiquement une action :
      <strong>🔴 Commander</strong> si aucun stock n'existe dans le réseau,
      ou <strong>🟢 Voir Cession</strong> si un site donneur possède un stock
      suffisant au-dessus du seuil de sécurité paramétrable.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Format du fichier attendu
    st.markdown("<div class='section-label'>Format du fichier attendu</div>",
                unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    cols_info = [
        ("🏷️", "Rayon",        "Libellé rayon",         "ex: 00011 - DROGUERIE"),
        ("🏷️", "Famille",      "Libellé famille",       "ex: 00110 - SOINS ET CONFORT"),
        ("🏷️", "Sous Famille", "Libellé sous-famille",  "ex: 880 - VITRES"),
        ("📋", "Article",      "Code + libellé article","ex: 11000001 - 750ML RECH LAVEVITRE"),
        ("🏪", "CODE - Site",  "Une colonne par magasin","ex: 10202 - Palmeraie"),
        ("🏪", "...",          "Autant de colonnes que de sites","Détection automatique"),
        ("📊", "Total",        "Colonne total réseau",  "Calculé automatiquement"),
    ]
    for i, (icon, col_name, desc, example) in enumerate(cols_info):
        with (c1 if i < 4 else c2):
            st.markdown(f"""
            <div class='col-required'>
              <div style='font-size:16px;margin-top:1px'>{icon}</div>
              <div>
                <div class='col-name'>{col_name}</div>
                <div class='col-desc'>{desc}</div>
                <div class='col-example'>{example}</div>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown(
        "<div style='background:#F9F9FB;border:0.5px solid #E5E5EA;"
        "border-radius:8px;padding:10px 14px;margin-top:4px;font-size:12px;"
        "color:#8E8E93'>💡 Les colonnes magasins sont détectées <strong>automatiquement</strong> "
        "via le format <code>CODE - NOM</code> — le module fonctionne quel que soit "
        "le nombre de sites ou le rayon analysé.</div>",
        unsafe_allow_html=True
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # Les 2 actions
    st.markdown("<div class='section-label'>Les 2 actions proposées</div>",
                unsafe_allow_html=True)
    a1, a2 = st.columns(2)
    for col, icon, title, color, bg, cond, detail, conseil in [
        (a1, "🔴", "COMMANDER", "#C0392B", "#FDECEA",
         "Rupture dans tous les magasins ou stock donneur ≤ seuil",
         "Aucune cession possible — réappro fournisseur obligatoire.",
         "Transmettre la liste à l'acheteur pour commande immédiate"),
        (a2, "🟢", "VOIR CESSION", "#196F3D", "#E8F8F0",
         "Au moins un magasin possède un stock > seuil de sécurité",
         "Le site donneur (max stock) est indiqué avec son niveau de stock.",
         "Initier un transfert inter-magasins sans attendre le fournisseur"),
    ]:
        with col:
            st.markdown(f"""
            <div style='background:#FFFFFF;border:0.5px solid #E5E5EA;
                        border-radius:12px;padding:14px 16px;margin-bottom:10px;
                        border-left:3px solid {color}'>
              <div style='display:flex;align-items:center;gap:8px;margin-bottom:8px'>
                <span class='badge' style='background:{bg};color:{color};
                      font-size:13px;padding:4px 12px'>{icon} {title}</span>
              </div>
              <div style='font-size:13px;font-weight:500;color:#1C1C1E;
                          margin-bottom:4px'>{cond}</div>
              <div style='font-size:12px;color:#6C6C70;margin-bottom:6px'>{detail}</div>
              <div style='font-size:11px;color:{color};font-weight:500'>
                → {conseil}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Seuil cession
    st.markdown("<div class='section-label'>Comprendre le seuil de cession</div>",
                unsafe_allow_html=True)
    st.markdown(f"""
    <div class='alert-box alert-orange'>
      <strong>⚙️ Seuil de sécurité (défaut : 100 unités)</strong><br>
      Une cession n'est proposée que si le stock du site donneur est
      <strong>strictement supérieur au seuil</strong> paramétrable dans la sidebar.<br><br>
      Exemple : seuil = 100 · site donneur stock = 80 → <strong>🔴 COMMANDER</strong>
      (céder 80u mettrait le site en rupture)<br>
      Exemple : seuil = 100 · site donneur stock = 342 → <strong>🟢 VOIR CESSION</strong>
      depuis ce site avec 342u disponibles
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("⬆️ Charge ton export stock Excel dans la sidebar pour lancer l'analyse.")
    st.stop()


# ─── CHARGEMENT ───────────────────────────────────────────────────────────────
file_bytes = uploaded.read()
with st.spinner("Analyse des ruptures en cours…"):
    try:
        plan_df, arts, mag_names = compute_plan(file_bytes, seuil)
    except ValueError as e:
        st.error(f"❌ Fichier invalide : {e}")
        st.stop()
    except Exception as e:
        st.error(f"❌ Erreur inattendue : {e}")
        st.stop()

nb_total  = len(arts)
nb_rupt   = len(plan_df)
nb_cmd    = int((plan_df['Action'] == 'COMMANDER').sum())
nb_cess   = int((plan_df['Action'] == 'VOIR CESSION').sum())
nb_sains  = nb_total - nb_rupt
rayons_dispo = sorted(arts['Rayon'].dropna().unique().tolist())

# Alerte critique si magasin > 40%
mag_pcts = {m: sum(1 for v in arts[m] if pd.isna(v) or v <= 0) / nb_total * 100
            for m in mag_names}
worst_mag = max(mag_pcts, key=mag_pcts.get)
if mag_pcts[worst_mag] >= 40:
    st.markdown(
        f"<div class='alert-box alert-red'>🔴 <strong>Alerte critique</strong> — "
        f"<strong>{worst_mag}</strong> affiche un taux de rupture de "
        f"<strong>{mag_pcts[worst_mag]:.0f}%</strong> ({int(mag_pcts[worst_mag]*nb_total/100)} références)."
        f"</div>", unsafe_allow_html=True
    )

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "📊 Synthèse",
    f"🔴 À Commander ({nb_cmd})",
    f"🟢 Cessions ({nb_cess})",
    "🔍 Plan complet & Export",
])

PL = dict(
    paper_bgcolor=PL_BG, plot_bgcolor=PL_BG,
    font=dict(family="-apple-system, Helvetica Neue, Arial",
              color=PL_FONT, size=11),
    margin=dict(t=16, b=16, l=8, r=16),
)


# ═══ TAB 1 — SYNTHÈSE ═════════════════════════════════════════════════════════
with tab1:
    # KPIs
    st.markdown("<div class='section-label'>Indicateurs globaux</div>",
                unsafe_allow_html=True)
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Références analysées", fmt(nb_total))
    k2.metric("En rupture (≥ 1 site)", fmt(nb_rupt),
              delta=f"{nb_rupt/nb_total*100:.0f}% du stock",
              delta_color="inverse")
    k3.metric("Saines (0 rupture)", fmt(nb_sains))
    k4.metric("🔴 À commander",     fmt(nb_cmd),
              help="Aucun stock dans le réseau ou sous le seuil")
    k5.metric("🟢 Cessions possibles", fmt(nb_cess),
              help=f"Stock donneur > {seuil} unités")

    st.markdown("<br>", unsafe_allow_html=True)

    col_g1, col_g2 = st.columns(2)

    # Graphe magasins
    with col_g1:
        st.markdown("<div class='section-label'>Taux de rupture par magasin</div>",
                    unsafe_allow_html=True)
        mag_rows = sorted(
            [{'Magasin': m, 'Pct': mag_pcts[m]} for m in mag_names],
            key=lambda x: x['Pct']
        )
        colors_bar = [
            COLOR_CMD if r['Pct'] >= 40 else
            (COLOR_WARN if r['Pct'] >= 25 else COLOR_CESS)
            for r in mag_rows
        ]
        fig_mag = go.Figure(go.Bar(
            x=[r['Pct'] for r in mag_rows],
            y=[r['Magasin'] for r in mag_rows],
            orientation='h',
            marker_color=colors_bar,
            marker_line_width=0,
            text=[f"{r['Pct']:.0f}%" for r in mag_rows],
            textposition='outside',
            hovertemplate='%{y}<br>%{x:.1f}% de ruptures<extra></extra>',
        ))
        fig_mag.update_layout(
            **PL, height=max(260, len(mag_names) * 32),
            xaxis=dict(showgrid=True, gridcolor=PL_GRID, zeroline=False,
                       ticksuffix='%', tickfont=dict(size=10)),
            yaxis=dict(showgrid=False, tickfont=dict(size=11)),
            showlegend=False,
        )
        st.plotly_chart(fig_mag, use_container_width=True,
                        config={"displayModeBar": False})

    # Graphe familles
    with col_g2:
        st.markdown("<div class='section-label'>Ruptures par famille</div>",
                    unsafe_allow_html=True)
        fam_rows = []
        for fam, grp in arts.groupby('Famille'):
            sub   = plan_df[plan_df['Famille'] == fam]
            cmd_n = int((sub['Action'] == 'COMMANDER').sum())
            ces_n = int((sub['Action'] == 'VOIR CESSION').sum())
            if cmd_n + ces_n > 0:
                fam_rows.append({'Famille': fam, 'Commander': cmd_n,
                                 'Cession': ces_n, 'Total': cmd_n + ces_n})
        fam_df = pd.DataFrame(fam_rows).sort_values('Total', ascending=True)

        fig_fam = go.Figure()
        fig_fam.add_trace(go.Bar(
            name='Commander', x=fam_df['Commander'], y=fam_df['Famille'],
            orientation='h', marker_color=COLOR_CMD, marker_line_width=0,
            hovertemplate='%{y}<br>%{x} à commander<extra></extra>',
        ))
        fig_fam.add_trace(go.Bar(
            name='Voir cession', x=fam_df['Cession'], y=fam_df['Famille'],
            orientation='h', marker_color=COLOR_CESS, marker_line_width=0,
            hovertemplate='%{y}<br>%{x} cessions possibles<extra></extra>',
        ))
        fig_fam.update_layout(
            **PL, barmode='stack',
            height=max(260, len(fam_df) * 32),
            xaxis=dict(showgrid=True, gridcolor=PL_GRID, zeroline=False,
                       tickfont=dict(size=10)),
            yaxis=dict(showgrid=False, tickfont=dict(size=11)),
            legend=dict(orientation='h', yanchor='bottom', y=1.02,
                        font=dict(size=10)),
        )
        st.plotly_chart(fig_fam, use_container_width=True,
                        config={"displayModeBar": False})

    # Répartition par rayon si plusieurs rayons
    if len(rayons_dispo) > 1:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Diagnostic par rayon</div>",
                    unsafe_allow_html=True)
        rcols = st.columns(min(len(rayons_dispo), 4))
        for i, rayon in enumerate(rayons_dispo):
            sub_arts = arts[arts['Rayon'] == rayon]
            sub_plan = plan_df[plan_df['Rayon'] == rayon]
            nb_r     = len(sub_arts)
            nb_r_rupt = len(sub_plan)
            nb_r_cmd  = int((sub_plan['Action'] == 'COMMANDER').sum())
            nb_r_ces  = int((sub_plan['Action'] == 'VOIR CESSION').sum())
            pct_r     = nb_r_rupt / nb_r * 100 if nb_r > 0 else 0
            alert     = "🔴" if pct_r >= 40 else ("🟠" if pct_r >= 25 else "🟢")
            with rcols[i % len(rcols)]:
                st.markdown(f"""
                <div class='kpi-card' style='margin-bottom:10px'>
                  <div style='font-size:11px;font-weight:600;color:{COLOR_INFO};
                              text-transform:uppercase;letter-spacing:.05em;
                              margin-bottom:6px'>{rayon}</div>
                  <div style='font-size:18px;font-weight:600;color:#1C1C1E'>
                    {fmt(nb_r_rupt)} rupt.</div>
                  <div style='font-size:12px;color:#8E8E93;margin-top:3px'>
                    sur {fmt(nb_r)} références · {pct_r:.0f}%</div>
                  <div style='font-size:12px;margin-top:5px'>
                    {alert}
                    <strong style='color:{COLOR_CMD}'>{fmt(nb_r_cmd)}</strong>
                    à commander ·
                    <strong style='color:{COLOR_CESS}'>{fmt(nb_r_ces)}</strong>
                    cessions</div>
                </div>""", unsafe_allow_html=True)


# ═══ TAB 2 — À COMMANDER ══════════════════════════════════════════════════════
with tab2:
    cmd_df = plan_df[plan_df['Action'] == 'COMMANDER'].copy()

    if len(cmd_df) == 0:
        st.markdown("""
        <div class='alert-box alert-green'>
          ✅ <strong>Aucune référence à commander</strong> — toutes les ruptures
          peuvent être couvertes par cession inter-magasins avec le seuil actuel.
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(
            f"<div class='alert-box alert-red'>🔴 <strong>{fmt(nb_cmd)} références</strong> "
            f"sans stock disponible dans le réseau — commande fournisseur requise.</div>",
            unsafe_allow_html=True
        )
        k1, k2, k3 = st.columns(3)
        k1.metric("Refs à commander", fmt(nb_cmd))
        k2.metric("Rupture 9/9 magasins",
                  fmt(int((cmd_df['Nb_Ruptures'] == len(mag_names)).sum())))
        k3.metric("Familles concernées",
                  str(cmd_df['Famille'].nunique()))

        st.markdown("<br>", unsafe_allow_html=True)

        # Filtres
        fc1, fc2 = st.columns([2, 2])
        with fc1:
            sel_fam_cmd = st.selectbox(
                "Famille", ['Toutes'] + sorted(cmd_df['Famille'].dropna().unique().tolist()),
                key='fam_cmd'
            )
        with fc2:
            sel_ray_cmd = st.selectbox(
                "Rayon", ['Tous'] + sorted(cmd_df['Rayon'].dropna().unique().tolist()),
                key='ray_cmd'
            )
        view_cmd = cmd_df.copy()
        if sel_fam_cmd != 'Toutes':
            view_cmd = view_cmd[view_cmd['Famille'] == sel_fam_cmd]
        if sel_ray_cmd != 'Tous':
            view_cmd = view_cmd[view_cmd['Rayon'] == sel_ray_cmd]

        st.markdown(
            f"<span style='font-size:12px;color:#8E8E93'>"
            f"{len(view_cmd):,} référence(s)</span>",
            unsafe_allow_html=True
        )
        disp_cmd = view_cmd[[
            'Code_Article', 'Libellé', 'Rayon', 'Famille',
            'Nb_Ruptures', 'Magasins_Rupture'
        ]].copy()
        disp_cmd.columns = [
            'Code', 'Libellé', 'Rayon', 'Famille',
            'Nb rupt.', 'Magasins en rupture'
        ]
        st.dataframe(
            disp_cmd, use_container_width=True, height=460,
            column_config={
                'Code':    st.column_config.TextColumn('Code',   width=100),
                'Libellé': st.column_config.TextColumn('Libellé', width=300),
                'Rayon':   st.column_config.TextColumn('Rayon',  width=140),
                'Famille': st.column_config.TextColumn('Famille', width=180),
                'Nb rupt.': st.column_config.NumberColumn('Nb rupt.', width=80,
                                format="%d"),
                'Magasins en rupture': st.column_config.TextColumn(
                    'Magasins en rupture', width=300),
            },
            hide_index=True,
        )


# ═══ TAB 3 — CESSIONS ═════════════════════════════════════════════════════════
with tab3:
    cess_df = plan_df[plan_df['Action'] == 'VOIR CESSION'].sort_values(
        ['Nb_Ruptures', 'Stock_Disponible'], ascending=[False, False]
    ).copy()

    if len(cess_df) == 0:
        st.markdown("""
        <div class='alert-box alert-orange'>
          ⚠️ Aucune cession possible avec le seuil actuel.
          Réduisez le seuil dans la sidebar pour déverrouiller des cessions.
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(
            f"<div class='alert-box alert-green'>🟢 <strong>{fmt(nb_cess)} références</strong> "
            f"peuvent être couverte(s) par cession — stock donneur &gt; {seuil} unités.</div>",
            unsafe_allow_html=True
        )

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Refs cédables",  fmt(nb_cess))
        k2.metric("Sites donneurs",
                  str(cess_df['Site_Donneur'].nunique()))
        k3.metric("Stock max disponible",
                  fmt(int(cess_df['Stock_Disponible'].max()))
                  if cess_df['Stock_Disponible'].notna().any() else "—")
        k4.metric("Familles concernées",
                  str(cess_df['Famille'].nunique()))

        st.markdown("<br>", unsafe_allow_html=True)

        # Graphe top sites donneurs
        donors = (cess_df.groupby('Site_Donneur')
                  .agg(Nb_Cessions=('Site_Donneur', 'count'),
                       Stock_Total=('Stock_Disponible', 'sum'))
                  .reset_index()
                  .sort_values('Nb_Cessions', ascending=False))

        st.markdown(
            "<div class='section-label'>Sites donneurs (nb de références cédables)</div>",
            unsafe_allow_html=True
        )
        fig_don = go.Figure(go.Bar(
            x=donors['Site_Donneur'], y=donors['Nb_Cessions'],
            marker_color=COLOR_CESS, marker_line_width=0,
            text=donors['Nb_Cessions'], textposition='outside',
            hovertemplate='%{x}<br>%{y} cessions<extra></extra>',
        ))
        fig_don.update_layout(
            **PL, height=240,
            xaxis=dict(showgrid=False, tickfont=dict(size=11)),
            yaxis=dict(showgrid=True, gridcolor=PL_GRID,
                       title='Nb références', tickfont=dict(size=10)),
            showlegend=False,
        )
        st.plotly_chart(fig_don, use_container_width=True,
                        config={"displayModeBar": False})

        # Filtres
        fc1, fc2, fc3 = st.columns([2, 2, 2])
        with fc1:
            sel_fam_c = st.selectbox(
                "Famille", ['Toutes'] + sorted(cess_df['Famille'].dropna().unique().tolist()),
                key='fam_cess'
            )
        with fc2:
            sel_don = st.selectbox(
                "Site donneur", ['Tous'] + sorted(cess_df['Site_Donneur'].unique().tolist()),
                key='don_cess'
            )
        with fc3:
            sel_mag_r = st.selectbox(
                "Magasin en rupture", ['Tous'] + sorted(mag_names),
                key='mag_cess'
            )
        view_cess = cess_df.copy()
        if sel_fam_c != 'Toutes':
            view_cess = view_cess[view_cess['Famille'] == sel_fam_c]
        if sel_don != 'Tous':
            view_cess = view_cess[view_cess['Site_Donneur'] == sel_don]
        if sel_mag_r != 'Tous':
            view_cess = view_cess[
                view_cess['Magasins_Rupture'].str.contains(sel_mag_r, regex=False)
            ]

        st.markdown(
            f"<span style='font-size:12px;color:#8E8E93'>"
            f"{len(view_cess):,} référence(s)</span>",
            unsafe_allow_html=True
        )
        disp_cess = view_cess[[
            'Code_Article', 'Libellé', 'Famille', 'Nb_Ruptures',
            'Magasins_Rupture', 'Site_Donneur', 'Stock_Disponible'
        ]].copy()
        disp_cess['Stock_Disponible'] = disp_cess['Stock_Disponible'].apply(
            lambda x: int(x) if pd.notna(x) else ''
        )
        disp_cess.columns = [
            'Code', 'Libellé', 'Famille', 'Nb rupt.',
            'Magasins en rupture', 'Site donneur', 'Stock dispo'
        ]
        st.dataframe(
            disp_cess, use_container_width=True, height=460,
            column_config={
                'Code':    st.column_config.TextColumn('Code',   width=100),
                'Libellé': st.column_config.TextColumn('Libellé', width=280),
                'Famille': st.column_config.TextColumn('Famille', width=160),
                'Nb rupt.': st.column_config.NumberColumn('Nb rupt.', width=80,
                                format="%d"),
                'Magasins en rupture': st.column_config.TextColumn(
                    'Magasins en rupture', width=260),
                'Site donneur': st.column_config.TextColumn('Site donneur',
                                    width=120),
                'Stock dispo': st.column_config.NumberColumn('Stock dispo',
                                   width=90, format="%d"),
            },
            hide_index=True,
        )


# ═══ TAB 4 — PLAN COMPLET & EXPORT ════════════════════════════════════════════
with tab4:
    st.markdown("<div class='section-label'>Récapitulatif avant export</div>",
                unsafe_allow_html=True)

    e1, e2, e3, e4 = st.columns(4)
    for col, val, label, color, bg in [
        (e1, nb_total,  "Refs analysées",  "#007AFF", "#EFF6FF"),
        (e2, nb_rupt,   "En rupture",      "#C0392B", "#FDECEA"),
        (e3, nb_cmd,    "À commander 🔴",  "#C0392B", "#FDECEA"),
        (e4, nb_cess,   "Cessions 🟢",     "#196F3D", "#E8F8F0"),
    ]:
        with col:
            st.markdown(
                f"<div class='kpi-card' style='border-top:3px solid {color};"
                f"text-align:center'>"
                f"<div style='font-size:24px;font-weight:700;color:{color}'>"
                f"{fmt(val)}</div>"
                f"<div style='font-size:11px;color:#8E8E93;margin-top:4px;"
                f"font-weight:600;text-transform:uppercase;letter-spacing:.04em'>"
                f"{label}</div></div>",
                unsafe_allow_html=True
            )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class='alert-box alert-blue'>
      <strong>📋 Contenu de l'export Excel</strong><br>
      · <strong>Plan d'action complet</strong> — {fmt(nb_rupt)} références avec
        colonne ACTION colorée, site donneur et stock disponible<br>
      · Filtre automatique sur toutes les colonnes<br>
      · Seuil cession appliqué : <strong>&gt; {seuil} unités</strong><br>
      · Compatible Excel 2016+
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Filtres plan complet
    pf1, pf2, pf3, pf4 = st.columns([2, 2, 2, 1])
    with pf1:
        sel_fam_p = st.selectbox(
            "Famille",
            ['Toutes'] + sorted(plan_df['Famille'].dropna().unique().tolist()),
            key='fam_plan'
        )
    with pf2:
        sel_act_p = st.selectbox(
            "Action", ['Toutes', 'COMMANDER', 'VOIR CESSION'], key='act_plan'
        )
    with pf3:
        sel_ray_p = st.selectbox(
            "Rayon",
            ['Tous'] + sorted(plan_df['Rayon'].dropna().unique().tolist()),
            key='ray_plan'
        )
    with pf4:
        nb_min_p = st.number_input(
            "Nb rupt. min.", min_value=1, max_value=len(mag_names),
            value=1, key='nb_plan'
        )

    view_plan = plan_df.copy()
    if sel_fam_p != 'Toutes':  view_plan = view_plan[view_plan['Famille'] == sel_fam_p]
    if sel_act_p != 'Toutes':  view_plan = view_plan[view_plan['Action']  == sel_act_p]
    if sel_ray_p != 'Tous':    view_plan = view_plan[view_plan['Rayon']   == sel_ray_p]
    view_plan = view_plan[view_plan['Nb_Ruptures'] >= nb_min_p]

    disp_plan = view_plan[[
        'Code_Article', 'Libellé', 'Rayon', 'Famille', 'Nb_Ruptures',
        'Magasins_Rupture', 'Action', 'Site_Donneur', 'Stock_Disponible'
    ]].copy()
    disp_plan['Action'] = disp_plan['Action'].apply(
        lambda a: '🔴 COMMANDER' if a == 'COMMANDER' else '🟢 VOIR CESSION'
    )
    disp_plan['Stock_Disponible'] = disp_plan['Stock_Disponible'].apply(
        lambda x: int(x) if pd.notna(x) and x else ''
    )
    disp_plan.columns = [
        'Code', 'Libellé', 'Rayon', 'Famille', 'Nb rupt.',
        'Magasins en rupture', 'Action', 'Site donneur', 'Stock dispo'
    ]

    st.markdown(
        f"<span style='font-size:12px;color:#8E8E93'>"
        f"{len(view_plan):,} référence(s) affichée(s)</span>",
        unsafe_allow_html=True
    )
    st.dataframe(
        disp_plan, use_container_width=True, height=400,
        column_config={
            'Code':    st.column_config.TextColumn('Code',   width=100),
            'Libellé': st.column_config.TextColumn('Libellé', width=280),
            'Rayon':   st.column_config.TextColumn('Rayon',  width=130),
            'Famille': st.column_config.TextColumn('Famille', width=160),
            'Nb rupt.': st.column_config.NumberColumn('Nb rupt.', width=80,
                            format="%d"),
            'Magasins en rupture': st.column_config.TextColumn(
                'Magasins en rupture', width=240),
            'Action': st.column_config.TextColumn('Action', width=150),
            'Site donneur': st.column_config.TextColumn('Site donneur', width=120),
            'Stock dispo': st.column_config.NumberColumn('Stock dispo',
                               width=90, format="%d"),
        },
        hide_index=True,
    )

    # Exports
    st.markdown("<br>", unsafe_allow_html=True)
    dl1, dl2, _ = st.columns([1, 1, 2])
    with dl1:
        st.download_button(
            "⬇️  Exporter le plan complet",
            data=to_excel_bytes(plan_df, seuil),
            file_name="SmartBuyer_Ruptures_Plan_Complet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl2:
        st.download_button(
            "⬇️  À commander uniquement",
            data=to_excel_bytes(
                plan_df[plan_df['Action'] == 'COMMANDER'].copy(), seuil
            ),
            file_name="SmartBuyer_Ruptures_A_Commander.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
