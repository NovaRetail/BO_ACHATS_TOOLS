"""
15_📋_CODIR_Hebdo.py — Module CODIR Hebdo · SmartBuyer Hub
Vue réseau (Rayon) + Destructeurs/Performeurs (Article) sur exports PowerBI hebdo.
Deux exports indépendants : Rayon→Famille→Sous-Famille, et Rayon→Famille→Sous-Famille→Article.
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG & CHARTE (identique au reste du Hub)
# ============================================================
st.set_page_config(page_title="CODIR Hebdo", page_icon="📋", layout="wide")

BLUE = "#007AFF"
GREEN = "#34C759"
RED = "#FF3B30"
AMBER = "#FF9500"
DARK = "#1C1C1E"
GREY = "#8E8E93"
BG = "#F2F2F7"

# ============================================================
# 🎯 CIBLES DE MARGE PAR RAYON — identiques au module 14_💰_Marge.py
# Clé = mot-clé contenu dans le libellé Rayon de l'export PBI.
# ============================================================
CIBLES_DEFAUT = {
    "BOISSONS": 19.5,
    "DROGUERIE": 25.0,
    "PARFUMERIE HYGIENE": 29.0,
    "EPICERIE": 16.0,
}
CIBLE_FALLBACK = 23.5

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', -apple-system, 'SF Pro Display', BlinkMacSystemFont, Calibri, sans-serif; }}
.stApp {{ background-color: {BG}; }}
.block-container {{ padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1280px; }}
[data-testid="stSidebar"] {{ background-color: #FFFFFF; border-right: 1px solid #E5E5EA; }}
hr {{ border-color: #E5E5EA !important; margin: 1rem 0 !important; }}

.page-title {{ font-size: 28px; font-weight: 700; color: {DARK}; letter-spacing: -0.03em; margin: 0; }}
.page-caption {{ font-size: 13px; color: {GREY}; margin-top: 3px; margin-bottom: 1.5rem; }}
.section-label {{ font-size: 11px; font-weight: 600; color: {GREY};
                 text-transform: uppercase; letter-spacing: 0.07em; margin: 18px 0 10px; }}

.card {{ background:#FFFFFF; border-radius:12px; padding:16px 18px; margin-bottom:10px;
        border:1px solid #E5E5EA; box-shadow:0 1px 3px rgba(0,0,0,0.04); }}

.kpi-card {{ background:#FFFFFF; border-radius:12px; padding:16px 18px;
            border:1px solid #E5E5EA; box-shadow:0 1px 3px rgba(0,0,0,0.04); }}
.kpi-label {{ font-size:11px; font-weight:500; color:{GREY};
             text-transform:uppercase; letter-spacing:0.04em; margin-bottom:3px; }}
.kpi-value {{ font-size:24px; font-weight:700; color:{DARK}; letter-spacing:-0.02em; line-height:1.1; }}
.kpi-sub {{ font-size:12px; color:{GREY}; margin-top:3px; }}
.kpi-sub.pos {{ color:{GREEN}; }}
.kpi-sub.neg {{ color:{RED}; }}

.info-box {{ background:#F0F8FF; border-left:3px solid {BLUE}; border-radius:10px;
            padding:16px 20px; margin-bottom:24px; }}
.info-box .it {{ font-size:15px; font-weight:700; color:{DARK}; margin-bottom:10px; }}
.info-box .ip {{ font-size:13px; color:#1C1C1E; line-height:1.6; }}
.info-box .iq {{ margin-top:14px; font-size:13px; color:#1C1C1E; line-height:1.9; }}

.badge {{ display:inline-block; padding:2px 10px; border-radius:6px; font-size:11px; font-weight:600; }}
</style>
""", unsafe_allow_html=True)

# ============================================================
# HELPERS DE FORMAT (convention SmartBuyer — cf. 06_💸_Marges_Negatives.py)
# ============================================================
def fmt(n):
    if n is None or (isinstance(n, float) and (pd.isna(n) or not np.isfinite(n))):
        return "—"
    a = abs(n)
    if a >= 1_000_000: return f"{n/1_000_000:.1f} M"
    if a >= 1_000:     return f"{int(n/1_000)} K"
    return f"{int(n):,}"

def fmt_pct(v, dec=1):
    if v is None or pd.isna(v): return "—"
    return f"{v:.{dec}f}%"

def fmt_delta(v):
    if v is None or pd.isna(v): return "—"
    return f"{v:+.1f} pts"

def safe_rate(num, den):
    den = np.where(den == 0, np.nan, den)
    r = num / den * 100
    return np.where(np.isnan(r), 0.0, r)

def rayon_key(libelle):
    """Mot-clé de rapprochement vers CIBLES_DEFAUT (insensible au code/préfixe numérique)."""
    s = str(libelle).upper()
    for k in CIBLES_DEFAUT:
        if k in s:
            return k
    return None

def kpi_card(label, value, sub=None, sub_class=""):
    sub_html = f"<div class='kpi-sub {sub_class}'>{sub}</div>" if sub else ""
    return (f"<div class='kpi-card'><div class='kpi-label'>{label}</div>"
            f"<div class='kpi-value'>{value}</div>{sub_html}</div>")

# ============================================================
# CHARGEMENT — EXPORT ARTICLE UNIQUE (Rayon → Famille → Sous Famille → Article)
# Toute la hiérarchie (réseau / rayon / famille / article) est dérivée de ce
# seul fichier — c'est lui que l'on charge chaque semaine.
# ============================================================
@st.cache_data(show_spinner=False)
def load_export(file_bytes):
    raw = pd.read_excel(io.BytesIO(file_bytes))
    raw.columns = [str(c).lstrip('\ufeff') for c in raw.columns]

    # La note "Filtres appliqués : ..." (dernière ligne) décrit le périmètre exact
    # (sites, enseigne, période) — utile pour vérifier la cohérence semaine après semaine.
    perimetre = None
    note_rows = raw[raw['Rayon'].astype(str).str.startswith('Filtres', na=False)]
    if not note_rows.empty:
        perimetre = str(note_rows.iloc[0]['Rayon'])

    df = raw[raw['Rayon'].notna()].copy()
    df = df[~df['Rayon'].astype(str).str.startswith('Filtres')]
    if 'Article' not in df.columns:
        df['Article'] = np.nan
    return df, perimetre

def kpis_globaux_rayon(df):
    g = df[df['Rayon'] == 'Total']
    if g.empty:
        return None
    g = g.iloc[0]
    ca, ca_n1, marge = g['CA'], g['CA N-1'], g['Marge']
    evol_marge_pct = g.get('%Vs N-1.1', np.nan)
    marge_n1 = marge / (1 + evol_marge_pct) if pd.notna(evol_marge_pct) and (1 + evol_marge_pct) != 0 else np.nan
    return {
        'ca': ca, 'ca_n1': ca_n1, 'evol_ca': ca/ca_n1 - 1 if ca_n1 else np.nan,
        'marge': marge, 'marge_n1': marge_n1,
        'tx_marge': marge/ca*100 if ca else np.nan,
        'tx_marge_n1': marge_n1/ca_n1*100 if ca_n1 and pd.notna(marge_n1) else np.nan,
        'qte': g.get('Qté Vente', np.nan), 'qte_n1': g.get('Qté Vente N-1', np.nan),
        'poids_promo': g.get('%CA Poids Promo', np.nan) * 100,
        'casse': g.get('Casse (Valeur)', np.nan),
    }

def perf_par_rayon(df, cibles):
    rows = []
    sub = df[(df['Famille'] == 'Total') & (df['Rayon'] != 'Total')]
    for _, r in sub.iterrows():
        key = rayon_key(r['Rayon'])
        cible = cibles.get(key, CIBLE_FALLBACK) if key else CIBLE_FALLBACK
        ca, ca_n1 = r['CA'], r['CA N-1']
        qte, qte_n1 = r.get('Qté Vente', np.nan), r.get('Qté Vente N-1', np.nan)
        tx = r['Marge']/ca*100 if ca else np.nan
        rows.append({
            'Rayon': str(r['Rayon']).split(' - ')[-1].strip(),
            'CA': ca, 'Évol CA %': (ca/ca_n1-1)*100 if ca_n1 else np.nan,
            'Évol Qté %': (qte/qte_n1-1)*100 if qte_n1 else np.nan,
            'Taux Marge %': tx, 'Objectif %': cible, 'Écart (pts)': tx - cible if pd.notna(tx) else np.nan,
        })
    return pd.DataFrame(rows).sort_values('CA', ascending=False)

def family_metrics(df):
    """Prépare toutes les métriques Top/Flop au niveau Famille (Sous Famille == Total)."""
    sub = df[(df['Sous Famille'] == 'Total') & (df['Famille'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    sub['Rayon_aff'] = sub['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    sub['Famille_aff'] = sub['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    sub['CA'] = sub['CA'].fillna(0)
    sub['CA N-1'] = sub['CA N-1'].fillna(0)
    sub['Perte CA'] = sub['CA'] - sub['CA N-1']
    sub['Évol CA %'] = np.where(sub['CA N-1'] > 0, (sub['CA']/sub['CA N-1']-1)*100, np.nan)
    sub['Tx Marge %'] = np.where(sub['CA'] > 0, sub['Marge']/sub['CA']*100, np.nan)
    evol_marge_pct = sub.get('%Vs N-1.1', pd.Series(np.nan, index=sub.index))
    with np.errstate(divide='ignore', invalid='ignore'):
        marge_n1 = sub['Marge'] / (1 + evol_marge_pct)
    marge_n1 = marge_n1.replace([np.inf, -np.inf], np.nan)
    sub['Tx Marge N-1 %'] = np.where(sub['CA N-1'] > 0, marge_n1/sub['CA N-1']*100, np.nan)
    sub['Écart Tx Marge (pts)'] = sub['Tx Marge %'] - sub['Tx Marge N-1 %']
    return sub

def top_familles(df, n=5, by='perte_ca'):
    """Top/Flop N familles selon le critère choisi (conserve la version courte pour compatibilité)."""
    sub = family_metrics(df)
    if by == 'perte_ca':
        out = sub.nsmallest(n, 'Perte CA')[['Rayon_aff','Famille_aff','CA','CA N-1','Perte CA','Tx Marge %']]
    elif by == 'casse':
        out = sub.nsmallest(n, 'Casse (Valeur)')[['Rayon_aff','Famille_aff','CA','Casse (Valeur)','%Casse (Valeur)']]
    elif by == 'promo':
        mat = sub[sub['CA'] > 1_000_000]
        out = mat.nlargest(n, '%CA Poids Promo')[['Rayon_aff','Famille_aff','CA','%CA Poids Promo','%Marge Promo','%Marge Hors Promo']]
    return out.reset_index(drop=True)

def top_flop_table(sub, metric, n, mode, cols, ca_floor=0):
    """mode='top' -> nlargest, mode='flop' -> nsmallest. ca_floor filtre les familles trop petites (bruit)."""
    base = sub[sub['CA'] > ca_floor] if ca_floor else sub
    base = base[base[metric].notna()]
    out = base.nlargest(n, metric) if mode == 'top' else base.nsmallest(n, metric)
    return out[['Rayon_aff','Famille_aff'] + cols].reset_index(drop=True)

# ============================================================
# DÉRIVATION — VUE ARTICLE (à partir du même dataframe)
# ============================================================
def prep_articles(df):
    # lignes article réelles uniquement (pas les sous-totaux Rayon/Famille/Sous-Famille)
    art = df[df['Article'].notna() & (df['Article'] != 'Total') & (df['Rayon'] != 'Total')].copy()
    art['Rayon_aff'] = art['Rayon'].astype(str).str.split(' - ').str[-1].str.strip()
    art['Famille_aff'] = art['Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    art['SousFamille_aff'] = art['Sous Famille'].astype(str).str.split(' - ').str[-1].str.strip()
    art['Article_aff'] = art['Article'].astype(str)
    art['CA'] = art['CA'].fillna(0)
    art['Marge'] = art['Marge'].fillna(0)
    art['Qté Vente'] = art['Qté Vente'].fillna(0)
    art['Qté Vente N-1'] = art['Qté Vente N-1'].fillna(0)
    art['CA N-1'] = art['CA N-1'].fillna(0)
    evol_marge = art.get('%Vs N-1.1', pd.Series(np.nan, index=art.index))
    with np.errstate(divide='ignore', invalid='ignore'):
        marge_n1 = art['Marge'] / (1 + evol_marge)
    marge_n1 = marge_n1.replace([np.inf, -np.inf], np.nan)
    art['Marge N-1 (calc)'] = marge_n1
    art['Tx Marge %'] = np.where(art['CA'] > 0, art['Marge']/art['CA']*100, np.nan)
    art['Tx Marge N-1 % (calc)'] = np.where(art['CA N-1'] > 0, marge_n1/art['CA N-1']*100, np.nan)
    art['Écart Tx Marge (pts)'] = art['Tx Marge %'] - art['Tx Marge N-1 % (calc)']
    art['Gain Marge (FCFA)'] = art['Marge'] - marge_n1
    art['Variation Qté'] = art['Qté Vente'] - art['Qté Vente N-1']
    art['Perte CA (FCFA)'] = art['CA'] - art['CA N-1']
    return art

def destructeurs_performeurs(art, n=15, seuil_ca=100_000):
    res = {}
    res['A_marge_neg'] = art[art['Marge'] < 0].nsmallest(n, 'Marge')
    pos = art[art['Marge'] >= 0].copy()
    res['B_degrad_marge'] = pos[pos['Écart Tx Marge (pts)'].notna()].nsmallest(n, 'Écart Tx Marge (pts)')
    mat = art[art['CA'] > seuil_ca]
    res['C_perf_gain_marge'] = mat.nlargest(n, 'Gain Marge (FCFA)')
    mat_n1 = mat[mat['CA N-1'] > 0].assign(_evol=lambda d: d['CA']/d['CA N-1']-1)
    res['D_croissance_ca'] = mat_n1.nlargest(n, '_evol')
    res['E_baisse_ca'] = art.nsmallest(n, 'Perte CA (FCFA)')
    res['F_hausse_qte'] = art.nlargest(n, 'Variation Qté')
    res['G_baisse_qte'] = art.nsmallest(n, 'Variation Qté')
    return res

DISPLAY_COLS = ['Rayon_aff','Famille_aff','SousFamille_aff','Article_aff']
DISPLAY_RENAME = {'Rayon_aff':'Rayon','Famille_aff':'Famille','SousFamille_aff':'Sous Famille','Article_aff':'Article'}

def show_table(d, extra_cols, fmt_map=None):
    if d.empty:
        st.caption("Aucune ligne ne correspond à ce critère sur la période.")
        return
    cols = DISPLAY_COLS + extra_cols
    disp = d[cols].rename(columns=DISPLAY_RENAME).reset_index(drop=True)
    disp.index = disp.index + 1
    if fmt_map:
        for c, f in fmt_map.items():
            if c in disp.columns:
                disp[c] = disp[c].map(f)
    st.dataframe(disp, use_container_width=True)

# ============================================================
# EXPORT EXCEL (valeurs figées — recalcul natif via Data_Semaine/Data_Articles
# conseillé pour le suivi hebdo continu, ce bouton sert aux exports ponctuels CODIR)
# ============================================================
def build_excel_codir(kpis, perf_rayon, tops):
    wb = Workbook()
    ws = wb.active
    ws.title = "CODIR Hebdo"
    bold = Font(bold=True, color="FFFFFFFF")
    fill = PatternFill("solid", fgColor="FF007AFF")
    ws.append(["Indicateur", "Valeur", "N-1", "Évolution"])
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = fill
    rows = [
        ("CA (FCFA)", kpis['ca'], kpis['ca_n1'], kpis['evol_ca']),
        ("Marge (FCFA)", kpis['marge'], kpis['marge_n1'], None),
        ("Taux de marge", kpis['tx_marge'], kpis['tx_marge_n1'], None),
        ("Qté vendue", kpis['qte'], kpis['qte_n1'], None),
        ("Casse (FCFA)", kpis['casse'], None, None),
    ]
    for label, v, n1, evo in rows:
        ws.append([label, v, n1, evo])
    ws.append([])
    ws.append(["Rayon", "CA", "Évol CA %", "Taux Marge %", "Objectif %", "Écart (pts)"])
    for c in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = bold
        cell.fill = fill
    for _, r in perf_rayon.iterrows():
        ws.append([r['Rayon'], r['CA'], r['Évol CA %'], r['Taux Marge %'], r['Objectif %'], r['Écart (pts)']])
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ============================================================
# INTERFACE
# ============================================================
st.markdown("<div class='page-title'>📋 Module CODIR Hebdo</div>"
            "<div class='page-caption'>Vue réseau PGC (Rayon) + Destructeurs/Performeurs (Article) · "
            "objectifs marge alignés Méti · une seule extraction à charger</div>", unsafe_allow_html=True)
st.markdown("<hr>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 📥 Import")
    up = st.file_uploader("Export Article hebdo (.xlsx)", type=['xlsx'], key="up_export")
    st.caption("L'extraction Article contient déjà les sous-totaux Rayon et Famille — "
               "tout le module en est dérivé, rien d'autre à charger.")
    st.markdown("---")
    st.markdown("##### ⚙️ Paramètres")
    n_top = st.slider("Nombre de lignes par classement", 5, 30, 15)
    seuil_ca = st.number_input("Seuil CA mini — performeurs/croissance (FCFA)", 0, 5_000_000, 100_000, step=10_000)
    st.markdown("---")
    st.caption("SmartBuyer Hub · Module CODIR Hebdo")

if up is None:
    st.markdown(
        f"<div class='info-box'>"
        f"<div class='it'>ℹ️ À quoi sert ce module ?</div>"
        f"<div class='ip'>Ce module prépare le point hebdo réseau pour le CODIR à partir de l'export PBI "
        f"<b>Rayon → Famille → Sous-Famille → Article</b>. Un seul fichier à charger chaque semaine, "
        f"dans la barre latérale.</div>"
        f"<div class='iq'>"
        f"<b>Onglet Dashboard CODIR</b> — CA, marge, quantités vs N-1 · performance par rayon vs objectifs Méti · "
        f"top familles en baisse de CA / casse / poids promo<br>"
        f"<b>Onglet Destructeurs &amp; Performeurs</b> — articles en marge négative, dégradation de taux de marge, "
        f"gain de marge, croissance/baisse de CA et de quantité</div>"
        f"</div>", unsafe_allow_html=True)
    st.stop()

df, perimetre = load_export(up.getvalue())
art = prep_articles(df)

if perimetre:
    with st.expander("🔎 Périmètre détecté dans le fichier"):
        st.code(perimetre, language=None)

tab1, tab2 = st.tabs(["📋 Dashboard CODIR", "💥 Destructeurs & Performeurs"])

# ---------------- TAB 1 : DASHBOARD CODIR (RAYON) ----------------
with tab1:
    k = kpis_globaux_rayon(df)
    if k is None:
        st.error("Ligne de total réseau ('Total') introuvable dans l'export — vérifiez le fichier.")
    else:
        st.markdown("<div class='section-label'>VUE D'ENSEMBLE RÉSEAU</div>", unsafe_allow_html=True)
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.markdown(kpi_card("CA", f"{fmt(k['ca'])} FCFA",
                    f"{k['evol_ca']*100:+.1f}% vs N-1", "pos" if k['evol_ca'] >= 0 else "neg"), unsafe_allow_html=True)
        c2.markdown(kpi_card("Marge", f"{fmt(k['marge'])} FCFA"), unsafe_allow_html=True)
        evo_tx = k['tx_marge'] - k['tx_marge_n1'] if pd.notna(k['tx_marge_n1']) else np.nan
        c3.markdown(kpi_card("Taux de marge", fmt_pct(k['tx_marge'], 2),
                    fmt_delta(evo_tx), "pos" if (pd.notna(evo_tx) and evo_tx >= 0) else "neg"), unsafe_allow_html=True)
        evo_qte = (k['qte']/k['qte_n1']-1)*100 if k['qte_n1'] else np.nan
        c4.markdown(kpi_card("Qté vendue", fmt(k['qte']),
                    f"{evo_qte:+.1f}% vs N-1" if pd.notna(evo_qte) else None,
                    "pos" if (pd.notna(evo_qte) and evo_qte >= 0) else "neg"), unsafe_allow_html=True)
        pct_casse = k['casse']/k['ca']*100 if k['ca'] else np.nan
        c5.markdown(kpi_card("Casse", f"{fmt(k['casse'])} FCFA", fmt_pct(pct_casse, 2) + " du CA"), unsafe_allow_html=True)

        st.markdown("<div class='section-label'>PERFORMANCE PAR RAYON VS OBJECTIFS MARGE (MÉTI)</div>", unsafe_allow_html=True)
        perf = perf_par_rayon(df, CIBLES_DEFAUT)
        disp = perf.copy()
        for c in ['Évol CA %', 'Évol Qté %', 'Taux Marge %', 'Objectif %']:
            disp[c] = disp[c].map(lambda v: fmt_pct(v))
        disp['Écart (pts)'] = perf['Écart (pts)'].map(fmt_delta)
        disp['CA'] = perf['CA'].map(lambda v: fmt(v))
        st.dataframe(disp, use_container_width=True, hide_index=True)

        st.markdown("<div class='section-label'>TOP & FLOP PAR FAMILLE</div>", unsafe_allow_html=True)
        fam = family_metrics(df)

        def pair(title_top, title_flop, metric, cols, fmt_map, ca_floor=0):
            cA, cB = st.columns(2)
            with cA:
                st.markdown(f"**🟢 {title_top}**")
                t = top_flop_table(fam, metric, n_top, 'top', cols, ca_floor)
                t = t.rename(columns={'Rayon_aff':'Rayon','Famille_aff':'Famille'})
                for c, f in fmt_map.items():
                    if c in t.columns: t[c] = t[c].map(f)
                st.dataframe(t, hide_index=True, use_container_width=True)
            with cB:
                st.markdown(f"**🔴 {title_flop}**")
                t = top_flop_table(fam, metric, n_top, 'flop', cols, ca_floor)
                t = t.rename(columns={'Rayon_aff':'Rayon','Famille_aff':'Famille'})
                for c, f in fmt_map.items():
                    if c in t.columns: t[c] = t[c].map(f)
                st.dataframe(t, hide_index=True, use_container_width=True)

        st.markdown("##### 📈 Évolution du CA (classement en valeur FCFA)")
        pair(f"Top {n_top} — Meilleur gain de CA", f"Flop {n_top} — Plus forte perte de CA",
             'Perte CA', ['CA','CA N-1','Évol CA %','Tx Marge %'],
             {'CA': fmt, 'CA N-1': fmt, 'Évol CA %': lambda v: fmt_pct(v), 'Tx Marge %': lambda v: fmt_pct(v)})

        st.markdown("##### 💰 Taux de marge")
        pair(f"Top {n_top} — Meilleur taux de marge", f"Flop {n_top} — Taux de marge le plus faible",
             'Tx Marge %', ['CA','Marge','Tx Marge %'],
             {'CA': fmt, 'Marge': fmt, 'Tx Marge %': lambda v: fmt_pct(v)}, ca_floor=1_000_000)

        st.markdown("##### 📊 Évolution du taux de marge (pts vs N-1)")
        pair(f"Top {n_top} — Meilleure progression", f"Flop {n_top} — Plus forte dégradation",
             'Écart Tx Marge (pts)', ['CA','Tx Marge %','Écart Tx Marge (pts)'],
             {'CA': fmt, 'Tx Marge %': lambda v: fmt_pct(v), 'Écart Tx Marge (pts)': fmt_delta}, ca_floor=1_000_000)

        st.markdown("##### 💔 Casse & 🎯 Poids promo")
        cC, cD = st.columns(2)
        with cC:
            st.markdown(f"**🔴 Flop {n_top} — Casse en valeur**")
            t = top_familles(df, n_top, 'casse').rename(
                columns={'Rayon_aff':'Rayon','Famille_aff':'Famille','Casse (Valeur)':'Casse','%Casse (Valeur)':'% Casse'})
            t['CA'] = t['CA'].map(fmt); t['Casse'] = t['Casse'].map(fmt)
            t['% Casse'] = t['% Casse'].map(lambda v: fmt_pct(v*100, 2))
            st.dataframe(t, hide_index=True, use_container_width=True)
        with cD:
            st.markdown(f"**🟠 Top {n_top} — Poids promo (CA&gt;1M)**")
            t = top_familles(df, n_top, 'promo').rename(
                columns={'Rayon_aff':'Rayon','Famille_aff':'Famille','%CA Poids Promo':'Poids Promo',
                         '%Marge Promo':'Tx M. Promo','%Marge Hors Promo':'Tx M. HP'})
            t['CA'] = t['CA'].map(fmt)
            for c in ['Poids Promo','Tx M. Promo','Tx M. HP']:
                t[c] = t[c].map(lambda v: fmt_pct(v*100))
            st.dataframe(t, hide_index=True, use_container_width=True)

        st.markdown("<div class='section-label'>EXPORT</div>", unsafe_allow_html=True)
        xls = build_excel_codir(k, perf, None)
        st.download_button("📥 Télécharger le récap CODIR (.xlsx)", xls,
                            file_name="CODIR_Hebdo_Rayon.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- TAB 2 : DESTRUCTEURS & PERFORMEURS (ARTICLE) ----------------
with tab2:
    st.caption(f"{len(art):,} lignes article chargées · seuil de matérialité : {fmt(seuil_ca)} FCFA "
               f"(modifiable dans la barre latérale)".replace(",", " "))
    res = destructeurs_performeurs(art, n=n_top, seuil_ca=seuil_ca)

    st.markdown(f"<span class='badge' style='background:#FFD6D4;color:{RED}'>A · Marge négative</span>", unsafe_allow_html=True)
    show_table(res['A_marge_neg'], ['CA','Marge','Tx Marge %','Qté Vente'],
               {'CA': fmt, 'Marge': fmt, 'Tx Marge %': lambda v: fmt_pct(v), 'Qté Vente': fmt})

    st.markdown(f"<span class='badge' style='background:#FFD6D4;color:{RED}'>B · Dégradation du taux de marge (marge encore positive)</span>", unsafe_allow_html=True)
    show_table(res['B_degrad_marge'], ['CA','Tx Marge %','Écart Tx Marge (pts)'],
               {'CA': fmt, 'Tx Marge %': lambda v: fmt_pct(v), 'Écart Tx Marge (pts)': fmt_delta})

    st.markdown(f"<span class='badge' style='background:#D7F5DE;color:#1A7A3A'>C · Performeurs — gain de marge en valeur</span>", unsafe_allow_html=True)
    show_table(res['C_perf_gain_marge'], ['CA','Gain Marge (FCFA)','Tx Marge %'],
               {'CA': fmt, 'Gain Marge (FCFA)': fmt, 'Tx Marge %': lambda v: fmt_pct(v)})

    st.markdown(f"<span class='badge' style='background:#D7F5DE;color:#1A7A3A'>D · Plus forte croissance de CA</span>", unsafe_allow_html=True)
    d4 = res['D_croissance_ca'].copy()
    if not d4.empty:
        d4['Évol CA %'] = (d4['CA']/d4['CA N-1']-1)*100
    show_table(d4, ['CA','CA N-1','Évol CA %','Tx Marge %'],
               {'CA': fmt, 'CA N-1': fmt, 'Évol CA %': lambda v: fmt_pct(v), 'Tx Marge %': lambda v: fmt_pct(v)})
    st.caption("⚠️ Une forte évolution % peut refléter un effet de base (article quasi absent en N-1) plutôt qu'une vraie dynamique.")

    st.markdown(f"<span class='badge' style='background:#FFD6D4;color:{RED}'>E · Plus forte baisse de CA</span>", unsafe_allow_html=True)
    show_table(res['E_baisse_ca'], ['CA','CA N-1','Perte CA (FCFA)'],
               {'CA': fmt, 'CA N-1': fmt, 'Perte CA (FCFA)': fmt})

    st.markdown(f"<span class='badge' style='background:#D7F5DE;color:#1A7A3A'>F · Plus forte hausse de quantité vendue</span>", unsafe_allow_html=True)
    show_table(res['F_hausse_qte'], ['Qté Vente','Qté Vente N-1','Variation Qté','CA'],
               {'Qté Vente': fmt, 'Qté Vente N-1': fmt, 'Variation Qté': lambda v: f"{v:+,.0f}".replace(",", " "), 'CA': fmt})

    st.markdown(f"<span class='badge' style='background:#FFD6D4;color:{RED}'>G · Plus forte baisse de quantité vendue</span>", unsafe_allow_html=True)
    show_table(res['G_baisse_qte'], ['Qté Vente','Qté Vente N-1','Variation Qté','CA'],
               {'Qté Vente': fmt, 'Qté Vente N-1': fmt, 'Variation Qté': lambda v: f"{v:+,.0f}".replace(",", " "), 'CA': fmt})
